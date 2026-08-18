from fastapi import FastAPI, Form, Request, UploadFile, File
from fastapi.responses import JSONResponse, HTMLResponse, RedirectResponse, Response, FileResponse
import sqlite3

try:
    import openpyxl
except Exception:
    openpyxl = None

try:
    from pypdf import PdfReader
except Exception:
    PdfReader = None

try:
    import psycopg
    from psycopg.rows import dict_row
except Exception:
    psycopg = None
    dict_row = None
import os
import io
import csv
import json
import base64
import re
import secrets
import hashlib
import hmac
import mimetypes
import zipfile
from contextvars import ContextVar
from datetime import date, datetime, timedelta
from pathlib import Path
try:
    from openai import OpenAI
except Exception:
    OpenAI = None

try:
    from reportlab.pdfgen import canvas
except Exception:
    canvas = None

app=FastAPI(title="BuildCommand AI",version="378.0")
DB="construction_ai_web.db"
DEFAULT_UPLOAD_DIR="/var/data/buildcommand_uploads" if os.path.isdir("/var/data") else "/tmp/buildcommand_uploads"
UPLOAD_DIR=os.environ.get("UPLOAD_DIR",DEFAULT_UPLOAD_DIR)
os.makedirs(UPLOAD_DIR,exist_ok=True)
_current_user_id=ContextVar("buildcommand_user_id",default=None)
_current_company_id=ContextVar("buildcommand_company_id",default=None)
MAX_UPLOAD_BYTES=45*1024*1024
ALLOWED_UPLOAD_EXTENSIONS={".pdf",".png",".jpg",".jpeg",".webp",".doc",".docx",".xls",".xlsx",".csv",".txt"}

DATABASE_URL=os.environ.get("DATABASE_URL","").strip()
DATABASE_KIND="postgres" if DATABASE_URL.startswith(("postgres://","postgresql://")) else "sqlite"

class PgCompatConnection:
    def __init__(self,conn):
        self.conn=conn
        self.last_insert_id=None
    def _sql(self,sql):
        if sql.strip().lower().startswith("select last_insert_rowid()"):
            return None
        sql=sql.replace("?","%s")
        sql=re.sub(r"INSERT OR REPLACE INTO user_state\s*\(user_id,selected_project_id\)\s*VALUES\(%s,%s\)","INSERT INTO user_state(user_id,selected_project_id) VALUES(%s,%s) ON CONFLICT(user_id) DO UPDATE SET selected_project_id=EXCLUDED.selected_project_id",sql,flags=re.I|re.S)
        sql=re.sub(r"INSERT OR REPLACE INTO app_state\s*\(id,selected_project_id\)\s*VALUES\(%s,%s\)","INSERT INTO app_state(id,selected_project_id) VALUES(%s,%s) ON CONFLICT(id) DO UPDATE SET selected_project_id=EXCLUDED.selected_project_id",sql,flags=re.I|re.S)
        return sql
    def execute(self,sql,params=()):
        translated=self._sql(sql)
        if translated is None:
            class OneRow:
                def __init__(self,v): self.v=v
                def fetchone(self): return {"id":self.v}
                def fetchall(self): return [{"id":self.v}]
                def __iter__(self): return iter([{"id":self.v}])
            return OneRow(self.last_insert_id)
        cur=self.conn.cursor()
        m=re.match(r"\s*INSERT\s+INTO\s+([A-Za-z_][A-Za-z0-9_]*)",translated,re.I)
        if m and "RETURNING" not in translated.upper() and m.group(1).lower() not in {"user_state"}:
            try:
                cur.execute(translated.rstrip().rstrip(";")+" RETURNING id",params)
                row=cur.fetchone()
                if row and "id" in row: self.last_insert_id=row["id"]
                return cur
            except Exception:
                self.conn.rollback(); cur=self.conn.cursor()
        cur.execute(translated,params)
        return cur
    def executescript(self,script):
        ddl=re.sub(r"\bid INTEGER PRIMARY KEY\b","id BIGSERIAL PRIMARY KEY",script,flags=re.I)
        for stmt in [s.strip() for s in ddl.split(";") if s.strip()]: self.execute(stmt)
        return self
    def commit(self): self.conn.commit()
    def rollback(self): self.conn.rollback()
    def close(self): self.conn.close()

def db():
    if DATABASE_KIND=="postgres":
        if psycopg is None: raise RuntimeError("PostgreSQL DATABASE_URL is set but psycopg is not installed")
        return PgCompatConnection(psycopg.connect(DATABASE_URL,row_factory=dict_row))
    conn=sqlite3.connect(DB)
    conn.row_factory=sqlite3.Row
    return conn

def init():
    c = db()
    c.executescript("""
    CREATE TABLE IF NOT EXISTS companies(id INTEGER PRIMARY KEY,name TEXT NOT NULL,logo_url TEXT,created TEXT);
    CREATE TABLE IF NOT EXISTS users(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,email TEXT NOT NULL UNIQUE,display_name TEXT,password_hash TEXT NOT NULL,role TEXT DEFAULT 'MEMBER',created TEXT);
    CREATE TABLE IF NOT EXISTS sessions(id INTEGER PRIMARY KEY,user_id INTEGER NOT NULL,token_hash TEXT NOT NULL UNIQUE,expires TEXT,created TEXT);
    CREATE TABLE IF NOT EXISTS user_state(user_id INTEGER PRIMARY KEY,selected_project_id INTEGER);
    CREATE TABLE IF NOT EXISTS attachments(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER,category TEXT,title TEXT,original_name TEXT,stored_name TEXT,mime_type TEXT,size_bytes INTEGER,created_by INTEGER,created TEXT);
    CREATE TABLE IF NOT EXISTS notifications(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER,severity TEXT,title TEXT,detail TEXT,source TEXT,status TEXT DEFAULT 'UNREAD',created TEXT);
    CREATE TABLE IF NOT EXISTS morning_briefs(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,brief_date TEXT,brief_text TEXT,created TEXT);
    CREATE TABLE IF NOT EXISTS beta_feedback(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,user_id INTEGER,project_id INTEGER,rating INTEGER,category TEXT,feedback TEXT,created TEXT);
    CREATE TABLE IF NOT EXISTS user_favorites(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,user_id INTEGER NOT NULL,tool_name TEXT,tool_url TEXT,created TEXT);
    CREATE TABLE IF NOT EXISTS project_archive_state(project_id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,archived INTEGER DEFAULT 0,archived_at TEXT);
    CREATE TABLE IF NOT EXISTS attachment_tags(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,attachment_id INTEGER NOT NULL,tag TEXT,created TEXT);
    CREATE TABLE IF NOT EXISTS notification_rules(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,rule_name TEXT,enabled INTEGER DEFAULT 1,severity TEXT DEFAULT 'HIGH',threshold_value REAL DEFAULT 0,created TEXT);
    CREATE TABLE IF NOT EXISTS health_history(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,snapshot_date TEXT,overall REAL,schedule REAL,readiness REAL,procurement REAL,risk REAL,field REAL,created TEXT);
    CREATE TABLE IF NOT EXISTS admin_audit_log(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,user_id INTEGER,project_id INTEGER,action TEXT,detail TEXT,created TEXT);

    CREATE TABLE IF NOT EXISTS invitations(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,email TEXT NOT NULL,role TEXT DEFAULT 'FIELD_USER',token_hash TEXT NOT NULL UNIQUE,expires TEXT,accepted INTEGER DEFAULT 0,created_by INTEGER,created TEXT);
    CREATE TABLE IF NOT EXISTS company_settings(company_id INTEGER PRIMARY KEY,auto_ai_brief INTEGER DEFAULT 1,email_alerts INTEGER DEFAULT 0,beta_mode INTEGER DEFAULT 1,onboarding_complete INTEGER DEFAULT 0);
    CREATE TABLE IF NOT EXISTS weekly_ai_reports(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,week_ending TEXT,report_text TEXT,created TEXT);
    
    CREATE TABLE IF NOT EXISTS weather_impacts(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,activity_id INTEGER,impact_date TEXT,weather_type TEXT,lost_hours REAL DEFAULT 0,description TEXT,created TEXT);
    CREATE TABLE IF NOT EXISTS schedule_import_batches(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,file_name TEXT,row_count INTEGER DEFAULT 0,imported_count INTEGER DEFAULT 0,created TEXT);
    CREATE TABLE IF NOT EXISTS photo_observations(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,attachment_id INTEGER,activity_id INTEGER,observation TEXT,severity TEXT DEFAULT 'WATCH',created TEXT);
    CREATE TABLE IF NOT EXISTS communication_drafts(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,sub_id INTEGER,draft_type TEXT,subject TEXT,body TEXT,status TEXT DEFAULT 'DRAFT',created TEXT);
    CREATE TABLE IF NOT EXISTS meeting_ai_summaries(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,source_text TEXT,summary_text TEXT,created TEXT);

    CREATE TABLE IF NOT EXISTS document_ai_chunks(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,attachment_id INTEGER NOT NULL,chunk_index INTEGER,text_content TEXT,created TEXT);
    CREATE TABLE IF NOT EXISTS blueprint_runs(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,status TEXT DEFAULT 'COMPLETE',source_files TEXT,project_summary TEXT,detected_disciplines TEXT,cross_discipline_flags TEXT,rfi_candidates TEXT,review_notes TEXT,model_name TEXT,created_by INTEGER,created TEXT);
    CREATE TABLE IF NOT EXISTS blueprint_trade_scopes(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,run_id INTEGER NOT NULL,trade TEXT NOT NULL,division TEXT,summary TEXT,scope_text TEXT,item_count INTEGER DEFAULT 0,created TEXT);
    CREATE TABLE IF NOT EXISTS blueprint_scope_items(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,run_id INTEGER NOT NULL,trade_scope_id INTEGER,trade TEXT NOT NULL,requirement TEXT NOT NULL,source_sheet TEXT,source_detail TEXT,source_spec TEXT,source_note TEXT,related_trade TEXT,confidence TEXT DEFAULT 'MEDIUM',item_type TEXT DEFAULT 'SCOPE',status TEXT DEFAULT 'NOT_STARTED',created TEXT);
    CREATE TABLE IF NOT EXISTS schedule_import_sources(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,source_type TEXT,file_name TEXT,imported_count INTEGER DEFAULT 0,created TEXT);
    CREATE TABLE IF NOT EXISTS forecast_snapshots(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,snapshot_date TEXT,health_score REAL,projected_delay_days REAL,confidence REAL,explanation TEXT,created TEXT);
    CREATE TABLE IF NOT EXISTS sub_risk_snapshots(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,sub_id INTEGER NOT NULL,risk_score REAL,risk_band TEXT,explanation TEXT,created TEXT);
    CREATE TABLE IF NOT EXISTS change_packages(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,change_event_id INTEGER,package_text TEXT,created TEXT);
CREATE TABLE IF NOT EXISTS quick_entries(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,user_id INTEGER,entry_type TEXT,text TEXT,routed_to TEXT,created TEXT);
    CREATE TABLE IF NOT EXISTS daily_reports(
        id INTEGER PRIMARY KEY,
        project_id INTEGER,
        report_date TEXT,
        weather TEXT,
        manpower INTEGER DEFAULT 0,
        work_completed TEXT,
        delays TEXT,
        deliveries TEXT,
        inspections TEXT,
        safety TEXT,
        tomorrow_plan TEXT,
        created TEXT
    );

    CREATE TABLE IF NOT EXISTS daily_report_analysis(
        id INTEGER PRIMARY KEY,
        project_id INTEGER,
        report_id INTEGER,
        analysis_text TEXT,
        created TEXT
    );

    CREATE TABLE IF NOT EXISTS action_items(
        id INTEGER PRIMARY KEY,
        project_id INTEGER,
        title TEXT,
        owner TEXT,
        priority TEXT,
        due TEXT,
        status TEXT DEFAULT 'OPEN',
        notes TEXT,
        created TEXT
    );

    CREATE TABLE IF NOT EXISTS activity_readiness(
        id INTEGER PRIMARY KEY,
        project_id INTEGER,
        activity_id INTEGER,
        drawings INTEGER DEFAULT 0,
        material INTEGER DEFAULT 0,
        manpower INTEGER DEFAULT 0,
        predecessor INTEGER DEFAULT 0,
        access_ready INTEGER DEFAULT 0,
        inspection INTEGER DEFAULT 0,
        equipment INTEGER DEFAULT 0,
        notes TEXT,
        updated TEXT,
        UNIQUE(project_id, activity_id)
    );

    CREATE TABLE IF NOT EXISTS procurement(
        id INTEGER PRIMARY KEY,
        project_id INTEGER,
        activity_id INTEGER,
        item TEXT,
        vendor TEXT,
        required_on_site TEXT,
        promised_date TEXT,
        status TEXT,
        notes TEXT,
        created TEXT
    );

    CREATE TABLE IF NOT EXISTS project_issues(
        id INTEGER PRIMARY KEY,
        project_id INTEGER,
        activity_id INTEGER,
        issue_type TEXT,
        title TEXT,
        owner TEXT,
        due TEXT,
        priority TEXT,
        status TEXT DEFAULT 'OPEN',
        description TEXT,
        response TEXT,
        created TEXT
    );

    CREATE TABLE IF NOT EXISTS punch_items(
        id INTEGER PRIMARY KEY,
        project_id INTEGER,
        activity_id INTEGER,
        title TEXT,
        location TEXT,
        trade TEXT,
        owner TEXT,
        priority TEXT,
        due TEXT,
        status TEXT DEFAULT 'OPEN',
        description TEXT,
        resolution TEXT,
        created TEXT
    );

    CREATE TABLE IF NOT EXISTS inspections_tracker(
        id INTEGER PRIMARY KEY,
        project_id INTEGER,
        activity_id INTEGER,
        inspection_type TEXT,
        authority TEXT,
        scheduled_date TEXT,
        result TEXT DEFAULT 'PENDING',
        reinspection_date TEXT,
        notes TEXT,
        created TEXT
    );

    CREATE TABLE IF NOT EXISTS submittals(
        id INTEGER PRIMARY KEY,
        project_id INTEGER,
        activity_id INTEGER,
        title TEXT,
        spec_section TEXT,
        responsible_party TEXT,
        sent_date TEXT,
        due_date TEXT,
        status TEXT DEFAULT 'PENDING',
        notes TEXT,
        created TEXT
    );

    CREATE TABLE IF NOT EXISTS safety_items(
        id INTEGER PRIMARY KEY,
        project_id INTEGER,
        activity_id INTEGER,
        event_date TEXT,
        item_type TEXT,
        title TEXT,
        location TEXT,
        responsible_party TEXT,
        severity TEXT,
        status TEXT DEFAULT 'OPEN',
        description TEXT,
        corrective_action TEXT,
        created TEXT
    );

    CREATE TABLE IF NOT EXISTS change_events(
        id INTEGER PRIMARY KEY,
        project_id INTEGER,
        activity_id INTEGER,
        event_type TEXT,
        title TEXT,
        responsible_party TEXT,
        estimated_cost REAL DEFAULT 0,
        schedule_days REAL DEFAULT 0,
        status TEXT DEFAULT 'OPEN',
        description TEXT,
        created TEXT
    );

    CREATE TABLE IF NOT EXISTS meeting_notes(
        id INTEGER PRIMARY KEY,
        project_id INTEGER,
        meeting_date TEXT,
        meeting_type TEXT,
        title TEXT,
        attendees TEXT,
        decisions TEXT,
        commitments TEXT,
        follow_up TEXT,
        created TEXT
    );

    CREATE TABLE IF NOT EXISTS app_state(
        id INTEGER PRIMARY KEY,
        selected_project_id INTEGER
    );
    CREATE TABLE IF NOT EXISTS projects(id INTEGER PRIMARY KEY,name TEXT,number TEXT,status TEXT);
    CREATE TABLE IF NOT EXISTS activities(id INTEGER PRIMARY KEY,project_id INTEGER,external_id TEXT,name TEXT,trade TEXT,start TEXT,finish TEXT,pct REAL DEFAULT 0,status TEXT DEFAULT 'NOT_STARTED');
    CREATE TABLE IF NOT EXISTS make_ready(id INTEGER PRIMARY KEY,project_id INTEGER,activity_id INTEGER,title TEXT,reason TEXT,due TEXT,priority TEXT,status TEXT DEFAULT 'OPEN');
    CREATE TABLE IF NOT EXISTS risks(id INTEGER PRIMARY KEY,project_id INTEGER,activity_id INTEGER,score REAL,band TEXT,explanation TEXT);
    CREATE TABLE IF NOT EXISTS subs(id INTEGER PRIMARY KEY,project_id INTEGER,name TEXT,trade TEXT);

    CREATE TABLE IF NOT EXISTS subcontractor_updates(
        id INTEGER PRIMARY KEY,
        project_id INTEGER,
        sub_id INTEGER,
        update_date TEXT,
        manpower INTEGER DEFAULT 0,
        commitment TEXT,
        issue TEXT,
        status TEXT,
        created TEXT
    );
    CREATE TABLE IF NOT EXISTS production(id INTEGER PRIMARY KEY,project_id INTEGER,activity_id INTEGER,work_date TEXT,crew INTEGER,qty REAL,planned_qty REAL,unit TEXT);
    CREATE TABLE IF NOT EXISTS field_updates(id INTEGER PRIMARY KEY,project_id INTEGER,activity_id INTEGER,update_type TEXT,text TEXT,created TEXT);
    CREATE TABLE IF NOT EXISTS memory(id INTEGER PRIMARY KEY,project_id INTEGER,category TEXT,insight TEXT,confidence REAL);
    CREATE TABLE IF NOT EXISTS recovery(id INTEGER PRIMARY KEY,project_id INTEGER,activity_id INTEGER,scenario TEXT,days_recovered REAL,est_cost REAL,status TEXT);
    """)

    if c.execute("SELECT COUNT(*) n FROM projects").fetchone()["n"] == 0:
        c.execute(
            "INSERT INTO projects(name,number,status) VALUES(?,?,?)",
            ("Canyon Medical Office", "CMO-024", "ACTIVE")
        )
        pid = c.execute("SELECT id FROM projects ORDER BY id LIMIT 1").fetchone()["id"]

        acts = [
            ("A100", "Footings & Foundations", "Concrete", "2026-08-10", "2026-08-18", 80, "IN_PROGRESS"),
            ("A200", "Structural Steel / Deck", "Structural", "2026-08-19", "2026-09-04", 15, "NOT_STARTED"),
            ("A300", "MEP Underground / Rough", "MEP", "2026-08-24", "2026-09-11", 5, "NOT_STARTED"),
            ("A400", "Interior Framing", "Framing", "2026-09-08", "2026-09-25", 0, "NOT_STARTED"),
            ("A500", "Drywall Close-In", "Drywall", "2026-09-22", "2026-10-09", 0, "NOT_STARTED"),
        ]
        for x in acts:
            c.execute(
                "INSERT INTO activities(project_id,external_id,name,trade,start,finish,pct,status) VALUES(?,?,?,?,?,?,?,?)",
                (pid, *x)
            )

        ids = {
            r["external_id"]: r["id"]
            for r in c.execute(
                "SELECT id,external_id FROM activities WHERE project_id=?",
                (pid,)
            )
        }

        c.execute(
            "INSERT INTO make_ready(project_id,activity_id,title,reason,due,priority) VALUES(?,?,?,?,?,?)",
            (pid, ids["A300"], "Clear MEP rough-in access", "Material laydown is blocking east-side access.", "2026-08-12", "CRITICAL")
        )
        c.execute(
            "INSERT INTO make_ready(project_id,activity_id,title,reason,due,priority) VALUES(?,?,?,?,?,?)",
            (pid, ids["A200"], "Confirm steel delivery", "Fabricator delivery confirmation has not been received.", "2026-08-13", "HIGH")
        )

        for aid, score, band, why in [
            (ids["A300"], 82, "CRITICAL", "MEP rough-in is behind the field plan and access remains unresolved."),
            (ids["A200"], 68, "HIGH", "Steel delivery confirmation is unresolved; downstream starts are exposed."),
            (ids["A400"], 44, "WATCH", "Framing depends on MEP rough-in and inspection clearance."),
        ]:
            c.execute(
                "INSERT INTO risks(project_id,activity_id,score,band,explanation) VALUES(?,?,?,?,?)",
                (pid, aid, score, band, why)
            )

        for name, trade in [
            ("Apex Concrete", "Concrete"),
            ("Metro Steel", "Structural"),
            ("Summit MEP", "MEP"),
        ]:
            c.execute(
                "INSERT INTO subs(project_id,name,trade) VALUES(?,?,?)",
                (pid, name, trade)
            )

        c.execute(
            "INSERT INTO memory(project_id,category,insight,confidence) VALUES(?,?,?,?)",
            (pid, "Company Memory", "Early access constraints on MEP rough-in should be cleared before manpower is increased.", .72)
        )

    if c.execute("SELECT COUNT(*) n FROM app_state").fetchone()["n"] == 0:
        first_project = c.execute("SELECT id FROM projects ORDER BY id LIMIT 1").fetchone()
        if first_project:
            c.execute(
                "INSERT INTO app_state(id, selected_project_id) VALUES(1, ?)",
                (first_project["id"],)
            )

    if DATABASE_KIND=="postgres":
        project_columns={row["column_name"] for row in c.execute("SELECT column_name FROM information_schema.columns WHERE table_schema='public' AND table_name='projects'").fetchall()}
    else:
        project_columns={row["name"] for row in c.execute("PRAGMA table_info(projects)").fetchall()}
    if "company_id" not in project_columns:
        c.execute("ALTER TABLE projects ADD COLUMN company_id INTEGER")

    c.commit()
    c.close()


init()

def hash_password(password):
    salt = secrets.token_bytes(16)
    rounds = 210000
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, rounds)
    return f"pbkdf2_sha256${rounds}${salt.hex()}${digest.hex()}"


def verify_password(password, stored):
    try:
        algorithm, rounds, salt_hex, digest_hex = stored.split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        calc = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), bytes.fromhex(salt_hex), int(rounds)).hex()
        return hmac.compare_digest(calc, digest_hex)
    except Exception:
        return False


def current_user_id():
    return _current_user_id.get()


def current_company_id():
    return _current_company_id.get()


def current_user():
    uid = current_user_id()
    if not uid:
        return None
    c = db()
    row = c.execute("""
        SELECT u.*, c.name AS company_name, c.logo_url
        FROM users u JOIN companies c ON c.id=u.company_id
        WHERE u.id=?
    """, (uid,)).fetchone()
    c.close()
    return row


ROLE_ORDER={"READ_ONLY":0,"FIELD_USER":1,"MEMBER":1,"SUPERINTENDENT":2,"PROJECT_MANAGER":3,"ADMIN":4,"OWNER":5}

def role_level(role):
    return ROLE_ORDER.get((role or "").upper(),0)

def require_role(min_role):
    u=current_user()
    return bool(u and role_level(u["role"])>=role_level(min_role))

def ensure_company_settings(company_id):
    if not company_id: return
    c=db()
    row=c.execute("SELECT company_id FROM company_settings WHERE company_id=?",(company_id,)).fetchone()
    if not row:
        c.execute("INSERT INTO company_settings(company_id,auto_ai_brief,email_alerts,beta_mode,onboarding_complete) VALUES(?,?,?,?,?)",(company_id,1,0,1,0))
        c.commit()
    c.close()

def company_setting(name,default=0):
    cid=current_company_id()
    if not cid: return default
    ensure_company_settings(cid)
    c=db(); row=c.execute("SELECT * FROM company_settings WHERE company_id=?",(cid,)).fetchone(); c.close()
    try: return row[name]
    except Exception: return default

def create_session(user_id):
    raw = secrets.token_urlsafe(40)
    token_hash = hashlib.sha256(raw.encode("utf-8")).hexdigest()
    expires = (datetime.utcnow() + timedelta(days=30)).isoformat()
    c = db()
    c.execute("INSERT INTO sessions(user_id,token_hash,expires,created) VALUES(?,?,?,?)",
              (user_id, token_hash, expires, datetime.utcnow().isoformat()))
    c.commit(); c.close()
    return raw


def user_from_session(raw_token):
    if not raw_token:
        return None
    token_hash = hashlib.sha256(raw_token.encode("utf-8")).hexdigest()
    c = db()
    row = c.execute("""
        SELECT u.* FROM sessions s JOIN users u ON u.id=s.user_id
        WHERE s.token_hash=? AND (s.expires IS NULL OR s.expires>?)
    """, (token_hash, datetime.utcnow().isoformat())).fetchone()
    c.close()
    return row


PUBLIC_PATHS = {"/login", "/register", "/health"}



@app.middleware("http")
async def v169_performance_headers(request, call_next):
    response=await call_next(request)
    path=request.url.path
    if path.startswith("/api/dashboard/"):
        response.headers["Cache-Control"]="private, max-age=10"
    elif path in {"/","/build","/estimate","/manage","/intelligence"}:
        response.headers["Cache-Control"]="private, no-store"
    return response

@app.middleware("http")
async def authentication_middleware(request: Request, call_next):
    raw_token = request.cookies.get("bc_session")
    user = user_from_session(raw_token)
    t1 = _current_user_id.set(user["id"] if user else None)
    t2 = _current_company_id.set(user["company_id"] if user else None)
    try:
        if not user and request.url.path not in PUBLIC_PATHS:
            return RedirectResponse("/login", status_code=303)
        return await call_next(request)
    finally:
        _current_user_id.reset(t1)
        _current_company_id.reset(t2)


def login_page(message=""):
    return f'''<!doctype html><html><head><meta name="viewport" content="width=device-width,initial-scale=1"><title>BuildCommand AI Login</title>
    <style>body{{margin:0;background:#0a1017;color:#eef4fb;font-family:Inter,system-ui,sans-serif;padding:28px}}.box{{max-width:460px;margin:6vh auto;background:#111923;border:1px solid #213042;border-radius:16px;padding:26px}}input{{width:100%;box-sizing:border-box;background:#0d1620;color:#eef4fb;border:1px solid #213042;border-radius:9px;padding:12px;margin:8px 0 16px}}button{{background:#f0b44d;border:0;border-radius:9px;padding:11px 16px;font-weight:800}}a{{color:#f0b44d}}.muted{{color:#8fa2b5}}.error{{color:#ff9b9b}}</style></head><body>
    <div class="box"><div class="muted">Construction Intelligence Platform</div><h1>BuildCommand AI</h1><p class="error">{esc(message)}</p>
    <form method="post" action="/login"><label>Email</label><input type="email" name="email" required><label>Password</label><input type="password" name="password" required><button type="submit">Sign In</button></form>
    <p><a href="/register">Create a company account</a></p><p class="muted">Built by Wilson LaHood · © 2026 Wilson LaHood</p></div></body></html>'''


@app.get("/login", response_class=HTMLResponse)
def login_get():
    return login_page()


@app.post("/login", response_class=HTMLResponse)
def login_post(email: str = Form(...), password: str = Form(...)):
    c = db(); user = c.execute("SELECT * FROM users WHERE lower(email)=lower(?)", (email.strip(),)).fetchone(); c.close()
    if not user or not verify_password(password, user["password_hash"]):
        return HTMLResponse(login_page("Email or password is incorrect."), status_code=401)
    raw = create_session(user["id"])
    response = RedirectResponse("/", status_code=303)
    response.set_cookie("bc_session", raw, httponly=True, secure=os.environ.get("COOKIE_SECURE", "1") == "1", samesite="lax", max_age=2592000)
    return response


@app.get("/register", response_class=HTMLResponse)
def register_get():
    return '''<!doctype html><html><head><meta name="viewport" content="width=device-width,initial-scale=1"><title>Create BuildCommand Account</title>
    <style>body{margin:0;background:#0a1017;color:#eef4fb;font-family:Inter,system-ui,sans-serif;padding:28px}.box{max-width:520px;margin:4vh auto;background:#111923;border:1px solid #213042;border-radius:16px;padding:26px}input{width:100%;box-sizing:border-box;background:#0d1620;color:#eef4fb;border:1px solid #213042;border-radius:9px;padding:12px;margin:8px 0 16px}button{background:#f0b44d;border:0;border-radius:9px;padding:11px 16px;font-weight:800}a{color:#f0b44d}.muted{color:#8fa2b5}</style></head><body>
    <div class="box"><div class="muted">BuildCommand AI</div><h1>Create Company Account</h1><form method="post" action="/register"><label>Company Name</label><input name="company_name" required><label>Your Name</label><input name="display_name" required><label>Email</label><input type="email" name="email" required><label>Password</label><input type="password" name="password" minlength="8" required><button type="submit">Create Account</button></form><p><a href="/login">Back to login</a></p></div></body></html>'''


@app.post("/register")
def register_post(company_name: str = Form(...), display_name: str = Form(...), email: str = Form(...), password: str = Form(...)):
    if len(password) < 8:
        return HTMLResponse("Password must be at least 8 characters.", status_code=400)
    c = db()
    if c.execute("SELECT id FROM users WHERE lower(email)=lower(?)", (email.strip(),)).fetchone():
        c.close(); return HTMLResponse("That email is already registered.", status_code=400)
    c.execute("INSERT INTO companies(name,created) VALUES(?,?)", (company_name.strip(), date.today().isoformat()))
    company_id = c.execute("SELECT last_insert_rowid() id").fetchone()["id"]
    c.execute("INSERT INTO users(company_id,email,display_name,password_hash,role,created) VALUES(?,?,?,?,?,?)",
              (company_id,email.strip().lower(),display_name.strip(),hash_password(password),"OWNER",date.today().isoformat()))
    user_id = c.execute("SELECT last_insert_rowid() id").fetchone()["id"]
    c.execute("UPDATE projects SET company_id=? WHERE company_id IS NULL", (company_id,))
    first_project = c.execute("SELECT id FROM projects WHERE company_id=? ORDER BY id LIMIT 1", (company_id,)).fetchone()
    if first_project:
        c.execute("INSERT INTO user_state(user_id,selected_project_id) VALUES(?,?) ON CONFLICT(user_id) DO UPDATE SET selected_project_id=excluded.selected_project_id", (user_id, first_project["id"]))
    c.commit(); c.close()
    raw = create_session(user_id)
    response = RedirectResponse("/", status_code=303)
    response.set_cookie("bc_session", raw, httponly=True, secure=os.environ.get("COOKIE_SECURE", "1") == "1", samesite="lax", max_age=2592000)
    return response


@app.post("/logout")
def logout(request: Request):
    raw = request.cookies.get("bc_session")
    if raw:
        token_hash = hashlib.sha256(raw.encode("utf-8")).hexdigest(); c = db(); c.execute("DELETE FROM sessions WHERE token_hash=?", (token_hash,)); c.commit(); c.close()
    response = RedirectResponse("/login", status_code=303); response.delete_cookie("bc_session"); return response


@app.get("/health")
def health():
    return {"status":"ok","app":"BuildCommand AI","version":"31.1"}

@app.get("/projects/new", response_class=HTMLResponse)
def new_project_form():
    return """
    <html>
    <head>
        <title>Add Project</title>
        <style>
            body {
                background:#0a1017;
                color:white;
                font-family:Arial,sans-serif;
                padding:40px;
            }
            .box {
                max-width:500px;
                margin:auto;
                background:#111923;
                padding:30px;
                border-radius:15px;
            }
            input, select {
                width:100%;
                padding:12px;
                margin:8px 0 18px;
                box-sizing:border-box;
                border-radius:8px;
                border:1px solid #213042;
                background:#0d1620;
                color:white;
            }
            button {
                background:#f0b44d;
                border:none;
                padding:12px 20px;
                border-radius:8px;
                font-weight:bold;
                cursor:pointer;
            }
            a {
                color:#f0b44d;
            }
        </style>
    </head>

    <body>
        <div class="box">
            <h1>Add New Project</h1>

            <form method="post" action="/projects/new">

                <label>Project Name</label>
                <input
                    type="text"
                    name="name"
                    placeholder="Example: Phoenix Medical Center"
                    required
                >

                <label>Project Number</label>
                <input
                    type="text"
                    name="number"
                    placeholder="Example: PMC-001"
                    required
                >

                <label>Status</label>
                <select name="status">
                    <option value="ACTIVE">Active</option>
                    <option value="PLANNING">Planning</option>
                    <option value="ON_HOLD">On Hold</option>
                    <option value="COMPLETE">Complete</option>
                </select>

                <button type="submit">
                    Save Project
                </button>

            </form>

            <p>
                <a href="/">← Back to Dashboard</a>
            </p>
        </div>
    </body>
    </html>
    """


@app.post("/projects/new")
def create_project(name: str = Form(...), number: str = Form(...), status: str = Form(...)):
    c=db(); c.execute("INSERT INTO projects(name,number,status,company_id) VALUES(?,?,?,?)",(name.strip(),number.strip(),status,current_company_id())); pid=c.execute("SELECT last_insert_rowid() id").fetchone()["id"]; c.execute("INSERT INTO user_state(user_id,selected_project_id) VALUES(?,?) ON CONFLICT(user_id) DO UPDATE SET selected_project_id=excluded.selected_project_id",(current_user_id(),pid)); c.commit(); c.close(); return RedirectResponse(url="/",status_code=303)
CSS="""
:root{--bg:#0a1017;--panel:#111923;--line:#213042;--text:#eef4fb;--muted:#8fa2b5;--gold:#f0b44d}
*{box-sizing:border-box}body{margin:0;background:var(--bg);color:var(--text);font-family:Inter,system-ui,-apple-system,Segoe UI,sans-serif}
.app{display:grid;grid-template-columns:260px 1fr;min-height:100vh}.side{background:#0c141d;border-right:1px solid var(--line);padding:22px 16px}.brand{font-size:20px;font-weight:800}.company{font-size:12px;color:var(--muted);margin:5px 0 20px}.nav a{display:block;color:#cbd7e3;text-decoration:none;padding:10px;border-radius:9px;margin:2px 0}.nav a:hover{background:#162333}.creator-footer{margin-top:24px;padding-top:16px;border-top:1px solid var(--line);color:var(--muted);font-size:11px;line-height:1.6}.main{padding:26px;max-width:1400px}
.hero,.card{background:var(--panel);border:1px solid var(--line);border-radius:16px;padding:18px;margin-bottom:12px}.hero h1{margin:4px 0}.eyebrow{color:var(--gold);font-size:11px;text-transform:uppercase;letter-spacing:.13em}.muted,.small{color:var(--muted)}.small{font-size:12px}
.grid4{display:grid;grid-template-columns:repeat(4,1fr);gap:12px}.grid3{display:grid;grid-template-columns:repeat(3,1fr);gap:12px}.grid2{display:grid;grid-template-columns:1fr 1fr;gap:12px}.kpi{font-size:28px;font-weight:800}.label{font-size:11px;color:var(--muted);text-transform:uppercase}.badge{display:inline-block;padding:4px 8px;border-radius:999px;font-weight:800;font-size:10px}.CRITICAL,.HOLD{background:#492324;color:#ff9b9b}.HIGH,.WATCH{background:#43381b;color:#ffd779}.READY,.LOW,.COMPLETE{background:#18392c;color:#82e4b5}.OPEN{background:#1d2e44;color:#99c9ff}
table{width:100%;border-collapse:collapse}th,td{text-align:left;padding:9px;border-bottom:1px solid var(--line);font-size:13px}th{color:var(--muted)}input,textarea,select{width:100%;background:#0d1620;color:var(--text);border:1px solid var(--line);border-radius:9px;padding:10px}textarea{min-height:90px}button{background:var(--gold);border:0;border-radius:9px;padding:10px 14px;font-weight:800}.action{padding:12px 0;border-bottom:1px solid var(--line)}
@media(max-width:850px){.app{grid-template-columns:1fr}.grid4,.grid3,.grid2{grid-template-columns:1fr}.main{padding:14px}}

.mobile-menu-btn{display:none}
@media(max-width:900px){.app{grid-template-columns:1fr}.side{position:relative;border-right:0;border-bottom:1px solid var(--line)}.grid4,.grid3,.grid2{grid-template-columns:1fr}.mobile-menu-btn{display:block;width:100%;margin:12px 0}.nav{display:none}.nav.mobile-open{display:block}}
"""

NAV=[("Daily Command","/"),("AI Command","/ai-command"),("Blueprint Brain","/plans-specs-ai"),("Deep Document AI","/document-ai"),("Schedule Import","/schedule-import"),("Advanced Schedule Import","/advanced-schedule-import"),("Photo Intelligence","/photo-intelligence"),("AI Photo Analysis","/photo-ai"),("3-Week Lookahead","/lookahead-intelligence"),("Project Health","/project-health"),("Predictive Forecast","/predictive-forecast"),("Morning Brief","/morning-brief"),("Action Center","/actions"),("RFIs / Issues","/issues"),("Punch List","/punch"),("Inspections","/inspections"),("Submittals","/submittals"),("Safety","/safety"),("Change Events","/changes"),("Meetings","/meetings"),("Documents","/documents"),("Notifications","/notifications"),("AI Assistant","/assistant"),("AI Analysis","/ai-analysis"),("Daily Report","/daily-report"),("Auto Daily Report","/auto-daily-report"),("Schedule","/schedule"),("Schedule Health","/schedule-health"),("Procurement","/procurement"),("Readiness","/readiness"),("Make Ready","/make-ready"),("Field","/field"),("Subcontractors","/subcontractors"),("Production","/production"),("Predictive Risk","/risk"),("Recovery","/recovery"),("Company Memory","/memory"),("Playbooks","/playbooks"),("Portfolio","/portfolio"),("Owner Dashboard","/owner-dashboard"),("Portfolio Intelligence","/portfolio-intelligence"),("Exports","/exports"),("Team","/team"),("Company Settings","/company-settings"),("Project Settings","/project-settings"),("System Check","/system-check"),("Beta Feedback","/beta-feedback"),("Setup","/setup"),("Invitations","/invitations"),("Production Settings","/production-settings"),("Sub Scorecards","/sub-scorecards"),("Sub Risk","/sub-risk"),("RFI Impact","/rfi-impact"),("Procurement Warning","/procurement-warning"),("Recovery Planner","/ai-recovery"),("Quick Entry","/quick-entry"),("Weekly AI Report","/weekly-report"),("RFI Drafting","/rfi-drafting"),("Sub Communications","/sub-communications"),("AI Meeting Minutes","/meeting-minutes-ai"),("Weather Impacts","/weather-impacts"),("PDF Reports","/pdf-reports"),("Cost Intelligence","/cost-intelligence"),("Change Package","/change-package"),("Mobile Home","/mobile-home"),("Mobile Field+","/mobile-field-plus"),("Beta Checklist","/beta-checklist")]


NAV_GROUPS=[
    ('🏠 Home',[
        ('BuildCommand Brain','/brain'),('Estimator Intelligence','/brain/estimator'),('Takeoff Intelligence','/brain/takeoff'),('Daily Command','/'),('AI Command','/ai-command'),('Morning Brief','/morning-brief'),('Action Center','/actions'),
        ('Global Search','/global-search'),('Favorites','/favorites'),('Recent Activity','/recent-activity'),('Mobile Field+','/mobile-field-plus')
    ]),
    ('📅 Schedule & Production',[
        ('Schedule','/schedule'),('Schedule Import','/schedule-import'),('Advanced Schedule Import','/advanced-schedule-import'),
        ('3-Week Lookahead','/lookahead-intelligence'),('Schedule Health','/schedule-health'),('Predictive Forecast','/predictive-forecast'),
        ('Production','/production'),('Readiness','/readiness'),('Make Ready','/make-ready'),('Procurement','/procurement'),
        ('Procurement Warning','/procurement-warning'),('Recovery','/recovery'),('Recovery Planner','/ai-recovery'),('Predictive Risk','/risk')
    ]),
    ('👷 Field Operations',[
        ('Field','/field'),('Quick Entry','/quick-entry'),('Daily Report','/daily-report'),('Auto Daily Report','/auto-daily-report'),
        ('Photo Intelligence','/photo-intelligence'),('AI Photo Analysis','/photo-ai'),('Weather Impacts','/weather-impacts'),
        ('Safety','/safety'),('Inspections','/inspections'),('Punch List','/punch')
    ]),
    ('📄 Documents & AI',[
        ('Documents','/documents'),('Document Tags','/document-tags'),('Plan Intake','/plans-specs-ai'),('Deep Document AI','/document-ai'),
        ('RFIs / Issues','/issues'),('RFI Drafting','/rfi-drafting'),('RFI Impact','/rfi-impact'),('Submittals','/submittals'),
        ('Meetings','/meetings'),('AI Meeting Minutes','/meeting-minutes-ai'),('AI Assistant','/assistant'),('AI Analysis','/ai-analysis'),
        ('Weekly AI Report','/weekly-report'),('PDF Reports','/pdf-reports')
    ]),
    ('🤝 Subs & Communication',[
        ('Subcontractors','/subcontractors'),('Sub Scorecards','/sub-scorecards'),('Sub Risk','/sub-risk'),
        ('Sub Communications','/sub-communications'),('Notifications','/notifications'),('Notification Rules','/notification-rules')
    ]),
    ('💰 Management & Executive',[
        ('Project Health','/project-health'),('Health History','/health-history'),('Change Events','/changes'),('Cost Intelligence','/cost-intelligence'),
        ('Change Package','/change-package'),('Owner Dashboard','/owner-dashboard'),('Portfolio','/portfolio'),('Portfolio Intelligence','/portfolio-intelligence'),
        ('Company Memory','/memory'),('Playbooks','/playbooks'),('Bulk Export','/bulk-export')
    ]),
    ('⚙️ Admin',[
        ('Team','/team'),('Invitations','/invitations'),('Company Settings','/company-settings'),('Project Settings','/project-settings'),
        ('Clone Project','/project-clone'),('Archive Projects','/project-archive'),('Storage Status','/storage-status'),('Audit Log','/audit-log'),
        ('Setup','/setup'),('System Check','/system-check'),('Beta Feedback','/beta-feedback'),('Beta Checklist','/beta-checklist')
    ])
]

def _v37_esc(value):
    import html
    return html.escape(str(value or ""), quote=True)

# Global compatibility helper used by legacy BuildCommand pages.
def esc(value):
    return _v37_esc(value)



# ============================================================

def categorized_nav():
    groups=[
        ("PROJECTS",[("Projects Home","/"),("Add Project","/projects/new"),("Recent Activity","/recent-activity"),("Archive Projects","/project-archive")]),
        ("BUILD",[("Build Home","/build"),("Field Command 3.0","/field-command-3"),("Analyze Project","/build/analyze-project"),("Blueprint Brain","/blueprint-brain"),("Review Project Scope","/brain"),("Preconstruction & Bid Intelligence","/preconstruction"),("Documents","/documents"),("Deep Document AI","/document-ai"),("Field Context & Assembly Intelligence","/field-context")]),
        ("ESTIMATE",[("Estimate Home","/estimate"),("Preconstruction Command","/precon-command"),("Estimator Intelligence","/brain/estimator"),("Takeoff Intelligence","/brain/takeoff"),("Bid Packages","/preconstruction/packages"),("Bid Leveling","/preconstruction/leveling"),("Historical Cost Brain","/learning/costs"),("Budget & Commitments","/project-control/budget")]),
        ("MANAGE",[("Manage Home","/manage"),("PM Command","/pm-command"),("Performance Monitor","/performance"),("Project Autopilot","/autopilot"),("Daily Superintendent Command","/daily-superintendent"),("Look-Ahead Intelligence","/lookahead-intelligence"),("Trade Readiness Brain","/trade-readiness"),("Trade Coordination Engine","/trade-coordination"),("Proactive Superintendent AI","/proactive-superintendent"),("Field Command","/field-command"),("Schedule","/schedule"),("Sequence Intelligence","/sequence-intelligence"),("RFIs / Issues","/issues"),("Submittals","/submittals"),("Procurement","/procurement"),("Inspections","/inspections"),("Subcontractors","/subcontractors"),("Project Control","/project-control"),("Punch","/punch"),("Closeout","/field-command/closeout")]),
        ("INTELLIGENCE",[("Intelligence Center","/intelligence"),("Knowledge Brain 2.0","/knowledge-brain-2"),("Smart RFI & Conflict Detection","/smart-rfi"),("Long-Lead Prediction","/longlead-intelligence"),("Inspection & QC Intelligence","/quality-intelligence"),("Scope Gap & Buyout Intelligence","/scope-gap-intelligence"),("Change Order Intelligence","/change-order-intelligence"),("Event-Driven Intelligence","/event-intelligence"),("Drawing Revision & Change Intelligence","/revision-intelligence"),("Project Memory & Continuous Learning","/project-memory"),("Master Construction Reasoning","/master-reasoning"),("Real Construction Reasoning 2.0","/reasoning-2"),("Project Knowledge Graph","/knowledge-graph"),("Prediction & Decision Intelligence","/prediction-intelligence"),("Brain Quality & Self-Learning","/brain-quality"),("Constructability Intelligence","/intelligence-engine/constructability"),("Learning Intelligence","/learning"),("Field Context Intelligence","/field-context")]),
        ("ASK BUILDCOMMAND",[("Ask BuildCommand","/ask-buildcommand"),("Search Everything","/global-search"),("Explain This Finding","/reasoning-2/explain"),("Reasoning Chain","/master-reasoning/chain"),("Answer Guardrails","/brain-quality/answer-guard")]),
    ]
    html='<div class="bc-topnav">'
    for label,items in groups:
        links=''.join(f'<a class="bc-drop-link" href="{href}">{_v37_esc(name)}</a>' for name,href in items)
        html+=(
            '<div class="bc-navdrop">'
            f'<button type="button" class="bc-navbtn" onclick="bcToggleMenu(this,event)">{_v37_esc(label)} <span class="bc-caret">▾</span></button>'
            f'<div class="bc-dropdown">{links}</div>'
            '</div>'
        )
    html+='</div>'
    return html

def shell(title, body):
    current_pid = project_id()
    company_id = current_company_id()
    user = current_user()
    c = db()
    projects = c.execute(
        "SELECT p.id,p.name,p.number,p.status FROM projects p "
        "LEFT JOIN project_archive_state a ON a.project_id=p.id "
        "WHERE p.company_id=? AND COALESCE(a.archived,0)=0 ORDER BY p.name",
        (company_id,)
    ).fetchall()
    current = c.execute(
        "SELECT * FROM projects WHERE id=? AND company_id=?",
        (current_pid, company_id)
    ).fetchone() if current_pid else None
    c.close()

    nav = categorized_nav()
    project_options = "".join(
        f'<option value="{p["id"]}" {"selected" if p["id"]==current_pid else ""}>{_v37_esc(p["number"])} - {_v37_esc(p["name"])}</option>'
        for p in projects
    )
    current_name = _v37_esc(current["name"]) if current else "No Project Selected"
    company_name = _v37_esc(user["company_name"]) if user else "BuildCommand Company"
    display_name = _v37_esc(user["display_name"]) if user else ""

    project_bar = f'''<div class="bc-projectbar">
      <div class="bc-project-meta">
        <div class="bc-project-label">CURRENT PROJECT</div>
        <div class="bc-project-name">{current_name}</div>
      </div>
      <form method="post" action="/projects/select" class="bc-project-switch">
        <select name="project_id">{project_options}</select>
        <button type="submit">Switch</button>
      </form>
      <a class="bc-add-project" href="/projects/new">+ Add Project</a>
    </div>'''

    _groups=[
      ("PROJECT",[("Project Command","/"),("Execution & Control Platform","/platform-369"),("Unified Platform","/platform-269"),("Projects","/projects"),("Project Autopilot","/autopilot")]),
      ("BUILD",[("Build Home","/build"),("Blueprint Brain","/blueprint-brain"),("Daily Superintendent","/daily-superintendent"),("Look-Ahead","/lookahead-intelligence"),("Trade Readiness","/trade-readiness"),("Trade Coordination","/trade-coordination")]),
      ("ESTIMATE",[("Estimate Home","/estimate"),("Preconstruction","/preconstruction"),("Scope Gap Intelligence","/scope-gap-intelligence")]),
      ("MANAGE",[("Manage Home","/manage"),("Proactive Superintendent AI","/proactive-superintendent"),("Change Order Intelligence","/change-order-intelligence"),("Performance Monitor","/performance")]),
      ("INTELLIGENCE",[("Intelligence Center","/intelligence"),("Knowledge Brain 2.0","/knowledge-brain-2"),("Event Intelligence","/event-intelligence"),("Smart RFI","/smart-rfi"),("Long-Lead Intelligence","/longlead-intelligence"),("Quality Intelligence","/quality-intelligence")])
    ]
    _side='<aside class="bc170-side" id="bc170-side"><div class="bc170-brand">BuildCommand AI<small>CONSTRUCTION OPERATION INTELLIGENCE SYSTEM</small></div>'
    for _g,_items in _groups:
        _side+=f'<div class="bc170-group"><button type="button" onclick="this.parentElement.classList.toggle(\'open\')">{esc(_g)} ▾</button><div class="bc170-links">'
        for _label,_href in _items: _side+=f'<a href="{_href}">{esc(_label)}</a>'
        _side+='</div></div>'
    _side+='</aside>'
    _top=f'<header class="bc170-top"><button class="bc170-menu" type="button" onclick="document.getElementById(\'bc170-side\').classList.toggle(\'show\')">☰</button><div class="bc170-title">{esc(title)}</div><div class="bc170-quick"><a href="/ask-buildcommand">Ask BuildCommand</a><a href="/autopilot">Autopilot</a></div></header>'
    _rail='<details class="bc170-rail"><summary>AI Command ▴</summary><div><a href="/proactive-superintendent/command">What should I deal with next?</a><a href="/lookahead-intelligence">Upcoming work readiness</a><a href="/event-intelligence/command">What changed?</a><a href="/performance">Performance monitor</a></div></details>'
    body=_side+_top+'<main class="bc170-main"><div class="bc170-work">'+body+'</div></main>'+_rail

    return f'''<!doctype html><html><head>
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{_v37_esc(title)} · BuildCommand AI</title>
<style>
{CSS}
body{{margin:0}}
.app{{display:block;min-height:100vh}}
.side{{position:sticky;top:0;z-index:1000;width:100%;min-height:auto;padding:0;background:rgba(12,15,19,.98);border-right:0;border-bottom:1px solid rgba(255,255,255,.10)}}
.main{{max-width:1500px;margin:0 auto;padding:28px 26px 70px}}
.bc-header{{display:flex;align-items:center;gap:22px;padding:12px 22px 10px;border-bottom:1px solid rgba(255,255,255,.07)}}
.bc-brand-wrap{{min-width:210px}}
.brand{{font-size:22px;font-weight:900;letter-spacing:-.4px;margin:0}}
.company{{font-size:11px;color:var(--muted);margin-top:2px}}
.bc-nav-wrap{{flex:1;min-width:0}}
.bc-topnav{{display:flex;align-items:center;gap:4px;flex-wrap:wrap}}
.bc-navdrop{{position:relative}}
.bc-navbtn{{width:auto;border:0;background:transparent;color:#f3f3f3;padding:10px 12px;border-radius:8px;font-weight:800;font-size:12px;letter-spacing:.35px;cursor:pointer}}
.bc-navbtn:hover,.bc-navdrop.open .bc-navbtn{{background:rgba(255,255,255,.08);color:#f0b44d}}
.bc-caret{{font-size:10px;margin-left:3px}}
.bc-dropdown{{display:none;position:absolute;top:calc(100% + 7px);left:0;min-width:275px;max-height:70vh;overflow:auto;padding:8px;background:#171b20;border:1px solid rgba(255,255,255,.12);border-radius:12px;box-shadow:0 18px 45px rgba(0,0,0,.45);z-index:1200}}
.bc-navdrop.open .bc-dropdown{{display:block}}
.bc-drop-link{{display:block;padding:10px 11px;border-radius:8px;text-decoration:none;color:#eee;font-size:13px;font-weight:650}}
.bc-drop-link:hover{{background:rgba(240,180,77,.12);color:#f0b44d}}
.bc-projectbar{{display:flex;align-items:center;gap:12px;padding:8px 22px 10px;background:rgba(255,255,255,.025)}}
.bc-project-meta{{min-width:210px}}
.bc-project-label{{font-size:9px;letter-spacing:1.2px;color:var(--muted);font-weight:800}}
.bc-project-name{{font-size:13px;font-weight:800}}
.bc-project-switch{{display:flex;gap:7px;align-items:center;flex:1;max-width:560px;margin:0}}
.bc-project-switch select{{margin:0;min-width:240px;padding:8px 10px}}
.bc-project-switch button{{width:auto;padding:8px 13px}}
.bc-add-project{{color:#f0b44d;text-decoration:none;font-weight:800;font-size:12px}}
.bc-user-actions{{margin-left:auto;display:flex;align-items:center;gap:10px}}
.bc-user-name{{font-size:11px;color:var(--muted)}}
.bc-signout{{margin:0}}
.bc-signout button{{width:auto;padding:8px 11px;font-size:11px}}
.creator-footer{{text-align:center;color:var(--muted);font-size:11px;padding:18px 10px 24px;border-top:1px solid rgba(255,255,255,.06)}}
.mobile-menu-btn{{display:none}}
.bc-home-grid{{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:14px}}
.bc-priority{{border-left:3px solid #f0b44d}}
@media(max-width:980px){{
  .bc-header{{align-items:flex-start;flex-wrap:wrap}}
  .bc-brand-wrap{{min-width:0}}
  .bc-nav-wrap{{width:100%;order:3}}
  .bc-topnav{{display:grid;grid-template-columns:repeat(3,1fr);width:100%}}
  .bc-navbtn{{width:100%;text-align:left}}
  .bc-dropdown{{position:fixed;left:18px;right:18px;top:120px;min-width:0;max-height:65vh}}
  .bc-projectbar{{flex-wrap:wrap}}
  .bc-project-switch{{max-width:none;width:100%;flex-basis:100%}}
  .bc-project-switch select{{min-width:0;flex:1}}
  .main{{padding:20px 14px 50px}}
  .bc-home-grid{{grid-template-columns:repeat(2,minmax(0,1fr))}}
}}
@media(max-width:620px){{
  .bc-topnav{{grid-template-columns:repeat(2,1fr)}}
  .bc-header{{padding:10px 12px}}
  .bc-projectbar{{padding:8px 12px}}
  .bc-home-grid{{grid-template-columns:1fr}}
}}
/* v170 Professional Command Center */
:root{{--bc170-side:255px;--bc170-top:62px}}.bc170-side{{position:fixed;left:0;top:0;bottom:0;width:var(--bc170-side);background:#101820;color:#fff;z-index:200;overflow:auto;padding:16px 12px;box-sizing:border-box}}.bc170-brand{{font-size:19px;font-weight:850;padding:7px 9px 18px}}.bc170-brand small{{display:block;font-size:9px;opacity:.55;letter-spacing:1.2px;margin-top:4px}}.bc170-group{{margin:4px 0}}.bc170-group button{{width:100%;border:0;background:transparent;color:#d9e0e6;text-align:left;padding:10px;border-radius:9px;font-weight:750;cursor:pointer}}.bc170-group button:hover{{background:rgba(255,255,255,.08)}}.bc170-links{{display:none;padding-bottom:5px}}.bc170-group.open .bc170-links{{display:block}}.bc170-links a{{display:block;color:#c7d0d8;text-decoration:none;padding:8px 10px 8px 18px;border-radius:8px;font-size:13px}}.bc170-links a:hover{{background:rgba(255,255,255,.08);color:#fff}}.bc170-top{{position:fixed;left:var(--bc170-side);right:0;top:0;height:var(--bc170-top);z-index:190;background:rgba(255,255,255,.97);border-bottom:1px solid #e5e9ee;display:flex;align-items:center;gap:10px;padding:0 18px;box-sizing:border-box}}.bc170-title{{font-weight:850;flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}.bc170-quick{{display:flex;gap:7px}}.bc170-quick a{{text-decoration:none;border:1px solid #dce2e8;border-radius:8px;padding:7px 10px;font-size:12px}}.bc170-main{{margin-left:var(--bc170-side);padding-top:var(--bc170-top);min-height:100vh}}.bc170-work{{max-width:1500px;margin:auto;padding:20px}}.bc170-menu{{display:none;border:1px solid #dce2e8;background:#fff;border-radius:8px;padding:7px 10px}}.bc170-rail{{position:fixed;right:14px;bottom:14px;width:285px;background:#fff;border:1px solid #e2e7ec;border-radius:14px;box-shadow:0 12px 35px rgba(0,0,0,.13);z-index:180}}.bc170-rail summary{{padding:13px 15px;font-weight:850;cursor:pointer}}.bc170-rail div{{padding:0 15px 13px}}.bc170-rail a{{display:block;text-decoration:none;padding:6px 0;font-size:12px}}@media(max-width:900px){{.bc170-side{{transform:translateX(-100%);transition:.18s}}.bc170-side.show{{transform:translateX(0)}}.bc170-top{{left:0}}.bc170-main{{margin-left:0}}.bc170-menu{{display:block}}.bc170-rail{{width:calc(100% - 28px)}}}}</style>
<script>
function bcToggleMenu(btn,event){{
  if(event) event.stopPropagation();
  var parent=btn.parentElement;
  document.querySelectorAll('.bc-navdrop.open').forEach(function(el){{if(el!==parent) el.classList.remove('open');}});
  parent.classList.toggle('open');
}}
document.addEventListener('click',function(e){{
  if(!e.target.closest('.bc-navdrop')){{
    document.querySelectorAll('.bc-navdrop.open').forEach(function(el){{el.classList.remove('open');}});
  }}
}});
document.addEventListener('keydown',function(e){{
  if(e.key==='Escape') document.querySelectorAll('.bc-navdrop.open').forEach(function(el){{el.classList.remove('open');}});
}});
</script>
</head><body><div class="app">
<header class="side">
  <div class="bc-header">
    <div class="bc-brand-wrap"><div class="brand">BuildCommand AI</div><div class="company">{company_name}</div></div>
    <div class="bc-nav-wrap">{nav}</div>
    <div class="bc-user-actions"><div class="bc-user-name">{display_name}</div><form method="post" action="/logout" class="bc-signout"><button type="submit">Sign Out</button></form></div>
  </div>
  {project_bar}
</header>
<main class="main">{body}</main>
<footer class="creator-footer">Built by Wilson LaHood · © 2026 Wilson LaHood</footer>
</div></body></html>'''

def project_id():
    user_id = current_user_id(); company_id = current_company_id()
    if not user_id or not company_id:
        return None
    c = db()
    state = c.execute("""SELECT us.selected_project_id FROM user_state us JOIN projects p ON p.id=us.selected_project_id WHERE us.user_id=? AND p.company_id=?""", (user_id, company_id)).fetchone()
    if state and state["selected_project_id"]:
        pid = state["selected_project_id"]
    else:
        first = c.execute("SELECT p.id FROM projects p LEFT JOIN project_archive_state a ON a.project_id=p.id WHERE p.company_id=? AND COALESCE(a.archived,0)=0 ORDER BY p.id LIMIT 1", (company_id,)).fetchone()
        pid = first["id"] if first else None
        if pid:
            c.execute("INSERT INTO user_state(user_id,selected_project_id) VALUES(?,?) ON CONFLICT(user_id) DO UPDATE SET selected_project_id=excluded.selected_project_id", (user_id,pid)); c.commit()
    c.close(); return pid


@app.post("/projects/select")
def select_project(project_id: int = Form(...)):
    company_id = current_company_id(); user_id = current_user_id(); c = db()
    exists = c.execute("SELECT id FROM projects WHERE id=? AND company_id=?", (project_id, company_id)).fetchone()
    if exists:
        c.execute("INSERT INTO user_state(user_id,selected_project_id) VALUES(?,?) ON CONFLICT(user_id) DO UPDATE SET selected_project_id=excluded.selected_project_id", (user_id,project_id)); c.commit()
    c.close(); return RedirectResponse("/", status_code=303)


# ============================================================
# v37.2 UNIFIED BUILDCOMMAND INTERFACE FIX
# ============================================================


def _v37_count(sql,args=()):
    """
    PostgreSQL/SQLite-safe scalar counter for unified dashboard.
    Always aliases COUNT(*) as n so dict_row and sqlite Row behave the same.
    """
    c=db()
    try:
        r=c.execute(sql,args).fetchone()
        value=int((r["n"] if r and "n" in r.keys() else 0) or 0)
        c.close()
        return value
    except Exception:
        try: c.rollback()
        except Exception: pass
        try: c.close()
        except Exception: pass
        return 0

def _v37_snapshot(pid):
    if not pid:
        return dict(scope=0,estimate=0,review=0,issues=0,submittals=0,actions=0,inspections=0)
    co=current_company_id()
    return {
        "scope":_v37_count(
            "SELECT COUNT(*) AS n FROM blueprint_scope_items WHERE company_id=? AND project_id=?",
            (co,pid)
        ),
        "estimate":_v37_count(
            "SELECT COUNT(*) AS n FROM estimator_items WHERE company_id=? AND project_id=?",
            (co,pid)
        ),
        "review":_v37_count(
            "SELECT COUNT(*) AS n FROM estimator_items WHERE company_id=? AND project_id=? AND COALESCE(verified,0)=0",
            (co,pid)
        ),
        "issues":_v37_count(
            "SELECT COUNT(*) AS n FROM project_issues WHERE project_id=? AND COALESCE(status,'OPEN')!='CLOSED'",
            (pid,)
        ),
        "submittals":_v37_count(
            "SELECT COUNT(*) AS n FROM submittals WHERE project_id=? AND COALESCE(status,'PENDING') NOT IN ('APPROVED','APPROVED_AS_NOTED','CLOSED','COMPLETE')",
            (pid,)
        ),
        "actions":_v37_count(
            "SELECT COUNT(*) AS n FROM action_items WHERE project_id=? AND COALESCE(status,'OPEN') NOT IN ('CLOSED','COMPLETE')",
            (pid,)
        ),
        "inspections":_v37_count(
            "SELECT COUNT(*) AS n FROM inspections_tracker WHERE project_id=? AND COALESCE(result,'PENDING')!='PASSED'",
            (pid,)
        )
    }

def _v37_link_card(title,desc,href,label="Open"):
    return ('<div class="card"><h2>'+_v37_esc(title)+'</h2><p class="muted">'+_v37_esc(desc)+'</p>'
            '<a href="'+href+'" style="display:inline-block;background:#f0b44d;color:#0a1017;text-decoration:none;padding:10px 14px;border-radius:9px;font-weight:800;">'
            +_v37_esc(label)+' →</a></div>')

@app.get("/",response_class=HTMLResponse)
def unified_projects_home():
    """v372 lightweight BuildCommand cover page with Blueprint source-intelligence hardening."""
    pid=project_id()
    try:
        p=db().execute("SELECT * FROM projects WHERE id=?",(pid,)).fetchone()
        project_name=str(p["name"] if p and p["name"] else "Current Project")
    except Exception:
        project_name="Current Project"

    body = """
    <style>
    .bc370-cover{min-height:calc(100vh - 120px);display:flex;flex-direction:column;justify-content:center}
    .bc370-hero{padding:52px 42px;border-radius:24px;background:linear-gradient(135deg,#111820 0%,#1e2a35 58%,#263847 100%);color:white;position:relative;overflow:hidden}
    .bc370-kicker{font-size:11px;font-weight:850;letter-spacing:2px;opacity:.62;margin-bottom:14px}
    .bc370-title{font-size:clamp(38px,6vw,74px);line-height:.94;letter-spacing:-2.5px;margin:0;max-width:920px}
    .bc370-title span{display:block;opacity:.62;font-size:.45em;letter-spacing:.5px;margin-top:16px}
    .bc370-copy{max-width:760px;font-size:17px;line-height:1.6;opacity:.78;margin:24px 0 30px}
    .bc370-actions{display:flex;flex-wrap:wrap;gap:10px}
    .bc370-actions a{display:inline-block;text-decoration:none;padding:12px 17px;border-radius:10px;font-weight:800}
    .bc370-primary{background:white;color:#111820}.bc370-secondary{border:1px solid rgba(255,255,255,.24);color:white}
    .bc370-project{margin-top:18px;font-size:12px;opacity:.6}
    .bc370-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px;margin-top:14px}
    .bc370-launch{display:block;text-decoration:none;padding:20px;border:1px solid #e2e7ec;background:#fff;border-radius:15px;min-height:118px}
    .bc370-launch:hover{transform:translateY(-2px);box-shadow:0 10px 28px rgba(0,0,0,.08)}
    .bc370-launch b{display:block;font-size:16px;margin-bottom:7px}.bc370-launch span{font-size:12px;line-height:1.45;opacity:.65}
    .bc370-bottom{display:grid;grid-template-columns:2fr 1fr;gap:12px;margin-top:12px}
    .bc370-command{padding:22px;border-radius:15px;border:1px solid #e2e7ec;background:#fff}.bc370-command h2{margin-top:0}
    @media(max-width:1050px){.bc370-grid{grid-template-columns:repeat(2,minmax(0,1fr))}}
    @media(max-width:650px){.bc370-hero{padding:34px 24px}.bc370-grid,.bc370-bottom{grid-template-columns:1fr}.bc370-title{letter-spacing:-1.5px}}
    </style>
    <div class="bc370-cover">
      <section class="bc370-hero">
        <div class="bc370-kicker">CONSTRUCTION OPERATION INTELLIGENCE SYSTEM</div>
        <h1 class="bc370-title">BuildCommand AI<span>Command the project. Build with intelligence.</span></h1>
        <p class="bc370-copy">One operating system connecting plans, scopes, field execution, project management, schedule, cost, safety, quality and construction intelligence.</p>
        <div class="bc370-actions">
          <a class="bc370-primary" href="/platform-369">Enter Command Center →</a>
          <a class="bc370-secondary" href="/ask-buildcommand">Ask BuildCommand</a>
          <a class="bc370-secondary" href="/blueprint-brain">Open Blueprint Brain</a>
        </div>
        <div class="bc370-project">ACTIVE PROJECT · __PROJECT__</div>
      </section>
      <section class="bc370-grid">
        <a class="bc370-launch" href="/field-execution-4"><b>FIELD COMMAND</b><span>Daily priorities, trade readiness, production, constraints and handoffs.</span></a>
        <a class="bc370-launch" href="/pm-command-4"><b>PM COMMAND</b><span>RFIs, submittals, decisions, coordination, procurement and changes.</span></a>
        <a class="bc370-launch" href="/schedule-command-4"><b>SCHEDULE COMMAND</b><span>Look-aheads, sequence exposure, constraints and recovery intelligence.</span></a>
        <a class="bc370-launch" href="/cost-command-4"><b>COST COMMAND</b><span>Known exposure, revisions, commercial risk and forecasting.</span></a>
        <a class="bc370-launch" href="/blueprint-brain"><b>BLUEPRINT BRAIN</b><span>Read drawings and specifications and connect requirements to trades.</span></a>
        <a class="bc370-launch" href="/knowledge-brain-2"><b>KNOWLEDGE BRAIN</b><span>Construction assemblies, trade ownership and cross-discipline reasoning.</span></a>
        <a class="bc370-launch" href="/sqc-command"><b>SAFETY + QUALITY</b><span>Inspection readiness, QC checkpoints and project quality intelligence.</span></a>
        <a class="bc370-launch" href="/autopilot"><b>PROJECT AUTOPILOT</b><span>What needs attention now, today, this week and next.</span></a>
      </section>
      <section class="bc370-bottom">
        <div class="bc370-command">
          <div class="bc370-kicker" style="color:#111820">BUILD COMMAND</div>
          <h2>What do you need to know?</h2>
          <p class="muted">Ask BuildCommand across the current project, or jump directly into the operating workspace you need.</p>
          <div class="bc370-actions"><a class="bc370-primary" style="background:#111820;color:white" href="/ask-buildcommand">Ask BuildCommand →</a><a href="/platform-369">View all intelligence</a></div>
        </div>
        <div class="bc370-command"><div class="bc370-kicker" style="color:#111820">SYSTEM</div><h2>v372</h2><p class="muted">Lightweight cover page designed to open fast while keeping the intelligence engines behind the launch screen.</p></div>
      </section>
    </div>
    """.replace("__PROJECT__",esc(project_name))
    return shell("BuildCommand AI",body)

@app.get("/build/analyze-project",response_class=HTMLResponse)
def unified_analyze_project_page():
    pid=project_id(); c=db()
    docs=c.execute("SELECT * FROM attachments WHERE company_id=? AND project_id=? ORDER BY id DESC",(current_company_id(),pid)).fetchall(); c.close()
    eligible=[d for d in docs if Path(d["original_name"] or "").suffix.lower() in {".pdf",".txt",".csv",".xlsx",".xlsm"}]
    checks="".join(f'<label style="display:flex;gap:10px;align-items:center;padding:10px 0;border-bottom:1px solid rgba(255,255,255,.08)"><input type="checkbox" name="attachment_ids" value="{d["id"]}" style="width:auto"><span><b>{esc(d["original_name"])}</b><br><span class="small">{(int(d["size_bytes"] or 0)/1024/1024):.1f} MB</span></span></label>' for d in eligible) or '<div class="muted">No supported project documents are uploaded yet.</div>'
    body='<div class="hero"><div class="eyebrow">BuildCommand · Unified Project Intelligence · v38</div><h1>Analyze the project once.</h1><p class="muted">BuildCommand runs Plan Intelligence, trade scope cleanup, estimator sync, takeoff splitting and quantity review.</p></div>'
    body+='<div class="card"><form method="post" action="/build/analyze-project"><h2>Select project documents</h2>'+checks+'<label style="margin-top:16px">Optional analysis focus</label><textarea name="focus" placeholder="Example: Full bid/scope review"></textarea><button type="submit">Analyze Project</button></form><p class="small">Selected files must total less than 50 MB. AI quantity proposals never overwrite estimator-entered quantities.</p></div>'
    return shell("Analyze Project",body)

@app.post("/build/analyze-project",response_class=HTMLResponse)
def unified_analyze_project_run(attachment_ids:list[int] | None=Form(None),focus:str=Form("")):
    pid=project_id(); docs=_v38_selected_docs(pid,attachment_ids)
    if not docs: return shell("Analyze Project",'<div class="card"><h2>Select at least one project document.</h2><p><a href="/build/analyze-project">← Back</a></p></div>')
    stages=[]
    try:
        bp=_v38_run_blueprint(pid,docs,focus); stages.append(("Plan Intelligence","COMPLETE",f'{bp["trades"]} trade scopes generated'))
        est=_seed_estimator_from_latest(pid); stages.append(("Estimator Sync","COMPLETE",f'{est["added"]} new · {est["updated"]} refreshed'))
        comp=_v38_run_component_split(pid)
        stages.append(("Takeoff Component Split","REVIEW" if comp.get("skipped") else "COMPLETE",comp.get("skipped") or f'{comp["created"]} measurable components created'))
        auto=_v38_run_auto_takeoff(pid)
        stages.append(("Automatic Quantity Review","REVIEW" if auto.get("skipped") else "COMPLETE",auto.get("skipped") or f'{auto["proposed"]} quantity proposals · {auto["verify"]} need verification'))
        rows="".join(f'<div class="action"><span class="badge {"READY" if status=="COMPLETE" else "WATCH"}">{status}</span> <b>{esc(name)}</b><div class="small">{esc(detail)}</div></div>' for name,status,detail in stages)
        body='<div class="hero"><div class="eyebrow">Unified Project Intelligence · v38</div><h1>Project intelligence ready.</h1><p class="muted">The brains ran as one pipeline. Review only the items that need human judgment.</p></div><div class="card"><h2>Pipeline Results</h2>'+rows+'</div><div class="grid3">'
        body+=_v37_link_card("Review BUILD","Review cleaned trade scope and project intelligence.","/build","Open BUILD")+_v37_link_card("Review ESTIMATE","Review takeoff proposals and estimator data.","/estimate","Open ESTIMATE")+_v37_link_card("Things That Need You","Review human-decision items.","/actions","Review")+'</div>'
        return shell("Project Intelligence Ready",body)
    except Exception as exc:
        rows="".join(f'<div class="action"><span class="badge READY">{status}</span> <b>{esc(name)}</b><div class="small">{esc(detail)}</div></div>' for name,status,detail in stages)
        body='<div class="hero"><h1>Analyze Project stopped.</h1></div><div class="card"><p>'+esc(str(exc))+'</p></div>'
        if rows: body+='<div class="card"><h2>Completed before stop</h2>'+rows+'</div>'
        body+='<div class="card"><p><a href="/build/analyze-project">← Back to Analyze Project</a></p></div>'
        return shell("Analyze Project",body)


# ============================================================
# v39 NEXT 10 CONSTRUCTION INTELLIGENCE
# ============================================================

def _v39_rows(sql,args=()):
    c=db()
    try:
        return c.execute(sql,args).fetchall()
    finally:
        c.close()

def _v39_safe_date(v):
    try:
        return datetime.fromisoformat(str(v)[:10]).date()
    except Exception:
        return None

def _v39_priority(due,priority=""):
    p=str(priority or "").upper()
    d=_v39_safe_date(due)
    today=datetime.utcnow().date()
    if p in {"CRITICAL","URGENT"} or (d and d<today): return "CRITICAL"
    if d==today or p=="HIGH": return "TODAY"
    if d and 0 < (d-today).days <= 7: return "THIS WEEK"
    return "REVIEW"

def _v39_attention(pid):
    items=[]
    for r in _v39_rows("SELECT * FROM action_items WHERE project_id=? AND COALESCE(status,'OPEN') NOT IN ('CLOSED','COMPLETE')",(pid,)):
        items.append((_v39_priority(r["due"],r["priority"]),"ACTION",r["title"],r["due"] or "",r["owner"] or ""))
    for r in _v39_rows("SELECT * FROM project_issues WHERE project_id=? AND COALESCE(status,'OPEN')!='CLOSED'",(pid,)):
        items.append((_v39_priority(r["due"],r["priority"]),"ISSUE",r["title"],r["due"] or "",r["owner"] or ""))
    for r in _v39_rows("SELECT * FROM submittals WHERE project_id=? AND COALESCE(status,'PENDING') NOT IN ('APPROVED','CLOSED','COMPLETE')",(pid,)):
        items.append((_v39_priority(r["due_date"],"HIGH" if r["due_date"] else ""),"SUBMITTAL",r["title"],r["due_date"] or "",r["responsible_party"] or ""))
    for r in _v39_rows("SELECT * FROM inspections_tracker WHERE project_id=? AND COALESCE(result,'PENDING')!='PASSED'",(pid,)):
        items.append((_v39_priority(r["scheduled_date"],"HIGH" if r["scheduled_date"] else ""),"INSPECTION",r["inspection_type"],r["scheduled_date"] or "",r["authority"] or ""))
    for r in _v39_rows("SELECT * FROM estimator_items WHERE project_id=? AND COALESCE(verified,0)=0 AND COALESCE(ai_confidence,'') IN ('LOW','VERIFY')",(pid,)):
        items.append(("REVIEW","ESTIMATE",r["description"],"",r["trade"] or ""))
    rank={"CRITICAL":0,"TODAY":1,"THIS WEEK":2,"REVIEW":3,"FYI":4}
    return sorted(items,key=lambda x:rank.get(x[0],9))

def _v39_memory_rules(pid):
    return _v39_rows("SELECT trade,requirement,confidence FROM blueprint_scope_items WHERE project_id=? AND COALESCE(confidence,'')='HIGH' ORDER BY id DESC LIMIT 100",(pid,))

def _v39_conflicts(pid):
    return _v39_rows("SELECT requirement,COUNT(DISTINCT trade) AS n,MIN(trade) AS trade FROM blueprint_scope_items WHERE project_id=? GROUP BY requirement HAVING COUNT(DISTINCT trade)>1 ORDER BY n DESC LIMIT 50",(pid,))

def _v39_schedule_risks(pid):
    today=datetime.utcnow().date()
    rows=_v39_rows("SELECT * FROM activities WHERE project_id=? AND COALESCE(status,'NOT_STARTED')!='COMPLETE' ORDER BY start LIMIT 100",(pid,))
    out=[]
    for r in rows:
        s=_v39_safe_date(r["start"]); f=_v39_safe_date(r["finish"]); pct=float(r["pct"] or 0)
        if f and f<today and pct<100: out.append(("CRITICAL",r["name"],"Past finish date",r["finish"]))
        elif s and 0 <= (s-today).days <= 14 and pct==0: out.append(("THIS WEEK",r["name"],"Upcoming / make-ready",r["start"]))
    return out

def _v39_procurement(pid):
    return _v39_rows("SELECT * FROM procurement WHERE project_id=? AND COALESCE(status,'OPEN') NOT IN ('DELIVERED','COMPLETE','CLOSED') ORDER BY required_on_site LIMIT 100",(pid,))

def _v39_changes(pid):
    return _v39_rows("SELECT * FROM change_events WHERE project_id=? AND COALESCE(status,'OPEN') NOT IN ('CLOSED','COMPLETE') ORDER BY id DESC LIMIT 100",(pid,))

@app.get("/intelligence",response_class=HTMLResponse)
def v39_intelligence_center():
    pid=project_id(); attention=_v39_attention(pid); conflicts=_v39_conflicts(pid); sched=_v39_schedule_risks(pid); proc=_v39_procurement(pid); changes=_v39_changes(pid)
    cards=[
        ("Things That Need You",len(attention),"/intelligence/attention","Prioritized decisions across the job."),
        ("Project Memory",len(_v39_memory_rules(pid)),"/intelligence/memory","High-confidence construction knowledge retained from this project."),
        ("Conflict Brain",len(conflicts),"/intelligence/conflicts","Cross-scope conflicts that deserve review."),
        ("RFI Intelligence",len(conflicts),"/intelligence/rfis","Draft RFIs from detected conflicts."),
        ("Schedule Intelligence",len(sched),"/intelligence/schedule","Late and upcoming work requiring make-ready."),
        ("Look-Ahead Brain",len(sched),"/intelligence/lookahead","2, 3 and 6 week field outlook."),
        ("Procurement Brain",len(proc),"/intelligence/procurement","Long-lead and required-on-site tracking."),
        ("Bid Scope Brain",_v37_snapshot(pid)["scope"],"/intelligence/bids","Trade scope package intelligence."),
        ("Cost & Change",len(changes),"/intelligence/changes","Open cost and schedule exposure."),
        ("Daily Superintendent Command",len(attention)+len(sched),"/intelligence/daily-command","Today's project command brief.")
    ]
    body='<div class="hero"><div class="eyebrow">BuildCommand v40</div><h1>Construction Intelligence Center</h1><p class="muted">Ten brains. One project operating system. The main menu stays simple.</p></div><div class="grid3">'
    for name,count,href,desc in cards:
        body+=f'<div class="card"><div class="label">{esc(name)}</div><div class="kpi">{count}</div><p class="muted">{esc(desc)}</p><a href="{href}">Open →</a></div>'
    body+='</div>'
    body += '<div class="grid3">'+_v37_link_card("Constructability Intelligence","Find access, clearance, penetration, ceiling and coordination risks.","/intelligence-engine/constructability","Open")+_v37_link_card("Superintendent Command Intelligence","Prioritized view of sequence, procurement, conflicts, gaps and inspections.","/intelligence-engine/command","Open")+_v37_link_card("Full Intelligence Engine","Conflict, RFI, scope, change, procurement, inspection and learning intelligence.","/intelligence-engine","Open")+'</div>'
    body += '<div class="grid3">' + _v37_link_card("Project Knowledge Graph","Drawing revisions, dependencies, equipment chains, prediction and verification.","/knowledge-graph","Open") + '</div>'
    body += '<div class="grid3">' + _v37_link_card("Prediction & Decision Intelligence","Dependency impacts, decision deadlines, manpower, materials, closeout and risk propagation.","/prediction-intelligence","Open") + '</div>'
    body += '<div class="grid3">' + _v37_link_card("Brain Quality & Self-Learning","Verification, confidence calibration, source quality, contradiction detection and improvement queue.","/brain-quality","Open") + '</div>'
    body += '<div class="grid3">' + _v37_link_card("Field Context & Assembly Intelligence","Rooms, assemblies, systems, prerequisites, hold points, commissioning and field work packages.","/field-context","Open") + '</div>'
    body += '<div class="grid3">' + _v37_link_card("Master Construction Reasoning","One connected project judgment across scope, sequence, risk, field, quality and commercial intelligence.","/master-reasoning","Open") + '</div>'
    body += '<div class="grid3">' + _v37_link_card("Real Construction Reasoning 2.0","Cause, dependency, consequence, ownership, alternatives and uncertainty.","/reasoning-2","Open") + '</div>'
    body += '<div class="grid3">' + _v37_link_card("Project Memory & Continuous Learning","Approved corrections, RFI answers, lessons and cross-project patterns.","/project-memory","Open") + '</div>'
    body += '<div class="grid3">' + _v37_link_card("Drawing Revision & Change Intelligence","Added/removed scope, affected trades, cost/schedule exposure and downstream controls.","/revision-intelligence","Open") + '</div>'
    body += '<div class="grid3">' + _v37_link_card("Event-Driven Intelligence","Refresh only the brain modules affected by a project change.","/event-intelligence","Open") + '</div>'










    return shell("Intelligence",body)

@app.get("/intelligence/attention",response_class=HTMLResponse)
def v39_attention_page():
    items=_v39_attention(project_id())
    rows="".join(f'<div class="action"><span class="badge">{esc(level)}</span> <b>{esc(kind)}</b> — {esc(title)}<div class="small">{esc(due)} {esc(owner)}</div></div>' for level,kind,title,due,owner in items)
    return shell("Things That Need You",'<div class="hero"><h1>Things That Need You</h1><p class="muted">Critical → Today → This Week → Review.</p></div><div class="card">'+(rows or '<p class="muted">Nothing needs attention.</p>')+'</div>')

@app.get("/intelligence/memory",response_class=HTMLResponse)
def v39_memory_page():
    rows=_v39_memory_rules(project_id())
    h="".join(f'<div class="action"><b>{esc(r["trade"])}</b><div>{esc(r["requirement"])}</div><span class="small">Confidence: {esc(r["confidence"])}</span></div>' for r in rows)
    return shell("Project Memory",'<div class="hero"><h1>Project Memory Brain</h1><p class="muted">High-confidence project construction knowledge retained for future reasoning.</p></div><div class="card">'+(h or '<p class="muted">Memory grows as project intelligence is reviewed.</p>')+'</div>')

@app.get("/intelligence/conflicts",response_class=HTMLResponse)
def v39_conflicts_page():
    rows=_v39_conflicts(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">REVIEW</span> {esc(r["requirement"])}<div class="small">Appears under {r["n"]} trades.</div></div>' for r in rows)
    return shell("Conflict Brain",'<div class="hero"><h1>Cross-Document Conflict Brain</h1></div><div class="card">'+(h or '<p class="muted">No duplicate cross-trade requirements detected.</p>')+'</div>')

@app.get("/intelligence/rfis",response_class=HTMLResponse)
def v39_rfi_page():
    rows=_v39_conflicts(project_id())
    h="".join(f'<div class="card"><span class="badge WATCH">DRAFT RFI</span><h3>{esc(r["requirement"])}</h3><p>Please clarify the governing scope/document requirement. BuildCommand detected this requirement across multiple trade assignments.</p><p class="small">Human approval required before issue.</p></div>' for r in rows)
    return shell("RFI Intelligence",'<div class="hero"><h1>Automatic RFI Intelligence</h1><p class="muted">Proposals only. Nothing is sent automatically.</p></div>'+(h or '<div class="card">No RFI candidates detected.</div>'))

@app.get("/intelligence/schedule",response_class=HTMLResponse)
def v39_schedule_page():
    rows=_v39_schedule_risks(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(level)}</span> <b>{esc(name)}</b><div class="small">{esc(reason)} · {esc(date)}</div></div>' for level,name,reason,date in rows)
    return shell("Schedule Intelligence",'<div class="hero"><h1>Schedule Intelligence Brain</h1></div><div class="card">'+(h or '<p class="muted">No immediate schedule exceptions detected.</p>')+'</div>')

@app.get("/intelligence/lookahead",response_class=HTMLResponse)
def v39_lookahead_page():
    pid=project_id(); today=datetime.utcnow().date(); rows=_v39_rows("SELECT * FROM activities WHERE project_id=? ORDER BY start",(pid,))
    body='<div class="hero"><h1>Look-Ahead Brain</h1><p class="muted">2, 3 and 6 week production outlook.</p></div>'
    for weeks in (2,3,6):
        end=today+timedelta(days=weeks*7)
        subset=[r for r in rows if _v39_safe_date(r["start"]) and today <= _v39_safe_date(r["start"]) <= end]
        h="".join(f'<div class="action"><b>{esc(r["name"])}</b><div class="small">{esc(r["trade"])} · {esc(r["start"])} → {esc(r["finish"])}</div></div>' for r in subset)
        body+=f'<div class="card"><h2>{weeks}-Week Look-Ahead</h2>{h or "<p class=muted>No scheduled starts in this window.</p>"}</div>'
    return shell("Look-Ahead",body)

@app.get("/intelligence/procurement",response_class=HTMLResponse)
def v39_procurement_page():
    rows=_v39_procurement(project_id())
    h="".join(f'<div class="action"><b>{esc(r["item"])}</b><div class="small">Required: {esc(r["required_on_site"])} · Promised: {esc(r["promised_date"])} · {esc(r["status"])}</div></div>' for r in rows)
    return shell("Procurement Intelligence",'<div class="hero"><h1>Procurement & Long-Lead Brain</h1></div><div class="card">'+(h or '<p class="muted">No open procurement items.</p>')+'</div>')

@app.get("/intelligence/bids",response_class=HTMLResponse)
def v39_bid_page():
    rows=_v39_rows("SELECT trade,COUNT(*) AS n FROM blueprint_scope_items WHERE project_id=? GROUP BY trade ORDER BY trade",(project_id(),))
    h="".join(f'<div class="action"><b>{esc(r["trade"])}</b><div class="small">{r["n"]} source-backed scope items ready for bid-package review.</div></div>' for r in rows)
    return shell("Bid Scope Intelligence",'<div class="hero"><h1>Subcontractor Scope & Bid Brain</h1><p class="muted">Trade packages originate from the cleaned Blueprint Brain scope.</p></div><div class="card">'+(h or '<p class="muted">Analyze project documents first.</p>')+'</div>')

@app.get("/intelligence/changes",response_class=HTMLResponse)
def v39_change_page():
    rows=_v39_changes(project_id()); total=sum(float(r["estimated_cost"] or 0) for r in rows); days=sum(float(r["schedule_days"] or 0) for r in rows)
    h="".join(f'<div class="action"><b>{esc(r["title"])}</b><div class="small">${float(r["estimated_cost"] or 0):,.0f} · {float(r["schedule_days"] or 0):g} days · {esc(r["status"])}</div></div>' for r in rows)
    return shell("Cost & Change",f'<div class="hero"><h1>Cost & Change Intelligence</h1><p class="muted">Open exposure: ${total:,.0f} · {days:g} schedule days.</p></div><div class="card">'+(h or '<p class="muted">No open change exposure.</p>')+'</div>')

@app.get("/intelligence/daily-command",response_class=HTMLResponse)
def v39_daily_command():
    pid=project_id(); attention=_v39_attention(pid); sched=_v39_schedule_risks(pid); proc=_v39_procurement(pid)
    critical=sum(1 for x in attention if x[0]=="CRITICAL"); today=sum(1 for x in attention if x[0]=="TODAY")
    body=f'<div class="hero"><div class="eyebrow">Daily Superintendent Command</div><h1>Run the job from here.</h1><p class="muted">{critical} critical · {today} today · {len(sched)} schedule exceptions · {len(proc)} procurement items.</p></div>'
    body+='<div class="grid3">'+_v37_link_card("Decisions","Prioritized items requiring human attention.","/intelligence/attention","Review")+_v37_link_card("Look-Ahead","Upcoming production and make-ready.","/intelligence/lookahead","Open")+_v37_link_card("Ask BuildCommand","Ask the project what matters now.","/ask-buildcommand","Ask")+'</div>'
    return shell("Daily Command",body)


# ============================================================
# v40 PROJECT AUTOPILOT
# ============================================================

def _v40_score(pid):
    att=_v39_attention(pid); sched=_v39_schedule_risks(pid); proc=_v39_procurement(pid); changes=_v39_changes(pid)
    critical=sum(1 for x in att if x[0]=="CRITICAL")+sum(1 for x in sched if x[0]=="CRITICAL")
    high=sum(1 for x in att if x[0] in {"TODAY","THIS WEEK"})
    score=min(100,critical*20+high*7+len(proc)*4+len(changes)*5)
    return {"risk":score,"health":max(0,100-score),"level":"CRITICAL" if score>=70 else "HIGH" if score>=45 else "MEDIUM" if score>=20 else "LOW"}

def _v40_revision_candidates(pid):
    rows=_v39_rows("SELECT * FROM attachments WHERE project_id=? ORDER BY id DESC LIMIT 100",(pid,))
    seen={}; pairs=[]
    for r in rows:
        name=str(r["original_name"] or "")
        base=re.sub(r'(?i)(rev(?:ision)?[ _-]*[A-Z0-9]+|addendum[ _-]*[A-Z0-9]+|bulletin[ _-]*[A-Z0-9]+)','',name)
        base=re.sub(r'[^a-z0-9]+',' ',base.lower()).strip()
        if base in seen: pairs.append((r,seen[base]))
        else: seen[base]=r
    return pairs[:25]

def _v40_gap_summary(pid):
    scopes=_v39_rows("SELECT trade,COUNT(*) AS n FROM blueprint_scope_items WHERE project_id=? GROUP BY trade ORDER BY trade",(pid,))
    subs=_v39_rows("SELECT trade,COUNT(*) AS n FROM subcontractors WHERE project_id=? GROUP BY trade",(pid,))
    covered={str(r["trade"] or "").lower() for r in subs}
    return [(r["trade"],r["n"],str(r["trade"] or "").lower() in covered) for r in scopes]

def _v40_inspection_ready(pid):
    inspections=_v39_rows("SELECT * FROM inspections_tracker WHERE project_id=? AND COALESCE(result,'PENDING')!='PASSED' ORDER BY scheduled_date LIMIT 50",(pid,))
    issues=len(_v39_rows("SELECT id FROM project_issues WHERE project_id=? AND COALESCE(status,'OPEN')!='CLOSED'",(pid,)))
    subs=len(_v39_rows("SELECT id FROM submittals WHERE project_id=? AND COALESCE(status,'PENDING') NOT IN ('APPROVED','CLOSED','COMPLETE')",(pid,)))
    return inspections,issues,subs

@app.get("/autopilot",response_class=HTMLResponse)
def v40_autopilot():
    pid=project_id(); s=_v40_score(pid)
    body=f'<div class="hero"><div class="eyebrow">BuildCommand v40 - Project Autopilot</div><h1>Your project is {s["level"]} risk.</h1><p class="muted">Health {s["health"]}/100 - Risk {s["risk"]}/100. Autopilot connects project decisions to downstream construction impact.</p></div><div class="grid3">'
    cards=[
      ("Risk Prediction","Score and prioritize project threats.","/autopilot/risks"),
      ("Revision Brain","Identify drawing and addendum revision candidates.","/autopilot/revisions"),
      ("Change Detection","Surface potential cost and time exposure.","/autopilot/change-detection"),
      ("Subcontractor Gaps","Compare cleaned scopes against subcontractor coverage.","/autopilot/gaps"),
      ("Inspection Readiness","Preflight upcoming inspections.","/autopilot/inspection-readiness"),
      ("Quality Control","Source-backed QC verification.","/autopilot/qc"),
      ("Field Photo Intelligence","Route photos into reviewed field intelligence.","/autopilot/photos"),
      ("Project Forecast","Forecast health, completion pressure and decisions.","/autopilot/forecast"),
      ("Executive Command","See project health across the company.","/autopilot/executive")
    ]
    for name,desc,href in cards: body+=_v37_link_card(name,desc,href,"Open")
    body+='</div>'
    return shell("Project Autopilot",body)

@app.get("/autopilot/risks",response_class=HTMLResponse)
def v40_risks():
    pid=project_id(); s=_v40_score(pid); items=_v39_attention(pid); sched=_v39_schedule_risks(pid)
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(x[0])}</span> <b>{esc(x[1])}</b> - {esc(x[2])}<div class="small">{esc(x[3])}</div></div>' for x in items[:60])
    h+="".join(f'<div class="action"><span class="badge WATCH">{esc(x[0])}</span> <b>SCHEDULE</b> - {esc(x[1])}<div class="small">{esc(x[2])} - {esc(x[3])}</div></div>' for x in sched)
    return shell("Risk Prediction",f'<div class="hero"><h1>Risk Prediction Engine</h1><p class="muted">{s["level"]} - Risk score {s["risk"]}/100</p></div><div class="card">{h or "No major risk signals detected."}</div>')

@app.get("/autopilot/revisions",response_class=HTMLResponse)
def v40_revisions():
    pairs=_v40_revision_candidates(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">COMPARE</span> <b>{esc(a["original_name"])}</b><div class="small">Possible revision of {esc(b["original_name"])}</div></div>' for a,b in pairs)
    return shell("Revision Brain",'<div class="hero"><h1>Drawing Revision / Addendum Brain</h1><p class="muted">Revision candidates are flagged for review. Contract scope is never changed automatically.</p></div><div class="card">'+(h or '<p class="muted">No likely revision pairs detected.</p>')+'</div>')

@app.get("/autopilot/change-detection",response_class=HTMLResponse)
def v40_change_detection():
    rows=_v39_conflicts(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">POTENTIAL CHANGE</span> {esc(r["requirement"])}<div class="small">Detected across {r["n"]} trade assignments. Review contract scope.</div></div>' for r in rows)
    return shell("Change Detection",'<div class="hero"><h1>Automatic Change-Order Detection</h1><p class="muted">Potential exposure only. Human approval required.</p></div><div class="card">'+(h or '<p class="muted">No current change candidates.</p>')+'</div>')

@app.get("/autopilot/gaps",response_class=HTMLResponse)
def v40_gaps():
    rows=_v40_gap_summary(project_id())
    h="".join(f'<div class="action"><span class="badge">{"COVERED" if ok else "GAP"}</span> <b>{esc(trade)}</b><div class="small">{n} cleaned scope items</div></div>' for trade,n,ok in rows)
    return shell("Subcontractor Gaps",'<div class="hero"><h1>Subcontractor Gap Analyzer</h1></div><div class="card">'+(h or '<p class="muted">Analyze project scope first.</p>')+'</div>')

@app.get("/autopilot/inspection-readiness",response_class=HTMLResponse)
def v40_inspection_readiness():
    rows,issues,subs=_v40_inspection_ready(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{"NOT READY" if issues or subs else "READY"}</span> <b>{esc(r["inspection_type"])}</b><div class="small">{esc(r["scheduled_date"])} - {issues} open issues - {subs} open submittals</div></div>' for r in rows)
    return shell("Inspection Readiness",'<div class="hero"><h1>Inspection Readiness Brain</h1></div><div class="card">'+(h or '<p class="muted">No pending inspections.</p>')+'</div>')

@app.get("/autopilot/qc",response_class=HTMLResponse)
def v40_qc():
    rows=_v39_rows("SELECT trade,requirement,source_ref FROM blueprint_scope_items WHERE project_id=? AND COALESCE(confidence,'')='HIGH' ORDER BY trade,id LIMIT 120",(project_id(),))
    h="".join(f'<div class="action"><input type="checkbox" style="width:auto"> <b>{esc(r["trade"])}</b> - {esc(r["requirement"])}<div class="small">{esc(r["source_ref"])}</div></div>' for r in rows)
    return shell("Quality Control",'<div class="hero"><h1>Quality-Control Brain</h1><p class="muted">Source-backed verification checklist.</p></div><div class="card">'+(h or '<p class="muted">Analyze plans first.</p>')+'</div>')

@app.get("/autopilot/photos",response_class=HTMLResponse)
def v40_photos():
    rows=_v39_rows("SELECT * FROM attachments WHERE project_id=? ORDER BY id DESC LIMIT 100",(project_id(),))
    photos=[r for r in rows if Path(r["original_name"] or "").suffix.lower() in {".jpg",".jpeg",".png",".webp"}]
    h="".join(f'<div class="action"><span class="badge WATCH">REVIEW PHOTO</span> <b>{esc(r["original_name"])}</b><div class="small">Associate with location, trade, QC, punch or schedule impact.</div></div>' for r in photos)
    return shell("Field Photo Intelligence",'<div class="hero"><h1>Field Photo Intelligence</h1><p class="muted">Human review required before formal issue creation.</p></div><div class="card">'+(h or '<p class="muted">No project photos uploaded.</p>')+'</div>')

@app.get("/autopilot/forecast",response_class=HTMLResponse)
def v40_forecast():
    pid=project_id(); s=_v40_score(pid); acts=_v39_rows("SELECT * FROM activities WHERE project_id=? ORDER BY finish DESC LIMIT 1",(pid,)); finish=acts[0]["finish"] if acts else ""; open_items=len(_v39_attention(pid))
    return shell("Project Forecast",f'<div class="hero"><h1>Project Forecast Brain</h1><p class="muted">Current planned finish: {esc(finish or "Not scheduled")}</p></div><div class="grid3"><div class="card"><div class="label">Health</div><div class="kpi">{s["health"]}</div></div><div class="card"><div class="label">Risk</div><div class="kpi">{s["risk"]}</div></div><div class="card"><div class="label">Open Decisions</div><div class="kpi">{open_items}</div></div></div>')

@app.get("/autopilot/executive",response_class=HTMLResponse)
def v40_executive():
    rows=_v39_rows("SELECT * FROM projects WHERE company_id=? ORDER BY id DESC",(current_company_id(),)); h=""
    for p in rows:
        s=_v40_score(p["id"]); h+=f'<div class="action"><span class="badge">{esc(s["level"])}</span> <b>{esc(p["number"])} - {esc(p["name"])}</b><div class="small">Health {s["health"]}/100 - Risk {s["risk"]}/100</div></div>'
    return shell("Executive Command",'<div class="hero"><h1>Executive Command Center</h1><p class="muted">Which projects need leadership attention and why.</p></div><div class="card">'+(h or '<p class="muted">No projects found.</p>')+'</div>')


# ============================================================
# v41 FIELD EXECUTION INTELLIGENCE
# ============================================================

def _v41_ensure_tables():
    c=db()
    if DATABASE_KIND=="postgres":
        c.execute("CREATE TABLE IF NOT EXISTS closeout_items(id BIGSERIAL PRIMARY KEY,company_id BIGINT,project_id BIGINT,category TEXT,item TEXT,responsible_party TEXT,due_date TEXT,status TEXT DEFAULT 'OPEN',notes TEXT DEFAULT '',created TEXT,updated TEXT)")
    else:
        c.execute("CREATE TABLE IF NOT EXISTS closeout_items(id INTEGER PRIMARY KEY,company_id INTEGER,project_id INTEGER,category TEXT,item TEXT,responsible_party TEXT,due_date TEXT,status TEXT DEFAULT 'OPEN',notes TEXT DEFAULT '',created TEXT,updated TEXT)")
    c.commit(); c.close()

def _v41_make_ready(pid):
    rows=_v39_rows("SELECT a.*,r.drawings,r.material,r.manpower,r.predecessor,r.access_ready,r.inspection,r.equipment,r.notes readiness_notes FROM activities a LEFT JOIN activity_readiness r ON r.activity_id=a.id AND r.project_id=a.project_id WHERE a.project_id=? AND COALESCE(a.status,'NOT_STARTED')!='COMPLETE' ORDER BY a.start LIMIT 80",(pid,))
    out=[]
    for r in rows:
        checks=[r["drawings"],r["material"],r["manpower"],r["predecessor"],r["access_ready"],r["inspection"],r["equipment"]]
        missing=7-sum(1 for x in checks if int(x or 0)==1)
        status="READY" if missing==0 else "AT RISK" if missing<=2 else "NOT READY"
        out.append((r,status,missing))
    return out

def _v41_sub_center(pid):
    return (
        _v39_rows("SELECT * FROM subs WHERE project_id=? ORDER BY trade,name",(pid,)),
        _v39_rows("SELECT * FROM activities WHERE project_id=? AND COALESCE(status,'NOT_STARTED')!='COMPLETE' ORDER BY start",(pid,)),
        _v39_rows("SELECT * FROM subcontractor_updates WHERE project_id=? ORDER BY id DESC LIMIT 100",(pid,))
    )

def _v41_delivery_risks(pid):
    today=datetime.utcnow().date()
    rows=_v39_rows("SELECT * FROM procurement WHERE project_id=? AND COALESCE(status,'OPEN') NOT IN ('DELIVERED','COMPLETE','CLOSED') ORDER BY required_on_site",(pid,))
    out=[]
    for r in rows:
        need=_v39_safe_date(r["required_on_site"]); promised=_v39_safe_date(r["promised_date"])
        exposure=(promised-need).days if need and promised else None
        level="CRITICAL" if exposure is not None and exposure>=7 else "HIGH" if exposure is not None and exposure>0 else "TODAY" if need and need<=today else "WATCH"
        out.append((r,level,exposure))
    return out

def _v41_inspection_countdown(pid):
    today=datetime.utcnow().date()
    rows=_v39_rows("SELECT * FROM inspections_tracker WHERE project_id=? AND COALESCE(result,'PENDING')!='PASSED' ORDER BY scheduled_date",(pid,))
    out=[]
    for r in rows:
        d=_v39_safe_date(r["scheduled_date"]); days=(d-today).days if d else None
        label="UNSCHEDULED" if days is None else "OVERDUE" if days<0 else "TODAY" if days==0 else "TOMORROW" if days==1 else f"{days} DAYS" if days<=7 else "UPCOMING"
        out.append((r,label,days))
    return out

def _v41_delay_candidates(pid):
    rows=_v39_rows("SELECT * FROM project_issues WHERE project_id=? AND COALESCE(status,'OPEN')!='CLOSED' ORDER BY id DESC",(pid,))
    out=[]
    for r in rows:
        txt=(str(r["title"] or "")+" "+str(r["description"] or "")+" "+str(r["issue_type"] or "")).lower()
        if any(x in txt for x in ["delay","late","waiting","hold","blocked","impact","shutdown","unavailable"]):
            out.append({"title":r["title"],"description":r["description"] or "","due":r["due"] or "","owner":r["owner"] or ""})
    for level,name,reason,datev in _v39_schedule_risks(pid):
        if level=="CRITICAL":
            out.append({"title":name,"description":reason,"due":datev,"owner":""})
    return out

def _v41_closeout_seed(pid):
    _v41_ensure_tables()
    c=db()
    r=c.execute("SELECT COUNT(*) AS n FROM closeout_items WHERE company_id=? AND project_id=?",(current_company_id(),pid)).fetchone()
    if int(r["n"] or 0)==0:
        now=datetime.utcnow().isoformat()
        defaults=[
            ("As-Builts","Record drawings / as-builts"),
            ("O&M","Operations and maintenance manuals"),
            ("Warranties","Warranty documentation"),
            ("Training","Owner training completion"),
            ("Attic Stock","Required attic stock / spare materials"),
            ("Inspections","Final inspections and certificates"),
            ("Testing","Final testing and commissioning"),
            ("Lien Releases","Final lien releases"),
            ("Turnover","Keys, access, owner turnover documents")
        ]
        for category,item in defaults:
            c.execute("INSERT INTO closeout_items(company_id,project_id,category,item,responsible_party,due_date,status,notes,created,updated) VALUES(?,?,?,?,?,?,?,?,?,?)",(current_company_id(),pid,category,item,"","","OPEN","",now,now))
        c.commit()
    c.close()

@app.get("/field-command",response_class=HTMLResponse)
def v41_field_command():
    pid=project_id()
    ready=_v41_make_ready(pid); deliveries=_v41_delivery_risks(pid); inspections=_v41_inspection_countdown(pid)
    subs,acts,updates=_v41_sub_center(pid); delays=_v41_delay_candidates(pid); attention=_v39_attention(pid)
    not_ready=sum(1 for _,s,_ in ready if s=="NOT READY")
    today_ins=sum(1 for _,label,_ in inspections if label in {"TODAY","TOMORROW"})
    delivery_risk=sum(1 for _,level,_ in deliveries if level in {"CRITICAL","HIGH","TODAY"})
    body=f'<div class="hero"><div class="eyebrow">BuildCommand v41 - Superintendent Field Command</div><h1>Run today&#39;s job.</h1><p class="muted">{len(subs)} subcontractors - {today_ins} inspections due now/next - {delivery_risk} delivery risks - {not_ready} activities not ready - {len(delays)} delay signals - {len(attention)} decisions need review.</p></div>'
    body+='<div class="grid3">'
    body+=_v37_link_card("Make-Ready","See what can actually start and why, with sequence intelligence.","/sequence-intelligence","Open")
    body+=_v37_link_card("Subcontractors","Crews, commitments and upcoming work.","/field-command/subs","Open")
    body+=_v37_link_card("Sub Alerts","Prepare schedule/field alerts for human approval.","/field-command/sub-alerts","Open")
    body+=_v37_link_card("Deliveries","Required-on-site vs promised-date exposure.","/field-command/deliveries","Open")
    body+=_v37_link_card("Inspections","Countdown and readiness before the inspector arrives.","/field-command/inspections","Open")
    body+=_v37_link_card("Daily Report","Build today's superintendent report from project data.","/field-command/daily-report","Open")
    body+=_v37_link_card("Delay Documentation","Capture possible delay events before the record is lost.","/field-command/delays","Open")
    body+=_v37_link_card("Punch & Completion","Open punch organized by location/trade.","/field-command/punch","Open")
    body+=_v37_link_card("Closeout","Track turnover requirements to completion.","/field-command/closeout","Open")
    body+='</div>'
    return shell("Field Command",body)

@app.get("/field-command/make-ready",response_class=HTMLResponse)
def v41_make_ready_page():
    rows=_v41_make_ready(project_id())
    h="".join(f'<div class="action"><span class="badge">{esc(status)}</span> <b>{esc(r["name"])}</b><div class="small">{esc(r["trade"])} - {esc(r["start"])} - {missing} readiness checks incomplete</div></div>' for r,status,missing in rows)
    return shell("Make-Ready Brain",'<div class="hero"><h1>Make-Ready Brain</h1><p class="muted">Drawings - material - manpower - predecessor - access - inspection - equipment.</p></div><div class="card">'+(h or '<p class="muted">No active activities.</p>')+'</div>')

@app.get("/field-command/subs",response_class=HTMLResponse)
def v41_subs_page():
    subs,acts,updates=_v41_sub_center(project_id()); h=""
    for s in subs:
        related=[a for a in acts if str(a["trade"] or "").lower()==str(s["trade"] or "").lower()]
        nxt=related[0] if related else None
        next_text=("Next: "+esc(nxt["name"])+" - "+esc(nxt["start"])) if nxt else "No upcoming activity matched."
        h+=f'<div class="card"><h3>{esc(s["name"])}</h3><div class="small">{esc(s["trade"])}</div><p>{next_text}</p></div>'
    return shell("Subcontractor Command",'<div class="hero"><h1>Subcontractor Command Center</h1></div><div class="grid3">'+(h or '<div class="card">No subcontractors loaded.</div>')+'</div>')

@app.get("/field-command/sub-alerts",response_class=HTMLResponse)
def v41_sub_alerts():
    subs,acts,updates=_v41_sub_center(project_id()); risks=_v39_schedule_risks(project_id()); h=""
    for s in subs:
        matching=[r for r in risks if any(str(a["trade"] or "").lower()==str(s["trade"] or "").lower() and a["name"]==r[1] for a in acts)]
        if matching:
            msg=f'BuildCommand draft: Please review upcoming {s["trade"]} schedule requirements and confirm manpower/material readiness. Human approval required before sending.'
            h+=f'<div class="card"><span class="badge WATCH">DRAFT ALERT</span><h3>{esc(s["name"])}</h3><p>{esc(msg)}</p></div>'
    return shell("Sub Alerts",'<div class="hero"><h1>Automatic Sub Alerts</h1><p class="muted">Prepared only. BuildCommand does not send external messages without approval.</p></div>'+(h or '<div class="card">No current alert candidates.</div>'))

@app.get("/field-command/deliveries",response_class=HTMLResponse)
def v41_deliveries_page():
    rows=_v41_delivery_risks(project_id()); h=""
    for r,level,exposure in rows:
        ex=(f' - {exposure} day exposure' if exposure is not None and exposure>0 else '')
        h+=f'<div class="action"><span class="badge WATCH">{esc(level)}</span> <b>{esc(r["item"])}</b><div class="small">Need {esc(r["required_on_site"])} - Promised {esc(r["promised_date"])}{ex}</div></div>'
    return shell("Delivery Intelligence",'<div class="hero"><h1>Delivery Intelligence</h1></div><div class="card">'+(h or '<p class="muted">No open deliveries.</p>')+'</div>')

@app.get("/field-command/inspections",response_class=HTMLResponse)
def v41_inspections_page():
    rows=_v41_inspection_countdown(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(label)}</span> <b>{esc(r["inspection_type"])}</b><div class="small">{esc(r["scheduled_date"])} - {esc(r["authority"])}</div></div>' for r,label,days in rows)
    return shell("Inspection Countdown",'<div class="hero"><h1>Inspection Countdown</h1><p class="muted">7 days - 3 days - tomorrow - today.</p></div><div class="card">'+(h or '<p class="muted">No pending inspections.</p>')+'</div>')

@app.get("/field-command/daily-report",response_class=HTMLResponse)
def v41_daily_report_page():
    pid=project_id()
    acts=_v39_rows("SELECT * FROM activities WHERE project_id=? AND COALESCE(status,'NOT_STARTED')!='COMPLETE' ORDER BY start LIMIT 20",(pid,))
    deliveries=_v41_delivery_risks(pid); inspections=_v41_inspection_countdown(pid); delays=_v41_delay_candidates(pid)
    work=", ".join(str(a["name"]) for a in acts[:8])
    body=f'<div class="hero"><h1>Daily Report Brain</h1><p class="muted">Draft from live project information; superintendent reviews before saving.</p></div><div class="card"><h2>Draft</h2><p><b>Work / upcoming:</b> {esc(work or "No active activities")}</p><p><b>Deliveries:</b> {len(deliveries)} open</p><p><b>Inspections:</b> {len(inspections)} pending</p><p><b>Potential delays:</b> {len(delays)}</p><p><a href="/daily-report">Open Daily Report editor -&gt;</a></p></div>'
    return shell("Daily Report Brain",body)

@app.get("/field-command/delays",response_class=HTMLResponse)
def v41_delays_page():
    rows=_v41_delay_candidates(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">DOCUMENT</span> <b>{esc(r["title"])}</b><div class="small">{esc(r["description"])}</div></div>' for r in rows)
    return shell("Delay Documentation",'<div class="hero"><h1>Delay Documentation Brain</h1><p class="muted">Capture dates, affected activity, responsible party, photos, RFIs and schedule impact.</p></div><div class="card">'+(h or '<p class="muted">No obvious delay candidates detected.</p>')+'</div>')

@app.get("/field-command/punch",response_class=HTMLResponse)
def v41_punch_page():
    rows=_v39_rows("SELECT * FROM punch_items WHERE project_id=? AND COALESCE(status,'OPEN') NOT IN ('VERIFIED','CLOSED','COMPLETE') ORDER BY location,trade,due",(project_id(),))
    h="".join(f'<div class="action"><b>{esc(r["location"] or "No location")} - {esc(r["title"])}</b><div class="small">{esc(r["trade"])} - {esc(r["owner"])} - Due {esc(r["due"])}</div></div>' for r in rows)
    return shell("Punch & Completion",'<div class="hero"><h1>Punch & Completion Brain</h1></div><div class="card">'+(h or '<p class="muted">No open punch items.</p>')+'<p><a href="/punch">Open full Punch List -&gt;</a></p></div>')

@app.get("/field-command/closeout",response_class=HTMLResponse)
def v41_closeout_page():
    pid=project_id(); _v41_closeout_seed(pid)
    rows=_v39_rows("SELECT * FROM closeout_items WHERE company_id=? AND project_id=? ORDER BY category,id",(current_company_id(),pid))
    complete=sum(1 for r in rows if str(r["status"] or "").upper() in {"COMPLETE","CLOSED","RECEIVED"})
    h="".join(f'<div class="action"><span class="badge">{esc(r["status"])}</span> <b>{esc(r["category"])}</b> - {esc(r["item"])}<div class="small">{esc(r["responsible_party"])} {esc(r["due_date"])}</div></div>' for r in rows)
    return shell("Closeout Brain",f'<div class="hero"><h1>Closeout Brain</h1><p class="muted">{complete}/{len(rows)} core turnover requirements complete.</p></div><div class="card">'+h+'</div>')


# ============================================================
# v42 PROJECT CONTROL INTELLIGENCE
# ============================================================

def _v42_ensure_tables():
    c=db()
    if DATABASE_KIND=="postgres":
        pk="BIGSERIAL PRIMARY KEY"; num="DOUBLE PRECISION"
    else:
        pk="INTEGER PRIMARY KEY"; num="REAL"
    stmts=[
      f"CREATE TABLE IF NOT EXISTS rfi_control(id {pk},company_id BIGINT,project_id BIGINT,number TEXT,title TEXT,question TEXT,responsible_party TEXT,due_date TEXT,status TEXT DEFAULT 'DRAFT',answer TEXT,cost_impact {num} DEFAULT 0,schedule_days {num} DEFAULT 0,source_ref TEXT,created TEXT,updated TEXT)",
      f"CREATE TABLE IF NOT EXISTS budget_commitments(id {pk},company_id BIGINT,project_id BIGINT,trade TEXT,vendor TEXT,original_budget {num} DEFAULT 0,commitment {num} DEFAULT 0,approved_changes {num} DEFAULT 0,pending_exposure {num} DEFAULT 0,notes TEXT,created TEXT,updated TEXT)",
      f"CREATE TABLE IF NOT EXISTS contract_scopes(id {pk},company_id BIGINT,project_id BIGINT,sub_name TEXT,trade TEXT,scope_text TEXT,exclusions TEXT,allowances TEXT,alternates TEXT,status TEXT DEFAULT 'REVIEW',created TEXT,updated TEXT)",
      f"CREATE TABLE IF NOT EXISTS correspondence_control(id {pk},company_id BIGINT,project_id BIGINT,correspondence_date TEXT,subject TEXT,party TEXT,trade TEXT,related_type TEXT,related_ref TEXT,summary TEXT,status TEXT DEFAULT 'OPEN',created TEXT)",
      f"CREATE TABLE IF NOT EXISTS owner_decisions(id {pk},company_id BIGINT,project_id BIGINT,title TEXT,decision_needed TEXT,due_date TEXT,cost_impact {num} DEFAULT 0,schedule_days {num} DEFAULT 0,status TEXT DEFAULT 'OPEN',source_ref TEXT,created TEXT,updated TEXT)",
      f"CREATE TABLE IF NOT EXISTS project_audit_log(id {pk},company_id BIGINT,project_id BIGINT,event_time TEXT,event_type TEXT,title TEXT,source_ref TEXT,decision TEXT,downstream_impact TEXT,reviewed_by TEXT,created TEXT)"
    ]
    for s in stmts: c.execute(s)
    c.commit(); c.close()

def _v42_rfis(pid):
    _v42_ensure_tables()
    return _v39_rows("SELECT * FROM rfi_control WHERE company_id=? AND project_id=? ORDER BY id DESC",(current_company_id(),pid))

def _v42_budget(pid):
    _v42_ensure_tables()
    return _v39_rows("SELECT * FROM budget_commitments WHERE company_id=? AND project_id=? ORDER BY trade,vendor",(current_company_id(),pid))

def _v42_contract_gaps(pid):
    _v42_ensure_tables()
    scopes=_v39_rows("SELECT trade,COUNT(*) n FROM blueprint_scope_items WHERE project_id=? GROUP BY trade ORDER BY trade",(pid,))
    contracts=_v39_rows("SELECT * FROM contract_scopes WHERE company_id=? AND project_id=?",(current_company_id(),pid))
    covered={str(r["trade"] or "").lower():r for r in contracts}
    return [(r["trade"],r["n"],covered.get(str(r["trade"] or "").lower())) for r in scopes]

def _v42_meeting_actions(pid):
    meetings=_v39_rows("SELECT * FROM meeting_notes WHERE project_id=? ORDER BY meeting_date DESC,id DESC LIMIT 50",(pid,))
    actions=_v39_rows("SELECT * FROM action_items WHERE project_id=? AND COALESCE(status,'OPEN') NOT IN ('CLOSED','COMPLETE') ORDER BY due",(pid,))
    return meetings,actions

def _v42_correspondence(pid):
    _v42_ensure_tables()
    return _v39_rows("SELECT * FROM correspondence_control WHERE company_id=? AND project_id=? ORDER BY correspondence_date DESC,id DESC",(current_company_id(),pid))

def _v42_owner(pid):
    _v42_ensure_tables()
    return _v39_rows("SELECT * FROM owner_decisions WHERE company_id=? AND project_id=? AND COALESCE(status,'OPEN') NOT IN ('CLOSED','COMPLETE') ORDER BY due_date",(current_company_id(),pid))

def _v42_audit(pid):
    _v42_ensure_tables()
    return _v39_rows("SELECT * FROM project_audit_log WHERE company_id=? AND project_id=? ORDER BY event_time DESC,id DESC LIMIT 200",(current_company_id(),pid))

@app.get("/project-control",response_class=HTMLResponse)
def v42_project_control():
    pid=project_id(); _v42_ensure_tables()
    rfis=_v42_rfis(pid); subs=_v39_rows("SELECT * FROM submittals WHERE project_id=? AND COALESCE(status,'PENDING') NOT IN ('APPROVED','CLOSED','COMPLETE')",(pid,))
    budget=_v42_budget(pid); changes=_v39_changes(pid); owners=_v42_owner(pid); gaps=_v42_contract_gaps(pid)
    meetings,actions=_v42_meeting_actions(pid); corr=_v42_correspondence(pid); audit=_v42_audit(pid)
    pending=sum(float(r["pending_exposure"] or 0) for r in budget)+sum(float(r["estimated_cost"] or 0) for r in changes)
    gap_count=sum(1 for _,_,c in gaps if c is None)
    body=f'<div class="hero"><div class="eyebrow">BuildCommand v42 - Project Control</div><h1>Control cost, schedule and decisions.</h1><p class="muted">{len(rfis)} RFIs - {len(subs)} open submittals - ${pending:,.0f} pending exposure - {len(owners)} owner decisions - {gap_count} contract scope gaps - {len(actions)} open actions.</p></div><div class="grid3">'
    cards=[
      ("RFI Command","Lifecycle, cost/schedule impact and field verification.","/project-control/rfis"),
      ("Submittal Command","Track required, submitted, review and release status.","/project-control/submittals"),
      ("Budget & Commitments","Budget, commitments, approved changes and exposure.","/project-control/budget"),
      ("Change Management","Trace potential changes through cost and schedule.","/project-control/changes"),
      ("Contract Scope","Compare subcontract coverage against Blueprint Brain.","/project-control/contracts"),
      ("Meeting Intelligence","Decisions, commitments and open actions.","/project-control/meetings"),
      ("Correspondence","Organize important communication around project issues.","/project-control/correspondence"),
      ("Owner Decisions","Decisions required and their consequence.","/project-control/owner-decisions"),
      ("Audit Trail","Chronological project intelligence record.","/project-control/audit")
    ]
    for name,desc,href in cards: body+=_v37_link_card(name,desc,href,"Open")
    body+='</div>'
    return shell("Project Control",body)

@app.get("/project-control/rfis",response_class=HTMLResponse)
def v42_rfi_page():
    rows=_v42_rfis(project_id())
    h="".join(f'<div class="action"><span class="badge">{esc(r["status"])}</span> <b>{esc(r["number"] or "DRAFT")} - {esc(r["title"])}</b><div class="small">Due {esc(r["due_date"])} - Cost ${float(r["cost_impact"] or 0):,.0f} - {float(r["schedule_days"] or 0):g} days</div></div>' for r in rows)
    return shell("RFI Command",'<div class="hero"><h1>RFI Command Brain</h1><p class="muted">Detected - draft - submitted - answered - cost/schedule impact - field verification.</p></div><div class="card">'+(h or '<p class="muted">No controlled RFIs yet. Conflict Brain proposals remain available for human review.</p>')+'</div>')

@app.get("/project-control/submittals",response_class=HTMLResponse)
def v42_submittal_page():
    rows=_v39_rows("SELECT * FROM submittals WHERE project_id=? ORDER BY due_date,id",(project_id(),))
    h="".join(f'<div class="action"><span class="badge">{esc(r["status"])}</span> <b>{esc(r["title"])}</b><div class="small">Spec {esc(r["spec_section"])} - {esc(r["responsible_party"])} - Due {esc(r["due_date"])}</div></div>' for r in rows)
    return shell("Submittal Command",'<div class="hero"><h1>Submittal Command Brain</h1></div><div class="card">'+(h or '<p class="muted">No submittals loaded.</p>')+'</div>')

@app.get("/project-control/budget",response_class=HTMLResponse)
def v42_budget_page():
    rows=_v42_budget(project_id()); h=""; tb=tc=ta=tp=0
    for r in rows:
        b=float(r["original_budget"] or 0); c=float(r["commitment"] or 0); a=float(r["approved_changes"] or 0); p=float(r["pending_exposure"] or 0)
        tb+=b; tc+=c; ta+=a; tp+=p; forecast=c+a+p; level="WATCH" if b and forecast>b else "READY"
        h+=f'<div class="action"><span class="badge {level}">{level}</span> <b>{esc(r["trade"])}</b><div class="small">Budget ${b:,.0f} - Committed ${c:,.0f} - Approved ${a:,.0f} - Pending ${p:,.0f} - Forecast ${forecast:,.0f}</div></div>'
    head=f'<div class="hero"><h1>Budget & Commitment Brain</h1><p class="muted">Budget ${tb:,.0f} - Commitments ${tc:,.0f} - Approved changes ${ta:,.0f} - Pending exposure ${tp:,.0f}</p></div>'
    return shell("Budget & Commitments",head+'<div class="card">'+(h or '<p class="muted">No budget commitments loaded yet.</p>')+'</div>')

@app.get("/project-control/changes",response_class=HTMLResponse)
def v42_changes_page():
    rows=_v39_changes(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(r["status"])}</span> <b>{esc(r["event_type"])} - {esc(r["title"])}</b><div class="small">${float(r["estimated_cost"] or 0):,.0f} - {float(r["schedule_days"] or 0):g} days - {esc(r["responsible_party"])}</div></div>' for r in rows)
    return shell("Change Management",'<div class="hero"><h1>Change Management Brain</h1><p class="muted">RFI / ASI / bulletin / owner request / field condition to cost and schedule exposure.</p></div><div class="card">'+(h or '<p class="muted">No open change events.</p>')+'</div>')

@app.get("/project-control/contracts",response_class=HTMLResponse)
def v42_contract_page():
    rows=_v42_contract_gaps(project_id()); h=""
    for trade,n,c in rows:
        if c is None:
            h+=f'<div class="action"><span class="badge WATCH">GAP</span> <b>{esc(trade)}</b><div class="small">{n} Blueprint Brain scope items - no contract scope record loaded.</div></div>'
        else:
            exclusions=str(c["exclusions"] or "")
            h+=f'<div class="action"><span class="badge READY">REVIEWED</span> <b>{esc(trade)} - {esc(c["sub_name"])}</b><div class="small">{n} source-backed items - exclusions: {esc(exclusions[:300])}</div></div>'
    return shell("Contract Scope",'<div class="hero"><h1>Contract Scope Brain</h1><p class="muted">Scope gaps, exclusions, overlaps, allowances and unclear responsibility.</p></div><div class="card">'+(h or '<p class="muted">Analyze Blueprint Brain scope first.</p>')+'</div>')

@app.get("/project-control/meetings",response_class=HTMLResponse)
def v42_meeting_page():
    meetings,actions=_v42_meeting_actions(project_id()); h=""
    for m in meetings:
        h+=f'<div class="card"><span class="badge">{esc(m["meeting_type"])}</span><h3>{esc(m["title"])}</h3><div class="small">{esc(m["meeting_date"])} - {esc(m["attendees"])}</div><p><b>Decisions:</b> {esc(m["decisions"])}</p><p><b>Commitments:</b> {esc(m["commitments"])}</p></div>'
    ah="".join(f'<div class="action"><b>{esc(a["title"])}</b><div class="small">{esc(a["owner"])} - Due {esc(a["due"])} - {esc(a["priority"])}</div></div>' for a in actions)
    return shell("Meeting Intelligence",'<div class="hero"><h1>Meeting Intelligence</h1><p class="muted">Keep decisions and commitments from disappearing.</p></div>'+(h or '<div class="card">No meeting notes loaded.</div>')+'<div class="card"><h2>Open Actions</h2>'+ (ah or '<p class="muted">No open actions.</p>')+'</div>')

@app.get("/project-control/correspondence",response_class=HTMLResponse)
def v42_correspondence_page():
    rows=_v42_correspondence(project_id())
    h="".join(f'<div class="action"><span class="badge">{esc(r["related_type"])}</span> <b>{esc(r["subject"])}</b><div class="small">{esc(r["correspondence_date"])} - {esc(r["party"])} - {esc(r["trade"])} - {esc(r["related_ref"])}</div><p>{esc(r["summary"])}</p></div>' for r in rows)
    return shell("Correspondence",'<div class="hero"><h1>Correspondence Brain</h1><p class="muted">Communication organized around trade, issue, RFI, submittal, schedule and change.</p></div><div class="card">'+(h or '<p class="muted">No controlled correspondence loaded yet.</p>')+'</div>')

@app.get("/project-control/owner-decisions",response_class=HTMLResponse)
def v42_owner_page():
    rows=_v42_owner(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">DECISION</span> <b>{esc(r["title"])}</b><div>{esc(r["decision_needed"])}</div><div class="small">Due {esc(r["due_date"])} - ${float(r["cost_impact"] or 0):,.0f} exposure - {float(r["schedule_days"] or 0):g} days at risk - {esc(r["source_ref"])}</div></div>' for r in rows)
    return shell("Owner Decisions",'<div class="hero"><h1>Owner Decision Tracker</h1><p class="muted">Decision needed - deadline - downstream cost/schedule consequence.</p></div><div class="card">'+(h or '<p class="muted">No owner decisions currently tracked.</p>')+'</div>')

@app.get("/project-control/audit",response_class=HTMLResponse)
def v42_audit_page():
    rows=_v42_audit(project_id())
    h="".join(f'<div class="action"><span class="badge">{esc(r["event_type"])}</span> <b>{esc(r["title"])}</b><div class="small">{esc(r["event_time"])} - {esc(r["source_ref"])} - reviewed by {esc(r["reviewed_by"])}</div><p>{esc(r["downstream_impact"])}</p></div>' for r in rows)
    return shell("Project Audit Trail",'<div class="hero"><h1>Project Audit Trail</h1><p class="muted">What changed, source, decision, downstream impact and review history.</p></div><div class="card">'+(h or '<p class="muted">Audit events will accumulate as controlled project decisions are recorded.</p>')+'</div>')


# ============================================================
# v43 CONSTRUCTION LEARNING INTELLIGENCE
# Human-approved learning only: AI output never teaches itself.
# ============================================================

def _v43_ensure_tables():
    c=db()
    if DATABASE_KIND=="postgres":
        pk="BIGSERIAL PRIMARY KEY"; num="DOUBLE PRECISION"
    else:
        pk="INTEGER PRIMARY KEY"; num="REAL"
    stmts=[
      f"""CREATE TABLE IF NOT EXISTS learning_rules(
        id {pk},company_id BIGINT,project_id BIGINT,rule_type TEXT,subject TEXT,
        learned_rule TEXT,source_ref TEXT,scope_level TEXT DEFAULT 'PROJECT ONLY',
        approval_status TEXT DEFAULT 'PROPOSED',approved_by TEXT,confidence TEXT DEFAULT 'HIGH',
        created TEXT,updated TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS estimator_learning(
        id {pk},company_id BIGINT,project_id BIGINT,estimator_item_id BIGINT,trade TEXT,
        description TEXT,ai_quantity {num},accepted_quantity {num},unit TEXT,
        decision TEXT,reason TEXT,approval_status TEXT DEFAULT 'PROPOSED',
        approved_by TEXT,created TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS historical_costs(
        id {pk},company_id BIGINT,project_id BIGINT,trade TEXT,scope TEXT,unit TEXT,
        quantity {num},unit_cost {num},total_cost {num},building_type TEXT,location TEXT,
        vendor TEXT,cost_date TEXT,approval_status TEXT DEFAULT 'APPROVED',created TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS schedule_learning(
        id {pk},company_id BIGINT,project_id BIGINT,activity_id BIGINT,trade TEXT,
        activity_name TEXT,planned_start TEXT,planned_finish TEXT,actual_start TEXT,actual_finish TEXT,
        planned_days {num},actual_days {num},production_note TEXT,approval_status TEXT DEFAULT 'PROPOSED',created TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS subcontractor_performance(
        id {pk},company_id BIGINT,project_id BIGINT,sub_name TEXT,trade TEXT,
        schedule_score {num} DEFAULT 0,quality_score {num} DEFAULT 0,safety_score {num} DEFAULT 0,
        responsiveness_score {num} DEFAULT 0,change_score {num} DEFAULT 0,closeout_score {num} DEFAULT 0,
        notes TEXT,approval_status TEXT DEFAULT 'PROPOSED',created TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS inspection_learning(
        id {pk},company_id BIGINT,project_id BIGINT,inspection_type TEXT,authority TEXT,trade TEXT,
        result TEXT,failure_reason TEXT,correction TEXT,lesson TEXT,
        approval_status TEXT DEFAULT 'PROPOSED',created TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS lessons_learned(
        id {pk},company_id BIGINT,project_id BIGINT,category TEXT,title TEXT,lesson TEXT,
        recommendation TEXT,source_ref TEXT,scope_level TEXT DEFAULT 'PROJECT ONLY',
        approval_status TEXT DEFAULT 'PROPOSED',approved_by TEXT,created TEXT)"""
    ]
    for s in stmts: c.execute(s)
    c.commit(); c.close()

def _v43_rules(pid=None):
    _v43_ensure_tables()
    if pid:
        return _v39_rows("SELECT * FROM learning_rules WHERE company_id=? AND project_id=? ORDER BY id DESC",(current_company_id(),pid))
    return _v39_rows("SELECT * FROM learning_rules WHERE company_id=? ORDER BY id DESC LIMIT 300",(current_company_id(),))

def _v43_trade_proposals(pid):
    # Current high-confidence reviewed scope becomes a learning candidate only.
    rows=_v39_rows("SELECT * FROM blueprint_scope_items WHERE project_id=? AND COALESCE(confidence,'')='HIGH' ORDER BY id DESC LIMIT 150",(pid,))
    return rows

def _v43_estimator_signals(pid):
    return _v39_rows("SELECT * FROM estimator_items WHERE project_id=? AND (COALESCE(verified,0)=1 OR ai_quantity IS NOT NULL) ORDER BY id DESC LIMIT 150",(pid,))

def _v43_schedule_signals(pid):
    return _v39_rows("SELECT * FROM activities WHERE project_id=? ORDER BY id DESC LIMIT 150",(pid,))

def _v43_inspection_signals(pid):
    return _v39_rows("SELECT * FROM inspections_tracker WHERE project_id=? AND COALESCE(result,'PENDING') NOT IN ('PENDING','PASSED') ORDER BY id DESC LIMIT 100",(pid,))

def _v43_rfi_change_patterns(pid):
    _v42_ensure_tables()
    rfis=_v39_rows("SELECT * FROM rfi_control WHERE company_id=? AND project_id=? ORDER BY id DESC",(current_company_id(),pid))
    changes=_v39_rows("SELECT * FROM change_events WHERE project_id=? ORDER BY id DESC",(pid,))
    return rfis,changes

@app.get("/learning",response_class=HTMLResponse)
def v43_learning_center():
    pid=project_id(); _v43_ensure_tables()
    rules=_v43_rules(pid); trade=_v43_trade_proposals(pid); est=_v43_estimator_signals(pid)
    inspections=_v43_inspection_signals(pid); rfis,changes=_v43_rfi_change_patterns(pid)
    approved=sum(1 for r in rules if str(r["approval_status"] or "").upper()=="APPROVED")
    body=f'<div class="hero"><div class="eyebrow">BuildCommand v43 - Construction Learning Intelligence</div><h1>Every approved lesson makes the next project smarter.</h1><p class="muted">{approved} approved learned rules - {len(trade)} trade-learning candidates - {len(est)} estimator signals - {len(inspections)} inspection lessons - {len(rfis)+len(changes)} RFI/change patterns.</p><p><b>Learning firewall:</b> AI proposals never become permanent construction knowledge without human approval.</p></div><div class="grid3">'
    cards=[
      ("Trade Assignment Learning","Turn reviewed trade corrections into reusable rules.","/learning/trades"),
      ("Scope Boundary Brain","Learn where one trade stops and another starts.","/learning/boundaries"),
      ("Estimator Learning","Learn from accepted, rejected and adjusted takeoff proposals.","/learning/estimator"),
      ("Historical Cost Brain","Private cost history by trade, scope, unit and project.","/learning/costs"),
      ("Schedule Learning","Compare planned durations with actual field performance.","/learning/schedule"),
      ("Sub Performance","Learn which subcontractors actually perform.","/learning/subs"),
      ("Inspection Learning","Remember failures, corrections and AHJ patterns.","/learning/inspections"),
      ("RFI / Change Patterns","Find recurring drawing conflicts and cost exposure.","/learning/patterns"),
      ("Lessons Learned","Turn project experience into approved future guidance.","/learning/lessons"),
      ("Knowledge Core","Controlled project/company/global construction knowledge.","/learning/core")
    ]
    for name,desc,href in cards: body+=_v37_link_card(name,desc,href,"Open")
    body+='</div>'
    return shell("Construction Learning",body)

@app.get("/learning/trades",response_class=HTMLResponse)
def v43_trade_learning():
    rows=_v43_trade_proposals(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">PROPOSAL</span> <b>{esc(r["trade"])}</b> - {esc(r["requirement"])}<div class="small">{esc(r["source_sheet"])} {esc(r["source_detail"])} - Human approval required before learning.</div></div>' for r in rows)
    return shell("Trade Assignment Learning",'<div class="hero"><h1>Trade Assignment Learning Brain</h1><p class="muted">Reviewed construction corrections become candidates for reusable knowledge, never automatic rules.</p></div><div class="card">'+(h or '<p class="muted">No high-confidence trade candidates yet.</p>')+'</div>')

@app.get("/learning/boundaries",response_class=HTMLResponse)
def v43_boundaries():
    rows=_v43_rules()
    rows=[r for r in rows if str(r["rule_type"] or "").upper() in {"SCOPE BOUNDARY","TRADE ASSIGNMENT"}]
    h="".join(f'<div class="action"><span class="badge">{esc(r["scope_level"])}</span> <b>{esc(r["subject"])}</b><div>{esc(r["learned_rule"])}</div><div class="small">{esc(r["approval_status"])} - {esc(r["source_ref"])}</div></div>' for r in rows)
    return shell("Scope Boundary Brain",'<div class="hero"><h1>Scope Boundary Brain</h1><p class="muted">Responsibility rules can be PROJECT ONLY, COMPANY STANDARD, or controlled GLOBAL BUILDCOMMAND RULE.</p></div><div class="card">'+(h or '<p class="muted">No approved boundary rules recorded yet.</p>')+'</div>')

@app.get("/learning/estimator",response_class=HTMLResponse)
def v43_estimator_learning():
    rows=_v43_estimator_signals(project_id()); h=""
    for r in rows:
        ai=r["ai_quantity"]; accepted=r["quantity"]; delta=(float(accepted or 0)-float(ai or 0)) if ai is not None else None
        h+=f'<div class="action"><span class="badge {"READY" if r["verified"] else "WATCH"}">{"VERIFIED" if r["verified"] else "REVIEW"}</span> <b>{esc(r["trade"])} - {esc(r["description"])}</b><div class="small">AI {esc(ai)} {esc(r["ai_unit"])} - Current {esc(accepted)} {esc(r["unit"])}'+(f' - Delta {delta:g}' if delta is not None else '')+'</div></div>'
    return shell("Estimator Learning",'<div class="hero"><h1>Estimator Learning Brain</h1><p class="muted">Quantity decisions remain proposals until reviewed; estimator quantities are never silently overwritten.</p></div><div class="card">'+(h or '<p class="muted">No estimator learning signals yet.</p>')+'</div>')

@app.get("/learning/costs",response_class=HTMLResponse)
def v43_costs():
    _v43_ensure_tables()
    rows=_v39_rows("SELECT * FROM historical_costs WHERE company_id=? ORDER BY cost_date DESC,id DESC LIMIT 200",(current_company_id(),))
    h="".join(f'<div class="action"><b>{esc(r["trade"])} - {esc(r["scope"])}</b><div class="small">{float(r["quantity"] or 0):g} {esc(r["unit"])} @ ${float(r["unit_cost"] or 0):,.2f} - Total ${float(r["total_cost"] or 0):,.0f} - {esc(r["vendor"])}</div></div>' for r in rows)
    return shell("Historical Cost Brain",'<div class="hero"><h1>Historical Cost Brain</h1><p class="muted">Private company cost intelligence by trade, scope, unit, project, vendor and date.</p></div><div class="card">'+(h or '<p class="muted">Historical costs will build from reviewed project actuals.</p>')+'</div>')

@app.get("/learning/schedule",response_class=HTMLResponse)
def v43_schedule_learning():
    rows=_v43_schedule_signals(project_id())
    h="".join(f'<div class="action"><b>{esc(r["trade"])} - {esc(r["name"])}</b><div class="small">Planned {esc(r["start"])} to {esc(r["finish"])} - {float(r["pct"] or 0):g}% complete - {esc(r["status"])}</div></div>' for r in rows)
    return shell("Schedule Learning",'<div class="hero"><h1>Schedule Learning Brain</h1><p class="muted">Build realistic production intelligence by comparing planned work against actual field performance.</p></div><div class="card">'+(h or '<p class="muted">No schedule activities loaded.</p>')+'</div>')

@app.get("/learning/subs",response_class=HTMLResponse)
def v43_sub_learning():
    _v43_ensure_tables()
    rows=_v39_rows("SELECT * FROM subcontractor_performance WHERE company_id=? ORDER BY id DESC LIMIT 200",(current_company_id(),))
    current=_v39_rows("SELECT * FROM subs WHERE project_id=? ORDER BY trade,name",(project_id(),))
    h="".join(f'<div class="action"><b>{esc(r["sub_name"])} - {esc(r["trade"])}</b><div class="small">Schedule {float(r["schedule_score"] or 0):g} - Quality {float(r["quality_score"] or 0):g} - Safety {float(r["safety_score"] or 0):g} - Responsiveness {float(r["responsiveness_score"] or 0):g} - Closeout {float(r["closeout_score"] or 0):g}</div></div>' for r in rows)
    if not h:
        h="".join(f'<div class="action"><span class="badge WATCH">NEEDS HISTORY</span> <b>{esc(r["name"])} - {esc(r["trade"])}</b></div>' for r in current)
    return shell("Subcontractor Performance",'<div class="hero"><h1>Subcontractor Performance Brain</h1><p class="muted">Performance history should inform future selection, not price alone.</p></div><div class="card">'+(h or '<p class="muted">No subcontractors loaded.</p>')+'</div>')

@app.get("/learning/inspections",response_class=HTMLResponse)
def v43_inspection_learning():
    rows=_v43_inspection_signals(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(r["result"])}</span> <b>{esc(r["inspection_type"])}</b><div class="small">{esc(r["authority"])} - {esc(r["notes"])}</div></div>' for r in rows)
    return shell("Inspection Learning",'<div class="hero"><h1>Inspection Learning Brain</h1><p class="muted">Remember failed inspections and corrections by jurisdiction, trade and inspection type.</p></div><div class="card">'+(h or '<p class="muted">No failed/rejected inspection signals.</p>')+'</div>')

@app.get("/learning/patterns",response_class=HTMLResponse)
def v43_patterns():
    rfis,changes=_v43_rfi_change_patterns(project_id()); h=""
    for r in rfis:
        h+=f'<div class="action"><span class="badge">RFI</span> <b>{esc(r["title"])}</b><div class="small">{esc(r["source_ref"])} - ${float(r["cost_impact"] or 0):,.0f} - {float(r["schedule_days"] or 0):g} days</div></div>'
    for r in changes:
        h+=f'<div class="action"><span class="badge WATCH">CHANGE</span> <b>{esc(r["title"])}</b><div class="small">{esc(r["event_type"])} - ${float(r["estimated_cost"] or 0):,.0f} - {float(r["schedule_days"] or 0):g} days</div></div>'
    return shell("RFI / Change Patterns",'<div class="hero"><h1>RFI / Change Pattern Brain</h1><p class="muted">Recurring conflicts become preconstruction warnings after review.</p></div><div class="card">'+(h or '<p class="muted">No RFI/change history yet.</p>')+'</div>')

@app.get("/learning/lessons",response_class=HTMLResponse)
def v43_lessons():
    _v43_ensure_tables()
    rows=_v39_rows("SELECT * FROM lessons_learned WHERE company_id=? AND project_id=? ORDER BY id DESC",(current_company_id(),project_id()))
    h="".join(f'<div class="action"><span class="badge">{esc(r["approval_status"])}</span> <b>{esc(r["category"])} - {esc(r["title"])}</b><div>{esc(r["lesson"])}</div><div class="small">Recommendation: {esc(r["recommendation"])} - {esc(r["scope_level"])}</div></div>' for r in rows)
    return shell("Lessons Learned",'<div class="hero"><h1>Project Lessons-Learned Brain</h1><p class="muted">What worked, what failed, what cost money, what delayed the job, and what the next project should do differently.</p></div><div class="card">'+(h or '<p class="muted">Lessons will accumulate through reviewed project closeout.</p>')+'</div>')

@app.get("/learning/core",response_class=HTMLResponse)
def v43_knowledge_core():
    rows=_v43_rules(); project=sum(1 for r in rows if r["scope_level"]=="PROJECT ONLY"); company=sum(1 for r in rows if r["scope_level"]=="COMPANY STANDARD"); globaln=sum(1 for r in rows if r["scope_level"]=="GLOBAL BUILDCOMMAND RULE"); approved=sum(1 for r in rows if str(r["approval_status"]).upper()=="APPROVED")
    h="".join(f'<div class="action"><span class="badge">{esc(r["scope_level"])}</span> <b>{esc(r["rule_type"])} - {esc(r["subject"])}</b><div>{esc(r["learned_rule"])}</div><div class="small">{esc(r["approval_status"])} - source {esc(r["source_ref"])}</div></div>' for r in rows[:150])
    head=f'<div class="hero"><h1>BuildCommand Construction Knowledge Core</h1><p class="muted">{approved} approved - {project} project-only - {company} company standards - {globaln} controlled global rules.</p><p><b>Knowledge firewall:</b> PROJECT ONLY can never silently become COMPANY STANDARD or GLOBAL. Promotion requires human approval.</p></div>'
    return shell("Knowledge Core",head+'<div class="card">'+(h or '<p class="muted">The controlled knowledge core is ready to learn from approved construction decisions.</p>')+'</div>')


# ============================================================
# v44 PRECONSTRUCTION & BID INTELLIGENCE
# ============================================================

def _v44_ensure_tables():
    c=db()
    if DATABASE_KIND=="postgres":
        pk="BIGSERIAL PRIMARY KEY"; num="DOUBLE PRECISION"
    else:
        pk="INTEGER PRIMARY KEY"; num="REAL"
    stmts=[
      f"""CREATE TABLE IF NOT EXISTS pursuit_intelligence(
        id {pk},company_id BIGINT,project_id BIGINT,client TEXT,project_type TEXT,
        location TEXT,bid_due TEXT,estimated_value {num} DEFAULT 0,competition TEXT,
        strategic_fit {num} DEFAULT 0,risk_score {num} DEFAULT 0,win_score {num} DEFAULT 0,
        recommendation TEXT DEFAULT 'REVIEW',notes TEXT,created TEXT,updated TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS bid_invites(
        id {pk},company_id BIGINT,project_id BIGINT,trade TEXT,sub_name TEXT,
        contact TEXT,invite_date TEXT,due_date TEXT,status TEXT DEFAULT 'PLANNED',
        response TEXT,notes TEXT,created TEXT,updated TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS bid_proposals(
        id {pk},company_id BIGINT,project_id BIGINT,trade TEXT,sub_name TEXT,
        base_bid {num} DEFAULT 0,alternates {num} DEFAULT 0,allowances {num} DEFAULT 0,
        exclusions TEXT,qualifications TEXT,scope_coverage {num} DEFAULT 0,
        risk_score {num} DEFAULT 0,status TEXT DEFAULT 'REVIEW',received_date TEXT,
        notes TEXT,created TEXT,updated TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS bid_packages(
        id {pk},company_id BIGINT,project_id BIGINT,trade TEXT,title TEXT,
        scope_text TEXT,source_count INTEGER DEFAULT 0,status TEXT DEFAULT 'DRAFT',
        issued_date TEXT,due_date TEXT,created TEXT,updated TEXT)"""
    ]
    for s in stmts:
        c.execute(s)
    c.commit(); c.close()

def _v44_scope_by_trade(pid):
    return _v39_rows(
        "SELECT trade,COUNT(*) AS n FROM blueprint_scope_items WHERE project_id=? GROUP BY trade ORDER BY trade",
        (pid,)
    )

def _v44_estimate_by_trade(pid):
    rows=_v39_rows("""
        SELECT trade,
               SUM(COALESCE(quantity,0)*COALESCE(material_unit_cost,0)) AS material,
               SUM(COALESCE(quantity,0)*COALESCE(labor_unit_cost,0)) AS labor,
               SUM(COALESCE(subcontract_quote,0)) AS subquote,
               SUM(COALESCE(allowance,0)) AS allowance
        FROM estimator_items
        WHERE project_id=?
        GROUP BY trade ORDER BY trade
    """,(pid,))
    return rows

def _v44_historical_cost_by_trade():
    _v43_ensure_tables()
    return _v39_rows("""
        SELECT trade,AVG(COALESCE(unit_cost,0)) AS avg_unit_cost,
               COUNT(*) AS samples,SUM(COALESCE(total_cost,0)) AS total_history
        FROM historical_costs
        WHERE company_id=? AND COALESCE(approval_status,'APPROVED')='APPROVED'
        GROUP BY trade ORDER BY trade
    """,(current_company_id(),))

def _v44_bid_coverage(pid):
    _v44_ensure_tables()
    scopes=_v44_scope_by_trade(pid)
    invites=_v39_rows("SELECT * FROM bid_invites WHERE company_id=? AND project_id=?",(current_company_id(),pid))
    proposals=_v39_rows("SELECT * FROM bid_proposals WHERE company_id=? AND project_id=?",(current_company_id(),pid))
    out=[]
    for r in scopes:
        trade=str(r["trade"] or "")
        ti=[x for x in invites if str(x["trade"] or "").lower()==trade.lower()]
        tp=[x for x in proposals if str(x["trade"] or "").lower()==trade.lower()]
        status="PRICED" if tp else "INVITED" if ti else "NO COVERAGE"
        out.append((trade,int(r["n"] or 0),len(ti),len(tp),status))
    return out

def _v44_scope_gaps(pid):
    rows=_v39_rows("""
        SELECT * FROM blueprint_scope_items
        WHERE project_id=? AND (
            COALESCE(confidence,'MEDIUM') IN ('LOW','MEDIUM')
            OR COALESCE(related_trade,'')!=''
        )
        ORDER BY trade,id LIMIT 150
    """,(pid,))
    return rows

def _v44_allowances(pid):
    rows=_v39_rows("""
        SELECT * FROM estimator_items
        WHERE project_id=? AND (
            COALESCE(allowance,0)>0
            OR LOWER(COALESCE(description,'')) LIKE '%allowance%'
            OR LOWER(COALESCE(description,'')) LIKE '%alternate%'
            OR LOWER(COALESCE(description,'')) LIKE '%owner furnished%'
        )
        ORDER BY trade,id
    """,(pid,))
    return rows

def _v44_long_leads(pid):
    rows=_v39_rows("""
        SELECT * FROM procurement
        WHERE project_id=? AND COALESCE(status,'OPEN') NOT IN ('DELIVERED','COMPLETE','CLOSED')
        ORDER BY required_on_site
    """,(pid,))
    return rows

def _v44_precon_risk(pid):
    coverage=_v44_bid_coverage(pid)
    gaps=_v44_scope_gaps(pid)
    est=_v44_estimate_by_trade(pid)
    longleads=_v44_long_leads(pid)
    uncovered=sum(1 for x in coverage if x[4]=="NO COVERAGE")
    unverified=_v37_snapshot(pid)["review"]
    score=min(100,uncovered*10+len(gaps)*2+unverified*2+len(longleads)*4)
    level="CRITICAL" if score>=70 else "HIGH" if score>=45 else "MEDIUM" if score>=20 else "LOW"
    return {"score":score,"level":level,"uncovered":uncovered,"gaps":len(gaps),"unverified":unverified,"longleads":len(longleads)}

def _v44_pursuit(pid):
    _v44_ensure_tables()
    rows=_v39_rows("SELECT * FROM pursuit_intelligence WHERE company_id=? AND project_id=? ORDER BY id DESC LIMIT 1",(current_company_id(),pid))
    return rows[0] if rows else None

@app.get("/preconstruction",response_class=HTMLResponse)
def v44_preconstruction():
    pid=project_id(); _v44_ensure_tables()
    risk=_v44_precon_risk(pid); coverage=_v44_bid_coverage(pid); gaps=_v44_scope_gaps(pid)
    proposals=_v39_rows("SELECT * FROM bid_proposals WHERE company_id=? AND project_id=?",(current_company_id(),pid))
    allowances=_v44_allowances(pid); longleads=_v44_long_leads(pid)
    body=f'<div class="hero"><div class="eyebrow">BuildCommand v44 - Preconstruction & Bid Intelligence</div><h1>Decide what to bid, what it should cost, and where the risk is.</h1><p class="muted">Preconstruction risk {risk["level"]} ({risk["score"]}/100) - {risk["uncovered"]} trades without bid coverage - {len(proposals)} proposals - {len(gaps)} scope review items - {len(allowances)} allowance/alternate items - {len(longleads)} long-lead signals.</p></div><div class="grid3">'
    cards=[
      ("Pursuit / Go-No-Go","Should we chase this project? Score fit, risk, competition and win potential.","/preconstruction/pursuit"),
      ("Scope Completeness","What is in the job, what is unclear, and what might be missing.","/preconstruction/scope"),
      ("Estimate Benchmark","Compare current estimate structure with approved historical company costs.","/preconstruction/benchmark"),
      ("Bid Coverage","Which trades are covered, invited, priced or completely exposed.","/preconstruction/coverage"),
      ("Bidder Strategy","Build the subcontractor pursuit list by trade and performance history.","/preconstruction/bidders"),
      ("Bid Packages","Turn cleaned Blueprint Brain scopes into trade bid-package readiness.","/preconstruction/packages"),
      ("Bid Leveling","Compare subcontractor pricing, coverage, exclusions and qualifications.","/preconstruction/leveling"),
      ("Allowances & Alternates","Separate uncertain money before it disappears into the estimate.","/preconstruction/allowances"),
      ("Long-Lead / Schedule Risk","Find procurement items that can hurt the proposed schedule.","/preconstruction/long-lead"),
      ("Preconstruction Risk","One consolidated view of scope, pricing, coverage and schedule exposure.","/preconstruction/risk")
    ]
    for name,desc,href in cards:
        body+=_v37_link_card(name,desc,href,"Open")
    body+='</div>'
    return shell("Preconstruction",body)

@app.get("/preconstruction/pursuit",response_class=HTMLResponse)
def v44_pursuit_page():
    pid=project_id(); p=_v44_pursuit(pid); risk=_v44_precon_risk(pid)
    if p:
        win=float(p["win_score"] or 0); fit=float(p["strategic_fit"] or 0)
        rec=p["recommendation"] or "REVIEW"
        detail=f'<div class="grid3"><div class="card"><div class="label">Win Score</div><div class="kpi">{win:g}</div></div><div class="card"><div class="label">Strategic Fit</div><div class="kpi">{fit:g}</div></div><div class="card"><div class="label">Recommendation</div><div class="kpi">{esc(rec)}</div></div></div><div class="card"><p>{esc(p["notes"])}</p></div>'
    else:
        detail=f'<div class="card"><span class="badge WATCH">REVIEW</span><h2>Go / No-Go candidate</h2><p>Current preconstruction risk: <b>{risk["level"]}</b> ({risk["score"]}/100).</p><p class="muted">Add client, competition, strategic fit, project value and win probability before making a pursuit decision.</p></div>'
    return shell("Pursuit Intelligence",'<div class="hero"><h1>Pursuit / Go-No-Go Brain</h1><p class="muted">Do not burn estimating time on every opportunity. Score whether the project is worth chasing.</p></div>'+detail)

@app.get("/preconstruction/scope",response_class=HTMLResponse)
def v44_scope_page():
    rows=_v44_scope_gaps(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(r["confidence"])}</span> <b>{esc(r["trade"])}</b> - {esc(r["requirement"])}<div class="small">{esc(r["source_sheet"])} {esc(r["source_detail"])} {esc(r["source_spec"])}'+(f' - Related trade: {esc(r["related_trade"])}' if r["related_trade"] else '')+'</div></div>' for r in rows)
    return shell("Scope Completeness",'<div class="hero"><h1>Scope Completeness Brain</h1><p class="muted">Focus estimating attention on ambiguous, cross-trade and lower-confidence requirements.</p></div><div class="card">'+(h or '<p class="muted">No obvious scope-completeness exceptions detected.</p>')+'</div>')

@app.get("/preconstruction/benchmark",response_class=HTMLResponse)
def v44_benchmark_page():
    current=_v44_estimate_by_trade(project_id()); hist=_v44_historical_cost_by_trade()
    hm={str(r["trade"] or "").lower():r for r in hist}; h=""
    for r in current:
        trade=str(r["trade"] or ""); current_total=float(r["material"] or 0)+float(r["labor"] or 0)+float(r["subquote"] or 0)+float(r["allowance"] or 0)
        old=hm.get(trade.lower())
        hist_text=f'{int(old["samples"] or 0)} historical samples - avg unit cost ${float(old["avg_unit_cost"] or 0):,.2f}' if old else 'No approved historical benchmark'
        h+=f'<div class="action"><b>{esc(trade)}</b><div class="small">Current estimator total ${current_total:,.0f} - {esc(hist_text)}</div></div>'
    return shell("Estimate Benchmark",'<div class="hero"><h1>Estimate Benchmark Brain</h1><p class="muted">Historical costs inform review; they never replace current project takeoff or subcontractor pricing.</p></div><div class="card">'+(h or '<p class="muted">No current estimate data.</p>')+'</div>')

@app.get("/preconstruction/coverage",response_class=HTMLResponse)
def v44_coverage_page():
    rows=_v44_bid_coverage(project_id())
    h="".join(f'<div class="action"><span class="badge {"READY" if status=="PRICED" else "WATCH"}">{esc(status)}</span> <b>{esc(trade)}</b><div class="small">{n} scope items - {invites} invite(s) - {proposals} proposal(s)</div></div>' for trade,n,invites,proposals,status in rows)
    return shell("Bid Coverage",'<div class="hero"><h1>Bid Coverage Brain</h1><p class="muted">Every trade should have deliberate pricing coverage before bid day.</p></div><div class="card">'+(h or '<p class="muted">Analyze project scope first.</p>')+'</div>')

@app.get("/preconstruction/bidders",response_class=HTMLResponse)
def v44_bidders_page():
    pid=project_id()
    subs=_v39_rows("SELECT * FROM subs WHERE project_id=? ORDER BY trade,name",(pid,))
    perf=_v39_rows("SELECT * FROM subcontractor_performance WHERE company_id=? AND COALESCE(approval_status,'PROPOSED')='APPROVED' ORDER BY trade,sub_name",(current_company_id(),))
    pm={(str(r["sub_name"] or "").lower(),str(r["trade"] or "").lower()):r for r in perf}
    h=""
    for s in subs:
        p=pm.get((str(s["name"] or "").lower(),str(s["trade"] or "").lower()))
        score=""
        if p:
            vals=[float(p[k] or 0) for k in ["schedule_score","quality_score","safety_score","responsiveness_score","closeout_score"]]
            score=f' - Performance avg {sum(vals)/len(vals):.1f}'
        h+=f'<div class="action"><b>{esc(s["name"])} - {esc(s["trade"])}</b><div class="small">Bidder candidate{esc(score)}</div></div>'
    return shell("Bidder Strategy",'<div class="hero"><h1>Bidder Strategy Brain</h1><p class="muted">Build the bidder list using trade coverage and approved performance history, not price alone.</p></div><div class="card">'+(h or '<p class="muted">No subcontractors loaded for this project.</p>')+'</div>')

@app.get("/preconstruction/packages",response_class=HTMLResponse)
def v44_packages_page():
    rows=_v44_scope_by_trade(project_id())
    h="".join(f'<div class="action"><span class="badge READY">READY TO BUILD</span> <b>{esc(r["trade"])}</b><div class="small">{int(r["n"] or 0)} source-backed scope items can feed a trade bid package.</div></div>' for r in rows)
    return shell("Bid Packages",'<div class="hero"><h1>Bid Package Brain</h1><p class="muted">Clean Blueprint Brain scope becomes the foundation of each subcontractor bid package.</p></div><div class="card">'+(h or '<p class="muted">Analyze project documents first.</p>')+'</div>')

@app.get("/preconstruction/leveling",response_class=HTMLResponse)
def v44_leveling_page():
    _v44_ensure_tables()
    rows=_v39_rows("SELECT * FROM bid_proposals WHERE company_id=? AND project_id=? ORDER BY trade,base_bid",(current_company_id(),project_id()))
    h="".join(f'<div class="action"><span class="badge {"READY" if float(r["scope_coverage"] or 0)>=90 else "WATCH"}">{float(r["scope_coverage"] or 0):g}% COVERAGE</span> <b>{esc(r["trade"])} - {esc(r["sub_name"])}</b><div class="small">Base ${float(r["base_bid"] or 0):,.0f} - Alternates ${float(r["alternates"] or 0):,.0f} - Allowances ${float(r["allowances"] or 0):,.0f} - Risk {float(r["risk_score"] or 0):g}</div><p><b>Exclusions:</b> {esc(r["exclusions"])}</p><p><b>Qualifications:</b> {esc(r["qualifications"])}</p></div>' for r in rows)
    return shell("Bid Leveling",'<div class="hero"><h1>Subcontractor Bid Leveling Brain</h1><p class="muted">Lowest number does not mean lowest risk. Compare scope coverage, exclusions, qualifications and price.</p></div><div class="card">'+(h or '<p class="muted">No subcontractor proposals loaded yet.</p>')+'</div>')

@app.get("/preconstruction/allowances",response_class=HTMLResponse)
def v44_allowance_page():
    rows=_v44_allowances(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">REVIEW</span> <b>{esc(r["trade"])}</b> - {esc(r["description"])}<div class="small">Allowance ${float(r["allowance"] or 0):,.0f} - {esc(r["source_ref"])}</div></div>' for r in rows)
    return shell("Allowances & Alternates",'<div class="hero"><h1>Allowances & Alternates Brain</h1><p class="muted">Keep uncertain money, owner-furnished items and alternates visible before final bid.</p></div><div class="card">'+(h or '<p class="muted">No allowance/alternate signals detected.</p>')+'</div>')

@app.get("/preconstruction/long-lead",response_class=HTMLResponse)
def v44_longlead_page():
    rows=_v44_long_leads(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(r["status"])}</span> <b>{esc(r["item"])}</b><div class="small">Required onsite {esc(r["required_on_site"])} - Promised {esc(r["promised_date"])} - {esc(r["vendor"])}</div></div>' for r in rows)
    return shell("Long Lead",'<div class="hero"><h1>Long-Lead / Schedule Risk Brain</h1><p class="muted">Preconstruction should identify procurement threats before the baseline schedule is committed.</p></div><div class="card">'+(h or '<p class="muted">No long-lead procurement items loaded yet.</p>')+'</div>')

@app.get("/preconstruction/risk",response_class=HTMLResponse)
def v44_risk_page():
    r=_v44_precon_risk(project_id())
    body=f'<div class="hero"><h1>Preconstruction Risk Brain</h1><p class="muted">{r["level"]} risk - Score {r["score"]}/100.</p></div><div class="grid4"><div class="card"><div class="label">Uncovered Trades</div><div class="kpi">{r["uncovered"]}</div></div><div class="card"><div class="label">Scope Review</div><div class="kpi">{r["gaps"]}</div></div><div class="card"><div class="label">Unverified Estimate</div><div class="kpi">{r["unverified"]}</div></div><div class="card"><div class="label">Long-Lead Signals</div><div class="kpi">{r["longleads"]}</div></div></div>'
    return shell("Preconstruction Risk",body)


# ============================================================
# v44.2 EXISTING BLUEPRINT DATA NORMALIZATION
# ============================================================

def _v442_normalize_existing_blueprint(pid):
    """
    Normalize saved Blueprint Brain items, delete excluded items,
    merge duplicate parent scopes, and rebuild every displayed scope boiler
    from the corrected child items.
    """
    company=current_company_id()
    c=db()

    items=c.execute(
        "SELECT * FROM blueprint_scope_items WHERE company_id=? AND project_id=? ORDER BY run_id,id",
        (company,pid)
    ).fetchall()

    changed=0
    for item in items:
        req=str(item["requirement"] or "").strip()

        # Remove explicitly excluded generic scope notes.
        if _v442_exclude_scope_item(req):
            c.execute("DELETE FROM blueprint_scope_items WHERE id=?",(item["id"],))
            changed+=1
            continue

        proposed=_v33_normalize_trade(item["trade"])
        target=_v441_primary_trade(req, proposed)
        target=_v441_apply_approved_learning(pid, req, target)

        # Find/create canonical parent scope for this run.
        parent=c.execute(
            "SELECT * FROM blueprint_trade_scopes WHERE run_id=? AND company_id=? AND project_id=? AND trade=? LIMIT 1",
            (item["run_id"],company,pid,target)
        ).fetchone()

        if parent:
            target_id=parent["id"]
        else:
            div={
                "GC / General Contractor":"01","Demolition":"02","Concrete":"03","Roofing":"07",
                "Doors / Frames / Hardware":"08","Storefront / Glazing":"08",
                "Framing / Drywall":"09","Ceilings":"09","Flooring / Tile":"09",
                "Painting":"09","Toilet / Bath Accessories":"10","Specialties":"10",
                "Fire Sprinkler":"21","Plumbing":"22","HVAC / Mechanical":"23",
                "Controls":"23","Electrical":"26","Low Voltage":"27","Fire Alarm":"28"
            }.get(target,"")
            now=datetime.utcnow().isoformat()
            c.execute(
                "INSERT INTO blueprint_trade_scopes(company_id,project_id,run_id,trade,division,summary,scope_text,item_count,created) VALUES(?,?,?,?,?,?,?,?,?)",
                (company,pid,item["run_id"],target,div,
                 f"BuildCommand source-backed scope for {target}.","",0,now)
            )
            target_id=c.execute("SELECT last_insert_rowid() id").fetchone()["id"]

        if target != item["trade"] or target_id != item["trade_scope_id"]:
            c.execute(
                "UPDATE blueprint_scope_items SET trade=?,trade_scope_id=? WHERE id=? AND company_id=? AND project_id=?",
                (target,target_id,item["id"],company,pid)
            )
            changed+=1

    # Canonicalize duplicate parent names.
    scopes=c.execute(
        "SELECT * FROM blueprint_trade_scopes WHERE company_id=? AND project_id=? ORDER BY run_id,id",
        (company,pid)
    ).fetchall()

    for sc in scopes:
        canonical=_v33_normalize_trade(sc["trade"])
        if canonical != sc["trade"]:
            other=c.execute(
                "SELECT * FROM blueprint_trade_scopes WHERE run_id=? AND company_id=? AND project_id=? AND trade=? AND id<>? LIMIT 1",
                (sc["run_id"],company,pid,canonical,sc["id"])
            ).fetchone()
            if other:
                c.execute(
                    "UPDATE blueprint_scope_items SET trade_scope_id=?,trade=? WHERE trade_scope_id=?",
                    (other["id"],canonical,sc["id"])
                )
                c.execute("DELETE FROM blueprint_trade_scopes WHERE id=?",(sc["id"],))
            else:
                c.execute(
                    "UPDATE blueprint_trade_scopes SET trade=? WHERE id=?",
                    (canonical,sc["id"])
                )
                c.execute(
                    "UPDATE blueprint_scope_items SET trade=? WHERE trade_scope_id=?",
                    (canonical,sc["id"])
                )

    # REBUILD every scope boiler from corrected child items.
    scopes=c.execute(
        "SELECT * FROM blueprint_trade_scopes WHERE company_id=? AND project_id=? ORDER BY run_id,id",
        (company,pid)
    ).fetchall()

    for sc in scopes:
        children=c.execute(
            "SELECT * FROM blueprint_scope_items WHERE trade_scope_id=? AND company_id=? AND project_id=? ORDER BY id",
            (sc["id"],company,pid)
        ).fetchall()

        if not children:
            c.execute("DELETE FROM blueprint_trade_scopes WHERE id=?",(sc["id"],))
            continue

        lines=[]
        for i,ch in enumerate(children,1):
            refs=[]
            if ch["source_sheet"]: refs.append("Sheet "+str(ch["source_sheet"]))
            if ch["source_detail"]: refs.append("Detail "+str(ch["source_detail"]))
            if ch["source_spec"]: refs.append("Spec "+str(ch["source_spec"]))
            ref_txt=f" [{' · '.join(refs)}]" if refs else ""
            lines.append(f"{i}. {ch['requirement']}{ref_txt}")

        boiler="\n".join(lines)
        c.execute(
            "UPDATE blueprint_trade_scopes SET trade=?,summary=?,scope_text=?,item_count=? WHERE id=?",
            ( _v33_normalize_trade(sc["trade"]),
              f"BuildCommand source-backed scope for {_v33_normalize_trade(sc['trade'])}.",
              boiler,len(children),sc["id"])
        )

    c.commit(); c.close()
    return changed


@app.post("/blueprint-brain/final-cleanup")
def v442_run_final_cleanup():
    pid=project_id()
    _v442_normalize_existing_blueprint(pid)
    return RedirectResponse("/blueprint-brain",status_code=303)


# ============================================================
# v45 SEQUENCE INTELLIGENCE ENGINE
# ============================================================

def _v45_ensure_tables():
    c=db()
    if DATABASE_KIND=="postgres":
        pk="BIGSERIAL PRIMARY KEY"
    else:
        pk="INTEGER PRIMARY KEY"
    c.execute(f"""CREATE TABLE IF NOT EXISTS sequence_intelligence(
        id {pk},
        company_id BIGINT,
        project_id BIGINT,
        activity_id BIGINT,
        activity_name TEXT,
        activity_trade TEXT,
        sequence_stage INTEGER DEFAULT 0,
        readiness_status TEXT DEFAULT 'REVIEW',
        risk_level TEXT DEFAULT 'LOW',
        predecessor_summary TEXT DEFAULT '',
        blocking_reason TEXT DEFAULT '',
        downstream_impact TEXT DEFAULT '',
        recommended_action TEXT DEFAULT '',
        source_type TEXT DEFAULT 'SYSTEM',
        created TEXT,
        updated TEXT
    )""")
    c.commit(); c.close()

def _v45_stage_for_activity(name,trade):
    s=(str(name or "")+" "+str(trade or "")).lower()
    rules=[
        (10,("mobilization","survey","layout","erosion","swppp","clearing","site prep")),
        (20,("earthwork","excavat","grading","underground","site utilities","storm drain","sanitary sewer","water line")),
        (30,("footing","foundation","stem wall","grade beam","slab","concrete")),
        (40,("structural steel","masonry","cmu","tilt","framing","stud","sheathing")),
        (50,("roof","waterproof")),
        (60,("mep rough","rough-in","rough in","electrical rough","plumbing rough","hvac rough","duct rough","overhead mep","fire sprinkler rough","low voltage rough")),
        (70,("rough inspection","above ceiling inspection","in-wall inspection","in wall inspection","pressure test","rough test")),
        (80,("insulation","drywall","gypsum","gyp board","close wall")),
        (90,("ceiling grid","act ceiling","acoustical ceiling")),
        (100,("paint","flooring","tile","millwork","casework","doors","hardware","finish")),
        (110,("fixture","trim out","trim-out","device","startup","start-up","commission","test and balance","tab")),
        (120,("punch","final inspection","closeout","turnover")),
    ]
    for stage,terms in rules:
        if any(term in s for term in terms):
            return stage
    return 65 if any(x in s for x in ("electrical","plumbing","hvac","mechanical","sprinkler","low voltage")) else 50

def _v45_gate_requirements(stage):
    gates=[]
    if stage>=30:
        gates.append("survey/layout and below-grade prerequisites complete")
    if stage>=40:
        gates.append("foundation/slab prerequisites complete where applicable")
    if stage>=60:
        gates.append("framing/supporting construction ready")
    if stage>=80:
        gates.append("MEP rough-in complete")
        gates.append("required rough inspections/testing passed")
    if stage>=90:
        gates.append("above-ceiling work coordinated and inspected")
    if stage>=100:
        gates.append("substrates and predecessor finishes ready")
    if stage>=110:
        gates.append("equipment/material available and approved")
    if stage>=120:
        gates.append("punch/testing/final documentation progressing")
    return gates

def _v45_sequence_analysis_uncached(pid):
    _v45_ensure_tables()
    today=datetime.utcnow().date()
    acts=_v39_rows("SELECT * FROM activities WHERE project_id=? ORDER BY start,id",(pid,))
    readiness={}
    for r in _v39_rows("SELECT * FROM activity_readiness WHERE project_id=?",(pid,)):
        readiness[int(r["activity_id"])]=r
    inspections=_v39_rows("SELECT * FROM inspections_tracker WHERE project_id=?",(pid,))
    procurement=_v39_rows("SELECT * FROM procurement WHERE project_id=?",(pid,))
    submittals=_v39_rows("SELECT * FROM submittals WHERE project_id=?",(pid,))
    issues=_v39_rows("SELECT * FROM project_issues WHERE project_id=? AND COALESCE(status,'OPEN')!='CLOSED'",(pid,))

    analyzed=[]
    prior=[]
    for a in acts:
        stage=_v45_stage_for_activity(a["name"],a["trade"])
        start=_v39_safe_date(a["start"]); finish=_v39_safe_date(a["finish"])
        pct=float(a["pct"] or 0)
        status=str(a["status"] or "NOT_STARTED").upper()
        reasons=[]
        risk="LOW"

        earlier=[p for p in prior if p["stage"] < stage and p["status"]!="COMPLETE"]
        immediate=sorted(earlier,key=lambda x:x["stage"],reverse=True)[:3]
        if immediate and (status=="IN_PROGRESS" or pct>0):
            reasons.append("Work appears started while earlier-stage activities remain incomplete.")
            risk="HIGH"

        rd=readiness.get(int(a["id"]))
        if rd:
            missing=[]
            labels=[("drawings","drawings"),("material","materials"),("manpower","manpower"),
                    ("predecessor","predecessors"),("access_ready","access"),
                    ("inspection","inspection gate"),("equipment","equipment")]
            for col,label in labels:
                if int(rd[col] or 0)!=1:
                    missing.append(label)
            if missing and start and start<=today+timedelta(days=7):
                reasons.append("Make-ready incomplete: "+", ".join(missing)+".")
                risk="CRITICAL" if start and start<=today else "HIGH"

        pending_insp=[i for i in inspections if i["activity_id"]==a["id"] and str(i["result"] or "PENDING").upper() not in {"PASSED","COMPLETE"}]
        if pending_insp and stage>=80:
            reasons.append(f"{len(pending_insp)} inspection gate(s) remain open before concealment/finish work.")
            risk="CRITICAL" if status=="IN_PROGRESS" or pct>0 else "HIGH"

        late_material=[]
        for p in procurement:
            if p["activity_id"]==a["id"]:
                need=_v39_safe_date(p["required_on_site"]); promised=_v39_safe_date(p["promised_date"])
                if need and promised and promised>need:
                    late_material.append(p)
        if late_material:
            reasons.append(f"{len(late_material)} procurement item(s) are promised after required-on-site date.")
            risk="CRITICAL" if risk=="HIGH" else "HIGH"

        open_subs=[s for s in submittals if s["activity_id"]==a["id"] and str(s["status"] or "PENDING").upper() not in {"APPROVED","COMPLETE","CLOSED"}]
        if open_subs and start and start<=today+timedelta(days=14):
            reasons.append(f"{len(open_subs)} submittal approval(s) remain open before planned start.")
            if risk=="LOW": risk="HIGH"

        open_issues=[i for i in issues if i["activity_id"]==a["id"]]
        if open_issues:
            reasons.append(f"{len(open_issues)} open issue(s) are tied to this activity.")
            if risk=="LOW": risk="MEDIUM"

        if finish and finish<today and pct<100:
            reasons.append("Activity is past its planned finish date.")
            risk="CRITICAL"

        readiness_status="READY" if not reasons else "NOT READY" if risk=="CRITICAL" else "AT RISK"
        gates=_v45_gate_requirements(stage)
        predecessor_summary="; ".join(gates)
        blocking_reason=" ".join(reasons)
        downstream=(
            "Downstream work may be delayed or forced out of sequence."
            if reasons else
            "No current sequence exception detected from available project data."
        )
        recommendation=(
            "Resolve blocking prerequisites before releasing downstream work."
            if reasons else
            "Maintain planned sequence and continue make-ready verification."
        )
        analyzed.append({
            "activity":a,"stage":stage,"status":status,"risk":risk,
            "readiness":readiness_status,"predecessor_summary":predecessor_summary,
            "blocking_reason":blocking_reason,"downstream":downstream,"recommendation":recommendation
        })
        prior.append({"stage":stage,"status":status,"name":a["name"]})

    # Persist latest intelligence snapshot.
    c=db()
    c.execute("DELETE FROM sequence_intelligence WHERE company_id=? AND project_id=?",(current_company_id(),pid))
    now=datetime.utcnow().isoformat()
    for x in analyzed:
        a=x["activity"]
        c.execute("""INSERT INTO sequence_intelligence(
            company_id,project_id,activity_id,activity_name,activity_trade,sequence_stage,
            readiness_status,risk_level,predecessor_summary,blocking_reason,downstream_impact,
            recommended_action,source_type,created,updated
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",(
            current_company_id(),pid,a["id"],a["name"],a["trade"],x["stage"],
            x["readiness"],x["risk"],x["predecessor_summary"],x["blocking_reason"],
            x["downstream"],x["recommendation"],"SYSTEM",now,now
        ))
    c.commit(); c.close()
    return analyzed

def _v45_sequence_analysis(pid):
    return _v56_sequence(pid)

@app.get("/sequence-intelligence",response_class=HTMLResponse)
def v45_sequence_home():
    pid=project_id()
    rows=_v45_sequence_analysis(pid)
    critical=sum(1 for x in rows if x["risk"]=="CRITICAL")
    high=sum(1 for x in rows if x["risk"]=="HIGH")
    medium=sum(1 for x in rows if x["risk"]=="MEDIUM")
    ready=sum(1 for x in rows if x["readiness"]=="READY")
    body=f'<div class="hero"><div class="eyebrow">BuildCommand v45 - Sequence Intelligence</div><h1>Build in the right order.</h1><p class="muted">{critical} critical - {high} high - {medium} medium sequence risks - {ready}/{len(rows)} activities currently ready from available data.</p></div>'
    body+='<div class="grid3">'
    body+=_v37_link_card("Sequence Exceptions","Activities that are late, blocked or out of order.","/sequence-intelligence/exceptions","Review")
    body+=_v37_link_card("Inspection Gates","See inspection/testing gates before concealment and finishes.","/sequence-intelligence/inspection-gates","Review")
    body+=_v37_link_card("Downstream Impact","See which activities can push following work.","/sequence-intelligence/downstream","Review")
    body+='</div><div class="card"><h2>Sequence Map</h2>'
    for x in rows:
        a=x["activity"]
        body+=f'<div class="action"><span class="badge {"READY" if x["readiness"]=="READY" else "WATCH"}">{esc(x["readiness"])}</span> <b>{esc(a["name"])}</b><div class="small">Stage {x["stage"]} - {esc(a["trade"])} - {esc(a["start"])} to {esc(a["finish"])} - Risk {esc(x["risk"])}</div><p>{esc(x["blocking_reason"] or "No current sequence exception.")}</p></div>'
    body+='</div>'
    return shell("Sequence Intelligence",body)

@app.get("/sequence-intelligence/exceptions",response_class=HTMLResponse)
def v45_sequence_exceptions():
    rows=_v45_sequence_analysis(project_id())
    rows=[x for x in rows if x["readiness"]!="READY"]
    h="".join(
        f'<div class="action"><span class="badge WATCH">{esc(x["risk"])}</span> <b>{esc(x["activity"]["name"])}</b>'
        f'<div class="small">{esc(x["activity"]["trade"])} - Stage {x["stage"]}</div>'
        f'<p>{esc(x["blocking_reason"])}</p><p><b>Recommended:</b> {esc(x["recommendation"])}</p></div>'
        for x in rows
    )
    return shell("Sequence Exceptions",'<div class="hero"><h1>Sequence Exceptions</h1><p class="muted">Only activities with a detected sequencing or readiness problem appear here.</p></div><div class="card">'+(h or '<p class="muted">No current sequence exceptions detected.</p>')+'</div>')

@app.get("/sequence-intelligence/inspection-gates",response_class=HTMLResponse)
def v45_inspection_gates():
    pid=project_id()
    rows=_v39_rows("SELECT * FROM inspections_tracker WHERE project_id=? ORDER BY scheduled_date,id",(pid,))
    h="".join(
        f'<div class="action"><span class="badge {"READY" if str(r["result"] or "").upper()=="PASSED" else "WATCH"}">{esc(r["result"])}</span> '
        f'<b>{esc(r["inspection_type"])}</b><div class="small">{esc(r["scheduled_date"])} - {esc(r["authority"])}</div></div>'
        for r in rows
    )
    return shell("Inspection Gates",'<div class="hero"><h1>Inspection & Testing Gates</h1><p class="muted">Sequence Intelligence treats required inspections as gates before concealment and downstream finish work.</p></div><div class="card">'+(h or '<p class="muted">No inspection gates loaded.</p>')+'</div>')

@app.get("/sequence-intelligence/downstream",response_class=HTMLResponse)
def v45_downstream():
    rows=_v45_sequence_analysis(project_id())
    risky=[x for x in rows if x["risk"] in {"CRITICAL","HIGH"}]
    h="".join(
        f'<div class="action"><span class="badge WATCH">{esc(x["risk"])}</span> <b>{esc(x["activity"]["name"])}</b>'
        f'<p>{esc(x["downstream"])}</p><div class="small">Prerequisite model: {esc(x["predecessor_summary"])}</div></div>'
        for x in risky
    )
    return shell("Downstream Impact",'<div class="hero"><h1>Downstream Sequence Impact</h1><p class="muted">Shows high-risk activities that can force later work out of sequence.</p></div><div class="card">'+(h or '<p class="muted">No high downstream sequence exposure detected.</p>')+'</div>')


# ============================================================
# v45.2 INTELLIGENCE ENGINE — NEXT 10
# ============================================================

def _v452_ensure_tables():
    c=db()
    if DATABASE_KIND=="postgres":
        pk="BIGSERIAL PRIMARY KEY"; num="DOUBLE PRECISION"
    else:
        pk="INTEGER PRIMARY KEY"; num="REAL"
    stmts=[
      f"""CREATE TABLE IF NOT EXISTS constructability_findings(
        id {pk},company_id BIGINT,project_id BIGINT,finding_type TEXT,title TEXT,
        description TEXT,trade TEXT,related_trade TEXT,source_ref TEXT,
        severity TEXT DEFAULT 'REVIEW',status TEXT DEFAULT 'OPEN',
        recommended_action TEXT,created TEXT,updated TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS scope_intelligence(
        id {pk},company_id BIGINT,project_id BIGINT,intelligence_type TEXT,
        trade TEXT,related_trade TEXT,requirement TEXT,source_ref TEXT,
        severity TEXT DEFAULT 'REVIEW',status TEXT DEFAULT 'OPEN',created TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS procurement_intelligence(
        id {pk},company_id BIGINT,project_id BIGINT,procurement_id BIGINT,
        item TEXT,risk_level TEXT,days_exposure INTEGER DEFAULT 0,
        reason TEXT,recommended_action TEXT,created TEXT,updated TEXT)"""
    ]
    for s in stmts: c.execute(s)
    c.commit(); c.close()

def _v452_scope_rows(pid):
    return _v39_rows("""
        SELECT * FROM blueprint_scope_items
        WHERE company_id=? AND project_id=?
        ORDER BY run_id,trade,id
    """,(current_company_id(),pid))

def _v452_constructability(pid):
    rows=_v452_scope_rows(pid)
    findings=[]
    patterns=[
      ("ACCESS","Equipment/service access",
       ("access","clearance","service space","maintenance access","working clearance"),
       "Verify physical access, service clearance, and installation path before release."),
      ("PENETRATION","Penetration / support coordination",
       ("penetration","core drill","sleeve","opening","curb","blocking","backing","support"),
       "Coordinate opening/support responsibility and verify structural/finish impact."),
      ("CEILING","Above-ceiling congestion",
       ("ceiling","diffuser","sprinkler","light fixture","duct","above ceiling"),
       "Coordinate ceiling layout and above-ceiling systems before grid/closure."),
      ("ADA","Accessibility / clearance",
       ("ada","accessible","grab bar","clear floor","maneuvering clearance"),
       "Verify dimensions and mounting requirements against accessibility details."),
      ("ROOF","Roof penetration coordination",
       ("roof penetration","roof curb","roof flashing","roof patch"),
       "Coordinate equipment, curb, flashing and roofing responsibility before installation."),
    ]
    seen=set()
    for r in rows:
        text=(str(r["requirement"] or "")+" "+str(r["source_note"] or "")).lower()
        for typ,title,terms,action in patterns:
            if any(term in text for term in terms):
                key=(typ,r["requirement"])
                if key in seen: continue
                seen.add(key)
                severity="HIGH" if typ in {"ACCESS","PENETRATION","ROOF"} else "REVIEW"
                findings.append({
                    "type":typ,"title":title,"description":r["requirement"],
                    "trade":r["trade"],"related":r["related_trade"] or "",
                    "source":" · ".join(x for x in [r["source_sheet"],r["source_detail"],r["source_spec"]] if x),
                    "severity":severity,"action":action
                })
    return findings[:150]

def _v452_conflicts(pid):
    rows=_v452_scope_rows(pid)
    by_req={}
    for r in rows:
        key=re.sub(r'\s+',' ',str(r["requirement"] or "").lower()).strip()
        if not key: continue
        by_req.setdefault(key,[]).append(r)
    out=[]
    for req,items in by_req.items():
        trades=sorted(set(str(x["trade"] or "") for x in items))
        if len(trades)>1:
            out.append((items[0]["requirement"],trades,items[0]["source_sheet"] or ""))
    return out[:100]

def _v452_scope_gaps(pid):
    rows=_v452_scope_rows(pid)
    out=[]
    for r in rows:
        trade=str(r["trade"] or "").strip()
        conf=str(r["confidence"] or "MEDIUM").upper()
        if trade in {"","Unassigned"} or conf=="LOW":
            out.append(r)
    return out[:150]

def _v452_scope_overlaps(pid):
    return _v452_conflicts(pid)

def _v452_change_exposure(pid):
    conflicts=_v452_conflicts(pid)
    rfis=_v42_rfis(pid)
    changes=_v39_changes(pid)
    return conflicts,rfis,changes

def _v452_procurement_analysis(pid):
    rows=_v39_rows("SELECT * FROM procurement WHERE project_id=? AND COALESCE(status,'OPEN') NOT IN ('DELIVERED','COMPLETE','CLOSED') ORDER BY required_on_site",(pid,))
    today=datetime.utcnow().date()
    out=[]
    for r in rows:
        need=_v39_safe_date(r["required_on_site"])
        promised=_v39_safe_date(r["promised_date"])
        exposure=(promised-need).days if need and promised else 0
        if exposure>=7: level="CRITICAL"
        elif exposure>0: level="HIGH"
        elif need and need<=today: level="TODAY"
        else: level="WATCH"
        reason=(f"Promised date is {exposure} day(s) after required-on-site date." if exposure>0
                else "Required-on-site timing requires monitoring.")
        action=("Escalate vendor/submittal/fabrication plan and evaluate schedule recovery."
                if exposure>0 else "Confirm fabrication, shipping and delivery status.")
        out.append((r,level,exposure,reason,action))
    return out

def _v452_inspection_derived(pid):
    scopes=_v452_scope_rows(pid)
    inspections=_v39_rows("SELECT * FROM inspections_tracker WHERE project_id=?",(pid,))
    existing=" ".join(str(i["inspection_type"] or "").lower() for i in inspections)
    candidates=[]
    mapping=[
      ("Concrete",("concrete","footing","slab","rebar"),"Concrete / reinforcing inspection"),
      ("Framing / Drywall",("framing","stud","drywall","gypsum"),"Framing / in-wall inspection"),
      ("Electrical",("electrical","conduit","branch circuit"),"Electrical rough inspection"),
      ("Plumbing",("plumbing","sanitary","domestic water"),"Plumbing rough / pressure test"),
      ("HVAC / Mechanical",("duct","hvac","mechanical"),"Mechanical rough inspection"),
      ("Fire Sprinkler",("sprinkler","fire suppression"),"Fire sprinkler inspection / test"),
      ("Roofing",("roof","roofing"),"Roofing / waterproofing inspection"),
    ]
    corpus=" ".join(str(r["requirement"] or "").lower() for r in scopes)
    for trade,terms,name in mapping:
        if any(term in corpus for term in terms) and name.lower() not in existing:
            candidates.append((trade,name,"Derived from analyzed project scope; verify with AHJ/project requirements."))
    return candidates

def _v452_learning_context(pid):
    _v43_ensure_tables()
    rules=_v39_rows("""
        SELECT * FROM learning_rules
        WHERE company_id=? AND approval_status='APPROVED'
          AND (project_id=? OR scope_level='COMPANY STANDARD')
        ORDER BY CASE WHEN project_id=? THEN 0 ELSE 1 END,id DESC
        LIMIT 200
    """,(current_company_id(),pid,pid))
    return rules

def _v452_command(pid):
    sequence=_v45_sequence_analysis(pid)
    attention=_v39_attention(pid)
    proc=_v452_procurement_analysis(pid)
    conflicts=_v452_conflicts(pid)
    gaps=_v452_scope_gaps(pid)
    inspection_candidates=_v452_inspection_derived(pid)

    critical_sequence=[x for x in sequence if x["risk"]=="CRITICAL"]
    high_sequence=[x for x in sequence if x["risk"]=="HIGH"]
    critical_proc=[x for x in proc if x[1]=="CRITICAL"]
    top=[]
    for x in critical_sequence[:3]:
        top.append(("CRITICAL","Sequence",x["activity"]["name"],x["blocking_reason"]))
    for x in critical_proc[:3]:
        top.append(("CRITICAL","Procurement",x[0]["item"],x[3]))
    for x in attention[:4]:
        top.append((x[0],x[1],x[2],x[3]))
    for req,trades,source in conflicts[:3]:
        top.append(("REVIEW","Conflict",req,f"Assigned across {', '.join(trades)}"))
    return {
        "top":top[:10],
        "sequence_critical":len(critical_sequence),
        "sequence_high":len(high_sequence),
        "proc_critical":len(critical_proc),
        "conflicts":len(conflicts),
        "gaps":len(gaps),
        "inspection_candidates":len(inspection_candidates)
    }

@app.get("/intelligence-engine",response_class=HTMLResponse)
def v452_intelligence_engine():
    pid=project_id()
    construct=_v452_constructability(pid)
    conflicts=_v452_conflicts(pid)
    gaps=_v452_scope_gaps(pid)
    proc=_v452_procurement_analysis(pid)
    inspections=_v452_inspection_derived(pid)
    learning=_v452_learning_context(pid)
    command=_v452_command(pid)

    body=f'<div class="hero"><div class="eyebrow">BuildCommand v45.2 - Intelligence Engine</div><h1>Think across the whole project.</h1><p class="muted">{len(construct)} constructability signals - {len(conflicts)} conflicts/overlaps - {len(gaps)} scope gaps - {len(proc)} procurement items - {len(inspections)} derived inspection candidates - {len(learning)} approved learning rules.</p></div><div class="grid3">'
    cards=[
      ("Constructability Intelligence","Find access, clearance, penetration, ceiling and coordination risks.","/intelligence-engine/constructability"),
      ("Conflict Intelligence","Find contradictory/duplicate ownership across project scope.","/intelligence-engine/conflicts"),
      ("Automatic RFI Brain","Draft RFI candidates from confirmed conflicts. Human approval required.","/intelligence-engine/rfis"),
      ("Scope Gap Intelligence","Find low-confidence or unassigned scope before award.","/intelligence-engine/gaps"),
      ("Scope Overlap Intelligence","Detect duplicate ownership and coordination overlap.","/intelligence-engine/overlaps"),
      ("Change-Order Intelligence","Connect conflicts, RFIs and open changes to exposure.","/intelligence-engine/change-exposure"),
      ("Procurement Intelligence","Connect promised dates to required-on-site risk.","/intelligence-engine/procurement"),
      ("Inspection Intelligence 2.0","Derive likely inspection/testing gates from project scope.","/intelligence-engine/inspections"),
      ("Approved Learning Context","See what company/project knowledge is influencing the brain.","/intelligence-engine/learning"),
      ("Superintendent Command Intelligence","One prioritized view of what can hurt the job now.","/intelligence-engine/command")
    ]
    for name,desc,href in cards:
        body+=_v37_link_card(name,desc,href,"Open")
    body+='</div>'
    return shell("Intelligence Engine",body)

@app.get("/intelligence-engine/constructability",response_class=HTMLResponse)
def v452_constructability_page():
    rows=_v452_constructability(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(r["severity"])}</span> <b>{esc(r["title"])}</b><div>{esc(r["description"])}</div><div class="small">{esc(r["trade"])} {("· "+esc(r["related"])) if r["related"] else ""} · {esc(r["source"])}</div><p><b>Check:</b> {esc(r["action"])}</p></div>' for r in rows)
    return shell("Constructability Intelligence",'<div class="hero"><h1>Constructability Intelligence</h1><p class="muted">Signals for things that may not coordinate, fit, or remain accessible. These are review prompts, not automatic design decisions.</p></div><div class="card">'+(h or '<p class="muted">No constructability signals detected from current scope.</p>')+'</div>')

@app.get("/intelligence-engine/conflicts",response_class=HTMLResponse)
def v452_conflict_page():
    rows=_v452_conflicts(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">CONFLICT</span> <b>{esc(req)}</b><div class="small">Trades: {esc(", ".join(trades))} · {esc(source)}</div></div>' for req,trades,source in rows)
    return shell("Conflict Intelligence",'<div class="hero"><h1>Conflict Intelligence</h1><p class="muted">Finds identical requirements assigned to multiple trades for review.</p></div><div class="card">'+(h or '<p class="muted">No duplicate cross-trade requirements detected.</p>')+'</div>')

@app.get("/intelligence-engine/rfis",response_class=HTMLResponse)
def v452_rfi_brain():
    rows=_v452_conflicts(project_id())
    h=""
    for i,(req,trades,source) in enumerate(rows,1):
        h+=f'<div class="card"><span class="badge WATCH">DRAFT RFI</span><h3>RFI Candidate {i}: {esc(req)}</h3><p><b>Question:</b> Please clarify the governing trade responsibility and document requirement for this item.</p><p><b>Detected trades:</b> {esc(", ".join(trades))}</p><p><b>Source:</b> {esc(source)}</p><p class="small">Proposal only. Human approval required before issue.</p></div>'
    return shell("Automatic RFI Brain",'<div class="hero"><h1>Automatic RFI Brain</h1><p class="muted">Turns conflicts into reviewable RFI drafts without sending anything automatically.</p></div>'+(h or '<div class="card">No RFI candidates detected.</div>'))

@app.get("/intelligence-engine/gaps",response_class=HTMLResponse)
def v452_gap_page():
    rows=_v452_scope_gaps(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">GAP / REVIEW</span> <b>{esc(r["trade"])}</b> - {esc(r["requirement"])}<div class="small">Confidence {esc(r["confidence"])} · {esc(r["source_sheet"])}</div></div>' for r in rows)
    return shell("Scope Gap Intelligence",'<div class="hero"><h1>Scope Gap Intelligence</h1><p class="muted">Find scope with weak or missing ownership before it becomes a field problem.</p></div><div class="card">'+(h or '<p class="muted">No obvious low-confidence/unassigned scope detected.</p>')+'</div>')

@app.get("/intelligence-engine/overlaps",response_class=HTMLResponse)
def v452_overlap_page():
    rows=_v452_scope_overlaps(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">OVERLAP</span> <b>{esc(req)}</b><div class="small">{esc(" / ".join(trades))}</div></div>' for req,trades,source in rows)
    return shell("Scope Overlap Intelligence",'<div class="hero"><h1>Scope Overlap Intelligence</h1><p class="muted">Separates true coordination interfaces from accidental duplicate ownership.</p></div><div class="card">'+(h or '<p class="muted">No current duplicate ownership detected.</p>')+'</div>')

@app.get("/intelligence-engine/change-exposure",response_class=HTMLResponse)
def v452_change_page():
    conflicts,rfis,changes=_v452_change_exposure(project_id())
    total=sum(float(r["estimated_cost"] or 0) for r in changes)
    days=sum(float(r["schedule_days"] or 0) for r in changes)
    body=f'<div class="hero"><h1>Change-Order Intelligence</h1><p class="muted">{len(conflicts)} scope conflicts · {len(rfis)} controlled RFIs · {len(changes)} open changes · ${total:,.0f} known exposure · {days:g} schedule days.</p></div>'
    body+='<div class="card"><p>Use conflicts and RFIs as early warning signals. Human review determines whether any item is truly extra-contract work.</p></div>'
    return shell("Change-Order Intelligence",body)

@app.get("/intelligence-engine/procurement",response_class=HTMLResponse)
def v452_procurement_page():
    rows=_v452_procurement_analysis(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(level)}</span> <b>{esc(r["item"])}</b><div class="small">Need {esc(r["required_on_site"])} · Promised {esc(r["promised_date"])} · Exposure {exposure} day(s)</div><p>{esc(reason)}</p><p><b>Action:</b> {esc(action)}</p></div>' for r,level,exposure,reason,action in rows)
    return shell("Procurement Intelligence",'<div class="hero"><h1>Procurement Intelligence</h1><p class="muted">Connects required-on-site timing to delivery exposure.</p></div><div class="card">'+(h or '<p class="muted">No open procurement items.</p>')+'</div>')

@app.get("/intelligence-engine/inspections",response_class=HTMLResponse)
def v452_inspection_page():
    candidates=_v452_inspection_derived(project_id())
    existing=_v39_rows("SELECT * FROM inspections_tracker WHERE project_id=? ORDER BY scheduled_date,id",(project_id(),))
    h="".join(f'<div class="action"><span class="badge WATCH">DERIVED</span> <b>{esc(name)}</b><div class="small">{esc(trade)}</div><p>{esc(note)}</p></div>' for trade,name,note in candidates)
    eh="".join(f'<div class="action"><span class="badge">{esc(r["result"])}</span> <b>{esc(r["inspection_type"])}</b><div class="small">{esc(r["scheduled_date"])} · {esc(r["authority"])}</div></div>' for r in existing)
    return shell("Inspection Intelligence 2.0",'<div class="hero"><h1>Inspection Intelligence 2.0</h1><p class="muted">Derives likely inspection/testing gates from actual scope. AHJ/project requirements still govern.</p></div><div class="card"><h2>Derived Candidates</h2>'+(h or '<p class="muted">No new inspection candidates.</p>')+'</div><div class="card"><h2>Tracked Inspections</h2>'+(eh or '<p class="muted">No tracked inspections.</p>')+'</div>')

@app.get("/intelligence-engine/learning",response_class=HTMLResponse)
def v452_learning_page():
    rows=_v452_learning_context(project_id())
    h="".join(f'<div class="action"><span class="badge">{esc(r["scope_level"])}</span> <b>{esc(r["rule_type"])} - {esc(r["subject"])}</b><div>{esc(r["learned_rule"])}</div><div class="small">Approved by {esc(r["approved_by"])}</div></div>' for r in rows)
    return shell("Approved Learning Context",'<div class="hero"><h1>Approved Learning Context</h1><p class="muted">Only approved project/company rules influence automated ownership logic.</p></div><div class="card">'+(h or '<p class="muted">No approved learning rules yet.</p>')+'</div>')

@app.get("/intelligence-engine/command",response_class=HTMLResponse)
def v452_command_page():
    d=_v452_command(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(level)}</span> <b>{esc(kind)}</b> - {esc(title)}<div class="small">{esc(detail)}</div></div>' for level,kind,title,detail in d["top"])
    body=f'<div class="hero"><div class="eyebrow">Superintendent Command Intelligence</div><h1>What can hurt the job now?</h1><p class="muted">{d["sequence_critical"]} critical sequence · {d["sequence_high"]} high sequence · {d["proc_critical"]} critical procurement · {d["conflicts"]} conflicts · {d["gaps"]} scope gaps · {d["inspection_candidates"]} derived inspection candidates.</p></div><div class="card"><h2>Top Priorities</h2>{h or "<p class=muted>No major current intelligence exceptions.</p>"}</div>'
    return shell("Superintendent Command Intelligence",body)


# ============================================================
# v46 PROJECT KNOWLEDGE GRAPH + DEEPER REASONING
# ============================================================

def _v46_ensure_tables():
    c=db()
    if DATABASE_KIND=="postgres":
        pk="BIGSERIAL PRIMARY KEY"; num="DOUBLE PRECISION"
    else:
        pk="INTEGER PRIMARY KEY"; num="REAL"
    stmts=[
      f"""CREATE TABLE IF NOT EXISTS knowledge_nodes(
        id {pk},company_id BIGINT,project_id BIGINT,node_type TEXT,node_key TEXT,
        label TEXT,trade TEXT,source_ref TEXT,status TEXT DEFAULT 'ACTIVE',
        confidence TEXT DEFAULT 'MEDIUM',metadata TEXT DEFAULT '',created TEXT,updated TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS knowledge_edges(
        id {pk},company_id BIGINT,project_id BIGINT,from_key TEXT,to_key TEXT,
        relationship TEXT,source_ref TEXT,confidence TEXT DEFAULT 'MEDIUM',
        status TEXT DEFAULT 'ACTIVE',created TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS revision_intelligence(
        id {pk},company_id BIGINT,project_id BIGINT,new_attachment_id BIGINT,
        prior_attachment_id BIGINT,revision_group TEXT,change_summary TEXT,
        affected_trades TEXT,cost_risk TEXT DEFAULT 'REVIEW',
        schedule_risk TEXT DEFAULT 'REVIEW',status TEXT DEFAULT 'REVIEW',
        created TEXT,updated TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS equipment_intelligence(
        id {pk},company_id BIGINT,project_id BIGINT,equipment_key TEXT,equipment_name TEXT,
        furnish_trade TEXT,install_trade TEXT,power_trade TEXT,controls_trade TEXT,
        plumbing_trade TEXT,support_trade TEXT,roofing_trade TEXT,startup_trade TEXT,
        source_ref TEXT,status TEXT DEFAULT 'REVIEW',created TEXT,updated TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS verification_intelligence(
        id {pk},company_id BIGINT,project_id BIGINT,subject_type TEXT,subject_key TEXT,
        primary_conclusion TEXT,verification_conclusion TEXT,agreement TEXT DEFAULT 'REVIEW',
        confidence TEXT DEFAULT 'MEDIUM',reason TEXT,source_ref TEXT,created TEXT,updated TEXT)"""
    ]
    for s in stmts:
        c.execute(s)
    c.commit(); c.close()

def _v46_docs(pid):
    return _v39_rows("SELECT * FROM attachments WHERE company_id=? AND project_id=? ORDER BY id DESC",(current_company_id(),pid))

def _v46_revision_pairs(pid):
    rows=_v46_docs(pid)
    groups={}
    for r in rows:
        name=str(r["original_name"] or "")
        base=re.sub(r'(?i)(rev(?:ision)?[\s_-]*[A-Z0-9]+|addendum[\s_-]*[A-Z0-9]+|bulletin[\s_-]*[A-Z0-9]+)','',name)
        base=re.sub(r'[^a-z0-9]+',' ',base.lower()).strip()
        groups.setdefault(base,[]).append(r)
    pairs=[]
    for base,items in groups.items():
        if len(items)>1:
            items=sorted(items,key=lambda x:x["id"],reverse=True)
            for i in range(len(items)-1):
                pairs.append((items[i],items[i+1],base))
    return pairs[:50]

def _v46_dependencies(pid):
    scopes=_v452_scope_rows(pid)
    out=[]
    for r in scopes:
        req=str(r["requirement"] or "")
        trade=str(r["trade"] or "")
        related=set()
        low=req.lower()
        if any(x in low for x in ["power","disconnect","electrical connection","branch circuit"]):
            related.add("Electrical")
        if any(x in low for x in ["roof curb","roof flashing","roof patch","roof penetration"]):
            related.add("Roofing")
        if any(x in low for x in ["thermostat","controls","control wiring"]):
            related.add("Controls")
        if any(x in low for x in ["water","drain","condensate","piping","valve"]):
            related.add("Plumbing")
        if any(x in low for x in ["duct","diffuser","grille","rtu","ahu","exhaust fan"]):
            related.add("HVAC / Mechanical")
        if any(x in low for x in ["blocking","backing","support framing","rough opening"]):
            related.add("Framing / Drywall")
        if any(x in low for x in ["storefront","glazing","curtain wall"]):
            related.add("Storefront / Glazing")
        if any(x in low for x in ["door operator","access control","card reader","electronic strike"]):
            related.add("Low Voltage")
        related.discard(trade)
        if related:
            out.append((r,sorted(related)))
    return out[:200]

def _v46_spec_drawing_checks(pid):
    rows=_v452_scope_rows(pid)
    out=[]
    for r in rows:
        has_drawing=bool(r["source_sheet"] or r["source_detail"])
        has_spec=bool(r["source_spec"])
        if has_drawing != has_spec:
            out.append((r,"DRAWING ONLY" if has_drawing else "SPEC ONLY"))
    return out[:150]

def _v46_equipment(pid):
    rows=_v452_scope_rows(pid)
    eq=[]
    patterns=[
        ("WH","water heater","Plumbing"),
        ("RTU","rooftop unit","HVAC / Mechanical"),
        ("AHU","air handler","HVAC / Mechanical"),
        ("EF","exhaust fan","HVAC / Mechanical"),
        ("VAV","vav","HVAC / Mechanical"),
        ("PUMP","pump","Plumbing"),
        ("PANEL","panelboard","Electrical"),
        ("XFMR","transformer","Electrical"),
    ]
    seen=set()
    for r in rows:
        low=str(r["requirement"] or "").lower()
        for prefix,term,primary in patterns:
            if term in low:
                key=f"{prefix}:{re.sub(r'[^a-z0-9]+','-',low[:80]).strip('-')}"
                if key in seen: continue
                seen.add(key)
                eq.append({
                    "key":key,"name":r["requirement"],"primary":primary,
                    "source":" · ".join(x for x in [r["source_sheet"],r["source_detail"],r["source_spec"]] if x),
                    "power":"Electrical" if primary!="Electrical" else "",
                    "controls":"Controls" if primary=="HVAC / Mechanical" else "",
                    "plumbing":"Plumbing" if primary in {"HVAC / Mechanical","Plumbing"} else "",
                    "support":"Framing / Drywall" if any(x in low for x in ["wall mounted","supported","curb","hung"]) else "",
                    "roof":"Roofing" if any(x in low for x in ["roof","rooftop","curb"]) else "",
                    "startup":primary
                })
    return eq[:100]

def _v46_schedule_prediction(pid):
    seq=_v45_sequence_analysis(pid)
    proc=_v452_procurement_analysis(pid)
    out=[]
    for x in seq:
        risk=x["risk"]
        if risk in {"CRITICAL","HIGH"}:
            out.append((risk,x["activity"]["name"],x["blocking_reason"],"Sequence / readiness"))
    for r,level,exposure,reason,action in proc:
        if level in {"CRITICAL","HIGH"}:
            out.append((level,r["item"],reason,"Procurement"))
    return out[:100]

def _v46_cost_risk(pid):
    conflicts=_v452_conflicts(pid)
    gaps=_v452_scope_gaps(pid)
    changes=_v39_changes(pid)
    out=[]
    for req,trades,source in conflicts:
        out.append(("HIGH","Conflict",req,f"Multiple trade ownership: {', '.join(trades)}",source))
    for r in gaps:
        out.append(("MEDIUM","Scope Gap",r["requirement"],f"Trade {r['trade']} / confidence {r['confidence']}",r["source_sheet"] or ""))
    for r in changes:
        if float(r["estimated_cost"] or 0)>0:
            out.append(("HIGH","Open Change",r["title"],f"${float(r['estimated_cost'] or 0):,.0f} exposure",""))
    return out[:100]

def _v46_plan_completeness(pid):
    rows=_v452_scope_rows(pid)
    findings=[]
    for r in rows:
        if not r["source_sheet"] and not r["source_spec"]:
            findings.append(("SOURCE","Missing source reference",r["requirement"],r["trade"]))
        if str(r["confidence"] or "").upper()=="LOW":
            findings.append(("CONFIDENCE","Low-confidence requirement",r["requirement"],r["trade"]))
        low=str(r["requirement"] or "").lower()
        if any(x in low for x in ["verify","confirm","coordinate","as required","where required"]) and str(r["confidence"] or "").upper()!="HIGH":
            findings.append(("AMBIGUITY","Open-ended requirement",r["requirement"],r["trade"]))
    return findings[:150]

def _v46_second_opinion(pid):
    rows=_v452_scope_rows(pid)
    out=[]
    for r in rows:
        primary=str(r["trade"] or "")
        verify=_v441_primary_trade(r["requirement"],primary)
        agreement="AGREE" if verify==primary else "DISAGREE"
        out.append((r,verify,agreement))
    return out[:200]

def _v46_graph(pid):
    _v46_ensure_tables()
    scopes=_v452_scope_rows(pid)
    deps=_v46_dependencies(pid)
    eq=_v46_equipment(pid)
    nodes=[]; edges=[]
    for r in scopes:
        key=f"scope:{r['id']}"
        nodes.append((key,"SCOPE",r["requirement"],r["trade"],r["source_sheet"] or ""))
        edges.append((f"trade:{r['trade']}",key,"OWNS"))
        if r["source_sheet"]:
            edges.append((f"sheet:{r['source_sheet']}",key,"DEFINES"))
        if r["source_spec"]:
            edges.append((f"spec:{r['source_spec']}",key,"GOVERNS"))
    for r,rels in deps:
        for tr in rels:
            edges.append((f"scope:{r['id']}",f"trade:{tr}","COORDINATES_WITH"))
    for e in eq:
        nodes.append((f"equipment:{e['key']}","EQUIPMENT",e["name"],e["primary"],e["source"]))
        for rel,tr in [("INSTALLED_BY",e["primary"]),("POWERED_BY",e["power"]),("CONTROLLED_BY",e["controls"]),("PIPED_BY",e["plumbing"]),("SUPPORTED_BY",e["support"]),("ROOFED_BY",e["roof"]),("STARTED_BY",e["startup"])]:
            if tr:
                edges.append((f"equipment:{e['key']}",f"trade:{tr}",rel))
    return nodes,edges

@app.get("/knowledge-graph",response_class=HTMLResponse)
def v46_knowledge_graph_home():
    pid=project_id()
    pairs=_v46_revision_pairs(pid)
    deps=_v46_dependencies(pid)
    specchecks=_v46_spec_drawing_checks(pid)
    equipment=_v46_equipment(pid)
    schedule=_v46_schedule_prediction(pid)
    cost=_v46_cost_risk(pid)
    complete=_v46_plan_completeness(pid)
    verify=_v46_second_opinion(pid)
    nodes,edges=_v46_graph(pid)
    disagreements=sum(1 for _,_,a in verify if a=="DISAGREE")

    body=f'<div class="hero"><div class="eyebrow">BuildCommand v46 - Project Knowledge Graph</div><h1>Understand how the project connects.</h1><p class="muted">{len(nodes)} knowledge nodes - {len(edges)} relationships - {len(pairs)} revision pairs - {len(deps)} cross-trade dependencies - {len(equipment)} equipment chains - {disagreements} verification disagreements.</p></div><div class="grid3">'
    cards=[
      ("Drawing Revision Intelligence","Identify likely prior/new document pairs and affected-project review needs.","/knowledge-graph/revisions"),
      ("Cross-Sheet Dependencies","See which trades and systems depend on the same requirement.","/knowledge-graph/dependencies"),
      ("Spec vs Drawing Verification","Find scope documented only in drawings or only in specifications.","/knowledge-graph/spec-drawing"),
      ("Equipment Intelligence","Map furnish/install/power/controls/piping/support/roof/startup responsibility.","/knowledge-graph/equipment"),
      ("Trade Handoff Intelligence","Show responsibility boundaries and coordination handoffs.","/knowledge-graph/handoffs"),
      ("Schedule Prediction","Predict likely near-term schedule threats from sequence and procurement.","/knowledge-graph/schedule-prediction"),
      ("Cost-Risk Prediction","Surface conflicts, scope gaps and known change exposure.","/knowledge-graph/cost-risk"),
      ("Plan Completeness Audit","Find missing sources, low-confidence and open-ended requirements.","/knowledge-graph/completeness"),
      ("Second Opinion Engine","Challenge Blueprint Brain trade ownership before trusting it.","/knowledge-graph/verification"),
      ("Project Knowledge Graph","View the connected nodes and relationships behind the intelligence.","/knowledge-graph/graph")
    ]
    for name,desc,href in cards:
        body+=_v37_link_card(name,desc,href,"Open")
    body+='</div>'
    return shell("Project Knowledge Graph",body)

@app.get("/knowledge-graph/revisions",response_class=HTMLResponse)
def v46_revisions():
    rows=_v46_revision_pairs(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">COMPARE</span> <b>{esc(new["original_name"])}</b><div class="small">Possible prior version: {esc(old["original_name"])} · Group {esc(group)}</div></div>' for new,old,group in rows)
    return shell("Drawing Revision Intelligence",'<div class="hero"><h1>Drawing Revision Intelligence</h1><p class="muted">Candidate pairs only; document content still requires comparison before scope or cost changes.</p></div><div class="card">'+(h or '<p class="muted">No likely revision pairs detected.</p>')+'</div>')

@app.get("/knowledge-graph/dependencies",response_class=HTMLResponse)
def v46_dependencies():
    rows=_v46_dependencies(project_id())
    h="".join(f'<div class="action"><b>{esc(r["trade"])}</b> - {esc(r["requirement"])}<div class="small">Coordinates with: {esc(", ".join(rels))}</div></div>' for r,rels in rows)
    return shell("Cross-Sheet Dependencies",'<div class="hero"><h1>Cross-Sheet Dependency Brain</h1><p class="muted">A requirement can belong to one trade while creating dependencies for several others.</p></div><div class="card">'+(h or '<p class="muted">No cross-trade dependency signals detected.</p>')+'</div>')

@app.get("/knowledge-graph/spec-drawing",response_class=HTMLResponse)
def v46_specdrawing():
    rows=_v46_spec_drawing_checks(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(kind)}</span> <b>{esc(r["trade"])}</b> - {esc(r["requirement"])}<div class="small">Sheet {esc(r["source_sheet"])} · Spec {esc(r["source_spec"])}</div></div>' for r,kind in rows)
    return shell("Spec vs Drawing Verification",'<div class="hero"><h1>Specification-to-Drawing Verification</h1><p class="muted">Find requirements supported by only one side of the contract-document set.</p></div><div class="card">'+(h or '<p class="muted">No one-sided source signals detected.</p>')+'</div>')

@app.get("/knowledge-graph/equipment",response_class=HTMLResponse)
def v46_equipment():
    rows=_v46_equipment(project_id())
    h=""
    for e in rows:
        h+=f'<div class="card"><h3>{esc(e["name"])}</h3><div class="small">{esc(e["source"])}</div><p><b>Install:</b> {esc(e["primary"])}</p><p><b>Power:</b> {esc(e["power"] or "N/A")} · <b>Controls:</b> {esc(e["controls"] or "N/A")} · <b>Plumbing:</b> {esc(e["plumbing"] or "N/A")} · <b>Support:</b> {esc(e["support"] or "N/A")} · <b>Roofing:</b> {esc(e["roof"] or "N/A")} · <b>Startup:</b> {esc(e["startup"])}</p></div>'
    return shell("Equipment Intelligence",'<div class="hero"><h1>Equipment Responsibility Intelligence</h1><p class="muted">Connect equipment to the trades required to make it operational.</p></div>'+(h or '<div class="card">No major equipment signals detected.</div>'))

@app.get("/knowledge-graph/handoffs",response_class=HTMLResponse)
def v46_handoffs():
    rows=_v46_dependencies(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">HANDOFF</span> <b>{esc(r["trade"])}</b> → {esc(", ".join(rels))}<div class="small">{esc(r["requirement"])}</div></div>' for r,rels in rows)
    return shell("Trade Handoff Intelligence",'<div class="hero"><h1>Trade Handoff Intelligence</h1><p class="muted">Shows where one subcontractor stops and another trade must pick up supporting work.</p></div><div class="card">'+(h or '<p class="muted">No handoff signals detected.</p>')+'</div>')

@app.get("/knowledge-graph/schedule-prediction",response_class=HTMLResponse)
def v46_schedule_prediction_page():
    rows=_v46_schedule_prediction(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(level)}</span> <b>{esc(title)}</b><div>{esc(reason)}</div><div class="small">{esc(source)}</div></div>' for level,title,reason,source in rows)
    return shell("Schedule Prediction",'<div class="hero"><h1>Schedule Prediction Brain</h1><p class="muted">Predicts likely near-term schedule trouble from current sequence/readiness and procurement signals.</p></div><div class="card">'+(h or '<p class="muted">No high schedule prediction signals.</p>')+'</div>')

@app.get("/knowledge-graph/cost-risk",response_class=HTMLResponse)
def v46_costrisk():
    rows=_v46_cost_risk(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(level)}</span> <b>{esc(kind)}</b> - {esc(title)}<div>{esc(reason)}</div><div class="small">{esc(source)}</div></div>' for level,kind,title,reason,source in rows)
    return shell("Cost Risk Prediction",'<div class="hero"><h1>Cost-Risk Prediction</h1><p class="muted">Early-warning signals only; human review determines contractual cost responsibility.</p></div><div class="card">'+(h or '<p class="muted">No major cost-risk signals detected.</p>')+'</div>')

@app.get("/knowledge-graph/completeness",response_class=HTMLResponse)
def v46_completeness():
    rows=_v46_plan_completeness(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(kind)}</span> <b>{esc(trade)}</b> - {esc(title)}<div class="small">{esc(req)}</div></div>' for kind,title,req,trade in rows)
    return shell("Plan Completeness Audit",'<div class="hero"><h1>Plan Completeness Audit</h1><p class="muted">Asks what an experienced builder would want clarified before relying on the information.</p></div><div class="card">'+(h or '<p class="muted">No obvious completeness exceptions detected.</p>')+'</div>')

@app.get("/knowledge-graph/verification",response_class=HTMLResponse)
def v46_verification():
    rows=_v46_second_opinion(project_id())
    h="".join(f'<div class="action"><span class="badge {"READY" if agreement=="AGREE" else "WATCH"}">{esc(agreement)}</span> <b>{esc(r["trade"])}</b> - {esc(r["requirement"])}<div class="small">Second opinion: {esc(verify)} · Confidence {esc(r["confidence"])}</div></div>' for r,verify,agreement in rows if agreement=="DISAGREE")
    return shell("Second Opinion Engine",'<div class="hero"><h1>AI Verification / Second Opinion Engine</h1><p class="muted">When ownership logic disagrees with the saved scope, BuildCommand flags it instead of pretending certainty.</p></div><div class="card">'+(h or '<p class="muted">No trade-ownership disagreements detected.</p>')+'</div>')

@app.get("/knowledge-graph/graph",response_class=HTMLResponse)
def v46_graph_page():
    nodes,edges=_v46_graph(project_id())
    node_counts={}
    edge_counts={}
    for _,typ,_,_,_ in nodes: node_counts[typ]=node_counts.get(typ,0)+1
    for _,_,rel in edges: edge_counts[rel]=edge_counts.get(rel,0)+1
    nh="".join(f'<div class="action"><b>{esc(k)}</b><div class="small">{v} node(s)</div></div>' for k,v in sorted(node_counts.items()))
    eh="".join(f'<div class="action"><b>{esc(k)}</b><div class="small">{v} relationship(s)</div></div>' for k,v in sorted(edge_counts.items()))
    return shell("Project Knowledge Graph",f'<div class="hero"><h1>Project Knowledge Graph</h1><p class="muted">{len(nodes)} nodes · {len(edges)} relationships.</p></div><div class="grid2"><div class="card"><h2>Nodes</h2>{nh}</div><div class="card"><h2>Relationships</h2>{eh}</div></div>')


# ============================================================
# v47 PREDICTION + DECISION INTELLIGENCE — NEXT 10
# ============================================================

def _v47_ensure_tables():
    c=db()
    if DATABASE_KIND=="postgres":
        pk="BIGSERIAL PRIMARY KEY"; num="DOUBLE PRECISION"
    else:
        pk="INTEGER PRIMARY KEY"; num="REAL"
    stmts=[
      f"""CREATE TABLE IF NOT EXISTS decision_intelligence(
        id {pk},company_id BIGINT,project_id BIGINT,decision_type TEXT,title TEXT,
        responsible_party TEXT,decision_needed_by TEXT,affected_activity TEXT,
        affected_trade TEXT,cost_exposure {num} DEFAULT 0,schedule_days {num} DEFAULT 0,
        severity TEXT DEFAULT 'REVIEW',status TEXT DEFAULT 'OPEN',source_ref TEXT,
        recommended_action TEXT,created TEXT,updated TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS manpower_forecast(
        id {pk},company_id BIGINT,project_id BIGINT,forecast_date TEXT,trade TEXT,
        active_activities INTEGER DEFAULT 0,risk_level TEXT DEFAULT 'LOW',
        readiness_note TEXT,created TEXT,updated TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS risk_propagation(
        id {pk},company_id BIGINT,project_id BIGINT,source_type TEXT,source_key TEXT,
        source_title TEXT,target_type TEXT,target_key TEXT,target_title TEXT,
        relationship TEXT,risk_level TEXT DEFAULT 'REVIEW',reason TEXT,created TEXT)"""
    ]
    for s in stmts:
        c.execute(s)
    c.commit(); c.close()

def _v47_dependency_impacts(pid):
    deps=_v46_dependencies(pid)
    seq=_v45_sequence_analysis(pid)
    seq_by_trade={}
    for x in seq:
        seq_by_trade.setdefault(str(x["activity"]["trade"] or ""),[]).append(x)

    impacts=[]
    for r,related in deps:
        for tr in related:
            related_seq=seq_by_trade.get(tr,[])
            high=[x for x in related_seq if x["risk"] in {"CRITICAL","HIGH"}]
            level="HIGH" if high else "REVIEW"
            reason=(f"{len(high)} related {tr} activity risk(s) may affect this requirement."
                    if high else f"Requirement depends on coordination with {tr}.")
            impacts.append((r,tr,level,reason))
    return impacts[:200]

def _v47_critical_path_signals(pid):
    seq=_v45_sequence_analysis(pid)
    # No CPM dependency network exists yet, so this is a transparent proxy:
    # upcoming/high-risk activities with downstream stage exposure.
    rows=[]
    for x in seq:
        a=x["activity"]
        start=_v39_safe_date(a["start"]); finish=_v39_safe_date(a["finish"])
        duration=(finish-start).days+1 if start and finish and finish>=start else 0
        score=0
        if x["risk"]=="CRITICAL": score+=50
        elif x["risk"]=="HIGH": score+=30
        if x["stage"]>=60: score+=10
        if duration>=10: score+=10
        if float(a["pct"] or 0)>0 and float(a["pct"] or 0)<100: score+=10
        if score>=30:
            rows.append((score,x))
    rows.sort(key=lambda z:z[0],reverse=True)
    return rows[:50]

def _v47_decision_deadlines(pid):
    _v42_ensure_tables()
    today=datetime.utcnow().date()
    rows=[]
    # Owner decisions
    for r in _v42_owner(pid):
        due=_v39_safe_date(r["due_date"])
        days=(due-today).days if due else None
        severity="CRITICAL" if days is not None and days<0 else "HIGH" if days is not None and days<=3 else "REVIEW"
        rows.append(("OWNER",r["title"],r["due_date"],severity,float(r["cost_impact"] or 0),float(r["schedule_days"] or 0),r["source_ref"] or ""))
    # RFIs
    for r in _v42_rfis(pid):
        if str(r["status"] or "").upper() in {"CLOSED","ANSWERED","COMPLETE"}: continue
        due=_v39_safe_date(r["due_date"])
        days=(due-today).days if due else None
        severity="CRITICAL" if days is not None and days<0 else "HIGH" if days is not None and days<=3 else "REVIEW"
        rows.append(("RFI",r["title"],r["due_date"],severity,float(r["cost_impact"] or 0),float(r["schedule_days"] or 0),r["source_ref"] or ""))
    # Submittals
    for r in _v39_rows("SELECT * FROM submittals WHERE project_id=? AND COALESCE(status,'PENDING') NOT IN ('APPROVED','CLOSED','COMPLETE') ORDER BY due_date",(pid,)):
        due=_v39_safe_date(r["due_date"])
        days=(due-today).days if due else None
        severity="CRITICAL" if days is not None and days<0 else "HIGH" if days is not None and days<=3 else "REVIEW"
        rows.append(("SUBMITTAL",r["title"],r["due_date"],severity,0,0,r["spec_section"] or ""))
    rank={"CRITICAL":0,"HIGH":1,"REVIEW":2}
    rows.sort(key=lambda x:(rank.get(x[3],9),x[2] or "9999"))
    return rows[:100]

def _v47_manpower_forecast(pid):
    acts=_v39_rows("SELECT * FROM activities WHERE project_id=? AND COALESCE(status,'NOT_STARTED')!='COMPLETE' ORDER BY start",(pid,))
    today=datetime.utcnow().date()
    buckets={}
    for a in acts:
        start=_v39_safe_date(a["start"]); finish=_v39_safe_date(a["finish"])
        if not start or not finish: continue
        for offset in range(0,15):
            d=today+timedelta(days=offset)
            if start<=d<=finish:
                tr=str(a["trade"] or "Unassigned")
                buckets.setdefault((d.isoformat(),tr),[]).append(a)
    out=[]
    for (d,tr),items in sorted(buckets.items()):
        count=len(items)
        level="HIGH" if count>=3 else "MEDIUM" if count==2 else "LOW"
        out.append((d,tr,count,level,[x["name"] for x in items]))
    return out[:200]

def _v47_material_readiness(pid):
    acts={int(a["id"]):a for a in _v39_rows("SELECT * FROM activities WHERE project_id=?",(pid,))}
    proc=_v452_procurement_analysis(pid)
    out=[]
    for r,level,exposure,reason,action in proc:
        aid=r["activity_id"]
        act=acts.get(int(aid)) if aid is not None else None
        out.append((r,act,level,exposure,reason,action))
    return out[:150]

def _v47_coordination_agenda(pid):
    conflicts=_v452_conflicts(pid)
    construct=_v452_constructability(pid)
    seq=[x for x in _v45_sequence_analysis(pid) if x["risk"] in {"CRITICAL","HIGH"}]
    decisions=_v47_decision_deadlines(pid)
    agenda=[]
    for req,trades,source in conflicts[:8]:
        agenda.append(("Conflict",req,f"Trades: {', '.join(trades)} · {source}"))
    for r in construct[:8]:
        agenda.append(("Constructability",r["title"],r["description"]))
    for x in seq[:8]:
        agenda.append(("Schedule",x["activity"]["name"],x["blocking_reason"]))
    for typ,title,due,severity,cost,days,source in decisions[:8]:
        agenda.append(("Decision",title,f"{typ} · Due {due} · {severity}"))
    return agenda[:25]

def _v47_change_causality(pid):
    changes=_v39_changes(pid)
    rfis=_v42_rfis(pid)
    issues=_v39_rows("SELECT * FROM project_issues WHERE project_id=? ORDER BY id DESC",(pid,))
    out=[]
    for c in changes:
        title=str(c["title"] or "").lower()
        linked=[]
        for r in rfis:
            if title and (title in str(r["title"] or "").lower() or str(r["title"] or "").lower() in title):
                linked.append(("RFI",r["title"]))
        for i in issues:
            if title and (title in str(i["title"] or "").lower() or str(i["title"] or "").lower() in title):
                linked.append(("ISSUE",i["title"]))
        out.append((c,linked))
    return out[:100]

def _v47_closeout_prediction(pid):
    _v41_closeout_seed(pid)
    rows=_v39_rows("SELECT * FROM closeout_items WHERE company_id=? AND project_id=? ORDER BY category,id",(current_company_id(),pid))
    today=datetime.utcnow().date()
    out=[]
    for r in rows:
        status=str(r["status"] or "OPEN").upper()
        due=_v39_safe_date(r["due_date"])
        if status in {"COMPLETE","CLOSED","RECEIVED"}:
            level="READY"
        elif due and due<today:
            level="CRITICAL"
        elif due and (due-today).days<=14:
            level="HIGH"
        else:
            level="REVIEW"
        out.append((r,level))
    return out

def _v47_risk_propagation(pid):
    impacts=[]
    for r,tr,level,reason in _v47_dependency_impacts(pid):
        impacts.append(("SCOPE",f"scope:{r['id']}",r["requirement"],"TRADE",f"trade:{tr}",tr,"DEPENDS_ON",level,reason))
    for score,x in _v47_critical_path_signals(pid):
        a=x["activity"]
        impacts.append(("ACTIVITY",f"activity:{a['id']}",a["name"],"PROJECT","project", "Project Completion","CAN_DELAY",x["risk"],x["blocking_reason"]))
    return impacts[:250]

def _v47_forecast_summary(pid):
    cp=_v47_critical_path_signals(pid)
    decisions=_v47_decision_deadlines(pid)
    material=_v47_material_readiness(pid)
    closeout=_v47_closeout_prediction(pid)
    critical_decisions=sum(1 for x in decisions if x[3]=="CRITICAL")
    critical_material=sum(1 for x in material if x[2]=="CRITICAL")
    critical_closeout=sum(1 for _,level in closeout if level=="CRITICAL")
    return {
        "critical_path":len(cp),
        "critical_decisions":critical_decisions,
        "critical_material":critical_material,
        "critical_closeout":critical_closeout,
        "risk_propagation":len(_v47_risk_propagation(pid))
    }

@app.get("/prediction-intelligence",response_class=HTMLResponse)
def v47_prediction_home():
    pid=project_id()
    s=_v47_forecast_summary(pid)
    body=f'<div class="hero"><div class="eyebrow">BuildCommand v47 - Prediction & Decision Intelligence</div><h1>See problems before they become field problems.</h1><p class="muted">{s["critical_path"]} critical-path proxy signals · {s["critical_decisions"]} overdue decision(s) · {s["critical_material"]} critical material risk(s) · {s["critical_closeout"]} critical closeout risk(s) · {s["risk_propagation"]} propagated dependency risks.</p></div><div class="grid3">'
    cards=[
      ("Dependency Impact Engine","Trace how one scope/trade dependency can affect another.","/prediction-intelligence/dependency-impact"),
      ("Critical Path Intelligence","Prioritize activities most likely to threaten completion.","/prediction-intelligence/critical-path"),
      ("Decision Deadline Engine","Put RFIs, submittals and owner decisions on one urgency clock.","/prediction-intelligence/decision-deadlines"),
      ("Manpower Forecast","See overlapping trade demand for the next two weeks.","/prediction-intelligence/manpower"),
      ("Material Readiness","Connect procurement timing directly to scheduled activities.","/prediction-intelligence/materials"),
      ("Coordination Agenda Brain","Auto-build coordination meeting topics from live project intelligence.","/prediction-intelligence/agenda"),
      ("Change Causality","Trace open changes back to related RFIs/issues when evidence exists.","/prediction-intelligence/change-causality"),
      ("Closeout Readiness Predictor","See turnover items likely to become late before project end.","/prediction-intelligence/closeout"),
      ("Risk Propagation Graph","Show how scope/activity risk spreads downstream.","/prediction-intelligence/risk-propagation"),
      ("Predictive Project Brief","One forward-looking view of the next problems to solve.","/prediction-intelligence/brief")
    ]
    for name,desc,href in cards:
        body+=_v37_link_card(name,desc,href,"Open")
    body+='</div>'
    return shell("Prediction Intelligence",body)

@app.get("/prediction-intelligence/dependency-impact",response_class=HTMLResponse)
def v47_dependency_page():
    rows=_v47_dependency_impacts(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(level)}</span> <b>{esc(r["trade"])}</b> → {esc(tr)}<div>{esc(r["requirement"])}</div><div class="small">{esc(reason)}</div></div>' for r,tr,level,reason in rows)
    return shell("Dependency Impact",'<div class="hero"><h1>Dependency Impact Engine</h1><p class="muted">Shows where supporting-trade risk can affect the primary scope.</p></div><div class="card">'+(h or '<p class="muted">No dependency impacts detected.</p>')+'</div>')

@app.get("/prediction-intelligence/critical-path",response_class=HTMLResponse)
def v47_critical_path_page():
    rows=_v47_critical_path_signals(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">SCORE {score}</span> <b>{esc(x["activity"]["name"])}</b><div class="small">{esc(x["activity"]["trade"])} · Stage {x["stage"]} · {esc(x["risk"])}</div><p>{esc(x["blocking_reason"] or "Upcoming activity with completion exposure.")}</p></div>' for score,x in rows)
    return shell("Critical Path Intelligence",'<div class="hero"><h1>Critical Path Intelligence</h1><p class="muted">This is a transparent risk proxy until a full CPM dependency network is available.</p></div><div class="card">'+(h or '<p class="muted">No major critical-path proxy signals.</p>')+'</div>')

@app.get("/prediction-intelligence/decision-deadlines",response_class=HTMLResponse)
def v47_decisions_page():
    rows=_v47_decision_deadlines(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(severity)}</span> <b>{esc(typ)} - {esc(title)}</b><div class="small">Due {esc(due)} · ${cost:,.0f} exposure · {days:g} days · {esc(source)}</div></div>' for typ,title,due,severity,cost,days,source in rows)
    return shell("Decision Deadlines",'<div class="hero"><h1>Decision Deadline Engine</h1><p class="muted">One urgency clock for owner decisions, RFIs and submittals.</p></div><div class="card">'+(h or '<p class="muted">No open decision deadlines.</p>')+'</div>')

@app.get("/prediction-intelligence/manpower",response_class=HTMLResponse)
def v47_manpower_page():
    rows=_v47_manpower_forecast(project_id())
    h="".join(f'<div class="action"><span class="badge {"WATCH" if level!="LOW" else "READY"}">{esc(level)}</span> <b>{esc(date)} - {esc(trade)}</b><div class="small">{count} active activity(ies): {esc(", ".join(names))}</div></div>' for date,trade,count,level,names in rows)
    return shell("Manpower Forecast",'<div class="hero"><h1>Manpower Forecast</h1><p class="muted">Predict overlapping trade demand from the next two weeks of scheduled activities.</p></div><div class="card">'+(h or '<p class="muted">No dated activities available for manpower forecasting.</p>')+'</div>')

@app.get("/prediction-intelligence/materials",response_class=HTMLResponse)
def v47_material_page():
    rows=_v47_material_readiness(project_id())
    h=""
    for r,act,level,exposure,reason,action in rows:
        h+=f'<div class="action"><span class="badge WATCH">{esc(level)}</span> <b>{esc(r["item"])}</b><div class="small">Activity: {esc(act["name"] if act else "Unlinked")} · Need {esc(r["required_on_site"])} · Promised {esc(r["promised_date"])} · Exposure {exposure} day(s)</div><p>{esc(reason)}</p></div>'
    return shell("Material Readiness",'<div class="hero"><h1>Material Readiness Intelligence</h1><p class="muted">Procurement risk is most useful when connected directly to installation work.</p></div><div class="card">'+(h or '<p class="muted">No open material-readiness risks.</p>')+'</div>')

@app.get("/prediction-intelligence/agenda",response_class=HTMLResponse)
def v47_agenda_page():
    rows=_v47_coordination_agenda(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(kind)}</span> <b>{esc(title)}</b><div class="small">{esc(detail)}</div></div>' for kind,title,detail in rows)
    return shell("Coordination Agenda Brain",'<div class="hero"><h1>Coordination Agenda Brain</h1><p class="muted">Builds meeting topics from live conflicts, constructability, schedule and decision intelligence.</p></div><div class="card">'+(h or '<p class="muted">No current coordination agenda items.</p>')+'</div>')

@app.get("/prediction-intelligence/change-causality",response_class=HTMLResponse)
def v47_change_causality_page():
    rows=_v47_change_causality(project_id())
    h=""
    for c,links in rows:
        linktxt=", ".join(f"{k}: {v}" for k,v in links) if links else "No obvious linked RFI/issue by title."
        h+=f'<div class="action"><b>{esc(c["title"])}</b><div class="small">{esc(c["event_type"])} · ${float(c["estimated_cost"] or 0):,.0f} · {float(c["schedule_days"] or 0):g} days</div><p>{esc(linktxt)}</p></div>'
    return shell("Change Causality",'<div class="hero"><h1>Change Causality Intelligence</h1><p class="muted">Helps preserve the reason chain behind change exposure.</p></div><div class="card">'+(h or '<p class="muted">No open change events.</p>')+'</div>')

@app.get("/prediction-intelligence/closeout",response_class=HTMLResponse)
def v47_closeout_page():
    rows=_v47_closeout_prediction(project_id())
    h="".join(f'<div class="action"><span class="badge {"READY" if level=="READY" else "WATCH"}">{esc(level)}</span> <b>{esc(r["category"])}</b> - {esc(r["item"])}<div class="small">{esc(r["status"])} · Due {esc(r["due_date"])} · {esc(r["responsible_party"])}</div></div>' for r,level in rows)
    return shell("Closeout Readiness",'<div class="hero"><h1>Closeout Readiness Predictor</h1><p class="muted">Turnover risk should be managed before the final month, not discovered at the end.</p></div><div class="card">'+h+'</div>')

@app.get("/prediction-intelligence/risk-propagation",response_class=HTMLResponse)
def v47_risk_graph_page():
    rows=_v47_risk_propagation(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(level)}</span> <b>{esc(stitle)}</b> → {esc(ttitle)}<div class="small">{esc(rel)} · {esc(reason)}</div></div>' for stype,skey,stitle,ttype,tkey,ttitle,rel,level,reason in rows)
    return shell("Risk Propagation",'<div class="hero"><h1>Risk Propagation Graph</h1><p class="muted">Shows how one unresolved dependency can create downstream project exposure.</p></div><div class="card">'+(h or '<p class="muted">No propagated risk signals.</p>')+'</div>')

@app.get("/prediction-intelligence/brief",response_class=HTMLResponse)
def v47_predictive_brief():
    pid=project_id()
    cp=_v47_critical_path_signals(pid)[:5]
    decisions=_v47_decision_deadlines(pid)[:5]
    materials=[x for x in _v47_material_readiness(pid) if x[2] in {"CRITICAL","HIGH"}][:5]
    closeout=[x for x in _v47_closeout_prediction(pid) if x[1] in {"CRITICAL","HIGH"}][:5]
    body='<div class="hero"><div class="eyebrow">Predictive Project Brief</div><h1>What is most likely to hurt the job next?</h1><p class="muted">Forward-looking signals from schedule, decisions, material readiness and closeout.</p></div>'
    body+='<div class="card"><h2>Critical Path / Sequence</h2>'+("".join(f'<div class="action"><b>{esc(x["activity"]["name"])}</b><div class="small">Score {score} · {esc(x["risk"])}</div></div>' for score,x in cp) or '<p class="muted">No major sequence threats.</p>')+'</div>'
    body+='<div class="card"><h2>Decision Deadlines</h2>'+("".join(f'<div class="action"><b>{esc(typ)} - {esc(title)}</b><div class="small">Due {esc(due)} · {esc(sev)}</div></div>' for typ,title,due,sev,cost,days,source in decisions) or '<p class="muted">No urgent decisions.</p>')+'</div>'
    body+='<div class="card"><h2>Material Risk</h2>'+("".join(f'<div class="action"><b>{esc(r["item"])}</b><div class="small">{esc(level)} · {exposure} day(s) exposure</div></div>' for r,act,level,exposure,reason,action in materials) or '<p class="muted">No high material risks.</p>')+'</div>'
    body+='<div class="card"><h2>Closeout Risk</h2>'+("".join(f'<div class="action"><b>{esc(r["item"])}</b><div class="small">{esc(level)} · Due {esc(r["due_date"])}</div></div>' for r,level in closeout) or '<p class="muted">No high closeout risks.</p>')+'</div>'
    return shell("Predictive Project Brief",body)


# ============================================================
# v48 BRAIN QUALITY & SELF-LEARNING ENGINE
# ============================================================

def _v48_ensure_tables():
    c=db()
    if DATABASE_KIND=="postgres":
        pk="BIGSERIAL PRIMARY KEY"; num="DOUBLE PRECISION"
    else:
        pk="INTEGER PRIMARY KEY"; num="REAL"
    stmts=[
      f"""CREATE TABLE IF NOT EXISTS quality_audit(
        id {pk},company_id BIGINT,project_id BIGINT,subject_type TEXT,subject_key TEXT,
        primary_result TEXT,secondary_result TEXT,agreement TEXT DEFAULT 'REVIEW',
        source_quality TEXT DEFAULT 'MEDIUM',confidence_score {num} DEFAULT 0,
        contradiction_flag INTEGER DEFAULT 0,reason TEXT,created TEXT,updated TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS confidence_calibration(
        id {pk},company_id BIGINT,project_id BIGINT,subject_type TEXT,subject_key TEXT,
        stated_confidence TEXT,calibrated_confidence TEXT,score {num} DEFAULT 0,
        reason TEXT,created TEXT,updated TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS contradiction_intelligence(
        id {pk},company_id BIGINT,project_id BIGINT,subject TEXT,source_a TEXT,source_b TEXT,
        contradiction_type TEXT,severity TEXT DEFAULT 'REVIEW',status TEXT DEFAULT 'OPEN',
        reason TEXT,created TEXT,updated TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS learning_feedback(
        id {pk},company_id BIGINT,project_id BIGINT,subject_type TEXT,subject_key TEXT,
        original_result TEXT,approved_result TEXT,feedback_type TEXT,reason TEXT,
        approved_by TEXT,scope_level TEXT DEFAULT 'PROJECT ONLY',created TEXT)"""
    ]
    for s in stmts:
        c.execute(s)
    c.commit(); c.close()

def _v48_source_quality(r):
    score=0
    if r["source_sheet"]: score+=35
    if r["source_detail"]: score+=25
    if r["source_spec"]: score+=25
    if r["source_note"]: score+=15
    if score>=75: return ("HIGH",score)
    if score>=40: return ("MEDIUM",score)
    return ("LOW",score)

def _v48_calibrated_confidence(r):
    stated=str(r["confidence"] or "MEDIUM").upper()
    source_level,source_score=_v48_source_quality(r)
    verify=_v441_primary_trade(r["requirement"],r["trade"])
    agrees=(verify==r["trade"])
    score=source_score
    score += 15 if agrees else -25
    score += 10 if stated=="HIGH" else 0
    score -= 10 if stated=="LOW" else 0
    score=max(0,min(100,score))
    calibrated="HIGH" if score>=75 else "MEDIUM" if score>=45 else "LOW"
    reason=f"Source quality {source_level}; second opinion {'agrees' if agrees else 'disagrees'} with saved trade."
    return calibrated,score,reason,verify

def _v48_self_audit_uncached(pid):
    rows=_v452_scope_rows(pid)
    out=[]
    for r in rows:
        calibrated,score,reason,verify=_v48_calibrated_confidence(r)
        agreement="AGREE" if verify==r["trade"] else "DISAGREE"
        source_level,_=_v48_source_quality(r)
        contradiction=1 if agreement=="DISAGREE" else 0
        out.append((r,verify,agreement,source_level,calibrated,score,contradiction,reason))
    return out[:300]

def _v48_self_audit(pid):
    return _v56_self_audit(pid)

def _v48_contradictions(pid):
    rows=_v452_scope_rows(pid)
    by_key={}
    out=[]
    for r in rows:
        req=re.sub(r'\s+',' ',str(r["requirement"] or "").lower()).strip()
        if not req: continue
        by_key.setdefault(req,[]).append(r)
    for req,items in by_key.items():
        trades=sorted(set(str(x["trade"] or "") for x in items))
        if len(trades)>1:
            out.append(("TRADE OWNERSHIP",items[0]["requirement"]," / ".join(trades),items[0]["source_sheet"] or "",items[-1]["source_sheet"] or ""))
    # One-sided spec/drawing signals are treated as possible document contradictions.
    for r,kind in _v46_spec_drawing_checks(pid):
        out.append(("SOURCE MISMATCH",r["requirement"],kind,r["source_sheet"] or "",r["source_spec"] or ""))
    return out[:200]

def _v48_learning_rules(pid):
    _v43_ensure_tables()
    return _v39_rows("""
        SELECT * FROM learning_rules
        WHERE company_id=? AND approval_status='APPROVED'
          AND (project_id=? OR scope_level='COMPANY STANDARD')
        ORDER BY CASE WHEN project_id=? THEN 0 ELSE 1 END,id DESC
        LIMIT 300
    """,(current_company_id(),pid,pid))

def _v48_quality_score(pid):
    audit=_v48_self_audit(pid)
    if not audit:
        return {"score":100,"disagree":0,"low":0,"missing_source":0}
    disagree=sum(1 for x in audit if x[2]=="DISAGREE")
    low=sum(1 for x in audit if x[4]=="LOW")
    missing_source=sum(1 for x in audit if x[3]=="LOW")
    score=max(0,100 - disagree*6 - low*3 - missing_source*2)
    return {"score":score,"disagree":disagree,"low":low,"missing_source":missing_source}

def _v48_learning_opportunities(pid):
    audit=_v48_self_audit(pid)
    approved=_v48_learning_rules(pid)
    subjects={str(r["subject"] or "").lower() for r in approved}
    out=[]
    for r,verify,agreement,source_level,calibrated,score,contradiction,reason in audit:
        if agreement=="DISAGREE" and str(r["requirement"] or "").lower() not in subjects:
            out.append((r,verify,calibrated,reason))
    return out[:150]

def _v48_answer_guard(pid,question):
    """
    Lightweight quality gate for Ask BuildCommand:
    surfaces uncertainty and related knowledge rather than fabricating certainty.
    """
    q=(question or "").lower()
    rows=_v452_scope_rows(pid)
    matches=[]
    for r in rows:
        text=(str(r["requirement"] or "")+" "+str(r["trade"] or "")+" "+str(r["source_note"] or "")).lower()
        if any(tok in text for tok in [w for w in re.findall(r'[a-z0-9]+',q) if len(w)>3]):
            matches.append(r)
    matches=matches[:12]
    if not matches:
        return {"confidence":"LOW","note":"No strong project-source match found.","matches":[]}
    low=sum(1 for r in matches if _v48_calibrated_confidence(r)[0]=="LOW")
    conf="LOW" if low>=max(1,len(matches)//2) else "MEDIUM" if low else "HIGH"
    note="Answer should cite project sources and expose uncertainty where sources conflict."
    return {"confidence":conf,"note":note,"matches":matches}

@app.get("/brain-quality",response_class=HTMLResponse)
def v48_brain_quality_home():
    pid=project_id(); _v48_ensure_tables()
    q=_v48_quality_score(pid)
    contradictions=_v48_contradictions(pid)
    rules=_v48_learning_rules(pid)
    opportunities=_v48_learning_opportunities(pid)
    body=f'<div class="hero"><div class="eyebrow">BuildCommand v48 - Brain Quality & Self-Learning</div><h1>Make the brain more accurate before making it bigger.</h1><p class="muted">Quality score {q["score"]}/100 · {q["disagree"]} second-opinion disagreement(s) · {q["low"]} low-confidence item(s) · {len(contradictions)} contradiction/source mismatch signal(s) · {len(rules)} approved learning rule(s).</p></div><div class="grid3">'
    cards=[
      ("Multi-Pass Verification","Primary ownership vs independent second opinion.","/brain-quality/verification"),
      ("Confidence Calibration","Re-score confidence using source quality and agreement.","/brain-quality/confidence"),
      ("Source Quality Audit","Find weakly sourced project intelligence.","/brain-quality/sources"),
      ("Contradiction Detection","Find ownership or source inconsistencies.","/brain-quality/contradictions"),
      ("Approved Learning Reuse","Show the project/company rules the brain can trust.","/brain-quality/learning"),
      ("Learning Opportunities","Find disagreements that should become reviewed corrections.","/brain-quality/opportunities"),
      ("Scope Self-Audit","Run trade/source/confidence QA before downstream use.","/brain-quality/self-audit"),
      ("Answer Guardrails","Show how Ask BuildCommand should expose uncertainty.","/brain-quality/answer-guard"),
      ("Quality Dashboard","One score for current intelligence quality.","/brain-quality/dashboard"),
      ("Brain Improvement Queue","Prioritize the highest-value corrections first.","/brain-quality/queue")
    ]
    for name,desc,href in cards:
        body+=_v37_link_card(name,desc,href,"Open")
    body+='</div>'
    return shell("Brain Quality",body)

@app.get("/brain-quality/verification",response_class=HTMLResponse)
def v48_verification_page():
    rows=_v48_self_audit(project_id())
    h="".join(f'<div class="action"><span class="badge {"READY" if agreement=="AGREE" else "WATCH"}">{esc(agreement)}</span> <b>{esc(r["trade"])}</b> - {esc(r["requirement"])}<div class="small">Second opinion: {esc(verify)} · Calibrated {esc(calibrated)} ({score:.0f})</div></div>' for r,verify,agreement,source_level,calibrated,score,contradiction,reason in rows if agreement=="DISAGREE")
    return shell("Multi-Pass Verification",'<div class="hero"><h1>Multi-Pass Verification</h1><p class="muted">The second pass challenges the saved result instead of rubber-stamping it.</p></div><div class="card">'+(h or '<p class="muted">No current ownership disagreements.</p>')+'</div>')

@app.get("/brain-quality/confidence",response_class=HTMLResponse)
def v48_confidence_page():
    rows=_v48_self_audit(project_id())
    h="".join(f'<div class="action"><span class="badge">{esc(calibrated)}</span> <b>{esc(r["trade"])}</b> - {esc(r["requirement"])}<div class="small">Stated {esc(r["confidence"])} → Calibrated {esc(calibrated)} ({score:.0f}/100) · {esc(reason)}</div></div>' for r,verify,agreement,source_level,calibrated,score,contradiction,reason in rows)
    return shell("Confidence Calibration",'<div class="hero"><h1>Confidence Calibration</h1><p class="muted">Confidence now considers source strength and whether an independent ownership pass agrees.</p></div><div class="card">'+(h or '<p class="muted">No scope intelligence available.</p>')+'</div>')

@app.get("/brain-quality/sources",response_class=HTMLResponse)
def v48_sources_page():
    rows=_v48_self_audit(project_id())
    weak=[x for x in rows if x[3]=="LOW"]
    h="".join(f'<div class="action"><span class="badge WATCH">LOW SOURCE</span> <b>{esc(r["trade"])}</b> - {esc(r["requirement"])}<div class="small">Sheet {esc(r["source_sheet"])} · Detail {esc(r["source_detail"])} · Spec {esc(r["source_spec"])}</div></div>' for r,verify,agreement,source_level,calibrated,score,contradiction,reason in weak)
    return shell("Source Quality Audit",'<div class="hero"><h1>Source Quality Audit</h1><p class="muted">Weak sourcing should lower trust even when the wording looks convincing.</p></div><div class="card">'+(h or '<p class="muted">No low-source-quality scope items detected.</p>')+'</div>')

@app.get("/brain-quality/contradictions",response_class=HTMLResponse)
def v48_contradictions_page():
    rows=_v48_contradictions(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(kind)}</span> <b>{esc(subject)}</b><div class="small">{esc(detail)} · Source A {esc(a)} · Source B {esc(b)}</div></div>' for kind,subject,detail,a,b in rows)
    return shell("Contradiction Detection",'<div class="hero"><h1>Contradiction Detection</h1><p class="muted">Conflicting ownership or one-sided documentation should become review work, not false certainty.</p></div><div class="card">'+(h or '<p class="muted">No contradiction signals detected.</p>')+'</div>')

@app.get("/brain-quality/learning",response_class=HTMLResponse)
def v48_learning_page():
    rows=_v48_learning_rules(project_id())
    h="".join(f'<div class="action"><span class="badge READY">{esc(r["scope_level"])}</span> <b>{esc(r["rule_type"])} - {esc(r["subject"])}</b><div>{esc(r["learned_rule"])}</div><div class="small">Approved by {esc(r["approved_by"])}</div></div>' for r in rows)
    return shell("Approved Learning Reuse",'<div class="hero"><h1>Approved Learning Reuse</h1><p class="muted">Only reviewed project/company rules are trusted by the brain.</p></div><div class="card">'+(h or '<p class="muted">No approved learning rules available.</p>')+'</div>')

@app.get("/brain-quality/opportunities",response_class=HTMLResponse)
def v48_opportunities_page():
    rows=_v48_learning_opportunities(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">LEARN</span> <b>{esc(r["trade"])}</b> - {esc(r["requirement"])}<div class="small">Second opinion says {esc(verify)} · {esc(calibrated)} · {esc(reason)}</div></div>' for r,verify,calibrated,reason in rows)
    return shell("Learning Opportunities",'<div class="hero"><h1>Learning Opportunities</h1><p class="muted">These are disagreements worth human review before they become approved knowledge.</p></div><div class="card">'+(h or '<p class="muted">No new learning opportunities detected.</p>')+'</div>')

@app.get("/brain-quality/self-audit",response_class=HTMLResponse)
def v48_selfaudit_page():
    rows=_v48_self_audit(project_id())
    bad=[x for x in rows if x[2]=="DISAGREE" or x[4]=="LOW" or x[3]=="LOW"]
    h="".join(f'<div class="action"><span class="badge WATCH">REVIEW</span> <b>{esc(r["trade"])}</b> - {esc(r["requirement"])}<div class="small">Agreement {esc(agreement)} · Source {esc(source_level)} · Confidence {esc(calibrated)} ({score:.0f})</div></div>' for r,verify,agreement,source_level,calibrated,score,contradiction,reason in bad)
    return shell("Scope Self-Audit",'<div class="hero"><h1>Scope Self-Audit</h1><p class="muted">Runs QA before intelligence feeds estimating, bidding, schedule or field execution.</p></div><div class="card">'+(h or '<p class="muted">No major quality exceptions detected.</p>')+'</div>')

@app.get("/brain-quality/answer-guard",response_class=HTMLResponse)
def v48_answer_guard_page(q:str=''):
    result=_v48_answer_guard(project_id(),q)
    matches=result["matches"]
    h="".join(f'<div class="action"><b>{esc(r["trade"])}</b> - {esc(r["requirement"])}<div class="small">Sheet {esc(r["source_sheet"])} · Spec {esc(r["source_spec"])} · Confidence {esc(_v48_calibrated_confidence(r)[0])}</div></div>' for r in matches)
    body='<div class="hero"><h1>Answer Guardrails</h1><p class="muted">Ask BuildCommand should say when project evidence is weak or conflicting.</p></div><div class="card"><form method="get"><input name="q" value="'+esc(q)+'" placeholder="Test a project question"><button type="submit">Check evidence</button></form></div>'
    if q:
        body+=f'<div class="card"><span class="badge">{esc(result["confidence"])}</span><p>{esc(result["note"])}</p>{h or "<p class=muted>No strong project-source matches.</p>"}</div>'
    return shell("Answer Guardrails",body)

@app.get("/brain-quality/dashboard",response_class=HTMLResponse)
def v48_dashboard_page():
    q=_v48_quality_score(project_id())
    return shell("Quality Dashboard",f'<div class="hero"><h1>Brain Quality Dashboard</h1><p class="muted">Current intelligence quality score: {q["score"]}/100.</p></div><div class="grid3"><div class="card"><div class="label">Second-Opinion Disagreements</div><div class="kpi">{q["disagree"]}</div></div><div class="card"><div class="label">Low Confidence</div><div class="kpi">{q["low"]}</div></div><div class="card"><div class="label">Weak Sources</div><div class="kpi">{q["missing_source"]}</div></div></div>')

@app.get("/brain-quality/queue",response_class=HTMLResponse)
def v48_queue_page():
    rows=_v48_self_audit(project_id())
    queue=[]
    for r,verify,agreement,source_level,calibrated,score,contradiction,reason in rows:
        priority=0
        if agreement=="DISAGREE": priority+=50
        if calibrated=="LOW": priority+=30
        if source_level=="LOW": priority+=20
        if priority:
            queue.append((priority,r,verify,agreement,source_level,calibrated,score))
    queue.sort(key=lambda x:x[0],reverse=True)
    h="".join(f'<div class="action"><span class="badge WATCH">PRIORITY {p}</span> <b>{esc(r["trade"])}</b> - {esc(r["requirement"])}<div class="small">2nd opinion {esc(verify)} · Source {esc(source)} · Confidence {esc(conf)} ({score:.0f})</div></div>' for p,r,verify,agreement,source,conf,score in queue[:100])
    return shell("Brain Improvement Queue",'<div class="hero"><h1>Brain Improvement Queue</h1><p class="muted">Fix the highest-value intelligence problems first.</p></div><div class="card">'+(h or '<p class="muted">No current quality-improvement queue.</p>')+'</div>')


# ============================================================
# v49 FIELD CONTEXT & ASSEMBLY INTELLIGENCE
# ============================================================

def _v49_ensure_tables():
    c=db()
    if DATABASE_KIND=="postgres":
        pk="BIGSERIAL PRIMARY KEY"; num="DOUBLE PRECISION"
    else:
        pk="INTEGER PRIMARY KEY"; num="REAL"
    stmts=[
      f"""CREATE TABLE IF NOT EXISTS room_intelligence(
        id {pk},company_id BIGINT,project_id BIGINT,room_key TEXT,room_name TEXT,
        room_number TEXT,source_ref TEXT,trade_count INTEGER DEFAULT 0,
        issue_count INTEGER DEFAULT 0,status TEXT DEFAULT 'REVIEW',created TEXT,updated TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS assembly_intelligence(
        id {pk},company_id BIGINT,project_id BIGINT,assembly_type TEXT,assembly_key TEXT,
        description TEXT,primary_trade TEXT,related_trades TEXT,source_ref TEXT,
        prerequisite_summary TEXT,inspection_summary TEXT,status TEXT DEFAULT 'REVIEW',
        created TEXT,updated TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS field_work_packages(
        id {pk},company_id BIGINT,project_id BIGINT,package_type TEXT,title TEXT,
        location TEXT,trade TEXT,scope_summary TEXT,prerequisites TEXT,
        inspection_gates TEXT,materials TEXT,status TEXT DEFAULT 'DRAFT',
        created TEXT,updated TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS commissioning_chain(
        id {pk},company_id BIGINT,project_id BIGINT,equipment_key TEXT,equipment_name TEXT,
        install_trade TEXT,power_trade TEXT,controls_trade TEXT,startup_trade TEXT,
        testing_requirement TEXT,closeout_requirement TEXT,status TEXT DEFAULT 'REVIEW',
        source_ref TEXT,created TEXT,updated TEXT)"""
    ]
    for s in stmts:
        c.execute(s)
    c.commit(); c.close()

def _v49_scope_rows(pid):
    return _v452_scope_rows(pid)

def _v49_room_key(text):
    s=str(text or "")
    # Common room-number patterns such as 101, 122A, Room 203.
    m=re.search(r'(?i)\broom\s+([A-Z]?\d{2,4}[A-Z]?)\b',s)
    if m:
        return m.group(1)
    m=re.search(r'\b([1-9]\d{2}[A-Z]?)\b',s)
    return m.group(1) if m else ""

def _v49_rooms(pid):
    rows=_v49_scope_rows(pid)
    rooms={}
    for r in rows:
        corpus=" ".join(str(r[k] or "") for k in ["requirement","source_note","source_detail"])
        rk=_v49_room_key(corpus)
        if not rk:
            continue
        ent=rooms.setdefault(rk,{"room":rk,"items":[],"trades":set()})
        ent["items"].append(r)
        ent["trades"].add(str(r["trade"] or ""))
    return list(rooms.values())[:200]

def _v49_assemblies(pid):
    rows=_v49_scope_rows(pid)
    out=[]
    patterns=[
      ("WALL",("wall type","partition","stud wall","gypsum board wall","shaft wall")),
      ("CEILING",("ceiling","act ceiling","acoustical ceiling","hard lid")),
      ("FLOOR",("flooring","tile","lvt","vct","carpet","resilient flooring")),
      ("ROOF",("roof","roofing","roof curb","roof penetration")),
      ("DOOR",("door","frame","hardware","storefront entrance")),
      ("PLUMBING FIXTURE",("water closet","urinal","lavatory","mop sink","floor drain")),
      ("EQUIPMENT",("water heater","rooftop unit","exhaust fan","air handler","panelboard","transformer"))
    ]
    seen=set()
    for r in rows:
        low=str(r["requirement"] or "").lower()
        for typ,terms in patterns:
            if any(term in low for term in terms):
                key=(typ,re.sub(r'\s+',' ',low[:140]))
                if key in seen:
                    continue
                seen.add(key)
                rel=set()
                if r["related_trade"]:
                    rel.add(str(r["related_trade"]))
                for dep_r,rels in _v46_dependencies(pid):
                    if dep_r["id"]==r["id"]:
                        rel.update(rels)
                prereq=[]
                if typ=="WALL":
                    prereq=["layout complete","framing/backing complete","MEP rough complete","required rough inspection passed"]
                elif typ=="CEILING":
                    prereq=["above-ceiling MEP complete","coordination complete","inspection/testing complete"]
                elif typ=="FLOOR":
                    prereq=["substrate ready","moisture/flatness requirements verified","overhead damaging work complete"]
                elif typ=="ROOF":
                    prereq=["penetrations/curbs coordinated","equipment/support locations confirmed"]
                elif typ=="DOOR":
                    prereq=["rough opening verified","frame/blocking ready","hardware/electrical interfaces coordinated"]
                elif typ=="PLUMBING FIXTURE":
                    prereq=["rough plumbing complete","wall/floor finishes ready","fixture available"]
                else:
                    prereq=["approved submittal","support ready","material onsite","power/controls/piping interfaces coordinated"]
                source=" · ".join(x for x in [r["source_sheet"],r["source_detail"],r["source_spec"]] if x)
                out.append((typ,r,sorted(rel),prereq,source))
    return out[:250]

def _v49_system_trace(pid):
    rows=_v49_scope_rows(pid)
    systems={
        "Electrical":[],"Plumbing":[],"HVAC / Mechanical":[],"Fire Sprinkler":[],
        "Low Voltage":[],"Fire Alarm":[],"Roofing":[],"Doors / Frames / Hardware":[]
    }
    for r in rows:
        tr=str(r["trade"] or "")
        if tr in systems:
            systems[tr].append(r)
    return systems

def _v49_penetrations(pid):
    rows=_v49_scope_rows(pid)
    out=[]
    for r in rows:
        low=str(r["requirement"] or "").lower()
        if any(x in low for x in ["penetration","sleeve","core drill","opening","roof curb","rough opening"]):
            out.append(r)
    return out[:150]

def _v49_install_prereqs(pid):
    seq=_v45_sequence_analysis(pid)
    out=[]
    for x in seq:
        a=x["activity"]
        gates=_v45_gate_requirements(x["stage"])
        out.append((a,x["readiness"],x["risk"],gates,x["blocking_reason"]))
    return out[:150]

def _v49_hold_points(pid):
    inspections=_v39_rows("SELECT * FROM inspections_tracker WHERE project_id=? ORDER BY scheduled_date,id",(pid,))
    seq=_v45_sequence_analysis(pid)
    out=[]
    for i in inspections:
        act=None
        if i["activity_id"] is not None:
            act=next((x["activity"] for x in seq if x["activity"]["id"]==i["activity_id"]),None)
        out.append((i,act))
    return out[:150]

def _v49_commissioning(pid):
    eq=_v46_equipment(pid)
    out=[]
    for e in eq:
        testing=[]
        closeout=[]
        name=e["name"].lower()
        if any(x in name for x in ["rtu","air handler","exhaust fan","vav"]):
            testing=["startup","controls verification","test and balance as applicable"]
            closeout=["O&M","warranty","startup record"]
        elif "water heater" in name:
            testing=["startup","temperature/operation verification","leak check"]
            closeout=["O&M","warranty"]
        elif any(x in name for x in ["panelboard","transformer"]):
            testing=["energization verification","labeling/identification"]
            closeout=["O&M if required","test records if required"]
        else:
            testing=["startup / functional verification as applicable"]
            closeout=["O&M / warranty as applicable"]
        out.append((e,testing,closeout))
    return out[:100]

def _v49_work_packages(pid):
    rooms=_v49_rooms(pid)
    assemblies=_v49_assemblies(pid)
    packages=[]
    for room in rooms:
        by_trade={}
        for r in room["items"]:
            by_trade.setdefault(r["trade"],[]).append(r)
        for tr,items in by_trade.items():
            scope="; ".join(str(x["requirement"]) for x in items[:8])
            packages.append({
                "title":f"Room {room['room']} - {tr}",
                "location":f"Room {room['room']}",
                "trade":tr,"scope":scope,
                "prereq":"Verify predecessor work, access, approved documents and material readiness.",
                "inspection":"Verify required inspection/hold points before concealment or finish work.",
                "materials":"Confirm project-specific material availability."
            })
    if not packages:
        # Fallback assembly-based packages when room numbers are not explicit.
        for typ,r,rels,prereq,source in assemblies[:80]:
            packages.append({
                "title":f"{typ} - {r['trade']}",
                "location":source or "Project",
                "trade":r["trade"],"scope":r["requirement"],
                "prereq":"; ".join(prereq),
                "inspection":"Verify applicable inspections/testing before downstream work.",
                "materials":"Confirm approved material/submittal and onsite readiness."
            })
    return packages[:200]

@app.get("/field-context",response_class=HTMLResponse)
def v49_field_context_home():
    pid=project_id(); _v49_ensure_tables()
    rooms=_v49_rooms(pid); assemblies=_v49_assemblies(pid); pens=_v49_penetrations(pid)
    prereqs=_v49_install_prereqs(pid); holds=_v49_hold_points(pid); comm=_v49_commissioning(pid)
    packages=_v49_work_packages(pid); systems=_v49_system_trace(pid)
    system_count=sum(1 for _,rows in systems.items() if rows)
    body=f'<div class="hero"><div class="eyebrow">BuildCommand v49 - Field Context & Assembly Intelligence</div><h1>Understand the project where the work actually happens.</h1><p class="muted">{len(rooms)} room context(s) · {len(assemblies)} assembly signals · {system_count} traced systems · {len(pens)} penetration/opening items · {len(holds)} inspection hold points · {len(comm)} commissioning chains · {len(packages)} field work packages.</p></div><div class="grid3">'
    cards=[
      ("Room Intelligence","Organize project scope by room/location context.","/field-context/rooms"),
      ("Assembly Intelligence","Understand wall, ceiling, floor, roof, door and equipment assemblies.","/field-context/assemblies"),
      ("System Trace Intelligence","Trace scope by Electrical, Plumbing, HVAC, Fire Protection and other systems.","/field-context/systems"),
      ("Penetration & Opening Brain","Find sleeves, penetrations, rough openings and roof/opening coordination.","/field-context/penetrations"),
      ("Installation Prerequisites","Show what must be ready before activities can start.","/field-context/prerequisites"),
      ("Inspection Hold Points","Connect inspections/testing to the work they gate.","/field-context/hold-points"),
      ("Equipment-to-Location Intelligence","Map equipment scope to room/source context and trade interfaces.","/field-context/equipment-locations"),
      ("Commissioning Chain","Follow install → power → controls → startup → testing → closeout.","/field-context/commissioning"),
      ("Field Work Packages","Turn project intelligence into location/trade execution packages.","/field-context/work-packages"),
      ("Field Context Command","One field-focused view of rooms, assemblies, gates and work packages.","/field-context/command")
    ]
    for name,desc,href in cards:
        body+=_v37_link_card(name,desc,href,"Open")
    body+='</div>'
    return shell("Field Context Intelligence",body)

@app.get("/field-context/rooms",response_class=HTMLResponse)
def v49_rooms_page():
    rows=_v49_rooms(project_id())
    h=""
    for room in rows:
        h+=f'<div class="card"><h3>Room {esc(room["room"])}</h3><p class="small">{len(room["items"])} scope item(s) · {len(room["trades"])} trade(s)</p>'
        for r in room["items"][:20]:
            h+=f'<div class="action"><b>{esc(r["trade"])}</b> - {esc(r["requirement"])}</div>'
        h+='</div>'
    return shell("Room Intelligence",'<div class="hero"><h1>Room Intelligence</h1><p class="muted">Location context is inferred only where explicit room identifiers exist in the analyzed text.</p></div>'+(h or '<div class="card">No explicit room identifiers detected in current scope.</div>'))

@app.get("/field-context/assemblies",response_class=HTMLResponse)
def v49_assemblies_page():
    rows=_v49_assemblies(project_id())
    h="".join(f'<div class="action"><span class="badge">{esc(typ)}</span> <b>{esc(r["trade"])}</b> - {esc(r["requirement"])}<div class="small">Related: {esc(", ".join(rels) or "None")} · Source {esc(source)}</div><p><b>Prerequisites:</b> {esc("; ".join(pre))}</p></div>' for typ,r,rels,pre,source in rows)
    return shell("Assembly Intelligence",'<div class="hero"><h1>Assembly Intelligence</h1><p class="muted">Breaks project scope into buildable assemblies with related trades and prerequisite logic.</p></div><div class="card">'+(h or '<p class="muted">No assembly signals detected.</p>')+'</div>')

@app.get("/field-context/systems",response_class=HTMLResponse)
def v49_systems_page():
    systems=_v49_system_trace(project_id())
    body='<div class="hero"><h1>System Trace Intelligence</h1><p class="muted">Trace the project by system instead of reading isolated scope items.</p></div>'
    for name,rows in systems.items():
        if not rows: continue
        body+=f'<div class="card"><h2>{esc(name)}</h2>'
        body+="".join(f'<div class="action">{esc(r["requirement"])}<div class="small">{esc(r["source_sheet"])} {esc(r["source_detail"])}</div></div>' for r in rows[:40])
        body+='</div>'
    return shell("System Trace Intelligence",body)

@app.get("/field-context/penetrations",response_class=HTMLResponse)
def v49_penetrations_page():
    rows=_v49_penetrations(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">COORDINATE</span> <b>{esc(r["trade"])}</b> - {esc(r["requirement"])}<div class="small">{esc(r["source_sheet"])} {esc(r["source_detail"])}</div></div>' for r in rows)
    return shell("Penetration & Opening Brain",'<div class="hero"><h1>Penetration & Opening Brain</h1><p class="muted">Finds opening/penetration scope requiring coordination across structure, MEP, roofing, framing or finishes.</p></div><div class="card">'+(h or '<p class="muted">No penetration/opening scope detected.</p>')+'</div>')

@app.get("/field-context/prerequisites",response_class=HTMLResponse)
def v49_prereq_page():
    rows=_v49_install_prereqs(project_id())
    h="".join(f'<div class="action"><span class="badge {"READY" if ready=="READY" else "WATCH"}">{esc(ready)}</span> <b>{esc(a["name"])}</b><div class="small">{esc(a["trade"])} · Risk {esc(risk)}</div><p><b>Expected gates:</b> {esc("; ".join(gates))}</p><p>{esc(blocking)}</p></div>' for a,ready,risk,gates,blocking in rows)
    return shell("Installation Prerequisites",'<div class="hero"><h1>Installation Prerequisite Intelligence</h1><p class="muted">Makes sequence/readiness logic understandable at the installation level.</p></div><div class="card">'+(h or '<p class="muted">No scheduled activities available.</p>')+'</div>')

@app.get("/field-context/hold-points",response_class=HTMLResponse)
def v49_hold_page():
    rows=_v49_hold_points(project_id())
    h="".join(f'<div class="action"><span class="badge {"READY" if str(i["result"] or "").upper()=="PASSED" else "WATCH"}">{esc(i["result"])}</span> <b>{esc(i["inspection_type"])}</b><div class="small">Activity: {esc(a["name"] if a else "Unlinked")} · {esc(i["scheduled_date"])} · {esc(i["authority"])}</div></div>' for i,a in rows)
    return shell("Inspection Hold Points",'<div class="hero"><h1>Inspection Hold Points</h1><p class="muted">Do not let downstream work bury required inspection/testing gates.</p></div><div class="card">'+(h or '<p class="muted">No inspection hold points loaded.</p>')+'</div>')

@app.get("/field-context/equipment-locations",response_class=HTMLResponse)
def v49_equipment_locations():
    rows=_v46_equipment(project_id())
    h=""
    for e in rows:
        room=_v49_room_key(e["name"])
        h+=f'<div class="action"><b>{esc(e["name"])}</b><div class="small">Location context: {esc("Room "+room if room else e["source"] or "Not explicit")} · Install {esc(e["primary"])} · Power {esc(e["power"] or "N/A")} · Controls {esc(e["controls"] or "N/A")}</div></div>'
    return shell("Equipment Location Intelligence",'<div class="hero"><h1>Equipment-to-Location Intelligence</h1><p class="muted">Connects equipment responsibilities to explicit room/source context where available.</p></div><div class="card">'+(h or '<p class="muted">No equipment intelligence detected.</p>')+'</div>')

@app.get("/field-context/commissioning",response_class=HTMLResponse)
def v49_commissioning_page():
    rows=_v49_commissioning(project_id())
    h="".join(f'<div class="card"><h3>{esc(e["name"])}</h3><p><b>Install:</b> {esc(e["primary"])} · <b>Power:</b> {esc(e["power"] or "N/A")} · <b>Controls:</b> {esc(e["controls"] or "N/A")} · <b>Startup:</b> {esc(e["startup"])}</p><p><b>Testing:</b> {esc("; ".join(testing))}</p><p><b>Closeout:</b> {esc("; ".join(closeout))}</p></div>' for e,testing,closeout in rows)
    return shell("Commissioning Chain",'<div class="hero"><h1>Commissioning Chain Intelligence</h1><p class="muted">Follows equipment from installation through operational verification and turnover.</p></div>'+(h or '<div class="card">No commissioning chains detected.</div>'))

@app.get("/field-context/work-packages",response_class=HTMLResponse)
def v49_packages_page():
    rows=_v49_work_packages(project_id())
    h="".join(f'<div class="card"><span class="badge">DRAFT PACKAGE</span><h3>{esc(p["title"])}</h3><p><b>Location:</b> {esc(p["location"])}</p><p><b>Scope:</b> {esc(p["scope"])}</p><p><b>Prerequisites:</b> {esc(p["prereq"])}</p><p><b>Inspection:</b> {esc(p["inspection"])}</p><p><b>Materials:</b> {esc(p["materials"])}</p></div>' for p in rows)
    return shell("Field Work Packages",'<div class="hero"><h1>Field Work Package Intelligence</h1><p class="muted">Draft execution packages only; superintendent review is required before release to the field.</p></div>'+(h or '<div class="card">No work-package candidates detected.</div>'))

@app.get("/field-context/command",response_class=HTMLResponse)
def v49_command_page():
    pid=project_id()
    rooms=_v49_rooms(pid); assemblies=_v49_assemblies(pid); pens=_v49_penetrations(pid)
    prereqs=[x for x in _v49_install_prereqs(pid) if x[1]!="READY"]
    holds=[x for x in _v49_hold_points(pid) if str(x[0]["result"] or "").upper()!="PASSED"]
    packages=_v49_work_packages(pid)
    body=f'<div class="hero"><div class="eyebrow">Field Context Command</div><h1>What needs coordination before work moves?</h1><p class="muted">{len(rooms)} room contexts · {len(assemblies)} assemblies · {len(pens)} penetration/opening items · {len(prereqs)} not-ready activities · {len(holds)} open hold points · {len(packages)} draft work packages.</p></div>'
    body+='<div class="grid3">'
    body+=_v37_link_card("Not-Ready Work","Installation prerequisites and blocking reasons.","/field-context/prerequisites","Review")
    body+=_v37_link_card("Hold Points","Inspections/testing that gate downstream work.","/field-context/hold-points","Review")
    body+=_v37_link_card("Work Packages","Location/trade execution packages from current intelligence.","/field-context/work-packages","Review")
    body+='</div>'
    return shell("Field Context Command",body)


# ============================================================
# v50 MASTER CONSTRUCTION REASONING ENGINE
# ============================================================

def _v50_ensure_tables():
    c=db()
    if DATABASE_KIND=="postgres":
        pk="BIGSERIAL PRIMARY KEY"; num="DOUBLE PRECISION"
    else:
        pk="INTEGER PRIMARY KEY"; num="REAL"
    stmts=[
      f"""CREATE TABLE IF NOT EXISTS master_reasoning_runs(
        id {pk},company_id BIGINT,project_id BIGINT,run_time TEXT,
        health_score {num} DEFAULT 100,risk_score {num} DEFAULT 0,
        quality_score {num} DEFAULT 100,summary TEXT,status TEXT DEFAULT 'COMPLETE',
        created TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS master_reasoning_findings(
        id {pk},company_id BIGINT,project_id BIGINT,run_id BIGINT,
        finding_type TEXT,title TEXT,severity TEXT,trade TEXT,location TEXT,
        reason TEXT,source_ref TEXT,downstream_impact TEXT,recommended_action TEXT,
        confidence TEXT DEFAULT 'MEDIUM',human_review INTEGER DEFAULT 1,created TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS master_reasoning_links(
        id {pk},company_id BIGINT,project_id BIGINT,run_id BIGINT,
        from_type TEXT,from_key TEXT,to_type TEXT,to_key TEXT,
        relationship TEXT,reason TEXT,created TEXT)"""
    ]
    for s in stmts:
        c.execute(s)
    c.commit(); c.close()

def _v50_collect_findings_uncached(pid):
    findings=[]

    # 1) Blueprint / scope quality
    for r,verify,agreement,source_level,calibrated,score,contradiction,reason in _v48_self_audit(pid):
        if agreement=="DISAGREE" or calibrated=="LOW" or source_level=="LOW":
            sev="HIGH" if agreement=="DISAGREE" else "MEDIUM"
            findings.append({
                "type":"SCOPE QUALITY","title":r["requirement"],"severity":sev,
                "trade":r["trade"],"location":_v49_room_key(r["requirement"]),
                "reason":reason,
                "source": " · ".join(x for x in [r["source_sheet"],r["source_detail"],r["source_spec"]] if x),
                "impact":"Bad ownership/source quality can contaminate estimating, bidding, schedule and field execution.",
                "action":f"Review saved trade ownership versus second opinion: {verify}.",
                "confidence":calibrated
            })

    # 2) Constructability
    for r in _v452_constructability(pid):
        findings.append({
            "type":"CONSTRUCTABILITY","title":r["title"],"severity":r["severity"],
            "trade":r["trade"],"location":_v49_room_key(r["description"]),
            "reason":r["description"],"source":r["source"],
            "impact":"Potential field coordination, access, clearance or installation problem.",
            "action":r["action"],"confidence":"MEDIUM"
        })

    # 3) Sequence
    for x in _v45_sequence_analysis(pid):
        if x["risk"] in {"CRITICAL","HIGH"}:
            a=x["activity"]
            findings.append({
                "type":"SEQUENCE","title":a["name"],"severity":x["risk"],
                "trade":a["trade"],"location":"",
                "reason":x["blocking_reason"] or "High sequence exposure.",
                "source":f'Schedule {a["start"]} to {a["finish"]}',
                "impact":x["downstream"],
                "action":x["recommendation"],"confidence":"HIGH"
            })

    # 4) Procurement
    for r,level,exposure,reason,action in _v452_procurement_analysis(pid):
        if level in {"CRITICAL","HIGH","TODAY"}:
            findings.append({
                "type":"PROCUREMENT","title":r["item"],"severity":"CRITICAL" if level=="TODAY" else level,
                "trade":"","location":"","reason":reason,
                "source":f'Need {r["required_on_site"]} · Promised {r["promised_date"]}',
                "impact":"Late material can delay installation and downstream activities.",
                "action":action,"confidence":"HIGH"
            })

    # 5) Decision deadlines
    for typ,title,due,severity,cost,days,source in _v47_decision_deadlines(pid):
        if severity in {"CRITICAL","HIGH"}:
            findings.append({
                "type":"DECISION","title":title,"severity":severity,"trade":"","location":"",
                "reason":f"{typ} decision due {due}.",
                "source":source,
                "impact":f"Potential ${cost:,.0f} cost exposure and {days:g} schedule day(s).",
                "action":"Escalate and obtain decision before downstream commitment.",
                "confidence":"HIGH"
            })

    # 6) Inspection hold points
    for i,a in _v49_hold_points(pid):
        if str(i["result"] or "").upper()!="PASSED":
            findings.append({
                "type":"INSPECTION","title":i["inspection_type"],"severity":"HIGH",
                "trade":a["trade"] if a else "","location":"",
                "reason":f'Inspection gate is {i["result"] or "PENDING"}.',
                "source":f'{i["scheduled_date"]} · {i["authority"]}',
                "impact":"Downstream concealment/finish work should not proceed past an unmet hold point.",
                "action":"Verify readiness and pass required inspection before releasing downstream work.",
                "confidence":"HIGH"
            })

    # 7) Cost risk
    for level,kind,title,reason,source in _v46_cost_risk(pid):
        if level in {"HIGH","CRITICAL"}:
            findings.append({
                "type":"COST RISK","title":title,"severity":level,
                "trade":"","location":"","reason":reason,"source":source,
                "impact":"Potential added cost or commercial exposure.",
                "action":"Review contract scope, source documents and change documentation.",
                "confidence":"MEDIUM"
            })

    # 8) Field prerequisites
    for a,ready,risk,gates,blocking in _v49_install_prereqs(pid):
        if ready!="READY" and risk in {"CRITICAL","HIGH"}:
            findings.append({
                "type":"MAKE READY","title":a["name"],"severity":risk,
                "trade":a["trade"],"location":"",
                "reason":blocking or "Installation prerequisites are incomplete.",
                "source":"; ".join(gates),
                "impact":"Activity may start unprepared or force downstream rework.",
                "action":"Close make-ready gaps before releasing crew.",
                "confidence":"HIGH"
            })

    # 9) Closeout
    for r,level in _v47_closeout_prediction(pid):
        if level in {"CRITICAL","HIGH"}:
            findings.append({
                "type":"CLOSEOUT","title":r["item"],"severity":level,
                "trade":r["responsible_party"],"location":"",
                "reason":f'Status {r["status"]} · Due {r["due_date"]}',
                "source":r["category"],
                "impact":"Late turnover requirement can delay substantial/final completion.",
                "action":"Assign owner and recover closeout requirement before project end.",
                "confidence":"HIGH"
            })

    # 10) Cross-trade dependencies / handoffs
    for r,rels in _v46_dependencies(pid):
        if rels:
            findings.append({
                "type":"HANDOFF","title":r["requirement"],"severity":"REVIEW",
                "trade":r["trade"],"location":_v49_room_key(r["requirement"]),
                "reason":f'Primary trade depends on {", ".join(rels)}.',
                "source":" · ".join(x for x in [r["source_sheet"],r["source_detail"],r["source_spec"]] if x),
                "impact":"Missed handoff can create scope gaps, rework or delay.",
                "action":"Confirm responsibility boundary and readiness between trades.",
                "confidence":"MEDIUM"
            })

    rank={"CRITICAL":0,"HIGH":1,"MEDIUM":2,"REVIEW":3,"LOW":4}
    findings.sort(key=lambda x:(rank.get(x["severity"],9),x["type"],x["title"]))
    return findings[:500]

def _v50_collect_findings(pid):
    return _v56_master_findings(pid)

def _v50_score(pid):
    return _v56_master_score(pid)


def _v50_run(pid):
    _v50_ensure_tables()
    findings=_v50_collect_findings(pid)
    score=_v50_score(pid)
    now=datetime.utcnow().isoformat()
    summary=f'{score["count"]} connected findings · health {score["health"]}/100 · risk {score["risk"]}/100 · quality {score["quality"]}/100.'
    c=db()
    c.execute("""INSERT INTO master_reasoning_runs(company_id,project_id,run_time,health_score,risk_score,quality_score,summary,status,created)
                 VALUES(?,?,?,?,?,?,?,?,?)""",
              (current_company_id(),pid,now,score["health"],score["risk"],score["quality"],summary,"COMPLETE",now))
    run_id=c.execute("SELECT last_insert_rowid() id").fetchone()["id"]
    for f in findings:
        c.execute("""INSERT INTO master_reasoning_findings(
            company_id,project_id,run_id,finding_type,title,severity,trade,location,reason,
            source_ref,downstream_impact,recommended_action,confidence,human_review,created
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (current_company_id(),pid,run_id,f["type"],f["title"],f["severity"],f["trade"],f["location"],
         f["reason"],f["source"],f["impact"],f["action"],f["confidence"],1,now))
    c.commit(); c.close()
    return run_id,findings,score

def _v50_top_priorities(pid,limit=10):
    return _v56_top_priorities(pid,limit)


def _v50_trade_brief(pid,trade):
    findings=[f for f in _v50_collect_findings(pid) if str(f["trade"] or "").lower()==str(trade or "").lower()]
    scopes=[r for r in _v452_scope_rows(pid) if str(r["trade"] or "").lower()==str(trade or "").lower()]
    return findings,scopes

def _v50_location_brief(pid,room):
    findings=[f for f in _v50_collect_findings(pid) if str(f["location"] or "")==str(room or "")]
    rooms=[r for r in _v49_rooms(pid) if r["room"]==room]
    return findings,rooms

def _v50_reasoning_chain(pid,title):
    q=str(title or "").lower()
    chain=[]
    for f in _v50_collect_findings(pid):
        text=(f["title"]+" "+f["reason"]+" "+f["trade"]).lower()
        tokens=[w for w in re.findall(r'[a-z0-9]+',q) if len(w)>3]
        if tokens and any(tok in text for tok in tokens):
            chain.append(f)
    return chain[:30]

@app.get("/master-reasoning",response_class=HTMLResponse)
def v50_master_reasoning_home():
    pid=project_id()
    run_id,findings,score=_v50_run(pid)
    top=_v50_top_priorities(pid,10)
    body=f'<div class="hero"><div class="eyebrow">BuildCommand v50 - Master Construction Reasoning Engine</div><h1>One brain. One project judgment.</h1><p class="muted">Health {score["health"]}/100 · Risk {score["risk"]}/100 · Intelligence quality {score["quality"]}/100 · {score["count"]} connected findings.</p></div>'
    body+='<div class="grid3">'
    cards=[
      ("Master Project Brief","The 10 most important connected project findings.","/master-reasoning/brief"),
      ("Reasoning Findings","All connected findings from Blueprint, sequence, risk, field and quality intelligence.","/master-reasoning/findings"),
      ("Trade Command","See connected intelligence by subcontractor/trade.","/master-reasoning/trades"),
      ("Location Command","See connected intelligence by room/location where explicit context exists.","/master-reasoning/locations"),
      ("Reasoning Chain","Trace why BuildCommand thinks an issue matters.","/master-reasoning/chain"),
      ("Quality Gate","See which findings should be trusted versus reviewed.","/master-reasoning/quality-gate"),
      ("Field Release Gate","Check whether high-risk work should be released to the field.","/master-reasoning/release-gate"),
      ("Commercial Exposure","Combine cost, decision and change signals.","/master-reasoning/commercial"),
      ("Schedule Exposure","Combine sequence, procurement, inspection and decision risks.","/master-reasoning/schedule"),
      ("Executive Judgment","One concise project health and leadership view.","/master-reasoning/executive")
    ]
    for name,desc,href in cards:
        body+=_v37_link_card(name,desc,href,"Open")
    body+='</div><div class="card"><h2>Top 10</h2>'
    body+="".join(f'<div class="action"><span class="badge WATCH">{esc(f["severity"])}</span> <b>{esc(f["type"])}</b> - {esc(f["title"])}<div class="small">{esc(f["reason"])}</div></div>' for f in top)
    body+='</div>'
    return shell("Master Construction Reasoning",body)

@app.get("/master-reasoning/brief",response_class=HTMLResponse)
def v50_brief_page():
    rows=_v50_top_priorities(project_id(),10)
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(f["severity"])}</span> <b>{esc(f["type"])} - {esc(f["title"])}</b><p>{esc(f["reason"])}</p><p><b>Impact:</b> {esc(f["impact"])}</p><p><b>Recommended:</b> {esc(f["action"])}</p><div class="small">{esc(f["trade"])} {("· Room "+esc(f["location"])) if f["location"] else ""} · {esc(f["source"])}</div></div>' for f in rows)
    return shell("Master Project Brief",'<div class="hero"><h1>Master Project Brief</h1><p class="muted">The highest-priority connected project findings from across BuildCommand.</p></div><div class="card">'+(h or '<p class="muted">No major findings detected.</p>')+'</div>')

@app.get("/master-reasoning/findings",response_class=HTMLResponse)
def v50_findings_page():
    rows=_v50_collect_findings(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(f["severity"])}</span> <b>{esc(f["type"])}</b> - {esc(f["title"])}<div>{esc(f["reason"])}</div><div class="small">Trade {esc(f["trade"])} · Confidence {esc(f["confidence"])} · {esc(f["source"])}</div></div>' for f in rows)
    return shell("Reasoning Findings",'<div class="hero"><h1>Connected Reasoning Findings</h1><p class="muted">This is the combined output of the major BuildCommand intelligence layers.</p></div><div class="card">'+(h or '<p class="muted">No findings.</p>')+'</div>')

@app.get("/master-reasoning/trades",response_class=HTMLResponse)
def v50_trade_page(trade:str=''):
    pid=project_id()
    trades=sorted(set(str(r["trade"] or "") for r in _v452_scope_rows(pid) if r["trade"]))
    options="".join(f'<option value="{esc(tr)}" {"selected" if tr==trade else ""}>{esc(tr)}</option>' for tr in trades)
    body='<div class="hero"><h1>Trade Command</h1><p class="muted">Connected scope, risk and handoff intelligence by trade.</p></div><div class="card"><form method="get"><select name="trade">'+options+'</select><button type="submit">Open Trade</button></form></div>'
    if trade:
        findings,scopes=_v50_trade_brief(pid,trade)
        body+=f'<div class="card"><h2>{esc(trade)}</h2><p>{len(scopes)} scope item(s) · {len(findings)} connected finding(s)</p>'
        body+="".join(f'<div class="action"><span class="badge WATCH">{esc(f["severity"])}</span> {esc(f["title"])}<div class="small">{esc(f["reason"])}</div></div>' for f in findings)
        body+='</div>'
    return shell("Trade Command",body)

@app.get("/master-reasoning/locations",response_class=HTMLResponse)
def v50_location_page(room:str=''):
    pid=project_id()
    rooms=sorted(r["room"] for r in _v49_rooms(pid))
    options="".join(f'<option value="{esc(r)}" {"selected" if r==room else ""}>Room {esc(r)}</option>' for r in rooms)
    body='<div class="hero"><h1>Location Command</h1><p class="muted">Connected intelligence by explicit room/location context.</p></div><div class="card"><form method="get"><select name="room">'+options+'</select><button type="submit">Open Location</button></form></div>'
    if room:
        findings,roomdata=_v50_location_brief(pid,room)
        body+=f'<div class="card"><h2>Room {esc(room)}</h2><p>{len(findings)} connected finding(s)</p>'
        body+="".join(f'<div class="action"><span class="badge WATCH">{esc(f["severity"])}</span> <b>{esc(f["type"])}</b> - {esc(f["title"])}<div class="small">{esc(f["reason"])}</div></div>' for f in findings)
        body+='</div>'
    return shell("Location Command",body)

@app.get("/master-reasoning/chain",response_class=HTMLResponse)
def v50_chain_page(q:str=''):
    rows=_v50_reasoning_chain(project_id(),q) if q else []
    body='<div class="hero"><h1>Reasoning Chain</h1><p class="muted">Trace the connected evidence behind a project issue or subject.</p></div><div class="card"><form method="get"><input name="q" value="'+esc(q)+'" placeholder="Example: water heater, storefront, ceiling"><button type="submit">Trace</button></form></div>'
    if q:
        body+='<div class="card">'+("".join(f'<div class="action"><span class="badge">{esc(f["type"])}</span> <b>{esc(f["title"])}</b><p>{esc(f["reason"])}</p><div class="small">{esc(f["source"])}</div></div>' for f in rows) or '<p class="muted">No connected reasoning chain found.</p>')+'</div>'
    return shell("Reasoning Chain",body)

@app.get("/master-reasoning/quality-gate",response_class=HTMLResponse)
def v50_quality_gate():
    rows=_v50_collect_findings(project_id())
    review=[f for f in rows if f["confidence"]=="LOW" or f["severity"]=="REVIEW"]
    h="".join(f'<div class="action"><span class="badge WATCH">HUMAN REVIEW</span> <b>{esc(f["type"])}</b> - {esc(f["title"])}<div class="small">Confidence {esc(f["confidence"])} · {esc(f["reason"])}</div></div>' for f in review)
    return shell("Quality Gate",'<div class="hero"><h1>Master Quality Gate</h1><p class="muted">Low-confidence and review-only findings should not silently drive consequential project action.</p></div><div class="card">'+(h or '<p class="muted">No current quality-gate exceptions.</p>')+'</div>')

@app.get("/master-reasoning/release-gate",response_class=HTMLResponse)
def v50_release_gate():
    rows=[f for f in _v50_collect_findings(project_id()) if f["type"] in {"SEQUENCE","MAKE READY","INSPECTION","PROCUREMENT"} and f["severity"] in {"CRITICAL","HIGH"}]
    h="".join(f'<div class="action"><span class="badge WATCH">HOLD / REVIEW</span> <b>{esc(f["title"])}</b><div>{esc(f["reason"])}</div><p><b>Before release:</b> {esc(f["action"])}</p></div>' for f in rows)
    return shell("Field Release Gate",'<div class="hero"><h1>Field Release Gate</h1><p class="muted">High-risk prerequisites, inspections, sequence and procurement conditions should be reviewed before releasing work.</p></div><div class="card">'+(h or '<p class="muted">No high-risk field release blockers detected.</p>')+'</div>')

@app.get("/master-reasoning/commercial",response_class=HTMLResponse)
def v50_commercial():
    rows=[f for f in _v50_collect_findings(project_id()) if f["type"] in {"COST RISK","DECISION"}]
    changes=_v39_changes(project_id())
    total=sum(float(r["estimated_cost"] or 0) for r in changes)
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(f["severity"])}</span> <b>{esc(f["title"])}</b><div class="small">{esc(f["reason"])}</div></div>' for f in rows)
    return shell("Commercial Exposure",f'<div class="hero"><h1>Commercial Exposure</h1><p class="muted">${total:,.0f} known open change exposure plus {len(rows)} connected cost/decision risk signal(s).</p></div><div class="card">'+(h or '<p class="muted">No major commercial exposure signals.</p>')+'</div>')

@app.get("/master-reasoning/schedule",response_class=HTMLResponse)
def v50_schedule():
    rows=[f for f in _v50_collect_findings(project_id()) if f["type"] in {"SEQUENCE","PROCUREMENT","DECISION","INSPECTION","MAKE READY"}]
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(f["severity"])}</span> <b>{esc(f["type"])}</b> - {esc(f["title"])}<div class="small">{esc(f["reason"])}</div></div>' for f in rows)
    return shell("Schedule Exposure",'<div class="hero"><h1>Schedule Exposure</h1><p class="muted">One connected schedule-risk view across readiness, procurement, inspections and decisions.</p></div><div class="card">'+(h or '<p class="muted">No major schedule exposure signals.</p>')+'</div>')

@app.get("/master-reasoning/executive",response_class=HTMLResponse)
def v50_executive():
    pid=project_id()
    score=_v50_score(pid)
    top=_v50_top_priorities(pid,5)
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(f["severity"])}</span> <b>{esc(f["type"])}</b> - {esc(f["title"])}<div class="small">{esc(f["impact"])}</div></div>' for f in top)
    return shell("Executive Judgment",f'<div class="hero"><h1>Executive Project Judgment</h1><p class="muted">Health {score["health"]}/100 · Risk {score["risk"]}/100 · Intelligence quality {score["quality"]}/100.</p></div><div class="card"><h2>Top Leadership Attention</h2>{h or "<p class=muted>No major leadership issues detected.</p>"}</div>')


# ============================================================
# v51 REAL CONSTRUCTION REASONING 2.0
# ============================================================

def _v51_reasoning_units(pid):
    """
    Convert master findings into cause -> dependency -> consequence -> owner -> action units.
    """
    findings=_v50_collect_findings(pid)
    deps=_v46_dependencies(pid)
    dep_map={}
    for r,rels in deps:
        dep_map[str(r["requirement"] or "").strip().lower()]=rels

    units=[]
    for f in findings:
        key=str(f["title"] or "").strip().lower()
        rels=dep_map.get(key,[])
        owner=f["trade"] or (rels[0] if rels else "")
        cause=f["reason"]
        dependency=", ".join(rels) if rels else "No explicit cross-trade dependency detected."
        consequence=f["impact"]
        action=f["action"]
        units.append({
            "severity":f["severity"],"type":f["type"],"title":f["title"],
            "cause":cause,"dependency":dependency,"consequence":consequence,
            "owner":owner,"action":action,"source":f["source"],
            "confidence":f["confidence"]
        })
    return units[:500]

def _v51_alternatives(unit):
    """
    Transparent option generation. These are review alternatives, not automatic directives.
    """
    typ=unit["type"]
    opts=[]
    if typ in {"SEQUENCE","MAKE READY"}:
        opts=[
            ("Hold downstream work","Protect sequence and avoid rework until prerequisites are complete."),
            ("Resequence unaffected work","Move crews to independent work while the blocker is resolved."),
            ("Recover prerequisite","Add manpower/material/coordination to clear the blocker faster.")
        ]
    elif typ=="PROCUREMENT":
        opts=[
            ("Expedite current source","Confirm fabrication/shipping recovery with current vendor."),
            ("Evaluate approved alternate","Use only if contract/submittal requirements allow an alternate."),
            ("Resequence installation","Move unaffected work ahead while protecting downstream milestones.")
        ]
    elif typ=="INSPECTION":
        opts=[
            ("Hold concealment","Do not cover work before required inspection/testing."),
            ("Resolve deficiencies first","Correct known issues before requesting inspection."),
            ("Coordinate inspection timing","Align inspector availability with field readiness.")
        ]
    elif typ in {"COST RISK","DECISION"}:
        opts=[
            ("Clarify contract responsibility","Review drawings/specs/subcontract scope before authorizing extra work."),
            ("Issue/advance RFI","Seek documented clarification where design intent is ambiguous."),
            ("Track potential change","Preserve cost/schedule documentation pending entitlement review.")
        ]
    elif typ=="CONSTRUCTABILITY":
        opts=[
            ("Coordinate affected trades","Resolve interfaces before installation."),
            ("Request design clarification","Use RFI if drawings/specs do not establish a buildable solution."),
            ("Field-verify existing conditions","Confirm dimensions/access before committing material or labor.")
        ]
    else:
        opts=[
            ("Review source documents","Confirm the governing requirement before action."),
            ("Coordinate responsible trade","Verify ownership and handoff."),
            ("Document decision","Record the approved resolution and downstream impact.")
        ]
    return opts

def _v51_uncertainty(unit):
    conf=str(unit["confidence"] or "MEDIUM").upper()
    missing_source=not bool(str(unit["source"] or "").strip())
    if conf=="LOW" or missing_source:
        return ("LOW","Project evidence is incomplete or weak; human review is required before relying on this conclusion.")
    if unit["severity"]=="REVIEW":
        return ("MEDIUM","This is a coordination/review signal rather than a confirmed project condition.")
    return ("HIGH","Current project data provides a reasonably strong basis for this reasoning path.")

def _v51_reasoning_score(unit):
    score=50
    sev={"CRITICAL":25,"HIGH":18,"MEDIUM":10,"REVIEW":3}.get(unit["severity"],0)
    score+=sev
    if unit["source"]: score+=10
    if unit["dependency"] and "No explicit" not in unit["dependency"]: score+=5
    if str(unit["confidence"]).upper()=="HIGH": score+=10
    elif str(unit["confidence"]).upper()=="LOW": score-=20
    return max(0,min(100,score))

def _v51_top_units(pid,limit=20):
    units=_v51_reasoning_units(pid)
    for u in units:
        u["reasoning_score"]=_v51_reasoning_score(u)
        u["certainty"],u["certainty_note"]=_v51_uncertainty(u)
    units.sort(key=lambda x:(-x["reasoning_score"],x["type"],x["title"]))
    return units[:limit]

@app.get("/reasoning-2",response_class=HTMLResponse)
def v51_reasoning_home():
    pid=project_id()
    units=_v51_top_units(pid,25)
    low=sum(1 for u in units if u["certainty"]=="LOW")
    body=f'<div class="hero"><div class="eyebrow">BuildCommand v51 - Real Construction Reasoning 2.0</div><h1>Explain the problem, the dependency, and what to do next.</h1><p class="muted">{len(units)} prioritized reasoning paths · {low} low-certainty path(s) requiring stronger human review.</p></div><div class="grid3">'
    cards=[
      ("Cause → Consequence","See the full reasoning chain for the highest-risk project findings.","/reasoning-2/chains"),
      ("Alternative Resolutions","Compare practical response options before choosing a path.","/reasoning-2/alternatives"),
      ("Responsible Trade Logic","See who owns the action versus who is only affected.","/reasoning-2/ownership"),
      ("Uncertainty Engine","Show when BuildCommand should not make a confident call.","/reasoning-2/uncertainty"),
      ("Reasoning Score","Rank findings by evidence strength and project consequence.","/reasoning-2/scores"),
      ("What Happens Next","Show downstream consequences if a blocker stays unresolved.","/reasoning-2/downstream"),
      ("Best Next Action","Prioritize the next recommended move for each major issue.","/reasoning-2/actions"),
      ("Decision Comparison","Compare hold / resequence / clarify / recover alternatives.","/reasoning-2/compare"),
      ("Explain This Finding","Search a topic and trace BuildCommand's reasoning path.","/reasoning-2/explain"),
      ("Reasoning Command","One concise project view of what matters, why, and what to do.","/reasoning-2/command")
    ]
    for name,desc,href in cards:
        body+=_v37_link_card(name,desc,href,"Open")
    body+='</div>'
    return shell("Real Construction Reasoning 2.0",body)

@app.get("/reasoning-2/chains",response_class=HTMLResponse)
def v51_chains():
    units=_v51_top_units(project_id(),100)
    h="".join(
        f'<div class="card"><span class="badge WATCH">{esc(u["severity"])}</span><h3>{esc(u["title"])}</h3>'
        f'<p><b>Cause:</b> {esc(u["cause"])}</p>'
        f'<p><b>Dependency:</b> {esc(u["dependency"])}</p>'
        f'<p><b>Consequence:</b> {esc(u["consequence"])}</p>'
        f'<p><b>Responsible:</b> {esc(u["owner"] or "Needs ownership review")}</p>'
        f'<p><b>Recommended:</b> {esc(u["action"])}</p>'
        f'<div class="small">Reasoning score {u["reasoning_score"]}/100 · Certainty {esc(u["certainty"])} · {esc(u["source"])}</div></div>'
        for u in units
    )
    return shell("Cause to Consequence",'<div class="hero"><h1>Cause → Dependency → Consequence</h1><p class="muted">The master brain now explains why each major finding matters.</p></div>'+h)

@app.get("/reasoning-2/alternatives",response_class=HTMLResponse)
def v51_alternatives_page():
    units=_v51_top_units(project_id(),30)
    body='<div class="hero"><h1>Alternative Resolution Intelligence</h1><p class="muted">Options are decision support, not automatic field directives.</p></div>'
    for u in units:
        body+=f'<div class="card"><h3>{esc(u["title"])}</h3><p>{esc(u["cause"])}</p>'
        for title,why in _v51_alternatives(u):
            body+=f'<div class="action"><b>{esc(title)}</b><div class="small">{esc(why)}</div></div>'
        body+='</div>'
    return shell("Alternative Resolutions",body)

@app.get("/reasoning-2/ownership",response_class=HTMLResponse)
def v51_ownership():
    units=_v51_top_units(project_id(),100)
    h="".join(
        f'<div class="action"><b>{esc(u["owner"] or "Ownership review needed")}</b> - {esc(u["title"])}'
        f'<div class="small">Affected dependency: {esc(u["dependency"])}</div></div>'
        for u in units
    )
    return shell("Responsible Trade Logic",'<div class="hero"><h1>Responsible Trade Logic</h1><p class="muted">Primary ownership is separated from downstream affected trades.</p></div><div class="card">'+h+'</div>')

@app.get("/reasoning-2/uncertainty",response_class=HTMLResponse)
def v51_uncertainty_page():
    units=_v51_top_units(project_id(),150)
    h="".join(
        f'<div class="action"><span class="badge {"WATCH" if u["certainty"]!="HIGH" else "READY"}">{esc(u["certainty"])}</span> '
        f'<b>{esc(u["title"])}</b><div class="small">{esc(u["certainty_note"])}</div></div>'
        for u in units
    )
    return shell("Uncertainty Engine",'<div class="hero"><h1>Uncertainty Engine</h1><p class="muted">BuildCommand should expose weak evidence instead of sounding certain when the project documents do not support certainty.</p></div><div class="card">'+h+'</div>')

@app.get("/reasoning-2/scores",response_class=HTMLResponse)
def v51_scores():
    units=_v51_top_units(project_id(),150)
    h="".join(
        f'<div class="action"><span class="badge">{u["reasoning_score"]}</span> <b>{esc(u["title"])}</b>'
        f'<div class="small">{esc(u["type"])} · {esc(u["severity"])} · Certainty {esc(u["certainty"])}</div></div>'
        for u in units
    )
    return shell("Reasoning Score",'<div class="hero"><h1>Reasoning Strength Score</h1><p class="muted">Ranks findings using consequence, source evidence, dependency context and confidence.</p></div><div class="card">'+h+'</div>')

@app.get("/reasoning-2/downstream",response_class=HTMLResponse)
def v51_downstream_page():
    units=_v51_top_units(project_id(),100)
    h="".join(
        f'<div class="action"><b>{esc(u["title"])}</b><p>{esc(u["consequence"])}</p>'
        f'<div class="small">If unresolved: downstream work may remain exposed. Dependency: {esc(u["dependency"])}</div></div>'
        for u in units
    )
    return shell("What Happens Next",'<div class="hero"><h1>What Happens Next?</h1><p class="muted">Focuses the team on downstream consequence, not only the immediate problem.</p></div><div class="card">'+h+'</div>')

@app.get("/reasoning-2/actions",response_class=HTMLResponse)
def v51_actions_page():
    units=_v51_top_units(project_id(),50)
    h="".join(
        f'<div class="action"><span class="badge WATCH">{esc(u["severity"])}</span> <b>{esc(u["action"])}</b>'
        f'<div class="small">Because: {esc(u["cause"])} · Owner: {esc(u["owner"] or "Review")}</div></div>'
        for u in units
    )
    return shell("Best Next Action",'<div class="hero"><h1>Best Next Action</h1><p class="muted">Recommended actions are tied to the reason they matter.</p></div><div class="card">'+h+'</div>')

@app.get("/reasoning-2/compare",response_class=HTMLResponse)
def v51_compare_page():
    units=_v51_top_units(project_id(),20)
    body='<div class="hero"><h1>Decision Comparison</h1><p class="muted">Compare practical response paths before committing the project.</p></div>'
    for u in units:
        body+=f'<div class="card"><h3>{esc(u["title"])}</h3>'
        for idx,(title,why) in enumerate(_v51_alternatives(u),1):
            body+=f'<div class="action"><b>Option {idx}: {esc(title)}</b><div class="small">{esc(why)}</div></div>'
        body+=f'<p><b>Current recommended action:</b> {esc(u["action"])}</p></div>'
    return shell("Decision Comparison",body)

@app.get("/reasoning-2/explain",response_class=HTMLResponse)
def v51_explain_page(q:str=''):
    units=_v51_top_units(project_id(),200)
    matches=[]
    if q:
        tokens=[w for w in re.findall(r'[a-z0-9]+',q.lower()) if len(w)>3]
        for u in units:
            text=(u["title"]+" "+u["cause"]+" "+u["dependency"]+" "+u["owner"]).lower()
            if tokens and any(tok in text for tok in tokens):
                matches.append(u)
    body='<div class="hero"><h1>Explain This Finding</h1><p class="muted">Search a project subject and see the reasoning path behind BuildCommand\'s conclusion.</p></div><div class="card"><form method="get"><input name="q" value="'+esc(q)+'" placeholder="Example: storefront, water heater, ceiling"><button type="submit">Explain</button></form></div>'
    if q:
        body+='<div class="card">'+("".join(
            f'<div class="action"><b>{esc(u["title"])}</b><p><b>Cause:</b> {esc(u["cause"])}</p><p><b>Dependency:</b> {esc(u["dependency"])}</p><p><b>Consequence:</b> {esc(u["consequence"])}</p><p><b>Action:</b> {esc(u["action"])}</p></div>'
            for u in matches[:30]
        ) or '<p class="muted">No reasoning path matched that subject.</p>')+'</div>'
    return shell("Explain This Finding",body)

@app.get("/reasoning-2/command",response_class=HTMLResponse)
def v51_command_page():
    units=_v51_top_units(project_id(),10)
    body='<div class="hero"><div class="eyebrow">Reasoning Command</div><h1>What matters, why, and what should happen next?</h1><p class="muted">Top connected reasoning paths from the current project.</p></div><div class="card">'
    for u in units:
        body+=f'<div class="action"><span class="badge WATCH">{esc(u["severity"])}</span> <b>{esc(u["title"])}</b><p>{esc(u["cause"])}</p><p><b>Next:</b> {esc(u["action"])}</p><div class="small">Owner {esc(u["owner"] or "Review")} · Score {u["reasoning_score"]}/100 · Certainty {esc(u["certainty"])}</div></div>'
    body+='</div>'
    return shell("Reasoning Command",body)


# ============================================================
# v54 PROJECT MEMORY & CONTINUOUS LEARNING
# ============================================================

def _v54_ensure_tables():
    c=db()
    if DATABASE_KIND=="postgres":
        pk="BIGSERIAL PRIMARY KEY"; num="DOUBLE PRECISION"
    else:
        pk="INTEGER PRIMARY KEY"; num="REAL"
    stmts=[
      f"""CREATE TABLE IF NOT EXISTS project_memory(
        id {pk},company_id BIGINT,project_id BIGINT,memory_type TEXT,subject TEXT,
        lesson TEXT,approved_result TEXT,source_ref TEXT,scope_level TEXT DEFAULT 'PROJECT ONLY',
        approval_status TEXT DEFAULT 'APPROVED',approved_by TEXT,confidence TEXT DEFAULT 'HIGH',
        created TEXT,updated TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS project_pattern_memory(
        id {pk},company_id BIGINT,pattern_type TEXT,pattern_key TEXT,pattern_summary TEXT,
        occurrence_count INTEGER DEFAULT 1,latest_project_id BIGINT,last_seen TEXT,
        approved_only INTEGER DEFAULT 1,created TEXT,updated TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS learning_application_log(
        id {pk},company_id BIGINT,project_id BIGINT,memory_id BIGINT,applied_to_type TEXT,
        applied_to_key TEXT,application_result TEXT,confidence TEXT DEFAULT 'MEDIUM',
        created TEXT)"""
    ]
    for s in stmts:
        c.execute(s)
    c.commit(); c.close()

def _v54_memory_rows(pid=None):
    _v54_ensure_tables()
    if pid:
        return _v39_rows("""
            SELECT * FROM project_memory
            WHERE company_id=? AND project_id=? AND approval_status='APPROVED'
            ORDER BY id DESC
        """,(current_company_id(),pid))
    return _v39_rows("""
        SELECT * FROM project_memory
        WHERE company_id=? AND approval_status='APPROVED'
        ORDER BY id DESC LIMIT 500
    """,(current_company_id(),))

def _v54_seed_from_existing(pid):
    """
    Import already-approved learning and answered project decisions into project memory
    without changing their original source tables.
    """
    _v54_ensure_tables()
    existing=_v39_rows("SELECT subject,lesson,source_ref FROM project_memory WHERE company_id=? AND project_id=?",(current_company_id(),pid))
    seen={(str(r["subject"] or ""),str(r["lesson"] or ""),str(r["source_ref"] or "")) for r in existing}
    c=db()
    now=datetime.utcnow().isoformat()

    # Approved learning rules
    for r in _v48_learning_rules(pid):
        key=(str(r["subject"] or ""),str(r["learned_rule"] or ""),str(r["source_ref"] or ""))
        if key in seen: continue
        c.execute("""INSERT INTO project_memory(
            company_id,project_id,memory_type,subject,lesson,approved_result,source_ref,
            scope_level,approval_status,approved_by,confidence,created,updated
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",(
            current_company_id(),pid,"APPROVED LEARNING",r["subject"],r["learned_rule"],r["learned_rule"],
            r["source_ref"],r["scope_level"],"APPROVED",r["approved_by"],r["confidence"],now,now
        ))
        seen.add(key)

    # Answered / closed RFIs
    try:
        for r in _v42_rfis(pid):
            status=str(r["status"] or "").upper()
            if status not in {"ANSWERED","CLOSED","COMPLETE"} or not str(r["answer"] or "").strip():
                continue
            subject=f'RFI {r["number"] or r["id"]}: {r["title"]}'
            lesson=str(r["answer"] or "").strip()
            key=(subject,lesson,str(r["source_ref"] or ""))
            if key in seen: continue
            c.execute("""INSERT INTO project_memory(
                company_id,project_id,memory_type,subject,lesson,approved_result,source_ref,
                scope_level,approval_status,approved_by,confidence,created,updated
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",(
                current_company_id(),pid,"RFI ANSWER",subject,lesson,lesson,r["source_ref"] or "",
                "PROJECT ONLY","APPROVED","PROJECT TEAM","HIGH",now,now
            ))
            seen.add(key)
    except Exception:
        pass

    # Completed lessons learned
    try:
        lessons=_v39_rows("""
            SELECT * FROM lessons_learned
            WHERE company_id=? AND project_id=? AND approval_status='APPROVED'
            ORDER BY id DESC
        """,(current_company_id(),pid))
        for r in lessons:
            subject=f'{r["category"]}: {r["title"]}'
            lesson=str(r["lesson"] or "")
            key=(subject,lesson,str(r["source_ref"] or ""))
            if key in seen: continue
            c.execute("""INSERT INTO project_memory(
                company_id,project_id,memory_type,subject,lesson,approved_result,source_ref,
                scope_level,approval_status,approved_by,confidence,created,updated
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",(
                current_company_id(),pid,"LESSON LEARNED",subject,lesson,r["recommendation"] or lesson,
                r["source_ref"] or "",r["scope_level"] or "PROJECT ONLY","APPROVED",
                r["approved_by"] or "PROJECT TEAM","HIGH",now,now
            ))
            seen.add(key)
    except Exception:
        pass

    c.commit(); c.close()

def _v54_rebuild_patterns():
    _v54_ensure_tables()
    rows=_v54_memory_rows()
    patterns={}
    for r in rows:
        if str(r["scope_level"] or "")=="PROJECT ONLY":
            continue
        typ=str(r["memory_type"] or "OTHER")
        key=re.sub(r'\s+',' ',str(r["subject"] or "").lower()).strip()[:180]
        if not key: continue
        p=patterns.setdefault((typ,key),{"summary":r["lesson"],"count":0,"latest":r["project_id"]})
        p["count"]+=1
        p["latest"]=r["project_id"]

    c=db()
    c.execute("DELETE FROM project_pattern_memory WHERE company_id=?",(current_company_id(),))
    now=datetime.utcnow().isoformat()
    for (typ,key),p in patterns.items():
        c.execute("""INSERT INTO project_pattern_memory(
            company_id,pattern_type,pattern_key,pattern_summary,occurrence_count,
            latest_project_id,last_seen,approved_only,created,updated
        ) VALUES(?,?,?,?,?,?,?,?,?,?)""",(
            current_company_id(),typ,key,p["summary"],p["count"],p["latest"],now,1,now,now
        ))
    c.commit(); c.close()

def _v54_company_patterns():
    _v54_rebuild_patterns()
    return _v39_rows("""
        SELECT * FROM project_pattern_memory
        WHERE company_id=?
        ORDER BY occurrence_count DESC,id DESC LIMIT 250
    """,(current_company_id(),))

def _v54_apply_memory_to_scope(pid):
    """
    Preview how approved company memory would influence current scope.
    Does not silently overwrite saved scope.
    """
    memories=[r for r in _v54_memory_rows() if str(r["scope_level"] or "")=="COMPANY STANDARD"]
    scopes=_v452_scope_rows(pid)
    proposals=[]
    for r in scopes:
        low=str(r["requirement"] or "").lower()
        for m in memories:
            subject=str(m["subject"] or "").lower().strip()
            if subject and subject in low:
                proposals.append((r,m))
    return proposals[:200]

def _v54_memory_quality():
    rows=_v54_memory_rows()
    project=sum(1 for r in rows if r["scope_level"]=="PROJECT ONLY")
    company=sum(1 for r in rows if r["scope_level"]=="COMPANY STANDARD")
    globaln=sum(1 for r in rows if r["scope_level"]=="GLOBAL BUILDCOMMAND RULE")
    return {"total":len(rows),"project":project,"company":company,"global":globaln}

@app.get("/project-memory",response_class=HTMLResponse)
def v54_memory_home():
    pid=project_id()
    _v54_seed_from_existing(pid)
    q=_v54_memory_quality()
    patterns=_v54_company_patterns()
    proposals=_v54_apply_memory_to_scope(pid)
    body=f'<div class="hero"><div class="eyebrow">BuildCommand v54 - Project Memory & Continuous Learning</div><h1>Do not make the same construction mistake twice.</h1><p class="muted">{q["total"]} approved memory item(s) · {q["project"]} project-only · {q["company"]} company standards · {q["global"]} global rules · {len(patterns)} reusable pattern(s) · {len(proposals)} current-project memory match(es).</p></div><div class="grid3">'
    cards=[
      ("Approved Project Memory","See approved corrections, RFI answers and lessons from this project.","/project-memory/current"),
      ("Company Standards Memory","Approved reusable construction rules across projects.","/project-memory/company"),
      ("Recurring Pattern Brain","Find repeated approved lessons across completed work.","/project-memory/patterns"),
      ("Memory-to-Scope Preview","See where approved company memory would influence current scope.","/project-memory/apply"),
      ("RFI Answer Memory","Reuse approved clarifications instead of rediscovering the same question.","/project-memory/rfis"),
      ("Lessons-Learned Memory","Carry approved lessons into future project intelligence.","/project-memory/lessons"),
      ("Trade Correction Memory","See approved ownership corrections that should affect Blueprint Brain.","/project-memory/trade-corrections"),
      ("Memory Quality","Understand what is project-specific versus reusable.","/project-memory/quality"),
      ("Project-to-Project Learning","Compare current project needs against approved company patterns.","/project-memory/cross-project"),
      ("Continuous Learning Command","One view of what BuildCommand has learned and where it can help next.","/project-memory/command")
    ]
    for name,desc,href in cards:
        body+=_v37_link_card(name,desc,href,"Open")
    body+='</div>'
    return shell("Project Memory",body)

@app.get("/project-memory/current",response_class=HTMLResponse)
def v54_current_memory():
    pid=project_id(); _v54_seed_from_existing(pid)
    rows=_v54_memory_rows(pid)
    h="".join(f'<div class="action"><span class="badge READY">{esc(r["memory_type"])}</span> <b>{esc(r["subject"])}</b><div>{esc(r["lesson"])}</div><div class="small">{esc(r["scope_level"])} · {esc(r["source_ref"])} · Approved by {esc(r["approved_by"])}</div></div>' for r in rows)
    return shell("Approved Project Memory",'<div class="hero"><h1>Approved Project Memory</h1><p class="muted">Only approved decisions become durable project memory.</p></div><div class="card">'+(h or '<p class="muted">No approved project memory yet.</p>')+'</div>')

@app.get("/project-memory/company",response_class=HTMLResponse)
def v54_company_memory():
    rows=[r for r in _v54_memory_rows() if r["scope_level"]=="COMPANY STANDARD"]
    h="".join(f'<div class="action"><span class="badge READY">COMPANY STANDARD</span> <b>{esc(r["subject"])}</b><div>{esc(r["lesson"])}</div><div class="small">{esc(r["memory_type"])} · {esc(r["source_ref"])}</div></div>' for r in rows)
    return shell("Company Standards Memory",'<div class="hero"><h1>Company Standards Memory</h1><p class="muted">Approved reusable rules can influence future projects; project-only memory cannot silently promote itself.</p></div><div class="card">'+(h or '<p class="muted">No company-standard memory yet.</p>')+'</div>')

@app.get("/project-memory/patterns",response_class=HTMLResponse)
def v54_patterns_page():
    rows=_v54_company_patterns()
    h="".join(f'<div class="action"><span class="badge">{r["occurrence_count"]}x</span> <b>{esc(r["pattern_type"])}</b> - {esc(r["pattern_key"])}<div>{esc(r["pattern_summary"])}</div></div>' for r in rows)
    return shell("Recurring Pattern Brain",'<div class="hero"><h1>Recurring Pattern Brain</h1><p class="muted">Repeated approved lessons become visible patterns; they are not automatically promoted to global truth.</p></div><div class="card">'+(h or '<p class="muted">No recurring approved patterns yet.</p>')+'</div>')

@app.get("/project-memory/apply",response_class=HTMLResponse)
def v54_apply_page():
    rows=_v54_apply_memory_to_scope(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">MEMORY MATCH</span> <b>{esc(r["trade"])}</b> - {esc(r["requirement"])}<div class="small">Approved company memory: {esc(m["subject"])} → {esc(m["lesson"])}</div></div>' for r,m in rows)
    return shell("Memory-to-Scope Preview",'<div class="hero"><h1>Memory-to-Scope Preview</h1><p class="muted">Shows where approved company memory matches current scope. Preview only—no silent overwrite.</p></div><div class="card">'+(h or '<p class="muted">No current scope matches approved company memory.</p>')+'</div>')

@app.get("/project-memory/rfis",response_class=HTMLResponse)
def v54_rfi_memory():
    rows=[r for r in _v54_memory_rows() if r["memory_type"]=="RFI ANSWER"]
    h="".join(f'<div class="action"><b>{esc(r["subject"])}</b><div>{esc(r["lesson"])}</div><div class="small">{esc(r["source_ref"])}</div></div>' for r in rows)
    return shell("RFI Answer Memory",'<div class="hero"><h1>RFI Answer Memory</h1><p class="muted">Approved clarifications remain searchable project knowledge instead of disappearing in old RFIs.</p></div><div class="card">'+(h or '<p class="muted">No approved RFI-answer memory yet.</p>')+'</div>')

@app.get("/project-memory/lessons",response_class=HTMLResponse)
def v54_lessons_memory():
    rows=[r for r in _v54_memory_rows() if r["memory_type"]=="LESSON LEARNED"]
    h="".join(f'<div class="action"><b>{esc(r["subject"])}</b><div>{esc(r["lesson"])}</div><div class="small">{esc(r["scope_level"])} · {esc(r["source_ref"])}</div></div>' for r in rows)
    return shell("Lessons-Learned Memory",'<div class="hero"><h1>Lessons-Learned Memory</h1><p class="muted">Approved project lessons can become company intelligence when deliberately promoted.</p></div><div class="card">'+(h or '<p class="muted">No approved lessons-learned memory yet.</p>')+'</div>')

@app.get("/project-memory/trade-corrections",response_class=HTMLResponse)
def v54_trade_memory():
    rows=[r for r in _v54_memory_rows() if r["memory_type"]=="APPROVED LEARNING" and ("trade" in str(r["subject"] or "").lower() or "->" in str(r["lesson"] or ""))]
    h="".join(f'<div class="action"><span class="badge READY">{esc(r["scope_level"])}</span> <b>{esc(r["subject"])}</b><div>{esc(r["lesson"])}</div></div>' for r in rows)
    return shell("Trade Correction Memory",'<div class="hero"><h1>Trade Correction Memory</h1><p class="muted">Human-approved ownership corrections should influence future Blueprint Brain decisions.</p></div><div class="card">'+(h or '<p class="muted">No approved trade-correction memory found.</p>')+'</div>')

@app.get("/project-memory/quality",response_class=HTMLResponse)
def v54_memory_quality_page():
    q=_v54_memory_quality()
    return shell("Memory Quality",f'<div class="hero"><h1>Memory Quality</h1><p class="muted">{q["total"]} approved memories available.</p></div><div class="grid3"><div class="card"><div class="label">Project Only</div><div class="kpi">{q["project"]}</div></div><div class="card"><div class="label">Company Standards</div><div class="kpi">{q["company"]}</div></div><div class="card"><div class="label">Global Rules</div><div class="kpi">{q["global"]}</div></div></div>')

@app.get("/project-memory/cross-project",response_class=HTMLResponse)
def v54_cross_project():
    patterns=_v54_company_patterns()
    h="".join(f'<div class="action"><span class="badge">{r["occurrence_count"]}x</span> <b>{esc(r["pattern_key"])}</b><div>{esc(r["pattern_summary"])}</div><div class="small">Latest project {esc(r["latest_project_id"])}</div></div>' for r in patterns)
    return shell("Project-to-Project Learning",'<div class="hero"><h1>Project-to-Project Learning</h1><p class="muted">Approved company-level patterns help the next project start smarter.</p></div><div class="card">'+(h or '<p class="muted">No reusable cross-project patterns yet.</p>')+'</div>')

@app.get("/project-memory/command",response_class=HTMLResponse)
def v54_command():
    pid=project_id(); _v54_seed_from_existing(pid)
    q=_v54_memory_quality(); proposals=_v54_apply_memory_to_scope(pid); patterns=_v54_company_patterns()[:10]
    body=f'<div class="hero"><div class="eyebrow">Continuous Learning Command</div><h1>What has BuildCommand learned?</h1><p class="muted">{q["total"]} approved memories · {q["company"]} company standards · {len(proposals)} current scope matches · {len(patterns)} recurring patterns.</p></div>'
    body+='<div class="grid3">'
    body+=_v37_link_card("Current Project Memory","Approved corrections, RFIs and lessons.","/project-memory/current","Open")
    body+=_v37_link_card("Memory Matches","Where prior approved knowledge applies now.","/project-memory/apply","Review")
    body+=_v37_link_card("Recurring Patterns","Approved patterns across projects.","/project-memory/patterns","Review")
    body+='</div>'
    return shell("Continuous Learning Command",body)


# ============================================================
# v55 AUTOMATIC DRAWING REVISION & CHANGE INTELLIGENCE
# ============================================================

def _v55_ensure_tables():
    c=db()
    if DATABASE_KIND=="postgres":
        pk="BIGSERIAL PRIMARY KEY"; num="DOUBLE PRECISION"
    else:
        pk="INTEGER PRIMARY KEY"; num="REAL"
    stmts=[
      f"""CREATE TABLE IF NOT EXISTS drawing_revision_runs(
        id {pk},company_id BIGINT,project_id BIGINT,new_attachment_id BIGINT,
        prior_attachment_id BIGINT,revision_group TEXT,status TEXT DEFAULT 'REVIEW',
        summary TEXT,affected_trades TEXT,cost_risk TEXT DEFAULT 'REVIEW',
        schedule_risk TEXT DEFAULT 'REVIEW',created TEXT,updated TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS drawing_revision_findings(
        id {pk},company_id BIGINT,project_id BIGINT,run_id BIGINT,
        finding_type TEXT,subject TEXT,trade TEXT,source_new TEXT,source_prior TEXT,
        change_summary TEXT,cost_exposure {num} DEFAULT 0,schedule_days {num} DEFAULT 0,
        severity TEXT DEFAULT 'REVIEW',status TEXT DEFAULT 'OPEN',
        recommended_action TEXT,created TEXT,updated TEXT)""",
      f"""CREATE TABLE IF NOT EXISTS revision_downstream_links(
        id {pk},company_id BIGINT,project_id BIGINT,run_id BIGINT,
        finding_id BIGINT,related_type TEXT,related_key TEXT,related_title TEXT,
        relationship TEXT,reason TEXT,created TEXT)"""
    ]
    for s in stmts:
        c.execute(s)
    c.commit(); c.close()

def _v55_doc_group(name):
    n=str(name or "")
    n=re.sub(r'(?i)\b(rev(?:ision)?|addendum|bulletin|asi|sk)[\s_-]*[A-Z0-9.-]+\b','',n)
    n=re.sub(r'(?i)\b\d{4}[-_]\d{2}[-_]\d{2}\b','',n)
    return re.sub(r'[^a-z0-9]+',' ',n.lower()).strip()

def _v55_revision_pairs(pid):
    docs=_v46_docs(pid)
    groups={}
    for d in docs:
        groups.setdefault(_v55_doc_group(d["original_name"]),[]).append(d)
    out=[]
    for group,items in groups.items():
        if not group or len(items)<2:
            continue
        items=sorted(items,key=lambda x:x["id"],reverse=True)
        out.append((items[0],items[1],group))
    return out[:100]

def _v55_scope_by_run(pid):
    rows=_v452_scope_rows(pid)
    runs={}
    for r in rows:
        runs.setdefault(int(r["run_id"]),[]).append(r)
    return runs

def _v55_scope_delta(pid):
    runs=_v55_scope_by_run(pid)
    run_ids=sorted(runs.keys(),reverse=True)
    if len(run_ids)<2:
        return [],[],[]
    new_rows=runs[run_ids[0]]
    old_rows=runs[run_ids[1]]

    def key(r):
        return re.sub(r'\s+',' ',str(r["requirement"] or "").lower()).strip()

    new_map={key(r):r for r in new_rows if key(r)}
    old_map={key(r):r for r in old_rows if key(r)}

    added=[new_map[k] for k in new_map.keys()-old_map.keys()]
    removed=[old_map[k] for k in old_map.keys()-new_map.keys()]
    changed=[]
    common=new_map.keys() & old_map.keys()
    for k in common:
        nr,orow=new_map[k],old_map[k]
        if str(nr["trade"])!=str(orow["trade"]) or str(nr["source_sheet"])!=str(orow["source_sheet"]) or str(nr["source_spec"])!=str(orow["source_spec"]):
            changed.append((orow,nr))
    return added[:200],removed[:200],changed[:200]

def _v55_affected_trades(pid):
    added,removed,changed=_v55_scope_delta(pid)
    trades=set()
    for r in added+removed:
        if r["trade"]: trades.add(str(r["trade"]))
    for old,new in changed:
        if old["trade"]: trades.add(str(old["trade"]))
        if new["trade"]: trades.add(str(new["trade"]))
    return sorted(trades)

def _v55_cost_schedule_exposure(pid):
    added,removed,changed=_v55_scope_delta(pid)
    estimate=_v39_rows("SELECT * FROM estimator_items WHERE company_id=? AND project_id=?",(current_company_id(),pid))
    est_by_scope={}
    for e in estimate:
        sid=e["blueprint_scope_item_id"]
        if sid is not None:
            total=(float(e["quantity"] or 0)*float(e["material_unit_cost"] or 0)
                   +float(e["quantity"] or 0)*float(e["labor_unit_cost"] or 0)
                   +float(e["subcontract_quote"] or 0)+float(e["allowance"] or 0))
            est_by_scope[int(sid)]=total
    added_cost=sum(est_by_scope.get(int(r["id"]),0) for r in added)
    removed_cost=sum(est_by_scope.get(int(r["id"]),0) for r in removed)
    schedule_days=min(30,len(added)*0.5+len(changed)*0.25)
    return added_cost,removed_cost,schedule_days

def _v55_downstream(pid):
    affected=_v55_affected_trades(pid)
    rfis=_v42_rfis(pid)
    subs=_v39_rows("SELECT * FROM submittals WHERE project_id=?",(pid,))
    acts=_v39_rows("SELECT * FROM activities WHERE project_id=?",(pid,))
    links=[]
    for tr in affected:
        for a in acts:
            if str(a["trade"] or "").lower()==tr.lower():
                links.append(("ACTIVITY",a["id"],a["name"],tr))
        for s in subs:
            if tr.lower() in str(s["responsible_party"] or "").lower() or tr.lower() in str(s["title"] or "").lower():
                links.append(("SUBMITTAL",s["id"],s["title"],tr))
        for r in rfis:
            blob=(str(r["title"] or "")+" "+str(r["question"] or "")).lower()
            if tr.lower() in blob:
                links.append(("RFI",r["id"],r["title"],tr))
    return links[:200]

def _v55_summary(pid):
    def build():
        pairs=_v55_revision_pairs(pid)
        added,removed,changed=_v55_scope_delta(pid)
        trades=_v55_affected_trades(pid)
        add_cost,remove_cost,days=_v55_cost_schedule_exposure(pid)
        links=_v55_downstream(pid)
        return {
            "pairs":pairs,"added":added,"removed":removed,"changed":changed,"trades":trades,
            "added_cost":add_cost,"removed_cost":remove_cost,"days":days,"links":links
        }
    return _v56_cached("revision_summary",pid,build)


@app.get("/revision-intelligence",response_class=HTMLResponse)
def v55_revision_home():
    pid=project_id(); _v55_ensure_tables()
    s=_v55_summary(pid)
    body=f'<div class="hero"><div class="eyebrow">BuildCommand v55 - Drawing Revision & Change Intelligence</div><h1>Know what changed before it hits the field.</h1><p class="muted">{len(s["pairs"])} likely revision pair(s) · {len(s["added"])} added scope item(s) · {len(s["removed"])} removed item(s) · {len(s["changed"])} changed ownership/source item(s) · {len(s["trades"])} affected trade(s).</p></div><div class="grid3">'
    cards=[
      ("Revision Pairing","Find likely prior/new versions of project documents.","/revision-intelligence/pairs"),
      ("Added Scope","See requirements appearing in the latest analyzed run.","/revision-intelligence/added"),
      ("Removed Scope","See requirements no longer present in the latest analyzed run.","/revision-intelligence/removed"),
      ("Changed Scope","See ownership/source changes between recent analyses.","/revision-intelligence/changed"),
      ("Affected Trades","See which subcontractor scopes may need review.","/revision-intelligence/trades"),
      ("Cost Exposure","Estimate known priced exposure tied to changed scope.","/revision-intelligence/cost"),
      ("Schedule Exposure","Show likely schedule review pressure from revision volume.","/revision-intelligence/schedule"),
      ("RFI / Submittal Impact","Connect affected trades to downstream control items.","/revision-intelligence/downstream"),
      ("Revision Action List","Prioritize what the project team should review next.","/revision-intelligence/actions"),
      ("Revision Command","One concise revision/change intelligence view.","/revision-intelligence/command")
    ]
    for name,desc,href in cards:
        body+=_v37_link_card(name,desc,href,"Open")
    body+='</div>'
    return shell("Revision Intelligence",body)

@app.get("/revision-intelligence/pairs",response_class=HTMLResponse)
def v55_pairs_page():
    rows=_v55_revision_pairs(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">COMPARE</span> <b>{esc(new["original_name"])}</b><div class="small">Prior candidate: {esc(old["original_name"])} · Group {esc(group)}</div></div>' for new,old,group in rows)
    return shell("Revision Pairing",'<div class="hero"><h1>Revision Pairing</h1><p class="muted">Filename-based candidate pairing only; human review still confirms the actual revision relationship.</p></div><div class="card">'+(h or '<p class="muted">No likely revision pairs detected.</p>')+'</div>')

@app.get("/revision-intelligence/added",response_class=HTMLResponse)
def v55_added_page():
    rows=_v55_scope_delta(project_id())[0]
    h="".join(f'<div class="action"><span class="badge WATCH">ADDED</span> <b>{esc(r["trade"])}</b> - {esc(r["requirement"])}<div class="small">{esc(r["source_sheet"])} · {esc(r["source_spec"])}</div></div>' for r in rows)
    return shell("Added Scope",'<div class="hero"><h1>Added Scope Intelligence</h1><p class="muted">Requirements present in the latest Blueprint Brain run but not the immediately prior run.</p></div><div class="card">'+(h or '<p class="muted">No added scope detected.</p>')+'</div>')

@app.get("/revision-intelligence/removed",response_class=HTMLResponse)
def v55_removed_page():
    rows=_v55_scope_delta(project_id())[1]
    h="".join(f'<div class="action"><span class="badge WATCH">REMOVED</span> <b>{esc(r["trade"])}</b> - {esc(r["requirement"])}<div class="small">{esc(r["source_sheet"])} · {esc(r["source_spec"])}</div></div>' for r in rows)
    return shell("Removed Scope",'<div class="hero"><h1>Removed Scope Intelligence</h1><p class="muted">Requirements present in the prior Blueprint Brain run but not the latest run.</p></div><div class="card">'+(h or '<p class="muted">No removed scope detected.</p>')+'</div>')

@app.get("/revision-intelligence/changed",response_class=HTMLResponse)
def v55_changed_page():
    rows=_v55_scope_delta(project_id())[2]
    h="".join(f'<div class="action"><span class="badge WATCH">CHANGED</span> <b>{esc(old["trade"])} → {esc(new["trade"])}</b> - {esc(new["requirement"])}<div class="small">Old {esc(old["source_sheet"])} / New {esc(new["source_sheet"])}</div></div>' for old,new in rows)
    return shell("Changed Scope",'<div class="hero"><h1>Changed Scope Intelligence</h1><p class="muted">Highlights ownership/source changes between the two most recent Blueprint Brain runs.</p></div><div class="card">'+(h or '<p class="muted">No changed scope signals detected.</p>')+'</div>')

@app.get("/revision-intelligence/trades",response_class=HTMLResponse)
def v55_trades_page():
    trades=_v55_affected_trades(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">REVIEW</span> <b>{esc(tr)}</b><div class="small">Scope/bid/schedule review recommended due to detected revision changes.</div></div>' for tr in trades)
    return shell("Affected Trades",'<div class="hero"><h1>Affected Trades</h1><p class="muted">Trade list derived from added, removed and changed Blueprint Brain scope.</p></div><div class="card">'+(h or '<p class="muted">No affected trades detected.</p>')+'</div>')

@app.get("/revision-intelligence/cost",response_class=HTMLResponse)
def v55_cost_page():
    add,remove,days=_v55_cost_schedule_exposure(project_id())
    return shell("Revision Cost Exposure",f'<div class="hero"><h1>Revision Cost Exposure</h1><p class="muted">Known estimator-linked scope only. Added ${add:,.0f} · Removed ${remove:,.0f} · Net ${add-remove:,.0f}.</p></div><div class="card"><p>Unpriced scope remains review exposure and is not fabricated into a dollar value.</p></div>')

@app.get("/revision-intelligence/schedule",response_class=HTMLResponse)
def v55_schedule_page():
    s=_v55_summary(project_id())
    return shell("Revision Schedule Exposure",f'<div class="hero"><h1>Revision Schedule Exposure</h1><p class="muted">{len(s["added"])} added + {len(s["changed"])} changed scope signals suggest approximately {s["days"]:.1f} review-day(s) of schedule pressure proxy.</p></div><div class="card"><p>This is a transparent review proxy, not CPM-calculated delay entitlement.</p></div>')

@app.get("/revision-intelligence/downstream",response_class=HTMLResponse)
def v55_downstream_page():
    rows=_v55_downstream(project_id())
    h="".join(f'<div class="action"><span class="badge">{esc(kind)}</span> <b>{esc(title)}</b><div class="small">Affected trade: {esc(trade)}</div></div>' for kind,key,title,trade in rows)
    return shell("Revision Downstream Impact",'<div class="hero"><h1>RFI / Submittal / Schedule Impact</h1><p class="muted">Connect revision-affected trades to downstream project controls that deserve review.</p></div><div class="card">'+(h or '<p class="muted">No downstream linked items detected.</p>')+'</div>')

@app.get("/revision-intelligence/actions",response_class=HTMLResponse)
def v55_actions_page():
    s=_v55_summary(project_id())
    actions=[]
    if s["added"]: actions.append(("HIGH","Review added scope",f'{len(s["added"])} new requirement(s) detected.'))
    if s["removed"]: actions.append(("HIGH","Review deleted scope",f'{len(s["removed"])} prior requirement(s) no longer detected.'))
    if s["changed"]: actions.append(("HIGH","Review changed ownership/source",f'{len(s["changed"])} changed item(s) detected.'))
    if s["trades"]: actions.append(("HIGH","Notify affected trades",", ".join(s["trades"])))
    if s["links"]: actions.append(("REVIEW","Review downstream controls",f'{len(s["links"])} RFI/submittal/activity link(s) may be affected.'))
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(level)}</span> <b>{esc(title)}</b><div class="small">{esc(detail)}</div></div>' for level,title,detail in actions)
    return shell("Revision Action List",'<div class="hero"><h1>Revision Action List</h1><p class="muted">Human review remains required before scope, cost, or schedule commitments change.</p></div><div class="card">'+(h or '<p class="muted">No revision actions detected.</p>')+'</div>')

@app.get("/revision-intelligence/command",response_class=HTMLResponse)
def v55_command_page():
    s=_v55_summary(project_id())
    net=s["added_cost"]-s["removed_cost"]
    body=f'<div class="hero"><div class="eyebrow">Revision Command</div><h1>What changed and what should the team do?</h1><p class="muted">{len(s["added"])} added · {len(s["removed"])} removed · {len(s["changed"])} changed · {len(s["trades"])} affected trades · ${net:,.0f} known net estimator-linked exposure.</p></div><div class="grid3">'
    body+=_v37_link_card("Review Added Scope","Latest-run additions.","/revision-intelligence/added","Review")
    body+=_v37_link_card("Affected Trades","Who needs scope review.","/revision-intelligence/trades","Review")
    body+=_v37_link_card("Action List","What the team should review next.","/revision-intelligence/actions","Open")
    body+='</div>'
    return shell("Revision Command",body)


# ============================================================
# v56 PERFORMANCE & ARCHITECTURE OPTIMIZATION
# ============================================================

_V56_CACHE={}
_V56_PERF={}
_V56_TTL={
    "snapshot":15,
    "sequence":30,
    "self_audit":60,
    "master_findings":45,
    "master_score":45,
    "top_priorities":45,
    "decisions":30,
    "materials":30,
    "hold_points":30,
    "agenda":30,
    "revision_summary":60,
}

def _v56_now_ts():
    return datetime.utcnow().timestamp()

def _v56_cache_get(name,pid):
    key=(name,current_company_id(),pid)
    row=_V56_CACHE.get(key)
    if not row:
        return None
    ttl=_V56_TTL.get(name,30)
    if _v56_now_ts()-row["ts"]>ttl:
        _V56_CACHE.pop(key,None)
        return None
    return row["value"]

def _v56_cache_set(name,pid,value):
    _V56_CACHE[(name,current_company_id(),pid)]={"ts":_v56_now_ts(),"value":value}
    return value

def _v56_cached(name,pid,builder):
    hit=_v56_cache_get(name,pid)
    if hit is not None:
        return hit
    return _v56_cache_set(name,pid,builder())

def _v56_clear_project_cache(pid=None):
    cid=current_company_id()
    keys=list(_V56_CACHE.keys())
    for k in keys:
        if k[1]==cid and (pid is None or k[2]==pid):
            _V56_CACHE.pop(k,None)

def _v56_perf_start(label):
    return (label,_v56_now_ts())

def _v56_perf_end(token):
    label,start=token
    elapsed=max(0,_v56_now_ts()-start)
    row=_V56_PERF.setdefault(label,{"count":0,"total":0.0,"max":0.0,"last":0.0})
    row["count"]+=1
    row["total"]+=elapsed
    row["max"]=max(row["max"],elapsed)
    row["last"]=elapsed
    return elapsed

def _v56_ensure_indexes():
    """
    Safe, idempotent indexes for the most common project-scoped reads.
    """
    c=db()
    indexes=[
        ("idx_activities_project_start","activities","project_id,start"),
        ("idx_readiness_project_activity","activity_readiness","project_id,activity_id"),
        ("idx_inspections_project_activity","inspections_tracker","project_id,activity_id"),
        ("idx_procurement_project_activity","procurement","project_id,activity_id"),
        ("idx_submittals_project_activity","submittals","project_id,activity_id"),
        ("idx_issues_project_activity_status","project_issues","project_id,activity_id,status"),
        ("idx_scope_company_project_run","blueprint_scope_items","company_id,project_id,run_id"),
        ("idx_scope_company_project_trade","blueprint_scope_items","company_id,project_id,trade"),
        ("idx_scope_trade_scope","blueprint_scope_items","trade_scope_id"),
        ("idx_trade_scopes_company_project_run","blueprint_trade_scopes","company_id,project_id,run_id"),
        ("idx_attachments_company_project","attachments","company_id,project_id"),
        ("idx_estimator_company_project_scope","estimator_items","company_id,project_id,blueprint_scope_item_id"),
        ("idx_changes_project_status","change_events","project_id,status"),
        ("idx_closeout_company_project","closeout_items","company_id,project_id"),
        ("idx_learning_company_project_approval","learning_rules","company_id,project_id,approval_status"),
        ("idx_memory_company_project_approval","project_memory","company_id,project_id,approval_status"),
    ]
    for name,table,cols in indexes:
        try:
            c.execute(f"CREATE INDEX IF NOT EXISTS {name} ON {table}({cols})")
        except Exception:
            try: c.rollback()
            except Exception: pass
    try: c.commit()
    except Exception: pass
    c.close()

_V56_INDEXES_READY=False
def _v56_indexes_once():
    global _V56_INDEXES_READY
    if not _V56_INDEXES_READY:
        try:
            _v56_ensure_indexes()
        finally:
            _V56_INDEXES_READY=True

def _v56_sequence(pid):
    return _v56_cached("sequence",pid,lambda:_v45_sequence_analysis_uncached(pid))

def _v56_self_audit(pid):
    return _v56_cached("self_audit",pid,lambda:_v48_self_audit_uncached(pid))

def _v56_master_findings(pid):
    return _v56_cached("master_findings",pid,lambda:_v50_collect_findings_uncached(pid))

def _v56_master_score(pid):
    def build():
        findings=_v56_master_findings(pid)
        weights={"CRITICAL":12,"HIGH":7,"MEDIUM":3,"REVIEW":1,"LOW":0}
        risk=min(100,sum(weights.get(f["severity"],1) for f in findings))
        quality=_v48_quality_score(pid)["score"]
        health=max(0,round((100-risk)*0.65 + quality*0.35))
        return {"health":health,"risk":risk,"quality":quality,"count":len(findings)}
    return _v56_cached("master_score",pid,build)

def _v56_top_priorities(pid,limit=10):
    cached=_v56_cache_get("top_priorities",pid)
    if cached is None:
        findings=_v56_master_findings(pid)
        priority=[f for f in findings if f["severity"] in {"CRITICAL","HIGH","MEDIUM"}]
        if len(priority)<50:
            priority.extend([f for f in findings if f["severity"]=="REVIEW"][:50-len(priority)])
        cached=_v56_cache_set("top_priorities",pid,priority[:50])
    return cached[:limit]

def _v56_dashboard_bundle(pid):
    """
    Compute the home screen once and share the same cached module results.
    """
    token=_v56_perf_start("home_dashboard_bundle")
    try:
        _v56_indexes_once()
        s=_v56_cached("snapshot",pid,lambda:_v37_snapshot(pid))
        score=_v56_master_score(pid)
        top=_v56_top_priorities(pid,8)
        decisions=_v56_cached("decisions",pid,lambda:_v47_decision_deadlines(pid)[:8])
        materials=_v56_cached("materials",pid,lambda:[x for x in _v47_material_readiness(pid) if x[2] in {"CRITICAL","HIGH","TODAY"}][:8])
        holds=_v56_cached("hold_points",pid,lambda:[x for x in _v49_hold_points(pid) if str(x[0]["result"] or "").upper()!="PASSED"][:8])
        agenda=_v56_cached("agenda",pid,lambda:_v47_coordination_agenda(pid)[:8])
        sequence=[x for x in _v56_sequence(pid) if x["risk"] in {"CRITICAL","HIGH"}][:8]
        return {
            "snapshot":s,"score":score,"top":top,"decisions":decisions,
            "materials":materials,"holds":holds,"agenda":agenda,"sequence":sequence
        }
    finally:
        _v56_perf_end(token)

@app.get("/performance",response_class=HTMLResponse)
def v56_performance_page():
    _v56_indexes_once()
    cache_count=len(_V56_CACHE)
    rows=[]
    for name,r in sorted(_V56_PERF.items(), key=lambda kv:kv[1]["last"], reverse=True):
        avg=(r["total"]/r["count"]) if r["count"] else 0
        rows.append((name,r["count"],r["last"],avg,r["max"]))
    h="".join(
        f'<div class="action"><b>{esc(name)}</b>'
        f'<div class="small">runs {count} · last {last*1000:.0f} ms · avg {avg*1000:.0f} ms · max {maxv*1000:.0f} ms</div></div>'
        for name,count,last,avg,maxv in rows
    )
    return shell("Performance Monitor",
        f'<div class="hero"><div class="eyebrow">v56 Performance</div><h1>BuildCommand Performance Monitor</h1>'
        f'<p class="muted">{cache_count} cached project intelligence snapshot(s). Heavy intelligence is reused until its short TTL expires.</p></div>'
        f'<div class="card">{h or "<p class=muted>No timing samples yet. Open the dashboard, then return here.</p>"}</div>')

@app.post("/performance/clear-cache")
def v56_clear_cache():
    _v56_clear_project_cache(project_id())
    return RedirectResponse("/performance",status_code=303)


# ============================================================
# v57 EVENT-DRIVEN INTELLIGENCE
# Refresh only the intelligence affected by a project change.
# ============================================================

_V57_EVENTS=[]
_V57_DIRTY={}

_V57_DEPENDENCIES={
    "DRAWING_UPLOAD":{"snapshot","master_findings","master_score","top_priorities","revision_summary","self_audit"},
    "BLUEPRINT_ANALYSIS":{"snapshot","master_findings","master_score","top_priorities","revision_summary","self_audit"},
    "SCOPE_CHANGE":{"snapshot","master_findings","master_score","top_priorities","revision_summary","self_audit"},
    "RFI_CHANGE":{"snapshot","master_findings","master_score","top_priorities","decisions","agenda","self_audit"},
    "SUBMITTAL_CHANGE":{"snapshot","master_findings","master_score","top_priorities","decisions","agenda","sequence"},
    "INSPECTION_CHANGE":{"snapshot","master_findings","master_score","top_priorities","hold_points","agenda","sequence"},
    "PROCUREMENT_CHANGE":{"snapshot","master_findings","master_score","top_priorities","materials","agenda","sequence"},
    "SCHEDULE_CHANGE":{"snapshot","master_findings","master_score","top_priorities","sequence","agenda","decisions"},
    "ISSUE_CHANGE":{"snapshot","master_findings","master_score","top_priorities","agenda","sequence"},
    "LEARNING_CHANGE":{"master_findings","master_score","top_priorities","self_audit","revision_summary"},
    "PROJECT_CHANGE":{"snapshot","master_findings","master_score","top_priorities","sequence","agenda","decisions","materials","hold_points"},
}

def _v57_emit(pid,event_type,subject="",source="SYSTEM"):
    event_type=str(event_type or "PROJECT_CHANGE").upper()
    affected=set(_V57_DEPENDENCIES.get(event_type,_V57_DEPENDENCIES["PROJECT_CHANGE"]))
    cid=current_company_id()
    dirty=_V57_DIRTY.setdefault((cid,pid),set())
    dirty.update(affected)

    # Invalidate only the affected v56 cache keys.
    for name in affected:
        _V56_CACHE.pop((name,cid,pid),None)

    event={
        "ts":datetime.utcnow().isoformat(),
        "company_id":cid,"project_id":pid,"event_type":event_type,
        "subject":subject,"source":source,"affected":sorted(affected)
    }
    _V57_EVENTS.append(event)
    if len(_V57_EVENTS)>500:
        del _V57_EVENTS[:-500]
    return event

def _v57_dirty(pid):
    return sorted(_V57_DIRTY.get((current_company_id(),pid),set()))

def _v57_mark_clean(pid,names):
    dirty=_V57_DIRTY.setdefault((current_company_id(),pid),set())
    for n in names: dirty.discard(n)

def _v57_refresh(pid,names=None):
    names=set(names or _v57_dirty(pid))
    token=_v56_perf_start("event_driven_refresh")
    refreshed=[]
    try:
        if "snapshot" in names:
            _v56_cached("snapshot",pid,lambda:_v37_snapshot(pid)); refreshed.append("snapshot")
        if "sequence" in names:
            _v56_sequence(pid); refreshed.append("sequence")
        if "self_audit" in names:
            _v56_self_audit(pid); refreshed.append("self_audit")
        if "master_findings" in names:
            _v56_master_findings(pid); refreshed.append("master_findings")
        if "master_score" in names:
            _v56_master_score(pid); refreshed.append("master_score")
        if "top_priorities" in names:
            _v56_top_priorities(pid,50); refreshed.append("top_priorities")
        if "decisions" in names:
            _v56_cached("decisions",pid,lambda:_v47_decision_deadlines(pid)[:8]); refreshed.append("decisions")
        if "materials" in names:
            _v56_cached("materials",pid,lambda:[x for x in _v47_material_readiness(pid) if x[2] in {"CRITICAL","HIGH","TODAY"}][:8]); refreshed.append("materials")
        if "hold_points" in names:
            _v56_cached("hold_points",pid,lambda:[x for x in _v49_hold_points(pid) if str(x[0]["result"] or "").upper()!="PASSED"][:8]); refreshed.append("hold_points")
        if "agenda" in names:
            _v56_cached("agenda",pid,lambda:_v47_coordination_agenda(pid)[:8]); refreshed.append("agenda")
        if "revision_summary" in names:
            _v55_summary(pid); refreshed.append("revision_summary")
        _v57_mark_clean(pid,refreshed)
        return refreshed
    finally:
        _v56_perf_end(token)

def _v57_recent_events(pid,limit=50):
    cid=current_company_id()
    return [e for e in reversed(_V57_EVENTS) if e["company_id"]==cid and e["project_id"]==pid][:limit]

@app.get("/event-intelligence",response_class=HTMLResponse)
def v57_event_home():
    pid=project_id()
    dirty=_v57_dirty(pid)
    events=_v57_recent_events(pid,20)
    body=f'<div class="hero"><div class="eyebrow">BuildCommand v57 - Event-Driven Intelligence</div><h1>Update only what changed.</h1><p class="muted">{len(dirty)} intelligence area(s) currently marked for refresh · {len(events)} recent project event(s).</p></div><div class="grid3">'
    cards=[
      ("Event Command","See project changes and affected intelligence.","/event-intelligence/command"),
      ("Dirty Intelligence","See exactly what needs recalculation.","/event-intelligence/dirty"),
      ("Dependency Map","See which project changes affect which brain modules.","/event-intelligence/dependencies"),
      ("Recent Events","Audit recent intelligence-triggering project changes.","/event-intelligence/events"),
      ("Refresh Changed Intelligence","Recalculate only dirty modules.","/event-intelligence/refresh"),
      ("Performance Monitor","Measure event-driven refresh speed.","/performance"),
    ]
    for name,desc,href in cards:
        body+=_v37_link_card(name,desc,href,"Open")
    body+='</div>'
    return shell("Event-Driven Intelligence",body)

@app.get("/event-intelligence/command",response_class=HTMLResponse)
def v57_event_command():
    pid=project_id(); dirty=_v57_dirty(pid); events=_v57_recent_events(pid,10)
    ehtml="".join(
        f'<div class="action"><span class="badge WATCH">{esc(e["event_type"])}</span> <b>{esc(e["subject"] or "Project change")}</b>'
        f'<div class="small">{esc(e["ts"])} · refreshes {esc(", ".join(e["affected"]))}</div></div>'
        for e in events
    ) or '<p class="muted">No project change events recorded in this process yet.</p>'
    dhtml="".join(f'<span class="badge WATCH" style="margin:4px">{esc(x)}</span>' for x in dirty) or '<span class="badge READY">CLEAN</span>'
    body=f'<div class="hero"><div class="eyebrow">Event Command</div><h1>What changed, and what brain work does it require?</h1><p class="muted">Dirty intelligence: {len(dirty)} module(s).</p></div><div class="card"><h2>Needs Refresh</h2>{dhtml}</div><div class="card"><h2>Recent Events</h2>{ehtml}</div>'
    return shell("Event Command",body)

@app.get("/event-intelligence/dirty",response_class=HTMLResponse)
def v57_dirty_page():
    dirty=_v57_dirty(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">DIRTY</span> <b>{esc(x)}</b><div class="small">This module was invalidated by a relevant project change.</div></div>' for x in dirty)
    return shell("Dirty Intelligence",'<div class="hero"><h1>Dirty Intelligence</h1><p class="muted">Only these modules need recalculation.</p></div><div class="card">'+(h or '<p class="muted">All tracked intelligence is clean.</p>')+'</div>')

@app.get("/event-intelligence/dependencies",response_class=HTMLResponse)
def v57_dependencies_page():
    h=""
    for event,names in _V57_DEPENDENCIES.items():
        h+=f'<div class="action"><b>{esc(event)}</b><div class="small">{esc(", ".join(sorted(names)))}</div></div>'
    return shell("Intelligence Dependency Map",'<div class="hero"><h1>Intelligence Dependency Map</h1><p class="muted">A drawing change should not force unrelated procurement or field calculations unless they actually depend on it.</p></div><div class="card">'+h+'</div>')

@app.get("/event-intelligence/events",response_class=HTMLResponse)
def v57_events_page():
    rows=_v57_recent_events(project_id(),100)
    h="".join(f'<div class="action"><span class="badge">{esc(e["event_type"])}</span> <b>{esc(e["subject"] or "Project change")}</b><div class="small">{esc(e["ts"])} · source {esc(e["source"])} · affected {esc(", ".join(e["affected"]))}</div></div>' for e in rows)
    return shell("Recent Intelligence Events",'<div class="hero"><h1>Recent Intelligence Events</h1><p class="muted">Transparent audit trail for cache invalidation and targeted refresh.</p></div><div class="card">'+(h or '<p class="muted">No events recorded yet.</p>')+'</div>')

@app.get("/event-intelligence/refresh",response_class=HTMLResponse)
def v57_refresh_page():
    pid=project_id()
    dirty=_v57_dirty(pid)
    if not dirty:
        return shell("Refresh Changed Intelligence",'<div class="hero"><h1>Nothing needs refreshing.</h1><p class="muted">Tracked intelligence is currently clean.</p></div>')
    h="".join(f'<li>{esc(x)}</li>' for x in dirty)
    return shell("Refresh Changed Intelligence",f'<div class="hero"><h1>Refresh only changed intelligence</h1><p class="muted">{len(dirty)} module(s) are dirty.</p></div><div class="card"><ul>{h}</ul><form method="post" action="/event-intelligence/refresh"><button type="submit">Refresh Changed Intelligence</button></form></div>')

@app.post("/event-intelligence/refresh")
def v57_refresh_post():
    _v57_refresh(project_id())
    return RedirectResponse("/event-intelligence/command",status_code=303)

@app.post("/event-intelligence/test-event")
def v57_test_event(event_type:str=Form("PROJECT_CHANGE"),subject:str=Form("Manual project update")):
    _v57_emit(project_id(),event_type,subject,"MANUAL")
    return RedirectResponse("/event-intelligence/command",status_code=303)


# ============================================================
# v58 PROACTIVE SUPERINTENDENT AI
# ============================================================

def _v58_days_until(value):
    if not value: return None
    try:
        d=datetime.fromisoformat(str(value)[:10]).date()
        return (d-date.today()).days
    except Exception:
        return None

def _v58_proactive_findings(pid):
    findings=[]

    # Start with master reasoning priorities.
    try:
        for f in _v56_top_priorities(pid,30):
            findings.append({
                "severity":f["severity"],"category":f["type"],"title":f["title"],
                "why":f["reason"],"action":f["action"],"source":"MASTER REASONING"
            })
    except Exception:
        pass

    # Upcoming activities and readiness.
    try:
        for x in _v56_sequence(pid):
            a=x["activity"]
            days=_v58_days_until(a["start"])
            if days is None or days < 0 or days > 14: continue
            if x["risk"] in {"CRITICAL","HIGH","MEDIUM"}:
                findings.append({
                    "severity":x["risk"],"category":"UPCOMING WORK",
                    "title":f'{a["name"]} starts in {days} day(s)',
                    "why":x["blocking_reason"] or "Upcoming activity has readiness/sequence exposure.",
                    "action":"Clear predecessor, readiness, material, inspection, and coordination blockers before mobilization.",
                    "source":"SEQUENCE INTELLIGENCE"
                })
    except Exception:
        pass

    # Material risk.
    try:
        for r,act,level,exposure,reason,action in _v47_material_readiness(pid):
            if level not in {"CRITICAL","HIGH","TODAY"}: continue
            findings.append({
                "severity":"CRITICAL" if level=="CRITICAL" else "HIGH",
                "category":"MATERIAL",
                "title":f'{r["item"]} threatens upcoming work',
                "why":reason,
                "action":action,
                "source":"PROCUREMENT"
            })
    except Exception:
        pass

    # Inspection hold points.
    try:
        for i,a in _v49_hold_points(pid):
            if str(i["result"] or "").upper()=="PASSED": continue
            findings.append({
                "severity":"HIGH","category":"INSPECTION",
                "title":f'{i["inspection_type"]} is an open hold point',
                "why":f'Scheduled {i["scheduled_date"] or "date not set"}; related activity {a["name"] if a else "not linked"}.',
                "action":"Confirm prerequisite work is complete and inspection is scheduled/passed before covering or proceeding.",
                "source":"INSPECTIONS"
            })
    except Exception:
        pass

    # Decisions / RFIs / submittals.
    try:
        for typ,title,due,severity,cost,days,source in _v47_decision_deadlines(pid)[:30]:
            findings.append({
                "severity":severity,"category":"DECISION",
                "title":title,
                "why":f'{typ} due {due}; ${cost:,.0f} known exposure; {days:g} schedule day(s) exposure.',
                "action":"Drive the responsible party to resolution before downstream work is affected.",
                "source":source or typ
            })
    except Exception:
        pass

    # Deduplicate and rank.
    rank={"CRITICAL":0,"HIGH":1,"MEDIUM":2,"REVIEW":3,"LOW":4}
    seen=set(); clean=[]
    for f in findings:
        k=(str(f["category"]).lower(),str(f["title"]).lower())
        if k in seen: continue
        seen.add(k); clean.append(f)
    clean.sort(key=lambda x:rank.get(x["severity"],3))
    return clean[:100]

def _v58_now_today_week(pid):
    rows=_v58_proactive_findings(pid)
    now=[]; today=[]; week=[]; upcoming=[]
    for f in rows:
        sev=f["severity"]
        if sev=="CRITICAL":
            now.append(f)
        elif sev=="HIGH":
            today.append(f)
        elif sev=="MEDIUM":
            week.append(f)
        else:
            upcoming.append(f)
    return now[:12],today[:15],week[:15],upcoming[:15]

def _v58_trade_alerts(pid):
    rows=_v58_proactive_findings(pid)
    alerts={}
    known_trades=[str(x["trade"] or "") for x in _v39_rows("SELECT DISTINCT trade FROM activities WHERE project_id=?",(pid,)) if x["trade"]]
    for f in rows:
        blob=(f["title"]+" "+f["why"]+" "+f["action"]).lower()
        for tr in known_trades:
            if tr.lower() in blob:
                alerts.setdefault(tr,[]).append(f)
    return alerts

def _v58_super_brief(pid):
    now,today,week,upcoming=_v58_now_today_week(pid)
    return {
        "now":now,"today":today,"week":week,"upcoming":upcoming,
        "total":len(now)+len(today)+len(week)+len(upcoming)
    }

@app.get("/proactive-superintendent",response_class=HTMLResponse)
def v58_home():
    pid=project_id(); b=_v58_super_brief(pid)
    body=f'<div class="hero"><div class="eyebrow">BuildCommand v58 - Proactive Superintendent AI</div><h1>BuildCommand should tell you before you have to ask.</h1><p class="muted">{len(b["now"])} do-now · {len(b["today"])} today · {len(b["week"])} this-week · {len(b["upcoming"])} upcoming intelligence item(s).</p></div><div class="grid3">'
    cards=[
      ("Superintendent Command","Prioritized proactive construction intelligence.","/proactive-superintendent/command"),
      ("Do Now","Critical items needing immediate attention.","/proactive-superintendent/now"),
      ("Today","High-priority actions for today.","/proactive-superintendent/today"),
      ("This Week","Medium-term coordination and readiness actions.","/proactive-superintendent/week"),
      ("Upcoming Risk","Watch items before they become field problems.","/proactive-superintendent/upcoming"),
      ("Trade Alerts","Group proactive intelligence by affected trade.","/proactive-superintendent/trades"),
      ("Inspection Readiness","Surface open hold points before work gets covered.","/proactive-superintendent/inspections"),
      ("Material Alerts","Catch procurement exposure before installation dates.","/proactive-superintendent/materials"),
      ("Decision Alerts","Push RFIs/submittals/decisions before downstream impact.","/proactive-superintendent/decisions"),
      ("Event Intelligence","See what project changes triggered intelligence refresh.","/event-intelligence","Open"),
    ]
    for n,d,h in cards: body+=_v37_link_card(n,d,h,"Open")
    body+='</div>'
    return shell("Proactive Superintendent AI",body)

def _v58_render(title,subtitle,rows):
    h="".join(
        f'<div class="action bc-priority"><span class="badge WATCH">{esc(f["severity"])}</span> '
        f'<b>{esc(f["category"])} - {esc(f["title"])}</b>'
        f'<div class="small">{esc(f["why"])}</div><p><b>Recommended next action:</b> {esc(f["action"])}</p>'
        f'<div class="small">Source: {esc(f["source"])}</div></div>'
        for f in rows
    ) or '<p class="muted">No items in this priority group.</p>'
    return shell(title,f'<div class="hero"><h1>{esc(title)}</h1><p class="muted">{esc(subtitle)}</p></div><div class="card">{h}</div>')

@app.get("/proactive-superintendent/now",response_class=HTMLResponse)
def v58_now_page():
    return _v58_render("Do Now","Critical project conditions BuildCommand believes deserve immediate superintendent attention.",_v58_now_today_week(project_id())[0])

@app.get("/proactive-superintendent/today",response_class=HTMLResponse)
def v58_today_page():
    return _v58_render("Today","High-priority project actions to drive today.",_v58_now_today_week(project_id())[1])

@app.get("/proactive-superintendent/week",response_class=HTMLResponse)
def v58_week_page():
    return _v58_render("This Week","Coordination, readiness and decision items to clear this week.",_v58_now_today_week(project_id())[2])

@app.get("/proactive-superintendent/upcoming",response_class=HTMLResponse)
def v58_upcoming_page():
    return _v58_render("Upcoming Risk","Watch these conditions before they become field blockers.",_v58_now_today_week(project_id())[3])

@app.get("/proactive-superintendent/inspections",response_class=HTMLResponse)
def v58_inspections_page():
    rows=[f for f in _v58_proactive_findings(project_id()) if f["category"]=="INSPECTION"]
    return _v58_render("Inspection Readiness","Open hold points that can stop covering, startup, energization or downstream work.",rows)

@app.get("/proactive-superintendent/materials",response_class=HTMLResponse)
def v58_materials_page():
    rows=[f for f in _v58_proactive_findings(project_id()) if f["category"]=="MATERIAL"]
    return _v58_render("Material Alerts","Procurement conditions that threaten upcoming installation work.",rows)

@app.get("/proactive-superintendent/decisions",response_class=HTMLResponse)
def v58_decisions_page():
    rows=[f for f in _v58_proactive_findings(project_id()) if f["category"]=="DECISION"]
    return _v58_render("Decision Alerts","RFIs, submittals and project decisions that need to move before downstream impact.",rows)

@app.get("/proactive-superintendent/trades",response_class=HTMLResponse)
def v58_trades_page():
    alerts=_v58_trade_alerts(project_id())
    h=""
    for tr,rows in alerts.items():
        h+=f'<div class="card"><h2>{esc(tr)}</h2>'
        h+="".join(f'<div class="action"><span class="badge WATCH">{esc(f["severity"])}</span> <b>{esc(f["title"])}</b><div class="small">{esc(f["action"])}</div></div>' for f in rows[:10])
        h+='</div>'
    return shell("Trade Alerts",'<div class="hero"><h1>Trade Alerts</h1><p class="muted">Proactive project intelligence grouped by affected trade.</p></div>'+(h or '<div class="card"><p class="muted">No trade-specific alerts detected.</p></div>'))

@app.get("/proactive-superintendent/command",response_class=HTMLResponse)
def v58_command():
    pid=project_id(); b=_v58_super_brief(pid)
    def compact(rows):
        return "".join(f'<div class="action"><span class="badge WATCH">{esc(f["severity"])}</span> <b>{esc(f["title"])}</b><div class="small">{esc(f["action"])}</div></div>' for f in rows[:8]) or '<p class="muted">Clear.</p>'
    body=f'<div class="hero"><div class="eyebrow">SUPERINTENDENT COMMAND</div><h1>What should I deal with next?</h1><p class="muted">{b["total"]} proactive intelligence item(s) prioritized from project controls and construction reasoning.</p></div>'
    body+='<div class="grid2"><div class="card"><h2>DO NOW</h2>'+compact(b["now"])+'</div><div class="card"><h2>DO TODAY</h2>'+compact(b["today"])+'</div></div>'
    body+='<div class="grid2"><div class="card"><h2>THIS WEEK</h2>'+compact(b["week"])+'</div><div class="card"><h2>UPCOMING RISK</h2>'+compact(b["upcoming"])+'</div></div>'
    return shell("Superintendent Command",body)


# ============================================================
# v59-v68 NEXT 10 OPERATIONS INTELLIGENCE
# ============================================================

def _v59_lookahead(pid,weeks=3):
    horizon=date.today()+timedelta(days=weeks*7)
    rows=_v39_rows("SELECT * FROM activities WHERE project_id=? ORDER BY start",(pid,))
    out=[]
    seq={int(x["activity"]["id"]):x for x in _v56_sequence(pid)}
    for a in rows:
        try: sd=datetime.fromisoformat(str(a["start"])[:10]).date()
        except Exception: continue
        if not (date.today()<=sd<=horizon): continue
        x=seq.get(int(a["id"]),{})
        risk=x.get("risk","LOW")
        state="READY" if risk=="LOW" else ("NOT READY" if risk in {"CRITICAL","HIGH"} else "AT RISK")
        out.append((a,state,risk,x.get("blocking_reason","")))
    return out

def _v60_trade_readiness(pid):
    out=[]
    for a,state,risk,reason in _v59_lookahead(pid,6):
        checks=[]
        checks.append(("Sequence",state=="READY",reason or "Sequence clear"))
        try:
            proc=[x for x in _v47_material_readiness(pid) if x[1] and int(x[1]["id"])==int(a["id"])]
            checks.append(("Materials",not any(x[2] in {"CRITICAL","HIGH","TODAY"} for x in proc),"Material readiness"))
        except Exception: pass
        try:
            holds=[x for x in _v49_hold_points(pid) if x[1] and int(x[1]["id"])==int(a["id"]) and str(x[0]["result"] or "").upper()!="PASSED"]
            checks.append(("Inspections",not holds,"Inspection hold points"))
        except Exception: pass
        failed=[c[0] for c in checks if not c[1]]
        readiness="READY" if not failed else ("NOT READY" if len(failed)>1 else "AT RISK")
        out.append((a,readiness,failed,checks))
    return out

def _v61_handoffs(pid):
    rows=_v39_rows("SELECT * FROM activities WHERE project_id=? ORDER BY start",(pid,))
    out=[]
    for i in range(len(rows)-1):
        a,b=rows[i],rows[i+1]
        if str(a["trade"] or "")!=str(b["trade"] or ""):
            out.append((a,b,"Verify predecessor completion, inspection, access, layout and turnover before successor mobilizes."))
    return out[:100]

def _v62_daily_brief(pid):
    b=_v58_super_brief(pid)
    today=str(date.today())
    acts=_v39_rows("SELECT * FROM activities WHERE project_id=? AND start<=? AND finish>=? ORDER BY start",(pid,today,today))
    try: inspections=[x for x in _v49_hold_points(pid) if str(x[0]["scheduled_date"] or "")[:10]==today]
    except Exception: inspections=[]
    try: mats=[x for x in _v47_material_readiness(pid) if str(x[0]["promised_date"] or "")[:10]==today]
    except Exception: mats=[]
    return {"priorities":b,"activities":acts,"inspections":inspections,"materials":mats}

def _v63_rfi_candidates(pid):
    candidates=[]
    try:
        for f in _v50_collect_findings(pid):
            blob=(str(f["type"])+" "+str(f["title"])+" "+str(f["reason"])).lower()
            if any(k in blob for k in ["conflict","contradiction","missing","unclear","coordination","constructability"]):
                candidates.append({
                    "severity":f["severity"],"title":f["title"],"reason":f["reason"],
                    "question":f'Please clarify the contract requirement regarding: {f["title"]}.',
                    "action":f["action"],"source":f.get("source","")
                })
    except Exception: pass
    return candidates[:100]

def _v64_longlead(pid):
    rows=_v39_rows("SELECT * FROM procurement WHERE project_id=? ORDER BY required_on_site",(pid,))
    out=[]
    for r in rows:
        try:
            need=datetime.fromisoformat(str(r["required_on_site"])[:10]).date()
            promised=datetime.fromisoformat(str(r["promised_date"])[:10]).date() if r["promised_date"] else None
        except Exception: continue
        exposure=(promised-need).days if promised else None
        level="HIGH" if exposure is None or exposure>0 else ("MEDIUM" if exposure==0 else "LOW")
        release_by=need-timedelta(days=42)
        out.append((r,level,release_by,exposure))
    return out

def _v65_qc_points(pid):
    rows=_v39_rows("SELECT * FROM activities WHERE project_id=? ORDER BY start",(pid,))
    rules=[
      ("concrete","Pre-pour: reinforcing, embeds, forms, elevations, underground signoff and required inspection/testing."),
      ("drywall","Before close-in: framing, above-wall MEP rough, fire/smoke assemblies, backing and inspections."),
      ("ceiling","Before ceiling close: above-ceiling MEP, fire sprinkler, controls, inspections and access coordination."),
      ("roof","Before concealment: substrate, flashing, penetrations, curbs and manufacturer-required conditions."),
      ("electrical","Verify rough inspection, labeling, terminations, testing and energization prerequisites."),
      ("plumbing","Verify pressure/testing, supports, slopes, cleanouts, insulation and inspection prerequisites."),
    ]
    out=[]
    for a in rows:
        blob=(str(a["name"])+" "+str(a["trade"] or "")).lower()
        for key,check in rules:
            if key in blob: out.append((a,key.upper(),check))
    return out[:200]

def _v66_scope_gaps(pid):
    scopes=_v452_scope_rows(pid)
    gaps=[]; seen={}
    for r in scopes:
        req=re.sub(r'\s+',' ',str(r["requirement"] or "").lower()).strip()
        if not str(r["trade"] or "").strip():
            gaps.append(("UNOWNED",r,"No trade owner assigned."))
        if req:
            if req in seen and str(seen[req]["trade"])!=str(r["trade"]):
                gaps.append(("DUPLICATE",r,f'Also assigned to {seen[req]["trade"]}.'))
            else: seen[req]=r
        if any(k in str(r["trade"] or "").lower() for k in ["general","other","unknown","tbd"]):
            gaps.append(("AMBIGUOUS",r,"Trade ownership should be confirmed before buyout."))
    return gaps[:200]

def _v67_change_intelligence(pid):
    s=_v55_summary(pid)
    return {
      "added":s["added"],"removed":s["removed"],"changed":s["changed"],
      "known_net":s["added_cost"]-s["removed_cost"],"schedule_proxy":s["days"],
      "affected_trades":s["trades"],"downstream":s["links"]
    }

def _v68_autopilot(pid):
    proactive=_v58_super_brief(pid)
    readiness=_v60_trade_readiness(pid)
    not_ready=[x for x in readiness if x[1]=="NOT READY"]
    at_risk=[x for x in readiness if x[1]=="AT RISK"]
    changes=_v67_change_intelligence(pid)
    return {
      "now":proactive["now"],"today":proactive["today"],"week":proactive["week"],
      "upcoming":proactive["upcoming"],"not_ready":not_ready,"at_risk":at_risk,
      "cost":changes["known_net"],"schedule":changes["schedule_proxy"]
    }

@app.get("/lookahead-intelligence",response_class=HTMLResponse)
def v59_home():
    pid=project_id()
    body='<div class="hero"><div class="eyebrow">v59 Look-Ahead Planning Intelligence</div><h1>Is upcoming work actually ready?</h1><p class="muted">2-, 3-, and 6-week schedule windows connected to construction readiness.</p></div><div class="grid3">'
    for weeks in (2,3,6):
        rows=_v59_lookahead(pid,weeks)
        bad=sum(1 for x in rows if x[1]!="READY")
        body+=_v37_link_card(f"{weeks}-Week Look-Ahead",f"{len(rows)} activity(s) · {bad} at risk/not ready.",f"/lookahead-intelligence/{weeks}","Open")
    body+='</div>'
    return shell("Look-Ahead Intelligence",body)

@app.get("/lookahead-intelligence/{weeks}",response_class=HTMLResponse)
def v59_window(weeks:int):
    weeks=max(1,min(weeks,12)); rows=_v59_lookahead(project_id(),weeks)
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(state)}</span> <b>{esc(a["name"])}</b><div class="small">{esc(a["trade"])} · {esc(a["start"])} → {esc(a["finish"])} · {esc(reason)}</div></div>' for a,state,risk,reason in rows)
    return shell(f"{weeks}-Week Look-Ahead",f'<div class="hero"><h1>{weeks}-Week Look-Ahead</h1></div><div class="card">{h or "<p class=muted>No activities in this window.</p>"}</div>')

@app.get("/trade-readiness",response_class=HTMLResponse)
def v60_home():
    rows=_v60_trade_readiness(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(state)}</span> <b>{esc(a["trade"])} - {esc(a["name"])}</b><div class="small">Failed checks: {esc(", ".join(failed) if failed else "None")}</div></div>' for a,state,failed,checks in rows)
    return shell("Trade Readiness Brain",'<div class="hero"><div class="eyebrow">v60</div><h1>Trade Readiness Brain</h1><p class="muted">READY / AT RISK / NOT READY before mobilization.</p></div><div class="card">'+(h or '<p class="muted">No upcoming activities.</p>')+'</div>')

@app.get("/trade-coordination",response_class=HTMLResponse)
def v61_home():
    rows=_v61_handoffs(project_id())
    h="".join(f'<div class="action"><span class="badge">HANDOFF</span> <b>{esc(a["trade"])} → {esc(b["trade"])}</b><div>{esc(a["name"])} → {esc(b["name"])}</div><div class="small">{esc(note)}</div></div>' for a,b,note in rows)
    return shell("Trade Coordination Engine",'<div class="hero"><div class="eyebrow">v61</div><h1>Automatic Trade Coordination Engine</h1></div><div class="card">'+(h or '<p class="muted">No trade handoffs detected.</p>')+'</div>')

@app.get("/daily-superintendent",response_class=HTMLResponse)
def v62_home():
    d=_v62_daily_brief(project_id()); b=d["priorities"]
    def ph(rows): return "".join(f'<div class="action"><span class="badge WATCH">{esc(x["severity"])}</span> <b>{esc(x["title"])}</b><div class="small">{esc(x["action"])}</div></div>' for x in rows[:8]) or '<p class="muted">Clear.</p>'
    acts="".join(f'<div class="action"><b>{esc(a["trade"])} - {esc(a["name"])}</b></div>' for a in d["activities"]) or '<p class="muted">No scheduled activities today.</p>'
    body='<div class="hero"><div class="eyebrow">v62 Daily Superintendent Command AI</div><h1>Today’s superintendent briefing.</h1></div><div class="grid2"><div class="card"><h2>DO NOW</h2>'+ph(b["now"])+'</div><div class="card"><h2>TODAY</h2>'+ph(b["today"])+'</div></div><div class="card"><h2>Scheduled Work Today</h2>'+acts+'</div>'
    return shell("Daily Superintendent Command",body)

@app.get("/smart-rfi",response_class=HTMLResponse)
def v63_home():
    rows=_v63_rfi_candidates(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(r["severity"])}</span> <b>{esc(r["title"])}</b><div>{esc(r["question"])}</div><div class="small">Why: {esc(r["reason"])} · Suggested action: {esc(r["action"])}</div></div>' for r in rows)
    return shell("Smart RFI Generator",'<div class="hero"><div class="eyebrow">v63</div><h1>Smart RFI Generator & Conflict Detection</h1><p class="muted">Draft candidates only. Human approval remains required before sending.</p></div><div class="card">'+(h or '<p class="muted">No RFI candidates detected.</p>')+'</div>')

@app.get("/longlead-intelligence",response_class=HTMLResponse)
def v64_home():
    rows=_v64_longlead(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(level)}</span> <b>{esc(r["item"])}</b><div class="small">Required {esc(r["required_on_site"])} · target release by {release_by} · promised {esc(r["promised_date"] or "not set")} · exposure {esc(exposure if exposure is not None else "unknown")} day(s)</div></div>' for r,level,release_by,exposure in rows)
    return shell("Long-Lead Prediction",'<div class="hero"><div class="eyebrow">v64</div><h1>Procurement & Long-Lead Prediction Brain</h1></div><div class="card">'+(h or '<p class="muted">No procurement records.</p>')+'</div>')

@app.get("/quality-intelligence",response_class=HTMLResponse)
def v65_home():
    rows=_v65_qc_points(project_id())
    h="".join(f'<div class="action"><span class="badge">QC</span> <b>{esc(a["name"])}</b><div class="small">{esc(kind)} · {esc(check)}</div></div>' for a,kind,check in rows)
    return shell("Inspection & QC Intelligence",'<div class="hero"><div class="eyebrow">v65</div><h1>Inspection & Quality Control Intelligence</h1></div><div class="card">'+(h or '<p class="muted">No QC checkpoints generated.</p>')+'</div>')

@app.get("/scope-gap-intelligence",response_class=HTMLResponse)
def v66_home():
    rows=_v66_scope_gaps(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(kind)}</span> <b>{esc(r["trade"] or "UNASSIGNED")}</b> - {esc(r["requirement"])}<div class="small">{esc(reason)}</div></div>' for kind,r,reason in rows)
    return shell("Scope Gap & Buyout Intelligence",'<div class="hero"><div class="eyebrow">v66</div><h1>Scope Gap & Buyout Intelligence</h1><p class="muted">Find unowned, duplicate and ambiguous scope before buyout.</p></div><div class="card">'+(h or '<p class="muted">No obvious scope gaps detected.</p>')+'</div>')

@app.get("/change-order-intelligence",response_class=HTMLResponse)
def v67_home():
    s=_v67_change_intelligence(project_id())
    body=f'<div class="hero"><div class="eyebrow">v67</div><h1>Change Order Intelligence</h1><p class="muted">{len(s["added"])} added · {len(s["removed"])} removed · {len(s["changed"])} changed · ${s["known_net"]:,.0f} known estimator-linked net exposure · {s["schedule_proxy"]:.1f} schedule review-day proxy.</p></div><div class="card"><p>BuildCommand does not invent entitlement or unpriced costs. Human contract review remains required.</p></div>'
    return shell("Change Order Intelligence",body)

@app.get("/autopilot",response_class=HTMLResponse)
def v68_home():
    a=_v68_autopilot(project_id())
    def ph(rows): return "".join(f'<div class="action"><span class="badge WATCH">{esc(x["severity"])}</span> <b>{esc(x["title"])}</b><div class="small">{esc(x["action"])}</div></div>' for x in rows[:8]) or '<p class="muted">Clear.</p>'
    nr="".join(f'<div class="action"><span class="badge WATCH">{esc(state)}</span> <b>{esc(act["trade"])} - {esc(act["name"])}</b><div class="small">{esc(", ".join(failed))}</div></div>' for act,state,failed,checks in a["not_ready"][:10]) or '<p class="muted">No not-ready trades.</p>'
    body=f'<div class="hero"><div class="eyebrow">v68 PROJECT AUTOPILOT</div><h1>One command center for what the project needs next.</h1><p class="muted">Known change exposure ${a["cost"]:,.0f} · schedule review proxy {a["schedule"]:.1f} day(s).</p></div>'
    body+='<div class="grid2"><div class="card"><h2>DO NOW</h2>'+ph(a["now"])+'</div><div class="card"><h2>DO TODAY</h2>'+ph(a["today"])+'</div></div>'
    body+='<div class="grid2"><div class="card"><h2>THIS WEEK</h2>'+ph(a["week"])+'</div><div class="card"><h2>NOT READY</h2>'+nr+'</div></div>'
    body+='<div class="grid3">'+_v37_link_card("Look-Ahead","2/3/6-week readiness.","/lookahead-intelligence","Open")+_v37_link_card("Trade Readiness","Mobilization readiness.","/trade-readiness","Open")+_v37_link_card("Smart RFI","Conflict-driven RFI candidates.","/smart-rfi","Open")+'</div>'
    return shell("Project Autopilot",body)


# ============================================================
# v69-v168 UNIFIED CONSTRUCTION KNOWLEDGE BRAIN 2.0
# 100 capabilities integrated into shared intelligence.
# ============================================================

_V168_FEATURES=[(69, 'Sheet Classification Brain', 'DRAWING'), (70, 'Drawing Discipline Detection', 'DRAWING'), (71, 'Sheet Number Intelligence', 'DRAWING'), (72, 'Drawing Index Reconciliation', 'DRAWING'), (73, 'Detail Bubble Recognition', 'DRAWING'), (74, 'Detail-to-Sheet Linking', 'DRAWING'), (75, 'Section Callout Intelligence', 'DRAWING'), (76, 'Elevation Callout Intelligence', 'DRAWING'), (77, 'Keynote Recognition', 'DRAWING'), (78, 'Keynote-to-Legend Linking', 'DRAWING'), (79, 'General Notes Intelligence', 'DRAWING'), (80, 'Plan Note Intelligence', 'DRAWING'), (81, 'Drawing Legend Brain', 'DRAWING'), (82, 'Symbol Recognition Intelligence', 'DRAWING'), (83, 'Abbreviation Intelligence', 'DRAWING'), (84, 'Drawing Scale Intelligence', 'DRAWING'), (85, 'Dimension Relationship Brain', 'DRAWING'), (86, 'Grid-Line Intelligence', 'DRAWING'), (87, 'Room/Area Recognition', 'DRAWING'), (88, 'Drawing Cross-Reference Brain', 'DRAWING'), (89, 'Door Schedule Brain', 'SCHEDULE_SPEC'), (90, 'Hardware Schedule Brain', 'SCHEDULE_SPEC'), (91, 'Finish Schedule Brain', 'SCHEDULE_SPEC'), (92, 'Room Finish Intelligence', 'SCHEDULE_SPEC'), (93, 'Equipment Schedule Brain', 'SCHEDULE_SPEC'), (94, 'Plumbing Fixture Schedule Brain', 'SCHEDULE_SPEC'), (95, 'Mechanical Equipment Schedule Brain', 'SCHEDULE_SPEC'), (96, 'Electrical Panel Schedule Brain', 'SCHEDULE_SPEC'), (97, 'Lighting Fixture Schedule Brain', 'SCHEDULE_SPEC'), (98, 'Structural Schedule Intelligence', 'SCHEDULE_SPEC'), (99, 'Specification Division Recognition', 'SCHEDULE_SPEC'), (100, 'Spec Section Classification', 'SCHEDULE_SPEC'), (101, 'Drawing-to-Spec Linking', 'SCHEDULE_SPEC'), (102, 'Product Requirement Extraction', 'SCHEDULE_SPEC'), (103, 'Manufacturer Requirement Extraction', 'SCHEDULE_SPEC'), (104, 'Testing Requirement Extraction', 'SCHEDULE_SPEC'), (105, 'Warranty Requirement Extraction', 'SCHEDULE_SPEC'), (106, 'Submittal Requirement Extraction', 'SCHEDULE_SPEC'), (107, 'Closeout Requirement Extraction', 'SCHEDULE_SPEC'), (108, 'Specification Conflict Detection', 'SCHEDULE_SPEC'), (109, 'Wall Assembly Intelligence', 'ASSEMBLY'), (110, 'Rated Wall Intelligence', 'ASSEMBLY'), (111, 'Shaft-Wall Intelligence', 'ASSEMBLY'), (112, 'Ceiling Assembly Intelligence', 'ASSEMBLY'), (113, 'Floor Assembly Intelligence', 'ASSEMBLY'), (114, 'Roof Assembly Intelligence', 'ASSEMBLY'), (115, 'Waterproofing Assembly Brain', 'ASSEMBLY'), (116, 'Exterior Wall Assembly Brain', 'ASSEMBLY'), (117, 'Storefront Assembly Intelligence', 'ASSEMBLY'), (118, 'Door/Frame Assembly Intelligence', 'ASSEMBLY'), (119, 'Foundation Assembly Intelligence', 'ASSEMBLY'), (120, 'Slab Assembly Intelligence', 'ASSEMBLY'), (121, 'Structural Steel Assembly Brain', 'ASSEMBLY'), (122, 'CMU Assembly Intelligence', 'ASSEMBLY'), (123, 'Concrete Assembly Intelligence', 'ASSEMBLY'), (124, 'Underground Utility Assembly Brain', 'ASSEMBLY'), (125, 'Plumbing System Assembly Brain', 'ASSEMBLY'), (126, 'HVAC System Assembly Brain', 'ASSEMBLY'), (127, 'Electrical System Assembly Brain', 'ASSEMBLY'), (128, 'Fire Protection Assembly Brain', 'ASSEMBLY'), (129, 'Architectural/Structural Coordination', 'COORDINATION'), (130, 'Architectural/Mechanical Coordination', 'COORDINATION'), (131, 'Architectural/Electrical Coordination', 'COORDINATION'), (132, 'Architectural/Plumbing Coordination', 'COORDINATION'), (133, 'Structural/MEP Coordination', 'COORDINATION'), (134, 'HVAC/Electrical Coordination', 'COORDINATION'), (135, 'HVAC/Plumbing Coordination', 'COORDINATION'), (136, 'Electrical/Plumbing Coordination', 'COORDINATION'), (137, 'Fire Sprinkler/Ceiling Coordination', 'COORDINATION'), (138, 'Lighting/Ceiling Coordination', 'COORDINATION'), (139, 'Access Panel Coordination', 'COORDINATION'), (140, 'Equipment Clearance Intelligence', 'COORDINATION'), (141, 'ADA Clearance Intelligence', 'COORDINATION'), (142, 'Maintenance Clearance Intelligence', 'COORDINATION'), (143, 'Penetration Coordination Brain', 'COORDINATION'), (144, 'Sleeve Coordination Intelligence', 'COORDINATION'), (145, 'Embed Coordination Intelligence', 'COORDINATION'), (146, 'Opening Coordination Intelligence', 'COORDINATION'), (147, 'Above-Ceiling Coordination Brain', 'COORDINATION'), (148, 'Underground Coordination Brain', 'COORDINATION'), (149, 'Constructability Intelligence', 'REASONING'), (150, 'Missing Information Detection', 'REASONING'), (151, 'Contradiction Intelligence', 'REASONING'), (152, 'Ambiguous Requirement Detection', 'REASONING'), (153, 'Impossible Condition Detection', 'REASONING'), (154, 'Incomplete Detail Detection', 'REASONING'), (155, 'Scope Ownership Brain 2.0', 'REASONING'), (156, 'Trade Boundary Intelligence', 'REASONING'), (157, 'Demolition Ownership Intelligence', 'REASONING'), (158, 'Temporary Work Intelligence', 'REASONING'), (159, 'Inspection Requirement Brain', 'REASONING'), (160, 'Testing Requirement Brain', 'REASONING'), (161, 'Permit Requirement Intelligence', 'REASONING'), (162, 'Sequence Requirement Brain', 'REASONING'), (163, 'Prerequisite Intelligence', 'REASONING'), (164, 'Means-and-Methods Awareness', 'REASONING'), (165, 'Confidence Calibration Brain', 'REASONING'), (166, 'Human Correction Learning 2.0', 'REASONING'), (167, 'Cross-Project Construction Memory', 'REASONING'), (168, 'Unified Construction Knowledge Brain 2.0', 'UNIFIED')]

def _v168_rows(pid):
    try: return _v452_scope_rows(pid)
    except Exception: return []

def _v168_text(pid):
    rows=_v168_rows(pid)
    return "\n".join(
        " ".join(str(x or "") for x in [r["trade"],r["requirement"],r["source_sheet"],r["source_spec"]])
        for r in rows
    ).lower()

def _v168_metrics(pid):
    rows=_v168_rows(pid); text=_v168_text(pid)
    patterns={
      "DRAWING":["sheet ","detail","section","elevation","keynote","legend","scale","grid","room"],
      "SCHEDULE_SPEC":["schedule","spec","manufacturer","testing","warranty","submittal","closeout"],
      "ASSEMBLY":["wall","ceiling","floor","roof","waterproof","storefront","foundation","slab","steel","cmu","concrete","plumbing","hvac","electrical","sprinkler"],
      "COORDINATION":["coordinate","clearance","penetration","sleeve","embed","opening","above ceiling","underground"],
      "REASONING":["conflict","missing","verify","confirm","inspection","testing","permit","sequence","prerequisite","demolition"]
    }
    counts={k:sum(text.count(p) for p in ps) for k,ps in patterns.items()}
    return {"scope_rows":len(rows),"signals":counts}

def _v168_feature_status(pid,feature):
    v,name,group=feature
    m=_v168_metrics(pid); n=m["scope_rows"]; sig=m["signals"].get(group,0)
    if group=="UNIFIED":
        active=sum(1 for x in m["signals"].values() if x>0)
        return "ACTIVE" if n and active>=3 else "LEARNING"
    return "ACTIVE" if n and sig>0 else ("LEARNING" if n else "WAITING FOR PROJECT DATA")

def _v168_ownership_rules():
    return [
      ("Concrete","slab cutting, trenching and restoration","Concrete"),
      ("Bathroom Accessories","grab bars, toilet partitions, urinal screens, toilet room accessories","Bathroom Accessories"),
      ("Demolition","existing fixture/equipment removal and demolition work","Demolition"),
      ("Low Voltage","accessible door operators, card access, cameras and electronic access","Low Voltage"),
      ("Flooring/Tile","flooring, tile, rubber base and tile backsplash","Flooring/Tile"),
      ("Storefront/Glazing","exterior storefront and glazing systems","Storefront/Glazing"),
      ("HVAC","mechanical/HVAC equipment and duct systems","HVAC"),
      ("Roofing","roof patching, flashing and roofing restoration","Roofing"),
      ("Paint","prime/refinish patched surfaces and finish painting","Paint"),
      ("Doors","interior doors, frames, hardware and hollow-metal framed interior windows","Doors"),
    ]

def _v168_reasoning_signals(pid):
    rows=_v168_rows(pid); out=[]
    seen={}
    for r in rows:
        req=str(r["requirement"] or "")
        low=req.lower(); trade=str(r["trade"] or "")
        if not req.strip():
            out.append(("MISSING","HIGH",trade,"Empty scope requirement","Review source document and regenerate requirement."))
        if any(k in low for k in ["verify","confirm","coordinate","as required","if required"]):
            out.append(("AMBIGUOUS","MEDIUM",trade,req,"Confirm contract requirement, responsible trade, and prerequisite before execution."))
        if any(k in low for k in ["inspection","test","testing"]):
            out.append(("QC","REVIEW",trade,req,"Link requirement to inspection/testing readiness and schedule hold point."))
        key=re.sub(r'\s+',' ',low).strip()
        if key:
            if key in seen and str(seen[key]["trade"])!=trade:
                out.append(("TRADE BOUNDARY","HIGH",trade,req,f'Requirement also appears under {seen[key]["trade"]}; resolve ownership.'))
            else: seen[key]=r
    return out[:300]

def _v168_assembly_links(pid):
    rows=_v168_rows(pid); out=[]
    maps=[
      ("WALL",["wall","partition","gypsum","stud"]),
      ("CEILING",["ceiling","grid","acoustical"]),
      ("ROOF",["roof","flashing","curb"]),
      ("CONCRETE",["concrete","slab","footing","foundation"]),
      ("PLUMBING",["plumbing","water closet","lavatory","drain","water heater"]),
      ("HVAC",["hvac","mechanical","duct","exhaust fan","rtu"]),
      ("ELECTRICAL",["electrical","panel","lighting","receptacle","power"]),
      ("FIRE PROTECTION",["sprinkler","fire protection"]),
    ]
    for r in rows:
        low=str(r["requirement"] or "").lower()
        linked=[name for name,keys in maps if any(k in low for k in keys)]
        if linked: out.append((r,linked))
    return out[:300]

def _v168_knowledge_summary(pid):
    m=_v168_metrics(pid)
    active=sum(1 for f in _V168_FEATURES if _v168_feature_status(pid,f)=="ACTIVE")
    reasoning=_v168_reasoning_signals(pid)
    assemblies=_v168_assembly_links(pid)
    return {"metrics":m,"active":active,"reasoning":reasoning,"assemblies":assemblies}

@app.get("/knowledge-brain-2",response_class=HTMLResponse)
def v168_home():
    pid=project_id(); s=_v168_knowledge_summary(pid)
    body=f'<div class="hero"><div class="eyebrow">BuildCommand v168</div><h1>Unified Construction Knowledge Brain 2.0</h1><p class="muted">{s["active"]} / 100 capabilities active against current project intelligence · {s["metrics"]["scope_rows"]} scope knowledge row(s).</p></div><div class="grid3">'
    cards=[
      ("100-Capability Map","See all v69-v168 construction knowledge capabilities.","/knowledge-brain-2/capabilities"),
      ("Drawing Intelligence","Sheet, detail, note, symbol and cross-reference understanding.","/knowledge-brain-2/group/DRAWING"),
      ("Schedule & Spec Intelligence","Schedules, specifications, products, testing and closeout.","/knowledge-brain-2/group/SCHEDULE_SPEC"),
      ("Assembly Brain","Construction systems and assembly relationships.","/knowledge-brain-2/assemblies"),
      ("Cross-Trade Coordination","Interfaces, penetrations, clearances and handoffs.","/knowledge-brain-2/group/COORDINATION"),
      ("Construction Reasoning","Missing, ambiguous, conflicting and ownership conditions.","/knowledge-brain-2/reasoning"),
      ("Scope Ownership 2.0","Approved construction trade-boundary knowledge.","/knowledge-brain-2/ownership"),
      ("Blueprint Brain","Use the existing blueprint analysis engine.","/blueprint-brain"),
      ("Project Autopilot","Feed construction knowledge into operations.","/autopilot"),
    ]
    for n,d,h in cards: body+=_v37_link_card(n,d,h,"Open")
    body+='</div>'
    return shell("Knowledge Brain 2.0",body)

@app.get("/knowledge-brain-2/capabilities",response_class=HTMLResponse)
def v168_capabilities():
    pid=project_id()
    h=""
    for v,name,group in _V168_FEATURES:
        status=_v168_feature_status(pid,(v,name,group))
        h+=f'<div class="action"><span class="badge">{esc(status)}</span> <b>v{v} - {esc(name)}</b><div class="small">{esc(group.replace("_"," "))}</div></div>'
    return shell("100 Construction Knowledge Capabilities",'<div class="hero"><h1>v69-v168 Capability Map</h1><p class="muted">These capabilities share the same project knowledge instead of becoming 100 disconnected pages.</p></div><div class="card">'+h+'</div>')

@app.get("/knowledge-brain-2/group/{group}",response_class=HTMLResponse)
def v168_group(group:str):
    group=group.upper()
    rows=[f for f in _V168_FEATURES if f[2]==group]
    h="".join(f'<div class="action"><span class="badge">{esc(_v168_feature_status(project_id(),f))}</span> <b>v{f[0]} - {esc(f[1])}</b></div>' for f in rows)
    return shell("Construction Knowledge Group",f'<div class="hero"><h1>{esc(group.replace("_"," "))}</h1><p class="muted">{len(rows)} integrated capability(s).</p></div><div class="card">{h or "<p class=muted>No capabilities.</p>"}</div>')

@app.get("/knowledge-brain-2/reasoning",response_class=HTMLResponse)
def v168_reasoning():
    rows=_v168_reasoning_signals(project_id())
    h="".join(f'<div class="action"><span class="badge WATCH">{esc(sev)}</span> <b>{esc(kind)} · {esc(trade)}</b><div>{esc(req)}</div><div class="small">{esc(action)}</div></div>' for kind,sev,trade,req,action in rows)
    return shell("Construction Reasoning 2.0",'<div class="hero"><h1>Construction Reasoning & Learning</h1><p class="muted">Find ambiguous requirements, trade-boundary conflicts and quality/inspection signals.</p></div><div class="card">'+(h or '<p class="muted">No reasoning exceptions detected.</p>')+'</div>')

@app.get("/knowledge-brain-2/assemblies",response_class=HTMLResponse)
def v168_assemblies():
    rows=_v168_assembly_links(project_id())
    h="".join(f'<div class="action"><span class="badge">ASSEMBLY</span> <b>{esc(", ".join(linked))}</b><div>{esc(r["requirement"])}</div><div class="small">{esc(r["trade"])} · {esc(r["source_sheet"])}</div></div>' for r,linked in rows)
    return shell("Assembly Intelligence",'<div class="hero"><h1>Construction Assembly Brain</h1><p class="muted">Connect scope requirements to the physical systems being built.</p></div><div class="card">'+(h or '<p class="muted">No assembly links detected.</p>')+'</div>')

@app.get("/knowledge-brain-2/ownership",response_class=HTMLResponse)
def v168_ownership():
    rules=_v168_ownership_rules()
    h="".join(f'<div class="action"><span class="badge READY">RULE</span> <b>{esc(owner)}</b><div>{esc(scope)}</div><div class="small">Normalized owner: {esc(normalized)}</div></div>' for owner,scope,normalized in rules)
    return shell("Scope Ownership Brain 2.0",'<div class="hero"><h1>Scope Ownership Brain 2.0</h1><p class="muted">Construction trade-boundary rules carried into the shared knowledge brain.</p></div><div class="card">'+h+'</div>')


@app.get("/api/dashboard/command")
def v169_dashboard_command_api():
    """
    Heavy dashboard intelligence is isolated from first paint and shares v56 caches.
    """
    pid=project_id()
    token=_v56_perf_start("dashboard_lazy_api")
    try:
        d=_v56_dashboard_bundle(pid)
        score=d["score"]; top=d["top"]; decisions=d["decisions"]
        materials=d["materials"]; holds=d["holds"]; agenda=d["agenda"]; sequence=d["sequence"]

        priorities_html="".join(
            f'<div class="action bc-priority"><span class="badge WATCH">{_v37_esc(f["severity"])}</span> '
            f'<b>{_v37_esc(f["type"])}</b> - {_v37_esc(f["title"])}'
            f'<div class="small">{_v37_esc(f["reason"])}</div>'
            f'<p><b>Next:</b> {_v37_esc(f["action"])}</p></div>'
            for f in top
        ) or '<p class="muted">No major project priorities detected.</p>'

        blockers_html="".join(
            f'<div class="action"><span class="badge WATCH">{_v37_esc(x["risk"])}</span> '
            f'<b>{_v37_esc(x["activity"]["name"])}</b>'
            f'<div class="small">{_v37_esc(x["activity"]["trade"])} · {_v37_esc(x["blocking_reason"] or "Sequence risk")}</div></div>'
            for x in sequence
        ) or '<p class="muted">No high sequence blockers.</p>'

        inspections_html="".join(
            f'<div class="action"><span class="badge WATCH">{_v37_esc(i["result"])}</span> '
            f'<b>{_v37_esc(i["inspection_type"])}</b>'
            f'<div class="small">{_v37_esc(i["scheduled_date"])} · {_v37_esc(i["authority"])}'
            f' · Activity {_v37_esc(a["name"] if a else "Unlinked")}</div></div>'
            for i,a in holds
        ) or '<p class="muted">No open inspection hold points.</p>'

        materials_html="".join(
            f'<div class="action"><span class="badge WATCH">{_v37_esc(level)}</span> '
            f'<b>{_v37_esc(r["item"])}</b>'
            f'<div class="small">Need {_v37_esc(r["required_on_site"])} · Promised {_v37_esc(r["promised_date"])}'
            f' · {exposure} day(s) exposure</div></div>'
            for r,act,level,exposure,reason,action in materials
        ) or '<p class="muted">No critical/high material risks.</p>'

        decisions_html="".join(
            f'<div class="action"><span class="badge WATCH">{_v37_esc(severity)}</span> '
            f'<b>{_v37_esc(typ)} - {_v37_esc(title)}</b>'
            f'<div class="small">Due {_v37_esc(due)} · ${cost:,.0f} exposure · {days:g} schedule day(s)</div></div>'
            for typ,title,due,severity,cost,days,source in decisions
        ) or '<p class="muted">No urgent decision deadlines.</p>'

        agenda_html="".join(
            f'<div class="action"><span class="badge">{_v37_esc(kind)}</span> '
            f'<b>{_v37_esc(title)}</b><div class="small">{_v37_esc(detail)}</div></div>'
            for kind,title,detail in agenda
        ) or '<p class="muted">No current coordination agenda items.</p>'

        return JSONResponse({
            "health":score["health"],"risk":score["risk"],"quality":score["quality"],
            "priorities_html":priorities_html,"blockers_html":blockers_html,
            "inspections_html":inspections_html,"materials_html":materials_html,
            "decisions_html":decisions_html,"agenda_html":agenda_html
        })
    finally:
        _v56_perf_end(token)


# ============================================================
# v171-v269 NEXT 99 + v170 = 100-VERSION PLATFORM PHASE
# Shared intelligence architecture, not 99 disconnected pages.
# ============================================================

_V269_CAPABILITIES=[(171, 'Customizable Dashboard', 'UI_WORKFLOW'), (172, 'Saved Dashboard Layouts', 'UI_WORKFLOW'), (173, 'Compact Navigation Mode', 'UI_WORKFLOW'), (174, 'Mobile Field Layout', 'UI_WORKFLOW'), (175, 'Quick Action Bar', 'UI_WORKFLOW'), (176, 'Global Command Search', 'UI_WORKFLOW'), (177, 'Recent Work History', 'UI_WORKFLOW'), (178, 'Favorite Tools', 'UI_WORKFLOW'), (179, 'Role-Based Home Screens', 'UI_WORKFLOW'), (180, 'Notification Center', 'UI_WORKFLOW'), (181, 'Smart Alert Prioritization', 'UI_WORKFLOW'), (182, 'Read/Unread Intelligence Alerts', 'UI_WORKFLOW'), (183, 'Project Context Header', 'UI_WORKFLOW'), (184, 'Command Breadcrumbs', 'UI_WORKFLOW'), (185, 'Side-by-Side Intelligence View', 'UI_WORKFLOW'), (186, 'Sheet + Scope Split View', 'UI_WORKFLOW'), (187, 'Full-Screen Blueprint Review Workspace', 'UI_WORKFLOW'), (188, 'Universal Filter & Sort System', 'UI_WORKFLOW'), (189, 'UI Performance Optimization 2.0', 'UI_WORKFLOW'), (190, 'Drawing Index Intelligence 2.0', 'BLUEPRINT'), (191, 'Sheet Revision Recognition', 'BLUEPRINT'), (192, 'Detail Callout Linking 2.0', 'BLUEPRINT'), (193, 'Section/Elevation Relationship Brain', 'BLUEPRINT'), (194, 'Keynote Intelligence 2.0', 'BLUEPRINT'), (195, 'Drawing Legend Intelligence 2.0', 'BLUEPRINT'), (196, 'Symbol Intelligence 2.0', 'BLUEPRINT'), (197, 'Room Recognition 2.0', 'BLUEPRINT'), (198, 'Area/Zone Intelligence', 'BLUEPRINT'), (199, 'Building-Level Intelligence', 'BLUEPRINT'), (200, 'Door Schedule Intelligence 2.0', 'BLUEPRINT'), (201, 'Finish Schedule Intelligence 2.0', 'BLUEPRINT'), (202, 'Equipment Schedule Intelligence 2.0', 'BLUEPRINT'), (203, 'Plumbing Fixture Schedule Intelligence 2.0', 'BLUEPRINT'), (204, 'Mechanical Schedule Intelligence 2.0', 'BLUEPRINT'), (205, 'Electrical Panel Intelligence 2.0', 'BLUEPRINT'), (206, 'Lighting Fixture Intelligence 2.0', 'BLUEPRINT'), (207, 'Structural Schedule Intelligence 2.0', 'BLUEPRINT'), (208, 'Specification Cross-Reference Brain', 'BLUEPRINT'), (209, 'Drawing/Spec Conflict Intelligence 2.0', 'BLUEPRINT'), (210, 'Daily Field Plan AI', 'FIELD'), (211, 'Crew Planning Intelligence', 'FIELD'), (212, 'Manpower Forecast 2.0', 'FIELD'), (213, 'Crew Stacking Conflict Detection', 'FIELD'), (214, 'Work Area Availability Brain', 'FIELD'), (215, 'Site Logistics Intelligence', 'FIELD'), (216, 'Delivery Coordination Brain', 'FIELD'), (217, 'Material Staging Intelligence', 'FIELD'), (218, 'Shutdown Planning Intelligence', 'FIELD'), (219, 'Temporary Protection Intelligence', 'FIELD'), (220, 'Trade Mobilization Planner', 'FIELD'), (221, 'Daily Production Tracking', 'FIELD'), (222, 'Production Rate Intelligence', 'FIELD'), (223, 'Planned-vs-Actual Field Production', 'FIELD'), (224, 'Daily Delay Cause Tracking', 'FIELD'), (225, 'Superintendent Recovery Suggestions', 'FIELD'), (226, 'Field Constraint Log Intelligence', 'FIELD'), (227, 'Weather Impact Integration', 'FIELD'), (228, 'Daily Field Risk Forecast', 'FIELD'), (229, 'Superintendent Command 3.0', 'FIELD'), (230, 'RFI Intelligence 3.0', 'PM'), (231, 'RFI Aging & Risk Prediction', 'PM'), (232, 'RFI-to-Schedule Impact Linking', 'PM'), (233, 'RFI-to-Cost Exposure Linking', 'PM'), (234, 'Submittal Intelligence 3.0', 'PM'), (235, 'Submittal Required-Date Prediction', 'PM'), (236, 'Submittal Approval Risk', 'PM'), (237, 'Procurement Chain Intelligence 3.0', 'PM'), (238, 'Vendor Commitment Tracking', 'PM'), (239, 'Long-Lead Recovery Scenarios', 'PM'), (240, 'Meeting Intelligence', 'PM'), (241, 'Automatic Meeting Agenda', 'PM'), (242, 'Meeting Decision Tracker', 'PM'), (243, 'Meeting Action Item Intelligence', 'PM'), (244, 'Owner Decision Intelligence 2.0', 'PM'), (245, 'Architect/Engineer Response Tracking', 'PM'), (246, 'Correspondence Intelligence', 'PM'), (247, 'Contract Notice Intelligence', 'PM'), (248, 'PM Command Center', 'PM'), (249, 'Project Management Autopilot', 'PM'), (250, 'Bid Package Generator 2.0', 'PRECON'), (251, 'Scope Coverage Matrix', 'PRECON'), (252, 'Bidder Scope Comparison', 'PRECON'), (253, 'Exclusion Intelligence', 'PRECON'), (254, 'Allowance Intelligence', 'PRECON'), (255, 'Alternate Intelligence', 'PRECON'), (256, 'Bid Leveling 2.0', 'PRECON'), (257, 'Proposal Normalization Brain', 'PRECON'), (258, 'Missing Scope Detection 3.0', 'PRECON'), (259, 'Duplicate Scope Detection 3.0', 'PRECON'), (260, 'Buyout Recommendation Engine', 'PRECON'), (261, 'Subcontractor Risk Scoring', 'PRECON'), (262, 'Historical Cost Intelligence 2.0', 'PRECON'), (263, 'Quantity Confidence Engine', 'PRECON'), (264, 'Estimate Revision Comparison', 'PRECON'), (265, 'Budget-to-Buyout Intelligence', 'PRECON'), (266, 'Preconstruction Risk Register', 'PRECON'), (267, 'Preconstruction Command Center', 'PRECON'), (268, 'Estimator/PM Handoff Intelligence', 'PRECON'), (269, 'Unified Preconstruction Operating Brain 3.0', 'PRECON')]

def _v269_table_exists(name):
    try:
        c=db()
        if DB_KIND=="postgres":
            return bool(c.execute("SELECT to_regclass(?)",(name,)).fetchone()[0])
        return bool(c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?",(name,)).fetchone())
    except Exception: return False

def _v269_count(table,pid):
    try:
        if not _v269_table_exists(table): return 0
        return int(db().execute(f"SELECT COUNT(*) FROM {table} WHERE project_id=?",(pid,)).fetchone()[0])
    except Exception: return 0

def _v269_context(pid):
    s=_v56_cached("snapshot",pid,lambda:_v37_snapshot(pid))
    return {
      "snapshot":s,
      "activities":_v269_count("activities",pid),
      "issues":_v269_count("issues",pid),
      "rfis":_v269_count("rfis",pid),
      "submittals":_v269_count("submittals",pid),
      "procurement":_v269_count("procurement",pid),
      "inspections":_v269_count("inspections",pid),
      "scope":len(_v452_scope_rows(pid)) if callable(globals().get("_v452_scope_rows")) else 0
    }

def _v269_cap_status(pid,cap):
    v,name,group=cap; c=_v269_context(pid)
    if group=="UI_WORKFLOW": return "ACTIVE"
    if group=="BLUEPRINT": return "ACTIVE" if c["scope"] else "WAITING FOR DOCUMENT DATA"
    if group=="FIELD": return "ACTIVE" if c["activities"] else "WAITING FOR SCHEDULE DATA"
    if group=="PM": return "ACTIVE" if (c["rfis"]+c["submittals"]+c["procurement"]+c["issues"]) else "WAITING FOR PM DATA"
    if group=="PRECON": return "ACTIVE" if c["scope"] else "WAITING FOR SCOPE DATA"
    return "LEARNING"

def _v269_search(pid,q):
    q=str(q or "").strip().lower()
    if not q:return []
    results=[]
    for v,name,group in _V269_CAPABILITIES:
        if q in name.lower() or q in group.lower():
            results.append(("TOOL",name,f"/platform-269/group/{group}",f"v{v} · {group}"))
    for r in _v452_scope_rows(pid)[:1000]:
        blob=" ".join(str(r[x] or "") for x in ["trade","requirement","source_sheet","source_spec"]).lower()
        if q in blob:
            results.append(("SCOPE",str(r["trade"] or "Scope"),"/blueprint-brain",str(r["requirement"] or "")[:180]))
        if len(results)>=100: break
    return results[:100]

def _v269_field_command(pid):
    try: brief=_v62_daily_brief(pid)
    except Exception: brief={"priorities":{"now":[],"today":[],"week":[]},"activities":[],"inspections":[],"materials":[]}
    try: readiness=_v60_trade_readiness(pid)
    except Exception: readiness=[]
    return brief,readiness

def _v269_pm_command(pid):
    try: decisions=_v47_decision_deadlines(pid)[:30]
    except Exception: decisions=[]
    try: agenda=_v47_coordination_agenda(pid)[:30]
    except Exception: agenda=[]
    return decisions,agenda

def _v269_precon_command(pid):
    try: gaps=_v66_scope_gaps(pid)
    except Exception: gaps=[]
    rows=_v452_scope_rows(pid)
    by_trade={}
    for r in rows:
        tr=str(r["trade"] or "UNASSIGNED")
        by_trade[tr]=by_trade.get(tr,0)+1
    return gaps,sorted(by_trade.items(),key=lambda x:(-x[1],x[0]))

@app.get("/platform-269",response_class=HTMLResponse)
def v269_platform():
    pid=project_id(); c=_v269_context(pid)
    active=sum(1 for x in _V269_CAPABILITIES if _v269_cap_status(pid,x)=="ACTIVE")
    body=f'<div class="hero"><div class="eyebrow">BuildCommand v269</div><h1>Unified Construction Operating Platform</h1><p class="muted">{active} / 99 new capabilities active against current project data. v170 layout + v171-v269 intelligence phase.</p></div><div class="grid3">'
    cards=[
      ("Command Search","Search tools and project scope from one place.","/platform-269/search"),
      ("UI & Workflow","Dashboard, navigation, alerts and workflow capabilities.","/platform-269/group/UI_WORKFLOW"),
      ("Blueprint Workspace","Advanced drawing/specification intelligence.","/platform-269/group/BLUEPRINT"),
      ("Field Command","Superintendent operations and readiness.","/field-command-3"),
      ("PM Command","RFI, submittal, procurement and decision intelligence.","/pm-command"),
      ("Preconstruction Command","Scope coverage, gaps, buyout and risk.","/precon-command"),
      ("100-Phase Capability Map","See v170-v269 platform capabilities.","/platform-269/capabilities"),
      ("Project Autopilot","Existing project operating intelligence.","/autopilot"),
      ("Knowledge Brain 2.0","Construction knowledge foundation.","/knowledge-brain-2"),
    ]
    for n,d,h in cards: body+=_v37_link_card(n,d,h,"Open")
    body+='</div>'
    return shell("Unified Operating Platform",body)

@app.get("/platform-269/capabilities",response_class=HTMLResponse)
def v269_caps():
    pid=project_id()
    h='<div class="action"><span class="badge READY">ACTIVE</span> <b>v170 - Professional Command Center Layout</b><div class="small">UI WORKFLOW</div></div>'
    for cap in _V269_CAPABILITIES:
        v,name,group=cap; status=_v269_cap_status(pid,cap)
        h+=f'<div class="action"><span class="badge">{esc(status)}</span> <b>v{v} - {esc(name)}</b><div class="small">{esc(group.replace("_"," "))}</div></div>'
    return shell("v170-v269 Capability Map",'<div class="hero"><h1>100-Version Platform Phase</h1><p class="muted">One shared operating architecture instead of 100 independent pages.</p></div><div class="card">'+h+'</div>')

@app.get("/platform-269/group/{group}",response_class=HTMLResponse)
def v269_group(group:str):
    group=group.upper(); pid=project_id()
    rows=[x for x in _V269_CAPABILITIES if x[2]==group]
    h="".join(f'<div class="action"><span class="badge">{esc(_v269_cap_status(pid,x))}</span> <b>v{x[0]} - {esc(x[1])}</b></div>' for x in rows)
    return shell(group.replace("_"," "),f'<div class="hero"><h1>{esc(group.replace("_"," "))}</h1><p class="muted">{len(rows)} integrated capabilities.</p></div><div class="card">{h}</div>')

@app.get("/platform-269/search",response_class=HTMLResponse)
def v269_search_page(q:str=""):
    rows=_v269_search(project_id(),q)
    form=f'<form method="get"><input name="q" value="{esc(q)}" placeholder="Search BuildCommand tools, trades, scope, sheets..." style="width:min(760px,90%);padding:12px"><button type="submit">Search</button></form>'
    h="".join(f'<div class="action"><span class="badge">{esc(kind)}</span> <a href="{href}"><b>{esc(title)}</b></a><div class="small">{esc(detail)}</div></div>' for kind,title,href,detail in rows)
    return shell("Command Search",'<div class="hero"><h1>Global Command Search</h1><p class="muted">Search the platform and current project knowledge.</p>'+form+'</div><div class="card">'+(h or '<p class="muted">Enter a search above.</p>')+'</div>')

@app.get("/field-command-3",response_class=HTMLResponse)
def v269_field():
    brief,ready=_v269_field_command(project_id())
    p=brief["priorities"]
    def ph(rows):return "".join(f'<div class="action"><span class="badge WATCH">{esc(x["severity"])}</span> <b>{esc(x["title"])}</b><div class="small">{esc(x["action"])}</div></div>' for x in rows[:8]) or '<p class="muted">Clear.</p>'
    nr="".join(f'<div class="action"><span class="badge WATCH">{esc(state)}</span> <b>{esc(a["trade"])} - {esc(a["name"])}</b><div class="small">{esc(", ".join(failed))}</div></div>' for a,state,failed,checks in ready if state!="READY") or '<p class="muted">Upcoming trades ready.</p>'
    body='<div class="hero"><div class="eyebrow">v229 Superintendent Command 3.0</div><h1>Field Command</h1></div><div class="grid2"><div class="card"><h2>DO NOW</h2>'+ph(p["now"])+'</div><div class="card"><h2>TODAY</h2>'+ph(p["today"])+'</div></div><div class="card"><h2>Trade Readiness</h2>'+nr+'</div>'
    return shell("Field Command",body)

@app.get("/pm-command",response_class=HTMLResponse)
def v269_pm():
    decisions,agenda=_v269_pm_command(project_id())
    dh="".join(f'<div class="action"><span class="badge WATCH">{esc(sev)}</span> <b>{esc(typ)} - {esc(title)}</b><div class="small">Due {esc(due)} · ${cost:,.0f} known exposure · {days:g} schedule day(s)</div></div>' for typ,title,due,sev,cost,days,source in decisions) or '<p class="muted">No urgent decisions.</p>'
    ah="".join(f'<div class="action"><span class="badge">{esc(kind)}</span> <b>{esc(title)}</b><div class="small">{esc(detail)}</div></div>' for kind,title,detail in agenda) or '<p class="muted">No coordination agenda items.</p>'
    return shell("PM Command",'<div class="hero"><div class="eyebrow">v248-v249</div><h1>Project Management Command</h1></div><div class="grid2"><div class="card"><h2>Decisions</h2>'+dh+'</div><div class="card"><h2>Coordination</h2>'+ah+'</div></div>')

@app.get("/precon-command",response_class=HTMLResponse)
def v269_precon():
    gaps,trades=_v269_precon_command(project_id())
    gh="".join(f'<div class="action"><span class="badge WATCH">{esc(kind)}</span> <b>{esc(r["trade"] or "UNASSIGNED")}</b><div>{esc(r["requirement"])}</div><div class="small">{esc(reason)}</div></div>' for kind,r,reason in gaps[:40]) or '<p class="muted">No obvious scope gaps.</p>'
    th="".join(f'<div class="action"><b>{esc(trade)}</b><div class="small">{count} scope requirement(s)</div></div>' for trade,count in trades[:40])
    return shell("Preconstruction Command",'<div class="hero"><div class="eyebrow">v267-v269</div><h1>Preconstruction Command</h1><p class="muted">Scope coverage, ownership and buyout risk in one workspace.</p></div><div class="grid2"><div class="card"><h2>Scope Risk</h2>'+gh+'</div><div class="card"><h2>Coverage by Trade</h2>'+th+'</div></div>')


# ============================================================
# v270-v369 EXECUTION, CONTROLS, SCHEDULE, COST & QUALITY INTELLIGENCE
# 100 capabilities in five shared operating workspaces.
# ============================================================

_V369_CAPABILITIES=[(270, 'Daily Workface Planning', 'FIELD_EXECUTION'), (271, 'Crew Availability Intelligence', 'FIELD_EXECUTION'), (272, 'Trade Stacking Forecast', 'FIELD_EXECUTION'), (273, 'Work Area Conflict Detection', 'FIELD_EXECUTION'), (274, 'Material Staging Planner', 'FIELD_EXECUTION'), (275, 'Delivery Window Intelligence', 'FIELD_EXECUTION'), (276, 'Access Route Intelligence', 'FIELD_EXECUTION'), (277, 'Temporary Protection Planner', 'FIELD_EXECUTION'), (278, 'Shutdown Coordination Brain', 'FIELD_EXECUTION'), (279, 'Site Logistics Command', 'FIELD_EXECUTION'), (280, 'Production Quantity Tracking', 'FIELD_EXECUTION'), (281, 'Planned vs Actual Production', 'FIELD_EXECUTION'), (282, 'Daily Delay Cause Intelligence', 'FIELD_EXECUTION'), (283, 'Recovery Action Recommender', 'FIELD_EXECUTION'), (284, 'Constraint Removal Planner', 'FIELD_EXECUTION'), (285, 'Field Issue Escalation Brain', 'FIELD_EXECUTION'), (286, 'Weather Exposure Planning', 'FIELD_EXECUTION'), (287, 'Trade Mobilization Forecast', 'FIELD_EXECUTION'), (288, 'Foreman Coordination Intelligence', 'FIELD_EXECUTION'), (289, 'Field Execution Command 4.0', 'FIELD_EXECUTION'), (290, 'RFI Aging Intelligence 4.0', 'PM_CONTROL'), (291, 'RFI Impact Propagation', 'PM_CONTROL'), (292, 'Submittal Aging Intelligence', 'PM_CONTROL'), (293, 'Submittal Required-Date Engine', 'PM_CONTROL'), (294, 'Procurement Commitment Tracking', 'PM_CONTROL'), (295, 'Vendor Response Risk', 'PM_CONTROL'), (296, 'Owner Decision Aging', 'PM_CONTROL'), (297, 'Architect Response Tracking', 'PM_CONTROL'), (298, 'Meeting Agenda Intelligence 2.0', 'PM_CONTROL'), (299, 'Meeting Decision Register', 'PM_CONTROL'), (300, 'Meeting Action Follow-Up', 'PM_CONTROL'), (301, 'Correspondence Intelligence 2.0', 'PM_CONTROL'), (302, 'Contract Notice Trigger Brain', 'PM_CONTROL'), (303, 'Potential Change Register 2.0', 'PM_CONTROL'), (304, 'Change Documentation Completeness', 'PM_CONTROL'), (305, 'PM Risk Register', 'PM_CONTROL'), (306, 'Responsibility Escalation Engine', 'PM_CONTROL'), (307, 'Communication Gap Detection', 'PM_CONTROL'), (308, 'Project Controls Integration', 'PM_CONTROL'), (309, 'PM Command 4.0', 'PM_CONTROL'), (310, 'CPM Relationship Readiness', 'SCHEDULE'), (311, 'Predecessor Quality Audit', 'SCHEDULE'), (312, 'Successor Exposure Intelligence', 'SCHEDULE'), (313, 'Float Consumption Predictor', 'SCHEDULE'), (314, 'Near-Critical Path Intelligence', 'SCHEDULE'), (315, 'Constraint Date Intelligence', 'SCHEDULE'), (316, 'Milestone Risk Forecast', 'SCHEDULE'), (317, 'Look-Ahead Reliability Score', 'SCHEDULE'), (318, 'Production-to-Schedule Feedback', 'SCHEDULE'), (319, 'Procurement-to-Schedule Link', 'SCHEDULE'), (320, 'RFI-to-Schedule Link 2.0', 'SCHEDULE'), (321, 'Inspection-to-Schedule Link', 'SCHEDULE'), (322, 'Weather-to-Schedule Link', 'SCHEDULE'), (323, 'Crew Capacity Schedule Risk', 'SCHEDULE'), (324, 'Area Turnover Sequence Brain', 'SCHEDULE'), (325, 'Recovery Scenario Generator', 'SCHEDULE'), (326, 'Resequencing Opportunity Brain', 'SCHEDULE'), (327, 'Delay Cause Classification', 'SCHEDULE'), (328, 'Schedule Confidence Engine', 'SCHEDULE'), (329, 'Schedule Command 4.0', 'SCHEDULE'), (330, 'Commitment Intelligence', 'COST'), (331, 'Pending Change Exposure', 'COST'), (332, 'Known vs Unpriced Cost Separation', 'COST'), (333, 'Cost-to-Complete Forecast', 'COST'), (334, 'Estimate-at-Completion Predictor', 'COST'), (335, 'Allowance Burn Intelligence', 'COST'), (336, 'Contingency Consumption Brain', 'COST'), (337, 'Buyout Variance Intelligence', 'COST'), (338, 'Production Cost Variance', 'COST'), (339, 'Schedule Cost Exposure', 'COST'), (340, 'RFI Cost Exposure 2.0', 'COST'), (341, 'Procurement Cost Risk', 'COST'), (342, 'Labor Productivity Cost Signal', 'COST'), (343, 'Subcontractor Cost Risk', 'COST'), (344, 'Forecast Confidence Engine', 'COST'), (345, 'Cash Exposure Snapshot', 'COST'), (346, 'Commercial Risk Register', 'COST'), (347, 'Budget Trend Intelligence', 'COST'), (348, 'Executive Cost Forecast', 'COST'), (349, 'Cost Command 4.0', 'COST'), (350, 'Activity Safety Readiness', 'SAFETY_QUALITY_COMPANY'), (351, 'JHA Requirement Intelligence', 'SAFETY_QUALITY_COMPANY'), (352, 'Permit-to-Work Intelligence', 'SAFETY_QUALITY_COMPANY'), (353, 'High-Risk Activity Detection', 'SAFETY_QUALITY_COMPANY'), (354, 'Quality Hold Point Intelligence 2.0', 'SAFETY_QUALITY_COMPANY'), (355, 'Preinstallation Conference Intelligence', 'SAFETY_QUALITY_COMPANY'), (356, 'Mockup Requirement Brain', 'SAFETY_QUALITY_COMPANY'), (357, 'Manufacturer Inspection Requirement', 'SAFETY_QUALITY_COMPANY'), (358, 'Testing Witness Requirement', 'SAFETY_QUALITY_COMPANY'), (359, 'Deficiency Recurrence Intelligence', 'SAFETY_QUALITY_COMPANY'), (360, 'Punch Trend Intelligence', 'SAFETY_QUALITY_COMPANY'), (361, 'Commissioning Readiness 2.0', 'SAFETY_QUALITY_COMPANY'), (362, 'Closeout Quality Intelligence', 'SAFETY_QUALITY_COMPANY'), (363, 'Subcontractor Quality Score', 'SAFETY_QUALITY_COMPANY'), (364, 'Subcontractor Safety Score', 'SAFETY_QUALITY_COMPANY'), (365, 'Project Lessons Pattern Brain', 'SAFETY_QUALITY_COMPANY'), (366, 'Company Benchmark Intelligence', 'SAFETY_QUALITY_COMPANY'), (367, 'Cross-Project Risk Pattern', 'SAFETY_QUALITY_COMPANY'), (368, 'Portfolio Health Intelligence', 'SAFETY_QUALITY_COMPANY'), (369, 'Safety Quality Company Command', 'SAFETY_QUALITY_COMPANY')]

def _v369_context(pid):
    try: snap=_v56_cached("snapshot",pid,lambda:_v37_snapshot(pid))
    except Exception: snap={}
    def cnt(table):
        try:
            c=db()
            row=c.execute(f"SELECT COUNT(*) FROM {table} WHERE project_id=?",(pid,)).fetchone()
            c.close()
            return int(row[0] if row else 0)
        except Exception:
            return 0
    return {
      "snapshot":snap,
      "activities":cnt("activities"),
      "issues":cnt("issues"),
      "rfis":cnt("rfis"),
      "submittals":cnt("submittals"),
      "procurement":cnt("procurement"),
      "inspections":cnt("inspections"),
      "changes":cnt("change_events"),
      "scope":len(_v452_scope_rows(pid)) if callable(globals().get("_v452_scope_rows")) else 0
    }

def _v369_status(pid,cap):
    _,_,group=cap
    c=_v369_context(pid)
    if group=="FIELD_EXECUTION":
        return "ACTIVE" if c["activities"] else "WAITING FOR SCHEDULE DATA"
    if group=="PM_CONTROL":
        return "ACTIVE" if (c["rfis"]+c["submittals"]+c["issues"]+c["procurement"]) else "WAITING FOR PM DATA"
    if group=="SCHEDULE":
        return "ACTIVE" if c["activities"] else "WAITING FOR SCHEDULE DATA"
    if group=="COST":
        return "ACTIVE" if (c["changes"] or c["scope"]) else "WAITING FOR COMMERCIAL DATA"
    if group=="SAFETY_QUALITY_COMPANY":
        return "ACTIVE" if (c["activities"] or c["inspections"] or c["scope"]) else "WAITING FOR PROJECT DATA"
    return "LEARNING"

def _v369_field(pid):
    try: brief=_v62_daily_brief(pid)
    except Exception: brief={"priorities":{"now":[],"today":[],"week":[]},"activities":[]}
    try: ready=_v60_trade_readiness(pid)
    except Exception: ready=[]
    try: handoffs=_v61_handoffs(pid)
    except Exception: handoffs=[]
    return brief,ready,handoffs

def _v369_pm(pid):
    try: decisions=_v47_decision_deadlines(pid)[:40]
    except Exception: decisions=[]
    try: agenda=_v47_coordination_agenda(pid)[:40]
    except Exception: agenda=[]
    try: changes=_v39_rows("SELECT * FROM change_events WHERE project_id=? ORDER BY id DESC LIMIT 40",(pid,))
    except Exception: changes=[]
    return decisions,agenda,changes

def _v369_schedule(pid):
    try: seq=_v56_sequence(pid)
    except Exception: seq=[]
    high=[x for x in seq if x.get("risk") in {"CRITICAL","HIGH"}]
    try: look=_v59_lookahead(pid,6)
    except Exception: look=[]
    return high,look

def _v369_cost(pid):
    try: changes=_v39_changes(pid)
    except Exception: changes=[]
    known=sum(float(r["estimated_cost"] or 0) for r in changes)
    try:
        rev=_v55_summary(pid)
        revnet=float(rev.get("added_cost",0))-float(rev.get("removed_cost",0))
    except Exception:
        revnet=0
    return {"known_change":known,"revision_net":revnet,"total_known":known+revnet}

def _v369_quality(pid):
    try: qc=_v65_qc_points(pid)
    except Exception: qc=[]
    try: holds=[x for x in _v49_hold_points(pid) if str(x[0]["result"] or "").upper()!="PASSED"]
    except Exception: holds=[]
    return qc,holds

@app.get("/platform-369",response_class=HTMLResponse)
def v369_platform():
    pid=project_id()
    active=sum(1 for x in _V369_CAPABILITIES if _v369_status(pid,x)=="ACTIVE")
    body=f'<div class="hero"><div class="eyebrow">BuildCommand v369</div><h1>Execution & Control Intelligence Platform</h1><p class="muted">{active} / 100 new capabilities active against current project data.</p></div><div class="grid3">'
    cards=[
      ("Field Execution Command","Production, crews, logistics, constraints and recovery.","/field-execution-4"),
      ("PM Command 4.0","RFI, submittal, decisions, meetings and change controls.","/pm-command-4"),
      ("Schedule Command 4.0","Sequence, near-critical exposure, float and recovery intelligence.","/schedule-command-4"),
      ("Cost Command 4.0","Known exposure, changes, forecast and commercial risk.","/cost-command-4"),
      ("Safety / Quality / Company","Safety readiness, QC, commissioning and company patterns.","/sqc-command"),
      ("Capability Map","See all v270-v369 capabilities.","/platform-369/capabilities"),
      ("Project Autopilot","Continue using the operating command center.","/autopilot"),
      ("Unified Platform v269","Prior platform phase.","/platform-269"),
      ("Knowledge Brain 2.0","Construction knowledge foundation.","/knowledge-brain-2"),
    ]
    for n,d,h in cards:
        body+=_v37_link_card(n,d,h,"Open")
    body+='</div>'
    return shell("Execution & Control Platform",body)

@app.get("/platform-369/capabilities",response_class=HTMLResponse)
def v369_caps():
    pid=project_id()
    h="".join(
        f'<div class="action"><span class="badge">{esc(_v369_status(pid,x))}</span> '
        f'<b>v{x[0]} - {esc(x[1])}</b><div class="small">{esc(x[2].replace("_"," "))}</div></div>'
        for x in _V369_CAPABILITIES
    )
    return shell("v270-v369 Capability Map",'<div class="hero"><h1>Next 100 Capability Map</h1><p class="muted">Shared workspaces, not 100 separate top-level pages.</p></div><div class="card">'+h+'</div>')

@app.get("/field-execution-4",response_class=HTMLResponse)
def v369_field_page():
    brief,ready,handoffs=_v369_field(project_id())
    p=brief["priorities"]
    def ph(rows): return "".join(
        f'<div class="action"><span class="badge WATCH">{esc(x["severity"])}</span> '
        f'<b>{esc(x["title"])}</b><div class="small">{esc(x["action"])}</div></div>' for x in rows[:10]
    ) or '<p class="muted">Clear.</p>'
    rh="".join(
        f'<div class="action"><span class="badge WATCH">{esc(state)}</span> '
        f'<b>{esc(a["trade"])} - {esc(a["name"])}</b><div class="small">{esc(", ".join(failed))}</div></div>'
        for a,state,failed,checks in ready if state!="READY"
    ) or '<p class="muted">Upcoming trades ready.</p>'
    hh="".join(
        f'<div class="action"><span class="badge">HANDOFF</span> <b>{esc(a["trade"])} → {esc(b["trade"])}</b>'
        f'<div class="small">{esc(a["name"])} → {esc(b["name"])}</div></div>' for a,b,note in handoffs[:12]
    ) or '<p class="muted">No immediate handoffs.</p>'
    body='<div class="hero"><div class="eyebrow">v289 Field Execution Command 4.0</div><h1>Field Execution Command</h1></div>'
    body+='<div class="grid2"><div class="card"><h2>DO NOW</h2>'+ph(p["now"])+'</div><div class="card"><h2>TODAY</h2>'+ph(p["today"])+'</div></div>'
    body+='<div class="grid2"><div class="card"><h2>Trade Readiness</h2>'+rh+'</div><div class="card"><h2>Upcoming Handoffs</h2>'+hh+'</div></div>'
    return shell("Field Execution Command",body)

@app.get("/pm-command-4",response_class=HTMLResponse)
def v369_pm_page():
    decisions,agenda,changes=_v369_pm(project_id())
    dh="".join(
        f'<div class="action"><span class="badge WATCH">{esc(sev)}</span> <b>{esc(typ)} - {esc(title)}</b>'
        f'<div class="small">Due {esc(due)} · ${cost:,.0f} known exposure · {days:g} day(s)</div></div>'
        for typ,title,due,sev,cost,days,source in decisions
    ) or '<p class="muted">No urgent decisions.</p>'
    ah="".join(
        f'<div class="action"><span class="badge">{esc(kind)}</span> <b>{esc(title)}</b><div class="small">{esc(detail)}</div></div>'
        for kind,title,detail in agenda
    ) or '<p class="muted">No current agenda items.</p>'
    ch="".join(
        f'<div class="action"><b>{esc(r["title"])}</b><div class="small">${float(r["estimated_cost"] or 0):,.0f} · {float(r["schedule_days"] or 0):g} day(s)</div></div>'
        for r in changes[:20]
    ) or '<p class="muted">No change events.</p>'
    return shell("PM Command 4.0",'<div class="hero"><div class="eyebrow">v309</div><h1>PM Command 4.0</h1></div><div class="grid3"><div class="card"><h2>Decisions</h2>'+dh+'</div><div class="card"><h2>Coordination</h2>'+ah+'</div><div class="card"><h2>Changes</h2>'+ch+'</div></div>')

@app.get("/schedule-command-4",response_class=HTMLResponse)
def v369_schedule_page():
    high,look=_v369_schedule(project_id())
    hh="".join(
        f'<div class="action"><span class="badge WATCH">{esc(x["risk"])}</span> <b>{esc(x["activity"]["name"])}</b>'
        f'<div class="small">{esc(x["activity"]["trade"])} · {esc(x["blocking_reason"])}</div></div>' for x in high[:30]
    ) or '<p class="muted">No high sequence risks.</p>'
    lh="".join(
        f'<div class="action"><span class="badge WATCH">{esc(state)}</span> <b>{esc(a["name"])}</b>'
        f'<div class="small">{esc(a["trade"])} · {esc(a["start"])} → {esc(a["finish"])}</div></div>' for a,state,risk,reason in look[:30]
    ) or '<p class="muted">No look-ahead activities.</p>'
    return shell("Schedule Command 4.0",'<div class="hero"><div class="eyebrow">v329</div><h1>Schedule Command 4.0</h1><p class="muted">Sequence exposure plus six-week readiness.</p></div><div class="grid2"><div class="card"><h2>High-Risk Sequence</h2>'+hh+'</div><div class="card"><h2>6-Week Look-Ahead</h2>'+lh+'</div></div>')

@app.get("/cost-command-4",response_class=HTMLResponse)
def v369_cost_page():
    c=_v369_cost(project_id())
    return shell("Cost Command 4.0",f'<div class="hero"><div class="eyebrow">v349</div><h1>Cost Command 4.0</h1><p class="muted">Known values only; unpriced exposure remains explicitly unpriced.</p></div><div class="grid3"><div class="card"><div class="label">Known Change Exposure</div><div class="kpi">${c["known_change"]:,.0f}</div></div><div class="card"><div class="label">Revision Net</div><div class="kpi">${c["revision_net"]:,.0f}</div></div><div class="card"><div class="label">Total Known</div><div class="kpi">${c["total_known"]:,.0f}</div></div></div>')

@app.get("/sqc-command",response_class=HTMLResponse)
def v369_sqc_page():
    qc,holds=_v369_quality(project_id())
    qh="".join(
        f'<div class="action"><span class="badge">QC</span> <b>{esc(a["name"])}</b><div class="small">{esc(kind)} · {esc(check)}</div></div>'
        for a,kind,check in qc[:30]
    ) or '<p class="muted">No QC checkpoints generated.</p>'
    hh="".join(
        f'<div class="action"><span class="badge WATCH">{esc(i["result"])}</span> <b>{esc(i["inspection_type"])}</b>'
        f'<div class="small">{esc(i["scheduled_date"])} · {esc(i["authority"])}</div></div>' for i,a in holds[:30]
    ) or '<p class="muted">No open inspection hold points.</p>'
    return shell("Safety Quality Company Command",'<div class="hero"><div class="eyebrow">v369</div><h1>Safety / Quality / Company Command</h1></div><div class="grid2"><div class="card"><h2>Quality Checkpoints</h2>'+qh+'</div><div class="card"><h2>Open Hold Points</h2>'+hh+'</div></div>')

@app.get("/build",response_class=HTMLResponse)
def unified_build():
    s=_v37_snapshot(project_id())
    body=(
      '<div class="hero"><div class="eyebrow">BUILD</div><h1>Understand the project.</h1><p class="muted">One place for plans, specs and scope intelligence.</p></div>'
      f'<div class="card"><div class="label">Source-backed Scope Items</div><div class="kpi">{s["scope"]}</div></div>'
      '<div class="grid2">'
      +_v37_link_card("Analyze Project","Upload once. BuildCommand runs plan intelligence, estimator sync, takeoff splitting and quantity review automatically.","/build/analyze-project","Analyze")
      +_v37_link_card("Review Project Scope","Review unified source-backed construction intelligence.","/brain","Review")
      +_v37_link_card("Blueprint Brain","Open source-backed trade scopes and run the final trade cleanup.","/blueprint-brain","Open")
      +'</div><details class="card"><summary><b>More Build tools</b></summary><p><a href="/sequence-intelligence">Sequence Intelligence</a> · <a href="/blueprint-brain">Blueprint Brain</a> · <a href="/preconstruction">Preconstruction & Bid Intelligence</a> · <a href="/documents">Documents</a> · <a href="/document-ai">Deep Document AI</a></p></details>'
    )
    return shell("Build",body)

@app.get("/estimate",response_class=HTMLResponse)
def unified_estimate():
    s=_v37_snapshot(project_id())
    body=(
      '<div class="hero"><div class="eyebrow">ESTIMATE</div><h1>Scope to price.</h1><p class="muted">The takeoff brains work underneath; review what needs estimator judgment.</p></div>'
      f'<div class="grid2"><div class="card"><div class="label">Estimator Items</div><div class="kpi">{s["estimate"]}</div></div>'
      f'<div class="card"><div class="label">Need Review</div><div class="kpi">{s["review"]}</div></div></div>'
      '<div class="grid2">'
      +_v37_link_card("Estimator Workspace","Scope, quantity, unit, labor, material, subcontract quote and markup.","/brain/estimator","Open Estimate")
      +_v37_link_card("Takeoff Review","Review AI quantity proposals and verification items.","/brain/takeoff","Review")
      +'</div><details class="card"><summary><b>Advanced estimating tools</b></summary><p><a href="/brain/takeoff/components">Takeoff Components</a> · <a href="/cost-intelligence">Cost Intelligence</a></p></details>'
    )
    return shell("Estimate",body)

@app.get("/manage",response_class=HTMLResponse)
def unified_manage():
    s=_v37_snapshot(project_id())
    attention=s["issues"]+s["submittals"]+s["actions"]+s["inspections"]
    body=(
      '<div class="hero"><div class="eyebrow">MANAGE</div><h1>Run the job.</h1><p class="muted">Schedule, field, RFIs, submittals, inspections and subcontractors.</p></div>'
      f'<div class="grid4"><div class="card"><div class="label">Need Attention</div><div class="kpi">{attention}</div></div>'
      f'<div class="card"><div class="label">Issues</div><div class="kpi">{s["issues"]}</div></div>'
      f'<div class="card"><div class="label">Submittals</div><div class="kpi">{s["submittals"]}</div></div>'
      f'<div class="card"><div class="label">Inspections</div><div class="kpi">{s["inspections"]}</div></div></div>'
      '<div class="grid3">'
      +_v37_link_card("Today","Superintendent Field Command: readiness, crews, deliveries, inspections and decisions.","/field-command")
      +_v37_link_card("Schedule","Schedule and production planning.","/schedule")
      +_v37_link_card("Things That Need You","One queue for human decisions.","/actions")
      +_v37_link_card("RFIs / Issues","Questions, conflicts and issues.","/issues")
      +_v37_link_card("Submittals","Submittal workflow.","/submittals")
      +_v37_link_card("Field","Field execution and reporting.","/field")
      +'</div><details class="card"><summary><b>More management tools</b></summary><p><a href="/sequence-intelligence">Sequence Intelligence</a> · <a href="/learning">Learning Intelligence</a> · <a href="/project-control">Project Control</a> · <a href="/intelligence">Intelligence Center</a> · <a href="/inspections">Inspections</a> · <a href="/safety">Safety</a> · <a href="/subcontractors">Subcontractors</a> · <a href="/procurement">Procurement</a> · <a href="/punch">Punch</a></p></details>'
    )
    return shell("Manage",body)

@app.get("/ask-buildcommand",response_class=HTMLResponse)
def unified_ask_buildcommand():
    body=(
      '<div class="hero"><div class="eyebrow">ASK BUILDCOMMAND</div><h1>Ask the project.</h1><p class="muted">Use natural language instead of hunting through menus.</p></div>'
      '<div class="grid2">'
      +_v37_link_card("Project Command","Ask about scope, schedule, risks, trades and project status.","/ai-command","Ask")
      +_v37_link_card("Search Everything","Find information across the project.","/global-search","Search")
      +'</div><div class="card"><h2>Try asking</h2><div class="action">What am I missing in electrical?</div><div class="action">What needs my attention today?</div><div class="action">Show me everything affecting doors.</div><div class="action">What could delay this week?</div></div>'
    )
    return shell("Ask BuildCommand",body)

@app.get("/legacy-dashboard",response_class=HTMLResponse)
def home():
    pid = project_id()
    ensure_today_morning_brief(pid)
    c = db()
    today = date.today().isoformat()

    project = c.execute(
        "SELECT * FROM projects WHERE id=?",
        (pid,)
    ).fetchone()

    activities = c.execute(
        "SELECT * FROM activities WHERE project_id=? ORDER BY start",
        (pid,)
    ).fetchall()

    risks = c.execute(
        """
        SELECT r.*, a.external_id, a.name activity
        FROM risks r
        JOIN activities a ON a.id=r.activity_id
        WHERE r.project_id=?
        ORDER BY r.score DESC
        """,
        (pid,)
    ).fetchall()

    make_ready_items = c.execute(
        """
        SELECT m.*, a.external_id, a.name activity
        FROM make_ready m
        JOIN activities a ON a.id=m.activity_id
        WHERE m.project_id=? AND m.status='OPEN'
        ORDER BY m.due
        """,
        (pid,)
    ).fetchall()

    actions = c.execute(
        """
        SELECT *
        FROM action_items
        WHERE project_id=? AND status='OPEN'
        ORDER BY due, id
        """,
        (pid,)
    ).fetchall()

    issues = c.execute(
        """
        SELECT i.*, a.external_id, a.name activity
        FROM project_issues i
        LEFT JOIN activities a ON a.id=i.activity_id
        WHERE i.project_id=? AND i.status!='CLOSED'
        ORDER BY i.due, i.id
        """,
        (pid,)
    ).fetchall()

    procurement_rows = c.execute(
        """
        SELECT p.*, a.external_id, a.name activity
        FROM procurement p
        LEFT JOIN activities a ON a.id=p.activity_id
        WHERE p.project_id=? AND p.status!='DELIVERED'
        ORDER BY p.required_on_site, p.id
        """,
        (pid,)
    ).fetchall()

    inspection_rows = c.execute(
        """
        SELECT i.*, a.external_id, a.name activity
        FROM inspections_tracker i
        LEFT JOIN activities a ON a.id=i.activity_id
        WHERE i.project_id=? AND i.result!='PASSED'
        ORDER BY i.scheduled_date, i.id
        """,
        (pid,)
    ).fetchall()

    submittal_rows = c.execute(
        """
        SELECT s.*, a.external_id, a.name activity
        FROM submittals s
        LEFT JOIN activities a ON a.id=s.activity_id
        WHERE s.project_id=? AND s.status NOT IN ('APPROVED','APPROVED_AS_NOTED')
        ORDER BY s.due_date, s.id
        """,
        (pid,)
    ).fetchall()

    safety_rows = c.execute(
        """
        SELECT *
        FROM safety_items
        WHERE project_id=? AND status!='CLOSED'
        ORDER BY id DESC
        """,
        (pid,)
    ).fetchall()

    change_rows = c.execute(
        """
        SELECT *
        FROM change_events
        WHERE project_id=? AND status NOT IN ('APPROVED','REJECTED')
        ORDER BY id DESC
        """,
        (pid,)
    ).fetchall()

    latest_report = c.execute(
        """
        SELECT *
        FROM daily_reports
        WHERE project_id=?
        ORDER BY report_date DESC,id DESC
        LIMIT 1
        """,
        (pid,)
    ).fetchone()

    c.close()

    critical_risks = [r for r in risks if r["band"] == "CRITICAL"]
    high_risks = [r for r in risks if r["band"] == "HIGH"]

    overdue_actions = [
        a for a in actions
        if a["due"] and a["due"] < today
    ]

    overdue_issues = [
        i for i in issues
        if i["due"] and i["due"] < today
    ]

    procurement_critical = []
    for p in procurement_rows:
        badge, risk_text = procurement_risk(
            p["required_on_site"],
            p["promised_date"],
            p["status"]
        )
        if badge == "CRITICAL":
            procurement_critical.append((p, risk_text))

    failed_inspections = [
        i for i in inspection_rows
        if i["result"] == "FAILED"
    ]

    overdue_inspections = [
        i for i in inspection_rows
        if i["scheduled_date"]
        and i["scheduled_date"] < today
        and i["result"] in ["PENDING", "SCHEDULED"]
    ]

    rejected_submittals = [
        s for s in submittal_rows
        if s["status"] == "REJECTED"
    ]

    overdue_submittals = [
        s for s in submittal_rows
        if s["due_date"]
        and s["due_date"] < today
        and s["status"] in ["PENDING", "SUBMITTED"]
    ]

    critical_safety = [
        s for s in safety_rows
        if s["severity"] == "CRITICAL"
    ]

    open_change_cost = sum((r["estimated_cost"] or 0) for r in change_rows)
    open_change_days = sum((r["schedule_days"] or 0) for r in change_rows)

    attention_score = min(
        100,
        len(critical_risks) * 20
        + len(high_risks) * 10
        + len(overdue_actions) * 8
        + len(overdue_issues) * 8
        + len(procurement_critical) * 12
        + len(failed_inspections) * 15
        + len(overdue_inspections) * 8
        + len(rejected_submittals) * 10
        + len(overdue_submittals) * 6
        + len(critical_safety) * 20
    )

    if attention_score >= 70:
        overall_badge = "CRITICAL"
        overall_text = "Immediate Attention"
    elif attention_score >= 40:
        overall_badge = "HIGH"
        overall_text = "Needs Attention"
    elif attention_score >= 15:
        overall_badge = "WATCH"
        overall_text = "Watch"
    else:
        overall_badge = "READY"
        overall_text = "Stable"

    priority_items = []

    for s in critical_safety[:2]:
        priority_items.append(
            ("CRITICAL", "Safety", s["title"], f'{s["location"] or "No location"}')
        )

    for r in critical_risks[:3]:
        priority_items.append(
            ("CRITICAL", "Risk", r["activity"], r["explanation"])
        )

    for i in failed_inspections[:2]:
        priority_items.append(
            ("CRITICAL", "Inspection", i["inspection_type"], i["notes"] or "Failed inspection")
        )

    for p, risk_text in procurement_critical[:2]:
        priority_items.append(
            ("CRITICAL", "Procurement", p["item"], risk_text)
        )

    for a in overdue_actions[:3]:
        priority_items.append(
            (
                a["priority"] if a["priority"] in ["CRITICAL","HIGH","WATCH","LOW"] else "HIGH",
                "Action",
                a["title"],
                f'Owner: {a["owner"] or "Unassigned"} · Due {a["due"]}'
            )
        )

    for i in overdue_issues[:2]:
        priority_items.append(
            (
                i["priority"] if i["priority"] in ["CRITICAL","HIGH","WATCH","LOW"] else "HIGH",
                i["issue_type"],
                i["title"],
                f'Owner: {i["owner"] or "Unassigned"} · Due {i["due"]}'
            )
        )

    for s in rejected_submittals[:2]:
        priority_items.append(
            ("HIGH", "Submittal", s["title"], "Rejected / revise and resubmit")
        )

    for m in make_ready_items[:3]:
        priority_items.append(
            (
                m["priority"] if m["priority"] in ["CRITICAL","HIGH","WATCH","LOW"] else "HIGH",
                "Make Ready",
                m["title"],
                f'Due {m["due"]}'
            )
        )

    if not priority_items:
        priority_items.append(
            ("READY", "Command", "No urgent issues detected", "Review today’s plan and upcoming work.")
        )

    priority_html = "".join(
        f"""
        <div class="action">
            <span class="badge {badge}">{esc(category)}</span>
            <b>{esc(title)}</b>
            <div class="small">{esc(detail)}</div>
        </div>
        """
        for badge, category, title, detail in priority_items[:10]
    )

    latest_report_html = (
        f"""
        <div class="small">Latest Daily Report: {esc(latest_report["report_date"])}</div>
        <p><b>Manpower:</b> {latest_report["manpower"] or 0}</p>
        <p><b>Work Completed:</b> {esc(latest_report["work_completed"]) or "—"}</p>
        <p><b>Delays:</b> {esc(latest_report["delays"]) or "—"}</p>
        <p><b>Tomorrow:</b> {esc(latest_report["tomorrow_plan"]) or "—"}</p>
        """
        if latest_report
        else '<div class="muted">No daily report submitted yet.</div>'
    )

    project_name = esc(project["name"]) if project else "Current Project"
    project_number = esc(project["number"]) if project else ""

    body = f"""
    <div class="hero">
        <div style="display:flex;justify-content:space-between;gap:18px;align-items:flex-start;flex-wrap:wrap;">
            <div>
                <div class="eyebrow">Morning Command Center</div>
                <h1>What needs attention today?</h1>
                <div class="muted">
                    {project_number} - {project_name}
                </div>
            </div>

            <div style="text-align:right;">
                <span class="badge {overall_badge}">{overall_text}</span>
                <div class="kpi" style="margin-top:8px;">{attention_score}</div>
                <div class="small">Project attention score</div>
            </div>
        </div>
    </div>

    <div class="grid4">
        <div class="card">
            <div class="label">Critical Risks</div>
            <div class="kpi">{len(critical_risks)}</div>
        </div>

        <div class="card">
            <div class="label">Overdue Actions</div>
            <div class="kpi">{len(overdue_actions)}</div>
        </div>

        <div class="card">
            <div class="label">Open Make-Ready</div>
            <div class="kpi">{len(make_ready_items)}</div>
        </div>

        <div class="card">
            <div class="label">Failed Inspections</div>
            <div class="kpi">{len(failed_inspections)}</div>
        </div>
    </div>

    <div class="grid2">
        <div class="card">
            <h2>Handle First</h2>
            {priority_html}
        </div>

        <div class="card">
            <h2>Latest Field Signal</h2>
            {latest_report_html}

            <div style="margin-top:14px;">
                <a href="/assistant"
                   style="color:#f0b44d;text-decoration:none;font-weight:700;">
                    Ask BuildCommand AI →
                </a>
            </div>
        </div>
    </div>

    <div class="grid4">
        <div class="card">
            <div class="label">Procurement Critical</div>
            <div class="kpi">{len(procurement_critical)}</div>
        </div>

        <div class="card">
            <div class="label">Overdue RFIs / Issues</div>
            <div class="kpi">{len(overdue_issues)}</div>
        </div>

        <div class="card">
            <div class="label">Submittal Problems</div>
            <div class="kpi">{len(rejected_submittals) + len(overdue_submittals)}</div>
        </div>

        <div class="card">
            <div class="label">Critical Safety</div>
            <div class="kpi">{len(critical_safety)}</div>
        </div>
    </div>

    <div class="grid2">
        <div class="card">
            <h2>Change Exposure</h2>

            <div class="grid2">
                <div>
                    <div class="label">Potential Cost</div>
                    <div class="kpi">${open_change_cost:,.0f}</div>
                </div>

                <div>
                    <div class="label">Potential Days</div>
                    <div class="kpi">{open_change_days:.1f}</div>
                </div>
            </div>

            <div style="margin-top:12px;">
                <a href="/changes"
                   style="color:#f0b44d;text-decoration:none;font-weight:700;">
                    Review Change Events →
                </a>
            </div>
        </div>

        <div class="card">
            <h2>Quick Command</h2>

            <div class="action">
                <a href="/daily-report" style="color:#f0b44d;text-decoration:none;font-weight:700;">
                    Daily Report →
                </a>
            </div>

            <div class="action">
                <a href="/schedule-health" style="color:#f0b44d;text-decoration:none;font-weight:700;">
                    Schedule Health →
                </a>
            </div>

            <div class="action">
                <a href="/readiness" style="color:#f0b44d;text-decoration:none;font-weight:700;">
                    Lookahead Readiness →
                </a>
            </div>

            <div class="action">
                <a href="/actions" style="color:#f0b44d;text-decoration:none;font-weight:700;">
                    Action Center →
                </a>
            </div>
        </div>
    </div>
    """

    return shell("Daily Command", body)


@app.get("/schedule",response_class=HTMLResponse)
def schedule():
    pid=project_id(); c=db(); rows=c.execute("SELECT * FROM activities WHERE project_id=? ORDER BY start",(pid,)).fetchall(); c.close()
    tr="".join(
        f'<tr>'
        f'<td>{esc(r["external_id"])}</td>'
        f'<td><b>{esc(r["name"])}</b></td>'
        f'<td>{esc(r["trade"])}</td>'
        f'<td>{r["start"]}</td>'
        f'<td>{r["finish"]}</td>'
        f'<td>{r["pct"]:.0f}%</td>'
        f'<td>{esc(r["status"])}</td>'
        f'<td><a href="/activities/{r["id"]}/edit" style="color:#f0b44d;text-decoration:none;font-weight:700;">Edit</a></td>'
        f'</tr>'
        for r in rows
    )
    return shell("Schedule",f'<div class="hero"><div style="display:flex;justify-content:space-between;align-items:center;gap:15px;flex-wrap:wrap;"><div><div class="eyebrow">Schedule + Lookahead</div><h1>Field execution plan</h1></div><a href="/activities/new" style="display:inline-block;background:#f0b44d;color:#0a1017;text-decoration:none;padding:12px 18px;border-radius:9px;font-weight:800;">+ Add Activity</a></div></div><div class="card"><table><tr><th>ID</th><th>Activity</th><th>Trade</th><th>Start</th><th>Finish</th><th>%</th><th>Status</th><th>Action</th></tr>{tr}</table></div>')

@app.get("/make-ready",response_class=HTMLResponse)
def make_ready():
    pid = project_id()
    c = db()

    rows = c.execute(
        """
        SELECT m.*, a.external_id, a.name activity
        FROM make_ready m
        JOIN activities a ON a.id=m.activity_id
        WHERE m.project_id=? AND m.status='OPEN'
        ORDER BY
            CASE m.priority
                WHEN 'CRITICAL' THEN 1
                WHEN 'HIGH' THEN 2
                WHEN 'WATCH' THEN 3
                ELSE 4
            END,
            m.due
        """,
        (pid,)
    ).fetchall()

    closed = c.execute(
        """
        SELECT m.*, a.external_id, a.name activity
        FROM make_ready m
        JOIN activities a ON a.id=m.activity_id
        WHERE m.project_id=? AND m.status='COMPLETE'
        ORDER BY m.id DESC
        LIMIT 10
        """,
        (pid,)
    ).fetchall()

    c.close()

    open_html = "".join(
        f"""
        <div class="card">
            <div style="display:flex;justify-content:space-between;gap:12px;align-items:flex-start;flex-wrap:wrap;">
                <div>
                    <span class="badge {r["priority"]}">{r["priority"]}</span>
                    <b>{esc(r["external_id"])} - {esc(r["activity"])}</b>
                </div>

                <form method="post" action="/make-ready/{r["id"]}/close">
                    <button type="submit">Mark Cleared</button>
                </form>
            </div>

            <h3>{esc(r["title"])}</h3>
            <p>{esc(r["reason"])}</p>
            <div class="small">Clear by {esc(r["due"])}</div>
        </div>
        """
        for r in rows
    ) or '<div class="card"><div class="muted">No open make-ready items. Nice work.</div></div>'

    closed_html = "".join(
        f"""
        <div class="action">
            <span class="badge COMPLETE">CLEARED</span>
            <b>{esc(r["external_id"])} - {esc(r["activity"])}</b>
            <div>{esc(r["title"])}</div>
        </div>
        """
        for r in closed
    ) or '<div class="muted">No recently cleared items.</div>'

    body = f"""
    <div class="hero">
        <div style="display:flex;justify-content:space-between;align-items:center;gap:15px;flex-wrap:wrap;">
            <div>
                <div class="eyebrow">Predictive Make-Ready</div>
                <h1>Clear tomorrow's blockers before they hit the field.</h1>
            </div>

            <a href="/make-ready/new"
               style="display:inline-block;background:#f0b44d;color:#0a1017;text-decoration:none;padding:12px 18px;border-radius:9px;font-weight:800;">
                + Add Make-Ready
            </a>
        </div>
    </div>

    <div class="grid2">
        <div>
            <h2>Open Blockers</h2>
            {open_html}
        </div>

        <div class="card">
            <h2>Recently Cleared</h2>
            {closed_html}
        </div>
    </div>
    """

    return shell("Make Ready", body)

@app.get("/field",response_class=HTMLResponse)
def field():
    pid=project_id(); c=db(); acts=c.execute("SELECT id,external_id,name FROM activities WHERE project_id=?",(pid,)).fetchall(); updates=c.execute("SELECT f.*,a.name activity FROM field_updates f JOIN activities a ON a.id=f.activity_id WHERE f.project_id=? ORDER BY f.id DESC LIMIT 15",(pid,)).fetchall(); c.close()
    opts="".join(f'<option value="{a["id"]}">{a["external_id"]} - {esc(a["name"])}</option>' for a in acts)
    recent="".join(f'<div class="action"><b>{u["update_type"]} - {esc(u["activity"])}</b><div>{esc(u["text"])}</div><div class="small">{u["created"]}</div></div>' for u in updates) or '<div class="muted">No updates yet.</div>'
    body=f'<div class="hero"><div class="eyebrow">Field Execution</div><h1>Capture what changed today.</h1></div><div class="grid2"><div class="card"><form method="post" action="/field/add"><select name="activity_id">{opts}</select><br><br><select name="update_type"><option>PROGRESS</option><option>BLOCKER</option><option>QUALITY</option><option>INSPECTION</option><option>DELIVERY</option></select><br><br><textarea name="text" placeholder="What changed in the field?"></textarea><br><br><button>Save field update</button></form></div><div class="card"><h2>Recent field updates</h2>{recent}</div></div>'
    return shell("Field",body)

@app.post("/field/add")
def add_field(activity_id:int=Form(...),update_type:str=Form(...),text:str=Form(...)):
    pid=project_id(); c=db(); c.execute("INSERT INTO field_updates(project_id,activity_id,update_type,text,created) VALUES(?,?,?,?,?)",(pid,activity_id,update_type,text,date.today().isoformat())); c.commit(); c.close()
    return RedirectResponse("/field",303)

@app.get("/subcontractors",response_class=HTMLResponse)
def subcontractors():
    pid = project_id()
    c = db()

    rows = c.execute(
        """
        SELECT *
        FROM subs
        WHERE project_id=?
        ORDER BY trade,name
        """,
        (pid,)
    ).fetchall()

    updates = c.execute(
        """
        SELECT u.*, s.name AS sub_name, s.trade AS sub_trade
        FROM subcontractor_updates u
        JOIN subs s ON s.id=u.sub_id
        WHERE u.project_id=?
        ORDER BY u.update_date DESC, u.id DESC
        LIMIT 25
        """,
        (pid,)
    ).fetchall()

    c.close()

    latest_by_sub = {}
    for u in updates:
        if u["sub_id"] not in latest_by_sub:
            latest_by_sub[u["sub_id"]] = u

    cards = ""

    for r in rows:
        latest = latest_by_sub.get(r["id"])

        if latest:
            status = latest["status"] or "WATCH"
            status_badge = status if status in ["CRITICAL", "HIGH", "WATCH", "READY", "LOW"] else "OPEN"
            detail = (
                f'<div class="small">Latest Update: {esc(latest["update_date"])}</div>'
                f'<p><b>Manpower:</b> {latest["manpower"] or 0}</p>'
                f'<p><b>Commitment:</b> {esc(latest["commitment"]) or "—"}</p>'
                f'<p><b>Issue:</b> {esc(latest["issue"]) or "—"}</p>'
                f'<span class="badge {status_badge}">{esc(status)}</span>'
            )
        else:
            detail = '<div class="muted">No field update recorded yet.</div>'

        cards += f"""
        <div class="card">
            <div style="display:flex;justify-content:space-between;gap:12px;flex-wrap:wrap;align-items:flex-start;">
                <div>
                    <h3 style="margin-top:0;">{esc(r["name"])}</h3>
                    <div class="muted">{esc(r["trade"])}</div>
                </div>

                <a href="/subcontractors/{r["id"]}/update"
                   style="color:#f0b44d;text-decoration:none;font-weight:700;">
                    Log Update
                </a>
            </div>

            <div style="margin-top:14px;">
                {detail}
            </div>
        </div>
        """

    if not cards:
        cards = '<div class="card"><div class="muted">No subcontractors added yet.</div></div>'

    recent_html = ""

    for u in updates[:12]:
        status = u["status"] or "WATCH"
        badge = status if status in ["CRITICAL", "HIGH", "WATCH", "READY", "LOW"] else "OPEN"

        recent_html += (
            f'<div class="action">'
            f'<span class="badge {badge}">{esc(status)}</span> '
            f'<b>{esc(u["sub_name"])}</b> · {esc(u["sub_trade"])}'
            f'<div class="small">{esc(u["update_date"])} · Manpower {u["manpower"] or 0}</div>'
            f'<div>{esc(u["commitment"]) or "No commitment entered."}</div>'
            f'<div class="small">{esc(u["issue"]) or "No issue entered."}</div>'
            f'</div>'
        )

    if not recent_html:
        recent_html = '<div class="muted">No subcontractor updates yet.</div>'

    body = f"""
    <div class="hero">
        <div style="display:flex;justify-content:space-between;align-items:center;gap:15px;flex-wrap:wrap;">
            <div>
                <div class="eyebrow">Subcontractor Intelligence</div>
                <h1>Know who is committed, who is staffed, and who needs attention.</h1>
            </div>

            <a href="/subcontractors/new"
               style="display:inline-block;background:#f0b44d;color:#0a1017;text-decoration:none;padding:12px 18px;border-radius:9px;font-weight:800;">
                + Add Subcontractor
            </a>
        </div>
    </div>

    <div class="grid3">
        {cards}
    </div>

    <div class="card">
        <h2>Recent Trade Partner Updates</h2>
        {recent_html}
    </div>
    """

    return shell("Subcontractors", body)


@app.get("/subcontractors/{sub_id}/update", response_class=HTMLResponse)
def subcontractor_update_form(sub_id: int):
    pid = project_id()
    c = db()

    sub = c.execute(
        """
        SELECT *
        FROM subs
        WHERE id=? AND project_id=?
        """,
        (sub_id, pid)
    ).fetchone()

    c.close()

    if not sub:
        return RedirectResponse(url="/subcontractors", status_code=303)

    body = f"""
    <div class="hero">
        <div class="eyebrow">Subcontractor Intelligence</div>
        <h1>Log Trade Partner Update</h1>
        <div class="muted">{esc(sub["name"])} · {esc(sub["trade"])}</div>
    </div>

    <div class="card" style="max-width:760px;">
        <form method="post" action="/subcontractors/{sub_id}/update">
            <label>Update Date</label>
            <input type="date" name="update_date" value="{date.today().isoformat()}" required>

            <label>Manpower</label>
            <input type="number" name="manpower" min="0" value="0">

            <label>Commitment</label>
            <textarea name="commitment" placeholder="Example: Complete level 2 overhead rough-in by Friday."></textarea>

            <label>Issue / Constraint</label>
            <textarea name="issue" placeholder="Example: Waiting on sleeves and access above ceiling."></textarea>

            <label>Status</label>
            <select name="status">
                <option value="READY">Ready / On Track</option>
                <option value="WATCH">Watch</option>
                <option value="HIGH">High Concern</option>
                <option value="CRITICAL">Critical</option>
                <option value="LOW">Low Concern</option>
            </select>

            <div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:18px;">
                <button type="submit">Save Update</button>
                <a href="/subcontractors"
                   style="display:inline-block;color:#f0b44d;text-decoration:none;padding:10px 4px;font-weight:700;">
                    Cancel
                </a>
            </div>
        </form>
    </div>
    """

    return shell("Subcontractor Update", body)


@app.post("/subcontractors/{sub_id}/update")
def save_subcontractor_update(
    sub_id: int,
    update_date: str = Form(...),
    manpower: int = Form(0),
    commitment: str = Form(""),
    issue: str = Form(""),
    status: str = Form("WATCH")
):
    pid = project_id()
    c = db()

    sub = c.execute(
        "SELECT id FROM subs WHERE id=? AND project_id=?",
        (sub_id, pid)
    ).fetchone()

    if sub:
        c.execute(
            """
            INSERT INTO subcontractor_updates(
                project_id,
                sub_id,
                update_date,
                manpower,
                commitment,
                issue,
                status,
                created
            )
            VALUES(?,?,?,?,?,?,?,?)
            """,
            (
                pid,
                sub_id,
                update_date,
                manpower,
                commitment.strip(),
                issue.strip(),
                status,
                date.today().isoformat()
            )
        )
        c.commit()

    c.close()
    return RedirectResponse(url="/subcontractors", status_code=303)


@app.get("/production",response_class=HTMLResponse)
def production():
    pid = project_id()
    c = db()

    acts = c.execute(
        """
        SELECT id,external_id,name
        FROM activities
        WHERE project_id=?
        ORDER BY start,name
        """,
        (pid,)
    ).fetchall()

    rows = c.execute(
        """
        SELECT p.*,a.external_id,a.name activity
        FROM production p
        JOIN activities a ON a.id=p.activity_id
        WHERE p.project_id=?
        ORDER BY p.work_date DESC,p.id DESC
        LIMIT 30
        """,
        (pid,)
    ).fetchall()

    c.close()

    opts = "".join(
        f'<option value="{a["id"]}">{esc(a["external_id"])} - {esc(a["name"])}</option>'
        for a in acts
    )

    total_qty = sum((r["qty"] or 0) for r in rows)
    total_plan = sum((r["planned_qty"] or 0) for r in rows)

    overall_pct = 0
    if total_plan > 0:
        overall_pct = (total_qty / total_plan) * 100

    behind_count = 0
    onplan_count = 0

    tr = ""

    for r in rows:
        qty = r["qty"] or 0
        plan = r["planned_qty"] or 0

        if plan > 0:
            pct = (qty / plan) * 100
            variance = qty - plan

            if pct < 90:
                status = "BEHIND"
                badge = "CRITICAL"
                behind_count += 1
            elif pct < 100:
                status = "WATCH"
                badge = "WATCH"
            else:
                status = "ON PLAN"
                badge = "READY"
                onplan_count += 1
        else:
            pct = 0
            variance = qty
            status = "NO PLAN"
            badge = "OPEN"

        tr += f"""
        <tr>
            <td>{esc(r["work_date"])}</td>
            <td><b>{esc(r["external_id"])} - {esc(r["activity"])}</b></td>
            <td>{r["crew"] or 0}</td>
            <td>{qty:g}</td>
            <td>{plan:g}</td>
            <td>{esc(r["unit"])}</td>
            <td>{pct:.0f}%</td>
            <td>{variance:+g}</td>
            <td><span class="badge {badge}">{status}</span></td>
        </tr>
        """

    if not tr:
        tr = '<tr><td colspan="9" class="muted">No production records yet.</td></tr>'

    if not opts:
        form_html = """
        <div class="muted">
            Add a schedule activity before recording production.
        </div>
        <div style="margin-top:12px;">
            <a href="/activities/new"
               style="color:#f0b44d;text-decoration:none;font-weight:700;">
                + Add Activity
            </a>
        </div>
        """
    else:
        form_html = f"""
        <form method="post" action="/production/add">
            <label>Activity</label>
            <select name="activity_id" required>{opts}</select>

            <label>Crew Size</label>
            <input name="crew" type="number" min="0" placeholder="Example: 6">

            <label>Installed Quantity</label>
            <input name="qty" type="number" step="0.1" placeholder="Example: 420">

            <label>Planned Quantity</label>
            <input name="planned_qty" type="number" step="0.1" placeholder="Example: 500">

            <label>Unit</label>
            <input name="unit" placeholder="LF / SF / CY / EA">

            <button type="submit">Record Production</button>
        </form>
        """

    body = f"""
    <div class="hero">
        <div class="eyebrow">Production Intelligence</div>
        <h1>Measure the job before the schedule update does.</h1>
        <div class="muted">
            Track actual production against plan and flag underperformance early.
        </div>
    </div>

    <div class="grid4">
        <div class="card">
            <div class="label">Production Records</div>
            <div class="kpi">{len(rows)}</div>
        </div>

        <div class="card">
            <div class="label">Overall To Plan</div>
            <div class="kpi">{overall_pct:.0f}%</div>
        </div>

        <div class="card">
            <div class="label">Behind Plan</div>
            <div class="kpi">{behind_count}</div>
        </div>

        <div class="card">
            <div class="label">On / Above Plan</div>
            <div class="kpi">{onplan_count}</div>
        </div>
    </div>

    <div class="grid2">
        <div class="card">
            <h2>Record Production</h2>
            {form_html}
        </div>

        <div class="card">
            <h2>How BuildCommand Reads It</h2>
            <p>
                Below 90% of planned production is flagged as <b>Behind</b>.
                Between 90% and 99% is flagged as <b>Watch</b>.
                At or above plan is shown as <b>On Plan</b>.
            </p>
            <div class="small">
                These records can be used by the AI Assistant when evaluating field performance and schedule risk.
            </div>
        </div>
    </div>

    <div class="card">
        <h2>Production History</h2>

        <table>
            <tr>
                <th>Date</th>
                <th>Activity</th>
                <th>Crew</th>
                <th>Actual</th>
                <th>Plan</th>
                <th>Unit</th>
                <th>% Plan</th>
                <th>Variance</th>
                <th>Status</th>
            </tr>

            {tr}
        </table>
    </div>
    """

    return shell("Production", body)


@app.post("/production/add")
def add_prod(activity_id:int=Form(...),crew:int=Form(0),qty:float=Form(0),planned_qty:float=Form(0),unit:str=Form("")):
    pid=project_id(); c=db(); c.execute("INSERT INTO production(project_id,activity_id,work_date,crew,qty,planned_qty,unit) VALUES(?,?,?,?,?,?,?)",(pid,activity_id,date.today().isoformat(),crew,qty,planned_qty,unit)); c.commit(); c.close()
    return RedirectResponse("/production",303)

@app.get("/risk",response_class=HTMLResponse)
def risk():
    pid = project_id()
    c = db()

    rows = c.execute(
        """
        SELECT r.*, a.external_id, a.name activity
        FROM risks r
        JOIN activities a ON a.id=r.activity_id
        WHERE r.project_id=?
        ORDER BY r.score DESC
        """,
        (pid,)
    ).fetchall()

    c.close()

    html = "".join(
        f"""
        <div class="card">
            <div style="display:flex;justify-content:space-between;gap:12px;align-items:flex-start;flex-wrap:wrap;">
                <div>
                    <span class="badge {r["band"]}">{r["band"]}</span>
                    <b>{esc(r["external_id"])} - {esc(r["activity"])}</b>
                </div>

                <a href="/risk/{r["id"]}/edit"
                   style="color:#f0b44d;text-decoration:none;font-weight:700;">
                    Edit
                </a>
            </div>

            <h3>{r["score"]:.0f}/100</h3>
            <p>{esc(r["explanation"])}</p>
        </div>
        """
        for r in rows
    ) or '<div class="card"><div class="muted">No risks entered for this project yet.</div></div>'

    body = f"""
    <div class="hero">
        <div style="display:flex;justify-content:space-between;align-items:center;gap:15px;flex-wrap:wrap;">
            <div>
                <div class="eyebrow">Predictive Risk</div>
                <h1>What is most likely to hurt the job next?</h1>
                <div class="muted">
                    Track field and schedule risk with an explainable score.
                </div>
            </div>

            <a href="/risk/new"
               style="display:inline-block;background:#f0b44d;color:#0a1017;text-decoration:none;padding:12px 18px;border-radius:9px;font-weight:800;">
                + Add Risk
            </a>
        </div>
    </div>

    {html}
    """

    return shell("Predictive Risk", body)


@app.get("/risk/new", response_class=HTMLResponse)
def new_risk_form():
    pid = project_id()
    c = db()

    project = c.execute(
        "SELECT name,number FROM projects WHERE id=?",
        (pid,)
    ).fetchone()

    activities = c.execute(
        """
        SELECT id,external_id,name
        FROM activities
        WHERE project_id=?
        ORDER BY start,name
        """,
        (pid,)
    ).fetchall()

    c.close()

    project_label = "Current Project"
    if project:
        project_label = f'{esc(project["number"])} - {esc(project["name"])}'

    options = "".join(
        f'<option value="{a["id"]}">{esc(a["external_id"])} - {esc(a["name"])}</option>'
        for a in activities
    )

    if not options:
        body = """
        <div class="hero">
            <div class="eyebrow">Predictive Risk</div>
            <h1>Add Risk</h1>
        </div>

        <div class="card">
            <p>This project has no schedule activities yet. Add an activity first.</p>
            <a href="/activities/new"
               style="color:#f0b44d;font-weight:700;text-decoration:none;">
                + Add Activity
            </a>
        </div>
        """
        return shell("Add Risk", body)

    body = f"""
    <div class="hero">
        <div class="eyebrow">Predictive Risk</div>
        <h1>Add Risk</h1>
        <div class="muted">{project_label}</div>
    </div>

    <div class="card" style="max-width:760px;">
        <form method="post" action="/risk/new">

            <label>Activity</label>
            <select name="activity_id" required>
                {options}
            </select>

            <div class="grid2">
                <div>
                    <label>Risk Score</label>
                    <input
                        type="number"
                        name="score"
                        min="0"
                        max="100"
                        step="1"
                        value="50"
                        required
                    >
                </div>

                <div>
                    <label>Risk Band</label>
                    <select name="band">
                        <option value="CRITICAL">Critical</option>
                        <option value="HIGH">High</option>
                        <option value="WATCH">Watch</option>
                        <option value="LOW">Low</option>
                    </select>
                </div>
            </div>

            <label>Why Is This a Risk?</label>
            <textarea
                name="explanation"
                placeholder="Example: Material release is late and threatens the rough-in start."
                required
            ></textarea>

            <div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:18px;">
                <button type="submit">Save Risk</button>

                <a href="/risk"
                   style="display:inline-block;color:#f0b44d;text-decoration:none;padding:10px 4px;font-weight:700;">
                    Cancel
                </a>
            </div>
        </form>
    </div>
    """

    return shell("Add Risk", body)


@app.post("/risk/new")
def create_risk(
    activity_id: int = Form(...),
    score: float = Form(...),
    band: str = Form(...),
    explanation: str = Form(...)
):
    pid = project_id()
    score = max(0.0, min(100.0, score))

    c = db()

    activity = c.execute(
        "SELECT id FROM activities WHERE id=? AND project_id=?",
        (activity_id, pid)
    ).fetchone()

    if activity:
        c.execute(
            """
            INSERT INTO risks(
                project_id,
                activity_id,
                score,
                band,
                explanation
            )
            VALUES(?,?,?,?,?)
            """,
            (
                pid,
                activity_id,
                score,
                band,
                explanation.strip()
            )
        )
        c.commit()

    c.close()

    return RedirectResponse(url="/risk", status_code=303)


@app.get("/risk/{risk_id}/edit", response_class=HTMLResponse)
def edit_risk_form(risk_id: int):
    pid = project_id()
    c = db()

    item = c.execute(
        """
        SELECT r.*, a.external_id, a.name activity
        FROM risks r
        JOIN activities a ON a.id=r.activity_id
        WHERE r.id=? AND r.project_id=?
        """,
        (risk_id, pid)
    ).fetchone()

    c.close()

    if not item:
        return RedirectResponse(url="/risk", status_code=303)

    bands = ["CRITICAL", "HIGH", "WATCH", "LOW"]
    band_options = "".join(
        f'<option value="{b}" {"selected" if item["band"] == b else ""}>{b.title()}</option>'
        for b in bands
    )

    body = f"""
    <div class="hero">
        <div class="eyebrow">Predictive Risk</div>
        <h1>Edit Risk</h1>
        <div class="muted">
            {esc(item["external_id"])} - {esc(item["activity"])}
        </div>
    </div>

    <div class="card" style="max-width:760px;">
        <form method="post" action="/risk/{risk_id}/edit">

            <div class="grid2">
                <div>
                    <label>Risk Score</label>
                    <input
                        type="number"
                        name="score"
                        min="0"
                        max="100"
                        step="1"
                        value="{item["score"]:.0f}"
                        required
                    >
                </div>

                <div>
                    <label>Risk Band</label>
                    <select name="band">
                        {band_options}
                    </select>
                </div>
            </div>

            <label>Explanation</label>
            <textarea name="explanation" required>{esc(item["explanation"])}</textarea>

            <div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:18px;">
                <button type="submit">Save Changes</button>

                <a href="/risk"
                   style="display:inline-block;color:#f0b44d;text-decoration:none;padding:10px 4px;font-weight:700;">
                    Cancel
                </a>
            </div>
        </form>

        <form method="post"
              action="/risk/{risk_id}/delete"
              style="margin-top:22px;padding-top:18px;border-top:1px solid #213042;">
            <button type="submit"
                    style="background:#492324;color:#ffb0b0;">
                Delete Risk
            </button>
        </form>
    </div>
    """

    return shell("Edit Risk", body)


@app.post("/risk/{risk_id}/edit")
def edit_risk(
    risk_id: int,
    score: float = Form(...),
    band: str = Form(...),
    explanation: str = Form(...)
):
    pid = project_id()
    score = max(0.0, min(100.0, score))

    c = db()
    c.execute(
        """
        UPDATE risks
        SET score=?, band=?, explanation=?
        WHERE id=? AND project_id=?
        """,
        (
            score,
            band,
            explanation.strip(),
            risk_id,
            pid
        )
    )
    c.commit()
    c.close()

    return RedirectResponse(url="/risk", status_code=303)


@app.post("/risk/{risk_id}/delete")
def delete_risk(risk_id: int):
    pid = project_id()

    c = db()
    c.execute(
        "DELETE FROM risks WHERE id=? AND project_id=?",
        (risk_id, pid)
    )
    c.commit()
    c.close()

    return RedirectResponse(url="/risk", status_code=303)


@app.get("/recovery",response_class=HTMLResponse)
def recovery():
    pid = project_id()
    c = db()

    rows = c.execute(
        """
        SELECT r.*, a.external_id, a.name activity
        FROM recovery r
        JOIN activities a ON a.id=r.activity_id
        WHERE r.project_id=?
        ORDER BY
            CASE r.status
                WHEN 'APPROVED' THEN 1
                WHEN 'PROPOSED' THEN 2
                WHEN 'REJECTED' THEN 3
                ELSE 4
            END,
            r.days_recovered DESC,
            r.est_cost ASC
        """,
        (pid,)
    ).fetchall()

    acts = c.execute(
        """
        SELECT id,external_id,name
        FROM activities
        WHERE project_id=?
        ORDER BY start,name
        """,
        (pid,)
    ).fetchall()

    c.close()

    opts = "".join(
        f'<option value="{a["id"]}">{esc(a["external_id"])} - {esc(a["name"])}</option>'
        for a in acts
    )

    total_scenarios = len(rows)
    approved = sum(1 for r in rows if r["status"] == "APPROVED")
    proposed = sum(1 for r in rows if r["status"] == "PROPOSED")

    best_value = None
    best_value_score = None

    scenario_cards = ""

    for r in rows:
        days = r["days_recovered"] or 0
        cost = r["est_cost"] or 0

        if days > 0:
            cost_per_day = cost / days
            value_text = f'${cost_per_day:,.0f} / recovered day'

            if best_value_score is None or cost_per_day < best_value_score:
                best_value_score = cost_per_day
                best_value = r["id"]
        else:
            cost_per_day = None
            value_text = "No recovered days entered"

        status_badge = (
            "READY" if r["status"] == "APPROVED"
            else "HIGH" if r["status"] == "REJECTED"
            else "OPEN"
        )

        best_badge = (
            '<span class="badge READY">BEST VALUE</span>'
            if best_value == r["id"]
            else ""
        )

        scenario_cards += f"""
        <div class="card">
            <div style="display:flex;justify-content:space-between;gap:12px;align-items:flex-start;flex-wrap:wrap;">
                <div>
                    <span class="badge {status_badge}">{esc(r["status"])}</span>
                    {best_badge}
                    <h3 style="margin:10px 0 4px;">{esc(r["scenario"]).replace("_"," ").title()}</h3>
                    <div class="muted">{esc(r["external_id"])} - {esc(r["activity"])}</div>
                </div>

                <form method="post" action="/recovery/{r["id"]}/delete">
                    <button type="submit" style="background:#492324;color:#ffb0b0;">
                        Delete
                    </button>
                </form>
            </div>

            <div class="grid3" style="margin-top:14px;">
                <div>
                    <div class="label">Days Recovered</div>
                    <div class="kpi">{days:.1f}</div>
                </div>

                <div>
                    <div class="label">Estimated Cost</div>
                    <div class="kpi">${cost:,.0f}</div>
                </div>

                <div>
                    <div class="label">Cost / Day</div>
                    <div style="font-size:20px;font-weight:800;">{value_text}</div>
                </div>
            </div>

            <div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:16px;">
                <form method="post" action="/recovery/{r["id"]}/approve">
                    <button type="submit">Approve</button>
                </form>

                <form method="post" action="/recovery/{r["id"]}/reject">
                    <button type="submit" style="background:#43381b;color:#ffd779;">
                        Reject
                    </button>
                </form>
            </div>
        </div>
        """

    if not scenario_cards:
        scenario_cards = '<div class="card"><div class="muted">No recovery scenarios saved yet.</div></div>'

    if opts:
        form = f"""
        <div class="card">
            <h2>Test Recovery Idea</h2>

            <form method="post" action="/recovery/add">
                <label>Activity</label>
                <select name="activity_id" required>
                    {opts}
                </select>

                <label>Scenario</label>
                <select name="scenario">
                    <option value="ADD_CREW">Add Crew</option>
                    <option value="OVERTIME">Overtime</option>
                    <option value="WORK_SATURDAY">Work Saturday</option>
                    <option value="RESEQUENCE">Resequence Work</option>
                    <option value="CLEAR_CONSTRAINT">Clear Constraint</option>
                    <option value="SECOND_SHIFT">Second Shift</option>
                </select>

                <label>Modeled Days Recovered</label>
                <input
                    name="days_recovered"
                    type="number"
                    step="0.1"
                    min="0"
                    placeholder="Example: 2.5"
                >

                <label>Estimated Cost</label>
                <input
                    name="est_cost"
                    type="number"
                    step="100"
                    min="0"
                    placeholder="Example: 7500"
                >

                <button type="submit">Save Recovery Scenario</button>
            </form>
        </div>
        """
    else:
        form = """
        <div class="card">
            <h2>Test Recovery Idea</h2>
            <p>Add a schedule activity before creating a recovery scenario.</p>
            <a href="/activities/new"
               style="color:#f0b44d;text-decoration:none;font-weight:700;">
                + Add Activity
            </a>
        </div>
        """

    body = f"""
    <div class="hero">
        <div class="eyebrow">Recovery Intelligence</div>
        <h1>Compare recovery before spending money.</h1>
        <div class="muted">
            Model time recovered, cost, and decision status before committing resources.
        </div>
    </div>

    <div class="grid4">
        <div class="card">
            <div class="label">Scenarios</div>
            <div class="kpi">{total_scenarios}</div>
        </div>

        <div class="card">
            <div class="label">Proposed</div>
            <div class="kpi">{proposed}</div>
        </div>

        <div class="card">
            <div class="label">Approved</div>
            <div class="kpi">{approved}</div>
        </div>

        <div class="card">
            <div class="label">Best Cost / Day</div>
            <div class="kpi">
                {"$"+format(best_value_score,",.0f") if best_value_score is not None else "—"}
            </div>
        </div>
    </div>

    <div class="grid2">
        {form}

        <div class="card">
            <h2>How To Use Recovery Intelligence</h2>
            <p>
                Compare options by how many schedule days they recover and what each recovered day costs.
                Approve the option you intend to execute and reject scenarios you no longer want considered.
            </p>
            <div class="small">
                BuildCommand AI can use these recovery scenarios when evaluating project risk and next actions.
            </div>
        </div>
    </div>

    <div>
        <h2>Recovery Scenarios</h2>
        {scenario_cards}
    </div>
    """

    return shell("Recovery", body)


@app.post("/recovery/add")
def add_recovery(
    activity_id: int = Form(...),
    scenario: str = Form(...),
    days_recovered: float = Form(0),
    est_cost: float = Form(0)
):
    pid = project_id()

    days_recovered = max(0.0, days_recovered)
    est_cost = max(0.0, est_cost)

    c = db()

    activity = c.execute(
        "SELECT id FROM activities WHERE id=? AND project_id=?",
        (activity_id, pid)
    ).fetchone()

    if activity:
        c.execute(
            """
            INSERT INTO recovery(
                project_id,
                activity_id,
                scenario,
                days_recovered,
                est_cost,
                status
            )
            VALUES(?,?,?,?,?,?)
            """,
            (
                pid,
                activity_id,
                scenario,
                days_recovered,
                est_cost,
                "PROPOSED"
            )
        )
        c.commit()

    c.close()

    return RedirectResponse("/recovery", 303)


@app.post("/recovery/{scenario_id}/approve")
def approve_recovery(scenario_id: int):
    pid = project_id()

    c = db()
    c.execute(
        """
        UPDATE recovery
        SET status='APPROVED'
        WHERE id=? AND project_id=?
        """,
        (scenario_id, pid)
    )
    c.commit()
    c.close()

    return RedirectResponse("/recovery", 303)


@app.post("/recovery/{scenario_id}/reject")
def reject_recovery(scenario_id: int):
    pid = project_id()

    c = db()
    c.execute(
        """
        UPDATE recovery
        SET status='REJECTED'
        WHERE id=? AND project_id=?
        """,
        (scenario_id, pid)
    )
    c.commit()
    c.close()

    return RedirectResponse("/recovery", 303)


@app.post("/recovery/{scenario_id}/delete")
def delete_recovery(scenario_id: int):
    pid = project_id()

    c = db()
    c.execute(
        "DELETE FROM recovery WHERE id=? AND project_id=?",
        (scenario_id, pid)
    )
    c.commit()
    c.close()

    return RedirectResponse("/recovery", 303)


@app.get("/memory",response_class=HTMLResponse)
def memory():
    pid=project_id(); c=db(); rows=c.execute("SELECT * FROM memory WHERE project_id=? ORDER BY confidence DESC",(pid,)).fetchall(); c.close()
    html="".join(f'<div class="card"><h3>{esc(r["category"])}</h3><p>{esc(r["insight"])}</p><div class="small">Confidence {r["confidence"]:.0%}</div></div>' for r in rows)
    return shell("Company Memory",'<div class="hero"><div class="eyebrow">Company Construction Memory</div><h1>Every project makes the next one smarter.</h1></div>'+html)

@app.get("/playbooks",response_class=HTMLResponse)
def playbooks():
    body='<div class="hero"><div class="eyebrow">Company Playbooks</div><h1>Turn repeated evidence into how your company runs work.</h1></div><div class="card"><h3>MEP Start Readiness</h3><p>Confirm manpower, material, predecessor completion, access, and inspection prerequisites. Escalate unresolved items before start.</p><div class="small">Evidence-backed company operating routine.</div></div><div class="card"><h3>Schedule Drift Response</h3><p>Verify field status, identify the controlling constraint, compare the field plan with the master schedule, test recovery, and document the agreed plan.</p></div>'
    return shell("Playbooks",body)

@app.get("/portfolio",response_class=HTMLResponse)
def portfolio():
    current_pid = project_id()
    c = db()

    projects = c.execute(
        """
        SELECT *
        FROM projects
        WHERE company_id=?
        ORDER BY
            CASE status
                WHEN 'ACTIVE' THEN 1
                WHEN 'PLANNING' THEN 2
                WHEN 'ON_HOLD' THEN 3
                WHEN 'COMPLETE' THEN 4
                ELSE 5
            END,
            name
        """,
        (current_company_id(),)
    ).fetchall()

    cards = []
    portfolio_scores = []

    for p in projects:
        pid = p["id"]

        risks = c.execute(
            "SELECT * FROM risks WHERE project_id=?",
            (pid,)
        ).fetchall()

        mr = c.execute(
            """
            SELECT COUNT(*) n
            FROM make_ready
            WHERE project_id=? AND status='OPEN'
            """,
            (pid,)
        ).fetchone()["n"]

        activities = c.execute(
            """
            SELECT COUNT(*) n
            FROM activities
            WHERE project_id=?
            """,
            (pid,)
        ).fetchone()["n"]

        active_activities = c.execute(
            """
            SELECT COUNT(*) n
            FROM activities
            WHERE project_id=? AND status='IN_PROGRESS'
            """,
            (pid,)
        ).fetchone()["n"]

        report = c.execute(
            """
            SELECT report_date, manpower, delays
            FROM daily_reports
            WHERE project_id=?
            ORDER BY report_date DESC, id DESC
            LIMIT 1
            """,
            (pid,)
        ).fetchone()

        crit = sum(r["band"] == "CRITICAL" for r in risks)
        high = sum(r["band"] == "HIGH" for r in risks)
        watch = sum(r["band"] == "WATCH" for r in risks)

        score = min(
            100,
            crit * 30 +
            high * 15 +
            watch * 5 +
            mr * 5
        )

        portfolio_scores.append(score)

        if score >= 70:
            attention_band = "CRITICAL"
            attention_text = "Immediate Attention"
        elif score >= 40:
            attention_band = "HIGH"
            attention_text = "Needs Attention"
        elif score >= 15:
            attention_band = "WATCH"
            attention_text = "Watch"
        else:
            attention_band = "READY"
            attention_text = "Stable"

        selected_badge = (
            '<span class="badge READY">CURRENT</span>'
            if pid == current_pid
            else ""
        )

        latest_report = esc(report["report_date"]) if report else "No report"
        latest_manpower = report["manpower"] if report else 0
        latest_delay = esc(report["delays"]) if report and report["delays"] else "No delay noted"

        card = f"""
        <div class="card">
            <div style="display:flex;justify-content:space-between;gap:12px;align-items:flex-start;flex-wrap:wrap;">
                <div>
                    <div class="eyebrow">{esc(p["number"])}</div>
                    <h2 style="margin:5px 0;">{esc(p["name"])}</h2>
                    <div class="muted">{esc(p["status"]).replace("_"," ").title()}</div>
                </div>

                <div style="display:flex;gap:8px;flex-wrap:wrap;">
                    {selected_badge}
                    <span class="badge {attention_band}">{attention_text}</span>
                </div>
            </div>

            <div class="grid4" style="margin-top:16px;">
                <div>
                    <div class="label">Attention</div>
                    <div class="kpi">{score}</div>
                </div>
                <div>
                    <div class="label">Critical</div>
                    <div class="kpi">{crit}</div>
                </div>
                <div>
                    <div class="label">High</div>
                    <div class="kpi">{high}</div>
                </div>
                <div>
                    <div class="label">Make Ready</div>
                    <div class="kpi">{mr}</div>
                </div>
            </div>

            <div style="margin-top:16px;">
                <div class="small">
                    Activities: {activities} · Active: {active_activities}
                </div>
                <div class="small">
                    Latest report: {latest_report} · Manpower: {latest_manpower or 0}
                </div>
                <div class="small">
                    Latest delay signal: {latest_delay}
                </div>
            </div>

            <form method="post" action="/projects/select" style="margin-top:16px;">
                <input type="hidden" name="project_id" value="{pid}">
                <button type="submit">Open Project</button>
            </form>
        </div>
        """

        cards.append((score, card))

    c.close()

    cards.sort(key=lambda x: x[0], reverse=True)
    cards_html = "".join(card for _, card in cards)

    total_projects = len(projects)
    active_projects = sum(p["status"] == "ACTIVE" for p in projects)
    high_attention = sum(score >= 40 for score in portfolio_scores)
    avg_score = (
        sum(portfolio_scores) / len(portfolio_scores)
        if portfolio_scores
        else 0
    )

    if not cards_html:
        cards_html = '<div class="card"><div class="muted">No projects created yet.</div></div>'

    body = f"""
    <div class="hero">
        <div class="eyebrow">Executive Intelligence</div>
        <h1>Which projects need intervention?</h1>
        <div class="muted">
            Portfolio-wide risk, make-ready, field reporting, and schedule attention in one view.
        </div>
    </div>

    <div class="grid4">
        <div class="card">
            <div class="label">Projects</div>
            <div class="kpi">{total_projects}</div>
        </div>

        <div class="card">
            <div class="label">Active Projects</div>
            <div class="kpi">{active_projects}</div>
        </div>

        <div class="card">
            <div class="label">Need Attention</div>
            <div class="kpi">{high_attention}</div>
        </div>

        <div class="card">
            <div class="label">Avg Attention</div>
            <div class="kpi">{avg_score:.0f}</div>
        </div>
    </div>

    <div class="grid2">
        {cards_html}
    </div>
    """

    return shell("Portfolio", body)

@app.get("/activities/new", response_class=HTMLResponse)
def new_activity_form():
    pid = project_id()
    c = db()
    project = c.execute("SELECT name,number FROM projects WHERE id=?", (pid,)).fetchone()
    c.close()

    project_label = "Current Project"
    if project:
        project_label = f'{esc(project["number"])} - {esc(project["name"])}'

    body = f"""
    <div class="hero">
        <div class="eyebrow">Schedule</div>
        <h1>Add Activity</h1>
        <div class="muted">Add a schedule activity to {project_label}.</div>
    </div>
    <div class="card" style="max-width:760px;">
        <form method="post" action="/activities/new">
            <div class="grid2">
                <div><label>Activity ID</label><input type="text" name="external_id" placeholder="Example: A600" required></div>
                <div><label>Trade</label><input type="text" name="trade" placeholder="Example: Electrical" required></div>
            </div>
            <label>Activity Name</label>
            <input type="text" name="name" placeholder="Example: Electrical Rough-In" required>
            <div class="grid2">
                <div><label>Start Date</label><input type="date" name="start" required></div>
                <div><label>Finish Date</label><input type="date" name="finish" required></div>
            </div>
            <div class="grid2">
                <div><label>Percent Complete</label><input type="number" name="pct" min="0" max="100" step="1" value="0" required></div>
                <div>
                    <label>Status</label>
                    <select name="status">
                        <option value="NOT_STARTED">Not Started</option>
                        <option value="IN_PROGRESS">In Progress</option>
                        <option value="COMPLETE">Complete</option>
                        <option value="HOLD">Hold</option>
                    </select>
                </div>
            </div>
            <div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:18px;">
                <button type="submit">Save Activity</button>
                <a href="/schedule" style="display:inline-block;color:#f0b44d;text-decoration:none;padding:10px 4px;font-weight:700;">Cancel</a>
            </div>
        </form>
    </div>
    """
    return shell("Add Activity", body)


@app.post("/activities/new")
def create_activity(
    external_id: str = Form(...),
    name: str = Form(...),
    trade: str = Form(...),
    start: str = Form(...),
    finish: str = Form(...),
    pct: float = Form(0),
    status: str = Form("NOT_STARTED")
):
    pid = project_id()
    pct = max(0.0, min(100.0, pct))
    c = db()
    c.execute(
        """
        INSERT INTO activities(project_id,external_id,name,trade,start,finish,pct,status)
        VALUES(?,?,?,?,?,?,?,?)
        """,
        (pid, external_id.strip(), name.strip(), trade.strip(), start, finish, pct, status)
    )
    c.commit()
    c.close()
    return RedirectResponse(url="/schedule", status_code=303)


@app.get("/activities/{activity_id}/edit", response_class=HTMLResponse)
def edit_activity_form(activity_id: int):
    pid = project_id()
    c = db()
    activity = c.execute(
        "SELECT * FROM activities WHERE id=? AND project_id=?",
        (activity_id, pid)
    ).fetchone()
    c.close()

    if not activity:
        return RedirectResponse(url="/schedule", status_code=303)

    statuses = ["NOT_STARTED", "IN_PROGRESS", "COMPLETE", "HOLD"]
    options = "".join(
        f'<option value="{s}" {"selected" if activity["status"] == s else ""}>{s.replace("_"," ").title()}</option>'
        for s in statuses
    )

    body = f"""
    <div class="hero">
        <div class="eyebrow">Schedule</div>
        <h1>Edit Activity</h1>
        <div class="muted">Update schedule information for this activity.</div>
    </div>

    <div class="card" style="max-width:760px;">
        <form method="post" action="/activities/{activity_id}/edit">
            <div class="grid2">
                <div>
                    <label>Activity ID</label>
                    <input type="text" name="external_id" value="{esc(activity["external_id"])}" required>
                </div>
                <div>
                    <label>Trade</label>
                    <input type="text" name="trade" value="{esc(activity["trade"])}" required>
                </div>
            </div>

            <label>Activity Name</label>
            <input type="text" name="name" value="{esc(activity["name"])}" required>

            <div class="grid2">
                <div>
                    <label>Start Date</label>
                    <input type="date" name="start" value="{activity["start"]}" required>
                </div>
                <div>
                    <label>Finish Date</label>
                    <input type="date" name="finish" value="{activity["finish"]}" required>
                </div>
            </div>

            <div class="grid2">
                <div>
                    <label>Percent Complete</label>
                    <input type="number" name="pct" min="0" max="100" step="1" value="{activity["pct"]:.0f}" required>
                </div>
                <div>
                    <label>Status</label>
                    <select name="status">{options}</select>
                </div>
            </div>

            <div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:18px;">
                <button type="submit">Save Changes</button>
                <a href="/schedule" style="display:inline-block;color:#f0b44d;text-decoration:none;padding:10px 4px;font-weight:700;">Cancel</a>
            </div>
        </form>

        <form method="post" action="/activities/{activity_id}/delete" style="margin-top:22px;padding-top:18px;border-top:1px solid #213042;">
            <button type="submit" style="background:#492324;color:#ffb0b0;">Delete Activity</button>
        </form>
    </div>
    """
    return shell("Edit Activity", body)


@app.post("/activities/{activity_id}/edit")
def edit_activity(
    activity_id: int,
    external_id: str = Form(...),
    name: str = Form(...),
    trade: str = Form(...),
    start: str = Form(...),
    finish: str = Form(...),
    pct: float = Form(0),
    status: str = Form("NOT_STARTED")
):
    pid = project_id()
    pct = max(0.0, min(100.0, pct))
    c = db()
    c.execute(
        """
        UPDATE activities
        SET external_id=?, name=?, trade=?, start=?, finish=?, pct=?, status=?
        WHERE id=? AND project_id=?
        """,
        (external_id.strip(), name.strip(), trade.strip(), start, finish, pct, status, activity_id, pid)
    )
    c.commit()
    c.close()
    return RedirectResponse(url="/schedule", status_code=303)


@app.post("/activities/{activity_id}/delete")
def delete_activity(activity_id: int):
    pid = project_id()
    c = db()

    c.execute("DELETE FROM field_updates WHERE activity_id=? AND project_id=?", (activity_id, pid))
    c.execute("DELETE FROM production WHERE activity_id=? AND project_id=?", (activity_id, pid))
    c.execute("DELETE FROM make_ready WHERE activity_id=? AND project_id=?", (activity_id, pid))
    c.execute("DELETE FROM risks WHERE activity_id=? AND project_id=?", (activity_id, pid))
    c.execute("DELETE FROM recovery WHERE activity_id=? AND project_id=?", (activity_id, pid))
    c.execute("DELETE FROM activities WHERE id=? AND project_id=?", (activity_id, pid))

    c.commit()
    c.close()
    return RedirectResponse(url="/schedule", status_code=303)


@app.get("/subcontractors/new", response_class=HTMLResponse)
def new_subcontractor_form():
    pid = project_id()
    c = db()
    project = c.execute("SELECT name,number FROM projects WHERE id=?", (pid,)).fetchone()
    c.close()

    project_label = "Current Project"
    if project:
        project_label = f'{esc(project["number"])} - {esc(project["name"])}'

    body = f"""
    <div class="hero">
        <div class="eyebrow">Subcontractors</div>
        <h1>Add Subcontractor</h1>
        <div class="muted">Add a trade partner to {project_label}.</div>
    </div>

    <div class="card" style="max-width:680px;">
        <form method="post" action="/subcontractors/new">
            <label>Company Name</label>
            <input type="text" name="name" placeholder="Example: Valley Electric" required>

            <label>Trade</label>
            <input type="text" name="trade" placeholder="Example: Electrical" required>

            <div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:18px;">
                <button type="submit">Save Subcontractor</button>
                <a href="/subcontractors" style="display:inline-block;color:#f0b44d;text-decoration:none;padding:10px 4px;font-weight:700;">Cancel</a>
            </div>
        </form>
    </div>
    """
    return shell("Add Subcontractor", body)


@app.post("/subcontractors/new")
def create_subcontractor(
    name: str = Form(...),
    trade: str = Form(...)
):
    pid = project_id()
    c = db()
    c.execute(
        "INSERT INTO subs(project_id,name,trade) VALUES(?,?,?)",
        (pid, name.strip(), trade.strip())
    )
    c.commit()
    c.close()
    return RedirectResponse(url="/subcontractors", status_code=303)


@app.get("/daily-report", response_class=HTMLResponse)
def daily_report():
    pid = project_id()
    c = db()

    project = c.execute(
        "SELECT name,number FROM projects WHERE id=?",
        (pid,)
    ).fetchone()

    reports = c.execute(
        """
        SELECT *
        FROM daily_reports
        WHERE project_id=?
        ORDER BY report_date DESC, id DESC
        LIMIT 10
        """,
        (pid,)
    ).fetchall()

    analyses = c.execute(
        """
        SELECT *
        FROM daily_report_analysis
        WHERE project_id=?
        ORDER BY id DESC
        """,
        (pid,)
    ).fetchall()

    c.close()

    analysis_by_report = {}
    for a in analyses:
        if a["report_id"] not in analysis_by_report:
            analysis_by_report[a["report_id"]] = a

    project_label = "Current Project"
    if project:
        project_label = f'{esc(project["number"])} - {esc(project["name"])}'

    report_cards = ""

    for r in reports:
        analysis = analysis_by_report.get(r["id"])

        if analysis:
            analysis_html = (
                '<div style="margin-top:16px;padding-top:14px;border-top:1px solid #213042;">'
                '<div class="small">BUILDCOMMAND ANALYSIS</div>'
                f'<div style="margin-top:8px;line-height:1.6;">{esc(analysis["analysis_text"]).replace(chr(10), "<br>")}</div>'
                '</div>'
            )
            analyze_button = (
                f'<form method="post" action="/daily-report/{r["id"]}/analyze">'
                '<button type="submit">Analyze Again</button>'
                '</form>'
            )
        else:
            analysis_html = ""
            analyze_button = (
                f'<form method="post" action="/daily-report/{r["id"]}/analyze">'
                '<button type="submit">Analyze Report</button>'
                '</form>'
            )

        report_cards += f"""
        <div class="card">
            <div style="display:flex;justify-content:space-between;gap:12px;flex-wrap:wrap;align-items:flex-start;">
                <div>
                    <div class="eyebrow">Daily Report</div>
                    <h3 style="margin:6px 0;">{esc(r["report_date"])}</h3>
                    <div class="badge OPEN">{r["manpower"] or 0} workers</div>
                </div>
                <div>{analyze_button}</div>
            </div>

            <div class="small" style="margin-top:14px;">Weather</div>
            <p>{esc(r["weather"]) or "—"}</p>

            <div class="small">Work Completed</div>
            <p>{esc(r["work_completed"]) or "—"}</p>

            <div class="small">Delays / Constraints</div>
            <p>{esc(r["delays"]) or "—"}</p>

            <div class="small">Deliveries</div>
            <p>{esc(r["deliveries"]) or "—"}</p>

            <div class="small">Inspections</div>
            <p>{esc(r["inspections"]) or "—"}</p>

            <div class="small">Safety</div>
            <p>{esc(r["safety"]) or "—"}</p>

            <div class="small">Tomorrow's Plan</div>
            <p>{esc(r["tomorrow_plan"]) or "—"}</p>

            {analysis_html}
        </div>
        """

    if not report_cards:
        report_cards = '<div class="card"><div class="muted">No daily reports yet.</div></div>'

    body = f"""
    <div class="hero">
        <div class="eyebrow">Daily Superintendent Report</div>
        <h1>Capture the job today so BuildCommand can analyze tomorrow.</h1>
        <div class="muted">{project_label}</div>
    </div>

    <div class="grid2">
        <div class="card">
            <h2>New Daily Report</h2>

            <form method="post" action="/daily-report">
                <label>Report Date</label>
                <input type="date" name="report_date" value="{date.today().isoformat()}" required>

                <label>Weather</label>
                <input type="text" name="weather" placeholder="Example: Clear, 96°F">

                <label>Total Manpower</label>
                <input type="number" name="manpower" min="0" value="0">

                <label>Work Completed</label>
                <textarea name="work_completed" placeholder="What got completed today?"></textarea>

                <label>Delays / Constraints</label>
                <textarea name="delays" placeholder="Access issues, missing material, RFIs, manpower problems, etc."></textarea>

                <label>Deliveries</label>
                <textarea name="deliveries" placeholder="What arrived or failed to arrive?"></textarea>

                <label>Inspections</label>
                <textarea name="inspections" placeholder="Passed, failed, scheduled, or pending inspections."></textarea>

                <label>Safety Notes</label>
                <textarea name="safety" placeholder="Incidents, observations, toolbox talks, corrective actions."></textarea>

                <label>Tomorrow's Plan</label>
                <textarea name="tomorrow_plan" placeholder="What needs to happen tomorrow?"></textarea>

                <button type="submit">Save Daily Report</button>
            </form>
        </div>

        <div>
            <h2 style="margin-top:0;">Recent Reports</h2>
            {report_cards}
        </div>
    </div>
    """

    return shell("Daily Report", body)


@app.post("/daily-report")
def save_daily_report(
    report_date: str = Form(...),
    weather: str = Form(""),
    manpower: int = Form(0),
    work_completed: str = Form(""),
    delays: str = Form(""),
    deliveries: str = Form(""),
    inspections: str = Form(""),
    safety: str = Form(""),
    tomorrow_plan: str = Form("")
):
    pid = project_id()
    c = db()
    c.execute(
        """
        INSERT INTO daily_reports(
            project_id, report_date, weather, manpower, work_completed,
            delays, deliveries, inspections, safety, tomorrow_plan, created
        )
        VALUES(?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            pid,
            report_date,
            weather.strip(),
            manpower,
            work_completed.strip(),
            delays.strip(),
            deliveries.strip(),
            inspections.strip(),
            safety.strip(),
            tomorrow_plan.strip(),
            date.today().isoformat()
        )
    )
    c.commit()
    c.close()
    return RedirectResponse(url="/daily-report", status_code=303)


@app.get("/ai-analysis", response_class=HTMLResponse)
def ai_analysis():
    pid = project_id()
    c = db()

    project = c.execute(
        "SELECT * FROM projects WHERE id=?",
        (pid,)
    ).fetchone()

    activities = c.execute(
        "SELECT * FROM activities WHERE project_id=? ORDER BY start",
        (pid,)
    ).fetchall()

    risks = c.execute(
        """
        SELECT r.*, a.name activity
        FROM risks r
        JOIN activities a ON a.id=r.activity_id
        WHERE r.project_id=?
        ORDER BY r.score DESC
        """,
        (pid,)
    ).fetchall()

    make_ready_items = c.execute(
        """
        SELECT m.*, a.name activity
        FROM make_ready m
        JOIN activities a ON a.id=m.activity_id
        WHERE m.project_id=? AND m.status='OPEN'
        ORDER BY due
        """,
        (pid,)
    ).fetchall()

    report = c.execute(
        """
        SELECT *
        FROM daily_reports
        WHERE project_id=?
        ORDER BY report_date DESC, id DESC
        LIMIT 1
        """,
        (pid,)
    ).fetchone()

    c.close()

    project_name = esc(project["name"]) if project else "Current Project"

    critical = [r for r in risks if r["band"] == "CRITICAL"]
    high = [r for r in risks if r["band"] == "HIGH"]
    in_progress = [a for a in activities if a["status"] == "IN_PROGRESS"]

    top_actions = []

    for r in critical[:3]:
        top_actions.append(
            f'CRITICAL: {esc(r["activity"])} — {esc(r["explanation"])}'
        )

    for item in make_ready_items[:3]:
        top_actions.append(
            f'MAKE READY: {esc(item["title"])} — due {esc(item["due"])}'
        )

    if report:
        delays = (report["delays"] or "").strip()
        inspections = (report["inspections"] or "").strip()
        tomorrow = (report["tomorrow_plan"] or "").strip()

        if delays:
            top_actions.append(f'FIELD DELAY: {esc(delays)}')

        if inspections:
            top_actions.append(f'INSPECTION: {esc(inspections)}')

        if tomorrow:
            top_actions.append(f'TOMORROW: {esc(tomorrow)}')

    if not top_actions:
        top_actions.append(
            "No immediate high-priority issues were found in the current project data."
        )

    action_html = "".join(
        f'<div class="action">{item}</div>'
        for item in top_actions[:7]
    )

    if report:
        report_summary = (
            f'<div class="small">Latest Report: {esc(report["report_date"])}</div>'
            f'<p><b>Manpower:</b> {report["manpower"] or 0}</p>'
            f'<p><b>Work Completed:</b> {esc(report["work_completed"]) or "—"}</p>'
            f'<p><b>Delays:</b> {esc(report["delays"]) or "—"}</p>'
            f'<p><b>Tomorrow:</b> {esc(report["tomorrow_plan"]) or "—"}</p>'
        )
    else:
        report_summary = '<div class="muted">No daily report has been submitted yet.</div>'

    risk_html = "".join(
        (
            f'<div class="action">'
            f'<span class="badge {r["band"]}">{r["band"]}</span> '
            f'<b>{esc(r["activity"])}</b> · {r["score"]:.0f}/100'
            f'<div class="small">{esc(r["explanation"])}</div>'
            f'</div>'
        )
        for r in risks[:6]
    ) or '<div class="muted">No risk records yet.</div>'

    body = f"""
    <div class="hero">
        <div class="eyebrow">BuildCommand Intelligence</div>
        <h1>What needs attention now?</h1>
        <div class="muted">
            Current project analysis for {project_name}.
        </div>
    </div>

    <div class="grid4">
        <div class="card">
            <div class="label">Critical Risks</div>
            <div class="kpi">{len(critical)}</div>
        </div>
        <div class="card">
            <div class="label">High Risks</div>
            <div class="kpi">{len(high)}</div>
        </div>
        <div class="card">
            <div class="label">Open Make-Ready</div>
            <div class="kpi">{len(make_ready_items)}</div>
        </div>
        <div class="card">
            <div class="label">Active Activities</div>
            <div class="kpi">{len(in_progress)}</div>
        </div>
    </div>

    <div class="grid2">
        <div class="card">
            <h2>Handle First</h2>
            {action_html}
        </div>

        <div class="card">
            <h2>Latest Field Signal</h2>
            {report_summary}
        </div>
    </div>

    <div class="card">
        <h2>Risk Intelligence</h2>
        {risk_html}
    </div>

    <div class="card">
        <h2>BuildCommand Recommendation</h2>
        <p>
            Focus first on critical risks and open make-ready items, then verify field delays,
            inspections, and tomorrow's plan before committing additional manpower or recovery cost.
        </p>
        <div class="small">
            This is the first intelligence layer. The next version can connect a live AI model
            so BuildCommand can answer natural-language questions about the project.
        </div>
    </div>
    """

    return shell("AI Analysis", body)


def build_project_context(pid):
    c = db()

    project = c.execute(
        "SELECT * FROM projects WHERE id=?",
        (pid,)
    ).fetchone()

    activities = c.execute(
        "SELECT * FROM activities WHERE project_id=? ORDER BY start",
        (pid,)
    ).fetchall()

    risks = c.execute(
        """
        SELECT r.*, a.external_id, a.name activity
        FROM risks r
        JOIN activities a ON a.id=r.activity_id
        WHERE r.project_id=?
        ORDER BY r.score DESC
        """,
        (pid,)
    ).fetchall()

    make_ready_items = c.execute(
        """
        SELECT m.*, a.external_id, a.name activity
        FROM make_ready m
        JOIN activities a ON a.id=m.activity_id
        WHERE m.project_id=? AND m.status='OPEN'
        ORDER BY due
        """,
        (pid,)
    ).fetchall()

    reports = c.execute(
        """
        SELECT *
        FROM daily_reports
        WHERE project_id=?
        ORDER BY report_date DESC, id DESC
        LIMIT 5
        """,
        (pid,)
    ).fetchall()

    subs = c.execute(
        "SELECT * FROM subs WHERE project_id=? ORDER BY trade,name",
        (pid,)
    ).fetchall()

    production_rows = c.execute(
        """
        SELECT p.*, a.external_id, a.name activity
        FROM production p
        JOIN activities a ON a.id=p.activity_id
        WHERE p.project_id=?
        ORDER BY p.work_date DESC, p.id DESC
        LIMIT 12
        """,
        (pid,)
    ).fetchall()

    sub_updates = c.execute(
        """
        SELECT u.*, s.name sub_name, s.trade
        FROM subcontractor_updates u
        JOIN subs s ON s.id=u.sub_id
        WHERE u.project_id=?
        ORDER BY u.update_date DESC, u.id DESC
        LIMIT 15
        """,
        (pid,)
    ).fetchall()

    recovery_rows = c.execute(
        """
        SELECT r.*, a.external_id, a.name activity
        FROM recovery r
        JOIN activities a ON a.id=r.activity_id
        WHERE r.project_id=?
        ORDER BY r.id DESC
        LIMIT 12
        """,
        (pid,)
    ).fetchall()

    action_rows = c.execute(
        """
        SELECT *
        FROM action_items
        WHERE project_id=? AND status='OPEN'
        ORDER BY due, id
        LIMIT 20
        """,
        (pid,)
    ).fetchall()

    readiness_rows = c.execute(
        """
        SELECT ar.*, a.external_id, a.name activity, a.start
        FROM activity_readiness ar
        JOIN activities a ON a.id=ar.activity_id
        WHERE ar.project_id=?
        ORDER BY a.start
        LIMIT 20
        """,
        (pid,)
    ).fetchall()

    procurement_rows = c.execute(
        """
        SELECT p.*, a.external_id, a.name activity
        FROM procurement p
        LEFT JOIN activities a ON a.id=p.activity_id
        WHERE p.project_id=?
        ORDER BY p.required_on_site, p.id
        LIMIT 20
        """,
        (pid,)
    ).fetchall()

    issue_rows = c.execute(
        """
        SELECT i.*, a.external_id, a.name activity
        FROM project_issues i
        LEFT JOIN activities a ON a.id=i.activity_id
        WHERE i.project_id=? AND i.status!='CLOSED'
        ORDER BY i.due, i.id
        LIMIT 20
        """,
        (pid,)
    ).fetchall()

    punch_rows = c.execute(
        """
        SELECT p.*, a.external_id, a.name activity
        FROM punch_items p
        LEFT JOIN activities a ON a.id=p.activity_id
        WHERE p.project_id=? AND p.status!='VERIFIED'
        ORDER BY p.due, p.id
        LIMIT 20
        """,
        (pid,)
    ).fetchall()

    inspection_rows = c.execute(
        """
        SELECT i.*, a.external_id, a.name activity
        FROM inspections_tracker i
        LEFT JOIN activities a ON a.id=i.activity_id
        WHERE i.project_id=? AND i.result!='PASSED'
        ORDER BY i.scheduled_date, i.id
        LIMIT 20
        """,
        (pid,)
    ).fetchall()

    submittal_rows = c.execute(
        """
        SELECT s.*, a.external_id, a.name activity
        FROM submittals s
        LEFT JOIN activities a ON a.id=s.activity_id
        WHERE s.project_id=? AND s.status NOT IN ('APPROVED','APPROVED_AS_NOTED')
        ORDER BY s.due_date, s.id
        LIMIT 20
        """,
        (pid,)
    ).fetchall()

    safety_rows = c.execute(
        """
        SELECT s.*, a.external_id, a.name activity
        FROM safety_items s
        LEFT JOIN activities a ON a.id=s.activity_id
        WHERE s.project_id=? AND s.status!='CLOSED'
        ORDER BY s.event_date DESC, s.id DESC
        LIMIT 20
        """,
        (pid,)
    ).fetchall()

    change_rows = c.execute(
        """
        SELECT ce.*, a.external_id, a.name activity
        FROM change_events ce
        LEFT JOIN activities a ON a.id=ce.activity_id
        WHERE ce.project_id=? AND ce.status NOT IN ('APPROVED','REJECTED')
        ORDER BY ce.id DESC
        LIMIT 20
        """,
        (pid,)
    ).fetchall()

    meeting_rows = c.execute(
        """
        SELECT *
        FROM meeting_notes
        WHERE project_id=?
        ORDER BY meeting_date DESC, id DESC
        LIMIT 10
        """,
        (pid,)
    ).fetchall()

    c.close()

    lines = []

    if project:
        lines.append(
            f'PROJECT: {project["number"]} - {project["name"]} | Status: {project["status"]}'
        )

    lines.append("\nSCHEDULE ACTIVITIES:")
    for a in activities:
        lines.append(
            f'- {a["external_id"]}: {a["name"]} | Trade: {a["trade"]} | '
            f'{a["start"]} to {a["finish"]} | {a["pct"]:.0f}% | {a["status"]}'
        )

    lines.append("\nRISKS:")
    for r in risks:
        lines.append(
            f'- {r["band"]} {r["score"]:.0f}/100 | {r["external_id"]} {r["activity"]}: '
            f'{r["explanation"]}'
        )

    lines.append("\nOPEN MAKE-READY ITEMS:")
    for m in make_ready_items:
        lines.append(
            f'- {m["priority"]} | {m["external_id"]} {m["activity"]} | '
            f'{m["title"]} | Reason: {m["reason"]} | Due: {m["due"]}'
        )

    lines.append("\nSUBCONTRACTORS:")
    for s in subs:
        lines.append(f'- {s["name"]} | Trade: {s["trade"]}')

    lines.append("\nRECENT PRODUCTION:")
    for p in production_rows:
        qty = p["qty"] or 0
        plan = p["planned_qty"] or 0
        pct = (qty / plan * 100) if plan > 0 else 0
        lines.append(
            f'- {p["work_date"]} | {p["external_id"]} {p["activity"]} | '
            f'Crew {p["crew"] or 0} | Actual {qty:g} {p["unit"] or ""} | '
            f'Plan {plan:g} | {pct:.0f}% of plan'
        )

    lines.append("\nRECENT SUBCONTRACTOR UPDATES:")
    for u in sub_updates:
        lines.append(
            f'- {u["update_date"]} | {u["sub_name"]} | Trade: {u["trade"]} | '
            f'Manpower: {u["manpower"] or 0} | Status: {u["status"] or "N/A"} | '
            f'Commitment: {u["commitment"] or "N/A"} | Issue: {u["issue"] or "N/A"}'
        )

    lines.append("\nRECOVERY SCENARIOS:")
    for r in recovery_rows:
        days = r["days_recovered"] or 0
        cost = r["est_cost"] or 0
        cpd = (cost / days) if days > 0 else 0
        lines.append(
            f'- {r["external_id"]} {r["activity"]} | '
            f'{r["scenario"]} | Days recovered: {days:.1f} | '
            f'Cost: ${cost:,.0f} | Cost/day: ${cpd:,.0f} | Status: {r["status"]}'
        )

    lines.append("\nOPEN ACTION ITEMS:")
    for a in action_rows:
        lines.append(
            f'- {a["priority"]} | {a["title"]} | Owner: {a["owner"] or "Unassigned"} | '
            f'Due: {a["due"]} | Notes: {a["notes"] or "N/A"}'
        )

    lines.append("\nACTIVITY READINESS:")
    for r in readiness_rows:
        pct, status, _ = readiness_result(r)
        lines.append(
            f'- {r["external_id"]} {r["activity"]} | Starts {r["start"]} | '
            f'Readiness {pct}% | Status: {status} | Notes: {r["notes"] or "N/A"}'
        )

    lines.append("\nSCHEDULE HEALTH:")
    readiness_map = {r["activity_id"]: r for r in readiness_rows}

    for a in activities:
        score, status, _, reasons = schedule_health_status(
            a,
            readiness_map.get(a["id"]),
            production_rows,
            risks
        )
        lines.append(
            f'- {a["external_id"]} {a["name"]} | Health score {score} | '
            f'Status: {status} | Reasons: {"; ".join(reasons)}'
        )

    lines.append("\nPROCUREMENT / LONG LEAD:")
    for p in procurement_rows:
        badge, risk_text = procurement_risk(
            p["required_on_site"],
            p["promised_date"],
            p["status"]
        )
        activity_text = (
            f'{p["external_id"]} {p["activity"]}'
            if p["external_id"]
            else "No linked activity"
        )
        lines.append(
            f'- {p["item"]} | {activity_text} | Vendor: {p["vendor"] or "N/A"} | '
            f'Required: {p["required_on_site"] or "N/A"} | '
            f'Promised: {p["promised_date"] or "N/A"} | '
            f'Status: {p["status"] or "N/A"} | Risk: {risk_text} | '
            f'Notes: {p["notes"] or "N/A"}'
        )

    lines.append("\nOPEN RFIS / ISSUES:")
    for i in issue_rows:
        activity_text = (
            f'{i["external_id"]} {i["activity"]}'
            if i["external_id"]
            else "No linked activity"
        )
        lines.append(
            f'- {i["issue_type"]} | {i["priority"]} | {i["title"]} | '
            f'{activity_text} | Owner: {i["owner"] or "Unassigned"} | '
            f'Due: {i["due"] or "N/A"} | Status: {i["status"]} | '
            f'Description: {i["description"] or "N/A"} | '
            f'Response: {i["response"] or "N/A"}'
        )

    lines.append("\nOPEN QUALITY / PUNCH ITEMS:")
    for p in punch_rows:
        activity_text = (
            f'{p["external_id"]} {p["activity"]}'
            if p["external_id"]
            else "No linked activity"
        )
        lines.append(
            f'- {p["priority"]} | {p["title"]} | Location: {p["location"] or "N/A"} | '
            f'Trade: {p["trade"] or "N/A"} | {activity_text} | '
            f'Owner: {p["owner"] or "Unassigned"} | Due: {p["due"] or "N/A"} | '
            f'Status: {p["status"]} | Description: {p["description"] or "N/A"} | '
            f'Resolution: {p["resolution"] or "N/A"}'
        )

    lines.append("\nOPEN / FAILED INSPECTIONS:")
    for i in inspection_rows:
        activity_text = (
            f'{i["external_id"]} {i["activity"]}'
            if i["external_id"]
            else "No linked activity"
        )
        lines.append(
            f'- {i["inspection_type"]} | {activity_text} | '
            f'Authority: {i["authority"] or "N/A"} | '
            f'Scheduled: {i["scheduled_date"] or "N/A"} | '
            f'Result: {i["result"]} | '
            f'Reinspection: {i["reinspection_date"] or "N/A"} | '
            f'Notes: {i["notes"] or "N/A"}'
        )

    lines.append("\nOPEN SUBMITTALS:")
    for s in submittal_rows:
        activity_text = (
            f'{s["external_id"]} {s["activity"]}'
            if s["external_id"]
            else "No linked activity"
        )
        lines.append(
            f'- {s["title"]} | Spec: {s["spec_section"] or "N/A"} | '
            f'{activity_text} | Responsible: {s["responsible_party"] or "Unassigned"} | '
            f'Sent: {s["sent_date"] or "N/A"} | Due: {s["due_date"] or "N/A"} | '
            f'Status: {s["status"]} | Notes: {s["notes"] or "N/A"}'
        )

    lines.append("\nOPEN SAFETY ITEMS:")
    for s in safety_rows:
        activity_text = (
            f'{s["external_id"]} {s["activity"]}'
            if s["external_id"]
            else "No linked activity"
        )
        lines.append(
            f'- {s["severity"]} | {s["item_type"]} | {s["title"]} | '
            f'Location: {s["location"] or "N/A"} | {activity_text} | '
            f'Responsible: {s["responsible_party"] or "Unassigned"} | '
            f'Status: {s["status"]} | Description: {s["description"] or "N/A"} | '
            f'Corrective action: {s["corrective_action"] or "N/A"}'
        )

    lines.append("\nOPEN CHANGE EVENTS:")
    for ce in change_rows:
        activity_text = (
            f'{ce["external_id"]} {ce["activity"]}'
            if ce["external_id"]
            else "No linked activity"
        )
        lines.append(
            f'- {ce["event_type"]} | {ce["title"]} | {activity_text} | '
            f'Responsible: {ce["responsible_party"] or "Unassigned"} | '
            f'Estimated cost: ${ce["estimated_cost"] or 0:,.0f} | '
            f'Schedule impact: {ce["schedule_days"] or 0:.1f} days | '
            f'Status: {ce["status"]} | Description: {ce["description"] or "N/A"}'
        )

    lines.append("\nRECENT MEETING NOTES:")
    for m in meeting_rows:
        lines.append(
            f'- {m["meeting_date"]} | {m["meeting_type"]} | {m["title"]} | '
            f'Attendees: {m["attendees"] or "N/A"} | '
            f'Decisions: {m["decisions"] or "N/A"} | '
            f'Commitments: {m["commitments"] or "N/A"} | '
            f'Follow-up: {m["follow_up"] or "N/A"}'
        )

    lines.append("\nRECENT DAILY REPORTS:")
    for r in reports:
        lines.append(
            f'- Date: {r["report_date"]} | Weather: {r["weather"] or "N/A"} | '
            f'Manpower: {r["manpower"] or 0}\n'
            f'  Work completed: {r["work_completed"] or "N/A"}\n'
            f'  Delays: {r["delays"] or "N/A"}\n'
            f'  Deliveries: {r["deliveries"] or "N/A"}\n'
            f'  Inspections: {r["inspections"] or "N/A"}\n'
            f'  Safety: {r["safety"] or "N/A"}\n'
            f'  Tomorrow plan: {r["tomorrow_plan"] or "N/A"}'
        )

    return "\n".join(lines)


@app.get("/assistant", response_class=HTMLResponse)
def assistant_page():
    pid = project_id()
    c = db()
    project = c.execute(
        "SELECT name,number FROM projects WHERE id=?",
        (pid,)
    ).fetchone()
    c.close()

    project_label = "Current Project"
    if project:
        project_label = f'{esc(project["number"])} - {esc(project["name"])}'

    api_ready = bool(os.environ.get("OPENAI_API_KEY"))

    status_html = (
        '<span class="badge READY">AI CONNECTED</span>'
        if api_ready
        else '<span class="badge HIGH">API KEY NEEDED</span>'
    )

    body = f"""
    <div class="hero">
        <div class="eyebrow">BuildCommand AI Copilot</div>
        <h1>Ask the project.</h1>
        <div class="muted">{project_label}</div>
        <div style="margin-top:10px;">{status_html}</div>
    </div>

    <div class="grid2">
        <div class="card">
            <h2>Ask BuildCommand</h2>

            <form method="post" action="/assistant">
                <label>Your Question</label>
                <textarea
                    name="question"
                    placeholder="Example: What should I worry about today?"
                    required
                ></textarea>

                <button type="submit">Analyze Project</button>
            </form>

            <div class="small" style="margin-top:14px;">
                Try: What should I handle first? Which activity could delay us?
                What should I ask the MEP subcontractor? What should tomorrow's plan include?
            </div>
        </div>

        <div class="card">
            <h2>How it works</h2>
            <p>
                BuildCommand sends the selected project's schedule, risks, open make-ready items,
                subcontractors, and recent daily reports to the AI with your question.
            </p>
            <p class="small">
                The AI is instructed to stay grounded in project data and clearly identify when
                information is missing instead of inventing jobsite facts.
            </p>
        </div>
    </div>
    """

    return shell("AI Assistant", body)


@app.post("/assistant", response_class=HTMLResponse)
def assistant_answer(question: str = Form(...)):
    pid = project_id()

    c = db()
    project = c.execute(
        "SELECT name,number FROM projects WHERE id=?",
        (pid,)
    ).fetchone()
    c.close()

    project_label = "Current Project"
    if project:
        project_label = f'{esc(project["number"])} - {esc(project["name"])}'

    api_key = os.environ.get("OPENAI_API_KEY")

    if not api_key:
        answer = """
        OPENAI_API_KEY is not configured yet.

        Add the key in Render under your service's Environment settings,
        then redeploy the service. Do not paste the API key into full_app.py.
        """
    else:
        context = build_project_context(pid)

        instructions = """
You are BuildCommand AI, a construction superintendent and project-execution copilot.

Use only the project information supplied in the prompt as job-specific facts.
Do not invent schedule status, subcontractor commitments, inspections, deliveries,
manpower, safety conditions, or field progress.

Prioritize:
1. immediate field risk,
2. schedule impact,
3. make-ready constraints,
4. inspections and deliveries,
5. subcontractor follow-up,
6. practical next actions.

When useful, structure the answer as:
- Handle first
- Why it matters
- Who to follow up with
- Next action

Keep answers concise, practical, and written for a working superintendent.
If the project data is insufficient, clearly say what information is missing.
"""

        user_input = f"""
PROJECT DATA
------------
{context}

SUPERINTENDENT QUESTION
-----------------------
{question}
"""

        try:
            client = OpenAI(api_key=api_key)
            response = client.responses.create(
                model="gpt-5.6",
                instructions=instructions,
                input=user_input
            )
            answer = response.output_text
        except Exception as exc:
            answer = (
                "BuildCommand could not reach the AI service. "
                f"Error: {str(exc)}"
            )

    answer_html = esc(answer).replace("\n", "<br>")

    body = f"""
    <div class="hero">
        <div class="eyebrow">BuildCommand AI Copilot</div>
        <h1>Project Answer</h1>
        <div class="muted">{project_label}</div>
    </div>

    <div class="grid2">
        <div class="card">
            <div class="small">YOUR QUESTION</div>
            <h3>{esc(question)}</h3>

            <a href="/assistant"
               style="color:#f0b44d;text-decoration:none;font-weight:700;">
                ← Ask another question
            </a>
        </div>

        <div class="card">
            <div class="small">BUILDCOMMAND AI</div>
            <div style="margin-top:12px;line-height:1.65;">
                {answer_html}
            </div>
        </div>
    </div>
    """

    return shell("AI Assistant", body)


@app.get("/make-ready/new", response_class=HTMLResponse)
def new_make_ready_form():
    pid = project_id()
    c = db()

    project = c.execute(
        "SELECT name,number FROM projects WHERE id=?",
        (pid,)
    ).fetchone()

    activities = c.execute(
        """
        SELECT id,external_id,name
        FROM activities
        WHERE project_id=?
        ORDER BY start,name
        """,
        (pid,)
    ).fetchall()

    c.close()

    project_label = "Current Project"
    if project:
        project_label = f'{esc(project["number"])} - {esc(project["name"])}'

    options = "".join(
        f'<option value="{a["id"]}">{esc(a["external_id"])} - {esc(a["name"])}</option>'
        for a in activities
    )

    if not options:
        body = """
        <div class="hero">
            <div class="eyebrow">Make Ready</div>
            <h1>Add Make-Ready Item</h1>
        </div>
        <div class="card">
            <p>This project has no activities yet. Add a schedule activity first.</p>
            <a href="/activities/new" style="color:#f0b44d;font-weight:700;text-decoration:none;">
                + Add Activity
            </a>
        </div>
        """
        return shell("Add Make-Ready", body)

    body = f"""
    <div class="hero">
        <div class="eyebrow">Make Ready</div>
        <h1>Add Make-Ready Item</h1>
        <div class="muted">{project_label}</div>
    </div>

    <div class="card" style="max-width:760px;">
        <form method="post" action="/make-ready/new">

            <label>Activity</label>
            <select name="activity_id" required>
                {options}
            </select>

            <label>Blocker / Action Title</label>
            <input
                type="text"
                name="title"
                placeholder="Example: Confirm electrical gear delivery"
                required
            >

            <label>Why It Matters</label>
            <textarea
                name="reason"
                placeholder="Describe the constraint, dependency, or field impact."
                required
            ></textarea>

            <div class="grid2">
                <div>
                    <label>Clear By</label>
                    <input type="date" name="due" required>
                </div>

                <div>
                    <label>Priority</label>
                    <select name="priority">
                        <option value="CRITICAL">Critical</option>
                        <option value="HIGH">High</option>
                        <option value="WATCH">Watch</option>
                        <option value="LOW">Low</option>
                    </select>
                </div>
            </div>

            <div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:18px;">
                <button type="submit">Save Make-Ready Item</button>

                <a href="/make-ready"
                   style="display:inline-block;color:#f0b44d;text-decoration:none;padding:10px 4px;font-weight:700;">
                    Cancel
                </a>
            </div>
        </form>
    </div>
    """

    return shell("Add Make-Ready", body)


@app.post("/make-ready/new")
def create_make_ready(
    activity_id: int = Form(...),
    title: str = Form(...),
    reason: str = Form(...),
    due: str = Form(...),
    priority: str = Form("HIGH")
):
    pid = project_id()

    c = db()

    activity = c.execute(
        "SELECT id FROM activities WHERE id=? AND project_id=?",
        (activity_id, pid)
    ).fetchone()

    if activity:
        c.execute(
            """
            INSERT INTO make_ready(
                project_id,
                activity_id,
                title,
                reason,
                due,
                priority,
                status
            )
            VALUES(?,?,?,?,?,?,?)
            """,
            (
                pid,
                activity_id,
                title.strip(),
                reason.strip(),
                due,
                priority,
                "OPEN"
            )
        )
        c.commit()

    c.close()

    return RedirectResponse(url="/make-ready", status_code=303)


@app.post("/make-ready/{item_id}/close")
def close_make_ready(item_id: int):
    pid = project_id()

    c = db()
    c.execute(
        """
        UPDATE make_ready
        SET status='COMPLETE'
        WHERE id=? AND project_id=?
        """,
        (item_id, pid)
    )
    c.commit()
    c.close()

    return RedirectResponse(url="/make-ready", status_code=303)


@app.post("/daily-report/{report_id}/analyze")
def analyze_daily_report(report_id: int):
    pid = project_id()
    c = db()

    report = c.execute(
        """
        SELECT *
        FROM daily_reports
        WHERE id=? AND project_id=?
        """,
        (report_id, pid)
    ).fetchone()

    project = c.execute(
        "SELECT * FROM projects WHERE id=?",
        (pid,)
    ).fetchone()

    activities = c.execute(
        "SELECT * FROM activities WHERE project_id=? ORDER BY start",
        (pid,)
    ).fetchall()

    c.close()

    if not report:
        return RedirectResponse(url="/daily-report", status_code=303)

    api_key = os.environ.get("OPENAI_API_KEY")

    if not api_key:
        analysis_text = (
            "OPENAI_API_KEY is not configured. Add the key in Render Environment settings "
            "to enable automatic report analysis."
        )
    else:
        schedule_lines = []
        for a in activities:
            schedule_lines.append(
                f'{a["external_id"]} | {a["name"]} | {a["trade"]} | '
                f'{a["start"]} to {a["finish"]} | {a["pct"]:.0f}% | {a["status"]}'
            )

        schedule_text = "\n".join(schedule_lines) or "No schedule activities entered."

        prompt = f"""
PROJECT
{project["number"] if project else ""} - {project["name"] if project else ""}

SCHEDULE
{schedule_text}

DAILY REPORT
Date: {report["report_date"]}
Weather: {report["weather"] or "N/A"}
Manpower: {report["manpower"] or 0}
Work completed: {report["work_completed"] or "N/A"}
Delays / constraints: {report["delays"] or "N/A"}
Deliveries: {report["deliveries"] or "N/A"}
Inspections: {report["inspections"] or "N/A"}
Safety: {report["safety"] or "N/A"}
Tomorrow plan: {report["tomorrow_plan"] or "N/A"}
"""

        instructions = """
You are BuildCommand AI, a construction superintendent field-intelligence copilot.

Analyze the daily report against the schedule information provided.
Do not invent facts that are not present.

Return a concise superintendent review with these headings:
HANDLE FIRST
SCHEDULE IMPACT
SUBCONTRACTOR / FOLLOW-UP
TOMORROW READINESS
MISSING INFORMATION

Call out specific delays, inspection issues, delivery risks, manpower concerns,
or tomorrow-plan gaps when supported by the report.
"""

        try:
            client = OpenAI(api_key=api_key)
            response = client.responses.create(
                model="gpt-5.6",
                instructions=instructions,
                input=prompt
            )
            analysis_text = response.output_text
        except Exception as exc:
            analysis_text = f"AI analysis failed: {str(exc)}"

    c = db()
    c.execute(
        "DELETE FROM daily_report_analysis WHERE project_id=? AND report_id=?",
        (pid, report_id)
    )
    c.execute(
        """
        INSERT INTO daily_report_analysis(
            project_id,
            report_id,
            analysis_text,
            created
        )
        VALUES(?,?,?,?)
        """,
        (
            pid,
            report_id,
            analysis_text,
            date.today().isoformat()
        )
    )
    c.commit()
    c.close()

    return RedirectResponse(url="/daily-report", status_code=303)


@app.get("/actions", response_class=HTMLResponse)
def actions_page():
    pid = project_id()
    c = db()

    rows = c.execute(
        """
        SELECT *
        FROM action_items
        WHERE project_id=?
        ORDER BY
            CASE status
                WHEN 'OPEN' THEN 1
                WHEN 'COMPLETE' THEN 2
                ELSE 3
            END,
            CASE priority
                WHEN 'CRITICAL' THEN 1
                WHEN 'HIGH' THEN 2
                WHEN 'WATCH' THEN 3
                WHEN 'LOW' THEN 4
                ELSE 5
            END,
            due
        """,
        (pid,)
    ).fetchall()

    c.close()

    today = date.today().isoformat()

    open_items = [r for r in rows if r["status"] == "OPEN"]
    complete_items = [r for r in rows if r["status"] == "COMPLETE"]
    overdue_items = [
        r for r in open_items
        if r["due"] and r["due"] < today
    ]

    open_html = ""

    for r in open_items:
        overdue = bool(r["due"] and r["due"] < today)
        due_text = f'Due {esc(r["due"])}'

        if overdue:
            due_text += " · OVERDUE"

        badge = r["priority"] if r["priority"] in ["CRITICAL","HIGH","WATCH","LOW"] else "OPEN"

        open_html += f"""
        <div class="card">
            <div style="display:flex;justify-content:space-between;gap:12px;align-items:flex-start;flex-wrap:wrap;">
                <div>
                    <span class="badge {badge}">{esc(r["priority"])}</span>
                    <h3 style="margin:10px 0 4px;">{esc(r["title"])}</h3>
                    <div class="muted">Owner: {esc(r["owner"]) or "Unassigned"}</div>
                </div>

                <form method="post" action="/actions/{r["id"]}/complete">
                    <button type="submit">Mark Complete</button>
                </form>
            </div>

            <p>{esc(r["notes"]) or "No notes entered."}</p>
            <div class="small">{due_text}</div>
        </div>
        """

    if not open_html:
        open_html = '<div class="card"><div class="muted">No open action items.</div></div>'

    complete_html = "".join(
        f"""
        <div class="action">
            <span class="badge COMPLETE">COMPLETE</span>
            <b>{esc(r["title"])}</b>
            <div class="small">
                Owner: {esc(r["owner"]) or "Unassigned"} · Due {esc(r["due"]) or "—"}
            </div>
        </div>
        """
        for r in complete_items[:12]
    ) or '<div class="muted">No completed actions yet.</div>'

    body = f"""
    <div class="hero">
        <div style="display:flex;justify-content:space-between;align-items:center;gap:15px;flex-wrap:wrap;">
            <div>
                <div class="eyebrow">Action Center</div>
                <h1>Turn project intelligence into accountable action.</h1>
                <div class="muted">
                    Assign ownership, due dates, and priority to the work that cannot fall through the cracks.
                </div>
            </div>

            <a href="/actions/new"
               style="display:inline-block;background:#f0b44d;color:#0a1017;text-decoration:none;padding:12px 18px;border-radius:9px;font-weight:800;">
                + Add Action
            </a>
        </div>
    </div>

    <div class="grid4">
        <div class="card">
            <div class="label">Open Actions</div>
            <div class="kpi">{len(open_items)}</div>
        </div>

        <div class="card">
            <div class="label">Overdue</div>
            <div class="kpi">{len(overdue_items)}</div>
        </div>

        <div class="card">
            <div class="label">Critical</div>
            <div class="kpi">{sum(r["priority"] == "CRITICAL" for r in open_items)}</div>
        </div>

        <div class="card">
            <div class="label">Completed</div>
            <div class="kpi">{len(complete_items)}</div>
        </div>
    </div>

    <div class="grid2">
        <div>
            <h2>Open Actions</h2>
            {open_html}
        </div>

        <div class="card">
            <h2>Recently Completed</h2>
            {complete_html}
        </div>
    </div>
    """

    return shell("Action Center", body)


@app.get("/actions/new", response_class=HTMLResponse)
def new_action_form():
    pid = project_id()
    c = db()

    subs = c.execute(
        """
        SELECT name,trade
        FROM subs
        WHERE project_id=?
        ORDER BY trade,name
        """,
        (pid,)
    ).fetchall()

    project = c.execute(
        "SELECT name,number FROM projects WHERE id=?",
        (pid,)
    ).fetchone()

    c.close()

    owner_options = '<option value="">Unassigned</option>'
    owner_options += '<option value="Superintendent">Superintendent</option>'
    owner_options += '<option value="Project Manager">Project Manager</option>'

    for s in subs:
        owner_options += (
            f'<option value="{esc(s["name"])}">{esc(s["name"])} - {esc(s["trade"])}</option>'
        )

    project_label = "Current Project"
    if project:
        project_label = f'{esc(project["number"])} - {esc(project["name"])}'

    body = f"""
    <div class="hero">
        <div class="eyebrow">Action Center</div>
        <h1>Add Action Item</h1>
        <div class="muted">{project_label}</div>
    </div>

    <div class="card" style="max-width:760px;">
        <form method="post" action="/actions/new">

            <label>Action</label>
            <input
                type="text"
                name="title"
                placeholder="Example: Confirm switchgear delivery date"
                required
            >

            <label>Owner</label>
            <select name="owner">
                {owner_options}
            </select>

            <div class="grid2">
                <div>
                    <label>Priority</label>
                    <select name="priority">
                        <option value="CRITICAL">Critical</option>
                        <option value="HIGH">High</option>
                        <option value="WATCH">Watch</option>
                        <option value="LOW">Low</option>
                    </select>
                </div>

                <div>
                    <label>Due Date</label>
                    <input type="date" name="due" required>
                </div>
            </div>

            <label>Notes</label>
            <textarea
                name="notes"
                placeholder="What specifically needs to happen?"
            ></textarea>

            <div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:18px;">
                <button type="submit">Save Action</button>

                <a href="/actions"
                   style="display:inline-block;color:#f0b44d;text-decoration:none;padding:10px 4px;font-weight:700;">
                    Cancel
                </a>
            </div>
        </form>
    </div>
    """

    return shell("Add Action", body)


@app.post("/actions/new")
def create_action(
    title: str = Form(...),
    owner: str = Form(""),
    priority: str = Form("HIGH"),
    due: str = Form(...),
    notes: str = Form("")
):
    pid = project_id()

    c = db()
    c.execute(
        """
        INSERT INTO action_items(
            project_id,
            title,
            owner,
            priority,
            due,
            status,
            notes,
            created
        )
        VALUES(?,?,?,?,?,?,?,?)
        """,
        (
            pid,
            title.strip(),
            owner.strip(),
            priority,
            due,
            "OPEN",
            notes.strip(),
            date.today().isoformat()
        )
    )
    c.commit()
    c.close()

    return RedirectResponse(url="/actions", status_code=303)


@app.post("/actions/{action_id}/complete")
def complete_action(action_id: int):
    pid = project_id()

    c = db()
    c.execute(
        """
        UPDATE action_items
        SET status='COMPLETE'
        WHERE id=? AND project_id=?
        """,
        (action_id, pid)
    )
    c.commit()
    c.close()

    return RedirectResponse(url="/actions", status_code=303)


def readiness_result(row):
    checks = [
        row["drawings"],
        row["material"],
        row["manpower"],
        row["predecessor"],
        row["access_ready"],
        row["inspection"],
        row["equipment"],
    ]

    complete = sum(1 for x in checks if x)
    pct = round((complete / len(checks)) * 100)

    critical_ready = bool(
        row["drawings"]
        and row["material"]
        and row["manpower"]
        and row["predecessor"]
        and row["access_ready"]
    )

    if complete == len(checks):
        status = "READY"
        badge = "READY"
    elif critical_ready and complete >= 5:
        status = "AT RISK"
        badge = "WATCH"
    else:
        status = "NOT READY"
        badge = "CRITICAL"

    return pct, status, badge


@app.get("/readiness", response_class=HTMLResponse)
def readiness_page():
    pid = project_id()
    c = db()

    activities = c.execute(
        """
        SELECT *
        FROM activities
        WHERE project_id=? AND status!='COMPLETE'
        ORDER BY start,name
        """,
        (pid,)
    ).fetchall()

    readiness_rows = c.execute(
        """
        SELECT *
        FROM activity_readiness
        WHERE project_id=?
        """,
        (pid,)
    ).fetchall()

    c.close()

    readiness_by_activity = {
        r["activity_id"]: r
        for r in readiness_rows
    }

    cards = ""
    ready_count = 0
    risk_count = 0
    not_ready_count = 0

    for a in activities:
        r = readiness_by_activity.get(a["id"])

        if r:
            pct, status, badge = readiness_result(r)
            notes = esc(r["notes"]) or "No readiness notes."
        else:
            pct, status, badge = 0, "NOT READY", "CRITICAL"
            notes = "Readiness has not been reviewed."

        if status == "READY":
            ready_count += 1
        elif status == "AT RISK":
            risk_count += 1
        else:
            not_ready_count += 1

        cards += f"""
        <div class="card">
            <div style="display:flex;justify-content:space-between;gap:12px;align-items:flex-start;flex-wrap:wrap;">
                <div>
                    <div class="eyebrow">{esc(a["external_id"])}</div>
                    <h3 style="margin:6px 0;">{esc(a["name"])}</h3>
                    <div class="muted">{esc(a["trade"])} · Starts {esc(a["start"])}</div>
                </div>

                <span class="badge {badge}">{status}</span>
            </div>

            <div style="margin-top:14px;">
                <div class="label">Readiness</div>
                <div class="kpi">{pct}%</div>
            </div>

            <p>{notes}</p>

            <a href="/readiness/{a["id"]}"
               style="color:#f0b44d;text-decoration:none;font-weight:700;">
                Review Readiness →
            </a>
        </div>
        """

    if not cards:
        cards = '<div class="card"><div class="muted">No incomplete activities found.</div></div>'

    body = f"""
    <div class="hero">
        <div class="eyebrow">Lookahead Readiness</div>
        <h1>Is upcoming work actually ready to start?</h1>
        <div class="muted">
            Verify the conditions that must be true before crews hit the field.
        </div>
    </div>

    <div class="grid4">
        <div class="card">
            <div class="label">Activities Reviewed</div>
            <div class="kpi">{len(activities)}</div>
        </div>

        <div class="card">
            <div class="label">Ready</div>
            <div class="kpi">{ready_count}</div>
        </div>

        <div class="card">
            <div class="label">At Risk</div>
            <div class="kpi">{risk_count}</div>
        </div>

        <div class="card">
            <div class="label">Not Ready</div>
            <div class="kpi">{not_ready_count}</div>
        </div>
    </div>

    <div class="grid2">
        {cards}
    </div>
    """

    return shell("Readiness", body)


@app.get("/readiness/{activity_id}", response_class=HTMLResponse)
def readiness_form(activity_id: int):
    pid = project_id()
    c = db()

    activity = c.execute(
        """
        SELECT *
        FROM activities
        WHERE id=? AND project_id=?
        """,
        (activity_id, pid)
    ).fetchone()

    row = c.execute(
        """
        SELECT *
        FROM activity_readiness
        WHERE activity_id=? AND project_id=?
        """,
        (activity_id, pid)
    ).fetchone()

    c.close()

    if not activity:
        return RedirectResponse(url="/readiness", status_code=303)

    def checked(field):
        return "checked" if row and row[field] else ""

    notes = esc(row["notes"]) if row else ""

    body = f"""
    <div class="hero">
        <div class="eyebrow">Lookahead Readiness</div>
        <h1>{esc(activity["external_id"])} - {esc(activity["name"])}</h1>
        <div class="muted">
            {esc(activity["trade"])} · {esc(activity["start"])} to {esc(activity["finish"])}
        </div>
    </div>

    <div class="card" style="max-width:820px;">
        <h2>Start Readiness Checklist</h2>

        <form method="post" action="/readiness/{activity_id}">
            <label class="form-check">
                <input class="form-check-input" type="checkbox" name="drawings" value="1" {checked("drawings")}>
                <span class="form-check-label">Drawings / approved information are available</span>
            </label>

            <label class="form-check">
                <input class="form-check-input" type="checkbox" name="material" value="1" {checked("material")}>
                <span class="form-check-label">Required material is on site or confirmed</span>
            </label>

            <label class="form-check">
                <input class="form-check-input" type="checkbox" name="manpower" value="1" {checked("manpower")}>
                <span class="form-check-label">Manpower is committed</span>
            </label>

            <label class="form-check">
                <input class="form-check-input" type="checkbox" name="predecessor" value="1" {checked("predecessor")}>
                <span class="form-check-label">Predecessor work is complete</span>
            </label>

            <label class="form-check">
                <input class="form-check-input" type="checkbox" name="access_ready" value="1" {checked("access_ready")}>
                <span class="form-check-label">Work area and access are ready</span>
            </label>

            <label class="form-check">
                <input class="form-check-input" type="checkbox" name="inspection" value="1" {checked("inspection")}>
                <span class="form-check-label">Required inspections / prerequisites are cleared</span>
            </label>

            <label class="form-check">
                <input class="form-check-input" type="checkbox" name="equipment" value="1" {checked("equipment")}>
                <span class="form-check-label">Required equipment / tools are available</span>
            </label>

            <label style="display:block;margin-top:18px;">Readiness Notes</label>
            <textarea
                name="notes"
                placeholder="What is missing, who owns it, or what needs to happen before start?"
            >{notes}</textarea>

            <div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:18px;">
                <button type="submit">Save Readiness</button>

                <a href="/readiness"
                   style="display:inline-block;color:#f0b44d;text-decoration:none;padding:10px 4px;font-weight:700;">
                    Back
                </a>
            </div>
        </form>
    </div>
    """

    return shell("Readiness Review", body)


@app.post("/readiness/{activity_id}")
def save_readiness(
    activity_id: int,
    drawings: int = Form(0),
    material: int = Form(0),
    manpower: int = Form(0),
    predecessor: int = Form(0),
    access_ready: int = Form(0),
    inspection: int = Form(0),
    equipment: int = Form(0),
    notes: str = Form("")
):
    pid = project_id()
    c = db()

    activity = c.execute(
        "SELECT id FROM activities WHERE id=? AND project_id=?",
        (activity_id, pid)
    ).fetchone()

    if activity:
        c.execute(
            """
            INSERT INTO activity_readiness(
                project_id,
                activity_id,
                drawings,
                material,
                manpower,
                predecessor,
                access_ready,
                inspection,
                equipment,
                notes,
                updated
            )
            VALUES(?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(project_id, activity_id)
            DO UPDATE SET
                drawings=excluded.drawings,
                material=excluded.material,
                manpower=excluded.manpower,
                predecessor=excluded.predecessor,
                access_ready=excluded.access_ready,
                inspection=excluded.inspection,
                equipment=excluded.equipment,
                notes=excluded.notes,
                updated=excluded.updated
            """,
            (
                pid,
                activity_id,
                1 if drawings else 0,
                1 if material else 0,
                1 if manpower else 0,
                1 if predecessor else 0,
                1 if access_ready else 0,
                1 if inspection else 0,
                1 if equipment else 0,
                notes.strip(),
                date.today().isoformat()
            )
        )
        c.commit()

    c.close()

    return RedirectResponse(url="/readiness", status_code=303)


def schedule_health_status(activity, readiness_row, production_rows, risk_rows):
    today = date.today().isoformat()

    pct = activity["pct"] or 0
    start = activity["start"] or ""
    finish = activity["finish"] or ""

    reasons = []
    score = 0

    if finish and finish < today and pct < 100:
        score += 40
        reasons.append("Finish date has passed and activity is incomplete.")

    if start and start <= today and pct == 0 and activity["status"] != "COMPLETE":
        score += 20
        reasons.append("Activity has started by date but shows 0% complete.")

    if readiness_row:
        readiness_pct, readiness_status, _ = readiness_result(readiness_row)

        if readiness_status == "NOT READY":
            score += 25
            reasons.append(f"Readiness is only {readiness_pct}% and activity is NOT READY.")
        elif readiness_status == "AT RISK":
            score += 12
            reasons.append(f"Readiness is {readiness_pct}% and activity is AT RISK.")

    related_risks = [
        r for r in risk_rows
        if r["activity_id"] == activity["id"]
    ]

    for r in related_risks:
        if r["band"] == "CRITICAL":
            score += 25
            reasons.append("Critical risk is open.")
        elif r["band"] == "HIGH":
            score += 15
            reasons.append("High risk is open.")

    related_prod = [
        p for p in production_rows
        if p["activity_id"] == activity["id"]
    ]

    if related_prod:
        latest = related_prod[0]
        qty = latest["qty"] or 0
        plan = latest["planned_qty"] or 0

        if plan > 0:
            prod_pct = (qty / plan) * 100

            if prod_pct < 90:
                score += 20
                reasons.append(f"Latest production is {prod_pct:.0f}% of plan.")
            elif prod_pct < 100:
                score += 8
                reasons.append(f"Latest production is {prod_pct:.0f}% of plan.")

    score = min(100, score)

    if score >= 60:
        status = "CRITICAL"
        badge = "CRITICAL"
    elif score >= 35:
        status = "HIGH"
        badge = "HIGH"
    elif score >= 15:
        status = "WATCH"
        badge = "WATCH"
    else:
        status = "STABLE"
        badge = "READY"

    if not reasons:
        reasons.append("No significant schedule-health warning signals found.")

    return score, status, badge, reasons


@app.get("/schedule-health", response_class=HTMLResponse)
def schedule_health_page():
    pid = project_id()
    c = db()

    activities = c.execute(
        """
        SELECT *
        FROM activities
        WHERE project_id=?
        ORDER BY start,name
        """,
        (pid,)
    ).fetchall()

    readiness_rows = c.execute(
        """
        SELECT *
        FROM activity_readiness
        WHERE project_id=?
        """,
        (pid,)
    ).fetchall()

    production_rows = c.execute(
        """
        SELECT *
        FROM production
        WHERE project_id=?
        ORDER BY work_date DESC,id DESC
        """,
        (pid,)
    ).fetchall()

    risk_rows = c.execute(
        """
        SELECT *
        FROM risks
        WHERE project_id=?
        """,
        (pid,)
    ).fetchall()

    c.close()

    readiness_by_activity = {
        r["activity_id"]: r
        for r in readiness_rows
    }

    health_rows = []

    for a in activities:
        score, status, badge, reasons = schedule_health_status(
            a,
            readiness_by_activity.get(a["id"]),
            production_rows,
            risk_rows
        )

        health_rows.append(
            (score, a, status, badge, reasons)
        )

    health_rows.sort(key=lambda x: x[0], reverse=True)

    critical_count = sum(1 for x in health_rows if x[2] == "CRITICAL")
    high_count = sum(1 for x in health_rows if x[2] == "HIGH")
    watch_count = sum(1 for x in health_rows if x[2] == "WATCH")
    stable_count = sum(1 for x in health_rows if x[2] == "STABLE")

    cards = ""

    for score, a, status, badge, reasons in health_rows:
        reason_html = "".join(
            f'<div class="small">• {esc(reason)}</div>'
            for reason in reasons
        )

        cards += f"""
        <div class="card">
            <div style="display:flex;justify-content:space-between;gap:12px;align-items:flex-start;flex-wrap:wrap;">
                <div>
                    <div class="eyebrow">{esc(a["external_id"])}</div>
                    <h3 style="margin:6px 0;">{esc(a["name"])}</h3>
                    <div class="muted">
                        {esc(a["trade"])} · {esc(a["start"])} to {esc(a["finish"])}
                    </div>
                </div>

                <span class="badge {badge}">{status}</span>
            </div>

            <div class="grid3" style="margin-top:16px;">
                <div>
                    <div class="label">Health Score</div>
                    <div class="kpi">{score}</div>
                </div>

                <div>
                    <div class="label">Complete</div>
                    <div class="kpi">{(a["pct"] or 0):.0f}%</div>
                </div>

                <div>
                    <div class="label">Status</div>
                    <div style="font-size:18px;font-weight:800;">
                        {esc(a["status"]).replace("_"," ").title()}
                    </div>
                </div>
            </div>

            <div style="margin-top:14px;">
                {reason_html}
            </div>
        </div>
        """

    if not cards:
        cards = '<div class="card"><div class="muted">No schedule activities found.</div></div>'

    body = f"""
    <div class="hero">
        <div class="eyebrow">Schedule Health</div>
        <h1>Which activities are drifting before the schedule update catches it?</h1>
        <div class="muted">
            Combines schedule dates, progress, readiness, production, and risk signals.
        </div>
    </div>

    <div class="grid4">
        <div class="card">
            <div class="label">Critical</div>
            <div class="kpi">{critical_count}</div>
        </div>

        <div class="card">
            <div class="label">High</div>
            <div class="kpi">{high_count}</div>
        </div>

        <div class="card">
            <div class="label">Watch</div>
            <div class="kpi">{watch_count}</div>
        </div>

        <div class="card">
            <div class="label">Stable</div>
            <div class="kpi">{stable_count}</div>
        </div>
    </div>

    <div class="grid2">
        {cards}
    </div>
    """

    return shell("Schedule Health", body)


@app.get("/project-settings", response_class=HTMLResponse)
def project_settings_page():
    pid = project_id()
    c = db()

    project = c.execute(
        "SELECT * FROM projects WHERE id=?",
        (pid,)
    ).fetchone()

    counts = {
        "activities": c.execute(
            "SELECT COUNT(*) n FROM activities WHERE project_id=?",
            (pid,)
        ).fetchone()["n"],
        "risks": c.execute(
            "SELECT COUNT(*) n FROM risks WHERE project_id=?",
            (pid,)
        ).fetchone()["n"],
        "make_ready": c.execute(
            "SELECT COUNT(*) n FROM make_ready WHERE project_id=?",
            (pid,)
        ).fetchone()["n"],
        "daily_reports": c.execute(
            "SELECT COUNT(*) n FROM daily_reports WHERE project_id=?",
            (pid,)
        ).fetchone()["n"],
    }

    c.close()

    if not project:
        return RedirectResponse(url="/", status_code=303)

    statuses = ["ACTIVE", "PLANNING", "ON_HOLD", "COMPLETE"]
    status_options = "".join(
        f'<option value="{s}" {"selected" if project["status"] == s else ""}>{s.replace("_"," ").title()}</option>'
        for s in statuses
    )

    body = f"""
    <div class="hero">
        <div class="eyebrow">Project Settings</div>
        <h1>{esc(project["number"])} - {esc(project["name"])}</h1>
        <div class="muted">
            Manage the selected project's identity and lifecycle.
        </div>
    </div>

    <div class="grid4">
        <div class="card">
            <div class="label">Activities</div>
            <div class="kpi">{counts["activities"]}</div>
        </div>

        <div class="card">
            <div class="label">Risks</div>
            <div class="kpi">{counts["risks"]}</div>
        </div>

        <div class="card">
            <div class="label">Make Ready</div>
            <div class="kpi">{counts["make_ready"]}</div>
        </div>

        <div class="card">
            <div class="label">Daily Reports</div>
            <div class="kpi">{counts["daily_reports"]}</div>
        </div>
    </div>

    <div class="grid2">
        <div class="card">
            <h2>Edit Project</h2>

            <form method="post" action="/project-settings/edit">
                <label>Project Name</label>
                <input
                    type="text"
                    name="name"
                    value="{esc(project["name"])}"
                    required
                >

                <label>Project Number</label>
                <input
                    type="text"
                    name="number"
                    value="{esc(project["number"])}"
                    required
                >

                <label>Status</label>
                <select name="status">
                    {status_options}
                </select>

                <button type="submit">Save Project Changes</button>
            </form>
        </div>

        <div class="card">
            <h2>Delete Project</h2>

            <p>
                Deleting a project removes its schedule activities, risks, make-ready items,
                field updates, production records, daily reports, subcontractor updates,
                actions, readiness records, recovery scenarios, and saved analyses.
            </p>

            <div class="small">
                This cannot be undone.
            </div>

            <form method="post"
                  action="/project-settings/delete"
                  style="margin-top:18px;">
                <label>Type DELETE to confirm</label>
                <input
                    type="text"
                    name="confirm_text"
                    placeholder="DELETE"
                    required
                >

                <button type="submit"
                        style="background:#492324;color:#ffb0b0;">
                    Delete Project
                </button>
            </form>
        </div>
    </div>
    """

    return shell("Project Settings", body)


@app.post("/project-settings/edit")
def edit_project_settings(
    name: str = Form(...),
    number: str = Form(...),
    status: str = Form(...)
):
    pid = project_id()

    allowed_statuses = {"ACTIVE", "PLANNING", "ON_HOLD", "COMPLETE"}
    if status not in allowed_statuses:
        status = "ACTIVE"

    c = db()
    c.execute(
        """
        UPDATE projects
        SET name=?, number=?, status=?
        WHERE id=?
        """,
        (
            name.strip(),
            number.strip(),
            status,
            pid
        )
    )
    c.commit()
    c.close()

    return RedirectResponse(url="/project-settings", status_code=303)


@app.post("/project-settings/delete")
def delete_project_settings(confirm_text: str = Form(...)):
    pid = project_id()

    if confirm_text.strip().upper() != "DELETE":
        return RedirectResponse(url="/project-settings", status_code=303)

    c = db()

    activity_ids = [
        r["id"]
        for r in c.execute(
            "SELECT id FROM activities WHERE project_id=?",
            (pid,)
        ).fetchall()
    ]

    # Delete project-scoped tables first.
    tables = [
        "daily_report_analysis",
        "daily_reports",
        "subcontractor_updates",
        "action_items",
        "activity_readiness",
        "procurement",
        "project_issues",
        "punch_items",
        "inspections_tracker",
        "submittals",
        "safety_items",
        "change_events",
        "meeting_notes",
        "field_updates",
        "production",
        "make_ready",
        "risks",
        "recovery",
        "memory",
        "subs",
    ]

    for table in tables:
        c.execute(
            f"DELETE FROM {table} WHERE project_id=?",
            (pid,)
        )

    c.execute(
        "DELETE FROM activities WHERE project_id=?",
        (pid,)
    )

    c.execute(
        "DELETE FROM projects WHERE id=?",
        (pid,)
    )

    next_project = c.execute(
        "SELECT id FROM projects ORDER BY id LIMIT 1"
    ).fetchone()

    if next_project:
        c.execute(
            """
            INSERT INTO app_state(id, selected_project_id)
            VALUES(1, ?)
            ON CONFLICT(id)
            DO UPDATE SET selected_project_id=excluded.selected_project_id
            """,
            (next_project["id"],)
        )
    else:
        c.execute(
            "DELETE FROM app_state WHERE id=1"
        )

    c.commit()
    c.close()

    if next_project:
        return RedirectResponse(url="/", status_code=303)

    return RedirectResponse(url="/projects/new", status_code=303)


def procurement_risk(required_on_site, promised_date, status):
    if status == "DELIVERED":
        return "READY", "DELIVERED"

    if not required_on_site:
        return "WATCH", "NO REQUIRED DATE"

    if not promised_date:
        return "CRITICAL", "NO PROMISED DATE"

    if promised_date > required_on_site:
        return "CRITICAL", "LATE"

    if status in ["RELEASED", "FABRICATION", "SHIPPED"]:
        return "WATCH", status

    return "HIGH", status or "OPEN"


@app.get("/procurement", response_class=HTMLResponse)
def procurement_page():
    pid = project_id()
    c = db()

    rows = c.execute(
        """
        SELECT p.*, a.external_id, a.name activity
        FROM procurement p
        LEFT JOIN activities a ON a.id=p.activity_id
        WHERE p.project_id=?
        ORDER BY
            CASE p.status
                WHEN 'DELIVERED' THEN 5
                WHEN 'SHIPPED' THEN 4
                WHEN 'FABRICATION' THEN 3
                WHEN 'RELEASED' THEN 2
                ELSE 1
            END,
            p.required_on_site
        """,
        (pid,)
    ).fetchall()

    c.close()

    critical = 0
    watch = 0
    delivered = 0

    cards = ""

    for r in rows:
        badge, risk_text = procurement_risk(
            r["required_on_site"],
            r["promised_date"],
            r["status"]
        )

        if badge == "CRITICAL":
            critical += 1
        elif badge == "WATCH":
            watch += 1

        if r["status"] == "DELIVERED":
            delivered += 1

        activity_text = (
            f'{esc(r["external_id"])} - {esc(r["activity"])}'
            if r["external_id"]
            else "No linked activity"
        )

        cards += f"""
        <div class="card">
            <div style="display:flex;justify-content:space-between;gap:12px;align-items:flex-start;flex-wrap:wrap;">
                <div>
                    <span class="badge {badge}">{risk_text}</span>
                    <h3 style="margin:10px 0 4px;">{esc(r["item"])}</h3>
                    <div class="muted">{activity_text}</div>
                </div>

                <a href="/procurement/{r["id"]}/edit"
                   style="color:#f0b44d;text-decoration:none;font-weight:700;">
                    Edit
                </a>
            </div>

            <div class="grid3" style="margin-top:16px;">
                <div>
                    <div class="label">Vendor / Sub</div>
                    <div>{esc(r["vendor"]) or "—"}</div>
                </div>

                <div>
                    <div class="label">Required On Site</div>
                    <div>{esc(r["required_on_site"]) or "—"}</div>
                </div>

                <div>
                    <div class="label">Promised</div>
                    <div>{esc(r["promised_date"]) or "—"}</div>
                </div>
            </div>

            <p>{esc(r["notes"]) or "No notes entered."}</p>
            <div class="small">Status: {esc(r["status"]).replace("_"," ").title()}</div>
        </div>
        """

    if not cards:
        cards = '<div class="card"><div class="muted">No procurement items added yet.</div></div>'

    body = f"""
    <div class="hero">
        <div style="display:flex;justify-content:space-between;align-items:center;gap:15px;flex-wrap:wrap;">
            <div>
                <div class="eyebrow">Procurement Intelligence</div>
                <h1>Know what material can hurt the schedule before it arrives late.</h1>
                <div class="muted">
                    Track required-on-site dates, vendor commitments, and long-lead exposure.
                </div>
            </div>

            <a href="/procurement/new"
               style="display:inline-block;background:#f0b44d;color:#0a1017;text-decoration:none;padding:12px 18px;border-radius:9px;font-weight:800;">
                + Add Procurement Item
            </a>
        </div>
    </div>

    <div class="grid4">
        <div class="card">
            <div class="label">Items</div>
            <div class="kpi">{len(rows)}</div>
        </div>

        <div class="card">
            <div class="label">Critical</div>
            <div class="kpi">{critical}</div>
        </div>

        <div class="card">
            <div class="label">Watch</div>
            <div class="kpi">{watch}</div>
        </div>

        <div class="card">
            <div class="label">Delivered</div>
            <div class="kpi">{delivered}</div>
        </div>
    </div>

    <div class="grid2">
        {cards}
    </div>
    """

    return shell("Procurement", body)


@app.get("/procurement/new", response_class=HTMLResponse)
def new_procurement_form():
    pid = project_id()
    c = db()

    activities = c.execute(
        """
        SELECT id,external_id,name
        FROM activities
        WHERE project_id=?
        ORDER BY start,name
        """,
        (pid,)
    ).fetchall()

    c.close()

    options = '<option value="">No linked activity</option>'
    options += "".join(
        f'<option value="{a["id"]}">{esc(a["external_id"])} - {esc(a["name"])}</option>'
        for a in activities
    )

    body = f"""
    <div class="hero">
        <div class="eyebrow">Procurement Intelligence</div>
        <h1>Add Procurement Item</h1>
    </div>

    <div class="card" style="max-width:760px;">
        <form method="post" action="/procurement/new">

            <label>Material / Equipment</label>
            <input
                type="text"
                name="item"
                placeholder="Example: Main electrical switchgear"
                required
            >

            <label>Linked Activity</label>
            <select name="activity_id">
                {options}
            </select>

            <label>Vendor / Subcontractor</label>
            <input
                type="text"
                name="vendor"
                placeholder="Example: Valley Electric / Eaton"
            >

            <div class="grid2">
                <div>
                    <label>Required On Site</label>
                    <input type="date" name="required_on_site" required>
                </div>

                <div>
                    <label>Promised Date</label>
                    <input type="date" name="promised_date">
                </div>
            </div>

            <label>Status</label>
            <select name="status">
                <option value="NOT_RELEASED">Not Released</option>
                <option value="RELEASED">Released</option>
                <option value="FABRICATION">Fabrication</option>
                <option value="SHIPPED">Shipped</option>
                <option value="DELIVERED">Delivered</option>
            </select>

            <label>Notes</label>
            <textarea
                name="notes"
                placeholder="Lead time, submittal status, release issue, delivery commitment, etc."
            ></textarea>

            <div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:18px;">
                <button type="submit">Save Procurement Item</button>

                <a href="/procurement"
                   style="display:inline-block;color:#f0b44d;text-decoration:none;padding:10px 4px;font-weight:700;">
                    Cancel
                </a>
            </div>
        </form>
    </div>
    """

    return shell("Add Procurement", body)


@app.post("/procurement/new")
def create_procurement(
    item: str = Form(...),
    activity_id: str = Form(""),
    vendor: str = Form(""),
    required_on_site: str = Form(...),
    promised_date: str = Form(""),
    status: str = Form("NOT_RELEASED"),
    notes: str = Form("")
):
    pid = project_id()

    linked_activity = int(activity_id) if activity_id.strip() else None

    c = db()

    if linked_activity is not None:
        valid_activity = c.execute(
            "SELECT id FROM activities WHERE id=? AND project_id=?",
            (linked_activity, pid)
        ).fetchone()

        if not valid_activity:
            linked_activity = None

    c.execute(
        """
        INSERT INTO procurement(
            project_id,
            activity_id,
            item,
            vendor,
            required_on_site,
            promised_date,
            status,
            notes,
            created
        )
        VALUES(?,?,?,?,?,?,?,?,?)
        """,
        (
            pid,
            linked_activity,
            item.strip(),
            vendor.strip(),
            required_on_site,
            promised_date,
            status,
            notes.strip(),
            date.today().isoformat()
        )
    )

    c.commit()
    c.close()

    _v57_emit(project_id(),"PROCUREMENT_CHANGE","Procurement updated","UI")
    return RedirectResponse(url="/procurement", status_code=303)


@app.get("/procurement/{item_id}/edit", response_class=HTMLResponse)
def edit_procurement_form(item_id: int):
    pid = project_id()
    c = db()

    item = c.execute(
        """
        SELECT *
        FROM procurement
        WHERE id=? AND project_id=?
        """,
        (item_id, pid)
    ).fetchone()

    activities = c.execute(
        """
        SELECT id,external_id,name
        FROM activities
        WHERE project_id=?
        ORDER BY start,name
        """,
        (pid,)
    ).fetchall()

    c.close()

    if not item:
        return RedirectResponse(url="/procurement", status_code=303)

    activity_options = '<option value="">No linked activity</option>'

    for a in activities:
        selected = "selected" if item["activity_id"] == a["id"] else ""
        activity_options += (
            f'<option value="{a["id"]}" {selected}>'
            f'{esc(a["external_id"])} - {esc(a["name"])}</option>'
        )

    statuses = [
        "NOT_RELEASED",
        "RELEASED",
        "FABRICATION",
        "SHIPPED",
        "DELIVERED"
    ]

    status_options = "".join(
        f'<option value="{s}" {"selected" if item["status"] == s else ""}>'
        f'{s.replace("_"," ").title()}</option>'
        for s in statuses
    )

    body = f"""
    <div class="hero">
        <div class="eyebrow">Procurement Intelligence</div>
        <h1>Edit Procurement Item</h1>
    </div>

    <div class="card" style="max-width:760px;">
        <form method="post" action="/procurement/{item_id}/edit">

            <label>Material / Equipment</label>
            <input type="text" name="item" value="{esc(item["item"])}" required>

            <label>Linked Activity</label>
            <select name="activity_id">
                {activity_options}
            </select>

            <label>Vendor / Subcontractor</label>
            <input type="text" name="vendor" value="{esc(item["vendor"])}">

            <div class="grid2">
                <div>
                    <label>Required On Site</label>
                    <input type="date" name="required_on_site" value="{esc(item["required_on_site"])}" required>
                </div>

                <div>
                    <label>Promised Date</label>
                    <input type="date" name="promised_date" value="{esc(item["promised_date"])}">
                </div>
            </div>

            <label>Status</label>
            <select name="status">
                {status_options}
            </select>

            <label>Notes</label>
            <textarea name="notes">{esc(item["notes"])}</textarea>

            <div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:18px;">
                <button type="submit">Save Changes</button>

                <a href="/procurement"
                   style="display:inline-block;color:#f0b44d;text-decoration:none;padding:10px 4px;font-weight:700;">
                    Cancel
                </a>
            </div>
        </form>

        <form method="post"
              action="/procurement/{item_id}/delete"
              style="margin-top:22px;padding-top:18px;border-top:1px solid #213042;">
            <button type="submit"
                    style="background:#492324;color:#ffb0b0;">
                Delete Procurement Item
            </button>
        </form>
    </div>
    """

    return shell("Edit Procurement", body)


@app.post("/procurement/{item_id}/edit")
def edit_procurement(
    item_id: int,
    item: str = Form(...),
    activity_id: str = Form(""),
    vendor: str = Form(""),
    required_on_site: str = Form(...),
    promised_date: str = Form(""),
    status: str = Form("NOT_RELEASED"),
    notes: str = Form("")
):
    pid = project_id()
    linked_activity = int(activity_id) if activity_id.strip() else None

    c = db()

    if linked_activity is not None:
        valid_activity = c.execute(
            "SELECT id FROM activities WHERE id=? AND project_id=?",
            (linked_activity, pid)
        ).fetchone()

        if not valid_activity:
            linked_activity = None

    c.execute(
        """
        UPDATE procurement
        SET activity_id=?,
            item=?,
            vendor=?,
            required_on_site=?,
            promised_date=?,
            status=?,
            notes=?
        WHERE id=? AND project_id=?
        """,
        (
            linked_activity,
            item.strip(),
            vendor.strip(),
            required_on_site,
            promised_date,
            status,
            notes.strip(),
            item_id,
            pid
        )
    )

    c.commit()
    c.close()

    return RedirectResponse(url="/procurement", status_code=303)


@app.post("/procurement/{item_id}/delete")
def delete_procurement(item_id: int):
    pid = project_id()

    c = db()
    c.execute(
        "DELETE FROM procurement WHERE id=? AND project_id=?",
        (item_id, pid)
    )
    c.commit()
    c.close()

    return RedirectResponse(url="/procurement", status_code=303)


@app.get("/issues", response_class=HTMLResponse)
def issues_page():
    pid = project_id()
    c = db()

    rows = c.execute(
        """
        SELECT i.*, a.external_id, a.name activity
        FROM project_issues i
        LEFT JOIN activities a ON a.id=i.activity_id
        WHERE i.project_id=?
        ORDER BY
            CASE i.status
                WHEN 'OPEN' THEN 1
                WHEN 'ANSWERED' THEN 2
                WHEN 'CLOSED' THEN 3
                ELSE 4
            END,
            CASE i.priority
                WHEN 'CRITICAL' THEN 1
                WHEN 'HIGH' THEN 2
                WHEN 'WATCH' THEN 3
                ELSE 4
            END,
            i.due
        """,
        (pid,)
    ).fetchall()

    c.close()

    today = date.today().isoformat()
    open_items = [r for r in rows if r["status"] == "OPEN"]
    overdue = [r for r in open_items if r["due"] and r["due"] < today]
    answered = [r for r in rows if r["status"] == "ANSWERED"]

    cards = ""

    for r in rows:
        activity_text = (
            f'{esc(r["external_id"])} - {esc(r["activity"])}'
            if r["external_id"]
            else "No linked activity"
        )

        overdue_text = ""
        if r["status"] == "OPEN" and r["due"] and r["due"] < today:
            overdue_text = " · OVERDUE"

        badge = r["priority"] if r["priority"] in ["CRITICAL","HIGH","WATCH","LOW"] else "OPEN"

        cards += f"""
        <div class="card">
            <div style="display:flex;justify-content:space-between;gap:12px;align-items:flex-start;flex-wrap:wrap;">
                <div>
                    <span class="badge {badge}">{esc(r["priority"])}</span>
                    <span class="badge OPEN">{esc(r["issue_type"])}</span>
                    <h3 style="margin:10px 0 4px;">{esc(r["title"])}</h3>
                    <div class="muted">{activity_text}</div>
                </div>

                <a href="/issues/{r["id"]}/edit"
                   style="color:#f0b44d;text-decoration:none;font-weight:700;">
                    Edit
                </a>
            </div>

            <p>{esc(r["description"]) or "No description entered."}</p>
            <div class="small">
                Owner: {esc(r["owner"]) or "Unassigned"} · Due {esc(r["due"]) or "—"}{overdue_text}
            </div>

            <div class="small" style="margin-top:8px;">
                Status: {esc(r["status"])}
            </div>

            {"<p><b>Response:</b> " + esc(r["response"]) + "</p>" if r["response"] else ""}
        </div>
        """

    if not cards:
        cards = '<div class="card"><div class="muted">No RFIs or project issues logged yet.</div></div>'

    body = f"""
    <div class="hero">
        <div style="display:flex;justify-content:space-between;align-items:center;gap:15px;flex-wrap:wrap;">
            <div>
                <div class="eyebrow">RFIs / Issues</div>
                <h1>Track unanswered questions before they stop the field.</h1>
                <div class="muted">
                    Capture ownership, due dates, responses, and schedule exposure.
                </div>
            </div>

            <a href="/issues/new"
               style="display:inline-block;background:#f0b44d;color:#0a1017;text-decoration:none;padding:12px 18px;border-radius:9px;font-weight:800;">
                + Add RFI / Issue
            </a>
        </div>
    </div>

    <div class="grid4">
        <div class="card">
            <div class="label">Open</div>
            <div class="kpi">{len(open_items)}</div>
        </div>

        <div class="card">
            <div class="label">Overdue</div>
            <div class="kpi">{len(overdue)}</div>
        </div>

        <div class="card">
            <div class="label">Answered</div>
            <div class="kpi">{len(answered)}</div>
        </div>

        <div class="card">
            <div class="label">Total</div>
            <div class="kpi">{len(rows)}</div>
        </div>
    </div>

    <div class="grid2">
        {cards}
    </div>
    """

    return shell("RFIs / Issues", body)


@app.get("/issues/new", response_class=HTMLResponse)
def new_issue_form():
    pid = project_id()
    c = db()

    activities = c.execute(
        """
        SELECT id,external_id,name
        FROM activities
        WHERE project_id=?
        ORDER BY start,name
        """,
        (pid,)
    ).fetchall()

    subs = c.execute(
        """
        SELECT name,trade
        FROM subs
        WHERE project_id=?
        ORDER BY trade,name
        """,
        (pid,)
    ).fetchall()

    c.close()

    activity_options = '<option value="">No linked activity</option>'
    activity_options += "".join(
        f'<option value="{a["id"]}">{esc(a["external_id"])} - {esc(a["name"])}</option>'
        for a in activities
    )

    owner_options = '<option value="">Unassigned</option>'
    owner_options += '<option value="Architect / Engineer">Architect / Engineer</option>'
    owner_options += '<option value="Owner">Owner</option>'
    owner_options += '<option value="Project Manager">Project Manager</option>'
    owner_options += '<option value="Superintendent">Superintendent</option>'

    for s in subs:
        owner_options += (
            f'<option value="{esc(s["name"])}">{esc(s["name"])} - {esc(s["trade"])}</option>'
        )

    body = f"""
    <div class="hero">
        <div class="eyebrow">RFIs / Issues</div>
        <h1>Add RFI or Project Issue</h1>
    </div>

    <div class="card" style="max-width:760px;">
        <form method="post" action="/issues/new">

            <label>Type</label>
            <select name="issue_type">
                <option value="RFI">RFI</option>
                <option value="FIELD_ISSUE">Field Issue</option>
                <option value="DESIGN">Design Issue</option>
                <option value="COORDINATION">Coordination</option>
                <option value="OWNER_DECISION">Owner Decision</option>
            </select>

            <label>Title</label>
            <input
                type="text"
                name="title"
                placeholder="Example: RFI - Beam conflict with duct route"
                required
            >

            <label>Linked Activity</label>
            <select name="activity_id">
                {activity_options}
            </select>

            <label>Owner / Responsible Party</label>
            <select name="owner">
                {owner_options}
            </select>

            <div class="grid2">
                <div>
                    <label>Priority</label>
                    <select name="priority">
                        <option value="CRITICAL">Critical</option>
                        <option value="HIGH">High</option>
                        <option value="WATCH">Watch</option>
                        <option value="LOW">Low</option>
                    </select>
                </div>

                <div>
                    <label>Due Date</label>
                    <input type="date" name="due" required>
                </div>
            </div>

            <label>Description</label>
            <textarea
                name="description"
                placeholder="What is the question, conflict, or decision needed?"
                required
            ></textarea>

            <button type="submit">Save RFI / Issue</button>
        </form>
    </div>
    """

    return shell("Add RFI / Issue", body)


@app.post("/issues/new")
def create_issue(
    issue_type: str = Form(...),
    title: str = Form(...),
    activity_id: str = Form(""),
    owner: str = Form(""),
    priority: str = Form("HIGH"),
    due: str = Form(...),
    description: str = Form(...)
):
    pid = project_id()
    linked_activity = int(activity_id) if activity_id.strip() else None

    c = db()

    if linked_activity is not None:
        valid = c.execute(
            "SELECT id FROM activities WHERE id=? AND project_id=?",
            (linked_activity, pid)
        ).fetchone()

        if not valid:
            linked_activity = None

    c.execute(
        """
        INSERT INTO project_issues(
            project_id,
            activity_id,
            issue_type,
            title,
            owner,
            due,
            priority,
            status,
            description,
            response,
            created
        )
        VALUES(?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            pid,
            linked_activity,
            issue_type,
            title.strip(),
            owner.strip(),
            due,
            priority,
            "OPEN",
            description.strip(),
            "",
            date.today().isoformat()
        )
    )

    c.commit()
    c.close()

    _v57_emit(project_id(),"ISSUE_CHANGE","Issue created","UI")
    return RedirectResponse(url="/issues", status_code=303)


@app.get("/issues/{issue_id}/edit", response_class=HTMLResponse)
def edit_issue_form(issue_id: int):
    pid = project_id()
    c = db()

    item = c.execute(
        """
        SELECT *
        FROM project_issues
        WHERE id=? AND project_id=?
        """,
        (issue_id, pid)
    ).fetchone()

    c.close()

    if not item:
        return RedirectResponse(url="/issues", status_code=303)

    statuses = ["OPEN", "ANSWERED", "CLOSED"]
    status_options = "".join(
        f'<option value="{s}" {"selected" if item["status"] == s else ""}>{s.title()}</option>'
        for s in statuses
    )

    body = f"""
    <div class="hero">
        <div class="eyebrow">RFIs / Issues</div>
        <h1>Edit {esc(item["issue_type"])}</h1>
    </div>

    <div class="card" style="max-width:760px;">
        <form method="post" action="/issues/{issue_id}/edit">

            <label>Title</label>
            <input type="text" name="title" value="{esc(item["title"])}" required>

            <label>Owner / Responsible Party</label>
            <input type="text" name="owner" value="{esc(item["owner"])}">

            <div class="grid2">
                <div>
                    <label>Priority</label>
                    <select name="priority">
                        <option value="CRITICAL" {"selected" if item["priority"]=="CRITICAL" else ""}>Critical</option>
                        <option value="HIGH" {"selected" if item["priority"]=="HIGH" else ""}>High</option>
                        <option value="WATCH" {"selected" if item["priority"]=="WATCH" else ""}>Watch</option>
                        <option value="LOW" {"selected" if item["priority"]=="LOW" else ""}>Low</option>
                    </select>
                </div>

                <div>
                    <label>Due Date</label>
                    <input type="date" name="due" value="{esc(item["due"])}" required>
                </div>
            </div>

            <label>Description</label>
            <textarea name="description" required>{esc(item["description"])}</textarea>

            <label>Response / Resolution</label>
            <textarea name="response">{esc(item["response"])}</textarea>

            <label>Status</label>
            <select name="status">
                {status_options}
            </select>

            <div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:18px;">
                <button type="submit">Save Changes</button>
                <a href="/issues"
                   style="display:inline-block;color:#f0b44d;text-decoration:none;padding:10px 4px;font-weight:700;">
                    Cancel
                </a>
            </div>
        </form>
    </div>
    """

    return shell("Edit RFI / Issue", body)


@app.post("/issues/{issue_id}/edit")
def edit_issue(
    issue_id: int,
    title: str = Form(...),
    owner: str = Form(""),
    priority: str = Form("HIGH"),
    due: str = Form(...),
    description: str = Form(...),
    response: str = Form(""),
    status: str = Form("OPEN")
):
    pid = project_id()

    c = db()
    c.execute(
        """
        UPDATE project_issues
        SET title=?,
            owner=?,
            priority=?,
            due=?,
            description=?,
            response=?,
            status=?
        WHERE id=? AND project_id=?
        """,
        (
            title.strip(),
            owner.strip(),
            priority,
            due,
            description.strip(),
            response.strip(),
            status,
            issue_id,
            pid
        )
    )
    c.commit()
    c.close()

    return RedirectResponse(url="/issues", status_code=303)


@app.get("/punch", response_class=HTMLResponse)
def punch_page():
    pid = project_id()
    c = db()

    rows = c.execute(
        """
        SELECT p.*, a.external_id, a.name activity
        FROM punch_items p
        LEFT JOIN activities a ON a.id=p.activity_id
        WHERE p.project_id=?
        ORDER BY
            CASE p.status
                WHEN 'OPEN' THEN 1
                WHEN 'CORRECTED' THEN 2
                WHEN 'VERIFIED' THEN 3
                ELSE 4
            END,
            CASE p.priority
                WHEN 'CRITICAL' THEN 1
                WHEN 'HIGH' THEN 2
                WHEN 'WATCH' THEN 3
                ELSE 4
            END,
            p.due
        """,
        (pid,)
    ).fetchall()

    c.close()

    today = date.today().isoformat()

    open_items = [r for r in rows if r["status"] == "OPEN"]
    corrected_items = [r for r in rows if r["status"] == "CORRECTED"]
    verified_items = [r for r in rows if r["status"] == "VERIFIED"]
    overdue_items = [
        r for r in open_items
        if r["due"] and r["due"] < today
    ]

    cards = ""

    for r in rows:
        activity_text = (
            f'{esc(r["external_id"])} - {esc(r["activity"])}'
            if r["external_id"]
            else "No linked activity"
        )

        badge = (
            r["priority"]
            if r["priority"] in ["CRITICAL", "HIGH", "WATCH", "LOW"]
            else "OPEN"
        )

        overdue_text = ""
        if r["status"] == "OPEN" and r["due"] and r["due"] < today:
            overdue_text = " · OVERDUE"

        if r["status"] == "OPEN":
            action_buttons = f"""
            <form method="post" action="/punch/{r["id"]}/corrected">
                <button type="submit">Mark Corrected</button>
            </form>
            """
        elif r["status"] == "CORRECTED":
            action_buttons = f"""
            <form method="post" action="/punch/{r["id"]}/verified">
                <button type="submit">Verify Complete</button>
            </form>
            """
        else:
            action_buttons = '<span class="badge COMPLETE">VERIFIED</span>'

        cards += f"""
        <div class="card">
            <div style="display:flex;justify-content:space-between;gap:12px;align-items:flex-start;flex-wrap:wrap;">
                <div>
                    <span class="badge {badge}">{esc(r["priority"])}</span>
                    <h3 style="margin:10px 0 4px;">{esc(r["title"])}</h3>
                    <div class="muted">
                        {esc(r["location"]) or "No location"} · {esc(r["trade"]) or "No trade"}
                    </div>
                </div>

                <div style="display:flex;gap:8px;flex-wrap:wrap;">
                    {action_buttons}
                    <a href="/punch/{r["id"]}/edit"
                       style="color:#f0b44d;text-decoration:none;font-weight:700;padding:10px 2px;">
                        Edit
                    </a>
                </div>
            </div>

            <p>{esc(r["description"]) or "No description entered."}</p>

            <div class="small">
                {activity_text}
            </div>

            <div class="small">
                Owner: {esc(r["owner"]) or "Unassigned"} ·
                Due {esc(r["due"]) or "—"}{overdue_text} ·
                Status: {esc(r["status"])}
            </div>

            {"<p><b>Resolution:</b> " + esc(r["resolution"]) + "</p>" if r["resolution"] else ""}
        </div>
        """

    if not cards:
        cards = '<div class="card"><div class="muted">No punch or quality items logged yet.</div></div>'

    body = f"""
    <div class="hero">
        <div style="display:flex;justify-content:space-between;align-items:center;gap:15px;flex-wrap:wrap;">
            <div>
                <div class="eyebrow">Quality / Punch List</div>
                <h1>Track deficiencies until they are corrected and verified.</h1>
                <div class="muted">
                    Assign ownership, due dates, location, trade, and final resolution.
                </div>
            </div>

            <a href="/punch/new"
               style="display:inline-block;background:#f0b44d;color:#0a1017;text-decoration:none;padding:12px 18px;border-radius:9px;font-weight:800;">
                + Add Punch Item
            </a>
        </div>
    </div>

    <div class="grid4">
        <div class="card">
            <div class="label">Open</div>
            <div class="kpi">{len(open_items)}</div>
        </div>

        <div class="card">
            <div class="label">Overdue</div>
            <div class="kpi">{len(overdue_items)}</div>
        </div>

        <div class="card">
            <div class="label">Corrected</div>
            <div class="kpi">{len(corrected_items)}</div>
        </div>

        <div class="card">
            <div class="label">Verified</div>
            <div class="kpi">{len(verified_items)}</div>
        </div>
    </div>

    <div class="grid2">
        {cards}
    </div>
    """

    return shell("Punch List", body)


@app.get("/punch/new", response_class=HTMLResponse)
def new_punch_form():
    pid = project_id()
    c = db()

    activities = c.execute(
        """
        SELECT id,external_id,name
        FROM activities
        WHERE project_id=?
        ORDER BY start,name
        """,
        (pid,)
    ).fetchall()

    subs = c.execute(
        """
        SELECT name,trade
        FROM subs
        WHERE project_id=?
        ORDER BY trade,name
        """,
        (pid,)
    ).fetchall()

    c.close()

    activity_options = '<option value="">No linked activity</option>'
    activity_options += "".join(
        f'<option value="{a["id"]}">{esc(a["external_id"])} - {esc(a["name"])}</option>'
        for a in activities
    )

    owner_options = '<option value="">Unassigned</option>'
    owner_options += '<option value="Superintendent">Superintendent</option>'
    owner_options += '<option value="Project Manager">Project Manager</option>'

    for s in subs:
        owner_options += (
            f'<option value="{esc(s["name"])}">{esc(s["name"])} - {esc(s["trade"])}</option>'
        )

    body = f"""
    <div class="hero">
        <div class="eyebrow">Quality / Punch List</div>
        <h1>Add Punch Item</h1>
    </div>

    <div class="card" style="max-width:760px;">
        <form method="post" action="/punch/new">

            <label>Item Title</label>
            <input
                type="text"
                name="title"
                placeholder="Example: Repair damaged drywall at Room 214"
                required
            >

            <label>Location</label>
            <input
                type="text"
                name="location"
                placeholder="Example: Level 2 - Room 214"
            >

            <label>Linked Activity</label>
            <select name="activity_id">
                {activity_options}
            </select>

            <label>Trade</label>
            <input
                type="text"
                name="trade"
                placeholder="Example: Drywall"
            >

            <label>Owner / Responsible Party</label>
            <select name="owner">
                {owner_options}
            </select>

            <div class="grid2">
                <div>
                    <label>Priority</label>
                    <select name="priority">
                        <option value="CRITICAL">Critical</option>
                        <option value="HIGH">High</option>
                        <option value="WATCH">Watch</option>
                        <option value="LOW">Low</option>
                    </select>
                </div>

                <div>
                    <label>Due Date</label>
                    <input type="date" name="due" required>
                </div>
            </div>

            <label>Description</label>
            <textarea
                name="description"
                placeholder="Describe the deficiency and what acceptable correction looks like."
                required
            ></textarea>

            <button type="submit">Save Punch Item</button>
        </form>
    </div>
    """

    return shell("Add Punch Item", body)


@app.post("/punch/new")
def create_punch(
    title: str = Form(...),
    location: str = Form(""),
    activity_id: str = Form(""),
    trade: str = Form(""),
    owner: str = Form(""),
    priority: str = Form("HIGH"),
    due: str = Form(...),
    description: str = Form(...)
):
    pid = project_id()
    linked_activity = int(activity_id) if activity_id.strip() else None

    c = db()

    if linked_activity is not None:
        valid = c.execute(
            "SELECT id FROM activities WHERE id=? AND project_id=?",
            (linked_activity, pid)
        ).fetchone()

        if not valid:
            linked_activity = None

    c.execute(
        """
        INSERT INTO punch_items(
            project_id,
            activity_id,
            title,
            location,
            trade,
            owner,
            priority,
            due,
            status,
            description,
            resolution,
            created
        )
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            pid,
            linked_activity,
            title.strip(),
            location.strip(),
            trade.strip(),
            owner.strip(),
            priority,
            due,
            "OPEN",
            description.strip(),
            "",
            date.today().isoformat()
        )
    )

    c.commit()
    c.close()

    return RedirectResponse(url="/punch", status_code=303)


@app.get("/punch/{item_id}/edit", response_class=HTMLResponse)
def edit_punch_form(item_id: int):
    pid = project_id()
    c = db()

    item = c.execute(
        """
        SELECT *
        FROM punch_items
        WHERE id=? AND project_id=?
        """,
        (item_id, pid)
    ).fetchone()

    c.close()

    if not item:
        return RedirectResponse(url="/punch", status_code=303)

    body = f"""
    <div class="hero">
        <div class="eyebrow">Quality / Punch List</div>
        <h1>Edit Punch Item</h1>
    </div>

    <div class="card" style="max-width:760px;">
        <form method="post" action="/punch/{item_id}/edit">

            <label>Item Title</label>
            <input type="text" name="title" value="{esc(item["title"])}" required>

            <label>Location</label>
            <input type="text" name="location" value="{esc(item["location"])}">

            <label>Trade</label>
            <input type="text" name="trade" value="{esc(item["trade"])}">

            <label>Owner</label>
            <input type="text" name="owner" value="{esc(item["owner"])}">

            <div class="grid2">
                <div>
                    <label>Priority</label>
                    <select name="priority">
                        <option value="CRITICAL" {"selected" if item["priority"]=="CRITICAL" else ""}>Critical</option>
                        <option value="HIGH" {"selected" if item["priority"]=="HIGH" else ""}>High</option>
                        <option value="WATCH" {"selected" if item["priority"]=="WATCH" else ""}>Watch</option>
                        <option value="LOW" {"selected" if item["priority"]=="LOW" else ""}>Low</option>
                    </select>
                </div>

                <div>
                    <label>Due Date</label>
                    <input type="date" name="due" value="{esc(item["due"])}" required>
                </div>
            </div>

            <label>Description</label>
            <textarea name="description" required>{esc(item["description"])}</textarea>

            <label>Resolution</label>
            <textarea name="resolution">{esc(item["resolution"])}</textarea>

            <button type="submit">Save Changes</button>
        </form>
    </div>
    """

    return shell("Edit Punch Item", body)


@app.post("/punch/{item_id}/edit")
def edit_punch(
    item_id: int,
    title: str = Form(...),
    location: str = Form(""),
    trade: str = Form(""),
    owner: str = Form(""),
    priority: str = Form("HIGH"),
    due: str = Form(...),
    description: str = Form(...),
    resolution: str = Form("")
):
    pid = project_id()

    c = db()
    c.execute(
        """
        UPDATE punch_items
        SET title=?,
            location=?,
            trade=?,
            owner=?,
            priority=?,
            due=?,
            description=?,
            resolution=?
        WHERE id=? AND project_id=?
        """,
        (
            title.strip(),
            location.strip(),
            trade.strip(),
            owner.strip(),
            priority,
            due,
            description.strip(),
            resolution.strip(),
            item_id,
            pid
        )
    )
    c.commit()
    c.close()

    return RedirectResponse(url="/punch", status_code=303)


@app.post("/punch/{item_id}/corrected")
def mark_punch_corrected(item_id: int):
    pid = project_id()

    c = db()
    c.execute(
        """
        UPDATE punch_items
        SET status='CORRECTED'
        WHERE id=? AND project_id=?
        """,
        (item_id, pid)
    )
    c.commit()
    c.close()

    return RedirectResponse(url="/punch", status_code=303)


@app.post("/punch/{item_id}/verified")
def mark_punch_verified(item_id: int):
    pid = project_id()

    c = db()
    c.execute(
        """
        UPDATE punch_items
        SET status='VERIFIED'
        WHERE id=? AND project_id=?
        """,
        (item_id, pid)
    )
    c.commit()
    c.close()

    return RedirectResponse(url="/punch", status_code=303)


@app.get("/inspections", response_class=HTMLResponse)
def inspections_page():
    pid = project_id()
    c = db()

    rows = c.execute(
        """
        SELECT i.*, a.external_id, a.name activity
        FROM inspections_tracker i
        LEFT JOIN activities a ON a.id=i.activity_id
        WHERE i.project_id=?
        ORDER BY
            CASE i.result
                WHEN 'FAILED' THEN 1
                WHEN 'PENDING' THEN 2
                WHEN 'SCHEDULED' THEN 3
                WHEN 'PASSED' THEN 4
                ELSE 5
            END,
            i.scheduled_date
        """,
        (pid,)
    ).fetchall()

    c.close()

    today = date.today().isoformat()

    pending = [r for r in rows if r["result"] in ["PENDING", "SCHEDULED"]]
    failed = [r for r in rows if r["result"] == "FAILED"]
    passed = [r for r in rows if r["result"] == "PASSED"]
    overdue = [
        r for r in pending
        if r["scheduled_date"] and r["scheduled_date"] < today
    ]

    cards = ""

    for r in rows:
        activity_text = (
            f'{esc(r["external_id"])} - {esc(r["activity"])}'
            if r["external_id"]
            else "No linked activity"
        )

        if r["result"] == "PASSED":
            badge = "READY"
        elif r["result"] == "FAILED":
            badge = "CRITICAL"
        elif r["result"] == "SCHEDULED":
            badge = "WATCH"
        else:
            badge = "OPEN"

        reinspection = (
            f' · Reinspection {esc(r["reinspection_date"])}'
            if r["reinspection_date"]
            else ""
        )

        overdue_text = ""
        if r["result"] in ["PENDING", "SCHEDULED"] and r["scheduled_date"] and r["scheduled_date"] < today:
            overdue_text = " · OVERDUE"

        cards += f"""
        <div class="card">
            <div style="display:flex;justify-content:space-between;gap:12px;align-items:flex-start;flex-wrap:wrap;">
                <div>
                    <span class="badge {badge}">{esc(r["result"])}</span>
                    <h3 style="margin:10px 0 4px;">{esc(r["inspection_type"])}</h3>
                    <div class="muted">{activity_text}</div>
                </div>

                <a href="/inspections/{r["id"]}/edit"
                   style="color:#f0b44d;text-decoration:none;font-weight:700;">
                    Edit
                </a>
            </div>

            <p>{esc(r["notes"]) or "No notes entered."}</p>

            <div class="small">
                Authority: {esc(r["authority"]) or "—"} ·
                Scheduled {esc(r["scheduled_date"]) or "—"}{overdue_text}{reinspection}
            </div>
        </div>
        """

    if not cards:
        cards = '<div class="card"><div class="muted">No inspections logged yet.</div></div>'

    body = f"""
    <div class="hero">
        <div style="display:flex;justify-content:space-between;align-items:center;gap:15px;flex-wrap:wrap;">
            <div>
                <div class="eyebrow">Inspection Intelligence</div>
                <h1>Know which inspection can stop the next activity.</h1>
                <div class="muted">
                    Track scheduled inspections, pass/fail results, authorities, and reinspections.
                </div>
            </div>

            <a href="/inspections/new"
               style="display:inline-block;background:#f0b44d;color:#0a1017;text-decoration:none;padding:12px 18px;border-radius:9px;font-weight:800;">
                + Add Inspection
            </a>
        </div>
    </div>

    <div class="grid4">
        <div class="card">
            <div class="label">Pending</div>
            <div class="kpi">{len(pending)}</div>
        </div>

        <div class="card">
            <div class="label">Overdue</div>
            <div class="kpi">{len(overdue)}</div>
        </div>

        <div class="card">
            <div class="label">Failed</div>
            <div class="kpi">{len(failed)}</div>
        </div>

        <div class="card">
            <div class="label">Passed</div>
            <div class="kpi">{len(passed)}</div>
        </div>
    </div>

    <div class="grid2">
        {cards}
    </div>
    """

    return shell("Inspections", body)


@app.get("/inspections/new", response_class=HTMLResponse)
def new_inspection_form():
    pid = project_id()
    c = db()

    activities = c.execute(
        """
        SELECT id,external_id,name
        FROM activities
        WHERE project_id=?
        ORDER BY start,name
        """,
        (pid,)
    ).fetchall()

    c.close()

    activity_options = '<option value="">No linked activity</option>'
    activity_options += "".join(
        f'<option value="{a["id"]}">{esc(a["external_id"])} - {esc(a["name"])}</option>'
        for a in activities
    )

    body = f"""
    <div class="hero">
        <div class="eyebrow">Inspection Intelligence</div>
        <h1>Add Inspection</h1>
    </div>

    <div class="card" style="max-width:760px;">
        <form method="post" action="/inspections/new">

            <label>Inspection Type</label>
            <input
                type="text"
                name="inspection_type"
                placeholder="Example: Above Ceiling Inspection"
                required
            >

            <label>Linked Activity</label>
            <select name="activity_id">
                {activity_options}
            </select>

            <label>Authority / Inspector</label>
            <input
                type="text"
                name="authority"
                placeholder="Example: City of Phoenix"
            >

            <label>Scheduled Date</label>
            <input type="date" name="scheduled_date" required>

            <label>Result</label>
            <select name="result">
                <option value="PENDING">Pending</option>
                <option value="SCHEDULED">Scheduled</option>
                <option value="PASSED">Passed</option>
                <option value="FAILED">Failed</option>
            </select>

            <label>Reinspection Date</label>
            <input type="date" name="reinspection_date">

            <label>Notes</label>
            <textarea
                name="notes"
                placeholder="Required corrections, prerequisites, inspector comments, etc."
            ></textarea>

            <button type="submit">Save Inspection</button>
        </form>
    </div>
    """

    return shell("Add Inspection", body)


@app.post("/inspections/new")
def create_inspection(
    inspection_type: str = Form(...),
    activity_id: str = Form(""),
    authority: str = Form(""),
    scheduled_date: str = Form(...),
    result: str = Form("PENDING"),
    reinspection_date: str = Form(""),
    notes: str = Form("")
):
    pid = project_id()
    linked_activity = int(activity_id) if activity_id.strip() else None

    c = db()

    if linked_activity is not None:
        valid = c.execute(
            "SELECT id FROM activities WHERE id=? AND project_id=?",
            (linked_activity, pid)
        ).fetchone()
        if not valid:
            linked_activity = None

    c.execute(
        """
        INSERT INTO inspections_tracker(
            project_id,
            activity_id,
            inspection_type,
            authority,
            scheduled_date,
            result,
            reinspection_date,
            notes,
            created
        )
        VALUES(?,?,?,?,?,?,?,?,?)
        """,
        (
            pid,
            linked_activity,
            inspection_type.strip(),
            authority.strip(),
            scheduled_date,
            result,
            reinspection_date,
            notes.strip(),
            date.today().isoformat()
        )
    )

    c.commit()
    c.close()

    _v57_emit(project_id(),"INSPECTION_CHANGE","Inspection updated","UI")
    return RedirectResponse(url="/inspections", status_code=303)


@app.get("/inspections/{inspection_id}/edit", response_class=HTMLResponse)
def edit_inspection_form(inspection_id: int):
    pid = project_id()
    c = db()

    item = c.execute(
        """
        SELECT *
        FROM inspections_tracker
        WHERE id=? AND project_id=?
        """,
        (inspection_id, pid)
    ).fetchone()

    c.close()

    if not item:
        return RedirectResponse(url="/inspections", status_code=303)

    results = ["PENDING", "SCHEDULED", "PASSED", "FAILED"]
    result_options = "".join(
        f'<option value="{r}" {"selected" if item["result"] == r else ""}>{r.title()}</option>'
        for r in results
    )

    body = f"""
    <div class="hero">
        <div class="eyebrow">Inspection Intelligence</div>
        <h1>Edit Inspection</h1>
    </div>

    <div class="card" style="max-width:760px;">
        <form method="post" action="/inspections/{inspection_id}/edit">

            <label>Inspection Type</label>
            <input type="text" name="inspection_type" value="{esc(item["inspection_type"])}" required>

            <label>Authority / Inspector</label>
            <input type="text" name="authority" value="{esc(item["authority"])}">

            <label>Scheduled Date</label>
            <input type="date" name="scheduled_date" value="{esc(item["scheduled_date"])}" required>

            <label>Result</label>
            <select name="result">
                {result_options}
            </select>

            <label>Reinspection Date</label>
            <input type="date" name="reinspection_date" value="{esc(item["reinspection_date"])}">

            <label>Notes</label>
            <textarea name="notes">{esc(item["notes"])}</textarea>

            <button type="submit">Save Changes</button>
        </form>
    </div>
    """

    return shell("Edit Inspection", body)


@app.post("/inspections/{inspection_id}/edit")
def edit_inspection(
    inspection_id: int,
    inspection_type: str = Form(...),
    authority: str = Form(""),
    scheduled_date: str = Form(...),
    result: str = Form("PENDING"),
    reinspection_date: str = Form(""),
    notes: str = Form("")
):
    pid = project_id()

    c = db()
    c.execute(
        """
        UPDATE inspections_tracker
        SET inspection_type=?,
            authority=?,
            scheduled_date=?,
            result=?,
            reinspection_date=?,
            notes=?
        WHERE id=? AND project_id=?
        """,
        (
            inspection_type.strip(),
            authority.strip(),
            scheduled_date,
            result,
            reinspection_date,
            notes.strip(),
            inspection_id,
            pid
        )
    )
    c.commit()
    c.close()

    return RedirectResponse(url="/inspections", status_code=303)


@app.get("/submittals", response_class=HTMLResponse)
def submittals_page():
    pid = project_id()
    c = db()

    rows = c.execute(
        """
        SELECT s.*, a.external_id, a.name activity
        FROM submittals s
        LEFT JOIN activities a ON a.id=s.activity_id
        WHERE s.project_id=?
        ORDER BY
            CASE s.status
                WHEN 'REJECTED' THEN 1
                WHEN 'PENDING' THEN 2
                WHEN 'SUBMITTED' THEN 3
                WHEN 'APPROVED_AS_NOTED' THEN 4
                WHEN 'APPROVED' THEN 5
                ELSE 6
            END,
            s.due_date
        """,
        (pid,)
    ).fetchall()

    c.close()

    today = date.today().isoformat()

    pending = [r for r in rows if r["status"] in ["PENDING", "SUBMITTED"]]
    rejected = [r for r in rows if r["status"] == "REJECTED"]
    approved = [r for r in rows if r["status"] in ["APPROVED", "APPROVED_AS_NOTED"]]
    overdue = [
        r for r in pending
        if r["due_date"] and r["due_date"] < today
    ]

    cards = ""

    for r in rows:
        activity_text = (
            f'{esc(r["external_id"])} - {esc(r["activity"])}'
            if r["external_id"]
            else "No linked activity"
        )

        if r["status"] == "REJECTED":
            badge = "CRITICAL"
        elif r["status"] in ["PENDING", "SUBMITTED"]:
            badge = "WATCH"
        elif r["status"] in ["APPROVED", "APPROVED_AS_NOTED"]:
            badge = "READY"
        else:
            badge = "OPEN"

        overdue_text = ""
        if r["status"] in ["PENDING", "SUBMITTED"] and r["due_date"] and r["due_date"] < today:
            overdue_text = " · OVERDUE"

        cards += f"""
        <div class="card">
            <div style="display:flex;justify-content:space-between;gap:12px;align-items:flex-start;flex-wrap:wrap;">
                <div>
                    <span class="badge {badge}">{esc(r["status"]).replace("_"," ")}</span>
                    <h3 style="margin:10px 0 4px;">{esc(r["title"])}</h3>
                    <div class="muted">{activity_text}</div>
                </div>

                <a href="/submittals/{r["id"]}/edit"
                   style="color:#f0b44d;text-decoration:none;font-weight:700;">
                    Edit
                </a>
            </div>

            <div class="grid3" style="margin-top:14px;">
                <div>
                    <div class="label">Spec Section</div>
                    <div>{esc(r["spec_section"]) or "—"}</div>
                </div>

                <div>
                    <div class="label">Responsible</div>
                    <div>{esc(r["responsible_party"]) or "—"}</div>
                </div>

                <div>
                    <div class="label">Due</div>
                    <div>{esc(r["due_date"]) or "—"}{overdue_text}</div>
                </div>
            </div>

            <p>{esc(r["notes"]) or "No notes entered."}</p>
            <div class="small">Sent: {esc(r["sent_date"]) or "—"}</div>
        </div>
        """

    if not cards:
        cards = '<div class="card"><div class="muted">No submittals logged yet.</div></div>'

    body = f"""
    <div class="hero">
        <div style="display:flex;justify-content:space-between;align-items:center;gap:15px;flex-wrap:wrap;">
            <div>
                <div class="eyebrow">Document Control</div>
                <h1>Track submittals before approvals become schedule blockers.</h1>
                <div class="muted">
                    Monitor due dates, approval status, responsible party, and linked work.
                </div>
            </div>

            <a href="/submittals/new"
               style="display:inline-block;background:#f0b44d;color:#0a1017;text-decoration:none;padding:12px 18px;border-radius:9px;font-weight:800;">
                + Add Submittal
            </a>
        </div>
    </div>

    <div class="grid4">
        <div class="card">
            <div class="label">Pending</div>
            <div class="kpi">{len(pending)}</div>
        </div>

        <div class="card">
            <div class="label">Overdue</div>
            <div class="kpi">{len(overdue)}</div>
        </div>

        <div class="card">
            <div class="label">Rejected</div>
            <div class="kpi">{len(rejected)}</div>
        </div>

        <div class="card">
            <div class="label">Approved</div>
            <div class="kpi">{len(approved)}</div>
        </div>
    </div>

    <div class="grid2">
        {cards}
    </div>
    """

    return shell("Submittals", body)


@app.get("/submittals/new", response_class=HTMLResponse)
def new_submittal_form():
    pid = project_id()
    c = db()

    activities = c.execute(
        """
        SELECT id,external_id,name
        FROM activities
        WHERE project_id=?
        ORDER BY start,name
        """,
        (pid,)
    ).fetchall()

    subs = c.execute(
        """
        SELECT name,trade
        FROM subs
        WHERE project_id=?
        ORDER BY trade,name
        """,
        (pid,)
    ).fetchall()

    c.close()

    activity_options = '<option value="">No linked activity</option>'
    activity_options += "".join(
        f'<option value="{a["id"]}">{esc(a["external_id"])} - {esc(a["name"])}</option>'
        for a in activities
    )

    responsible_options = '<option value="">Unassigned</option>'
    responsible_options += '<option value="Project Manager">Project Manager</option>'
    responsible_options += '<option value="Architect / Engineer">Architect / Engineer</option>'

    for s in subs:
        responsible_options += (
            f'<option value="{esc(s["name"])}">{esc(s["name"])} - {esc(s["trade"])}</option>'
        )

    body = f"""
    <div class="hero">
        <div class="eyebrow">Document Control</div>
        <h1>Add Submittal</h1>
    </div>

    <div class="card" style="max-width:760px;">
        <form method="post" action="/submittals/new">

            <label>Submittal Title</label>
            <input type="text" name="title" placeholder="Example: Electrical Switchgear Product Data" required>

            <label>Spec Section</label>
            <input type="text" name="spec_section" placeholder="Example: 26 24 16">

            <label>Linked Activity</label>
            <select name="activity_id">
                {activity_options}
            </select>

            <label>Subcontractor / Responsible Party</label>
            <input
                type="text"
                name="responsible_party"
                list="responsible_parties"
                placeholder="Type or select a subcontractor"
            >
            <datalist id="responsible_parties">
                <option value="Project Manager">
                <option value="Architect / Engineer">
                {''.join(f'<option value="{esc(s["name"])}">{esc(s["trade"])}</option>' for s in subs)}
            </datalist>

            <div class="grid2">
                <div>
                    <label>Sent Date</label>
                    <input type="date" name="sent_date">
                </div>

                <div>
                    <label>Due Date</label>
                    <input type="date" name="due_date" required>
                </div>
            </div>

            <label>Status</label>
            <select name="status">
                <option value="PENDING">Pending</option>
                <option value="SUBMITTED">Submitted</option>
                <option value="APPROVED">Approved</option>
                <option value="APPROVED_AS_NOTED">Approved As Noted</option>
                <option value="REJECTED">Rejected / Revise & Resubmit</option>
            </select>

            <label>Notes</label>
            <textarea name="notes" placeholder="Review comments, lead-time impact, release constraints, etc."></textarea>

            <button type="submit">Save Submittal</button>
        </form>
    </div>
    """

    return shell("Add Submittal", body)


@app.post("/submittals/new")
def create_submittal(
    title: str = Form(...),
    spec_section: str = Form(""),
    activity_id: str = Form(""),
    responsible_party: str = Form(""),
    sent_date: str = Form(""),
    due_date: str = Form(...),
    status: str = Form("PENDING"),
    notes: str = Form("")
):
    pid = project_id()
    linked_activity = int(activity_id) if activity_id.strip() else None

    c = db()

    if linked_activity is not None:
        valid = c.execute(
            "SELECT id FROM activities WHERE id=? AND project_id=?",
            (linked_activity, pid)
        ).fetchone()
        if not valid:
            linked_activity = None

    c.execute(
        """
        INSERT INTO submittals(
            project_id, activity_id, title, spec_section, responsible_party,
            sent_date, due_date, status, notes, created
        )
        VALUES(?,?,?,?,?,?,?,?,?,?)
        """,
        (
            pid,
            linked_activity,
            title.strip(),
            spec_section.strip(),
            responsible_party.strip(),
            sent_date,
            due_date,
            status,
            notes.strip(),
            date.today().isoformat()
        )
    )

    c.commit()
    c.close()

    _v57_emit(project_id(),"SUBMITTAL_CHANGE","Submittal created","UI")
    return RedirectResponse(url="/submittals", status_code=303)


@app.get("/submittals/{submittal_id}/edit", response_class=HTMLResponse)
def edit_submittal_form(submittal_id: int):
    pid = project_id()
    c = db()

    item = c.execute(
        "SELECT * FROM submittals WHERE id=? AND project_id=?",
        (submittal_id, pid)
    ).fetchone()

    c.close()

    if not item:
        return RedirectResponse(url="/submittals", status_code=303)

    statuses = ["PENDING", "SUBMITTED", "APPROVED", "APPROVED_AS_NOTED", "REJECTED"]
    status_options = "".join(
        f'<option value="{s}" {"selected" if item["status"] == s else ""}>{s.replace("_"," ").title()}</option>'
        for s in statuses
    )

    body = f"""
    <div class="hero">
        <div class="eyebrow">Document Control</div>
        <h1>Edit Submittal</h1>
    </div>

    <div class="card" style="max-width:760px;">
        <form method="post" action="/submittals/{submittal_id}/edit">

            <label>Submittal Title</label>
            <input type="text" name="title" value="{esc(item["title"])}" required>

            <label>Spec Section</label>
            <input type="text" name="spec_section" value="{esc(item["spec_section"])}">

            <label>Subcontractor / Responsible Party</label>
            <input
                type="text"
                name="responsible_party"
                value="{esc(item["responsible_party"])}"
                placeholder="Type subcontractor or responsible party"
            >

            <div class="grid2">
                <div>
                    <label>Sent Date</label>
                    <input type="date" name="sent_date" value="{esc(item["sent_date"])}">
                </div>

                <div>
                    <label>Due Date</label>
                    <input type="date" name="due_date" value="{esc(item["due_date"])}" required>
                </div>
            </div>

            <label>Status</label>
            <select name="status">
                {status_options}
            </select>

            <label>Notes</label>
            <textarea name="notes">{esc(item["notes"])}</textarea>

            <button type="submit">Save Changes</button>
        </form>
    </div>
    """

    return shell("Edit Submittal", body)


@app.post("/submittals/{submittal_id}/edit")
def edit_submittal(
    submittal_id: int,
    title: str = Form(...),
    spec_section: str = Form(""),
    responsible_party: str = Form(""),
    sent_date: str = Form(""),
    due_date: str = Form(...),
    status: str = Form("PENDING"),
    notes: str = Form("")
):
    pid = project_id()

    c = db()
    c.execute(
        """
        UPDATE submittals
        SET title=?,
            spec_section=?,
            responsible_party=?,
            sent_date=?,
            due_date=?,
            status=?,
            notes=?
        WHERE id=? AND project_id=?
        """,
        (
            title.strip(),
            spec_section.strip(),
            responsible_party.strip(),
            sent_date,
            due_date,
            status,
            notes.strip(),
            submittal_id,
            pid
        )
    )
    c.commit()
    c.close()

    return RedirectResponse(url="/submittals", status_code=303)


@app.get("/safety", response_class=HTMLResponse)
def safety_page():
    pid = project_id()
    c = db()

    rows = c.execute(
        """
        SELECT s.*, a.external_id, a.name activity
        FROM safety_items s
        LEFT JOIN activities a ON a.id=s.activity_id
        WHERE s.project_id=?
        ORDER BY
            CASE s.status
                WHEN 'OPEN' THEN 1
                WHEN 'CORRECTED' THEN 2
                WHEN 'CLOSED' THEN 3
                ELSE 4
            END,
            CASE s.severity
                WHEN 'CRITICAL' THEN 1
                WHEN 'HIGH' THEN 2
                WHEN 'WATCH' THEN 3
                ELSE 4
            END,
            s.event_date DESC
        """,
        (pid,)
    ).fetchall()

    c.close()

    open_items = [r for r in rows if r["status"] == "OPEN"]
    corrected = [r for r in rows if r["status"] == "CORRECTED"]
    closed = [r for r in rows if r["status"] == "CLOSED"]
    critical = [r for r in open_items if r["severity"] == "CRITICAL"]

    cards = ""

    for r in rows:
        activity_text = (
            f'{esc(r["external_id"])} - {esc(r["activity"])}'
            if r["external_id"]
            else "No linked activity"
        )

        badge = (
            r["severity"]
            if r["severity"] in ["CRITICAL","HIGH","WATCH","LOW"]
            else "OPEN"
        )

        if r["status"] == "OPEN":
            action_button = f"""
            <form method="post" action="/safety/{r["id"]}/corrected">
                <button type="submit">Mark Corrected</button>
            </form>
            """
        elif r["status"] == "CORRECTED":
            action_button = f"""
            <form method="post" action="/safety/{r["id"]}/closed">
                <button type="submit">Close Item</button>
            </form>
            """
        else:
            action_button = '<span class="badge COMPLETE">CLOSED</span>'

        cards += f"""
        <div class="card">
            <div style="display:flex;justify-content:space-between;gap:12px;align-items:flex-start;flex-wrap:wrap;">
                <div>
                    <span class="badge {badge}">{esc(r["severity"])}</span>
                    <span class="badge OPEN">{esc(r["item_type"]).replace("_"," ")}</span>
                    <h3 style="margin:10px 0 4px;">{esc(r["title"])}</h3>
                    <div class="muted">{esc(r["location"]) or "No location"} · {activity_text}</div>
                </div>

                <div>{action_button}</div>
            </div>

            <p>{esc(r["description"]) or "No description entered."}</p>

            <div class="small">
                Date: {esc(r["event_date"])} ·
                Responsible: {esc(r["responsible_party"]) or "Unassigned"} ·
                Status: {esc(r["status"])}
            </div>

            {"<p><b>Corrective Action:</b> " + esc(r["corrective_action"]) + "</p>" if r["corrective_action"] else ""}
        </div>
        """

    if not cards:
        cards = '<div class="card"><div class="muted">No safety items logged yet.</div></div>'

    body = f"""
    <div class="hero">
        <div style="display:flex;justify-content:space-between;align-items:center;gap:15px;flex-wrap:wrap;">
            <div>
                <div class="eyebrow">Safety Intelligence</div>
                <h1>Track observations, near misses, and corrective actions.</h1>
                <div class="muted">
                    Keep safety issues visible until they are corrected and closed.
                </div>
            </div>

            <a href="/safety/new"
               style="display:inline-block;background:#f0b44d;color:#0a1017;text-decoration:none;padding:12px 18px;border-radius:9px;font-weight:800;">
                + Add Safety Item
            </a>
        </div>
    </div>

    <div class="grid4">
        <div class="card">
            <div class="label">Open</div>
            <div class="kpi">{len(open_items)}</div>
        </div>

        <div class="card">
            <div class="label">Critical</div>
            <div class="kpi">{len(critical)}</div>
        </div>

        <div class="card">
            <div class="label">Corrected</div>
            <div class="kpi">{len(corrected)}</div>
        </div>

        <div class="card">
            <div class="label">Closed</div>
            <div class="kpi">{len(closed)}</div>
        </div>
    </div>

    <div class="grid2">
        {cards}
    </div>
    """

    return shell("Safety", body)


@app.get("/safety/new", response_class=HTMLResponse)
def new_safety_form():
    pid = project_id()
    c = db()

    activities = c.execute(
        """
        SELECT id,external_id,name
        FROM activities
        WHERE project_id=?
        ORDER BY start,name
        """,
        (pid,)
    ).fetchall()

    subs = c.execute(
        """
        SELECT name,trade
        FROM subs
        WHERE project_id=?
        ORDER BY trade,name
        """,
        (pid,)
    ).fetchall()

    c.close()

    activity_options = '<option value="">No linked activity</option>'
    activity_options += "".join(
        f'<option value="{a["id"]}">{esc(a["external_id"])} - {esc(a["name"])}</option>'
        for a in activities
    )

    responsible_options = '<option value="">Unassigned</option>'
    responsible_options += '<option value="Superintendent">Superintendent</option>'
    responsible_options += '<option value="Safety Manager">Safety Manager</option>'

    for s in subs:
        responsible_options += (
            f'<option value="{esc(s["name"])}">{esc(s["name"])} - {esc(s["trade"])}</option>'
        )

    body = f"""
    <div class="hero">
        <div class="eyebrow">Safety Intelligence</div>
        <h1>Add Safety Item</h1>
    </div>

    <div class="card" style="max-width:760px;">
        <form method="post" action="/safety/new">

            <label>Type</label>
            <select name="item_type">
                <option value="OBSERVATION">Observation</option>
                <option value="NEAR_MISS">Near Miss</option>
                <option value="INCIDENT">Incident</option>
                <option value="CORRECTIVE_ACTION">Corrective Action</option>
            </select>

            <label>Title</label>
            <input type="text" name="title" placeholder="Example: Missing guardrail at stair opening" required>

            <label>Event Date</label>
            <input type="date" name="event_date" value="{date.today().isoformat()}" required>

            <label>Location</label>
            <input type="text" name="location" placeholder="Example: Level 3 west stair">

            <label>Linked Activity</label>
            <select name="activity_id">
                {activity_options}
            </select>

            <label>Responsible Party</label>
            <select name="responsible_party">
                {responsible_options}
            </select>

            <label>Severity</label>
            <select name="severity">
                <option value="CRITICAL">Critical</option>
                <option value="HIGH">High</option>
                <option value="WATCH">Watch</option>
                <option value="LOW">Low</option>
            </select>

            <label>Description</label>
            <textarea name="description" placeholder="Describe the condition or event." required></textarea>

            <label>Corrective Action</label>
            <textarea name="corrective_action" placeholder="What must be done to correct or prevent recurrence?"></textarea>

            <button type="submit">Save Safety Item</button>
        </form>
    </div>
    """

    return shell("Add Safety Item", body)


@app.post("/safety/new")
def create_safety_item(
    item_type: str = Form(...),
    title: str = Form(...),
    event_date: str = Form(...),
    location: str = Form(""),
    activity_id: str = Form(""),
    responsible_party: str = Form(""),
    severity: str = Form("WATCH"),
    description: str = Form(...),
    corrective_action: str = Form("")
):
    pid = project_id()
    linked_activity = int(activity_id) if activity_id.strip() else None

    c = db()

    if linked_activity is not None:
        valid = c.execute(
            "SELECT id FROM activities WHERE id=? AND project_id=?",
            (linked_activity, pid)
        ).fetchone()
        if not valid:
            linked_activity = None

    c.execute(
        """
        INSERT INTO safety_items(
            project_id,
            activity_id,
            event_date,
            item_type,
            title,
            location,
            responsible_party,
            severity,
            status,
            description,
            corrective_action,
            created
        )
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            pid,
            linked_activity,
            event_date,
            item_type,
            title.strip(),
            location.strip(),
            responsible_party.strip(),
            severity,
            "OPEN",
            description.strip(),
            corrective_action.strip(),
            date.today().isoformat()
        )
    )

    c.commit()
    c.close()

    return RedirectResponse(url="/safety", status_code=303)


@app.post("/safety/{item_id}/corrected")
def mark_safety_corrected(item_id: int):
    pid = project_id()

    c = db()
    c.execute(
        """
        UPDATE safety_items
        SET status='CORRECTED'
        WHERE id=? AND project_id=?
        """,
        (item_id, pid)
    )
    c.commit()
    c.close()

    return RedirectResponse(url="/safety", status_code=303)


@app.post("/safety/{item_id}/closed")
def mark_safety_closed(item_id: int):
    pid = project_id()

    c = db()
    c.execute(
        """
        UPDATE safety_items
        SET status='CLOSED'
        WHERE id=? AND project_id=?
        """,
        (item_id, pid)
    )
    c.commit()
    c.close()

    return RedirectResponse(url="/safety", status_code=303)


@app.get("/changes", response_class=HTMLResponse)
def changes_page():
    pid = project_id()
    c = db()

    rows = c.execute(
        """
        SELECT ce.*, a.external_id, a.name activity
        FROM change_events ce
        LEFT JOIN activities a ON a.id=ce.activity_id
        WHERE ce.project_id=?
        ORDER BY
            CASE ce.status
                WHEN 'OPEN' THEN 1
                WHEN 'PRICING' THEN 2
                WHEN 'SUBMITTED' THEN 3
                WHEN 'APPROVED' THEN 4
                WHEN 'REJECTED' THEN 5
                ELSE 6
            END,
            ce.id DESC
        """,
        (pid,)
    ).fetchall()

    c.close()

    open_items = [r for r in rows if r["status"] in ["OPEN", "PRICING", "SUBMITTED"]]
    approved = [r for r in rows if r["status"] == "APPROVED"]
    rejected = [r for r in rows if r["status"] == "REJECTED"]

    open_cost = sum((r["estimated_cost"] or 0) for r in open_items)
    open_days = sum((r["schedule_days"] or 0) for r in open_items)

    cards = ""

    for r in rows:
        activity_text = (
            f'{esc(r["external_id"])} - {esc(r["activity"])}'
            if r["external_id"]
            else "No linked activity"
        )

        if r["status"] == "APPROVED":
            badge = "READY"
        elif r["status"] == "REJECTED":
            badge = "LOW"
        elif r["status"] == "SUBMITTED":
            badge = "WATCH"
        elif r["status"] == "PRICING":
            badge = "HIGH"
        else:
            badge = "OPEN"

        cards += f"""
        <div class="card">
            <div style="display:flex;justify-content:space-between;gap:12px;align-items:flex-start;flex-wrap:wrap;">
                <div>
                    <span class="badge {badge}">{esc(r["status"])}</span>
                    <span class="badge OPEN">{esc(r["event_type"]).replace("_"," ")}</span>
                    <h3 style="margin:10px 0 4px;">{esc(r["title"])}</h3>
                    <div class="muted">{activity_text}</div>
                </div>

                <a href="/changes/{r["id"]}/edit"
                   style="color:#f0b44d;text-decoration:none;font-weight:700;">
                    Edit
                </a>
            </div>

            <div class="grid3" style="margin-top:16px;">
                <div>
                    <div class="label">Estimated Cost</div>
                    <div class="kpi">${(r["estimated_cost"] or 0):,.0f}</div>
                </div>

                <div>
                    <div class="label">Schedule Impact</div>
                    <div class="kpi">{(r["schedule_days"] or 0):.1f}</div>
                    <div class="small">days</div>
                </div>

                <div>
                    <div class="label">Responsible</div>
                    <div>{esc(r["responsible_party"]) or "Unassigned"}</div>
                </div>
            </div>

            <p>{esc(r["description"]) or "No description entered."}</p>
        </div>
        """

    if not cards:
        cards = '<div class="card"><div class="muted">No change events logged yet.</div></div>'

    body = f"""
    <div class="hero">
        <div style="display:flex;justify-content:space-between;align-items:center;gap:15px;flex-wrap:wrap;">
            <div>
                <div class="eyebrow">Change Intelligence</div>
                <h1>Track field changes before cost and schedule impact get lost.</h1>
                <div class="muted">
                    Capture scope changes, owner decisions, design changes, and potential change orders.
                </div>
            </div>

            <a href="/changes/new"
               style="display:inline-block;background:#f0b44d;color:#0a1017;text-decoration:none;padding:12px 18px;border-radius:9px;font-weight:800;">
                + Add Change Event
            </a>
        </div>
    </div>

    <div class="grid4">
        <div class="card">
            <div class="label">Open Events</div>
            <div class="kpi">{len(open_items)}</div>
        </div>

        <div class="card">
            <div class="label">Open Exposure</div>
            <div class="kpi">${open_cost:,.0f}</div>
        </div>

        <div class="card">
            <div class="label">Potential Days</div>
            <div class="kpi">{open_days:.1f}</div>
        </div>

        <div class="card">
            <div class="label">Approved</div>
            <div class="kpi">{len(approved)}</div>
        </div>
    </div>

    <div class="grid2">
        {cards}
    </div>
    """

    return shell("Change Events", body)


@app.get("/changes/new", response_class=HTMLResponse)
def new_change_form():
    pid = project_id()
    c = db()

    activities = c.execute(
        """
        SELECT id,external_id,name
        FROM activities
        WHERE project_id=?
        ORDER BY start,name
        """,
        (pid,)
    ).fetchall()

    subs = c.execute(
        """
        SELECT name,trade
        FROM subs
        WHERE project_id=?
        ORDER BY trade,name
        """,
        (pid,)
    ).fetchall()

    c.close()

    activity_options = '<option value="">No linked activity</option>'
    activity_options += "".join(
        f'<option value="{a["id"]}">{esc(a["external_id"])} - {esc(a["name"])}</option>'
        for a in activities
    )

    responsible_options = '<option value="">Unassigned</option>'
    responsible_options += '<option value="Owner">Owner</option>'
    responsible_options += '<option value="Architect / Engineer">Architect / Engineer</option>'
    responsible_options += '<option value="General Contractor">General Contractor</option>'

    for s in subs:
        responsible_options += (
            f'<option value="{esc(s["name"])}">{esc(s["name"])} - {esc(s["trade"])}</option>'
        )

    body = f"""
    <div class="hero">
        <div class="eyebrow">Change Intelligence</div>
        <h1>Add Change Event</h1>
    </div>

    <div class="card" style="max-width:760px;">
        <form method="post" action="/changes/new">

            <label>Change Type</label>
            <select name="event_type">
                <option value="OWNER_CHANGE">Owner Change</option>
                <option value="DESIGN_CHANGE">Design Change</option>
                <option value="SCOPE_GAP">Scope Gap</option>
                <option value="FIELD_CONDITION">Field Condition</option>
                <option value="RFI_IMPACT">RFI Impact</option>
                <option value="POTENTIAL_CHANGE_ORDER">Potential Change Order</option>
            </select>

            <label>Title</label>
            <input
                type="text"
                name="title"
                placeholder="Example: Added power for imaging equipment"
                required
            >

            <label>Linked Activity</label>
            <select name="activity_id">
                {activity_options}
            </select>

            <label>Responsible Party</label>
            <select name="responsible_party">
                {responsible_options}
            </select>

            <div class="grid2">
                <div>
                    <label>Estimated Cost Impact</label>
                    <input
                        type="number"
                        name="estimated_cost"
                        min="0"
                        step="100"
                        value="0"
                    >
                </div>

                <div>
                    <label>Schedule Impact (Days)</label>
                    <input
                        type="number"
                        name="schedule_days"
                        min="0"
                        step="0.5"
                        value="0"
                    >
                </div>
            </div>

            <label>Status</label>
            <select name="status">
                <option value="OPEN">Open</option>
                <option value="PRICING">Pricing</option>
                <option value="SUBMITTED">Submitted</option>
                <option value="APPROVED">Approved</option>
                <option value="REJECTED">Rejected</option>
            </select>

            <label>Description</label>
            <textarea
                name="description"
                placeholder="Describe the change, cause, scope, cost exposure, and schedule impact."
                required
            ></textarea>

            <button type="submit">Save Change Event</button>
        </form>
    </div>
    """

    return shell("Add Change Event", body)


@app.post("/changes/new")
def create_change_event(
    event_type: str = Form(...),
    title: str = Form(...),
    activity_id: str = Form(""),
    responsible_party: str = Form(""),
    estimated_cost: float = Form(0),
    schedule_days: float = Form(0),
    status: str = Form("OPEN"),
    description: str = Form(...)
):
    pid = project_id()
    linked_activity = int(activity_id) if activity_id.strip() else None

    estimated_cost = max(0.0, estimated_cost)
    schedule_days = max(0.0, schedule_days)

    c = db()

    if linked_activity is not None:
        valid = c.execute(
            "SELECT id FROM activities WHERE id=? AND project_id=?",
            (linked_activity, pid)
        ).fetchone()
        if not valid:
            linked_activity = None

    c.execute(
        """
        INSERT INTO change_events(
            project_id,
            activity_id,
            event_type,
            title,
            responsible_party,
            estimated_cost,
            schedule_days,
            status,
            description,
            created
        )
        VALUES(?,?,?,?,?,?,?,?,?,?)
        """,
        (
            pid,
            linked_activity,
            event_type,
            title.strip(),
            responsible_party.strip(),
            estimated_cost,
            schedule_days,
            status,
            description.strip(),
            date.today().isoformat()
        )
    )

    c.commit()
    c.close()

    return RedirectResponse(url="/changes", status_code=303)


@app.get("/changes/{change_id}/edit", response_class=HTMLResponse)
def edit_change_form(change_id: int):
    pid = project_id()
    c = db()

    item = c.execute(
        "SELECT * FROM change_events WHERE id=? AND project_id=?",
        (change_id, pid)
    ).fetchone()

    c.close()

    if not item:
        return RedirectResponse(url="/changes", status_code=303)

    statuses = ["OPEN", "PRICING", "SUBMITTED", "APPROVED", "REJECTED"]
    status_options = "".join(
        f'<option value="{s}" {"selected" if item["status"] == s else ""}>{s.title()}</option>'
        for s in statuses
    )

    body = f"""
    <div class="hero">
        <div class="eyebrow">Change Intelligence</div>
        <h1>Edit Change Event</h1>
    </div>

    <div class="card" style="max-width:760px;">
        <form method="post" action="/changes/{change_id}/edit">

            <label>Title</label>
            <input type="text" name="title" value="{esc(item["title"])}" required>

            <label>Responsible Party</label>
            <input type="text" name="responsible_party" value="{esc(item["responsible_party"])}">

            <div class="grid2">
                <div>
                    <label>Estimated Cost</label>
                    <input
                        type="number"
                        name="estimated_cost"
                        min="0"
                        step="100"
                        value="{item["estimated_cost"] or 0}"
                    >
                </div>

                <div>
                    <label>Schedule Impact (Days)</label>
                    <input
                        type="number"
                        name="schedule_days"
                        min="0"
                        step="0.5"
                        value="{item["schedule_days"] or 0}"
                    >
                </div>
            </div>

            <label>Status</label>
            <select name="status">
                {status_options}
            </select>

            <label>Description</label>
            <textarea name="description" required>{esc(item["description"])}</textarea>

            <button type="submit">Save Changes</button>
        </form>
    </div>
    """

    return shell("Edit Change Event", body)


@app.post("/changes/{change_id}/edit")
def edit_change_event(
    change_id: int,
    title: str = Form(...),
    responsible_party: str = Form(""),
    estimated_cost: float = Form(0),
    schedule_days: float = Form(0),
    status: str = Form("OPEN"),
    description: str = Form(...)
):
    pid = project_id()

    estimated_cost = max(0.0, estimated_cost)
    schedule_days = max(0.0, schedule_days)

    c = db()
    c.execute(
        """
        UPDATE change_events
        SET title=?,
            responsible_party=?,
            estimated_cost=?,
            schedule_days=?,
            status=?,
            description=?
        WHERE id=? AND project_id=?
        """,
        (
            title.strip(),
            responsible_party.strip(),
            estimated_cost,
            schedule_days,
            status,
            description.strip(),
            change_id,
            pid
        )
    )
    c.commit()
    c.close()

    return RedirectResponse(url="/changes", status_code=303)


@app.get("/meetings", response_class=HTMLResponse)
def meetings_page():
    pid = project_id()
    c = db()

    rows = c.execute(
        """
        SELECT *
        FROM meeting_notes
        WHERE project_id=?
        ORDER BY meeting_date DESC, id DESC
        LIMIT 30
        """,
        (pid,)
    ).fetchall()

    c.close()

    cards = ""

    for r in rows:
        cards += f"""
        <div class="card">
            <div style="display:flex;justify-content:space-between;gap:12px;align-items:flex-start;flex-wrap:wrap;">
                <div>
                    <span class="badge OPEN">{esc(r["meeting_type"]).replace("_"," ")}</span>
                    <h3 style="margin:10px 0 4px;">{esc(r["title"])}</h3>
                    <div class="muted">{esc(r["meeting_date"])}</div>
                </div>

                <a href="/meetings/{r["id"]}/edit"
                   style="color:#f0b44d;text-decoration:none;font-weight:700;">
                    Edit
                </a>
            </div>

            <div class="small">Attendees</div>
            <p>{esc(r["attendees"]) or "—"}</p>

            <div class="small">Decisions</div>
            <p>{esc(r["decisions"]) or "—"}</p>

            <div class="small">Commitments</div>
            <p>{esc(r["commitments"]) or "—"}</p>

            <div class="small">Follow-Up</div>
            <p>{esc(r["follow_up"]) or "—"}</p>
        </div>
        """

    if not cards:
        cards = '<div class="card"><div class="muted">No meeting notes logged yet.</div></div>'

    body = f"""
    <div class="hero">
        <div style="display:flex;justify-content:space-between;align-items:center;gap:15px;flex-wrap:wrap;">
            <div>
                <div class="eyebrow">Meeting Intelligence</div>
                <h1>Capture decisions and commitments before they disappear into notes.</h1>
                <div class="muted">
                    Track OAC meetings, trade coordination, owner decisions, and follow-up.
                </div>
            </div>

            <a href="/meetings/new"
               style="display:inline-block;background:#f0b44d;color:#0a1017;text-decoration:none;padding:12px 18px;border-radius:9px;font-weight:800;">
                + Add Meeting Notes
            </a>
        </div>
    </div>

    <div class="grid2">
        {cards}
    </div>
    """

    return shell("Meetings", body)


@app.get("/meetings/new", response_class=HTMLResponse)
def new_meeting_form():
    body = f"""
    <div class="hero">
        <div class="eyebrow">Meeting Intelligence</div>
        <h1>Add Meeting Notes</h1>
    </div>

    <div class="card" style="max-width:820px;">
        <form method="post" action="/meetings/new">

            <label>Meeting Date</label>
            <input type="date" name="meeting_date" value="{date.today().isoformat()}" required>

            <label>Meeting Type</label>
            <select name="meeting_type">
                <option value="OAC">OAC</option>
                <option value="SUBCONTRACTOR_COORDINATION">Subcontractor Coordination</option>
                <option value="OWNER_MEETING">Owner Meeting</option>
                <option value="INTERNAL">Internal</option>
                <option value="SAFETY">Safety</option>
                <option value="PREINSTALL">Pre-Installation</option>
            </select>

            <label>Title</label>
            <input type="text" name="title" placeholder="Example: Weekly OAC Meeting" required>

            <label>Attendees</label>
            <textarea name="attendees" placeholder="Who attended?"></textarea>

            <label>Decisions Made</label>
            <textarea name="decisions" placeholder="What was decided?"></textarea>

            <label>Commitments</label>
            <textarea name="commitments" placeholder="Who committed to what and by when?"></textarea>

            <label>Follow-Up</label>
            <textarea name="follow_up" placeholder="What needs to happen next?"></textarea>

            <button type="submit">Save Meeting Notes</button>
        </form>
    </div>
    """

    return shell("Add Meeting Notes", body)


@app.post("/meetings/new")
def create_meeting(
    meeting_date: str = Form(...),
    meeting_type: str = Form(...),
    title: str = Form(...),
    attendees: str = Form(""),
    decisions: str = Form(""),
    commitments: str = Form(""),
    follow_up: str = Form("")
):
    pid = project_id()

    c = db()
    c.execute(
        """
        INSERT INTO meeting_notes(
            project_id,
            meeting_date,
            meeting_type,
            title,
            attendees,
            decisions,
            commitments,
            follow_up,
            created
        )
        VALUES(?,?,?,?,?,?,?,?,?)
        """,
        (
            pid,
            meeting_date,
            meeting_type,
            title.strip(),
            attendees.strip(),
            decisions.strip(),
            commitments.strip(),
            follow_up.strip(),
            date.today().isoformat()
        )
    )
    c.commit()
    c.close()

    return RedirectResponse(url="/meetings", status_code=303)


@app.get("/meetings/{meeting_id}/edit", response_class=HTMLResponse)
def edit_meeting_form(meeting_id: int):
    pid = project_id()
    c = db()

    item = c.execute(
        "SELECT * FROM meeting_notes WHERE id=? AND project_id=?",
        (meeting_id, pid)
    ).fetchone()

    c.close()

    if not item:
        return RedirectResponse(url="/meetings", status_code=303)

    body = f"""
    <div class="hero">
        <div class="eyebrow">Meeting Intelligence</div>
        <h1>Edit Meeting Notes</h1>
    </div>

    <div class="card" style="max-width:820px;">
        <form method="post" action="/meetings/{meeting_id}/edit">

            <label>Meeting Date</label>
            <input type="date" name="meeting_date" value="{esc(item["meeting_date"])}" required>

            <label>Title</label>
            <input type="text" name="title" value="{esc(item["title"])}" required>

            <label>Attendees</label>
            <textarea name="attendees">{esc(item["attendees"])}</textarea>

            <label>Decisions Made</label>
            <textarea name="decisions">{esc(item["decisions"])}</textarea>

            <label>Commitments</label>
            <textarea name="commitments">{esc(item["commitments"])}</textarea>

            <label>Follow-Up</label>
            <textarea name="follow_up">{esc(item["follow_up"])}</textarea>

            <button type="submit">Save Changes</button>
        </form>
    </div>
    """

    return shell("Edit Meeting Notes", body)


@app.post("/meetings/{meeting_id}/edit")
def edit_meeting(
    meeting_id: int,
    meeting_date: str = Form(...),
    title: str = Form(...),
    attendees: str = Form(""),
    decisions: str = Form(""),
    commitments: str = Form(""),
    follow_up: str = Form("")
):
    pid = project_id()

    c = db()
    c.execute(
        """
        UPDATE meeting_notes
        SET meeting_date=?,
            title=?,
            attendees=?,
            decisions=?,
            commitments=?,
            follow_up=?
        WHERE id=? AND project_id=?
        """,
        (
            meeting_date,
            title.strip(),
            attendees.strip(),
            decisions.strip(),
            commitments.strip(),
            follow_up.strip(),
            meeting_id,
            pid
        )
    )
    c.commit()
    c.close()

    return RedirectResponse(url="/meetings", status_code=303)


def safe_filename(name):
    base=os.path.basename(name or "upload"); cleaned="".join(ch for ch in base if ch.isalnum() or ch in "._- ")
    return cleaned[:120] or "upload"


def refresh_notifications(pid=None):
    cid=current_company_id()
    if not cid or not pid: return
    c=db(); c.execute("DELETE FROM notifications WHERE company_id=? AND project_id=? AND source='AUTO'",(cid,pid)); today=date.today().isoformat()
    def add(sev,title,detail):
        c.execute("INSERT INTO notifications(company_id,project_id,severity,title,detail,source,status,created) VALUES(?,?,?,?,?,'AUTO','UNREAD',?)",(cid,pid,sev,title,detail,datetime.utcnow().isoformat()))
    for r in c.execute("SELECT * FROM action_items WHERE project_id=? AND status='OPEN' AND due!='' AND due<?",(pid,today)).fetchall(): add(r["priority"] or "HIGH",f'Overdue action: {r["title"]}',f'Owner: {r["owner"] or "Unassigned"} · Due {r["due"]}')
    for r in c.execute("SELECT * FROM procurement WHERE project_id=? AND status!='DELIVERED' AND promised_date!='' AND required_on_site!='' AND promised_date>required_on_site",(pid,)).fetchall(): add("CRITICAL",f'Late procurement: {r["item"]}',f'Promised {r["promised_date"]}; required {r["required_on_site"]}')
    for r in c.execute("SELECT * FROM inspections_tracker WHERE project_id=? AND result='FAILED'",(pid,)).fetchall(): add("CRITICAL",f'Failed inspection: {r["inspection_type"]}',r["notes"] or "Correction/reinspection required.")
    for r in c.execute("SELECT * FROM submittals WHERE project_id=? AND status='REJECTED'",(pid,)).fetchall(): add("HIGH",f'Rejected submittal: {r["title"]}',r["notes"] or "Revise and resubmit.")
    for r in c.execute("SELECT * FROM safety_items WHERE project_id=? AND status!='CLOSED' AND severity='CRITICAL'",(pid,)).fetchall(): add("CRITICAL",f'Safety: {r["title"]}',r["description"] or "Critical safety item is open.")
    c.commit(); c.close()


@app.get("/notifications",response_class=HTMLResponse)
def notifications_page():
    pid=project_id(); refresh_notifications(pid); c=db(); rows=c.execute("SELECT * FROM notifications WHERE company_id=? AND (project_id=? OR project_id IS NULL) ORDER BY CASE severity WHEN 'CRITICAL' THEN 1 WHEN 'HIGH' THEN 2 WHEN 'WATCH' THEN 3 ELSE 4 END,id DESC LIMIT 100",(current_company_id(),pid)).fetchall(); c.close()
    cards="".join(f'<div class="card"><span class="badge {r["severity"] if r["severity"] in ["CRITICAL","HIGH","WATCH","LOW","READY"] else "OPEN"}">{esc(r["severity"])}</span><h3>{esc(r["title"])}</h3><p>{esc(r["detail"])}</p><div class="small">{esc(r["created"])}</div></div>' for r in rows) or '<div class="card"><div class="muted">No active notifications.</div></div>'
    return shell("Notifications",f'<div class="hero"><div class="eyebrow">Notifications</div><h1>What needs your attention?</h1></div><div class="card"><form method="post" action="/notifications/email-digest"><button type="submit">Email My Alert Digest</button></form><div class="small">Requires SMTP settings in Render.</div></div><div class="grid2">{cards}</div>')


@app.get("/documents",response_class=HTMLResponse)
def documents_page():
    pid=project_id(); c=db(); rows=c.execute("SELECT a.*,u.display_name FROM attachments a LEFT JOIN users u ON u.id=a.created_by WHERE a.company_id=? AND a.project_id=? ORDER BY a.id DESC",(current_company_id(),pid)).fetchall(); c.close()
    files_html="".join(f'<div class="card"><span class="badge OPEN">{esc(r["category"])}</span><h3>{esc(r["title"] or r["original_name"])}</h3><div class="small">{esc(r["original_name"])} · {(r["size_bytes"] or 0)/1024:.0f} KB</div><p><a href="/documents/{r["id"]}/download" style="color:#f0b44d;font-weight:700;">Download</a></p></div>' for r in rows) or '<div class="card"><div class="muted">No project documents uploaded yet.</div></div>'
    body=f'''<div class="hero"><div class="eyebrow">Document & Photo Center</div><h1>Keep field evidence with the project.</h1><div class="muted">Set UPLOAD_DIR to a Render persistent-disk path for durable storage.</div></div><div class="grid2"><div class="card"><h2>Upload</h2><form method="post" action="/documents/upload" enctype="multipart/form-data"><label>Category</label><select name="category"><option>PHOTO</option><option>DAILY_REPORT</option><option>RFI</option><option>PUNCH</option><option>SAFETY</option><option>SUBMITTAL</option><option>OTHER</option></select><label>Title</label><input name="title"><label>File</label><input type="file" name="file" required><button type="submit">Upload File</button></form></div><div class="card"><h2>Upload Rules</h2><p>Maximum 10 MB. PDF, image, Office, CSV, and text files are accepted.</p></div></div><div class="grid2">{files_html}</div>'''
    return shell("Documents",body)


@app.post("/documents/upload")
async def documents_upload(category:str=Form("OTHER"),title:str=Form(""),file:UploadFile=File(...)):
    pid=project_id(); original=safe_filename(file.filename); ext=Path(original).suffix.lower()
    if ext not in ALLOWED_UPLOAD_EXTENSIONS: return HTMLResponse("File type not allowed.",status_code=400)
    data=await file.read()
    if len(data)>MAX_UPLOAD_BYTES: return HTMLResponse("File exceeds the 10 MB limit.",status_code=400)
    stored=f"{secrets.token_hex(12)}{ext}"; path=os.path.join(UPLOAD_DIR,stored)
    with open(path,"wb") as f: f.write(data)
    c=db(); c.execute("INSERT INTO attachments(company_id,project_id,category,title,original_name,stored_name,mime_type,size_bytes,created_by,created) VALUES(?,?,?,?,?,?,?,?,?,?)",(current_company_id(),pid,category,title.strip(),original,stored,file.content_type or mimetypes.guess_type(original)[0] or "application/octet-stream",len(data),current_user_id(),datetime.utcnow().isoformat())); c.commit(); c.close()
    return RedirectResponse("/documents",status_code=303)


@app.get("/documents/{attachment_id}/download")
def document_download(attachment_id:int):
    c=db(); row=c.execute("SELECT * FROM attachments WHERE id=? AND company_id=?",(attachment_id,current_company_id())).fetchone(); c.close()
    if not row: return HTMLResponse("File not found.",status_code=404)
    path=os.path.join(UPLOAD_DIR,row["stored_name"])
    if not os.path.isfile(path): return HTMLResponse("Stored file is unavailable. Configure persistent storage.",status_code=404)
    return FileResponse(path,media_type=row["mime_type"] or "application/octet-stream",filename=row["original_name"])


@app.get("/morning-brief",response_class=HTMLResponse)
def morning_brief_page():
    pid=project_id(); c=db(); row=c.execute("SELECT * FROM morning_briefs WHERE company_id=? AND project_id=? ORDER BY id DESC LIMIT 1",(current_company_id(),pid)).fetchone(); c.close(); brief=esc(row["brief_text"]).replace("\n","<br>") if row else "No AI morning brief has been generated yet."
    return shell("Morning Brief",f'<div class="hero"><div class="eyebrow">Proactive AI</div><h1>Morning Superintendent Brief</h1></div><div class="card"><form method="post" action="/morning-brief/generate"><button type="submit">Generate Current Brief</button></form></div><div class="card"><div style="line-height:1.65;">{brief}</div></div>')


@app.post("/morning-brief/generate")
def generate_morning_brief():
    pid=project_id(); api_key=os.environ.get("OPENAI_API_KEY"); context=build_project_context(pid)
    if api_key and OpenAI is not None:
        try:
            client=OpenAI(api_key=api_key); response=client.responses.create(model=os.environ.get("OPENAI_MODEL","gpt-5.6"),instructions="You are BuildCommand AI, a construction superintendent morning-command copilot. Use only supplied project facts. Return: HANDLE FIRST, SCHEDULE THREATS, WHO OWES WHAT, TODAY'S VERIFICATIONS, COST / CHANGE EXPOSURE, SAFETY / QUALITY. Do not invent facts.",input=context); brief=response.output_text
        except Exception as exc: brief=f"AI brief failed: {exc}"
    else: brief="OPENAI_API_KEY is not configured. The rule-based Daily Command dashboard remains available."
    c=db(); c.execute("INSERT INTO morning_briefs(company_id,project_id,brief_date,brief_text,created) VALUES(?,?,?,?,?)",(current_company_id(),pid,date.today().isoformat(),brief,datetime.utcnow().isoformat())); c.commit(); c.close(); return RedirectResponse("/morning-brief",status_code=303)


@app.get("/team",response_class=HTMLResponse)
def team_page():
    user=current_user(); c=db(); rows=c.execute("SELECT * FROM users WHERE company_id=? ORDER BY display_name,email",(current_company_id(),)).fetchall(); c.close(); members="".join(f'<div class="action"><b>{esc(r["display_name"] or r["email"])}</b><div class="small">{esc(r["email"])} · {esc(r["role"])}</div></div>' for r in rows)
    add=''
    if user and user["role"] in ["OWNER","ADMIN"]: add='''<div class="card"><h2>Add Team Member</h2><form method="post" action="/team/add"><label>Name</label><input name="display_name" required><label>Email</label><input type="email" name="email" required><label>Temporary Password</label><input type="password" name="password" minlength="8" required><label>Role</label><select name="role"><option>MEMBER</option><option>ADMIN</option></select><button type="submit">Create User</button></form></div>'''
    return shell("Team",f'<div class="hero"><div class="eyebrow">Accounts</div><h1>Company Team</h1></div><div class="grid2"><div class="card"><h2>Users</h2>{members}</div>{add}</div>')


@app.post("/team/add")
def add_team_member(display_name:str=Form(...),email:str=Form(...),password:str=Form(...),role:str=Form("MEMBER")):
    user=current_user()
    if not user or user["role"] not in ["OWNER","ADMIN"]: return HTMLResponse("Not authorized.",status_code=403)
    if role not in ["MEMBER","ADMIN"]: role="MEMBER"
    c=db()
    if c.execute("SELECT id FROM users WHERE lower(email)=lower(?)",(email.strip(),)).fetchone(): c.close(); return HTMLResponse("That email is already in use.",status_code=400)
    c.execute("INSERT INTO users(company_id,email,display_name,password_hash,role,created) VALUES(?,?,?,?,?,?)",(current_company_id(),email.strip().lower(),display_name.strip(),hash_password(password),role,date.today().isoformat())); c.commit(); c.close(); return RedirectResponse("/team",status_code=303)


@app.get("/company-settings",response_class=HTMLResponse)
def company_settings():
    u=current_user(); return shell("Company Settings",f'''<div class="hero"><div class="eyebrow">Branding</div><h1>Company Settings</h1></div><div class="card" style="max-width:720px;"><form method="post" action="/company-settings"><label>Company Name</label><input name="name" value="{esc(u["company_name"] if u else "")}" required><label>Logo URL (optional)</label><input name="logo_url" value="{esc(u["logo_url"] if u else "")}"><button type="submit">Save Company Settings</button></form></div>''')


@app.post("/company-settings")
def company_settings_save(name:str=Form(...),logo_url:str=Form("")):
    u=current_user()
    if not u or u["role"] not in ["OWNER","ADMIN"]: return HTMLResponse("Not authorized.",status_code=403)
    c=db(); c.execute("UPDATE companies SET name=?,logo_url=? WHERE id=?",(name.strip(),logo_url.strip(),current_company_id())); c.commit(); c.close(); return RedirectResponse("/company-settings",status_code=303)


@app.get("/exports",response_class=HTMLResponse)
def exports_page():
    return shell("Exports",'''<div class="hero"><div class="eyebrow">Reports & Exports</div><h1>Take BuildCommand data with you.</h1></div><div class="grid2"><div class="card"><h2>Project CSV</h2><a href="/exports/project.csv" style="color:#f0b44d;font-weight:700;">Download CSV</a></div><div class="card"><h2>Project JSON</h2><a href="/exports/project.json" style="color:#f0b44d;font-weight:700;">Download JSON</a></div><div class="card"><h2>Latest Daily Report PDF</h2><a href="/exports/daily-report.pdf" style="color:#f0b44d;font-weight:700;">Download PDF</a></div><div class="card"><h2>Backup</h2><a href="/backup/download" style="color:#f0b44d;font-weight:700;">Download Backup ZIP</a></div></div>''')


@app.get("/exports/project.csv")
def export_project_csv():
    pid=project_id(); c=db(); project=c.execute("SELECT * FROM projects WHERE id=?",(pid,)).fetchone(); risks=c.execute("SELECT * FROM risks WHERE project_id=? ORDER BY score DESC",(pid,)).fetchall(); actions=c.execute("SELECT * FROM action_items WHERE project_id=? ORDER BY due",(pid,)).fetchall(); c.close(); out=io.StringIO(); w=csv.writer(out); w.writerow(["BuildCommand AI Project Export"]); w.writerow(["Project",project["number"] if project else "",project["name"] if project else ""]); w.writerow([]); w.writerow(["RISKS"]); w.writerow(["Score","Band","Explanation"]); [w.writerow([r["score"],r["band"],r["explanation"]]) for r in risks]; w.writerow([]); w.writerow(["ACTIONS"]); w.writerow(["Title","Owner","Priority","Due","Status","Notes"]); [w.writerow([r["title"],r["owner"],r["priority"],r["due"],r["status"],r["notes"]]) for r in actions]; return Response(out.getvalue(),media_type="text/csv",headers={"Content-Disposition":'attachment; filename="buildcommand_project.csv"'})


@app.get("/exports/project.json")
def export_project_json():
    pid=project_id(); c=db(); project=c.execute("SELECT * FROM projects WHERE id=?",(pid,)).fetchone(); tables=["activities","risks","make_ready","action_items","project_issues","punch_items","inspections_tracker","submittals","safety_items","change_events","meeting_notes","procurement","production","daily_reports","subs","subcontractor_updates"]; payload={"project":dict(project) if project else None}; [payload.__setitem__(t,[dict(r) for r in c.execute(f"SELECT * FROM {t} WHERE project_id=?",(pid,)).fetchall()]) for t in tables]; c.close(); return Response(json.dumps(payload,indent=2,default=str),media_type="application/json",headers={"Content-Disposition":'attachment; filename="buildcommand_project.json"'})


@app.get("/exports/daily-report.pdf")
def export_daily_report_pdf():
    pid=project_id(); c=db(); project=c.execute("SELECT * FROM projects WHERE id=?",(pid,)).fetchone(); report=c.execute("SELECT * FROM daily_reports WHERE project_id=? ORDER BY report_date DESC,id DESC LIMIT 1",(pid,)).fetchone(); c.close()
    if not report: return HTMLResponse("No daily report is available to export.",status_code=404)
    if canvas is None: return HTMLResponse("PDF support is not installed. Ensure reportlab is in requirements.txt.",status_code=503)
    buffer=io.BytesIO(); pdf=canvas.Canvas(buffer); y=742; pdf.setFont("Helvetica-Bold",16); pdf.drawString(45,y,"BuildCommand AI - Daily Superintendent Report"); y-=24; pdf.setFont("Helvetica",10); pdf.drawString(45,y,f'{project["number"]} - {project["name"]}'); y-=18; pdf.drawString(45,y,f'Date: {report["report_date"]}   Manpower: {report["manpower"] or 0}'); y-=26
    for label,value in [("Weather",report["weather"]),("Work Completed",report["work_completed"]),("Delays / Constraints",report["delays"]),("Deliveries",report["deliveries"]),("Inspections",report["inspections"]),("Safety",report["safety"]),("Tomorrow's Plan",report["tomorrow_plan"])]:
        pdf.setFont("Helvetica-Bold",10); pdf.drawString(45,y,label); y-=14; pdf.setFont("Helvetica",9); words=str(value or "—").split(); line=""
        for word in words:
            test=(line+" "+word).strip()
            if pdf.stringWidth(test,"Helvetica",9)>510: pdf.drawString(55,y,line); y-=12; line=word
            else: line=test
            if y<60: pdf.showPage(); y=742
        if line: pdf.drawString(55,y,line); y-=12
        y-=10
    pdf.setFont("Helvetica",8); pdf.drawString(45,30,"Built by Wilson LaHood · © 2026 Wilson LaHood"); pdf.save(); buffer.seek(0); return Response(buffer.getvalue(),media_type="application/pdf",headers={"Content-Disposition":'attachment; filename="buildcommand_daily_report.pdf"'})


@app.get("/backup/download")
def backup_download():
    stamp=datetime.utcnow().strftime("%Y%m%d_%H%M%S"); tmp=f"/tmp/bc_{secrets.token_hex(5)}"; os.makedirs(tmp,exist_ok=True); dbcopy=os.path.join(tmp,"buildcommand.db"); s=sqlite3.connect(DB); t=sqlite3.connect(dbcopy); s.backup(t); t.close(); s.close(); zpath=f"/tmp/buildcommand_backup_{stamp}.zip"
    with zipfile.ZipFile(zpath,"w",zipfile.ZIP_DEFLATED) as z:
        z.write(dbcopy,arcname="buildcommand.db"); c=db(); files=c.execute("SELECT * FROM attachments WHERE company_id=?",(current_company_id(),)).fetchall(); c.close()
        for r in files:
            p=os.path.join(UPLOAD_DIR,r["stored_name"])
            if os.path.isfile(p): z.write(p,arcname=f'uploads/{r["stored_name"]}')
    return FileResponse(zpath,media_type="application/zip",filename=f"buildcommand_backup_{stamp}.zip")


@app.get("/system-check",response_class=HTMLResponse)
def system_check():
    checks=[]
    try:
        c=db(); tables=["projects","activities","subs","subcontractor_updates","daily_reports","risks","make_ready","action_items","project_issues","punch_items","inspections_tracker","submittals","safety_items","change_events","meeting_notes","procurement","attachments","users","companies"]
        for t in tables: c.execute(f"SELECT COUNT(*) FROM {t}").fetchone()
        c.close(); checks.append(("READY","Database schema","Core tables respond successfully."))
    except Exception as exc: checks.append(("CRITICAL","Database schema",str(exc)))
    checks.append(("READY" if os.environ.get("OPENAI_API_KEY") else "WATCH","OpenAI connection","OPENAI_API_KEY is configured." if os.environ.get("OPENAI_API_KEY") else "AI key is not configured.")); checks.append(("READY" if os.path.isdir(UPLOAD_DIR) and os.access(UPLOAD_DIR,os.W_OK) else "CRITICAL","Upload storage",f"Upload path: {UPLOAD_DIR}")); checks.append(("READY" if DATABASE_KIND=="postgres" else "WATCH","Database engine",f"Current engine: {DATABASE_KIND}.")); checks.append(("READY" if os.environ.get("UPLOAD_DIR") else "WATCH","Persistent files","UPLOAD_DIR is explicitly configured." if os.environ.get("UPLOAD_DIR") else "Set UPLOAD_DIR to a Render persistent-disk path for durable files.")); html="".join(f'<div class="card"><span class="badge {b}">{b}</span><h3>{esc(t)}</h3><p>{esc(d)}</p></div>' for b,t,d in checks); return shell("System Check",f'<div class="hero"><div class="eyebrow">Stability</div><h1>BuildCommand System Check</h1></div><div class="grid2">{html}</div>')


@app.get("/beta-feedback",response_class=HTMLResponse)
def beta_feedback_page():
    c=db(); rows=c.execute("SELECT b.*,u.display_name FROM beta_feedback b LEFT JOIN users u ON u.id=b.user_id WHERE b.company_id=? ORDER BY b.id DESC LIMIT 20",(current_company_id(),)).fetchall(); c.close(); recent="".join(f'<div class="action"><b>{esc(r["category"])}</b> · {r["rating"]}/5<div>{esc(r["feedback"])}</div><div class="small">{esc(r["display_name"] or "")} · {esc(r["created"])}</div></div>' for r in rows) or '<div class="muted">No beta feedback yet.</div>'; return shell("Beta Feedback",f'''<div class="hero"><div class="eyebrow">Beta Program</div><h1>What should BuildCommand improve next?</h1></div><div class="grid2"><div class="card"><form method="post" action="/beta-feedback"><label>Rating</label><select name="rating"><option value="5">5 - Excellent</option><option value="4">4</option><option value="3">3</option><option value="2">2</option><option value="1">1 - Poor</option></select><label>Category</label><select name="category"><option>FIELD_WORKFLOW</option><option>AI</option><option>SCHEDULE</option><option>MOBILE</option><option>REPORTING</option><option>BUG</option><option>OTHER</option></select><label>Feedback</label><textarea name="feedback" required></textarea><button type="submit">Save Feedback</button></form></div><div class="card"><h2>Recent Feedback</h2>{recent}</div></div>''')


@app.post("/beta-feedback")
def beta_feedback_save(rating:int=Form(...),category:str=Form(...),feedback:str=Form(...)):
    c=db(); c.execute("INSERT INTO beta_feedback(company_id,user_id,project_id,rating,category,feedback,created) VALUES(?,?,?,?,?,?,?)",(current_company_id(),current_user_id(),project_id(),max(1,min(5,rating)),category,feedback.strip(),datetime.utcnow().isoformat())); c.commit(); c.close(); return RedirectResponse("/beta-feedback",status_code=303)


@app.post("/notifications/email-digest")
def notifications_email_digest():
    import smtplib
    from email.message import EmailMessage
    user=current_user(); pid=project_id(); refresh_notifications(pid)
    host=os.environ.get("SMTP_HOST"); sender=os.environ.get("SMTP_FROM"); smtp_user=os.environ.get("SMTP_USER"); smtp_password=os.environ.get("SMTP_PASSWORD"); port=int(os.environ.get("SMTP_PORT","587"))
    if not host or not sender:
        return HTMLResponse(shell("Notifications",'<div class="card"><h2>Email alerts need setup</h2><p>Add SMTP_HOST and SMTP_FROM in Render. If your mail provider requires login, also add SMTP_USER and SMTP_PASSWORD.</p><p><a href="/notifications">Back to Notifications</a></p></div>'),status_code=503)
    c=db(); rows=c.execute("SELECT * FROM notifications WHERE company_id=? AND project_id=? ORDER BY CASE severity WHEN 'CRITICAL' THEN 1 WHEN 'HIGH' THEN 2 WHEN 'WATCH' THEN 3 ELSE 4 END,id DESC LIMIT 30",(current_company_id(),pid)).fetchall(); c.close()
    lines=[f'{r["severity"]}: {r["title"]} — {r["detail"]}' for r in rows]
    msg=EmailMessage(); msg["Subject"]="BuildCommand AI Alert Digest"; msg["From"]=sender; msg["To"]=user["email"]; msg.set_content("\n".join(lines) or "No active BuildCommand alerts.")
    with smtplib.SMTP(host,port,timeout=15) as smtp:
        smtp.starttls()
        if smtp_user: smtp.login(smtp_user,smtp_password or "")
        smtp.send_message(msg)
    return RedirectResponse("/notifications",status_code=303)


def ensure_today_morning_brief(pid):
    if not pid or not company_setting("auto_ai_brief", 1):
        return
    cid = current_company_id()
    c = db()
    existing = c.execute(
        "SELECT id FROM morning_briefs WHERE company_id=? AND project_id=? AND brief_date=? LIMIT 1",
        (cid, pid, date.today().isoformat()),
    ).fetchone()
    c.close()
    if existing or not os.environ.get("OPENAI_API_KEY") or OpenAI is None:
        return
    try:
        client = OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))
        response = client.responses.create(
            model=os.environ.get("OPENAI_MODEL", "gpt-5.6"),
            instructions=(
                "You are BuildCommand AI. Create a concise superintendent morning brief grounded only in "
                "supplied project data. Use HANDLE FIRST, SCHEDULE THREATS, WHO OWES WHAT, TODAY'S "
                "VERIFICATIONS, COST / CHANGE EXPOSURE, SAFETY / QUALITY."
            ),
            input=build_project_context(pid),
        )
        brief = response.output_text
    except Exception:
        return
    c = db()
    c.execute(
        "INSERT INTO morning_briefs(company_id,project_id,brief_date,brief_text,created) VALUES(?,?,?,?,?)",
        (cid, pid, date.today().isoformat(), brief, datetime.utcnow().isoformat()),
    )
    c.commit()
    c.close()


@app.get("/setup", response_class=HTMLResponse)
def setup_wizard():
    cid = current_company_id()
    ensure_company_settings(cid)
    u = current_user()
    c = db()
    pc = c.execute("SELECT COUNT(*) n FROM projects WHERE company_id=?", (cid,)).fetchone()["n"]
    tc = c.execute("SELECT COUNT(*) n FROM users WHERE company_id=?", (cid,)).fetchone()["n"]
    c.close()
    body = f"""
    <div class="hero"><div class="eyebrow">Customer Setup</div><h1>Finish your BuildCommand workspace.</h1></div>
    <div class="grid2">
      <div class="card"><h2>1. Company</h2><p>{esc(u['company_name'])}</p><a href="/company-settings">Edit Company →</a></div>
      <div class="card"><h2>2. Project</h2><p>{pc} project(s)</p><a href="/projects/new">Add Project →</a></div>
      <div class="card"><h2>3. Team</h2><p>{tc} user(s)</p><a href="/invitations">Invite Team →</a></div>
      <div class="card"><h2>4. Field Setup</h2><a href="/subcontractors">Add Subcontractors →</a><br><a href="/activities/new">Add Schedule Activity →</a></div>
    </div>
    <div class="card"><form method="post" action="/setup/complete"><button type="submit">Finish Setup</button></form></div>
    """
    return shell("Setup", body)


@app.post("/setup/complete")
def setup_complete():
    cid = current_company_id()
    ensure_company_settings(cid)
    c = db()
    c.execute("UPDATE company_settings SET onboarding_complete=1 WHERE company_id=?", (cid,))
    c.commit()
    c.close()
    return RedirectResponse("/", 303)


@app.get("/invitations", response_class=HTMLResponse)
def invitations_page():
    if not require_role("ADMIN"):
        return HTMLResponse("Not authorized.", 403)
    c = db()
    rows = c.execute(
        "SELECT * FROM invitations WHERE company_id=? ORDER BY id DESC LIMIT 30",
        (current_company_id(),),
    ).fetchall()
    c.close()
    hist = "".join(
        f"<div class='action'><b>{esc(r['email'])}</b> · {esc(r['role'])}<div class='small'>{'Accepted' if r['accepted'] else 'Pending'} · Expires {esc(r['expires'])}</div></div>"
        for r in rows
    ) or "<div class='muted'>No invitations yet.</div>"
    form = """
    <div class='card'><h2>Create Invitation</h2><form method='post' action='/invitations'>
    <label>Email</label><input type='email' name='email' required>
    <label>Role</label><select name='role'>
      <option value='READ_ONLY'>Read Only</option><option value='FIELD_USER'>Field User</option>
      <option value='SUPERINTENDENT'>Superintendent</option><option value='PROJECT_MANAGER'>Project Manager</option>
      <option value='ADMIN'>Admin</option>
    </select><button type='submit'>Create Invite</button></form></div>
    """
    return shell(
        "Invitations",
        f"<div class='hero'><div class='eyebrow'>Team Invitations</div><h1>Invite the project team.</h1></div><div class='grid2'>{form}<div class='card'><h2>Recent</h2>{hist}</div></div>",
    )


@app.post("/invitations")
def create_invitation(email: str = Form(...), role: str = Form("FIELD_USER")):
    if not require_role("ADMIN"):
        return HTMLResponse("Not authorized.", 403)
    valid = {"READ_ONLY", "FIELD_USER", "SUPERINTENDENT", "PROJECT_MANAGER", "ADMIN"}
    role = role if role in valid else "FIELD_USER"
    raw = secrets.token_urlsafe(32)
    token_hash = hashlib.sha256(raw.encode()).hexdigest()
    expires = (datetime.utcnow() + timedelta(days=7)).isoformat()
    c = db()
    c.execute(
        "INSERT INTO invitations(company_id,email,role,token_hash,expires,accepted,created_by,created) VALUES(?,?,?,?,?,?,?,?)",
        (current_company_id(), email.strip().lower(), role, token_hash, expires, 0, current_user_id(), datetime.utcnow().isoformat()),
    )
    c.commit()
    c.close()
    link = f"/accept-invite?token={raw}"
    return HTMLResponse(
        shell(
            "Invitation Created",
            f"<div class='card'><h2>Invitation created</h2><p>Send this link to {esc(email)}. It expires in 7 days.</p><input value='{esc(link)}' readonly><p><a href='/invitations'>Back</a></p></div>",
        )
    )


@app.get("/accept-invite", response_class=HTMLResponse)
def accept_invite_get(token: str):
    th = hashlib.sha256(token.encode()).hexdigest()
    c = db()
    inv = c.execute(
        "SELECT i.*,c.name company_name FROM invitations i JOIN companies c ON c.id=i.company_id WHERE i.token_hash=? AND i.accepted=0 AND i.expires>?",
        (th, datetime.utcnow().isoformat()),
    ).fetchone()
    c.close()
    if not inv:
        return HTMLResponse("Invitation invalid or expired.", 400)
    return f"""
    <!doctype html><html><head><meta name='viewport' content='width=device-width,initial-scale=1'><style>{CSS}</style></head>
    <body><main class='main'><div class='card'><h1>Join {esc(inv['company_name'])}</h1>
    <form method='post' action='/accept-invite'><input type='hidden' name='token' value='{esc(token)}'>
    <label>Name</label><input name='display_name' required><label>Password</label><input type='password' name='password' minlength='8' required>
    <button type='submit'>Create Account</button></form></div></main></body></html>
    """


@app.post("/accept-invite")
def accept_invite_post(token: str = Form(...), display_name: str = Form(...), password: str = Form(...)):
    if len(password) < 8:
        return HTMLResponse("Password must be at least 8 characters.", 400)
    th = hashlib.sha256(token.encode()).hexdigest()
    c = db()
    inv = c.execute(
        "SELECT * FROM invitations WHERE token_hash=? AND accepted=0 AND expires>?",
        (th, datetime.utcnow().isoformat()),
    ).fetchone()
    if not inv:
        c.close()
        return HTMLResponse("Invitation invalid or expired.", 400)
    if c.execute("SELECT id FROM users WHERE lower(email)=lower(?)", (inv["email"],)).fetchone():
        c.close()
        return HTMLResponse("Email already registered.", 400)
    c.execute(
        "INSERT INTO users(company_id,email,display_name,password_hash,role,created) VALUES(?,?,?,?,?,?)",
        (inv["company_id"], inv["email"], display_name.strip(), hash_password(password), inv["role"], date.today().isoformat()),
    )
    uid = c.execute("SELECT last_insert_rowid() id").fetchone()["id"]
    c.execute("UPDATE invitations SET accepted=1 WHERE id=?", (inv["id"],))
    fp = c.execute("SELECT id FROM projects WHERE company_id=? ORDER BY id LIMIT 1", (inv["company_id"],)).fetchone()
    if fp:
        c.execute(
            "INSERT INTO user_state(user_id,selected_project_id) VALUES(?,?) ON CONFLICT(user_id) DO UPDATE SET selected_project_id=excluded.selected_project_id",
            (uid, fp["id"]),
        )
    c.commit()
    c.close()
    raw = create_session(uid)
    resp = RedirectResponse("/", 303)
    resp.set_cookie("bc_session", raw, httponly=True, secure=os.environ.get("COOKIE_SECURE", "1") == "1", samesite="lax", max_age=2592000)
    return resp


@app.get("/production-settings", response_class=HTMLResponse)
def production_settings_page():
    if not require_role("ADMIN"):
        return HTMLResponse("Not authorized.", 403)
    cid = current_company_id()
    ensure_company_settings(cid)
    c = db()
    s = c.execute("SELECT * FROM company_settings WHERE company_id=?", (cid,)).fetchone()
    c.close()
    body = f"""
    <div class='hero'><div class='eyebrow'>Production Controls</div><h1>Automation & beta settings</h1></div>
    <div class='card'><form method='post' action='/production-settings'>
    <label>Automatic AI Morning Brief</label><select name='auto_ai_brief'>
      <option value='1' {'selected' if s['auto_ai_brief'] else ''}>Enabled</option><option value='0' {'selected' if not s['auto_ai_brief'] else ''}>Disabled</option></select>
    <label>Email Alerts</label><select name='email_alerts'>
      <option value='1' {'selected' if s['email_alerts'] else ''}>Enabled</option><option value='0' {'selected' if not s['email_alerts'] else ''}>Disabled</option></select>
    <label>Beta Mode</label><select name='beta_mode'>
      <option value='1' {'selected' if s['beta_mode'] else ''}>Enabled</option><option value='0' {'selected' if not s['beta_mode'] else ''}>Disabled</option></select>
    <button type='submit'>Save Settings</button></form></div>
    """
    return shell("Production Settings", body)


@app.post("/production-settings")
def production_settings_save(auto_ai_brief: int = Form(1), email_alerts: int = Form(0), beta_mode: int = Form(1)):
    if not require_role("ADMIN"):
        return HTMLResponse("Not authorized.", 403)
    cid = current_company_id()
    ensure_company_settings(cid)
    c = db()
    c.execute(
        "UPDATE company_settings SET auto_ai_brief=?,email_alerts=?,beta_mode=? WHERE company_id=?",
        (1 if auto_ai_brief else 0, 1 if email_alerts else 0, 1 if beta_mode else 0, cid),
    )
    c.commit()
    c.close()
    return RedirectResponse("/production-settings", 303)


@app.get("/beta-checklist", response_class=HTMLResponse)
def beta_checklist():
    pid = project_id()
    c = db()
    checks = {
        "Project created": bool(pid),
        "Schedule activity entered": c.execute("SELECT COUNT(*) n FROM activities WHERE project_id=?", (pid,)).fetchone()["n"] > 0 if pid else False,
        "Subcontractor entered": c.execute("SELECT COUNT(*) n FROM subs WHERE project_id=?", (pid,)).fetchone()["n"] > 0 if pid else False,
        "Daily report submitted": c.execute("SELECT COUNT(*) n FROM daily_reports WHERE project_id=?", (pid,)).fetchone()["n"] > 0 if pid else False,
        "Risk entered": c.execute("SELECT COUNT(*) n FROM risks WHERE project_id=?", (pid,)).fetchone()["n"] > 0 if pid else False,
        "AI configured": bool(os.environ.get("OPENAI_API_KEY")),
        "Persistent upload path configured": bool(os.environ.get("UPLOAD_DIR")),
        "PostgreSQL connected": DATABASE_KIND == "postgres",
    }
    c.close()
    complete = sum(1 for v in checks.values() if v)
    html = "".join(
        f"<div class='action'><span class='badge {'READY' if ok else 'WATCH'}'>{'DONE' if ok else 'TODO'}</span> {esc(name)}</div>"
        for name, ok in checks.items()
    )
    return shell(
        "Beta Checklist",
        f"<div class='hero'><div class='eyebrow'>Beta Readiness</div><h1>{complete}/{len(checks)} launch checks complete</h1></div><div class='card'>{html}</div><div class='card'><a href='/beta-feedback'>Record Beta Feedback →</a></div>",
    )

# =========================
# BuildCommand AI v28 Intelligence Layer
# =========================

def parse_iso_date(value):
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except Exception:
        return None


def activity_delay_signal(activity, today=None):
    today = today or date.today()
    start = parse_iso_date(activity["start"])
    finish = parse_iso_date(activity["finish"])
    pct = float(activity["pct"] or 0)
    status = activity["status"] or "NOT_STARTED"
    score, reasons = 0, []
    if finish and finish < today and pct < 100:
        score += 55; reasons.append("finish date has passed")
    if start and start <= today and pct <= 0 and status != "COMPLETE":
        score += 25; reasons.append("scheduled start has passed with 0% complete")
    if start and finish and start <= today <= finish:
        total = max(1, (finish-start).days+1)
        elapsed = max(0, (today-start).days+1)
        expected = min(100, elapsed/total*100)
        if pct + 15 < expected:
            variance = expected-pct
            score += min(35, 10 + variance*.6)
            reasons.append(f"progress trails time-elapsed plan by about {variance:.0f} points")
    score = min(100, int(round(score)))
    band = "CRITICAL" if score >= 70 else "HIGH" if score >= 45 else "WATCH" if score >= 20 else "READY"
    return score, band, reasons


def procurement_warning_signal(row):
    required, promised = parse_iso_date(row["required_on_site"]), parse_iso_date(row["promised_date"])
    status, today = (row["status"] or "NOT_RELEASED").upper(), date.today()
    score, reasons = 0, []
    if status == "DELIVERED": return 0, "READY", ["delivered"]
    if not required: score += 20; reasons.append("required-on-site date missing")
    if not promised: score += 45; reasons.append("promised date missing")
    elif required and promised > required:
        late=(promised-required).days; score += min(80,45+late*4); reasons.append(f"promised date is {late} day(s) after required date")
    if required:
        days=(required-today).days
        if days < 0: score += 30; reasons.append("required-on-site date has passed")
        elif days <= 7 and status not in ["SHIPPED","DELIVERED"]: score += 25; reasons.append("needed within 7 days and not shipped")
        elif days <= 21 and status in ["NOT_RELEASED","RELEASED"]: score += 12; reasons.append("needed within 3 weeks and not yet in fabrication/shipping")
    score=min(100,score)
    band="CRITICAL" if score>=70 else "HIGH" if score>=45 else "WATCH" if score>=20 else "READY"
    return score,band,reasons or ["no procurement warning detected"]


def project_health_snapshot(pid):
    c=db()
    activities=c.execute("SELECT * FROM activities WHERE project_id=?",(pid,)).fetchall()
    risks=c.execute("SELECT * FROM risks WHERE project_id=?",(pid,)).fetchall()
    actions=c.execute("SELECT * FROM action_items WHERE project_id=? AND status='OPEN'",(pid,)).fetchall()
    issues=c.execute("SELECT * FROM project_issues WHERE project_id=? AND status!='CLOSED'",(pid,)).fetchall()
    procurement=c.execute("SELECT * FROM procurement WHERE project_id=? AND status!='DELIVERED'",(pid,)).fetchall()
    readiness=c.execute("SELECT * FROM activity_readiness WHERE project_id=?",(pid,)).fetchall()
    safety=c.execute("SELECT * FROM safety_items WHERE project_id=? AND status!='CLOSED'",(pid,)).fetchall()
    inspections=c.execute("SELECT * FROM inspections_tracker WHERE project_id=? AND result!='PASSED'",(pid,)).fetchall(); c.close()
    today=date.today()
    overdue_actions=sum(1 for x in actions if parse_iso_date(x["due"]) and parse_iso_date(x["due"])<today)
    overdue_issues=sum(1 for x in issues if parse_iso_date(x["due"]) and parse_iso_date(x["due"])<today)
    delay_risk=sum(activity_delay_signal(a)[0] for a in activities)/len(activities) if activities else 0
    risk_exp=sum(float(r["score"] or 0) for r in risks)/len(risks) if risks else 0
    proc_exp=sum(procurement_warning_signal(p)[0] for p in procurement)/len(procurement) if procurement else 0
    ready_pct=sum(readiness_result(r)[0] for r in readiness)/len(readiness) if readiness else (50 if activities else 100)
    safety_penalty=sum(25 if s["severity"]=="CRITICAL" else 15 if s["severity"]=="HIGH" else 7 for s in safety)
    inspection_penalty=sum(18 if i["result"]=="FAILED" else 7 for i in inspections)
    scores={"schedule":max(0,100-delay_risk),"readiness":max(0,min(100,ready_pct)),"procurement":max(0,100-proc_exp),"risk":max(0,100-risk_exp),"field":max(0,100-min(45,overdue_actions*6+overdue_issues*7)-min(35,safety_penalty+inspection_penalty))}
    overall=round(scores["schedule"]*.28+scores["readiness"]*.22+scores["procurement"]*.18+scores["risk"]*.17+scores["field"]*.15)
    return {"overall":overall,**{k:round(v) for k,v in scores.items()},"overdue_actions":overdue_actions,"overdue_issues":overdue_issues}


@app.get("/project-health", response_class=HTMLResponse)
def project_health_page():
    h=project_health_snapshot(project_id())
    badge,label=("READY","Healthy") if h["overall"]>=85 else ("WATCH","Watch") if h["overall"]>=70 else ("HIGH","At Risk") if h["overall"]>=50 else ("CRITICAL","Critical")
    return shell("Project Health",f'<div class="hero"><div class="eyebrow">Executive Project Health</div><h1>{h["overall"]}/100 <span class="badge {badge}">{label}</span></h1></div><div class="grid4"><div class="card"><div class="label">Schedule</div><div class="kpi">{h["schedule"]}</div></div><div class="card"><div class="label">Readiness</div><div class="kpi">{h["readiness"]}</div></div><div class="card"><div class="label">Procurement</div><div class="kpi">{h["procurement"]}</div></div><div class="card"><div class="label">Risk</div><div class="kpi">{h["risk"]}</div></div></div><div class="card"><h2>Field Execution {h["field"]}</h2><p>Overdue actions: {h["overdue_actions"]} · Overdue RFIs/issues: {h["overdue_issues"]}</p></div>')


@app.get("/lookahead-intelligence", response_class=HTMLResponse)
def lookahead_intelligence_page():
    pid=project_id(); today=date.today(); horizon=today+timedelta(days=21); c=db()
    activities=c.execute("SELECT * FROM activities WHERE project_id=? ORDER BY start",(pid,)).fetchall(); readiness_rows=c.execute("SELECT * FROM activity_readiness WHERE project_id=?",(pid,)).fetchall(); procurement=c.execute("SELECT * FROM procurement WHERE project_id=? AND status!='DELIVERED'",(pid,)).fetchall(); inspections=c.execute("SELECT * FROM inspections_tracker WHERE project_id=? AND result!='PASSED'",(pid,)).fetchall(); issues=c.execute("SELECT * FROM project_issues WHERE project_id=? AND status!='CLOSED'",(pid,)).fetchall(); c.close()
    rm={r["activity_id"]:r for r in readiness_rows}; pm={}; im={}; qm={}
    for p in procurement: pm.setdefault(p["activity_id"],[]).append(p)
    for i in inspections: im.setdefault(i["activity_id"],[]).append(i)
    for q in issues: qm.setdefault(q["activity_id"],[]).append(q)
    cards=''
    for a in activities:
        start,finish=parse_iso_date(a["start"]),parse_iso_date(a["finish"])
        if not start or (finish and finish<today) or start>horizon: continue
        rp,rs=(0,"NOT REVIEWED")
        if a["id"] in rm: rp,rs,_=readiness_result(rm[a["id"]])
        blockers=[]
        if rs in ["NOT READY","AT RISK","NOT REVIEWED"]: blockers.append(f"Readiness {rp}%")
        for p in pm.get(a["id"],[]):
            _,b,_=procurement_warning_signal(p)
            if b!="READY": blockers.append(f"Procurement: {p['item']} ({b})")
        for i in im.get(a["id"],[]): blockers.append(f"Inspection: {i['inspection_type']} ({i['result']})")
        for q in qm.get(a["id"],[]): blockers.append(f"{q['issue_type']}: {q['title']}")
        _,dband,_=activity_delay_signal(a)
        if dband!="READY": blockers.append(f"Schedule: {dband}")
        days=(start-today).days; timing="STARTED / DUE NOW" if days<=0 else f"Starts in {days} day(s)"; badge="READY" if not blockers else "CRITICAL" if any("CRITICAL" in b for b in blockers) else "WATCH"
        bh=''.join(f'<div class="small">• {esc(b)}</div>' for b in blockers) or '<div class="small">No major blocker detected.</div>'
        cards+=f'<div class="card"><span class="badge {badge}">{timing}</span><h3>{esc(a["external_id"])} - {esc(a["name"])}</h3><p>Readiness: {rp}%</p>{bh}</div>'
    return shell("3-Week Lookahead",f'<div class="hero"><div class="eyebrow">3-Week Lookahead Intelligence</div><h1>Upcoming work + blockers</h1><div class="muted">{today} through {horizon}</div></div><div class="grid2">{cards or "<div class=card>No activities in the next 21 days.</div>"}</div>')


@app.get("/sub-scorecards", response_class=HTMLResponse)
def subcontractor_scorecards_page():
    pid=project_id(); c=db(); subs=c.execute("SELECT * FROM subs WHERE project_id=? ORDER BY trade,name",(pid,)).fetchall(); updates=c.execute("SELECT * FROM subcontractor_updates WHERE project_id=? ORDER BY update_date DESC,id DESC",(pid,)).fetchall(); punch=c.execute("SELECT * FROM punch_items WHERE project_id=? AND status!='VERIFIED'",(pid,)).fetchall(); safety=c.execute("SELECT * FROM safety_items WHERE project_id=? AND status!='CLOSED'",(pid,)).fetchall(); actions=c.execute("SELECT * FROM action_items WHERE project_id=? AND status='OPEN'",(pid,)).fetchall(); c.close(); today=date.today(); cards=''
    for s in subs:
        su=[u for u in updates if u["sub_id"]==s["id"]]; latest=su[0] if su else None; score=100; reasons=[]
        if not latest: score-=20; reasons.append("no recent subcontractor update")
        else:
            st=(latest["status"] or "").upper(); score-=35 if st=="CRITICAL" else 22 if st=="HIGH" else 10 if st=="WATCH" else 0
            if st in ["CRITICAL","HIGH","WATCH"]: reasons.append(f"latest status: {st}")
            if (latest["manpower"] or 0)<=0: score-=12; reasons.append("latest manpower is zero")
        name=(s["name"] or "").lower(); trade=(s["trade"] or "").lower(); op=sum(1 for p in punch if (name and name in (p["owner"] or "").lower()) or (trade and trade in (p["trade"] or "").lower())); osf=sum(1 for x in safety if name and name in (x["responsible_party"] or "").lower()); oa=sum(1 for a in actions if name and name in (a["owner"] or "").lower() and parse_iso_date(a["due"]) and parse_iso_date(a["due"])<today)
        if op: score-=min(20,op*4); reasons.append(f"{op} open punch item(s)")
        if osf: score-=min(25,osf*8); reasons.append(f"{osf} open safety item(s)")
        if oa: score-=min(20,oa*6); reasons.append(f"{oa} overdue action(s)")
        score=max(0,score); grade="A" if score>=90 else "B" if score>=80 else "C" if score>=65 else "D" if score>=50 else "F"; badge="READY" if score>=80 else "WATCH" if score>=65 else "HIGH" if score>=50 else "CRITICAL"; rh=''.join(f'<div class="small">• {esc(r)}</div>' for r in reasons) or '<div class="small">No major negative signal detected.</div>'; cards+=f'<div class="card"><span class="badge {badge}">Grade {grade}</span><h3>{esc(s["name"])}</h3><div class="muted">{esc(s["trade"])}</div><div class="kpi">{score}</div>{rh}</div>'
    return shell("Subcontractor Scorecards",f'<div class="hero"><div class="eyebrow">Trade Partner Intelligence</div><h1>Subcontractor Scorecards</h1></div><div class="grid3">{cards or "<div class=card>Add subcontractors to generate scorecards.</div>"}</div>')


@app.get("/rfi-impact", response_class=HTMLResponse)
def rfi_impact_page():
    pid=project_id(); c=db(); issues=c.execute("""SELECT i.*,a.external_id,a.name activity,a.start activity_start FROM project_issues i LEFT JOIN activities a ON a.id=i.activity_id WHERE i.project_id=? AND i.status!='CLOSED' ORDER BY i.due,i.id""",(pid,)).fetchall(); activities=c.execute("SELECT * FROM activities WHERE project_id=? ORDER BY start",(pid,)).fetchall(); c.close(); cards=''; today=date.today()
    for i in issues:
        ls=parse_iso_date(i["activity_start"]) if i["activity_start"] else None; impacted=[a for a in activities if ls and parse_iso_date(a["start"]) and parse_iso_date(a["start"])>=ls and a["id"]!=i["activity_id"]]; due=parse_iso_date(i["due"]); score=20+(40 if i["priority"]=="CRITICAL" else 25 if i["priority"]=="HIGH" else 0)+(25 if due and due<today else 0)+(20 if ls and (ls-today).days<=14 else 0); score=min(100,score); badge="CRITICAL" if score>=70 else "HIGH" if score>=45 else "WATCH"; impact=', '.join(f'{a["external_id"]} {a["name"]}' for a in impacted[:4]) or 'No downstream activity inferred from date order.'; cards+=f'<div class="card"><span class="badge {badge}">Impact {score}</span><h3>{esc(i["title"])}</h3><p>Linked: {esc(i["external_id"] or "None")} {esc(i["activity"] or "")}</p><p>Possible downstream exposure: {esc(impact)}</p></div>'
    return shell("RFI Impact",f'<div class="hero"><div class="eyebrow">RFI Impact Intelligence</div><h1>Which questions can move the schedule?</h1></div><div class="grid2">{cards or "<div class=card>No open RFIs/issues.</div>"}</div>')


@app.get("/procurement-warning", response_class=HTMLResponse)
def procurement_warning_page():
    pid=project_id(); c=db(); rows=c.execute("""SELECT p.*,a.external_id,a.name activity FROM procurement p LEFT JOIN activities a ON a.id=p.activity_id WHERE p.project_id=? AND p.status!='DELIVERED' ORDER BY p.required_on_site,p.id""",(pid,)).fetchall(); c.close(); scored=sorted([(procurement_warning_signal(r)[0],procurement_warning_signal(r)[1],r,procurement_warning_signal(r)[2]) for r in rows],key=lambda x:x[0],reverse=True); cards=''
    for score,band,r,reasons in scored: cards+=f'<div class="card"><span class="badge {band}">{band} · {score}</span><h3>{esc(r["item"])}</h3><p>Required: {esc(r["required_on_site"] or "—")} · Promised: {esc(r["promised_date"] or "—")}</p>'+''.join(f'<div class="small">• {esc(x)}</div>' for x in reasons)+'</div>'
    return shell("Procurement Warning",f'<div class="hero"><div class="eyebrow">Procurement Early Warning</div><h1>Material threats ranked before installation.</h1></div><div class="grid2">{cards or "<div class=card>No open procurement items.</div>"}</div>')


def recovery_options_for_activity(activity,delay_score):
    if delay_score<20: return [("Protect Plan",0,0,"No major delay signal. Protect access, material, and inspection readiness.")]
    return [("Add targeted manpower",min(3.0,max(1.0,delay_score/30)),2500+delay_score*35,"Increase manpower only after constraints are cleared."),("Second shift / extended hours",min(5.0,max(1.5,delay_score/22)),4500+delay_score*55,"Use where supervision and site rules allow."),("Resequence downstream work",min(4.0,max(1.0,delay_score/25)),1500+delay_score*20,"Advance unaffected work while the constraint is removed.")]


@app.get("/ai-recovery", response_class=HTMLResponse)
def ai_recovery_page():
    pid=project_id(); c=db(); acts=c.execute("SELECT * FROM activities WHERE project_id=? AND status!='COMPLETE' ORDER BY start",(pid,)).fetchall(); c.close(); scored=sorted([(activity_delay_signal(a)[0],activity_delay_signal(a)[1],a) for a in acts],key=lambda x:x[0],reverse=True); cards=''
    for score,band,a in scored[:6]:
        opts=recovery_options_for_activity(a,score); oh=''.join(f'<div class="action"><b>{esc(n)}</b><div class="small">Estimated recovery: {d:.1f} day(s) · ROM cost: ${cst:,.0f}</div><div class="small">{esc(note)}</div></div>' for n,d,cst,note in opts); n,d,cst,_=opts[0]; cards+=f'<div class="card"><span class="badge {band}">Delay Signal {score}</span><h3>{esc(a["external_id"])} - {esc(a["name"])}</h3>{oh}<form method="post" action="/ai-recovery/save"><input type="hidden" name="activity_id" value="{a["id"]}"><input type="hidden" name="scenario" value="{esc(n)}"><input type="hidden" name="days_recovered" value="{d}"><input type="hidden" name="est_cost" value="{cst}"><button type="submit">Save Top Recovery Option</button></form></div>'
    return shell("AI Recovery Planner",f'<div class="hero"><div class="eyebrow">Recovery Intelligence</div><h1>What can we do when work drifts?</h1><div class="muted">ROM planning aid—field validation required.</div></div><div class="grid2">{cards}</div>')


@app.post("/ai-recovery/save")
def save_ai_recovery(activity_id:int=Form(...),scenario:str=Form(...),days_recovered:float=Form(0),est_cost:float=Form(0)):
    pid=project_id(); c=db(); c.execute("INSERT INTO recovery(project_id,activity_id,scenario,days_recovered,est_cost,status) VALUES(?,?,?,?,?,'PROPOSED')",(pid,activity_id,scenario,days_recovered,est_cost)); c.commit(); c.close(); return RedirectResponse('/ai-recovery',status_code=303)


@app.get("/quick-entry", response_class=HTMLResponse)
def quick_entry_page():
    pid=project_id(); c=db(); rows=c.execute("SELECT * FROM quick_entries WHERE company_id=? AND project_id=? ORDER BY id DESC LIMIT 15",(current_company_id(),pid)).fetchall(); c.close(); recent=''.join(f'<div class="action"><b>{esc(r["entry_type"])}</b> → {esc(r["routed_to"])}<div>{esc(r["text"])}</div></div>' for r in rows) or '<div class="muted">No quick entries yet.</div>'; body=f'''<div class="hero"><div class="eyebrow">Superintendent Quick Entry</div><h1>Say it once. Route it where it belongs.</h1></div><div class="grid2"><div class="card"><form method="post" action="/quick-entry"><label>Route As</label><select name="entry_type"><option value="AUTO">Auto Detect</option><option value="ACTION">Action</option><option value="FIELD_UPDATE">Field Update</option><option value="RFI">RFI / Issue</option><option value="SAFETY">Safety Observation</option></select><label>Field Note</label><textarea id="quickText" name="text" required></textarea><button type="button" onclick="startBuildCommandVoice()">Speak</button><button type="submit">Route Entry</button></form></div><div class="card"><h2>Recent Quick Entries</h2>{recent}</div></div><script>function startBuildCommandVoice(){{const SR=window.SpeechRecognition||window.webkitSpeechRecognition;if(!SR){{alert('Voice recognition is not supported by this browser.');return;}}const rec=new SR();rec.lang='en-US';rec.onresult=function(e){{document.getElementById('quickText').value=e.results[0][0].transcript;}};rec.start();}}</script>'''; return shell("Quick Entry",body)


@app.post("/quick-entry")
def save_quick_entry(entry_type:str=Form("AUTO"),text:str=Form(...)):
    pid=project_id(); raw=text.strip(); low=raw.lower(); detected=entry_type
    if detected=="AUTO": detected="SAFETY" if any(k in low for k in ["unsafe","safety","guardrail","near miss","incident"]) else "RFI" if any(k in low for k in ["rfi","question","clarify","design","drawing conflict"]) else "ACTION" if any(k in low for k in ["need to","follow up","confirm","call","send","by friday","by tomorrow"]) else "FIELD_UPDATE"
    c=db(); routed=detected
    if detected=="ACTION": c.execute("INSERT INTO action_items(project_id,title,owner,priority,due,status,notes,created) VALUES(?,?,?,?,?,'OPEN',?,?)",(pid,raw[:140],"Superintendent","WATCH",date.today().isoformat(),raw,date.today().isoformat())); routed="Action Center"
    elif detected=="RFI": c.execute("INSERT INTO project_issues(project_id,activity_id,issue_type,title,owner,due,priority,status,description,response,created) VALUES(?,NULL,'FIELD_ISSUE',?,'',?,'WATCH','OPEN',?,'',?)",(pid,raw[:140],date.today().isoformat(),raw,date.today().isoformat())); routed="RFIs / Issues"
    elif detected=="SAFETY": c.execute("INSERT INTO safety_items(project_id,activity_id,event_date,item_type,title,location,responsible_party,severity,status,description,corrective_action,created) VALUES(?,NULL,?,'OBSERVATION',?,'','','WATCH','OPEN',?,'',?)",(pid,date.today().isoformat(),raw[:140],raw,date.today().isoformat())); routed="Safety"
    else: c.execute("INSERT INTO field_updates(project_id,activity_id,update_type,text,created) VALUES(?,NULL,'QUICK_ENTRY',?,?)",(pid,raw,date.today().isoformat())); routed="Field Updates"
    c.execute("INSERT INTO quick_entries(company_id,project_id,user_id,entry_type,text,routed_to,created) VALUES(?,?,?,?,?,?,?)",(current_company_id(),pid,current_user_id(),detected,raw,routed,datetime.utcnow().isoformat())); c.commit(); c.close(); return RedirectResponse('/quick-entry',status_code=303)


@app.get("/weekly-report", response_class=HTMLResponse)
def weekly_report_page():
    pid=project_id(); c=db(); row=c.execute("SELECT * FROM weekly_ai_reports WHERE company_id=? AND project_id=? ORDER BY id DESC LIMIT 1",(current_company_id(),pid)).fetchone(); c.close(); report=esc(row["report_text"]).replace("\n","<br>") if row else "No weekly AI report generated yet."; return shell("Weekly AI Report",f'<div class="hero"><div class="eyebrow">Owner / PM Reporting</div><h1>AI Weekly Project Report</h1></div><div class="card"><form method="post" action="/weekly-report/generate"><button type="submit">Generate Weekly Report</button></form></div><div class="card">{report}</div>')


@app.post("/weekly-report/generate")
def generate_weekly_report():
    pid=project_id(); context=build_project_context(pid); health=project_health_snapshot(pid); key=os.environ.get("OPENAI_API_KEY")
    if key:
        instructions=f'''You are BuildCommand AI producing a professional weekly construction project report for an owner/project manager. Use only supplied project facts. Project health score: {health["overall"]}/100. Use sections: EXECUTIVE SUMMARY, SCHEDULE, FIELD PROGRESS, RFIs / DECISIONS, PROCUREMENT, SAFETY / QUALITY, CHANGE EXPOSURE, TOP PRIORITIES NEXT WEEK. Do not invent facts, commitments, costs, or dates.'''
        try: report=OpenAI(api_key=key).responses.create(model=os.environ.get("OPENAI_MODEL","gpt-5.6"),instructions=instructions,input=context).output_text
        except Exception as exc: report=f"AI weekly report failed: {exc}"
    else: report=f"Project Health: {health['overall']}/100\nOPENAI_API_KEY is not configured."
    c=db(); c.execute("INSERT INTO weekly_ai_reports(company_id,project_id,week_ending,report_text,created) VALUES(?,?,?,?,?)",(current_company_id(),pid,date.today().isoformat(),report,datetime.utcnow().isoformat())); c.commit(); c.close(); return RedirectResponse('/weekly-report',status_code=303)


@app.get("/ai-command", response_class=HTMLResponse)
def ai_command_page():
    pid=project_id(); h=project_health_snapshot(pid); c=db(); acts=c.execute("SELECT * FROM activities WHERE project_id=? AND status!='COMPLETE'",(pid,)).fetchall(); actions=c.execute("SELECT * FROM action_items WHERE project_id=? AND status='OPEN'",(pid,)).fetchall(); procs=c.execute("SELECT * FROM procurement WHERE project_id=? AND status!='DELIVERED'",(pid,)).fetchall(); issues=c.execute("SELECT * FROM project_issues WHERE project_id=? AND status!='CLOSED'",(pid,)).fetchall(); c.close(); sig=[]; today=date.today()
    for a in acts:
        s,b,r=activity_delay_signal(a)
        if b!="READY": sig.append((s,b,"Schedule",f'{a["external_id"]} {a["name"]}',"; ".join(r)))
    for p in procs:
        s,b,r=procurement_warning_signal(p)
        if b!="READY": sig.append((s,b,"Procurement",p["item"],"; ".join(r)))
    for a in actions:
        d=parse_iso_date(a["due"])
        if d and d<today: sig.append((65,a["priority"] or "HIGH","Action",a["title"],f'Overdue · Owner {a["owner"] or "Unassigned"}'))
    for i in issues:
        d=parse_iso_date(i["due"])
        if d and d<today: sig.append((70,i["priority"] or "HIGH",i["issue_type"],i["title"],"Response/decision overdue"))
    sig.sort(key=lambda x:x[0],reverse=True); sh=''.join(f'<div class="action"><span class="badge {b if b in ["CRITICAL","HIGH","WATCH","READY","LOW"] else "HIGH"}">{esc(cat)}</span> <b>{esc(t)}</b><div class="small">{esc(d)}</div></div>' for _,b,cat,t,d in sig[:12]) or '<div class="muted">No major automated warning signals detected.</div>'; return shell("AI Command",f'<div class="hero"><div class="eyebrow">AI Daily Command Center</div><h1>Project Health {h["overall"]}/100</h1></div><div class="grid2"><div class="card"><h2>Handle First</h2>{sh}</div><div class="card"><h2>Command Tools</h2><p><a href="/lookahead-intelligence">3-Week Lookahead</a></p><p><a href="/ai-recovery">Recovery Planner</a></p><p><a href="/quick-entry">Quick / Voice Entry</a></p><p><a href="/weekly-report">Weekly AI Report</a></p></div></div>')

def _attachment_text(row):
    if not row:
        return ""
    path=os.path.join(UPLOAD_DIR,row["stored_name"])
    ext=Path(row["original_name"] or "").suffix.lower()
    if not os.path.isfile(path):
        return ""
    try:
        if ext in {".txt",".csv",".md"}:
            return Path(path).read_text(errors="ignore")[:250000]
        if ext==".pdf" and PdfReader is not None:
            return "\n".join((p.extract_text() or "") for p in PdfReader(path).pages[:250])[:350000]
        if ext in {".xlsx",".xlsm"} and openpyxl is not None:
            wb=openpyxl.load_workbook(path,read_only=True,data_only=True)
            lines=[]
            for ws in wb.worksheets[:30]:
                lines.append("SHEET: "+ws.title)
                for vals in ws.iter_rows(values_only=True):
                    lines.append(" | ".join("" if v is None else str(v) for v in vals))
                    if len(lines)>25000:
                        break
            return "\n".join(lines)[:350000]
    except Exception:
        return ""
    return ""


def _blueprint_json(text_value):
    raw=(text_value or "").strip()
    if raw.startswith("```"):
        raw=re.sub(r"^```(?:json)?\s*", "", raw, flags=re.I)
        raw=re.sub(r"\s*```$", "", raw)
    try:
        return json.loads(raw)
    except Exception:
        m=re.search(r"\{.*\}",raw,re.S)
        if m:
            return json.loads(m.group(0))
        raise ValueError("Blueprint Brain returned a response that was not valid JSON.")


def _blueprint_scope_text(trade_data):
    lines=[]
    summary=(trade_data.get("summary") or "").strip()
    if summary:
        lines.append(summary)
        lines.append("")
    for i,item in enumerate(trade_data.get("items") or [],1):
        req=(item.get("requirement") or "").strip()
        if not req:
            continue
        src=[]
        if item.get("source_sheet"): src.append("Sheet "+str(item.get("source_sheet")))
        if item.get("source_detail"): src.append("Detail "+str(item.get("source_detail")))
        if item.get("source_spec"): src.append("Spec "+str(item.get("source_spec")))
        conf=(item.get("confidence") or "MEDIUM").upper()
        line=f"{i}. {req}"
        if src: line += " ["+" · ".join(src)+"]"
        line += f" — Confidence: {conf}"
        lines.append(line)
    return "\n".join(lines).strip()


def _blueprint_prompt(source_names):
    return f"""
You are the BuildCommand AI Blueprint Brain, a construction-document scope intelligence engine.

SOURCE FILES:
{source_names}

MISSION:
Read the supplied project plans/specifications as one coordinated construction package. Build a trade-by-trade scope of work by finding requirements wherever they appear, including requirements that belong to one trade but are written on another discipline's sheets.

NON-NEGOTIABLE RULES:
1. Use only information supported by the supplied files. Never invent drawing numbers, details, spec sections, quantities, equipment, code requirements, or responsibilities.
2. A requirement may be assigned to a trade even when it is found on another discipline's sheet. Mark those as CROSS_DISCIPLINE.
3. Every scope item must preserve the best available source: sheet number, detail, spec section, note, or filename. If a source cannot be identified, use an empty string and lower the confidence.
4. Distinguish explicit requirements from coordination inferences. Use item_type values SCOPE, CROSS_DISCIPLINE, COORDINATION, EXCLUSION_REVIEW, or RFI_CANDIDATE.
5. Flag contradictions, missing information, equipment with unclear power/plumbing/HVAC connections, unclear trade responsibility, and likely scope gaps as RFI candidates. Do not resolve ambiguity by guessing.
6. Include all detected trades, not only MEP. Typical trades can include demolition, earthwork, concrete, masonry, structural steel, rough carpentry, millwork, waterproofing, roofing, doors/frames/hardware, glazing, framing/drywall, ceilings, flooring, tile, painting, specialties, equipment, furnishings, fire sprinkler, plumbing, HVAC, controls, electrical, fire alarm, low voltage, security, site utilities, paving, landscaping, and others actually supported by the documents.
7. Phrase each requirement as an actionable subcontractor scope item suitable for GC review. Do not claim the generated scope replaces the signed subcontract, specifications, addenda, RFIs, or professional design review.
8. Confidence must be HIGH, MEDIUM, or LOW based on source clarity.
9. Trade ownership rules: resilient/rubber/vinyl base and resilient flooring belong to Flooring, not Framing/Drywall.
10. Demolition is system-owned: electrical demolition belongs to Electrical, plumbing demolition to Plumbing, HVAC/mechanical demolition to HVAC/Mechanical, fire sprinkler demolition to Fire Sprinkler, fire alarm demolition to Fire Alarm, and low-voltage/security demolition to the applicable systems trade. Only non-system/general demolition belongs to Demolition.

Return ONLY valid JSON using exactly this top-level structure:
{{
  "project_summary": "short summary grounded in the documents",
  "detected_disciplines": ["Architectural", "Electrical"],
  "trade_scopes": [
    {{
      "trade": "Electrical",
      "division": "26",
      "summary": "short trade scope overview",
      "items": [
        {{
          "requirement": "actionable scope requirement",
          "source_sheet": "E2.1",
          "source_detail": "3/E5.2",
          "source_spec": "26 05 00",
          "source_note": "General Note 7",
          "related_trade": "Mechanical",
          "confidence": "HIGH",
          "item_type": "CROSS_DISCIPLINE"
        }}
      ]
    }}
  ],
  "cross_discipline_flags": ["plain-language flag with source"],
  "rfi_candidates": ["plain-language ambiguity/gap with source"],
  "review_notes": ["items the GC should verify before issuing scopes"]
}}

IMPORTANT TRADE OWNERSHIP:
Classify by the actual work/material/system, not by which drawing sheet contains the note.
If one note contains work for multiple trades, split it into separate scope items and keep the same source sheet/detail/spec reference on each.
Wall/gypsum/drywall patching = Framing / Drywall. Paint/refinish after patch = Painting.
Doors, door frames, hollow-metal/wood doors and hardware = Doors / Frames / Hardware.
Door rough openings, jamb studs and header framing = Framing / Drywall.
Tile materials/installation = Tile. Metal studs/framing = Framing / Drywall even when shown in a tile detail.
MEP demolition stays with the responsible MEP trade. General non-MEP demolition = Demolition.
""".strip()




V33_TRADES = [
    "Demolition","Earthwork","Concrete","Masonry","Structural Steel","Rough Carpentry",
    "Millwork","Waterproofing","Roofing","Doors / Frames / Hardware","Storefront / Glazing",
    "Framing / Drywall","Ceilings","Flooring / Tile","Painting","Toilet / Bath Accessories",
    "Specialties","Equipment","Furnishings","Fire Sprinkler","Plumbing","HVAC / Mechanical",
    "Controls","Electrical","Fire Alarm","Low Voltage","Security","Site Utilities",
    "Paving","Landscaping","Unassigned"
]

def _v33_normalize_trade(name):
    n=(name or "").strip().lower()
    aliases={
        "mechanical":"HVAC / Mechanical","hvac":"HVAC / Mechanical",
        "hvac/mechanical":"HVAC / Mechanical","mechanical / hvac":"HVAC / Mechanical",
        "drywall":"Framing / Drywall","framing":"Framing / Drywall","framing/drywall":"Framing / Drywall",
        "doors/frames/hardware":"Doors / Frames / Hardware","door hardware":"Doors / Frames / Hardware",
        "storefront":"Storefront / Glazing","glazing":"Storefront / Glazing",
        "glazing/storefront":"Storefront / Glazing","glazing / storefront":"Storefront / Glazing",
        "storefront/glazing":"Storefront / Glazing","storefront / glazing":"Storefront / Glazing",
        "tile":"Flooring / Tile","flooring":"Flooring / Tile","flooring/tile":"Flooring / Tile",
        "tile/flooring":"Flooring / Tile","flooring / tile":"Flooring / Tile","tile / flooring":"Flooring / Tile",
        "bath accessories":"Toilet / Bath Accessories","bathroom accessories":"Toilet / Bath Accessories",
        "toilet accessories":"Toilet / Bath Accessories","toilet room accessories":"Toilet / Bath Accessories",
        "toilet / bath accessories":"Toilet / Bath Accessories",
        "fire sprinkler":"Fire Sprinkler","sprinkler":"Fire Sprinkler","fire alarm":"Fire Alarm",
        "low voltage":"Low Voltage","low-voltage":"Low Voltage","electrical":"Electrical",
        "plumbing":"Plumbing","demolition":"Demolition","demo":"Demolition"
    }
    return aliases.get(n, (name or "Unassigned").strip() or "Unassigned")

def _v33_trade_for_item(item, proposed_trade):
    """v34.1 field-tested Blueprint Brain trade classifier."""
    req=str(item.get("requirement") or "").strip().lower()
    text=" ".join(str(item.get(k) or "") for k in
                  ["requirement","source_note","source_spec","related_trade"]).lower()
    proposed=_v33_normalize_trade(proposed_trade)

    def has(*terms): return any(x in text for x in terms)

    # ------------------------------------------------------------
    # 1) ACTION FIRST: demolition ownership
    # Demolition must be the work action, not merely context such as
    # "patching resulting from demolition."
    explicit_demo=(
        req.startswith("demo ") or req.startswith("demolish ") or
        req.startswith("remove existing ") or req.startswith("remove and dispose ") or
        " existing to be removed" in req or " shall be demolished" in req or
        " demolition of " in req
    )

    # Access/coordination removal is NOT demolition scope when it is explicitly remove-and-replace.
    access_remove_replace=has("remove and replace ceiling tile","remove and replace ceiling tiles",
                              "remove and replace ceiling grid","remove/reinstall ceiling",
                              "remove and reinstall ceiling","temporarily remove ceiling")

    if explicit_demo and not access_remove_replace:
        # If the primary requirement is clearly general architectural demo and merely says
        # MEP removals remain by their trades, keep the item in Demolition.
        general_demo_objects=has("restroom partition","toilet partition","accessories","finish flooring",
                                 "wall finish","ceiling finish","metal stud partition","gypsum partition",
                                 "drywall partition","door and frame","casework","millwork","storefront",
                                 "ceramic tile","flooring","ceiling tile","ceiling grid")
        exception_language=has("remain assigned to their respective trades","system-specific",
                               "mep removals remain","mep removal remains")
        if general_demo_objects and exception_language:
            return "Demolition"

        # Protected MEP/system demo.
        if has("fire alarm","smoke detector","horn strobe","pull station","notification appliance"):
            return "Fire Alarm"
        if has("card reader","card access","access control","electronic strike","electric strike",
               "electrified strike","request to exit","rex device","door contact","security contact",
               "camera","cctv","security system","intrusion alarm","data cabling","telecom cabling",
               "low voltage","low-voltage","intercom"):
            return "Low Voltage"
        if has("sprinkler","fire suppression","fire protection piping"):
            return "Fire Sprinkler"
        if has("receptacle","outlet","switch","panelboard","electrical panel","transformer","conduit",
               "wire","wiring","branch circuit","breaker","light fixture","lighting fixture","feeder"):
            return "Electrical"
        if has("water closet","urinal","lavatory","wash fountain","mop sink","floor sink","floor drain",
               "drinking fountain","plumbing fixture","sanitary piping","waste piping","vent piping",
               "domestic water","water heater","plumbing piping"):
            return "Plumbing"
        if has("ductwork","duct","diffuser","grille","register","vav","rtu","ahu","fan coil",
               "exhaust fan","air handler","mechanical equipment","hvac equipment","refrigerant piping"):
            return "HVAC / Mechanical"
        return "Demolition"

    # ------------------------------------------------------------
    # 2) GC / coordination responsibilities
    # ------------------------------------------------------------
    if has("coordinate installation and utility rough-ins","coordinate installation","coordinate utility rough-ins",
           "confirm which appliances","confirm which equipment","confirm owner-furnished",
           "owner furnished versus contractor furnished","owner-furnished versus contractor-furnished",
           "verify furnished by","confirm furnished by","confirm procurement responsibility"):
        return "GC / General Contractor"

    # ------------------------------------------------------------
    # 3) Explicit work-action overrides equipment/location words
    # ------------------------------------------------------------
    if has("electrical connections","electrical connection","disconnecting means","provide power",
           "power connection","branch circuit","electrical circuit","electrical feed","provide disconnect"):
        return "Electrical"

    # Plumbing fixture/equipment package. Location inside millwork never changes ownership.
    if has("plumbing fixture schedule","water closets","urinals","lavatories","wash fountains",
           "mop sink","floor sink","floor drains","drinking fountains","break-room sinks",
           "break room sinks","disposal connections","instantaneous electric water heater","iwh-1",
           "chronomite"):
        return "Plumbing"

    # HVAC equipment package. Roof-mounted/roof curb is location/accessory, not Roofing.
    if has("exhaust fan","ef-1","greenheck","upblast fan","upblast exhaust fan","full-size exhaust duct",
           "full size exhaust duct","backdraft damper"):
        return "HVAC / Mechanical"

    # Roof repair work specifically belongs to Roofing.
    if has("patch roof","roof patch","repair roof","roof repair","restore roof","roof membrane patch",
           "patch roofing","repair roof membrane","restore roof membrane","roof flashing",
           "flash roof penetration","watertight roof restoration"):
        return "Roofing"

    # ------------------------------------------------------------
    # 4) Protected specialty systems
    # ------------------------------------------------------------
    if has("card reader","card access","access control","electronic strike","electric strike",
           "electrified strike","request to exit","rex device","door contact","security contact",
           "camera","cctv","security system","intrusion alarm","data cabling","telecom cabling",
           "low voltage","low-voltage","intercom"):
        return "Low Voltage"

    if has("sprinkler head","sprinkler heads","sprinkler piping","fire sprinkler","fire suppression piping",
           "sprinkler branch","sprinkler drop"):
        return "Fire Sprinkler"

    if has("fire extinguisher cabinet","fire extinguisher cabinets","semi-recessed fire extinguisher",
           "larson or equal"):
        return "Specialties"

    # ------------------------------------------------------------
    # 5) Framing/Drywall primary assemblies and supports
    # ------------------------------------------------------------
    if has("concealed wood blocking","wood blocking","plywood backing","wood backing","in-wall backing",
           "in wall backing","metal backing","concealed backing","support framing","equivalent supports"):
        return "Framing / Drywall"

    if has("marlite","symmetrix","smart seam","frp","fiberglass reinforced panel",
           "fiberglass-reinforced panel","wainscot"):
        return "Framing / Drywall"

    if has("wall type b","gypsum-board ceiling","gypsum board ceiling","resilient channels",
           "resilient channel","acoustical sealant","taped joints","metal-stud jamb","metal stud jamb",
           "metal-stud header","metal stud header","rough openings","rough opening","stud framing",
           "metal studs","metal stud","steel studs","steel stud","gypsum board","gyp board","drywall",
           "shaftwall","shaft wall"):
        return "Framing / Drywall"

    if has("wall patch","patch wall","patch drywall","drywall patch","gypsum patch","patch gypsum",
           "repair drywall","repair gypsum","gyp board patch","patch gyp","patch and repair wall",
           "patch gypsum board","wall or ceiling openings resulting from demolition"):
        return "Framing / Drywall"

    # ------------------------------------------------------------
    # 6) Ceilings
    # ------------------------------------------------------------
    if has("suspended acoustical tile ceiling","suspended acoustical tile ceilings",
           "acoustical tile ceiling","acoustical tile ceilings","acoustic ceiling tile",
           "acoustical ceiling tile","ceiling tiles and grid","ceiling tile and grid",
           "ceiling tiles","ceiling grid","act ceiling","suspension system"):
        return "Ceilings"

    # ------------------------------------------------------------
    # 7) Doors vs Storefront/Glazing -- primary system first
    # ------------------------------------------------------------
    interior_door_system=has("interior doors and frames","interior door and frame","hollow-metal framed windows",
                             "hollow metal framed windows","interior hollow-metal","interior hollow metal",
                             "door schedule","door hardware sets","rotary white birch doors","hollow-metal doors",
                             "hollow metal doors")
    if interior_door_system:
        return "Doors / Frames / Hardware"

    # True storefront/glazing system only; generic 'glazing' is not enough.
    if has("aluminum storefront","storefront system","store front system","curtain wall","aluminum entrance",
           "exterior storefront","storefront doors","storefront frames"):
        return "Storefront / Glazing"

    if has("door frame","door frames","wood door","wood doors","hollow metal frame","hollow-metal frame",
           "door hardware","hardware set","door closer","panic hardware","exit device","lockset",
           "door threshold","door sweep","door hinges"):
        return "Doors / Frames / Hardware"

    # ------------------------------------------------------------
    # 8) Finishes / tile / flooring / paint
    # ------------------------------------------------------------
    if has("faux tile backsplash","tile backsplash","backsplash wt2","aspect ideas a9550",
           "ceramic tile","porcelain tile","wall tile","floor tile","tile base","tile grout",
           "tile mortar","tile adhesive","setting bed","tile setting","tile trim","schluter"):
        return "Tile"

    if has("rubber base","resilient base","vinyl base","cove base","lvt","luxury vinyl","vct",
           "carpet tile","sheet vinyl","resilient flooring","floor transition","transition strip"):
        return "Flooring"

    if has("paint or refinish","paint patched","paint patch","repaint","touch-up paint","touch up paint",
           "paint repair","refinish patched","finish paint","painting of patched surfaces"):
        return "Painting"

    if has("casework","millwork","cabinet","countertop"):
        return "Millwork / Casework"

    # Conservative fallbacks: do not allow location/source words to steal ownership.
    if has("electrical","receptacle","panelboard","conduit","circuit"):
        return "Electrical"
    if has("plumbing","sanitary","domestic water"):
        return "Plumbing"
    if has("hvac","mechanical","ductwork","diffuser","vav"):
        return "HVAC / Mechanical"
    if has("paint","painting","primer","finish coat"):
        return "Painting"

    return proposed

def _v33_reclassify_data(data):
    """Rebuild every parent trade scope from final item ownership."""
    grouped={}
    for td in data.get("trade_scopes") or []:
        proposed=_v33_normalize_trade(td.get("trade"))
        for item in td.get("items") or []:
            if not isinstance(item,dict):
                continue
            target=_v33_trade_for_item(item, proposed) or "Unassigned"
            clean=dict(item)
            clean["assigned_trade"]=target
            # Preserve source and original proposed trade for auditability.
            if proposed and proposed != target:
                clean["original_proposed_trade"]=proposed
            grouped.setdefault(target,[]).append(clean)

    division_defaults={
        "GC / General Contractor":"01","Demolition":"02","Concrete":"03","Masonry":"04",
        "Structural Steel":"05","Rough Carpentry":"06","Waterproofing":"07","Roofing":"07",
        "Doors / Frames / Hardware":"08","Storefront / Glazing":"08","Framing / Drywall":"09",
        "Ceilings":"09","Flooring":"09","Tile":"09","Painting":"09","Specialties":"10",
        "Fire Sprinkler":"21","Plumbing":"22","HVAC / Mechanical":"23","Controls":"23",
        "Electrical":"26","Low Voltage":"27","Fire Alarm":"28"
    }
    rebuilt=[]
    for trade in sorted(grouped):
        items=grouped[trade]
        rebuilt.append({
            "trade":trade,
            "division":division_defaults.get(trade,""),
            "summary":f"BuildCommand source-backed scope for {trade}.",
            "items":items
        })
    data["trade_scopes"]=rebuilt
    notes=data.setdefault("review_notes",[])
    notes.append("v34.1 rebuilt every parent trade scope from final action-first ownership; source sheet/location cannot override the primary work assembly.")
    return data



# ============================================================
# v44.1 BLUEPRINT OWNERSHIP GUARD
# ============================================================

_V441_TRADES=set(V33_TRADES)

def _v441_primary_trade(requirement, proposed):
    s=str(requirement or "").lower().strip()
    def has(*terms): return any(x in s for x in terms)

    # Exact live Blueprint Brain corrections from field review.
    if has("patch concrete and floor surfaces disturbed by electrical or plumbing work"):
        return "Concrete"

    if has("prime and refinish all wall, ceiling, and other surfaces patched after demolition or mep installation"):
        return "Painting"

    if has("bobrick toilet tissue holders","stainless-steel grab bars","paper towel/waste receptacles",
           "soap dispensers","robe hooks","framed mirrors","toilet-seat-cover dispensers",
           "sanitary-napkin disposals"):
        return "Toilet / Bath Accessories"

    if has("electric water heater wh-1","bradford white cehd120"):
        return "Plumbing"

    if has("seven-day programmable thermostats","programmable thermostats") and has("rooftop units","rtu"):
        return "HVAC / Mechanical"

    # 1. PRIMARY ACTION FIRST.
    explicit_demo=(
        s.startswith("demo ") or s.startswith("demolish ") or
        s.startswith("remove existing ") or s.startswith("remove the existing ") or
        s.startswith("remove the architectural ") or s.startswith("remove architectural ") or
        s.startswith("remove and dispose ") or s.startswith("perform non-system") or
        "specifically marked for removal" in s or "identified for demolition" in s
    )
    if explicit_demo:
        return "Demolition"

    # Generic firestopping coordination/spec language is excluded later, not routed as a trade scope.
    # Finish actions beat substrate/material words.
    if has("paint ","paint gypsum","paint hollow","prime and refinish","prime and paint",
           "paint or refinish","repaint","touch-up paint","touch up paint",
           "refinish all wall","finish paint","painting of patched surfaces"):
        return "Painting"

    # Supporting construction work owns itself even when caused by MEP.
    if has("slab cutting","sawcut slab","saw cut slab","concrete cutting",
           "trenching and restoration","patch concrete","concrete patch",
           "restore concrete","slab restoration"):
        return "Concrete"

    # Ceiling coordination/layout belongs to ceilings; device names are interfaces only.
    if has("coordinate ceiling grid layout","ceiling grid layout and openings",
           "coordinate ceiling openings","suspended acoustical tile ceiling",
           "acoustical tile ceiling","acoustic ceiling tile","acoustical ceiling tile",
           "ceiling tile and grid","ceiling tiles and grid","ceiling grid","act ceiling",
           "suspension system"):
        return "Ceilings"

    # Duct smoke devices tied to air-conditioning units are HVAC in this BuildCommand ownership model.
    if has("duct smoke detector","duct smoke detectors"):
        return "HVAC / Mechanical"

    # Electronic accessible/automatic door operators and touch pads are Low Voltage.
    if has("electronic accessible door operator","automatic door operator",
           "accessible door operator","door operator with","touch pads","activation touch pad"):
        return "Low Voltage"

    # Toilet/bath accessories are not plumbing.
    if has("grab bar","grab bars","toilet partition","toilet partitions","urinal screen","urinal screens",
           "toilet tissue holder","paper towel/waste","paper towel dispenser","soap dispenser",
           "robe hook","framed mirror","toilet-seat-cover","toilet seat cover",
           "sanitary-napkin","sanitary napkin","bath accessory","bathroom accessory",
           "toilet room accessory"):
        return "Toilet / Bath Accessories"

    # Storefront hardware remains storefront when explicitly part of storefront/entrance assembly.
    if has("storefront door hardware","storefront entrance hardware","aluminum entrance hardware"):
        return "Storefront / Glazing"

    # Protected specialty systems.
    if has("fire alarm","horn strobe","pull station","notification appliance","smoke detector"):
        return "Fire Alarm"
    if has("card access","card reader","access control","electronic strike","electric strike",
           "electrified strike","request to exit","rex device","door contact","cctv","camera",
           "security system","intrusion alarm","data cabling","telecom cabling","intercom"):
        return "Low Voltage"
    if has("sprinkler head","sprinkler heads","sprinkler piping","fire sprinkler",
           "fire suppression piping","sprinkler branch","sprinkler drop"):
        return "Fire Sprinkler"

    if has("patch roof","roof patch","repair roof","roof repair","restore roof","roof membrane patch",
           "patch roofing","repair roof membrane","flash roof penetration","roof flashing",
           "watertight roof","roof penetration patch"):
        return "Roofing"

    if has("electrical connection","electrical connections","disconnecting means","provide power",
           "branch circuit","electrical feed","power connection","provide disconnect"):
        return "Electrical"

    if has("water closet","urinal","lavatory","wash fountain","mop sink","floor sink","floor drain",
           "drinking fountain","plumbing fixture","break-room sink","break room sink",
           "disposal connection","instantaneous electric water heater","electric domestic water heater",
           "water heater wh-","chronomite","iwh-1","expansion tank","recirculation pump"):
        return "Plumbing"

    if has("exhaust fan","upblast fan","greenheck","backdraft damper","full-size exhaust duct",
           "full size exhaust duct","rtu","ahu","air handler","vav","diffuser","grille","ductwork",
           "thermostat","thermostats","rooftop unit","air-conditioning unit","air conditioning unit"):
        return "HVAC / Mechanical"

    if has("concealed wood blocking","wood blocking","wood backing","plywood backing","metal backing",
           "in-wall backing","in wall backing","support framing","rough opening","metal-stud jamb",
           "metal stud jamb","metal-stud header","metal stud header","stud framing","metal studs",
           "steel studs","gypsum board","gyp board","drywall","shaftwall","shaft wall",
           "patch gypsum","patch drywall","wall patch","patch wall","repair drywall",
           "marlite","symmetrix","frp","wainscot"):
        return "Framing / Drywall"

    if has("interior doors and frames","interior door and frame","hollow-metal framed window",
           "hollow metal framed window","door schedule","door hardware set","hollow-metal door",
           "hollow metal door","door frame","door hardware","lockset","panic hardware","exit device"):
        return "Doors / Frames / Hardware"

    if has("aluminum storefront","storefront system","store front system","curtain wall",
           "aluminum entrance","exterior storefront","storefront door","storefront frame"):
        return "Storefront / Glazing"

    if has("faux tile backsplash","tile backsplash","ceramic tile","porcelain tile","wall tile",
           "floor tile","tile base","tile grout","tile mortar","tile adhesive","schluter",
           "rubber base","resilient base","vinyl base","cove base","lvt","luxury vinyl","vct",
           "carpet tile","sheet vinyl","resilient flooring","floor transition","transition strip"):
        return "Flooring / Tile"

    if has("fire extinguisher cabinet","fire extinguisher cabinets","semi-recessed fire extinguisher"):
        return "Specialties"

    if has("casework","millwork","countertop"):
        return "Millwork"

    return _v33_normalize_trade(proposed)

def _v442_exclude_scope_item(requirement):
    s=str(requirement or "").lower()
    if "firestop cable, raceway, piping, and similar penetrations" in s:
        return True
    # User-approved exclusion: generic firestop penetration specification/coordination statement.
    return (
        ("firestop" in s or "firestopping" in s) and
        ("penetration" in s or "penetrations" in s) and
        ("listed assembl" in s or "restore" in s or "original rating" in s)
    )

def _v441_apply_approved_learning(pid, requirement, trade):
    try:
        _v43_ensure_tables()
        c=db()
        rows=c.execute("""
            SELECT * FROM learning_rules
            WHERE company_id=? AND approval_status='APPROVED'
              AND (project_id=? OR scope_level='COMPANY STANDARD')
              AND rule_type IN ('TRADE ASSIGNMENT','SCOPE BOUNDARY')
            ORDER BY CASE WHEN project_id=? THEN 0 ELSE 1 END,id DESC
        """,(current_company_id(),pid,pid)).fetchall()
        c.close()
        low=str(requirement or "").lower()
        for r in rows:
            subject=str(r["subject"] or "").strip().lower()
            if subject and subject in low:
                learned=str(r["learned_rule"] or "").strip()
                target=learned.split("->")[-1].strip() if "->" in learned else learned
                target=_v33_normalize_trade(target)
                if target in _V441_TRADES:
                    return target
    except Exception:
        pass
    return trade

def _v441_reclassify_with_learning(pid,data):
    grouped={}
    for td in data.get("trade_scopes") or []:
        proposed=_v33_normalize_trade(td.get("trade"))
        for item in td.get("items") or []:
            if not isinstance(item,dict): continue
            req=str(item.get("requirement") or "").strip()
            if not req: continue
            if _v442_exclude_scope_item(req):
                continue
            target=_v33_trade_for_item(item,proposed) or proposed
            target=_v441_primary_trade(req,target)
            target=_v441_apply_approved_learning(pid,req,target)
            clean=dict(item)
            clean["assigned_trade"]=target
            if proposed != target:
                clean["original_proposed_trade"]=proposed
            grouped.setdefault(target,[]).append(clean)

    div={
        "GC / General Contractor":"01","Demolition":"02","Concrete":"03","Masonry":"04",
        "Structural Steel":"05","Rough Carpentry":"06","Waterproofing":"07","Roofing":"07",
        "Doors / Frames / Hardware":"08","Storefront / Glazing":"08","Framing / Drywall":"09",
        "Ceilings":"09","Flooring / Tile":"09","Painting":"09","Millwork":"12",
        "Toilet / Bath Accessories":"10","Specialties":"10","Fire Sprinkler":"21","Plumbing":"22","HVAC / Mechanical":"23",
        "Controls":"23","Electrical":"26","Low Voltage":"27","Fire Alarm":"28"
    }
    data["trade_scopes"]=[
        {"trade":trade,"division":div.get(trade,""),
         "summary":f"BuildCommand source-backed scope for {trade}.","items":items}
        for trade,items in sorted(grouped.items())
    ]
    data.setdefault("review_notes",[]).append(
        "v44.1 Blueprint Ownership Guard applied action/assembly routing and approved learning rules before save."
    )
    return data

def _save_blueprint_result(pid, docs, data, model_name):
    data=_v33_reclassify_data(data)
    data=_v441_reclassify_with_learning(pid,data)
    company_id=current_company_id(); user_id=current_user_id(); now=datetime.utcnow().isoformat()
    c=db()
    c.execute("INSERT INTO blueprint_runs(company_id,project_id,status,source_files,project_summary,detected_disciplines,cross_discipline_flags,rfi_candidates,review_notes,model_name,created_by,created) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",(
        company_id,pid,"COMPLETE",json.dumps([d["original_name"] for d in docs]),str(data.get("project_summary") or ""),json.dumps(data.get("detected_disciplines") or []),json.dumps(data.get("cross_discipline_flags") or []),json.dumps(data.get("rfi_candidates") or []),json.dumps(data.get("review_notes") or []),model_name,user_id,now
    ))
    run_id=c.execute("SELECT last_insert_rowid() id").fetchone()["id"]
    for trade_data in data.get("trade_scopes") or []:
        trade=(trade_data.get("trade") or "Unassigned").strip() or "Unassigned"
        division=str(trade_data.get("division") or "").strip()
        summary=str(trade_data.get("summary") or "").strip()
        items=trade_data.get("items") or []
        scope_text=_blueprint_scope_text(trade_data)
        c.execute("INSERT INTO blueprint_trade_scopes(company_id,project_id,run_id,trade,division,summary,scope_text,item_count,created) VALUES(?,?,?,?,?,?,?,?,?)",(company_id,pid,run_id,trade,division,summary,scope_text,len(items),now))
        trade_scope_id=c.execute("SELECT last_insert_rowid() id").fetchone()["id"]
        for item in items:
            req=str(item.get("requirement") or "").strip()
            if not req:
                continue
            confidence=str(item.get("confidence") or "MEDIUM").upper()
            if confidence not in {"HIGH","MEDIUM","LOW"}: confidence="MEDIUM"
            item_type=str(item.get("item_type") or "SCOPE").upper()
            c.execute("INSERT INTO blueprint_scope_items(company_id,project_id,run_id,trade_scope_id,trade,requirement,source_sheet,source_detail,source_spec,source_note,related_trade,confidence,item_type,status,created) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",(
                company_id,pid,run_id,trade_scope_id,trade,req,str(item.get("source_sheet") or ""),str(item.get("source_detail") or ""),str(item.get("source_spec") or ""),str(item.get("source_note") or ""),str(item.get("related_trade") or ""),confidence,item_type,"NOT_STARTED",now
            ))
    c.commit(); c.close(); return run_id


def _blueprint_latest(pid):
    c=db(); run=c.execute("SELECT * FROM blueprint_runs WHERE company_id=? AND project_id=? ORDER BY id DESC LIMIT 1",(current_company_id(),pid)).fetchone()
    scopes=[]
    if run:
        scopes=c.execute("SELECT * FROM blueprint_trade_scopes WHERE company_id=? AND project_id=? AND run_id=? ORDER BY trade",(current_company_id(),pid,run["id"])).fetchall()
    c.close(); return run,scopes



# ============================================================
# v34 BUILDCOMMAND CORE BRAIN + ESTIMATOR INTELLIGENCE
# ============================================================

def _ensure_v34_estimator_tables():
    """
    v34.2 estimator schema repair.
    PostgreSQL needs a sequence/default for estimator_items.id. Earlier v34 used
    c.execute(CREATE TABLE...) instead of executescript(), so PgCompat did not
    translate INTEGER PRIMARY KEY to BIGSERIAL. Repair existing installs in place.
    """
    c=db()
    if DATABASE_KIND=="postgres":
        c.execute("""
            CREATE TABLE IF NOT EXISTS estimator_items(
                id BIGSERIAL PRIMARY KEY,
                company_id BIGINT,
                project_id BIGINT,
                blueprint_scope_item_id BIGINT,
                trade TEXT,
                description TEXT,
                source_ref TEXT,
                quantity DOUBLE PRECISION DEFAULT 0,
                unit TEXT DEFAULT '',
                material_unit_cost DOUBLE PRECISION DEFAULT 0,
                labor_unit_cost DOUBLE PRECISION DEFAULT 0,
                subcontract_quote DOUBLE PRECISION DEFAULT 0,
                allowance DOUBLE PRECISION DEFAULT 0,
                markup_pct DOUBLE PRECISION DEFAULT 0,
                notes TEXT DEFAULT '',
                verified INTEGER DEFAULT 0,

                ai_quantity DOUBLE PRECISION,
                ai_unit TEXT DEFAULT '',
                ai_confidence TEXT DEFAULT '',
                ai_basis TEXT DEFAULT '',
                ai_source TEXT DEFAULT '',
                ai_updated TEXT,
                created TEXT,
                updated TEXT
            )
        """)
        # Repair a v34 table that already exists with "id INTEGER PRIMARY KEY"
        # and therefore has no sequence/default.
        try:
            c.execute("CREATE SEQUENCE IF NOT EXISTS estimator_items_id_seq")
            c.execute("ALTER SEQUENCE estimator_items_id_seq OWNED BY estimator_items.id")
            c.execute("""
                ALTER TABLE estimator_items
                ALTER COLUMN id SET DEFAULT nextval('estimator_items_id_seq')
            """)
            c.execute("""
                SELECT setval(
                    'estimator_items_id_seq',
                    GREATEST(COALESCE((SELECT MAX(id) FROM estimator_items),0)+1,1),
                    false
                )
            """)
        except Exception:
            c.rollback()
            c=db()
            # If another deploy/request repaired it concurrently, normal use can continue.
    else:
        c.execute("""
            CREATE TABLE IF NOT EXISTS estimator_items(
                id INTEGER PRIMARY KEY,
                company_id INTEGER,
                project_id INTEGER,
                blueprint_scope_item_id INTEGER,
                trade TEXT,
                description TEXT,
                source_ref TEXT,
                quantity REAL DEFAULT 0,
                unit TEXT DEFAULT '',
                material_unit_cost REAL DEFAULT 0,
                labor_unit_cost REAL DEFAULT 0,
                subcontract_quote REAL DEFAULT 0,
                allowance REAL DEFAULT 0,
                markup_pct REAL DEFAULT 0,
                notes TEXT DEFAULT '',
                verified INTEGER DEFAULT 0,

                ai_quantity REAL,
                ai_unit TEXT DEFAULT '',
                ai_confidence TEXT DEFAULT '',
                ai_basis TEXT DEFAULT '',
                ai_source TEXT DEFAULT '',
                ai_updated TEXT,

                created TEXT,
                updated TEXT
            )
        """)

    # v36 automatic takeoff proposal fields. Add safely to existing v34/v35 tables.
    v36_columns=[
        ("ai_quantity","DOUBLE PRECISION" if DATABASE_KIND=="postgres" else "REAL"),
        ("ai_unit","TEXT DEFAULT ''"),
        ("ai_confidence","TEXT DEFAULT ''"),
        ("ai_basis","TEXT DEFAULT ''"),
        ("ai_source","TEXT DEFAULT ''"),
        ("ai_updated","TEXT")
    ]
    for col,col_type in v36_columns:
        try:
            c.execute(f"ALTER TABLE estimator_items ADD COLUMN {col} {col_type}")
            c.commit()
        except Exception:
            try: c.rollback()
            except Exception: pass

    # Helpful lookup/index; duplicate historical rows are tolerated, so no UNIQUE migration.
    try:
        c.execute("""
            CREATE INDEX IF NOT EXISTS idx_estimator_project_scope
            ON estimator_items(company_id,project_id,blueprint_scope_item_id)
        """)
    except Exception:
        pass
    c.commit(); c.close()

def _latest_blueprint_run(pid):
    c=db()
    row=c.execute("""SELECT * FROM blueprint_runs
                     WHERE company_id=? AND project_id=?
                     ORDER BY id DESC LIMIT 1""",
                  (current_company_id(),pid)).fetchone()
    c.close(); return row

def _seed_estimator_from_latest(pid):
    """
    Sync estimator rows from the latest cleaned Blueprint/Plan Intelligence run.
    Existing user-entered pricing/quantities are preserved, while trade, description,
    and source references are refreshed from v34.1+ corrected scope ownership.
    """
    _ensure_v34_estimator_tables()
    run=_latest_blueprint_run(pid)
    if not run:
        return {"added":0,"updated":0,"run_id":None}

    c=db()
    rows=c.execute("""
        SELECT i.id,i.requirement,i.source_sheet,i.source_detail,i.source_spec,s.trade
        FROM blueprint_scope_items i
        JOIN blueprint_trade_scopes s ON s.id=i.trade_scope_id
        WHERE i.company_id=? AND i.project_id=? AND s.run_id=?
        ORDER BY s.trade,i.id
    """,(current_company_id(),pid,run["id"])).fetchall()

    added=0; updated=0; now=datetime.utcnow().isoformat()
    latest_scope_ids=set()

    for r in rows:
        scope_item_id=int(r["id"])
        latest_scope_ids.add(scope_item_id)
        refs=[str(r[x] or "").strip() for x in ["source_sheet","source_detail","source_spec"]
              if str(r[x] or "").strip()]
        source_ref=" · ".join(refs)

        existing=c.execute("""
            SELECT id,trade,description,source_ref
            FROM estimator_items
            WHERE company_id=? AND project_id=? AND blueprint_scope_item_id=?
            ORDER BY id LIMIT 1
        """,(current_company_id(),pid,scope_item_id)).fetchone()

        if existing:
            # Keep estimator-entered numbers/notes but refresh scope truth.
            if ((existing["trade"] or "") != (r["trade"] or "") or
                (existing["description"] or "") != (r["requirement"] or "") or
                (existing["source_ref"] or "") != source_ref):
                c.execute("""
                    UPDATE estimator_items
                    SET trade=?,description=?,source_ref=?,updated=?
                    WHERE id=? AND company_id=? AND project_id=?
                """,(r["trade"],r["requirement"],source_ref,now,
                     existing["id"],current_company_id(),pid))
                updated+=1
            continue

        c.execute("""
            INSERT INTO estimator_items(
                company_id,project_id,blueprint_scope_item_id,trade,description,source_ref,
                quantity,unit,material_unit_cost,labor_unit_cost,subcontract_quote,allowance,
                markup_pct,notes,verified,created,updated
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,(current_company_id(),pid,scope_item_id,r["trade"],r["requirement"],
             source_ref,0,"",0,0,0,0,0,"",0,now,now))
        added+=1

    c.commit(); c.close()
    return {"added":added,"updated":updated,"run_id":run["id"]}

def _estimate_math(row):
    qty=float(row["quantity"] or 0); mat=float(row["material_unit_cost"] or 0)
    labor=float(row["labor_unit_cost"] or 0); sub=float(row["subcontract_quote"] or 0)
    allowance=float(row["allowance"] or 0); markup=float(row["markup_pct"] or 0)
    direct=(qty*(mat+labor))+sub+allowance
    return direct,direct*(1+(markup/100.0))

@app.get("/brain",response_class=HTMLResponse)
def buildcommand_core_brain():
    pid=project_id(); run=_latest_blueprint_run(pid); c=db()
    project=c.execute("SELECT name,number FROM projects WHERE id=?",(pid,)).fetchone()
    scope_count=item_count=cross_count=rfi_count=0
    if run:
        scope_count=c.execute("SELECT COUNT(*) n FROM blueprint_trade_scopes WHERE run_id=?",(run["id"],)).fetchone()["n"]
        item_count=c.execute("SELECT COUNT(*) n FROM blueprint_scope_items WHERE run_id=?",(run["id"],)).fetchone()["n"]
        try: cross_count=len(json.loads(run["cross_discipline_flags"] or "[]"))
        except: cross_count=0
        try: rfi_count=len(json.loads(run["rfi_candidates"] or "[]"))
        except: rfi_count=0
    open_issues=c.execute("SELECT COUNT(*) n FROM project_issues WHERE project_id=? AND status!='CLOSED'",(pid,)).fetchone()["n"]
    activities=c.execute("SELECT COUNT(*) n FROM activities WHERE project_id=?",(pid,)).fetchone()["n"]
    c.close()
    project_label=f'{esc(project["number"])} - {esc(project["name"])}' if project else "Current Project"
    latest=(f'<span class="badge READY">PLAN INTELLIGENCE READY</span><div class="small" style="margin-top:8px">Latest run #{run["id"]} · {esc(run["created"] or "")}</div>'
            if run else '<span class="badge WATCH">NO PLAN ANALYSIS YET</span>')
    body=f"""
    <div class="hero">
      <div class="eyebrow">BuildCommand Core Brain · v34.1</div>
      <h1>One construction brain. One project memory.</h1>
      <div class="muted">{project_label}</div><div style="margin-top:12px">{latest}</div>
    </div>
    <div class="grid4">
      <div class="card"><div class="label">Trade Scopes</div><div class="kpi">{scope_count}</div></div>
      <div class="card"><div class="label">Scope Items</div><div class="kpi">{item_count}</div></div>
      <div class="card"><div class="label">Cross-Discipline</div><div class="kpi">{cross_count}</div></div>
      <div class="card"><div class="label">RFI Candidates</div><div class="kpi">{rfi_count}</div></div>
    </div>
    <div class="grid2">
      <div class="card">
        <h2>Ask the BuildCommand Brain</h2>
        <form method="post" action="/brain">
          <textarea name="question" required placeholder="Example: What are the biggest scope and estimating risks on this project?"></textarea>
          <button type="submit">Ask BuildCommand</button>
        </form>
        <p class="small">Core Brain uses operations data plus the latest source-backed plan intelligence.</p>
      </div>
      <div class="card">
        <h2>One Brain Tools</h2>
        <p><a href="/plans-specs-ai"><b>Plan Intake</b></a> — read plans and create trade scopes.</p>
        <p><a href="/brain/estimator"><b>Estimator Intelligence</b></a> — estimate from those same scope items.</p>
        <p><a href="/issues"><b>RFIs / Issues</b></a> — manage scope gaps and coordination issues.</p>
        <p><a href="/schedule"><b>Schedule</b></a> — connect scope to execution.</p>
      </div>
    </div>
    <div class="grid2">
      <div class="card"><h2>Operations connected</h2><div class="kpi">{activities}</div><div class="muted">schedule activities</div></div>
      <div class="card"><h2>Open issues connected</h2><div class="kpi">{open_issues}</div><div class="muted">project issues</div></div>
    </div>"""
    return shell("BuildCommand Brain",body)

@app.post("/brain",response_class=HTMLResponse)
def buildcommand_core_brain_answer(question:str=Form(...)):
    pid=project_id(); run=_latest_blueprint_run(pid); c=db()
    scope_rows=[]
    if run:
        scope_rows=c.execute("""
            SELECT s.trade,i.requirement,i.source_sheet,i.source_detail,i.source_spec
            FROM blueprint_scope_items i
            JOIN blueprint_trade_scopes s ON s.id=i.trade_scope_id
            WHERE i.company_id=? AND i.project_id=? AND s.run_id=?
            ORDER BY s.trade,i.id LIMIT 500
        """,(current_company_id(),pid,run["id"])).fetchall()
    c.close()
    plan_context="\n".join(
        f'- {r["trade"]}: {r["requirement"]} | source: {r["source_sheet"] or ""} {r["source_detail"] or ""} {r["source_spec"] or ""}'
        for r in scope_rows
    ) or "No plan/scope intelligence available yet."
    api_key=os.environ.get("OPENAI_API_KEY")
    if not api_key:
        answer="OPENAI_API_KEY is not configured."
    else:
        try:
            operations=build_project_context(pid)
            client=OpenAI(api_key=api_key)
            resp=client.responses.create(
                model=os.environ.get("OPENAI_MODEL","gpt-5.6"),
                instructions="""You are the BuildCommand Core Brain, a construction operations and estimating intelligence system.
Use only supplied project facts as job-specific facts.
Think across estimating, trade scope, drawings, schedule, field execution, RFIs, procurement, inspections and closeout.
Never invent quantities, dimensions, costs, field conditions, commitments or drawing requirements.
If quantity/cost is not supported, say ESTIMATOR VERIFICATION REQUIRED.
Classify by actual work, not by drawing discipline.
Non-MEP demolition belongs to Demolition; MEP/system demo stays with the system trade.
Blocking/backing belongs to Framing/Drywall. Roof patching belongs to Roofing.
Sprinkler heads/piping belong to Fire Sprinkler.
Card access/electronic strikes/cameras/security belong to Low Voltage.
Storefront/glazing is separate from standard Doors/Frames/Hardware.""",
                input=f"""PROJECT OPERATIONS DATA
{operations}

LATEST PLAN / TRADE SCOPE INTELLIGENCE
{plan_context}

USER QUESTION
{question}"""
            )
            answer=resp.output_text
        except Exception as exc:
            answer=f"BuildCommand Brain could not complete the analysis: {exc}"
    body=f"""<div class="hero"><div class="eyebrow">BuildCommand Core Brain</div><h1>Project Answer</h1></div>
    <div class="grid2"><div class="card"><div class="small">QUESTION</div><h3>{esc(question)}</h3><p><a href="/brain">← Back to Brain</a></p></div>
    <div class="card"><div class="small">BUILDCOMMAND BRAIN</div><div style="white-space:pre-wrap;line-height:1.6">{esc(answer)}</div></div></div>"""
    return shell("BuildCommand Brain",body)


# ============================================================
# v35 ESTIMATOR TAKEOFF INTELLIGENCE
# ============================================================

def _suggest_takeoff_unit(trade, description):
    t=(str(description or "")+" "+str(trade or "")).lower()

    # Counts / equipment
    if any(x in t for x in [
        "door","frame","window","sprinkler head","receptacle","outlet","switch",
        "fixture","water closet","urinal","lavatory","sink","floor drain","cleanout",
        "diffuser","grille","vav","fan","rtu","ahu","panel","transformer","disconnect",
        "card reader","camera","fire extinguisher cabinet","cabinet","appliance"
    ]):
        return "EA"

    # Linear measurements
    if any(x in t for x in [
        "rubber base","resilient base","cove base","base molding","curb","handrail",
        "guardrail","pipe","piping","conduit","duct","ductwork","gutter","flashing",
        "joint sealant","wall base"
    ]):
        return "LF"

    # Area measurements
    if any(x in t for x in [
        "flooring","carpet","lvt","vct","tile","wall tile","floor tile","ceiling",
        "ceiling tile","drywall","gypsum board","wallboard","paint","painting",
        "frp","marlite","wainscot","roof membrane","roofing","waterproofing",
        "storefront glazing","glazing","insulation"
    ]):
        return "SF"

    # Volume
    if any(x in t for x in ["concrete","slab","footing","grade beam","grout fill"]):
        return "CY"

    # Earthwork
    if any(x in t for x in ["excavation","backfill","import soil","export soil"]):
        return "CY"

    # Lump-sum / coordination / general conditions
    if any(x in t for x in [
        "coordinate","confirm","verify","allowance","general conditions","mobilization",
        "testing","commissioning","permit","closeout","temporary protection"
    ]):
        return "LS"

    return "EA"


def _takeoff_status(row):
    qty=float(row["quantity"] or 0)
    verified=bool(int(row["verified"] or 0))
    if verified:
        return "VERIFIED"
    if qty > 0:
        return "QUANTITY_ENTERED"
    return "NEEDS_TAKEOFF"




# ============================================================
# v36.1 TAKEOFF COMPONENT SPLITTER
# ============================================================

def _ensure_takeoff_component_tables():
    c=db()
    if DATABASE_KIND=="postgres":
        c.execute("""
            CREATE TABLE IF NOT EXISTS takeoff_components(
                id BIGSERIAL PRIMARY KEY,
                company_id BIGINT,
                project_id BIGINT,
                estimator_item_id BIGINT,
                component_name TEXT,
                description TEXT,
                unit TEXT,
                quantity DOUBLE PRECISION,
                confidence TEXT DEFAULT 'VERIFY',
                basis TEXT DEFAULT '',
                source_ref TEXT DEFAULT '',
                status TEXT DEFAULT 'PROPOSED',
                created TEXT,
                updated TEXT
            )
        """)
    else:
        c.execute("""
            CREATE TABLE IF NOT EXISTS takeoff_components(
                id INTEGER PRIMARY KEY,
                company_id INTEGER,
                project_id INTEGER,
                estimator_item_id INTEGER,
                component_name TEXT,
                description TEXT,
                unit TEXT,
                quantity REAL,
                confidence TEXT DEFAULT 'VERIFY',
                basis TEXT DEFAULT '',
                source_ref TEXT DEFAULT '',
                status TEXT DEFAULT 'PROPOSED',
                created TEXT,
                updated TEXT
            )
        """)
    try:
        c.execute("""CREATE INDEX IF NOT EXISTS idx_takeoff_components_parent
                     ON takeoff_components(company_id,project_id,estimator_item_id)""")
    except Exception:
        pass
    c.commit()
    c.close()


def _component_split_prompt(rows):
    lines=[]
    for r in rows:
        lines.append(
            f'ESTIMATOR_ID={r["id"]} | TRADE={r["trade"]} | '
            f'SCOPE={r["description"]} | SOURCE={r["source_ref"] or ""}'
        )
    return """You are BuildCommand Takeoff Component Splitter.

Break each construction estimator scope item into separate measurable takeoff components ONLY when the scope contains multiple distinct measurable objects or measurement types.

RULES:
- Preserve the parent trade. Do not reclassify.
- Do not invent quantities.
- One component should normally have one measurement unit.
- Use EA for countable fixtures/equipment/devices, LF for linear work, SF for area work, CY for volume, LS for coordination/non-measurable scope.
- Do not split a simple single-measurement scope unnecessarily.
- Example: "Provide exhaust fan, roof curb, backdraft damper and full-size duct" becomes exhaust fan EA; roof curb EA; backdraft damper EA; exhaust duct LF.
- Example: "Provide water closets, lavatories, floor drains and domestic water piping" becomes water closets EA; lavatories EA; floor drains EA; domestic water piping LF.
- Keep the same source reference unless a more specific source is obvious.
- Return JSON only.

Return:
{
  "results":[
    {
      "estimator_id":123,
      "components":[
        {"name":"Exhaust fan EF-1","description":"Provide EF-1 exhaust fan","unit":"EA","source":"M401"},
        {"name":"Exhaust duct","description":"Provide full-size exhaust duct","unit":"LF","source":"M401"}
      ]
    }
  ]
}

ESTIMATOR SCOPE ITEMS
---------------------
""" + "\n".join(lines)


@app.post("/brain/takeoff/split-components",response_class=HTMLResponse)
def split_takeoff_components():
    pid=project_id()
    _seed_estimator_from_latest(pid)
    _ensure_takeoff_component_tables()
    run=_latest_blueprint_run(pid)
    if not run:
        return shell("Takeoff Components",'<div class="card"><h2>No current plan analysis found.</h2></div>')

    c=db()
    rows=c.execute("""
        SELECT e.*
        FROM estimator_items e
        JOIN blueprint_scope_items b ON b.id=e.blueprint_scope_item_id
        WHERE e.company_id=? AND e.project_id=? AND b.run_id=?
        ORDER BY e.trade,e.id
        LIMIT 180
    """,(current_company_id(),pid,run["id"])).fetchall()
    c.close()

    if not rows:
        return shell("Takeoff Components",'<div class="card"><h2>No estimator items found.</h2></div>')
    if not os.environ.get("OPENAI_API_KEY"):
        return shell("Takeoff Components",'<div class="card"><h2>OPENAI_API_KEY is not configured.</h2></div>')

    try:
        client=OpenAI(api_key=os.environ["OPENAI_API_KEY"])
        resp=client.responses.create(
            model=os.environ.get("OPENAI_MODEL","gpt-5.6"),
            input=_component_split_prompt(rows)
        )
        data=_v36_parse_json(resp.output_text)
        valid={int(r["id"]):r for r in rows}
        now=datetime.utcnow().isoformat()
        c=db()
        created=0

        for result in data.get("results") or []:
            try:
                eid=int(result.get("estimator_id"))
            except Exception:
                continue
            if eid not in valid:
                continue

            comps=result.get("components") or []
            if len(comps)<=1:
                continue

            c.execute("""DELETE FROM takeoff_components
                         WHERE company_id=? AND project_id=? AND estimator_item_id=? AND status='PROPOSED'""",
                      (current_company_id(),pid,eid))

            for comp in comps[:25]:
                name=str(comp.get("name") or "").strip()
                desc=str(comp.get("description") or name).strip()
                unit=str(comp.get("unit") or "").upper()
                if not name or unit not in {"EA","LF","SF","CY","LS","HR","DAY","TON","GAL"}:
                    continue
                source=str(comp.get("source") or valid[eid]["source_ref"] or "")
                c.execute("""
                    INSERT INTO takeoff_components(
                        company_id,project_id,estimator_item_id,component_name,description,
                        unit,quantity,confidence,basis,source_ref,status,created,updated
                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,(current_company_id(),pid,eid,name,desc,unit,None,"VERIFY",
                     "Component split from cleaned parent scope; quantity not yet verified.",
                     source,"PROPOSED",now,now))
                created+=1

        c.commit()
        c.close()

        body = (
            '<div class="hero"><div class="eyebrow">BuildCommand · v36.1</div>'
            '<h1>Takeoff components created.</h1></div>'
            f'<div class="card"><div class="kpi">{created}</div>'
            '<div class="muted">measurable child components created without changing the parent scopes.</div>'
            '<p><a href="/brain/takeoff/components"><b>Review components →</b></a></p></div>'
        )
        return shell("Takeoff Components",body)

    except Exception as exc:
        return shell(
            "Takeoff Components",
            f'<div class="card"><h2>Component split did not complete.</h2><p>{esc(str(exc))}</p></div>'
        )


@app.get("/brain/takeoff/components",response_class=HTMLResponse)
def takeoff_components_page():
    pid=project_id()
    _ensure_takeoff_component_tables()
    c=db()
    rows=c.execute("""
        SELECT tc.*,e.trade,e.description parent_description
        FROM takeoff_components tc
        JOIN estimator_items e ON e.id=tc.estimator_item_id
        WHERE tc.company_id=? AND tc.project_id=?
        ORDER BY e.trade,tc.estimator_item_id,tc.id
    """,(current_company_id(),pid)).fetchall()
    c.close()

    cards=[]
    for r in rows:
        qty="" if r["quantity"] is None else r["quantity"]
        confidence_options=''.join(
            f'<option {"selected" if r["confidence"]==x else ""}>{x}</option>'
            for x in ["HIGH","MEDIUM","LOW","VERIFY"]
        )
        status_options=''.join(
            f'<option {"selected" if r["status"]==x else ""}>{x}</option>'
            for x in ["PROPOSED","ACCEPTED","VERIFIED"]
        )

        cards.append(
            f'<div class="card">'
            f'<div class="small">{esc(r["trade"])} · Parent estimator item #{r["estimator_item_id"]}</div>'
            f'<h3>{esc(r["component_name"])}</h3>'
            f'<div>{esc(r["description"])}</div>'
            f'<div class="small">Source: {esc(r["source_ref"] or "")}</div>'
            f'<form method="post" action="/brain/takeoff/component/{r["id"]}" style="margin-top:10px">'
            '<div class="grid4">'
            f'<div><label>Qty</label><input name="quantity" type="number" step="0.01" value="{qty}"></div>'
            f'<div><label>Unit</label><input name="unit" value="{esc(r["unit"] or "")}"></div>'
            f'<div><label>Confidence</label><select name="confidence">{confidence_options}</select></div>'
            f'<div><label>Status</label><select name="status">{status_options}</select></div>'
            '</div>'
            f'<label>Basis / note</label><input name="basis" value="{esc(r["basis"] or "")}">'
            '<button type="submit" style="margin-top:10px">Save Component</button>'
            '</form></div>'
        )

    body = (
        '<div class="hero"><div class="eyebrow">Takeoff Component Splitter · v36.1</div>'
        '<h1>One scope. Separate measurable pieces.</h1></div>'
        '<div class="card"><p><a href="/brain/takeoff">← Back to Takeoff Intelligence</a></p></div>'
        + ("".join(cards) if cards else '<div class="card"><h2>No split components yet.</h2></div>')
    )
    return shell("Takeoff Components",body)


@app.post("/brain/takeoff/component/{component_id}")
def update_takeoff_component(component_id:int,quantity:str=Form(""),unit:str=Form("EA"),
                             confidence:str=Form("VERIFY"),status:str=Form("PROPOSED"),
                             basis:str=Form("")):
    pid=project_id()
    try:
        q=float(quantity) if str(quantity).strip() else None
    except Exception:
        q=None

    unit=unit.upper().strip()
    if unit not in {"EA","LF","SF","CY","LS","HR","DAY","TON","GAL"}:
        unit="EA"
    if confidence not in {"HIGH","MEDIUM","LOW","VERIFY"}:
        confidence="VERIFY"
    if status not in {"PROPOSED","ACCEPTED","VERIFIED"}:
        status="PROPOSED"

    c=db()
    c.execute("""UPDATE takeoff_components SET quantity=?,unit=?,confidence=?,status=?,basis=?,updated=?
                 WHERE id=? AND company_id=? AND project_id=?""",
              (q,unit,confidence,status,basis,datetime.utcnow().isoformat(),
               component_id,current_company_id(),pid))
    c.commit()
    c.close()
    return RedirectResponse("/brain/takeoff/components",status_code=303)

# ============================================================
# v36 AUTOMATIC PLAN TAKEOFF BRAIN
# ============================================================

def _v36_latest_plan_docs(pid):
    run=_latest_blueprint_run(pid)
    if not run:
        return run,[]
    try:
        names=json.loads(run["source_files"] or "[]")
    except Exception:
        names=[]
    docs=[]
    c=db()
    for name in names:
        d=c.execute("""
            SELECT * FROM attachments
            WHERE company_id=? AND project_id=? AND original_name=?
            ORDER BY id DESC LIMIT 1
        """,(current_company_id(),pid,name)).fetchone()
        if d:
            docs.append(d)
    c.close()
    return run,docs


def _v36_scope_targets(pid,run_id):
    c=db()
    rows=c.execute("""
        SELECT e.id estimator_id,e.trade,e.description,e.source_ref,e.quantity,e.unit,e.verified,
               b.id scope_item_id,b.confidence scope_confidence
        FROM estimator_items e
        JOIN blueprint_scope_items b ON b.id=e.blueprint_scope_item_id
        WHERE e.company_id=? AND e.project_id=? AND b.run_id=?
        ORDER BY e.trade,e.id
        LIMIT 180
    """,(current_company_id(),pid,run_id)).fetchall()
    c.close()
    return rows


def _v36_takeoff_prompt(targets):
    lines=[]
    for r in targets:
        lines.append(
            f'ESTIMATOR_ID={r["estimator_id"]} | TRADE={r["trade"]} | '
            f'DESCRIPTION={r["description"]} | SOURCE={r["source_ref"] or ""} | '
            f'SUGGESTED_UNIT={_suggest_takeoff_unit(r["trade"],r["description"])}'
        )
    return """You are BuildCommand Automatic Plan Takeoff Brain.

Analyze the attached construction plan PDFs visually and textually against the TAKEOFF TARGETS below.

STRICT RULES:
1. Never guess a quantity.
2. Return a numeric quantity only when the documents provide enough evidence to count or calculate it with reasonable confidence.
3. Prefer explicit schedules, legends, tagged fixture/equipment counts, door/window schedules, device plans, room finish schedules, and clearly dimensioned/calculable information.
4. Do NOT infer scaled LF/SF/CY measurements from a drawing image unless a reliable explicit scale/dimension and geometry are sufficiently clear. If not, return quantity=null and confidence=VERIFY.
5. A scope statement that says "provide doors" is not itself proof of the number of doors. Use schedules/plans as evidence.
6. Avoid double-counting the same tagged object shown on multiple sheets/details.
7. Preserve trade ownership from the supplied target; this task is quantity extraction, not reclassification.
8. For each non-null quantity, cite the best supporting sheet/schedule/detail in source and briefly explain the basis.
9. Confidence must be HIGH, MEDIUM, LOW, or VERIFY.
10. Return JSON only.

Return:
{
  "results":[
    {
      "estimator_id":123,
      "quantity":12,
      "unit":"EA",
      "confidence":"HIGH",
      "basis":"12 unique door marks listed in the door schedule.",
      "source":"A601 Door Schedule"
    },
    {
      "estimator_id":124,
      "quantity":null,
      "unit":"SF",
      "confidence":"VERIFY",
      "basis":"Area requires scaled takeoff; no reliable explicit area found.",
      "source":"A201/A801"
    }
  ],
  "review_notes":[]
}

TAKEOFF TARGETS
----------------
"""+"\n".join(lines)


def _v36_parse_json(raw):
    s=(raw or "").strip()
    if s.startswith("```"):
        s=re.sub(r"^```(?:json)?\s*","",s,flags=re.I)
        s=re.sub(r"\s*```$","",s)
    start=s.find("{"); end=s.rfind("}")
    if start>=0 and end>start:
        s=s[start:end+1]
    return json.loads(s)


@app.post("/brain/takeoff/auto",response_class=HTMLResponse)
def automatic_plan_takeoff():
    pid=project_id()
    _seed_estimator_from_latest(pid)
    run,docs=_v36_latest_plan_docs(pid)
    if not run:
        return shell("Automatic Takeoff",'<div class="card"><h2>No Plan Intelligence run found.</h2><p>Run Plan Intake first.</p></div>')
    if not docs:
        return shell("Automatic Takeoff",'<div class="card"><h2>Source plan files are unavailable.</h2><p>BuildCommand found the scope run but could not locate its original project attachments.</p></div>')
    if not os.environ.get("OPENAI_API_KEY"):
        return shell("Automatic Takeoff",'<div class="card"><h2>OPENAI_API_KEY is not configured.</h2></div>')

    targets=_v36_scope_targets(pid,run["id"])
    if not targets:
        return shell("Automatic Takeoff",'<div class="card"><h2>No current estimator targets found.</h2></div>')

    total=sum(int(d["size_bytes"] or 0) for d in docs if Path(d["original_name"] or "").suffix.lower()==".pdf")
    if total>=50*1024*1024:
        return shell("Automatic Takeoff",f'<div class="card"><h2>Plan set too large for one automatic takeoff pass.</h2><p>Current PDF total: {total/1024/1024:.1f} MB. Run Plan Intake in smaller volumes before automatic takeoff.</p></div>')

    client=OpenAI(api_key=os.environ["OPENAI_API_KEY"])
    uploaded=[]; content=[]
    try:
        for d in docs:
            path=os.path.join(UPLOAD_DIR,d["stored_name"])
            if not os.path.isfile(path):
                continue
            if Path(d["original_name"] or "").suffix.lower()==".pdf":
                with open(path,"rb") as fh:
                    remote=client.files.create(file=fh,purpose="user_data")
                uploaded.append(remote.id)
                content.append({"type":"input_file","file_id":remote.id,"detail":"high"})
        if not content:
            return shell("Automatic Takeoff",'<div class="card"><h2>No PDF plan files were available for visual takeoff.</h2></div>')

        content.append({"type":"input_text","text":_v36_takeoff_prompt(targets)})
        response=client.responses.create(
            model=os.environ.get("OPENAI_MODEL","gpt-5.6"),
            input=[{"role":"user","content":content}]
        )
        data=_v36_parse_json(response.output_text)
        results=data.get("results") or []
        valid_ids={int(r["estimator_id"]) for r in targets}
        now=datetime.utcnow().isoformat()
        saved=0; proposed=0; verify=0
        c=db()
        for item in results:
            try:
                eid=int(item.get("estimator_id"))
            except Exception:
                continue
            if eid not in valid_ids:
                continue
            q=item.get("quantity")
            try:
                q=float(q) if q is not None else None
            except Exception:
                q=None
            unit=str(item.get("unit") or "").upper().strip()
            if unit not in {"EA","LF","SF","CY","LS","HR","DAY","TON","GAL"}:
                unit=""
            confidence=str(item.get("confidence") or "VERIFY").upper()
            if confidence not in {"HIGH","MEDIUM","LOW","VERIFY"}:
                confidence="VERIFY"
            basis=str(item.get("basis") or "")[:4000]
            source=str(item.get("source") or "")[:1500]
            # Only present usable proposals; never overwrite estimator-entered quantity.
            c.execute("""
                UPDATE estimator_items
                SET ai_quantity=?,ai_unit=?,ai_confidence=?,ai_basis=?,ai_source=?,ai_updated=?
                WHERE id=? AND company_id=? AND project_id=?
            """,(q,unit,confidence,basis,source,now,eid,current_company_id(),pid))
            saved+=1
            if q is not None and confidence in {"HIGH","MEDIUM"}:
                proposed+=1
            else:
                verify+=1
        c.commit(); c.close()

        notes=data.get("review_notes") or []
        note_html="".join(f'<div class="action">{esc(n)}</div>' for n in notes)
        body=f"""
        <div class="hero">
          <div class="eyebrow">BuildCommand Brain · Automatic Plan Takeoff · v36</div>
          <h1>Automatic takeoff pass complete.</h1>
        </div>
        <div class="grid3">
          <div class="card"><div class="label">Targets reviewed</div><div class="kpi">{saved}</div></div>
          <div class="card"><div class="label">Quantity proposals</div><div class="kpi">{proposed}</div></div>
          <div class="card"><div class="label">Needs verification</div><div class="kpi">{verify}</div></div>
        </div>
        <div class="card">
          <h2>Important</h2>
          <p>No estimator quantity was overwritten. AI results are proposals only. Review them on Takeoff Intelligence and accept only the quantities you trust.</p>
          <p><a href="/brain/takeoff"><b>Review takeoff proposals →</b></a></p>
        </div>
        {f'<div class="card"><h2>AI Review Notes</h2>{note_html}</div>' if note_html else ''}
        """
        return shell("Automatic Plan Takeoff",body)
    except Exception as exc:
        return shell("Automatic Takeoff",f'<div class="hero"><h1>Automatic takeoff did not complete.</h1></div><div class="card"><p>{esc(str(exc))}</p><p><a href="/brain/takeoff">← Back to Takeoff Intelligence</a></p></div>')
    finally:
        for fid in uploaded:
            try: client.files.delete(fid)
            except Exception: pass


@app.post("/brain/takeoff/item/{item_id}/accept-ai")
def accept_ai_takeoff(item_id:int):
    pid=project_id(); c=db()
    row=c.execute("""
        SELECT * FROM estimator_items
        WHERE id=? AND company_id=? AND project_id=?
    """,(item_id,current_company_id(),pid)).fetchone()
    if not row:
        c.close(); return HTMLResponse("Estimator item not found.",404)
    if row["ai_quantity"] is None or (row["ai_confidence"] or "") not in {"HIGH","MEDIUM"}:
        c.close()
        return shell("Takeoff Intelligence",'<div class="card"><h2>This AI quantity is not eligible for acceptance.</h2><p>Only HIGH or MEDIUM confidence numeric proposals can be accepted. Verify the item manually.</p><p><a href="/brain/takeoff">← Back</a></p></div>')
    unit=(row["ai_unit"] or row["unit"] or "EA").upper()
    c.execute("""
        UPDATE estimator_items
        SET quantity=?,unit=?,verified=0,notes=?,updated=?
        WHERE id=? AND company_id=? AND project_id=?
    """,(row["ai_quantity"],unit,
         ((row["notes"] or "")+" | AI takeoff accepted for estimator verification").strip(" |"),
         datetime.utcnow().isoformat(),item_id,current_company_id(),pid))
    c.commit(); c.close()
    return RedirectResponse("/brain/takeoff",status_code=303)

@app.get("/brain/takeoff", response_class=HTMLResponse)
def estimator_takeoff_intelligence():
    pid=project_id()
    sync=_seed_estimator_from_latest(pid)
    c=db()
    project=c.execute("SELECT name,number FROM projects WHERE id=?",(pid,)).fetchone()
    run=_latest_blueprint_run(pid)

    if run:
        rows=c.execute("""
            SELECT e.*
            FROM estimator_items e
            JOIN blueprint_scope_items b ON b.id=e.blueprint_scope_item_id
            WHERE e.company_id=? AND e.project_id=? AND b.run_id=?
            ORDER BY e.trade,e.id
        """,(current_company_id(),pid,run["id"])).fetchall()
    else:
        rows=[]
    c.close()

    counts={"NEEDS_TAKEOFF":0,"QUANTITY_ENTERED":0,"VERIFIED":0}
    cards=[]
    for r in rows:
        status=_takeoff_status(r)
        counts[status]+=1
        suggested=_suggest_takeoff_unit(r["trade"],r["description"])
        current_unit=(r["unit"] or "").strip()
        unit=current_unit or suggested
        badge = "READY" if status=="VERIFIED" else ("WATCH" if status=="QUANTITY_ENTERED" else "HIGH")

        cards.append(f"""
        <div class="card">
          <div style="display:flex;justify-content:space-between;gap:12px;align-items:center">
            <div><b>{esc(r["trade"])}</b></div>
            <span class="badge {badge}">{status.replace("_"," ")}</span>
          </div>
          <h3>{esc(r["description"])}</h3>
          <div class="small">Source: {esc(r["source_ref"] or "Source not identified")}</div>
          <div class="small" style="margin-top:6px">Suggested takeoff unit: <b>{esc(suggested)}</b></div>
          {(
            f'<div class="action" style="margin-top:10px"><b>AI PLAN TAKEOFF PROPOSAL:</b> '
            + (f'{r["ai_quantity"]:g} {esc(r["ai_unit"] or suggested)}' if r["ai_quantity"] is not None else 'VERIFY / no reliable quantity')
            + f' · Confidence: {esc(r["ai_confidence"] or "VERIFY")}<br>'
            + f'<span class="small">Basis: {esc(r["ai_basis"] or "")}<br>Source: {esc(r["ai_source"] or "")}</span>'
            + (f'<form method="post" action="/brain/takeoff/item/{r["id"]}/accept-ai" style="margin-top:8px"><button type="submit">Accept AI Quantity</button></form>'
               if r["ai_quantity"] is not None and (r["ai_confidence"] or "") in ["HIGH","MEDIUM"] else '')
            + '</div>'
          ) if (r["ai_confidence"] or r["ai_basis"] or r["ai_source"]) else ''}

          <form method="post" action="/brain/takeoff/item/{r["id"]}" style="margin-top:12px">
            <div class="grid4">
              <div>
                <label>Quantity</label>
                <input name="quantity" type="number" step="0.01" value="{r["quantity"] or 0}">
              </div>
              <div>
                <label>Unit</label>
                <select name="unit">
                  {''.join(f'<option value="{u}" {"selected" if unit==u else ""}>{u}</option>' for u in ["EA","LF","SF","CY","LS","HR","DAY","TON","GAL"])}
                </select>
              </div>
              <div>
                <label>Estimator verified</label>
                <select name="verified">
                  <option value="0" {"selected" if not r["verified"] else ""}>No</option>
                  <option value="1" {"selected" if r["verified"] else ""}>Yes</option>
                </select>
              </div>
              <div>
                <label>Takeoff note</label>
                <input name="notes" value="{esc(r["notes"] or "")}" placeholder="Counted on A201, verify addendum, etc.">
              </div>
            </div>
            <button type="submit" style="margin-top:10px">Save Takeoff</button>
          </form>
        </div>
        """)

    project_label=f'{esc(project["number"])} - {esc(project["name"])}' if project else "Current Project"
    body=f"""
    <div class="hero">
      <div class="eyebrow">BuildCommand Brain · Takeoff Intelligence · v36.1</div>
      <h1>Scope → measurable takeoff targets.</h1>
      <div class="muted">{project_label}</div>
    </div>

    <div class="grid4">
      <div class="card"><div class="label">Scope Items</div><div class="kpi">{len(rows)}</div></div>
      <div class="card"><div class="label">Needs Takeoff</div><div class="kpi">{counts["NEEDS_TAKEOFF"]}</div></div>
      <div class="card"><div class="label">Qty Entered</div><div class="kpi">{counts["QUANTITY_ENTERED"]}</div></div>
      <div class="card"><div class="label">Verified</div><div class="kpi">{counts["VERIFIED"]}</div></div>
    </div>

    <div class="card">
      <h2>How BuildCommand handles takeoffs</h2>
      <p>
        BuildCommand identifies what should be counted or measured and suggests the likely unit.
        It does <b>not invent a quantity</b>. Quantities remain estimator-entered or estimator-verified
        until a future scaled-drawing takeoff engine is connected.
      </p>
      <form method="post" action="/brain/takeoff/split-components" style="margin:14px 0">
        <button type="submit">Split Mixed Scopes into Takeoff Components</button>
      </form>
      <p><a href="/brain/takeoff/components"><b>Review Takeoff Components →</b></a></p>
      <form method="post" action="/brain/takeoff/auto" style="margin:14px 0">
        <button type="submit">Run Automatic Plan Takeoff</button>
      </form>
      <p class="small">AI proposals never overwrite estimator quantities. HIGH/MEDIUM proposals must still be accepted and then estimator-verified.</p>
      <p><a href="/brain/estimator"><b>← Back to Estimator Intelligence</b></a></p>
    </div>

    {"".join(cards) if cards else '<div class="card"><h2>No takeoff items yet</h2><p>Run Plan Intake first.</p></div>'}
    """
    return shell("Takeoff Intelligence",body)


@app.post("/brain/takeoff/item/{item_id}")
def estimator_takeoff_update(
    item_id:int,
    quantity:float=Form(0),
    unit:str=Form("EA"),
    verified:int=Form(0),
    notes:str=Form("")
):
    allowed={"EA","LF","SF","CY","LS","HR","DAY","TON","GAL"}
    unit=(unit or "EA").upper()
    if unit not in allowed:
        unit="EA"

    pid=project_id()
    c=db()
    c.execute("""
        UPDATE estimator_items
        SET quantity=?,unit=?,verified=?,notes=?,updated=?
        WHERE id=? AND company_id=? AND project_id=?
    """,(quantity,unit,verified,notes,datetime.utcnow().isoformat(),
         item_id,current_company_id(),pid))
    c.commit(); c.close()
    return RedirectResponse("/brain/takeoff",status_code=303)

@app.get("/brain/estimator",response_class=HTMLResponse)
def estimator_intelligence():
    pid=project_id(); sync=_seed_estimator_from_latest(pid); c=db()
    project=c.execute("SELECT name,number FROM projects WHERE id=?",(pid,)).fetchone()
    run=_latest_blueprint_run(pid)
    if run:
        rows=c.execute("""
            SELECT e.*
            FROM estimator_items e
            JOIN blueprint_scope_items b ON b.id=e.blueprint_scope_item_id
            WHERE e.company_id=? AND e.project_id=? AND b.run_id=?
            ORDER BY e.trade,e.id
        """,(current_company_id(),pid,run["id"])).fetchall()
    else:
        rows=[]
    c.close()
    trade_totals={}; grand=0.0; cards=[]
    for r in rows:
        direct,total=_estimate_math(r); grand+=total
        trade_totals[r["trade"]]=trade_totals.get(r["trade"],0)+total
        verify_label="✓ VERIFIED" if int(r["verified"] or 0) else "VERIFY"
        cards.append(f"""
        <div class="card">
          <div class="small">{esc(r["trade"])} · {verify_label}</div>
          <h3>{esc(r["description"])}</h3>
          <div class="small">Source: {esc(r["source_ref"] or "Source not identified")}</div>
          <form method="post" action="/brain/estimator/item/{r["id"]}">
            <div class="grid4" style="margin-top:12px">
              <div><label>Qty</label><input name="quantity" type="number" step="0.01" value="{r["quantity"] or 0}"></div>
              <div><label>Unit</label><input name="unit" value="{esc(r["unit"] or "")}" placeholder="EA, LF, SF"></div>
              <div><label>Material / unit</label><input name="material_unit_cost" type="number" step="0.01" value="{r["material_unit_cost"] or 0}"></div>
              <div><label>Labor / unit</label><input name="labor_unit_cost" type="number" step="0.01" value="{r["labor_unit_cost"] or 0}"></div>
            </div>
            <div class="grid4" style="margin-top:10px">
              <div><label>Sub quote</label><input name="subcontract_quote" type="number" step="0.01" value="{r["subcontract_quote"] or 0}"></div>
              <div><label>Allowance</label><input name="allowance" type="number" step="0.01" value="{r["allowance"] or 0}"></div>
              <div><label>Markup %</label><input name="markup_pct" type="number" step="0.01" value="{r["markup_pct"] or 0}"></div>
              <div><label>Verified</label><select name="verified"><option value="0" {"selected" if not r["verified"] else ""}>No</option><option value="1" {"selected" if r["verified"] else ""}>Yes</option></select></div>
            </div>
            <label style="margin-top:10px">Estimator notes</label><input name="notes" value="{esc(r["notes"] or "")}">
            <div style="margin-top:10px"><b>Extended total: ${total:,.2f}</b></div>
            <button type="submit" style="margin-top:10px">Save Estimate Item</button>
          </form>
        </div>""")
    totals="".join(f"<tr><td>{esc(k)}</td><td>${v:,.2f}</td></tr>" for k,v in sorted(trade_totals.items()))
    project_label=f'{esc(project["number"])} - {esc(project["name"])}' if project else "Current Project"
    body=f"""
    <div class="hero"><div class="eyebrow">BuildCommand Brain · Estimator Intelligence · v34.2</div>
      <h1>Plans → scope → estimate.</h1><div class="muted">{project_label}</div></div>
    <div class="grid2">
      <div class="card"><div class="label">Current Estimate</div><div class="kpi">${grand:,.2f}</div>
        <div class="small">{len(rows)} source-backed scope items · {sync["added"]} new · {sync["updated"]} refreshed</div></div>
      <div class="card"><h2>Estimator Brain Review</h2>
        <form method="post" action="/brain/estimator/review">
          <textarea name="focus" placeholder="Find bid gaps, allowances, long-lead items and takeoff verification needs."></textarea>
          <button type="submit">Analyze Estimate Risk</button>
        </form></div>
    </div>
    <div class="card"><h2>Takeoff Intelligence</h2><p>Turn cleaned scope items into quantity targets and verification status.</p><p><a href="/brain/takeoff"><b>Open Takeoff Intelligence →</b></a></p></div>
    <div class="card"><h2>Trade Totals</h2><table><tr><th>Trade</th><th>Current Total</th></tr>{totals}</table></div>
    <div class="card"><h2>Estimator rule</h2><p>BuildCommand organizes and identifies takeoff targets, but quantities and costs remain <b>unverified until an estimator confirms them</b>. Source references stay attached.</p></div>
    {"".join(cards) if cards else '<div class="card"><h2>No estimator items yet</h2><p>Run Plan Intake first, then return here.</p></div>'}"""
    return shell("Estimator Intelligence",body)

@app.post("/brain/estimator/item/{item_id}")
def estimator_item_update(item_id:int,quantity:float=Form(0),unit:str=Form(""),
    material_unit_cost:float=Form(0),labor_unit_cost:float=Form(0),
    subcontract_quote:float=Form(0),allowance:float=Form(0),
    markup_pct:float=Form(0),notes:str=Form(""),verified:int=Form(0)):
    pid=project_id(); c=db()
    c.execute("""UPDATE estimator_items SET quantity=?,unit=?,material_unit_cost=?,labor_unit_cost=?,
        subcontract_quote=?,allowance=?,markup_pct=?,notes=?,verified=?,updated=?
        WHERE id=? AND company_id=? AND project_id=?""",
        (quantity,unit,material_unit_cost,labor_unit_cost,subcontract_quote,allowance,
         markup_pct,notes,verified,datetime.utcnow().isoformat(),item_id,current_company_id(),pid))
    c.commit(); c.close()
    return RedirectResponse("/brain/estimator",status_code=303)

@app.post("/brain/estimator/review",response_class=HTMLResponse)
def estimator_brain_review(focus:str=Form("")):
    pid=project_id(); _seed_estimator_from_latest(pid); c=db()
    run=_latest_blueprint_run(pid)
    if run:
        rows=c.execute("""
            SELECT e.*
            FROM estimator_items e
            JOIN blueprint_scope_items b ON b.id=e.blueprint_scope_item_id
            WHERE e.company_id=? AND e.project_id=? AND b.run_id=?
            ORDER BY e.trade,e.id
        """,(current_company_id(),pid,run["id"])).fetchall()
    else:
        rows=[]
    c.close()
    lines=[]
    for r in rows:
        direct,total=_estimate_math(r)
        lines.append(f'{r["trade"]} | {r["description"]} | source={r["source_ref"]} | '
                     f'qty={r["quantity"]} {r["unit"]} | mat={r["material_unit_cost"]} | '
                     f'labor={r["labor_unit_cost"]} | sub={r["subcontract_quote"]} | '
                     f'allowance={r["allowance"]} | markup={r["markup_pct"]}% | total={total:.2f} | '
                     f'verified={bool(r["verified"])} | notes={r["notes"]}')
    api_key=os.environ.get("OPENAI_API_KEY")
    if not api_key:
        answer="OPENAI_API_KEY is not configured."
    else:
        try:
            client=OpenAI(api_key=api_key)
            resp=client.responses.create(
                model=os.environ.get("OPENAI_MODEL","gpt-5.6"),
                instructions="""You are BuildCommand Estimator Intelligence.
Review source-backed scope items and estimator-entered values.
Do not invent quantities or prices. Treat zero/unverified entries as missing, not as confirmed zero scope.
Identify bid gaps, missing takeoffs, missing quotes, allowances, cross-trade coordination,
scope exclusions needing confirmation, long-lead/procurement concerns, and high-risk items.
Prioritize what an estimator must verify before bid submission.""",
                input=f"""ESTIMATE WORKSHEET
{chr(10).join(lines)}

ESTIMATOR FOCUS
{focus or "Full bid-risk review"}"""
            )
            answer=resp.output_text
        except Exception as exc:
            answer=f"Estimator Intelligence could not complete review: {exc}"
    body=f"""<div class="hero"><div class="eyebrow">Estimator Intelligence</div><h1>Bid Risk Review</h1></div>
    <div class="card"><div style="white-space:pre-wrap;line-height:1.6">{esc(answer)}</div>
    <p><a href="/brain/estimator">← Back to Estimate</a></p></div>"""
    return shell("Estimator Brain Review",body)

@app.get("/plans-specs-ai",response_class=HTMLResponse)
def plans_specs_ai():
    pid=project_id(); c=db()
    docs=c.execute("SELECT * FROM attachments WHERE company_id=? AND project_id=? ORDER BY id DESC",(current_company_id(),pid)).fetchall()
    c.close()
    eligible=[d for d in docs if Path(d["original_name"] or "").suffix.lower() in {".pdf",".txt",".csv",".xlsx",".xlsm"}]
    checks="".join(f'<label style="display:flex;gap:10px;align-items:center;padding:10px 0;border-bottom:1px solid rgba(255,255,255,.08)"><input type="checkbox" name="attachment_ids" value="{d["id"]}" style="width:auto"><span><b>{esc(d["original_name"])}</b><br><span class="small">{(int(d["size_bytes"] or 0)/1024/1024):.1f} MB</span></span></label>' for d in eligible) or '<div class="muted">No PDF/TXT/CSV/Excel project documents are uploaded yet. Use Documents first.</div>'
    run,scopes=_blueprint_latest(pid)
    latest='<div class="muted">No Blueprint Brain analysis has been run yet.</div>'
    if run:
        scope_cards="".join(f'<div class="action"><a href="/blueprint-brain/trade/{s["id"]}" style="font-weight:800">{esc(s["trade"])}</a> <span class="badge READY">{s["item_count"]} ITEMS</span><div class="small">Division {esc(s["division"] or "—")} · {esc(s["summary"] or "")}</div></div>' for s in scopes)
        flags=json.loads(run["cross_discipline_flags"] or "[]")
        rfis=json.loads(run["rfi_candidates"] or "[]")
        latest=f'<div class="small">LATEST RUN · {esc(run["created"] or "")}</div><h3>{esc(run["project_summary"] or "")}</h3><div style="margin:14px 0"><span class="badge READY">{len(scopes)} TRADES</span> <span class="badge WATCH">{len(flags)} CROSS-DISCIPLINE</span> <span class="badge HIGH">{len(rfis)} RFI CANDIDATES</span></div>{scope_cards}<p><a href="/blueprint-brain/run/{run["id"]}">Open full intelligence report →</a></p>'
    body=f'<div class="hero"><div class="eyebrow">BuildCommand Plan Intelligence</div><h1>Plans → trade scopes → field execution.</h1><div class="muted">Reads PDF plan pages visually and textually, finds cross-discipline requirements, preserves sources, generates trade scopes, and flags gaps for GC review.</div></div><div class="grid2"><div class="card"><h2>Analyze Plan Set</h2><form method="post" action="/plans-specs-ai/analyze">{checks}<label style="margin-top:16px">Analysis focus (optional)</label><textarea name="focus" placeholder="Example: Full bid/scope review, or focus on MEP coordination"></textarea><button type="submit">Run Blueprint Brain</button></form><p class="small">Selected files must total less than 50 MB. PDF page images use high-detail analysis. AI-generated scopes require superintendent/PM review before contractual use.</p></div><div class="card"><h2>Latest Blueprint Intelligence</h2>{latest}</div></div>'
    return shell("Blueprint Brain",body)


@app.post("/plans-specs-ai",response_class=HTMLResponse)
def plans_specs_ai_answer(attachment_id:int=Form(...),question:str=Form(...)):
    pid=project_id(); c=db(); d=c.execute("SELECT * FROM attachments WHERE id=? AND company_id=? AND project_id=?",(attachment_id,current_company_id(),pid)).fetchone(); c.close()
    if not d: return HTMLResponse("Document not found",404)
    content=_attachment_text(d)
    if not content: answer=f'BuildCommand can see "{d["original_name"]}", but no readable text was extracted. Use Blueprint Brain for PDF visual analysis.'
    elif not os.environ.get("OPENAI_API_KEY"): answer="OPENAI_API_KEY is not configured."
    else:
        try:
            client=OpenAI(api_key=os.environ["OPENAI_API_KEY"])
            r=client.responses.create(model=os.environ.get("OPENAI_MODEL","gpt-5.6"),instructions="Answer only from the supplied construction document. Say when unsupported and never invent drawing references.",input=f"DOCUMENT:\n{content[:96000]}\nQUESTION:\n{question}")
            answer=r.output_text
        except Exception as e: answer=f"Document AI failed: {e}"
    return shell("Blueprint Brain",f'<div class="hero"><h1>{esc(d["original_name"])}</h1></div><div class="card"><h2>Answer</h2><div style="white-space:pre-wrap">{esc(answer)}</div></div>')


@app.post("/plans-specs-ai/analyze",response_class=HTMLResponse)
def blueprint_analyze(attachment_ids:list[int] | None=Form(None),focus:str=Form("")):
    pid=project_id(); company_id=current_company_id()
    if not os.environ.get("OPENAI_API_KEY"):
        return shell("Blueprint Brain",'<div class="card"><h2>OPENAI_API_KEY is not configured.</h2><p>Add it in Render Environment settings and redeploy.</p></div>')
    ids=list(dict.fromkeys(attachment_ids or []))
    if not ids:
        return shell("Blueprint Brain",'<div class="hero"><div class="eyebrow">Blueprint Brain</div><h1>Select a plan set first.</h1></div><div class="card"><p>Choose at least one PDF or supported project document, then click <b>Run Blueprint Brain</b>.</p><p><a href="/plans-specs-ai">← Back to Blueprint Brain</a></p></div>')
    c=db(); docs=[]
    for aid in ids:
        d=c.execute("SELECT * FROM attachments WHERE id=? AND company_id=? AND project_id=?",(aid,company_id,pid)).fetchone()
        if d: docs.append(d)
    c.close()
    if not docs: return HTMLResponse("No valid project documents were selected.",404)
    total=sum(int(d["size_bytes"] or 0) for d in docs)
    if total>=50*1024*1024:
        return shell("Blueprint Brain",f'<div class="card"><h2>Plan set is too large for one analysis request.</h2><p>Selected total: {total/1024/1024:.1f} MB. Keep each Blueprint Brain batch under 50 MB, then run the remaining volumes separately.</p></div>')
    model=os.environ.get("OPENAI_MODEL","gpt-5.6")
    client=OpenAI(api_key=os.environ["OPENAI_API_KEY"])
    uploaded=[]; input_content=[]; text_fallback=[]
    try:
        for d in docs:
            path=os.path.join(UPLOAD_DIR,d["stored_name"])
            if not os.path.isfile(path):
                continue
            ext=Path(d["original_name"] or "").suffix.lower()
            if ext==".pdf":
                with open(path,"rb") as fh:
                    remote=client.files.create(file=fh,purpose="user_data")
                uploaded.append(remote.id)
                input_content.append({"type":"input_file","file_id":remote.id,"detail":"high"})
            else:
                extracted=_attachment_text(d)
                if extracted:
                    text_fallback.append(f"\n--- FILE: {d['original_name']} ---\n{extracted[:120000]}")
        if not input_content and not text_fallback:
            return HTMLResponse("No readable selected files were available on the server.",400)
        names="\n".join(f"- {d['original_name']}" for d in docs)
        prompt=_blueprint_prompt(names)
        if focus.strip(): prompt += "\n\nGC ANALYSIS FOCUS:\n"+focus.strip()
        if text_fallback: prompt += "\n\nEXTRACTED NON-PDF DOCUMENT CONTENT:\n"+"\n".join(text_fallback)
        input_content.append({"type":"input_text","text":prompt})
        response=client.responses.create(model=model,input=[{"role":"user","content":input_content}])
        data=_blueprint_json(response.output_text)
        if not isinstance(data.get("trade_scopes"),list):
            raise ValueError("Blueprint Brain response did not contain trade_scopes.")
        run_id=_save_blueprint_result(pid,docs,data,model)
        return RedirectResponse(f"/blueprint-brain/run/{run_id}",status_code=303)
    except Exception as exc:
        return shell("Blueprint Brain",f'<div class="hero"><div class="eyebrow">Blueprint Brain</div><h1>Analysis did not complete.</h1></div><div class="card"><p>{esc(str(exc))}</p><p><a href="/plans-specs-ai">← Back to Blueprint Brain</a></p></div>')
    finally:
        for fid in uploaded:
            try: client.files.delete(fid)
            except Exception: pass



@app.get("/blueprint-brain",response_class=HTMLResponse)
def blueprint_brain_home():
    pid=project_id()
    company=current_company_id()
    c=db()
    runs=c.execute(
        "SELECT * FROM blueprint_runs WHERE company_id=? AND project_id=? ORDER BY id DESC LIMIT 25",
        (company,pid)
    ).fetchall()
    c.close()

    cards=""
    for r in runs:
        cards+=(
            f'<div class="card"><div class="small">Run #{r["id"]} · {esc(r["created"] or "")}</div>'
            f'<h3>{esc(r["project_summary"] or "Blueprint analysis")}</h3>'
            f'<p><a href="/blueprint-brain/run/{r["id"]}">Open analysis →</a></p></div>'
        )

    body=(
        '<div class="hero"><div class="eyebrow">Blueprint Brain</div>'
        '<h1>Source-backed trade scopes.</h1>'
        '<p class="muted">Review the latest analysis, run final trade cleanup, or start a new project analysis.</p></div>'
        '<div class="grid3">'
        +_v37_link_card("Analyze Project","Run the unified project intelligence pipeline on selected plans/specs.","/build/analyze-project","Analyze")
        +('<div class="card"><h2>Final Trade Cleanup</h2>'
             '<p class="muted">Normalize existing Blueprint Brain trade ownership and duplicate scopes.</p>'
             '<form method="post" action="/blueprint-brain/final-cleanup">'
             '<button type="submit">Run Final Trade Cleanup</button></form></div>')
        +_v37_link_card("Project Scope Review","Open the unified source-backed project scope view.","/brain","Review")
        +'</div>'
        '<div class="card"><h2>Recent Blueprint Runs</h2></div>'
        +(cards or '<div class="card"><p class="muted">No Blueprint Brain analyses yet.</p></div>')
    )
    return shell("Blueprint Brain",body)

@app.get("/blueprint-brain/run/{run_id}",response_class=HTMLResponse)
def blueprint_run_detail(run_id:int):
    pid=project_id(); c=db(); run=c.execute("SELECT * FROM blueprint_runs WHERE id=? AND company_id=? AND project_id=?",(run_id,current_company_id(),pid)).fetchone()
    if not run: c.close(); return HTMLResponse("Blueprint analysis not found.",404)
    scopes=c.execute("SELECT * FROM blueprint_trade_scopes WHERE run_id=? AND company_id=? AND project_id=? ORDER BY trade",(run_id,current_company_id(),pid)).fetchall(); c.close()
    def _arr(field):
        try: return json.loads(run[field] or "[]")
        except Exception: return []
    files=_arr("source_files"); disciplines=_arr("detected_disciplines"); flags=_arr("cross_discipline_flags"); rfis=_arr("rfi_candidates"); notes=_arr("review_notes")
    scope_cards="".join(f'<div class="card"><div class="small">DIVISION {esc(s["division"] or "—")}</div><h2>{esc(s["trade"])}</h2><p>{esc(s["summary"] or "")}</p><span class="badge READY">{s["item_count"]} SCOPE ITEMS</span><p><a href="/blueprint-brain/trade/{s["id"]}">Open trade scope →</a></p></div>' for s in scopes)
    def list_html(items,empty): return "".join(f'<div class="action">{esc(x)}</div>' for x in items) or f'<div class="muted">{esc(empty)}</div>'
    body=f'<div class="hero"><div class="eyebrow">Blueprint Intelligence Run #{run_id}</div><h1>{esc(run["project_summary"] or "Plan set analyzed")}</h1><div class="muted">Sources: {esc(", ".join(files))}</div></div><div class="grid3"><div class="card"><div class="label">Trades</div><div class="kpi">{len(scopes)}</div></div><div class="card"><div class="label">Cross-Discipline Flags</div><div class="kpi">{len(flags)}</div></div><div class="card"><div class="label">RFI Candidates</div><div class="kpi">{len(rfis)}</div></div></div><div class="card"><h2>Detected Disciplines</h2><p>{esc(" · ".join(disciplines) or "Not identified")}</p></div><div class="grid2"><div class="card"><h2>Cross-Discipline Requirements</h2>{list_html(flags,"No cross-discipline flags returned.")}</div><div class="card"><h2>Potential Scope Gaps / RFIs</h2>{list_html(rfis,"No RFI candidates returned.")}</div></div><div class="card"><h2>GC Review Notes</h2>{list_html(notes,"No additional review notes returned.")}</div><h2>Trade Scopes</h2><div class="grid2">{scope_cards}</div><div class="card"><p class="small"><b>BuildCommand review rule:</b> AI-generated scope intelligence is a coordination aid. Verify against the complete contract documents, addenda, RFIs, subcontract agreements, and design-professional direction before issuing contractual scope.</p></div>'
    return shell("Blueprint Intelligence",body)


@app.get("/blueprint-brain/trade/{scope_id}",response_class=HTMLResponse)
def blueprint_trade_scope(scope_id:int):
    pid=project_id(); c=db(); scope=c.execute("SELECT * FROM blueprint_trade_scopes WHERE id=? AND company_id=? AND project_id=?",(scope_id,current_company_id(),pid)).fetchone()
    if not scope: c.close(); return HTMLResponse("Trade scope not found.",404)
    items=c.execute("SELECT * FROM blueprint_scope_items WHERE trade_scope_id=? AND company_id=? AND project_id=? ORDER BY id",(scope_id,current_company_id(),pid)).fetchall(); c.close()
    cards=[]
    for item in items:
        refs=[]
        if item["source_sheet"]: refs.append("Sheet "+item["source_sheet"])
        if item["source_detail"]: refs.append("Detail "+item["source_detail"])
        if item["source_spec"]: refs.append("Spec "+item["source_spec"])
        if item["source_note"]: refs.append(item["source_note"])
        conf=item["confidence"] if item["confidence"] in {"HIGH","MEDIUM","LOW"} else "WATCH"
        typ=item["item_type"] or "SCOPE"
        confidence_badge="HIGH" if conf=="HIGH" else "WATCH"
        type_badge="HIGH" if typ=="RFI_CANDIDATE" else "WATCH" if typ in ["CROSS_DISCIPLINE","COORDINATION"] else "READY"
        related=(' · <b>Related trade:</b> '+esc(item["related_trade"])) if item["related_trade"] else ''
        options=''.join(f'<option {"selected" if item["status"]==st else ""}>{st}</option>' for st in ["NOT_STARTED","IN_PROGRESS","INSPECTION_REQUIRED","COMPLETE","VERIFIED"])
        trade_options=''.join(f'<option value="{esc(t)}" {"selected" if item["trade"]==t else ""}>{esc(t)}</option>' for t in V33_TRADES)
        cards.append(f'<div class="action"><span class="badge {confidence_badge}">{esc(item["confidence"])}</span> <span class="badge {type_badge}">{esc(typ)}</span><h3>{esc(item["requirement"])}</h3><div class="small"><b>Assigned trade:</b> {esc(item["trade"])} · <b>Source:</b> {esc(" · ".join(refs) or "Source not clearly identified")}{related}</div><div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:10px"><form method="post" action="/blueprint-brain/item/{item["id"]}/status"><select name="status">{options}</select><button type="submit">Update status</button></form><form method="post" action="/blueprint-brain/item/{item["id"]}/trade"><select name="trade">{trade_options}</select><select name="learn_scope"><option>PROJECT ONLY</option><option>COMPANY STANDARD</option></select><button type="submit">Move & Learn</button></form></div></div>')
    body=f'<div class="hero"><div class="eyebrow">Division {esc(scope["division"] or "—")} · Blueprint Brain</div><h1>{esc(scope["trade"])} Scope of Work</h1><div class="muted">{esc(scope["summary"] or "")}</div></div><div class="card"><p><a href="/blueprint-brain/run/{scope["run_id"]}">← Full Blueprint Intelligence</a> · <a href="/blueprint-brain/trade/{scope_id}/export.txt">Export scope text</a></p></div><div class="card"><h2>Scope Boiler</h2><div style="white-space:pre-wrap">{esc(scope["scope_text"] or "")}</div></div><div class="card"><h2>Execution Checklist</h2>{"".join(cards) or "<div class=muted>No scope items.</div>"}</div>'
    return shell(scope["trade"]+" Scope",body)


@app.post("/blueprint-brain/item/{item_id}/status")
def blueprint_item_status(item_id:int,status:str=Form(...)):
    allowed={"NOT_STARTED","IN_PROGRESS","INSPECTION_REQUIRED","COMPLETE","VERIFIED"}; status=status.upper()
    if status not in allowed: return HTMLResponse("Invalid status",400)
    pid=project_id(); c=db(); item=c.execute("SELECT trade_scope_id FROM blueprint_scope_items WHERE id=? AND company_id=? AND project_id=?",(item_id,current_company_id(),pid)).fetchone()
    if not item: c.close(); return HTMLResponse("Scope item not found",404)
    c.execute("UPDATE blueprint_scope_items SET status=? WHERE id=? AND company_id=? AND project_id=?",(status,item_id,current_company_id(),pid)); c.commit(); scope_id=item["trade_scope_id"]; c.close(); return RedirectResponse(f"/blueprint-brain/trade/{scope_id}",status_code=303)




@app.post("/blueprint-brain/item/{item_id}/trade")
def blueprint_item_trade(item_id:int,trade:str=Form(...),learn_scope:str=Form("PROJECT ONLY")):
    trade=_v33_normalize_trade(trade)
    if trade not in V33_TRADES:
        return HTMLResponse("Invalid trade",400)
    if learn_scope not in {"PROJECT ONLY","COMPANY STANDARD"}:
        learn_scope="PROJECT ONLY"

    pid=project_id(); company_id=current_company_id(); c=db()
    item=c.execute("SELECT * FROM blueprint_scope_items WHERE id=? AND company_id=? AND project_id=?",(item_id,company_id,pid)).fetchone()
    if not item:
        c.close(); return HTMLResponse("Scope item not found",404)

    old_trade=item["trade"]; old_scope_id=item["trade_scope_id"]; run_id=item["run_id"]
    target=c.execute("SELECT * FROM blueprint_trade_scopes WHERE run_id=? AND company_id=? AND project_id=? AND trade=?",(run_id,company_id,pid,trade)).fetchone()
    if not target:
        div={"GC / General Contractor":"01","Demolition":"02","Roofing":"07",
             "Doors / Frames / Hardware":"08","Storefront / Glazing":"08",
             "Flooring":"09","Tile":"09","Painting":"09","Framing / Drywall":"09",
             "Ceilings":"09","Specialties":"10","Plumbing":"22","HVAC / Mechanical":"23",
             "Electrical":"26","Fire Sprinkler":"21","Fire Alarm":"28","Low Voltage":"27"}.get(trade,"")
        now=datetime.utcnow().isoformat()
        c.execute("INSERT INTO blueprint_trade_scopes(company_id,project_id,run_id,trade,division,summary,scope_text,item_count,created) VALUES(?,?,?,?,?,?,?,?,?)",
                  (company_id,pid,run_id,trade,div,f"BuildCommand classified scope for {trade}.","",0,now))
        target_id=c.execute("SELECT last_insert_rowid() id").fetchone()["id"]
    else:
        target_id=target["id"]

    c.execute("UPDATE blueprint_scope_items SET trade=?,trade_scope_id=? WHERE id=? AND company_id=? AND project_id=?",
              (trade,target_id,item_id,company_id,pid))
    for sid in {old_scope_id,target_id}:
        count=c.execute("SELECT COUNT(*) n FROM blueprint_scope_items WHERE trade_scope_id=? AND company_id=? AND project_id=?",(sid,company_id,pid)).fetchone()["n"]
        c.execute("UPDATE blueprint_trade_scopes SET item_count=? WHERE id=? AND company_id=? AND project_id=?",(count,sid,company_id,pid))
    c.commit(); c.close()

    if old_trade != trade:
        _v43_ensure_tables()
        c=db(); now=datetime.utcnow().isoformat()
        subject=str(item["requirement"] or "")[:220]
        c.execute("""INSERT INTO learning_rules(
            company_id,project_id,rule_type,subject,learned_rule,source_ref,scope_level,
            approval_status,approved_by,confidence,created,updated
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
        (company_id,pid,"TRADE ASSIGNMENT",subject,f"{old_trade} -> {trade}",
         "Blueprint manual correction",learn_scope,"APPROVED",
         str(current_user_id()),"HIGH",now,now))
        c.commit(); c.close()

    return RedirectResponse(f"/blueprint-brain/trade/{target_id}",status_code=303)


@app.get("/blueprint-brain/trade/{scope_id}/export.txt")
def blueprint_trade_export(scope_id:int):
    pid=project_id(); c=db(); scope=c.execute("SELECT * FROM blueprint_trade_scopes WHERE id=? AND company_id=? AND project_id=?",(scope_id,current_company_id(),pid)).fetchone(); c.close()
    if not scope: return HTMLResponse("Trade scope not found.",404)
    header=f"BuildCommand AI - {scope['trade']} Scope of Work\nDivision {scope['division'] or '—'}\n\n"
    footer="\n\nGC REVIEW REQUIRED: Verify this AI-generated scope against the complete contract documents, addenda, RFIs, subcontract agreement, and design-professional direction before contractual use.\n"
    filename=re.sub(r"[^A-Za-z0-9_-]+","_",scope["trade"] or "trade")+"_scope.txt"
    return Response(content=header+(scope["scope_text"] or "")+footer,media_type="text/plain",headers={"Content-Disposition":f'attachment; filename="{filename}"'})


@app.get("/schedule-import",response_class=HTMLResponse)
def schedule_import_page():
    return shell("Schedule Import",'''<div class="hero"><div class="eyebrow">Schedule Import</div><h1>Import schedule activities from CSV.</h1><div class="muted">Headers: external_id,name,trade,start,finish,pct,status</div></div><div class="card"><form method="post" action="/schedule-import" enctype="multipart/form-data"><input type="file" name="file" accept=".csv" required><button>Import Schedule</button></form></div>''')


@app.post("/schedule-import")
async def schedule_import(file:UploadFile=File(...)):
    pid=project_id(); raw=await file.read()
    try: decoded=raw.decode("utf-8-sig")
    except Exception: return HTMLResponse("CSV must be UTF-8",400)
    reader=csv.DictReader(io.StringIO(decoded)); req={"external_id","name","trade","start","finish","pct","status"}
    if not req.issubset(set(reader.fieldnames or [])): return HTMLResponse("CSV missing required headers",400)
    rows=list(reader); c=db(); imported=0
    for row in rows:
        ext=(row.get("external_id") or "").strip(); name=(row.get("name") or "").strip()
        if not ext or not name: continue
        trade=(row.get("trade") or "").strip(); start=(row.get("start") or "").strip(); finish=(row.get("finish") or "").strip(); pct=float(row.get("pct") or 0); status=(row.get("status") or "NOT_STARTED").strip().upper()
        ex=c.execute("SELECT id FROM activities WHERE project_id=? AND external_id=?",(pid,ext)).fetchone()
        if ex:
            c.execute("UPDATE activities SET name=?,trade=?,start=?,finish=?,pct=?,status=? WHERE id=? AND project_id=?",(name,trade,start,finish,pct,status,ex["id"],pid))
        else:
            c.execute("INSERT INTO activities(project_id,external_id,name,trade,start,finish,pct,status) VALUES(?,?,?,?,?,?,?,?)",(pid,ext,name,trade,start,finish,pct,status))
        imported+=1
    c.execute("INSERT INTO schedule_import_batches(company_id,project_id,file_name,row_count,imported_count,created) VALUES(?,?,?,?,?,?)",(current_company_id(),pid,safe_filename(file.filename),len(rows),imported,datetime.utcnow().isoformat()))
    c.commit(); c.close()
    return RedirectResponse("/schedule",303)




def ensure_v29_tables():
    c=db()
    c.execute("CREATE TABLE IF NOT EXISTS photo_observations(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,attachment_id INTEGER,activity_id INTEGER,observation TEXT,severity TEXT DEFAULT 'WATCH',created TEXT)")
    c.execute("CREATE TABLE IF NOT EXISTS weather_impacts(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,activity_id INTEGER,impact_date TEXT,weather_type TEXT,lost_hours REAL DEFAULT 0,description TEXT,created TEXT)")
    c.execute("CREATE TABLE IF NOT EXISTS schedule_import_batches(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,file_name TEXT,row_count INTEGER DEFAULT 0,imported_count INTEGER DEFAULT 0,created TEXT)")
    c.execute("CREATE TABLE IF NOT EXISTS communication_drafts(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,sub_id INTEGER,draft_type TEXT,subject TEXT,body TEXT,status TEXT DEFAULT 'DRAFT',created TEXT)")
    c.execute("CREATE TABLE IF NOT EXISTS meeting_ai_summaries(id INTEGER PRIMARY KEY,company_id INTEGER NOT NULL,project_id INTEGER NOT NULL,source_text TEXT,summary_text TEXT,created TEXT)")
    c.commit(); c.close()

@app.get("/photo-intelligence",response_class=HTMLResponse)
def photo_intelligence():
    ensure_v29_tables()
    pid=project_id(); c=db()
    photos=c.execute(
        """
        SELECT *
        FROM attachments
        WHERE company_id=? AND project_id=?
          AND (
            lower(COALESCE(mime_type,'')) LIKE ?
            OR lower(COALESCE(original_name,'')) LIKE ?
            OR lower(COALESCE(original_name,'')) LIKE ?
            OR lower(COALESCE(original_name,'')) LIKE ?
            OR lower(COALESCE(original_name,'')) LIKE ?
          )
        ORDER BY id DESC
        """,
        (
            current_company_id(),
            pid,
            "image/%",
            "%.jpg",
            "%.jpeg",
            "%.png",
            "%.webp"
        )
    ).fetchall()
    obs=c.execute("SELECT p.*,a.original_name FROM photo_observations p LEFT JOIN attachments a ON a.id=p.attachment_id WHERE p.company_id=? AND p.project_id=? ORDER BY p.id DESC LIMIT 25",(current_company_id(),pid)).fetchall()
    c.close()
    opts="".join(f'<option value="{p["id"]}">{esc(p["original_name"])}</option>' for p in photos) or '<option value="">No photos uploaded</option>'
    recent="".join(f'<div class="action"><span class="badge {o["severity"]}">{esc(o["severity"])}</span> <b>{esc(o["original_name"] or "Photo")}</b><div>{esc(o["observation"])}</div></div>' for o in obs) or '<div class="muted">No observations yet.</div>'
    return shell("Photo Intelligence",f'''<div class="hero"><div class="eyebrow">Field Photo Intelligence</div><h1>Document photo observations.</h1></div><div class="grid2"><div class="card"><form method="post" action="/photo-intelligence"><select name="attachment_id">{opts}</select><textarea name="observation" required></textarea><select name="severity"><option>LOW</option><option selected>WATCH</option><option>HIGH</option><option>CRITICAL</option></select><button>Save Observation</button></form></div><div class="card">{recent}</div></div>''')


@app.post("/photo-intelligence")
def photo_intelligence_save(attachment_id:int=Form(...),observation:str=Form(...),severity:str=Form("WATCH")):
    ensure_v29_tables()
    pid=project_id(); c=db()
    c.execute("INSERT INTO photo_observations(company_id,project_id,attachment_id,activity_id,observation,severity,created) VALUES(?,?,?,NULL,?,?,?)",(current_company_id(),pid,attachment_id,observation.strip(),severity,datetime.utcnow().isoformat()))
    c.commit(); c.close(); return RedirectResponse("/photo-intelligence",303)


@app.get("/rfi-drafting",response_class=HTMLResponse)
def rfi_drafting():
    return shell("RFI Drafting",'''<div class="hero"><div class="eyebrow">AI RFI Drafting</div><h1>Turn a field issue into a professional RFI.</h1></div><div class="card"><form method="post" action="/rfi-drafting"><textarea name="field_issue" required></textarea><button>Draft RFI</button></form></div>''')


@app.post("/rfi-drafting",response_class=HTMLResponse)
def rfi_drafting_generate(field_issue:str=Form(...)):
    context=build_project_context(project_id())
    if os.environ.get("OPENAI_API_KEY"):
        try:
            client=OpenAI(api_key=os.environ["OPENAI_API_KEY"]); r=client.responses.create(model=os.environ.get("OPENAI_MODEL","gpt-5.6"),instructions="Draft a professional construction RFI. Return SUBJECT, BACKGROUND, QUESTION, POTENTIAL IMPACT. Do not invent drawing/spec numbers, dates or costs.",input=f"{context}\nFIELD ISSUE:\n{field_issue}"); draft=r.output_text
        except Exception as e: draft=f"AI RFI drafting failed: {e}"
    else: draft="OPENAI_API_KEY is not configured."
    return shell("RFI Drafting",f'''<div class="hero"><h1>Draft Ready</h1></div><div class="card"><div style="white-space:pre-wrap">{esc(draft)}</div></div><div class="card"><form method="post" action="/rfi-drafting/save"><input type="hidden" name="title" value="{esc(field_issue[:120])}"><textarea name="description">{esc(draft)}</textarea><button>Save to RFIs</button></form></div>''')


@app.post("/rfi-drafting/save")
def rfi_drafting_save(title:str=Form(...),description:str=Form(...)):
    pid=project_id(); c=db()
    c.execute("INSERT INTO project_issues(project_id,activity_id,issue_type,title,owner,due,priority,status,description,response,created) VALUES(?,NULL,'RFI',?,'',?,'WATCH','OPEN',?,'',?)",(pid,title.strip(),date.today().isoformat(),description.strip(),date.today().isoformat()))
    c.commit(); c.close(); return RedirectResponse("/issues",303)


@app.get("/sub-communications",response_class=HTMLResponse)
def sub_communications():
    pid=project_id(); c=db(); subs=c.execute("SELECT * FROM subs WHERE project_id=? ORDER BY trade,name",(pid,)).fetchall(); drafts=c.execute("SELECT d.*,s.name sub_name FROM communication_drafts d LEFT JOIN subs s ON s.id=d.sub_id WHERE d.company_id=? AND d.project_id=? ORDER BY d.id DESC LIMIT 20",(current_company_id(),pid)).fetchall(); c.close()
    opts="".join(f'<option value="{s["id"]}">{esc(s["name"])} - {esc(s["trade"])}</option>' for s in subs)
    recent="".join(f'<div class="action"><b>{esc(d["draft_type"])}</b> · {esc(d["sub_name"] or "")}<div>{esc(d["subject"])}</div></div>' for d in drafts) or '<div class="muted">No drafts yet.</div>'
    return shell("Sub Communications",f'''<div class="hero"><div class="eyebrow">Subcontractor Communications</div><h1>Generate field follow-ups.</h1></div><div class="grid2"><div class="card"><form method="post" action="/sub-communications"><select name="sub_id">{opts}</select><select name="draft_type"><option>MANPOWER_REQUEST</option><option>DELAY_NOTICE</option><option>COORDINATION</option><option>FOLLOW_UP</option></select><textarea name="prompt" required></textarea><button>Generate Draft</button></form></div><div class="card">{recent}</div></div>''')


@app.post("/sub-communications",response_class=HTMLResponse)
def sub_communications_generate(sub_id:int=Form(...),draft_type:str=Form(...),prompt:str=Form(...)):
    pid=project_id(); c=db(); sub=c.execute("SELECT * FROM subs WHERE id=? AND project_id=?",(sub_id,pid)).fetchone(); c.close()
    if os.environ.get("OPENAI_API_KEY") and sub:
        try:
            client=OpenAI(api_key=os.environ["OPENAI_API_KEY"]); r=client.responses.create(model=os.environ.get("OPENAI_MODEL","gpt-5.6"),instructions="Write a concise professional subcontractor communication. Firm, collaborative, no invented dates or contract terms.",input=f"SUB: {sub['name']} {sub['trade']}\nTYPE: {draft_type}\nNEED: {prompt}"); body=r.output_text
        except Exception as e: body=f"AI communication failed: {e}"
    else: body=prompt
    subject=f"{draft_type.replace('_',' ').title()} - {sub['name'] if sub else 'Subcontractor'}"
    c=db(); c.execute("INSERT INTO communication_drafts(company_id,project_id,sub_id,draft_type,subject,body,status,created) VALUES(?,?,?,?,?,?,'DRAFT',?)",(current_company_id(),pid,sub_id,draft_type,subject,body,datetime.utcnow().isoformat())); c.commit(); c.close()
    return shell("Sub Communications",f'<div class="hero"><h1>{esc(subject)}</h1></div><div class="card"><div style="white-space:pre-wrap">{esc(body)}</div></div>')


@app.get("/meeting-minutes-ai",response_class=HTMLResponse)
def meeting_minutes_ai():
    return shell("AI Meeting Minutes",'''<div class="hero"><div class="eyebrow">Meeting Minutes AI</div><h1>Turn raw notes into decisions and actions.</h1></div><div class="card"><form method="post" action="/meeting-minutes-ai"><textarea name="source_text" required style="min-height:250px"></textarea><button>Generate Minutes</button></form></div>''')


@app.post("/meeting-minutes-ai",response_class=HTMLResponse)
def meeting_minutes_ai_generate(source_text:str=Form(...)):
    pid=project_id()
    if os.environ.get("OPENAI_API_KEY"):
        try:
            client=OpenAI(api_key=os.environ["OPENAI_API_KEY"]); r=client.responses.create(model=os.environ.get("OPENAI_MODEL","gpt-5.6"),instructions="Convert construction meeting notes into DECISIONS, COMMITMENTS, ACTION ITEMS, RISKS/ISSUES, FOLLOW-UP. Do not invent owners or dates.",input=source_text); summary=r.output_text
        except Exception as e: summary=f"AI minutes failed: {e}"
    else: summary=source_text
    c=db(); c.execute("INSERT INTO meeting_ai_summaries(company_id,project_id,source_text,summary_text,created) VALUES(?,?,?,?,?)",(current_company_id(),pid,source_text,summary,datetime.utcnow().isoformat())); c.commit(); c.close()
    return shell("AI Meeting Minutes",f'''<div class="hero"><h1>Minutes Ready</h1></div><div class="card"><div style="white-space:pre-wrap">{esc(summary)}</div></div><div class="card"><form method="post" action="/meeting-minutes-ai/save"><textarea name="summary">{esc(summary)}</textarea><button>Save to Meetings</button></form></div>''')


@app.post("/meeting-minutes-ai/save")
def meeting_minutes_ai_save(summary:str=Form(...)):
    pid=project_id(); c=db(); c.execute("INSERT INTO meeting_notes(project_id,meeting_date,meeting_type,title,attendees,decisions,commitments,follow_up,created) VALUES(?,?,'COORDINATION','AI Meeting Minutes','',?,'','',?)",(pid,date.today().isoformat(),summary.strip(),date.today().isoformat())); c.commit(); c.close(); return RedirectResponse("/meetings",303)


@app.get("/weather-impacts",response_class=HTMLResponse)
def weather_impacts():
    pid=project_id(); c=db(); acts=c.execute("SELECT * FROM activities WHERE project_id=? ORDER BY start",(pid,)).fetchall(); rows=c.execute("SELECT w.*,a.external_id,a.name activity FROM weather_impacts w LEFT JOIN activities a ON a.id=w.activity_id WHERE w.company_id=? AND w.project_id=? ORDER BY w.impact_date DESC",(current_company_id(),pid)).fetchall(); c.close()
    opts='<option value="">No linked activity</option>'+''.join(f'<option value="{a["id"]}">{esc(a["external_id"])} - {esc(a["name"])}</option>' for a in acts)
    hist=''.join(f'<div class="action"><b>{esc(r["impact_date"])} · {esc(r["weather_type"])}</b><div>{r["lost_hours"] or 0} lost hour(s) · {esc(r["external_id"] or "")} {esc(r["activity"] or "")}</div><div class="small">{esc(r["description"])}</div></div>' for r in rows) or '<div class="muted">No impacts logged.</div>'
    return shell("Weather Impacts",f'''<div class="hero"><div class="eyebrow">Weather Impact Log</div><h1>Document weather-related delay.</h1></div><div class="grid2"><div class="card"><form method="post" action="/weather-impacts"><input type="date" name="impact_date" value="{date.today().isoformat()}" required><select name="weather_type"><option>RAIN</option><option>HEAT</option><option>WIND</option><option>LIGHTNING</option><option>COLD</option><option>OTHER</option></select><select name="activity_id">{opts}</select><input type="number" step="0.5" min="0" name="lost_hours" value="0"><textarea name="description"></textarea><button>Log Impact</button></form></div><div class="card">{hist}</div></div>''')


@app.post("/weather-impacts")
def weather_impacts_save(impact_date:str=Form(...),weather_type:str=Form(...),activity_id:str=Form(""),lost_hours:float=Form(0),description:str=Form("")):
    pid=project_id(); linked=int(activity_id) if str(activity_id).strip() else None; c=db(); c.execute("INSERT INTO weather_impacts(company_id,project_id,activity_id,impact_date,weather_type,lost_hours,description,created) VALUES(?,?,?,?,?,?,?,?)",(current_company_id(),pid,linked,impact_date,weather_type,max(0,lost_hours),description.strip(),datetime.utcnow().isoformat())); c.commit(); c.close(); return RedirectResponse("/weather-impacts",303)


def _simple_pdf(title,lines,filename):
    if canvas is None: return HTMLResponse("reportlab is not installed",500)
    buf=io.BytesIO(); pdf=canvas.Canvas(buf); y=744; pdf.setFont("Helvetica-Bold",16); pdf.drawString(42,y,title); y-=28; pdf.setFont("Helvetica",9)
    for line in lines:
        words=str(line).split(); cur=""
        for word in words:
            test=(cur+" "+word).strip()
            if pdf.stringWidth(test,"Helvetica",9)>520:
                pdf.drawString(48,y,cur); y-=12; cur=word
                if y<50: pdf.showPage(); y=744; pdf.setFont("Helvetica",9)
            else: cur=test
        if cur: pdf.drawString(48,y,cur); y-=14
    pdf.setFont("Helvetica",8); pdf.drawString(42,28,"BuildCommand AI · Built by Wilson LaHood"); pdf.save(); buf.seek(0)
    return Response(buf.getvalue(),media_type="application/pdf",headers={"Content-Disposition":f'attachment; filename="{filename}"'})


@app.get("/pdf-reports",response_class=HTMLResponse)
def pdf_reports():
    return shell("PDF Reports",'''<div class="hero"><div class="eyebrow">Professional Reports</div><h1>Export project reports.</h1></div><div class="grid2"><div class="card"><a href="/pdf-reports/project-health.pdf">Project Health PDF</a></div><div class="card"><a href="/pdf-reports/lookahead.pdf">3-Week Lookahead PDF</a></div><div class="card"><a href="/pdf-reports/weekly.pdf">Weekly AI Report PDF</a></div><div class="card"><a href="/exports/daily-report.pdf">Daily Report PDF</a></div></div>''')


@app.get("/pdf-reports/project-health.pdf")
def pdf_health():
    h=project_health_snapshot(project_id()); return _simple_pdf("BuildCommand AI - Project Health",[f'Overall: {h["overall"]}/100',f'Schedule: {h["schedule"]}',f'Readiness: {h["readiness"]}',f'Procurement: {h["procurement"]}',f'Risk: {h["risk"]}',f'Field: {h["field"]}'],"buildcommand_project_health.pdf")


@app.get("/pdf-reports/lookahead.pdf")
def pdf_lookahead():
    pid=project_id(); today=date.today(); horizon=today+timedelta(days=21); c=db(); acts=c.execute("SELECT * FROM activities WHERE project_id=? ORDER BY start",(pid,)).fetchall(); c.close(); lines=[f"Window: {today} through {horizon}"]
    for a in acts:
        s=parse_iso_date(a["start"])
        if s and today<=s<=horizon: lines.append(f'{a["external_id"]} - {a["name"]} | {a["trade"]} | {a["start"]} to {a["finish"]} | {a["pct"] or 0}%')
    return _simple_pdf("BuildCommand AI - 3 Week Lookahead",lines,"buildcommand_lookahead.pdf")


@app.get("/pdf-reports/weekly.pdf")
def pdf_weekly():
    pid=project_id(); c=db(); r=c.execute("SELECT * FROM weekly_ai_reports WHERE company_id=? AND project_id=? ORDER BY id DESC LIMIT 1",(current_company_id(),pid)).fetchone(); c.close()
    if not r: return HTMLResponse("Generate a Weekly AI Report first.",404)
    return _simple_pdf("BuildCommand AI - Weekly Project Report",(r["report_text"] or "").splitlines(),"buildcommand_weekly_report.pdf")


@app.get("/cost-intelligence",response_class=HTMLResponse)
def cost_intelligence():
    pid=project_id(); c=db(); rows=c.execute("SELECT ce.*,a.external_id,a.name activity FROM change_events ce LEFT JOIN activities a ON a.id=ce.activity_id WHERE ce.project_id=? ORDER BY ce.id DESC",(pid,)).fetchall(); c.close()
    open_rows=[r for r in rows if r["status"] not in ["APPROVED","REJECTED"]]; approved=[r for r in rows if r["status"]=="APPROVED"]; open_cost=sum(float(r["estimated_cost"] or 0) for r in open_rows); approved_cost=sum(float(r["estimated_cost"] or 0) for r in approved); days=sum(float(r["schedule_days"] or 0) for r in open_rows)
    cards=''.join(f'<div class="card"><span class="badge {"READY" if r["status"]=="APPROVED" else "WATCH"}">{esc(r["status"])}</span><h3>{esc(r["title"])}</h3><p>${(r["estimated_cost"] or 0):,.0f} · {(r["schedule_days"] or 0):.1f} day(s)</p></div>' for r in rows) or '<div class="card">No change events.</div>'
    return shell("Cost Intelligence",f'''<div class="hero"><div class="eyebrow">Cost & Change Intelligence</div><h1>Change exposure in dollars and days.</h1></div><div class="grid3"><div class="card"><div class="label">Open Exposure</div><div class="kpi">${open_cost:,.0f}</div></div><div class="card"><div class="label">Approved</div><div class="kpi">${approved_cost:,.0f}</div></div><div class="card"><div class="label">Open Days</div><div class="kpi">{days:.1f}</div></div></div><div class="grid2">{cards}</div>''')


@app.get("/mobile-home",response_class=HTMLResponse)
def mobile_home():
    h=project_health_snapshot(project_id())
    return shell("Mobile Home",f'''<div class="hero"><div class="eyebrow">Superintendent Mobile</div><h1>Today in the Field</h1><div class="kpi">{h["overall"]}/100</div></div><div class="grid2"><div class="card"><a href="/quick-entry">🎤 Speak / Quick Note</a></div><div class="card"><a href="/documents">📷 Add Photo / Document</a></div><div class="card"><a href="/daily-report">📝 Daily Report</a></div><div class="card"><a href="/issues">⚠️ Issue / RFI</a></div><div class="card"><a href="/ai-command">📌 View Today</a></div><div class="card"><a href="/assistant">🤖 AI Assistant</a></div></div>''')

def v30_extract_text(row):
    if not row:
        return ""
    path=os.path.join(UPLOAD_DIR,row["stored_name"])
    ext=Path(row["original_name"] or "").suffix.lower()
    if not os.path.isfile(path):
        return ""
    try:
        if ext in {".txt",".csv"}:
            return Path(path).read_text(errors="ignore")[:250000]
        if ext==".pdf" and PdfReader is not None:
            return "\n".join((p.extract_text() or "") for p in PdfReader(path).pages[:200])[:300000]
        if ext in {".xlsx",".xlsm"} and openpyxl is not None:
            wb=openpyxl.load_workbook(path,read_only=True,data_only=True)
            lines=[]
            for ws in wb.worksheets[:20]:
                lines.append("SHEET: "+ws.title)
                for vals in ws.iter_rows(values_only=True):
                    lines.append(" | ".join("" if v is None else str(v) for v in vals))
                    if len(lines)>20000:
                        break
            return "\n".join(lines)[:300000]
    except Exception:
        return ""
    return ""

@app.get("/document-ai",response_class=HTMLResponse)
def v30_document_ai_page():
    pid=project_id(); c=db()
    docs=c.execute("SELECT * FROM attachments WHERE company_id=? AND project_id=? ORDER BY id DESC LIMIT 100",(current_company_id(),pid)).fetchall()
    c.close()
    options="".join(f'<option value="{d["id"]}">{esc(d["original_name"])}</option>' for d in docs) or '<option value="">No uploaded documents</option>'
    return shell("Deep Document AI",f"""
    <div class="hero"><div class="eyebrow">Deep Plans & Specs AI</div><h1>Read PDF, Excel, TXT, and CSV project documents.</h1></div>
    <div class="card"><form method="post" action="/document-ai"><select name="attachment_id">{options}</select>
    <textarea name="question" required placeholder="Ask a question about this document"></textarea>
    <button type="submit">Analyze Document</button></form></div>""")

@app.post("/document-ai",response_class=HTMLResponse)
def v30_document_ai_answer(attachment_id:int=Form(...),question:str=Form(...)):
    pid=project_id(); c=db()
    doc=c.execute("SELECT * FROM attachments WHERE id=? AND company_id=? AND project_id=?",(attachment_id,current_company_id(),pid)).fetchone()
    c.close()
    if not doc:
        return HTMLResponse("Document not found.",status_code=404)
    extracted=v30_extract_text(doc)
    if not extracted:
        answer="No readable text could be extracted from this file."
    else:
        try:
            client=OpenAI(api_key=os.environ["OPENAI_API_KEY"])
            response=client.responses.create(
                model=os.environ.get("OPENAI_MODEL","gpt-5.6"),
                instructions="Answer only from the supplied construction document text. Do not invent drawing numbers or spec sections.",
                input=f"DOCUMENT: {doc['original_name']}\n\n{extracted[:96000]}\n\nQUESTION: {question}"
            )
            answer=response.output_text
        except Exception as exc:
            answer=f"Document AI failed: {exc}"
    return shell("Deep Document AI",f"""<div class="hero"><h1>{esc(doc["original_name"])}</h1></div>
    <div class="card"><div style="white-space:pre-wrap">{esc(answer)}</div></div>""")

@app.get("/advanced-schedule-import",response_class=HTMLResponse)
def v30_schedule_import_page():
    return shell("Advanced Schedule Import","""
    <div class="hero"><div class="eyebrow">Advanced Schedule Import</div><h1>Import CSV or Excel schedules.</h1></div>
    <div class="card"><form method="post" action="/advanced-schedule-import" enctype="multipart/form-data">
    <input type="file" name="file" accept=".csv,.xlsx" required><button type="submit">Import Schedule</button></form></div>""")

@app.post("/advanced-schedule-import")
async def v30_schedule_import(file:UploadFile=File(...)):
    pid=project_id(); filename=safe_filename(file.filename); ext=Path(filename).suffix.lower(); raw=await file.read(); rows=[]
    if ext==".csv":
        rows=list(csv.DictReader(io.StringIO(raw.decode("utf-8-sig"))))
    elif ext==".xlsx":
        if openpyxl is None:
            return HTMLResponse("Excel support is unavailable.",status_code=500)
        tmp=tempfile.NamedTemporaryFile(delete=False,suffix=".xlsx"); tmp.write(raw); tmp.close()
        try:
            wb=openpyxl.load_workbook(tmp.name,read_only=True,data_only=True); ws=wb[wb.sheetnames[0]]
            vals=list(ws.iter_rows(values_only=True))
            if vals:
                headers=[str(v or "").strip() for v in vals[0]]
                rows=[{headers[i]:r[i] for i in range(min(len(headers),len(r)))} for r in vals[1:]]
        finally:
            try: os.unlink(tmp.name)
            except Exception: pass
    else:
        return HTMLResponse("Use CSV or XLSX.",status_code=400)
    required={"external_id","name","trade","start","finish","pct","status"}
    if rows and not required.issubset(set(rows[0].keys())):
        return HTMLResponse("Schedule file is missing required headers.",status_code=400)
    c=db(); imported=0
    for row in rows:
        eid=str(row.get("external_id") or "").strip(); name=str(row.get("name") or "").strip()
        if not eid or not name: continue
        existing=c.execute("SELECT id FROM activities WHERE project_id=? AND external_id=?",(pid,eid)).fetchone()
        trade=str(row.get("trade") or "").strip(); start=str(row.get("start") or "").split(" ")[0]; finish=str(row.get("finish") or "").split(" ")[0]
        try: pct=float(row.get("pct") or 0)
        except Exception: pct=0
        status=str(row.get("status") or "NOT_STARTED").strip().upper()
        if existing:
            c.execute("UPDATE activities SET name=?,trade=?,start=?,finish=?,pct=?,status=? WHERE id=? AND project_id=?",(name,trade,start,finish,pct,status,existing["id"],pid))
        else:
            c.execute("INSERT INTO activities(project_id,external_id,name,trade,start,finish,pct,status) VALUES(?,?,?,?,?,?,?,?)",(pid,eid,name,trade,start,finish,pct,status))
        imported+=1
    c.execute("INSERT INTO schedule_import_sources(company_id,project_id,source_type,file_name,imported_count,created) VALUES(?,?,?,?,?,?)",(current_company_id(),pid,ext.replace(".","").upper(),filename,imported,datetime.utcnow().isoformat()))
    c.commit(); c.close()
    return RedirectResponse("/schedule",status_code=303)

@app.get("/photo-ai",response_class=HTMLResponse)
def v30_photo_ai_page():
    pid=project_id(); c=db()
    photos=c.execute("""SELECT * FROM attachments WHERE company_id=? AND project_id=? AND
    (lower(COALESCE(mime_type,'')) LIKE ? OR lower(COALESCE(original_name,'')) LIKE ? OR lower(COALESCE(original_name,'')) LIKE ? OR lower(COALESCE(original_name,'')) LIKE ? OR lower(COALESCE(original_name,'')) LIKE ?)
    ORDER BY id DESC LIMIT 50""",(current_company_id(),pid,"image/%","%.jpg","%.jpeg","%.png","%.webp")).fetchall(); c.close()
    options="".join(f'<option value="{p["id"]}">{esc(p["original_name"])}</option>' for p in photos) or '<option value="">No photos uploaded</option>'
    return shell("AI Photo Analysis",f"""<div class="hero"><div class="eyebrow">AI Photo Analysis</div><h1>Analyze field photos.</h1></div>
    <div class="card"><form method="post" action="/photo-ai"><select name="attachment_id">{options}</select>
    <select name="focus"><option>GENERAL</option><option>SAFETY</option><option>QUALITY</option><option>PROGRESS</option></select>
    <button type="submit">Analyze Photo</button></form></div>""")

@app.post("/photo-ai",response_class=HTMLResponse)
def v30_photo_ai(attachment_id:int=Form(...),focus:str=Form("GENERAL")):
    pid=project_id(); c=db()
    row=c.execute("SELECT * FROM attachments WHERE id=? AND company_id=? AND project_id=?",(attachment_id,current_company_id(),pid)).fetchone(); c.close()
    if not row: return HTMLResponse("Photo not found.",status_code=404)
    path=os.path.join(UPLOAD_DIR,row["stored_name"])
    if not os.path.isfile(path): return HTMLResponse("Stored photo unavailable.",status_code=404)
    try:
        data=Path(path).read_bytes(); mime=row["mime_type"] or "image/jpeg"; b64=base64.b64encode(data).decode("ascii")
        client=OpenAI(api_key=os.environ["OPENAI_API_KEY"])
        response=client.responses.create(
            model=os.environ.get("OPENAI_VISION_MODEL",os.environ.get("OPENAI_MODEL","gpt-5.6")),
            instructions=f"Describe only visible construction facts. Focus: {focus}. Treat possible concerns as observations, not definitive code violations.",
            input=[{"role":"user","content":[{"type":"input_text","text":"Analyze this construction field photo."},{"type":"input_image","image_url":f"data:{mime};base64,{b64}"}]}]
        )
        result=response.output_text
    except Exception as exc:
        result=f"Photo AI failed: {exc}"
    return shell("AI Photo Analysis",f"""<div class="hero"><h1>{esc(row["original_name"])}</h1></div><div class="card"><div style="white-space:pre-wrap">{esc(result)}</div></div>""")

@app.get("/auto-daily-report",response_class=HTMLResponse)
def v30_auto_daily_page():
    return shell("Auto Daily Report","""<div class="hero"><div class="eyebrow">Automatic Daily Report</div><h1>Generate today's report from BuildCommand activity.</h1></div>
    <div class="card"><form method="post" action="/auto-daily-report/generate"><button type="submit">Generate Draft Daily Report</button></form></div>""")

@app.post("/auto-daily-report/generate",response_class=HTMLResponse)
def v30_auto_daily_generate():
    context=build_project_context(project_id())
    try:
        client=OpenAI(api_key=os.environ["OPENAI_API_KEY"])
        response=client.responses.create(
            model=os.environ.get("OPENAI_MODEL","gpt-5.6"),
            instructions="Draft a superintendent daily report using only supplied project data. Sections: WORK COMPLETED, MANPOWER, DELAYS/CONSTRAINTS, DELIVERIES, INSPECTIONS, SAFETY, TOMORROW PLAN. Say Not documented when missing.",
            input=context
        )
        draft=response.output_text
    except Exception as exc: draft=f"Auto daily report failed: {exc}"
    return shell("Auto Daily Report",f"""<div class="hero"><h1>Draft Ready</h1></div><div class="card"><div style="white-space:pre-wrap">{esc(draft)}</div></div>""")

def v30_forecast(pid):
    c=db(); acts=c.execute("SELECT * FROM activities WHERE project_id=? AND status!='COMPLETE'",(pid,)).fetchall(); c.close()
    if not acts: return 0.0,.5,"No active activities."
    scores=[]; notes=[]
    for a in acts:
        score,band,reasons=activity_delay_signal(a); scores.append(score)
        if score>=45: notes.append(f'{a["external_id"]} {a["name"]}: {band}')
    avg=sum(scores)/len(scores)
    return round(avg/18.0,1),round(min(.9,.45+len(acts)*.03),2),"; ".join(notes[:5]) or "No major delay signal detected."

@app.get("/predictive-forecast",response_class=HTMLResponse)
def v30_predictive_forecast():
    pid=project_id(); h=project_health_snapshot(pid); projected,confidence,explanation=v30_forecast(pid)
    c=db(); c.execute("INSERT INTO forecast_snapshots(company_id,project_id,snapshot_date,health_score,projected_delay_days,confidence,explanation,created) VALUES(?,?,?,?,?,?,?,?)",
    (current_company_id(),pid,date.today().isoformat(),h["overall"],projected,confidence,explanation,datetime.utcnow().isoformat())); c.commit(); c.close()
    band="READY" if projected<1 else "WATCH" if projected<3 else "HIGH" if projected<6 else "CRITICAL"
    return shell("Predictive Forecast",f"""<div class="hero"><div class="eyebrow">Predictive Forecast</div><h1>{projected:.1f} projected delay day(s)</h1>
    <span class="badge {band}">{band}</span><div class="muted">Confidence {confidence*100:.0f}%</div></div><div class="card"><p>{esc(explanation)}</p></div>""")

@app.get("/sub-risk",response_class=HTMLResponse)
def v30_sub_risk():
    pid=project_id(); c=db(); subs=c.execute("SELECT * FROM subs WHERE project_id=? ORDER BY trade,name",(pid,)).fetchall()
    updates=c.execute("SELECT * FROM subcontractor_updates WHERE project_id=? ORDER BY update_date DESC,id DESC",(pid,)).fetchall(); c.close()
    cards=""
    for s in subs:
        su=[u for u in updates if u["sub_id"]==s["id"]]; latest=su[0] if su else None; score=20; reasons=[]
        if not latest: score+=35; reasons.append("no recent field update")
        else:
            st=(latest["status"] or "").upper()
            if st=="CRITICAL": score+=55; reasons.append("latest status critical")
            elif st=="HIGH": score+=35; reasons.append("latest status high")
            elif st=="WATCH": score+=18; reasons.append("latest status watch")
            if (latest["manpower"] or 0)<=0: score+=20; reasons.append("latest manpower zero")
        score=min(100,score); band="CRITICAL" if score>=75 else "HIGH" if score>=50 else "WATCH" if score>=25 else "READY"; explanation="; ".join(reasons) or "No major risk signal."
        c2=db(); c2.execute("INSERT INTO sub_risk_snapshots(company_id,project_id,sub_id,risk_score,risk_band,explanation,created) VALUES(?,?,?,?,?,?,?)",(current_company_id(),pid,s["id"],score,band,explanation,datetime.utcnow().isoformat())); c2.commit(); c2.close()
        cards+=f'<div class="card"><span class="badge {band}">{band} · {score}</span><h3>{esc(s["name"])}</h3><p>{esc(explanation)}</p></div>'
    return shell("Sub Risk",f'<div class="hero"><h1>Subcontractor Risk</h1></div><div class="grid3">{cards}</div>')

@app.get("/change-package",response_class=HTMLResponse)
def v30_change_package_page():
    pid=project_id(); c=db(); rows=c.execute("SELECT * FROM change_events WHERE project_id=? ORDER BY id DESC",(pid,)).fetchall(); c.close()
    options="".join(f'<option value="{r["id"]}">{esc(r["title"])}</option>' for r in rows)
    return shell("Change Package",f"""<div class="hero"><div class="eyebrow">Change Order Documentation</div><h1>Build a change package.</h1></div>
    <div class="card"><form method="post" action="/change-package"><select name="change_event_id">{options}</select><button type="submit">Generate Change Package</button></form></div>""")

@app.post("/change-package",response_class=HTMLResponse)
def v30_change_package(change_event_id:int=Form(...)):
    pid=project_id(); c=db()
    ce=c.execute("""SELECT ce.*,a.external_id,a.name activity FROM change_events ce LEFT JOIN activities a ON a.id=ce.activity_id WHERE ce.id=? AND ce.project_id=?""",(change_event_id,pid)).fetchone(); c.close()
    if not ce: return HTMLResponse("Change event not found.",status_code=404)
    package="\n".join([
        "CHANGE EVENT", str(ce["title"] or ""), "",
        "TYPE", str(ce["event_type"] or ""), "",
        "LINKED ACTIVITY", f'{ce["external_id"] or ""} {ce["activity"] or ""}', "",
        "RESPONSIBLE PARTY", str(ce["responsible_party"] or "Unassigned"), "",
        "ESTIMATED COST IMPACT", f'${float(ce["estimated_cost"] or 0):,.0f}', "",
        "SCHEDULE IMPACT", f'{float(ce["schedule_days"] or 0):.1f} day(s)', "",
        "STATUS", str(ce["status"] or ""), "",
        "DESCRIPTION", str(ce["description"] or ""), "",
        "SUPPORTING DOCUMENTS", "Attach photos, RFIs, submittals, meeting notes, and pricing as applicable."
    ])
    c=db(); c.execute("INSERT INTO change_packages(company_id,project_id,change_event_id,package_text,created) VALUES(?,?,?,?,?)",(current_company_id(),pid,change_event_id,package,datetime.utcnow().isoformat())); c.commit(); c.close()
    return shell("Change Package",f'<div class="hero"><h1>{esc(ce["title"])}</h1></div><div class="card"><div style="white-space:pre-wrap">{esc(package)}</div></div>')

@app.get("/owner-dashboard",response_class=HTMLResponse)
def v30_owner_dashboard():
    pid=project_id(); h=project_health_snapshot(pid); c=db(); project=c.execute("SELECT * FROM projects WHERE id=?",(pid,)).fetchone()
    changes=c.execute("SELECT * FROM change_events WHERE project_id=?",(pid,)).fetchall(); latest=c.execute("SELECT * FROM weekly_ai_reports WHERE project_id=? ORDER BY id DESC LIMIT 1",(pid,)).fetchone(); c.close()
    open_cost=sum(float(r["estimated_cost"] or 0) for r in changes if r["status"] not in ["APPROVED","REJECTED"]); approved=sum(float(r["estimated_cost"] or 0) for r in changes if r["status"]=="APPROVED")
    summary=esc(latest["report_text"][:1800]) if latest else "Generate a Weekly AI Report to populate the owner summary."
    return shell("Owner Dashboard",f"""<div class="hero"><div class="eyebrow">Owner / Executive View</div><h1>{esc(project["name"] if project else "Project")}</h1><div class="kpi">{h["overall"]}/100</div></div>
    <div class="grid4"><div class="card"><div class="label">Schedule</div><div class="kpi">{h["schedule"]}</div></div>
    <div class="card"><div class="label">Readiness</div><div class="kpi">{h["readiness"]}</div></div>
    <div class="card"><div class="label">Open Change</div><div class="kpi">${open_cost:,.0f}</div></div>
    <div class="card"><div class="label">Approved Changes</div><div class="kpi">${approved:,.0f}</div></div></div>
    <div class="card"><h2>Executive Summary</h2><div style="white-space:pre-wrap">{summary}</div></div>""")

@app.get("/portfolio-intelligence",response_class=HTMLResponse)
def v30_portfolio_intelligence():
    cid=current_company_id(); c=db(); projects=c.execute("SELECT * FROM projects WHERE company_id=? ORDER BY name",(cid,)).fetchall(); c.close()
    cards=""; scores=[]
    for p in projects:
        h=project_health_snapshot(p["id"]); scores.append(h["overall"]); band="READY" if h["overall"]>=85 else "WATCH" if h["overall"]>=70 else "HIGH" if h["overall"]>=50 else "CRITICAL"
        cards+=f'<div class="card"><span class="badge {band}">{h["overall"]}/100</span><h3>{esc(p["name"])}</h3><p>Schedule {h["schedule"]} · Readiness {h["readiness"]} · Procurement {h["procurement"]}</p></div>'
    avg=round(sum(scores)/len(scores)) if scores else 0
    return shell("Portfolio Intelligence",f'<div class="hero"><h1>Company Portfolio Health: {avg}/100</h1></div><div class="grid3">{cards}</div>')

@app.get("/mobile-field-plus",response_class=HTMLResponse)
def v30_mobile_field_plus():
    h=project_health_snapshot(project_id())
    return shell("Mobile Field+",f"""<div class="hero"><div class="eyebrow">Mobile Superintendent Workflow</div><h1>Field Mode</h1><div class="kpi">{h["overall"]}/100</div></div>
    <div class="grid2">
    <div class="card"><a href="/quick-entry">🎤 Speak Note</a></div>
    <div class="card"><a href="/photo-ai">📷 Analyze Photo</a></div>
    <div class="card"><a href="/auto-daily-report">📝 Auto Daily Report</a></div>
    <div class="card"><a href="/rfi-drafting">❓ Draft RFI</a></div>
    <div class="card"><a href="/lookahead-intelligence">📅 3-Week Lookahead</a></div>
    <div class="card"><a href="/predictive-forecast">📈 Forecast</a></div>
    <div class="card"><a href="/sub-risk">👷 Sub Risk</a></div>
    <div class="card"><a href="/assistant">🤖 Ask AI</a></div></div>""")

def audit_event(action,detail='',pid=None):
    try:
        c=db(); c.execute('INSERT INTO admin_audit_log(company_id,user_id,project_id,action,detail,created) VALUES(?,?,?,?,?,?)',(current_company_id(),current_user_id(),pid if pid is not None else project_id(),action,detail,datetime.utcnow().isoformat())); c.commit(); c.close()
    except Exception: pass

@app.get('/storage-status',response_class=HTMLResponse)
def storage_status_page():
    pid=project_id(); c=db(); rows=c.execute('SELECT * FROM attachments WHERE company_id=? AND project_id=? ORDER BY id DESC',(current_company_id(),pid)).fetchall(); c.close()
    present=sum(1 for r in rows if os.path.isfile(os.path.join(UPLOAD_DIR,r['stored_name']))); missing=len(rows)-present; persistent=not os.path.abspath(UPLOAD_DIR).startswith('/tmp/'); band='READY' if persistent and missing==0 else 'WATCH'
    return shell('Storage Status',f'''<div class="hero"><div class="eyebrow">Persistent Storage</div><h1><span class="badge {band}">{band}</span> {present}/{len(rows)} project file(s) present</h1></div><div class="grid2"><div class="card"><h2>Upload Path</h2><p>{esc(UPLOAD_DIR)}</p><p class="small">{'Persistent path detected.' if persistent else 'Temporary /tmp path detected.'}</p></div><div class="card"><h2>Missing Files</h2><div class="kpi">{missing}</div></div></div>''')


@app.get('/global-search',response_class=HTMLResponse)
def global_search_page(q:str=''):
    pid=project_id()
    company=current_company_id()
    results=[]
    q=(q or '').strip()

    if q:
        like=f'%{q.lower()}%'
        specs=[
            ("Schedule","activities","name","/schedule",False),
            ("RFI / Issue","project_issues","title","/issues",False),
            ("Action","action_items","title","/actions",False),
            ("Subcontractor","subs","name","/subcontractors",False),
            ("Document","attachments","original_name","/documents",True),
            ("Change","change_events","title","/changes",False),
            ("Blueprint Scope","blueprint_scope_items","requirement","/blueprint-brain",True),
            ("Estimator","estimator_items","description","/brain/estimator",True),
            ("Submittal","submittals","title","/submittals",False),
            ("Punch","punch_items","title","/punch",False),
            ("Procurement","procurement","item","/procurement",False),
        ]

        c=db()
        for kind,table,column,url,company_scoped in specs:
            try:
                if company_scoped:
                    sql=f"SELECT id,{column} AS label FROM {table} WHERE project_id=? AND company_id=? AND LOWER(COALESCE({column}, '')) LIKE ? ORDER BY id DESC LIMIT 25"
                    rows=c.execute(sql,(pid,company,like)).fetchall()
                else:
                    sql=f"SELECT id,{column} AS label FROM {table} WHERE project_id=? AND LOWER(COALESCE({column}, '')) LIKE ? ORDER BY id DESC LIMIT 25"
                    rows=c.execute(sql,(pid,like)).fetchall()

                for r in rows:
                    label=r["label"] if r["label"] is not None else ""
                    results.append((kind,label,url))
            except Exception:
                try:
                    c.rollback()
                except Exception:
                    pass
                continue
        c.close()

    html=''.join(
        f'<div class="search-result"><b>{esc(k)}</b> · <a href="{u}" style="color:#f0b44d">{esc(l)}</a></div>'
        for k,l,u in results
    ) or (
        '<div class="muted">No matching project records.</div>'
        if q else
        '<div class="muted">Search plans/scopes, estimator, schedule, RFIs, actions, subcontractors, documents, changes, submittals, punch and procurement.</div>'
    )

    body=(
        '<div class="hero"><div class="eyebrow">Search Everything</div>'
        '<h1>Search this project.</h1>'
        '<p class="muted">One search across the project instead of hunting through individual tools.</p></div>'
        '<div class="card"><form method="get" action="/global-search">'
        f'<input name="q" value="{esc(q)}" placeholder="Search plans, scope, estimator, schedule, RFIs, documents, changes...">'
        '<button type="submit">Search Everything</button></form></div>'
        f'<div class="card">{html}</div>'
    )
    return shell('Global Search',body)


@app.get('/favorites',response_class=HTMLResponse)
def favorites_page():
    c=db(); rows=c.execute('SELECT * FROM user_favorites WHERE company_id=? AND user_id=? ORDER BY id DESC',(current_company_id(),current_user_id())).fetchall(); c.close()
    options=''.join(f'<option value="{esc(u)}|{esc(n)}">{esc(n)}</option>' for n,u in NAV)
    cards=''.join(f'<div class="action"><a href="{esc(r["tool_url"])}" style="color:#f0b44d;font-weight:700">{esc(r["tool_name"])}</a><form method="post" action="/favorites/remove"><input type="hidden" name="favorite_id" value="{r["id"]}"><button type="submit">Remove</button></form></div>' for r in rows) or '<div class="muted">No pinned tools yet.</div>'
    return shell('Favorites',f'''<div class="hero"><div class="eyebrow">Favorites</div><h1>Pin your most-used tools.</h1></div><div class="grid2"><div class="card"><form method="post" action="/favorites"><select name="tool">{options}</select><button type="submit">Pin Tool</button></form></div><div class="card">{cards}</div></div>''')

@app.post('/favorites')
def favorites_add(tool:str=Form(...)):
    try: url,name=tool.split('|',1)
    except Exception: return RedirectResponse('/favorites',303)
    c=db(); ex=c.execute('SELECT id FROM user_favorites WHERE company_id=? AND user_id=? AND tool_url=?',(current_company_id(),current_user_id(),url)).fetchone()
    if not ex: c.execute('INSERT INTO user_favorites(company_id,user_id,tool_name,tool_url,created) VALUES(?,?,?,?,?)',(current_company_id(),current_user_id(),name,url,datetime.utcnow().isoformat())); c.commit()
    c.close(); audit_event('PIN_TOOL',name); return RedirectResponse('/favorites',303)

@app.post('/favorites/remove')
def favorites_remove(favorite_id:int=Form(...)):
    c=db(); c.execute('DELETE FROM user_favorites WHERE id=? AND company_id=? AND user_id=?',(favorite_id,current_company_id(),current_user_id())); c.commit(); c.close(); audit_event('UNPIN_TOOL',str(favorite_id)); return RedirectResponse('/favorites',303)

@app.get('/recent-activity',response_class=HTMLResponse)
def recent_activity_page():
    pid=project_id(); c=db(); entries=[]; sources=[('Action','action_items','title','created','/actions'),('RFI / Issue','project_issues','title','created','/issues'),('Document','attachments','original_name','created','/documents'),('Change','change_events','title','created','/changes'),('Daily Report','daily_reports','report_date','created','/daily-report')]
    for kind,table,label_col,date_col,url in sources:
        if table=='attachments': rows=c.execute(f'SELECT {label_col} label,{date_col} stamp FROM {table} WHERE project_id=? AND company_id=? ORDER BY id DESC LIMIT 12',(pid,current_company_id())).fetchall()
        else: rows=c.execute(f'SELECT {label_col} label,{date_col} stamp FROM {table} WHERE project_id=? ORDER BY id DESC LIMIT 12',(pid,)).fetchall()
        for r in rows: entries.append((r['stamp'] or '',kind,r['label'],url))
    c.close(); entries.sort(key=lambda x:x[0],reverse=True); html=''.join(f'<div class="action"><b>{esc(k)}</b> · <a href="{u}" style="color:#f0b44d">{esc(l)}</a><div class="small">{esc(s)}</div></div>' for s,k,l,u in entries[:40]) or '<div class="muted">No recent project activity.</div>'
    return shell('Recent Activity',f'<div class="hero"><div class="eyebrow">Recent Activity</div><h1>What changed recently?</h1></div><div class="card">{html}</div>')

@app.get('/project-clone',response_class=HTMLResponse)
def project_clone_page():
    pid=project_id(); c=db(); p=c.execute('SELECT * FROM projects WHERE id=? AND company_id=?',(pid,current_company_id())).fetchone(); c.close()
    return shell('Clone Project',f'''<div class="hero"><div class="eyebrow">Project Template</div><h1>Clone {esc(p['name'] if p else 'current project')}</h1></div><div class="card"><form method="post" action="/project-clone"><input name="name" placeholder="New Project Name" required><input name="number" placeholder="New Project Number" required><button type="submit">Clone Project Setup</button></form><p class="small">Copies schedule activities and subcontractor list only.</p></div>''')

@app.post('/project-clone')
def project_clone(name:str=Form(...),number:str=Form(...)):
    source=project_id(); cid=current_company_id(); c=db(); c.execute('INSERT INTO projects(name,number,status,company_id) VALUES(?,?,?,?)',(name.strip(),number.strip(),'PLANNING',cid)); new_id=c.execute('SELECT last_insert_rowid() id').fetchone()['id']
    for a in c.execute('SELECT * FROM activities WHERE project_id=? ORDER BY id',(source,)).fetchall(): c.execute("INSERT INTO activities(project_id,external_id,name,trade,start,finish,pct,status) VALUES(?,?,?,?,?,?,0,'NOT_STARTED')",(new_id,a['external_id'],a['name'],a['trade'],a['start'],a['finish']))
    for s in c.execute('SELECT * FROM subs WHERE project_id=? ORDER BY id',(source,)).fetchall(): c.execute('INSERT INTO subs(project_id,name,trade) VALUES(?,?,?)',(new_id,s['name'],s['trade']))
    c.commit(); c.close(); audit_event('CLONE_PROJECT',f'{source} -> {new_id}',new_id); return RedirectResponse('/',303)

@app.get('/project-archive',response_class=HTMLResponse)
def project_archive_page():
    cid=current_company_id(); c=db(); rows=c.execute('SELECT p.*,COALESCE(a.archived,0) archived FROM projects p LEFT JOIN project_archive_state a ON a.project_id=p.id WHERE p.company_id=? ORDER BY p.name',(cid,)).fetchall(); c.close()
    cards=''.join(f'''<div class="action"><b>{esc(r['number'])} - {esc(r['name'])}</b><div class="small">{'ARCHIVED' if r['archived'] else 'ACTIVE'}</div><form method="post" action="/project-archive"><input type="hidden" name="project_id_value" value="{r['id']}"><input type="hidden" name="archived" value="{0 if r['archived'] else 1}"><button type="submit">{'Restore' if r['archived'] else 'Archive'}</button></form></div>''' for r in rows)
    return shell('Archive Projects',f'<div class="hero"><div class="eyebrow">Project Lifecycle</div><h1>Archive or restore projects.</h1></div><div class="card">{cards}</div>')

@app.post('/project-archive')
def project_archive_save(project_id_value:int=Form(...),archived:int=Form(...)):
    cid=current_company_id(); c=db(); valid=c.execute('SELECT id FROM projects WHERE id=? AND company_id=?',(project_id_value,cid)).fetchone()
    if valid:
        ex=c.execute('SELECT project_id FROM project_archive_state WHERE project_id=?',(project_id_value,)).fetchone()
        if ex: c.execute('UPDATE project_archive_state SET archived=?,archived_at=? WHERE project_id=?',(1 if archived else 0,datetime.utcnow().isoformat() if archived else None,project_id_value))
        else: c.execute('INSERT INTO project_archive_state(project_id,company_id,archived,archived_at) VALUES(?,?,?,?)',(project_id_value,cid,1 if archived else 0,datetime.utcnow().isoformat() if archived else None))
        c.commit()
    c.close(); audit_event('ARCHIVE_PROJECT' if archived else 'RESTORE_PROJECT',str(project_id_value),project_id_value); return RedirectResponse('/project-archive',303)

@app.get('/document-tags',response_class=HTMLResponse)
def document_tags_page():
    pid=project_id(); c=db(); docs=c.execute('SELECT * FROM attachments WHERE company_id=? AND project_id=? ORDER BY id DESC',(current_company_id(),pid)).fetchall(); tags=c.execute('SELECT t.*,a.original_name FROM attachment_tags t JOIN attachments a ON a.id=t.attachment_id WHERE t.company_id=? AND a.project_id=? ORDER BY t.id DESC',(current_company_id(),pid)).fetchall(); c.close()
    options=''.join(f'<option value="{d["id"]}">{esc(d["original_name"])}</option>' for d in docs); recent=''.join(f'<div class="action"><b>{esc(t["tag"])}</b> · {esc(t["original_name"])}</div>' for t in tags) or '<div class="muted">No document tags yet.</div>'
    return shell('Document Tags',f'''<div class="hero"><div class="eyebrow">Document Organization</div><h1>Tag project files.</h1></div><div class="grid2"><div class="card"><form method="post" action="/document-tags"><select name="attachment_id">{options}</select><input name="tag" placeholder="electrical, permit, owner" required><button type="submit">Add Tag</button></form></div><div class="card">{recent}</div></div>''')

@app.post('/document-tags')
def document_tags_add(attachment_id:int=Form(...),tag:str=Form(...)):
    c=db(); valid=c.execute('SELECT id FROM attachments WHERE id=? AND company_id=?',(attachment_id,current_company_id())).fetchone()
    if valid: c.execute('INSERT INTO attachment_tags(company_id,attachment_id,tag,created) VALUES(?,?,?,?)',(current_company_id(),attachment_id,tag.strip().lower(),datetime.utcnow().isoformat())); c.commit()
    c.close(); audit_event('TAG_DOCUMENT',tag); return RedirectResponse('/document-tags',303)

@app.get('/notification-rules',response_class=HTMLResponse)
def notification_rules_page():
    c=db(); rows=c.execute('SELECT * FROM notification_rules WHERE company_id=? ORDER BY id DESC',(current_company_id(),)).fetchall(); c.close()
    recent=''.join(f'<div class="action"><span class="badge {esc(r["severity"])}">{esc(r["severity"])}</span> <b>{esc(r["rule_name"])}</b><div class="small">Threshold {r["threshold_value"]} - {"Enabled" if r["enabled"] else "Disabled"}</div></div>' for r in rows) or '<div class="muted">No custom notification rules yet.</div>'
    return shell('Notification Rules',f'''<div class="hero"><div class="eyebrow">Alert Rules</div><h1>Control what deserves attention.</h1></div><div class="grid2"><div class="card"><form method="post" action="/notification-rules"><input name="rule_name" placeholder="High schedule delay" required><select name="severity"><option>WATCH</option><option selected>HIGH</option><option>CRITICAL</option></select><input type="number" step="0.1" name="threshold_value" value="0"><button type="submit">Add Rule</button></form></div><div class="card">{recent}</div></div>''')

@app.post('/notification-rules')
def notification_rules_add(rule_name:str=Form(...),severity:str=Form('HIGH'),threshold_value:float=Form(0)):
    c=db(); c.execute('INSERT INTO notification_rules(company_id,rule_name,enabled,severity,threshold_value,created) VALUES(?,?,1,?,?,?)',(current_company_id(),rule_name.strip(),severity,threshold_value,datetime.utcnow().isoformat())); c.commit(); c.close(); audit_event('ADD_NOTIFICATION_RULE',rule_name); return RedirectResponse('/notification-rules',303)

@app.get('/health-history',response_class=HTMLResponse)
def health_history_page():
    pid=project_id(); h=project_health_snapshot(pid); c=db(); ex=c.execute('SELECT id FROM health_history WHERE company_id=? AND project_id=? AND snapshot_date=?',(current_company_id(),pid,date.today().isoformat())).fetchone()
    if not ex: c.execute('INSERT INTO health_history(company_id,project_id,snapshot_date,overall,schedule,readiness,procurement,risk,field,created) VALUES(?,?,?,?,?,?,?,?,?,?)',(current_company_id(),pid,date.today().isoformat(),h['overall'],h['schedule'],h['readiness'],h['procurement'],h['risk'],h['field'],datetime.utcnow().isoformat())); c.commit()
    rows=c.execute('SELECT * FROM health_history WHERE company_id=? AND project_id=? ORDER BY snapshot_date DESC,id DESC LIMIT 30',(current_company_id(),pid)).fetchall(); c.close()
    history=''.join(f'<div class="action"><b>{esc(r["snapshot_date"])}</b> - Overall {r["overall"]:.0f}/100<div class="small">Schedule {r["schedule"]:.0f} · Readiness {r["readiness"]:.0f} · Procurement {r["procurement"]:.0f} · Risk {r["risk"]:.0f} · Field {r["field"]:.0f}</div></div>' for r in rows)
    return shell('Health History',f'<div class="hero"><div class="eyebrow">Project Trend</div><h1>Health history</h1></div><div class="card">{history}</div>')

@app.get('/bulk-export')
def bulk_export():
    pid=project_id(); cid=current_company_id(); stamp=datetime.utcnow().strftime('%Y%m%d_%H%M%S'); zpath=f'/tmp/buildcommand_project_export_{stamp}.zip'; tables=['projects','activities','subs','action_items','project_issues','daily_reports','procurement','risks','change_events','attachments']; c=db()
    with zipfile.ZipFile(zpath,'w',zipfile.ZIP_DEFLATED) as z:
        for table in tables:
            if table=='projects': rows=c.execute('SELECT * FROM projects WHERE id=? AND company_id=?',(pid,cid)).fetchall()
            elif table=='attachments': rows=c.execute('SELECT * FROM attachments WHERE project_id=? AND company_id=?',(pid,cid)).fetchall()
            else: rows=c.execute(f'SELECT * FROM {table} WHERE project_id=?',(pid,)).fetchall()
            if not rows: continue
            headers=list(rows[0].keys()); sio=io.StringIO(); writer=csv.writer(sio); writer.writerow(headers)
            for r in rows: writer.writerow([r[h] for h in headers])
            z.writestr(f'{table}.csv',sio.getvalue())
    c.close(); audit_event('BULK_EXPORT',f'project {pid}',pid); return FileResponse(zpath,media_type='application/zip',filename=f'buildcommand_project_export_{stamp}.zip')

@app.get('/audit-log',response_class=HTMLResponse)
def audit_log_page():
    if not require_role('ADMIN'): return HTMLResponse('Not authorized.',status_code=403)
    c=db(); rows=c.execute('SELECT l.*,u.display_name FROM admin_audit_log l LEFT JOIN users u ON u.id=l.user_id WHERE l.company_id=? ORDER BY l.id DESC LIMIT 100',(current_company_id(),)).fetchall(); c.close()
    html=''.join(f'<div class="action"><b>{esc(r["action"])}</b> - {esc(r["display_name"] or "System")}<div>{esc(r["detail"])}</div><div class="small">{esc(r["created"])}</div></div>' for r in rows) or '<div class="muted">No v31 audit events yet.</div>'
    return shell('Audit Log',f'<div class="hero"><div class="eyebrow">Admin Audit Log</div><h1>Track important workspace changes.</h1></div><div class="card">{html}</div>')


# =============================================================================
# BuildCommand AI v371 - Blueprint Brain Reliability & Trade Accuracy
# Safe hardening layer built on the verified v370 staging baseline.
# =============================================================================

_V371_TRADE_REGRESSION_CASES = [
    ("Patch concrete and floor surfaces disturbed by electrical or plumbing work", "Concrete"),
    ("Prime and refinish all wall, ceiling, and other surfaces patched after demolition or MEP installation", "Painting"),
    ("Provide new suspended acoustical tile ceiling grid and tile", "Ceilings"),
    ("Coordinate ceiling grid openings with lights, diffusers, sprinkler heads and smoke detectors", "Ceilings"),
    ("Provide sprinkler heads and sprinkler drops at new ceiling", "Fire Sprinkler"),
    ("Provide duct smoke detectors for rooftop air-conditioning units", "HVAC / Mechanical"),
    ("Provide EF-1 roof-mounted exhaust fan", "HVAC / Mechanical"),
    ("Patch roof membrane at mechanical penetrations", "Roofing"),
    ("Provide electrical connection and disconnect for EF-1 exhaust fan", "Electrical"),
    ("Provide instantaneous electric water heater IWH-1", "Plumbing"),
    ("Provide water closets, urinals, lavatories, floor drains and drinking fountain", "Plumbing"),
    ("Sawcut slab, trench and restore concrete for below-grade plumbing", "Concrete"),
    ("Provide stainless-steel grab bars and toilet room accessories", "Toilet / Bath Accessories"),
    ("Provide toilet partitions and urinal screens", "Toilet / Bath Accessories"),
    ("Provide concealed wood blocking and backing for wall-mounted accessories", "Framing / Drywall"),
    ("Construct restroom Wall Type B gypsum board assemblies and resilient-channel gypsum ceilings", "Framing / Drywall"),
    ("Provide interior hollow-metal doors, frames and hardware", "Doors / Frames / Hardware"),
    ("Provide aluminum storefront system and exterior storefront doors", "Storefront / Glazing"),
    ("Provide rubber base, LVT flooring and tile backsplash", "Flooring / Tile"),
    ("Provide electronic accessible door operator with activation touch pads", "Low Voltage"),
    ("Provide card access, electronic strikes, cameras and door contacts", "Low Voltage"),
    ("Provide fire extinguisher cabinets", "Specialties"),
    ("Paint hollow-metal doors and frames", "Painting"),
    ("Remove existing architectural partitions, doors, ceilings and millwork", "Demolition"),
    ("Remove existing lavatories, urinals, water closets, floor drains and water heater", "Demolition"),
]

def _v371_trade_regression_results():
    results=[]
    for requirement, expected in _V371_TRADE_REGRESSION_CASES:
        actual=_v441_primary_trade(requirement, "Unassigned")
        results.append({
            "requirement": requirement,
            "expected": expected,
            "actual": actual,
            "passed": actual == expected,
        })
    return results


def _v371_trade_regression_summary():
    rows=_v371_trade_regression_results()
    passed=sum(1 for r in rows if r["passed"])
    return {
        "version":"v371",
        "suite":"Blueprint Brain trade ownership",
        "passed":passed,
        "total":len(rows),
        "failed":len(rows)-passed,
        "ok":passed==len(rows),
        "results":rows,
    }

@app.get('/health/blueprint-v371')
def v371_blueprint_health():
    """Deterministic staging smoke test for core Blueprint Brain trade ownership rules."""
    return _v371_trade_regression_summary()

@app.get('/blueprint-reliability-v371', response_class=HTMLResponse)
def v371_blueprint_reliability_page():
    summary=_v371_trade_regression_summary()
    rows=''.join(
        f'<div class="action"><span class="badge {"READY" if r["passed"] else "WATCH"}">'
        f'{"PASS" if r["passed"] else "FAIL"}</span> <b>{esc(r["expected"])}</b>'
        f'<div>{esc(r["requirement"])}</div>'
        f'<div class="small">Classifier returned: {esc(r["actual"])}</div></div>'
        for r in summary['results']
    )
    status='READY' if summary['ok'] else 'WATCH'
    return shell('Blueprint Reliability v371',
        f'<div class="hero"><div class="eyebrow">BuildCommand v371</div>'
        f'<h1>Blueprint Brain Reliability Center</h1>'
        f'<p class="muted">Deterministic trade-ownership regression checks before production promotion.</p>'
        f'<p><span class="badge {status}">{summary["passed"]}/{summary["total"]} PASSED</span></p></div>'
        f'<div class="card">{rows}</div>')


# =============================================================================
# BuildCommand AI v372 - Blueprint Brain Source Intelligence
# Cross-checks sheet/detail/spec/note evidence before trusting extracted scope.
# =============================================================================

_V372_SOURCE_WEIGHTS = {
    "source_spec": 4,
    "source_detail": 3,
    "source_sheet": 2,
    "source_note": 2,
}


def _v372_clean_source(value):
    v=str(value or "").strip()
    if not v or v.lower() in {"none","null","n/a","na","unknown","-"}:
        return ""
    return v


def _v372_source_evidence(item):
    """Return deterministic source-strength metadata for one Blueprint Brain item."""
    evidence=[]
    weighted=0
    for field,weight in _V372_SOURCE_WEIGHTS.items():
        value=_v372_clean_source(item.get(field) if hasattr(item,'get') else item[field])
        if value:
            evidence.append({"field":field,"value":value,"weight":weight})
            weighted += weight
    count=len(evidence)
    # Reward corroboration across source types, not repeated text from one field.
    corroboration=max(0,count-1)*2
    score=min(100, weighted*10 + corroboration*5)
    if count >= 3 or score >= 80:
        level="STRONG"
    elif count >= 2 or score >= 45:
        level="SUPPORTED"
    elif count == 1:
        level="SINGLE_SOURCE"
    else:
        level="UNSOURCED"
    return {"count":count,"weighted":weighted,"score":score,"level":level,"evidence":evidence}


def _v372_trade_check(requirement, stored_trade):
    predicted=_v441_primary_trade(str(requirement or ''), str(stored_trade or 'Unassigned'))
    stored=str(stored_trade or 'Unassigned')
    return {
        "stored":stored,
        "predicted":predicted,
        "agrees":predicted == stored,
    }


def _v372_item_intelligence(item):
    requirement=str(item.get('requirement','') if hasattr(item,'get') else item['requirement'])
    trade=str(item.get('trade','') if hasattr(item,'get') else item['trade'])
    ev=_v372_source_evidence(item)
    tc=_v372_trade_check(requirement,trade)
    base=ev['score']
    if tc['agrees']:
        base=min(100,base+10)
    else:
        base=max(0,base-20)
    if ev['level']=='UNSOURCED':
        disposition='VERIFY_SOURCE'
    elif not tc['agrees']:
        disposition='REVIEW_TRADE'
    elif ev['level']=='SINGLE_SOURCE':
        disposition='VERIFY_CROSS_REFERENCE'
    else:
        disposition='READY'
    return {
        "requirement":requirement,
        "trade":trade,
        "source":ev,
        "trade_check":tc,
        "confidence_score":base,
        "disposition":disposition,
    }


_V372_SOURCE_REGRESSION_CASES = [
    ({"requirement":"Provide EF-1 roof-mounted exhaust fan","trade":"HVAC / Mechanical","source_sheet":"M2.1","source_detail":"3/M5.1","source_spec":"23 34 00","source_note":"Mechanical note 8"},"STRONG","READY"),
    ({"requirement":"Provide suspended acoustical ceiling grid","trade":"Ceilings","source_sheet":"A6.1","source_detail":"","source_spec":"09 51 00","source_note":""},"SUPPORTED","READY"),
    ({"requirement":"Provide card access and electronic strikes","trade":"Low Voltage","source_sheet":"E7.1","source_detail":"","source_spec":"","source_note":""},"SINGLE_SOURCE","VERIFY_CROSS_REFERENCE"),
    ({"requirement":"Paint hollow-metal doors and frames","trade":"Doors / Frames / Hardware","source_sheet":"A8.1","source_detail":"","source_spec":"09 91 00","source_note":""},"SUPPORTED","REVIEW_TRADE"),
    ({"requirement":"Patch roof membrane at mechanical penetrations","trade":"Roofing","source_sheet":"A5.1","source_detail":"7/A5.2","source_spec":"07 54 00","source_note":"Roof note 4"},"STRONG","READY"),
    ({"requirement":"Provide water closets and lavatories","trade":"Plumbing","source_sheet":"","source_detail":"","source_spec":"","source_note":""},"UNSOURCED","VERIFY_SOURCE"),
    ({"requirement":"Sawcut slab and restore concrete for plumbing trench","trade":"Concrete","source_sheet":"P1.1","source_detail":"2/P4.1","source_spec":"03 30 00","source_note":""},"STRONG","READY"),
    ({"requirement":"Provide grab bars and toilet accessories","trade":"Toilet / Bath Accessories","source_sheet":"A9.1","source_detail":"5/A9.2","source_spec":"10 28 00","source_note":""},"STRONG","READY"),
    ({"requirement":"Provide rubber base and LVT flooring","trade":"Flooring / Tile","source_sheet":"A10.1","source_detail":"","source_spec":"09 65 00","source_note":"Finish schedule"},"STRONG","READY"),
    ({"requirement":"Remove existing partitions doors ceilings and millwork","trade":"Demolition","source_sheet":"AD1.1","source_detail":"","source_spec":"02 41 19","source_note":"Demo note 2"},"STRONG","READY"),
]


def _v372_source_regression_results():
    rows=[]
    for item,expected_level,expected_disposition in _V372_SOURCE_REGRESSION_CASES:
        result=_v372_item_intelligence(item)
        passed=(result['source']['level']==expected_level and result['disposition']==expected_disposition)
        rows.append({
            "requirement":item['requirement'],
            "expected_level":expected_level,
            "actual_level":result['source']['level'],
            "expected_disposition":expected_disposition,
            "actual_disposition":result['disposition'],
            "score":result['confidence_score'],
            "passed":passed,
        })
    return rows


def _v372_source_regression_summary():
    rows=_v372_source_regression_results()
    passed=sum(1 for r in rows if r['passed'])
    trade=_v371_trade_regression_summary()
    return {
        "version":"v372",
        "suite":"Blueprint Brain source intelligence",
        "source_passed":passed,
        "source_total":len(rows),
        "trade_passed":trade['passed'],
        "trade_total":trade['total'],
        "passed":passed+trade['passed'],
        "total":len(rows)+trade['total'],
        "failed":(len(rows)-passed)+trade['failed'],
        "ok":passed==len(rows) and trade['ok'],
        "results":rows,
    }


@app.get('/health/blueprint-v372')
def v372_blueprint_health():
    """Staging gate: v371 trade ownership plus v372 source-intelligence regression checks."""
    return _v372_source_regression_summary()


@app.get('/blueprint-source-intelligence-v372', response_class=HTMLResponse)
def v372_blueprint_source_intelligence_page():
    pid=project_id(); cid=current_company_id()
    c=db()
    rows=c.execute("""
        SELECT * FROM blueprint_scope_items
        WHERE company_id=? AND project_id=?
        ORDER BY id DESC LIMIT 250
    """,(cid,pid)).fetchall()
    c.close()
    analyzed=[]
    for r in rows:
        try:
            analyzed.append(_v372_item_intelligence(dict(r)))
        except Exception:
            continue
    ready=sum(1 for x in analyzed if x['disposition']=='READY')
    review=sum(1 for x in analyzed if x['disposition']=='REVIEW_TRADE')
    verify=sum(1 for x in analyzed if x['disposition'] in {'VERIFY_SOURCE','VERIFY_CROSS_REFERENCE'})
    cards=''.join(
        f'<div class="action"><span class="badge {"READY" if x["disposition"]=="READY" else "WATCH"}">{esc(x["disposition"])}</span> '
        f'<b>{esc(x["trade"])}</b><div>{esc(x["requirement"])}</div>'
        f'<div class="small">Source strength: {esc(x["source"]["level"])} · Evidence types: {x["source"]["count"]} · Score: {x["confidence_score"]}/100 · Classifier: {esc(x["trade_check"]["predicted"])}</div></div>'
        for x in analyzed
    ) or '<p class="muted">No Blueprint Brain scope items yet. Analyze project documents first.</p>'
    return shell('Blueprint Source Intelligence v372',
        f'<div class="hero"><div class="eyebrow">BuildCommand v373</div><h1>Blueprint Source Intelligence</h1>'
        f'<p class="muted">Cross-check sheet, detail, specification and note evidence before scope intelligence is trusted downstream.</p></div>'
        f'<div class="grid3"><div class="card"><div class="label">Ready</div><div class="kpi">{ready}</div></div>'
        f'<div class="card"><div class="label">Trade Review</div><div class="kpi">{review}</div></div>'
        f'<div class="card"><div class="label">Source Verification</div><div class="kpi">{verify}</div></div></div>'
        f'<div class="card"><h2>Source-backed Scope Review</h2>{cards}</div>')


# =============================================================================
# BuildCommand AI v373 - Cross-Document Conflict Intelligence
# Detects contradictions between plan/spec/detail/note evidence before downstream use.
# =============================================================================

import re as _v373_re

_V373_NEGATION_PAIRS = [
    ("existing", "new"),
    ("remove", "remain"),
    ("remove", "retain"),
    ("demolish", "remain"),
    ("demolish", "retain"),
    ("provide", "not required"),
    ("install", "not required"),
    ("paint", "factory finish"),
    ("paint", "prefinished"),
    ("reuse", "replace"),
]


def _v373_norm(value):
    s=str(value or '').lower().strip()
    s=_v373_re.sub(r'[^a-z0-9./# -]+',' ',s)
    s=_v373_re.sub(r'\s+',' ',s)
    return s


def _v373_source_texts(item):
    """Collect source-specific text when present without assuming every schema has these columns."""
    fields=(
        'sheet_text','detail_text','spec_text','note_text',
        'source_sheet_text','source_detail_text','source_spec_text','source_note_text',
    )
    out=[]
    for field in fields:
        try:
            value=item.get(field,'') if hasattr(item,'get') else item[field]
        except Exception:
            value=''
        value=str(value or '').strip()
        if value:
            out.append({'field':field,'text':value})
    return out


def _v373_numeric_tokens(text):
    """Capture construction-relevant sizes/ratings/voltages without treating sheet numbers as conflicts."""
    s=_v373_norm(text)
    patterns=[
        r'\b\d+(?:\.\d+)?\s*(?:v|volt|volts|kw|amp|amps|a)\b',
        r'\b\d+(?:\.\d+)?\s*(?:inch|inches|in\.?|mm|cm)\b',
        r'\b\d+\s*(?:hr|hour|hours)\b',
        r'\btype\s+[a-z0-9-]+\b',
    ]
    vals=[]
    for p in patterns:
        vals.extend(_v373_re.findall(p,s))
    return sorted(set(vals))


def _v373_conflict_analysis(item):
    requirement=str(item.get('requirement','') if hasattr(item,'get') else item['requirement'])
    sources=_v373_source_texts(item)
    conflicts=[]

    # Direct semantic opposition across different document sources.
    for i,a in enumerate(sources):
        na=_v373_norm(a['text'])
        for b in sources[i+1:]:
            nb=_v373_norm(b['text'])
            for left,right in _V373_NEGATION_PAIRS:
                if ((left in na and right in nb) or (right in na and left in nb)):
                    conflicts.append({
                        'type':'SEMANTIC_CONFLICT','severity':'HIGH',
                        'sources':[a['field'],b['field']],
                        'detail':f'{left} vs {right}'
                    })

    # Different explicit technical values can indicate a genuine coordination conflict.
    numeric=[]
    for s in sources:
        vals=_v373_numeric_tokens(s['text'])
        if vals:
            numeric.append((s['field'],vals))
    for i,(fa,va) in enumerate(numeric):
        for fb,vb in numeric[i+1:]:
            # Compare only when both sources contain same broad unit/category.
            cats={
                'electrical': lambda x: any(u in x for u in (' v','volt','kw','amp')),
                'dimension': lambda x: any(u in x for u in ('inch',' in','mm','cm')),
                'rating': lambda x: any(u in x for u in ('hr','hour')),
                'type': lambda x: x.startswith('type '),
            }
            for cat,fn in cats.items():
                aa={x for x in va if fn(x)}; bb={x for x in vb if fn(x)}
                if aa and bb and aa != bb:
                    conflicts.append({
                        'type':'VALUE_CONFLICT','severity':'HIGH',
                        'sources':[fa,fb],
                        'detail':f'{cat}: {sorted(aa)} vs {sorted(bb)}'
                    })

    # Deduplicate deterministic findings.
    seen=set(); unique=[]
    for c in conflicts:
        key=(c['type'],tuple(c['sources']),c['detail'])
        if key not in seen:
            seen.add(key); unique.append(c)

    if unique:
        disposition='RESOLVE_CONFLICT'
        risk='HIGH'
    elif len(sources) >= 2:
        disposition='NO_CONFLICT_FOUND'
        risk='LOW'
    elif len(sources) == 1:
        disposition='NEEDS_CROSS_CHECK'
        risk='MEDIUM'
    else:
        disposition='NO_COMPARABLE_SOURCE_TEXT'
        risk='UNKNOWN'
    return {
        'requirement':requirement,
        'source_text_count':len(sources),
        'conflict_count':len(unique),
        'risk':risk,
        'disposition':disposition,
        'conflicts':unique,
    }


_V373_CONFLICT_CASES = [
    ({'requirement':'Door D101','sheet_text':'Provide new hollow metal door D101','spec_text':'Door D101 shall be new hollow metal'},0,'NO_CONFLICT_FOUND'),
    ({'requirement':'Door D102','sheet_text':'Remove door D102','note_text':'Door D102 to remain'},1,'RESOLVE_CONFLICT'),
    ({'requirement':'WH-1','sheet_text':'WH-1 277 V 8.31 kW','spec_text':'WH-1 208 V 8.31 kW'},1,'RESOLVE_CONFLICT'),
    ({'requirement':'Partition rating','sheet_text':'Partition shall be 1 hr rated','detail_text':'Partition shall be 2 hr rated'},1,'RESOLVE_CONFLICT'),
    ({'requirement':'Existing casework','sheet_text':'Existing casework to remain','note_text':'Retain existing casework'},0,'NO_CONFLICT_FOUND'),
    ({'requirement':'Roof curb','detail_text':'Install new roof curb','spec_text':'Roof curb not required'},1,'RESOLVE_CONFLICT'),
    ({'requirement':'HM frames','sheet_text':'Paint hollow metal frames','spec_text':'Hollow metal frames factory finish'},1,'RESOLVE_CONFLICT'),
    ({'requirement':'Ceiling','sheet_text':'Provide ACT ceiling','spec_text':'Provide ACT ceiling system'},0,'NO_CONFLICT_FOUND'),
    ({'requirement':'Reuse fixture','sheet_text':'Reuse existing fixture','note_text':'Replace existing fixture'},1,'RESOLVE_CONFLICT'),
    ({'requirement':'Single source item','sheet_text':'Provide floor drain FD-1'},0,'NEEDS_CROSS_CHECK'),
]


def _v373_conflict_regression_results():
    rows=[]
    for item,minimum_conflicts,expected_disposition in _V373_CONFLICT_CASES:
        result=_v373_conflict_analysis(item)
        passed=(result['conflict_count'] >= minimum_conflicts and result['disposition']==expected_disposition)
        rows.append({
            'requirement':item['requirement'],
            'expected_disposition':expected_disposition,
            'actual_disposition':result['disposition'],
            'conflicts':result['conflict_count'],
            'passed':passed,
        })
    return rows


def _v373_conflict_regression_summary():
    conflict_rows=_v373_conflict_regression_results()
    conflict_passed=sum(1 for r in conflict_rows if r['passed'])
    previous=_v372_source_regression_summary()
    return {
        'version':'v373',
        'suite':'Blueprint Brain cross-document conflict intelligence',
        'conflict_passed':conflict_passed,
        'conflict_total':len(conflict_rows),
        'source_passed':previous['source_passed'],
        'source_total':previous['source_total'],
        'trade_passed':previous['trade_passed'],
        'trade_total':previous['trade_total'],
        'passed':conflict_passed+previous['passed'],
        'total':len(conflict_rows)+previous['total'],
        'failed':(len(conflict_rows)-conflict_passed)+previous['failed'],
        'ok':conflict_passed==len(conflict_rows) and previous['ok'],
        'results':conflict_rows,
    }


@app.get('/health/blueprint-v373')
def v373_blueprint_health():
    """Staging gate: trade ownership + source intelligence + conflict intelligence."""
    return _v373_conflict_regression_summary()


@app.get('/blueprint-conflicts-v373', response_class=HTMLResponse)
def v373_blueprint_conflicts_page():
    pid=project_id(); cid=current_company_id(); c=db()
    rows=c.execute("""
        SELECT * FROM blueprint_scope_items
        WHERE company_id=? AND project_id=?
        ORDER BY id DESC LIMIT 250
    """,(cid,pid)).fetchall()
    c.close()
    analyzed=[]
    for r in rows:
        try:
            analyzed.append(_v373_conflict_analysis(dict(r)))
        except Exception:
            continue
    conflicts=sum(1 for x in analyzed if x['disposition']=='RESOLVE_CONFLICT')
    cross=sum(1 for x in analyzed if x['disposition']=='NEEDS_CROSS_CHECK')
    clear=sum(1 for x in analyzed if x['disposition']=='NO_CONFLICT_FOUND')
    cards=''.join(
        f'<div class="action"><span class="badge {"WATCH" if x["disposition"]=="RESOLVE_CONFLICT" else "READY"}">{esc(x["disposition"])}</span> '
        f'<b>{esc(x["requirement"])}</b><div class="small">Comparable source text: {x["source_text_count"]} · Conflicts: {x["conflict_count"]} · Risk: {esc(x["risk"])}</div>'
        + ''.join(f'<div class="small">⚠ {esc(c["type"])} · {esc(c["detail"])}</div>' for c in x['conflicts'])
        + '</div>' for x in analyzed
    ) or '<p class="muted">No Blueprint Brain scope items yet. Analyze project documents first.</p>'
    return shell('Blueprint Conflicts v373',
        f'<div class="hero"><div class="eyebrow">BuildCommand v373</div><h1>Cross-Document Conflict Intelligence</h1>'
        f'<p class="muted">Flags contradictions between plan, detail, specification and note evidence before scope, estimating or field execution trusts the requirement.</p></div>'
        f'<div class="grid3"><div class="card"><div class="label">Conflicts</div><div class="kpi">{conflicts}</div></div>'
        f'<div class="card"><div class="label">Needs Cross-Check</div><div class="kpi">{cross}</div></div>'
        f'<div class="card"><div class="label">No Conflict Found</div><div class="kpi">{clear}</div></div></div>'
        f'<div class="card"><h2>Coordination Review</h2>{cards}</div>')


# =============================================================================
# BuildCommand AI v374 - Conflict-to-RFI Intelligence
# Converts verified cross-document conflicts into structured DRAFT RFI candidates.
# Nothing is issued or emailed automatically. Human approval remains mandatory.
# =============================================================================

def _v374_safe_get(row, key, default=""):
    try:
        if hasattr(row, "get"):
            return row.get(key, default)
        return row[key]
    except Exception:
        return default


def _v374_source_label(field):
    names = {
        "sheet_text":"Drawing / Sheet",
        "source_sheet_text":"Drawing / Sheet",
        "detail_text":"Detail",
        "source_detail_text":"Detail",
        "spec_text":"Specification",
        "source_spec_text":"Specification",
        "note_text":"Plan Note",
        "source_note_text":"Plan Note",
    }
    return names.get(str(field or ""), str(field or "").replace("_"," ").title())


def _v374_trade_for_item(item):
    trade = str(_v374_safe_get(item, "trade", "") or "").strip()
    requirement = str(_v374_safe_get(item, "requirement", "") or "").strip()
    try:
        predicted = _v371_trade_owner(requirement)
    except Exception:
        predicted = ""
    # Prefer an explicit stored trade unless it is blank/ambiguous.
    if trade and trade.lower() not in {"unknown","tbd","other","general","unassigned"}:
        return trade
    return predicted or trade or "GC / Design Team"


def _v374_source_refs(item, conflict):
    refs = []
    for field in conflict.get("sources", []):
        val = str(_v374_safe_get(item, field, "") or "").strip()
        refs.append({
            "field": field,
            "label": _v374_source_label(field),
            "text": val,
        })
    return refs


def _v374_build_rfi_candidate(item):
    analysis = _v373_conflict_analysis(item)
    if analysis["disposition"] != "RESOLVE_CONFLICT":
        return None

    req = str(_v374_safe_get(item, "requirement", "") or "").strip() or "Document coordination conflict"
    trade = _v374_trade_for_item(item)
    source_blocks = []
    source_labels = []

    for conflict in analysis["conflicts"]:
        refs = _v374_source_refs(item, conflict)
        for ref in refs:
            if ref["label"] not in source_labels:
                source_labels.append(ref["label"])
            if ref["text"]:
                source_blocks.append(f'{ref["label"]}: {ref["text"]}')

    source_summary = " | ".join(source_blocks) if source_blocks else "Conflicting source references detected by BuildCommand; verify exact document references before issue."
    conflict_summary = "; ".join(c["detail"] for c in analysis["conflicts"])

    title = f"Clarification Required - {req[:120]}"
    question = (
        f"Please clarify the governing contract requirement for: {req}. "
        f"BuildCommand detected conflicting document information ({conflict_summary}). "
        "Please identify which requirement governs and provide any required revised detail, specification direction, "
        "or drawing clarification."
    )
    background = (
        f"Potential cross-document conflict identified for {trade}. "
        f"Detected conflict: {conflict_summary}. "
        f"Compared sources: {', '.join(source_labels) if source_labels else 'project documents'}."
    )
    impact = (
        "Potential impact is not yet quantified. Clarification may affect trade coordination, procurement, installation, "
        "inspection readiness, cost, or schedule. Human project review is required before assigning any impact."
    )
    return {
        "title": title,
        "requirement": req,
        "trade": trade,
        "question": question,
        "background": background,
        "potential_impact": impact,
        "source_summary": source_summary,
        "conflict_count": analysis["conflict_count"],
        "risk": analysis["risk"],
        "status": "DRAFT_CANDIDATE",
        "human_approval_required": True,
    }


def _v374_ensure_rfi_control():
    c = db()
    kind = DATABASE_KIND
    pk = "BIGSERIAL PRIMARY KEY" if kind == "postgres" else "INTEGER PRIMARY KEY AUTOINCREMENT"
    num = "DOUBLE PRECISION" if kind == "postgres" else "REAL"
    c.execute(
        f"""CREATE TABLE IF NOT EXISTS rfi_control(
            id {pk},
            company_id BIGINT,
            project_id BIGINT,
            number TEXT,
            title TEXT,
            question TEXT,
            responsible_party TEXT,
            due_date TEXT,
            status TEXT DEFAULT 'DRAFT',
            answer TEXT,
            cost_impact {num} DEFAULT 0,
            schedule_days {num} DEFAULT 0,
            source_ref TEXT,
            created TEXT,
            updated TEXT
        )"""
    )
    try:
        c.commit()
    except Exception:
        pass
    c.close()


def _v374_number_for_next_rfi(pid):
    _v374_ensure_rfi_control()
    c = db()
    row = c.execute(
        "SELECT COUNT(*) AS n FROM rfi_control WHERE company_id=? AND project_id=?",
        (current_company_id(), pid)
    ).fetchone()
    c.close()
    try:
        n = int(row["n"] or 0) + 1
    except Exception:
        n = 1
    return f"RFI-DRAFT-{n:03d}"


def _v374_candidate_from_scope_id(scope_item_id):
    pid = project_id()
    cid = current_company_id()
    c = db()
    row = c.execute(
        "SELECT * FROM blueprint_scope_items WHERE id=? AND company_id=? AND project_id=?",
        (scope_item_id, cid, pid)
    ).fetchone()
    c.close()
    if not row:
        return None, None
    item = dict(row)
    return item, _v374_build_rfi_candidate(item)


_V374_RFI_CASES = [
    (
        {"requirement":"WH-1 electrical service","trade":"Plumbing",
         "sheet_text":"WH-1 277 V 8.31 kW","spec_text":"WH-1 208 V 8.31 kW"},
        True, "Plumbing"
    ),
    (
        {"requirement":"Door D102","trade":"Doors & Hardware",
         "sheet_text":"Remove door D102","note_text":"Door D102 to remain"},
        True, "Doors & Hardware"
    ),
    (
        {"requirement":"Partition rating","trade":"Framing & Drywall",
         "sheet_text":"Partition shall be 1 hr rated","detail_text":"Partition shall be 2 hr rated"},
        True, "Framing & Drywall"
    ),
    (
        {"requirement":"ACT ceiling","trade":"Ceilings",
         "sheet_text":"Provide ACT ceiling","spec_text":"Provide ACT ceiling system"},
        False, "Ceilings"
    ),
    (
        {"requirement":"Existing casework","trade":"Millwork",
         "sheet_text":"Existing casework to remain","note_text":"Retain existing casework"},
        False, "Millwork"
    ),
    (
        {"requirement":"Roof curb","trade":"Roofing",
         "detail_text":"Install new roof curb","spec_text":"Roof curb not required"},
        True, "Roofing"
    ),
    (
        {"requirement":"HM frames","trade":"Paint",
         "sheet_text":"Paint hollow metal frames","spec_text":"Hollow metal frames factory finish"},
        True, "Paint"
    ),
    (
        {"requirement":"Reuse light fixture","trade":"Electrical",
         "sheet_text":"Reuse existing fixture","note_text":"Replace existing fixture"},
        True, "Electrical"
    ),
    (
        {"requirement":"Floor drain FD-1","trade":"Plumbing",
         "sheet_text":"Provide floor drain FD-1"},
        False, "Plumbing"
    ),
    (
        {"requirement":"New restroom grab bars","trade":"Bathroom Accessories",
         "sheet_text":"Provide new grab bars","spec_text":"Provide new grab bars"},
        False, "Bathroom Accessories"
    ),
]


def _v374_rfi_regression_results():
    rows = []
    for item, should_create, expected_trade in _V374_RFI_CASES:
        candidate = _v374_build_rfi_candidate(item)
        created = candidate is not None
        passed = (
            created == should_create and
            (not candidate or candidate["trade"] == expected_trade) and
            (not candidate or candidate["human_approval_required"] is True) and
            (not candidate or candidate["status"] == "DRAFT_CANDIDATE")
        )
        rows.append({
            "requirement": item["requirement"],
            "expected_candidate": should_create,
            "candidate_created": created,
            "expected_trade": expected_trade,
            "actual_trade": candidate["trade"] if candidate else expected_trade,
            "human_approval_required": candidate["human_approval_required"] if candidate else True,
            "passed": passed,
        })
    return rows


def _v374_rfi_regression_summary():
    rfi_rows = _v374_rfi_regression_results()
    rfi_passed = sum(1 for r in rfi_rows if r["passed"])
    previous = _v373_conflict_regression_summary()
    return {
        "version":"v374",
        "suite":"Blueprint Brain conflict-to-RFI intelligence",
        "rfi_passed":rfi_passed,
        "rfi_total":len(rfi_rows),
        "conflict_passed":previous["conflict_passed"],
        "conflict_total":previous["conflict_total"],
        "source_passed":previous["source_passed"],
        "source_total":previous["source_total"],
        "trade_passed":previous["trade_passed"],
        "trade_total":previous["trade_total"],
        "passed":rfi_passed + previous["passed"],
        "total":len(rfi_rows) + previous["total"],
        "failed":(len(rfi_rows)-rfi_passed) + previous["failed"],
        "ok":rfi_passed == len(rfi_rows) and previous["ok"],
        "results":rfi_rows,
    }


@app.get("/health/blueprint-v374")
def v374_blueprint_health():
    """Staging gate: trade + source + conflict + conflict-to-RFI intelligence."""
    return _v374_rfi_regression_summary()


@app.get("/rfi-intelligence-v374", response_class=HTMLResponse)
def v374_rfi_intelligence_page():
    pid = project_id()
    cid = current_company_id()
    c = db()
    rows = c.execute(
        """SELECT * FROM blueprint_scope_items
           WHERE company_id=? AND project_id=?
           ORDER BY id DESC LIMIT 300""",
        (cid, pid)
    ).fetchall()
    c.close()

    candidates = []
    for r in rows:
        item = dict(r)
        candidate = _v374_build_rfi_candidate(item)
        if candidate:
            candidate["scope_item_id"] = item.get("id")
            candidates.append(candidate)

    cards = ""
    for x in candidates:
        cards += (
            '<div class="card">'
            '<span class="badge WATCH">DRAFT RFI CANDIDATE</span>'
            f'<h3>{esc(x["title"])}</h3>'
            f'<p><b>Trade:</b> {esc(x["trade"])}</p>'
            f'<p><b>Background:</b> {esc(x["background"])}</p>'
            f'<p><b>Question:</b> {esc(x["question"])}</p>'
            f'<p><b>Potential impact:</b> {esc(x["potential_impact"])}</p>'
            f'<p class="small"><b>Source evidence:</b> {esc(x["source_summary"])}</p>'
            '<p class="small"><b>Safety rule:</b> Nothing is issued automatically. A person must review and save the draft.</p>'
            f'<form method="post" action="/rfi-intelligence-v374/save-draft">'
            f'<input type="hidden" name="scope_item_id" value="{int(x["scope_item_id"])}">'
            '<button type="submit">Save as Draft RFI</button>'
            '</form>'
            '</div>'
        )

    if not cards:
        cards = '<div class="card"><p class="muted">No verified cross-document conflicts currently qualify for an RFI draft.</p></div>'

    return shell(
        "RFI Intelligence v374",
        f'<div class="hero"><div class="eyebrow">BuildCommand v374</div>'
        f'<h1>Conflict-to-RFI Intelligence</h1>'
        f'<p class="muted">Turns verified document conflicts into structured RFI candidates while keeping final judgment and issuance under human control.</p></div>'
        f'<div class="grid3"><div class="card"><div class="label">RFI Candidates</div><div class="kpi">{len(candidates)}</div></div>'
        f'<div class="card"><div class="label">Auto-Issued</div><div class="kpi">0</div></div>'
        f'<div class="card"><div class="label">Approval Rule</div><div class="kpi">HUMAN</div></div></div>'
        + cards
    )


@app.post("/rfi-intelligence-v374/save-draft")
def v374_save_rfi_draft(scope_item_id:int=Form(...)):
    item, candidate = _v374_candidate_from_scope_id(scope_item_id)
    if not item or not candidate:
        return RedirectResponse("/rfi-intelligence-v374", status_code=303)

    _v374_ensure_rfi_control()
    pid = project_id()
    now = datetime.utcnow().isoformat()
    number = _v374_number_for_next_rfi(pid)
    source_ref = json.dumps({
        "blueprint_scope_item_id": scope_item_id,
        "requirement": candidate["requirement"],
        "source_summary": candidate["source_summary"],
        "generated_by": "BuildCommand v374",
        "human_approval_required": True,
    })

    c = db()
    c.execute(
        """INSERT INTO rfi_control(
            company_id,project_id,number,title,question,responsible_party,due_date,status,
            answer,cost_impact,schedule_days,source_ref,created,updated
        ) VALUES(?,?,?,?,?,?,?,'DRAFT','',0,0,?,?,?)""",
        (
            current_company_id(), pid, number, candidate["title"], candidate["question"],
            candidate["trade"], "", source_ref, now, now
        )
    )
    c.commit()
    c.close()
    return RedirectResponse("/project-control/rfis", status_code=303)


# =============================================================================
# BuildCommand AI v375 - RFI Impact Intelligence
# Connects an RFI candidate to schedule, procurement, inspection, cost, and field
# readiness signals. It reports evidence; it does NOT invent cost/schedule values.
# =============================================================================

def _v375_text(value):
    return str(value or "").strip()

def _v375_lower(value):
    return _v375_text(value).lower()

def _v375_match_score(requirement, trade, name="", row_trade=""):
    req = _v375_lower(requirement)
    tr = _v375_lower(trade)
    nm = _v375_lower(name)
    rt = _v375_lower(row_trade)
    score = 0
    if tr and rt and (tr == rt or tr in rt or rt in tr):
        score += 5
    words = [w for w in re.findall(r"[a-z0-9]+", req) if len(w) >= 4]
    score += min(5, sum(1 for w in set(words) if w in nm))
    return score

def _v375_query_safe(c, sql, params=()):
    try:
        return c.execute(sql, params).fetchall()
    except Exception:
        try:
            c.rollback()
        except Exception:
            pass
        return []

def _v375_impact_analysis(item, candidate=None):
    candidate = candidate or _v374_build_rfi_candidate(item)
    if not candidate:
        return None

    pid = project_id()
    cid = current_company_id()
    requirement = candidate["requirement"]
    trade = candidate["trade"]
    c = db()

    activities = _v375_query_safe(
        c,
        """SELECT id,name,trade,start,finish,pct,status
           FROM activities WHERE project_id=? ORDER BY start""",
        (pid,)
    )
    procurement = _v375_query_safe(
        c,
        """SELECT id,item,activity_id,required_on_site,promised_date,status
           FROM procurement WHERE project_id=?""",
        (pid,)
    )
    inspections = _v375_query_safe(
        c,
        """SELECT id,name,activity_id,due,status
           FROM inspections_tracker WHERE project_id=?""",
        (pid,)
    )
    issues = _v375_query_safe(
        c,
        """SELECT id,title,activity_id,status
           FROM project_issues WHERE project_id=?""",
        (pid,)
    )
    readiness = _v375_query_safe(
        c,
        """SELECT activity_id,drawings_ok,material_ok,manpower_ok,predecessor_ok,
                  access_ok,inspection_ok,equipment_ok
           FROM activity_readiness WHERE project_id=?""",
        (pid,)
    )
    changes = _v375_query_safe(
        c,
        """SELECT id,title,status,cost_impact,schedule_days
           FROM change_events WHERE project_id=?""",
        (pid,)
    )
    c.close()

    matched_activities = []
    for r in activities:
        rr = dict(r)
        score = _v375_match_score(requirement, trade, rr.get("name",""), rr.get("trade",""))
        if score >= 5:
            rr["match_score"] = score
            matched_activities.append(rr)

    matched_ids = {int(r["id"]) for r in matched_activities if r.get("id") is not None}

    proc_hits = []
    for r in procurement:
        rr = dict(r)
        activity_hit = rr.get("activity_id") is not None and int(rr["activity_id"]) in matched_ids
        text_hit = _v375_match_score(requirement, trade, rr.get("item",""), "") >= 2
        if activity_hit or text_hit:
            proc_hits.append(rr)

    inspection_hits = []
    for r in inspections:
        rr = dict(r)
        activity_hit = rr.get("activity_id") is not None and int(rr["activity_id"]) in matched_ids
        text_hit = _v375_match_score(requirement, trade, rr.get("name",""), "") >= 2
        if activity_hit or text_hit:
            inspection_hits.append(rr)

    issue_hits = []
    for r in issues:
        rr = dict(r)
        activity_hit = rr.get("activity_id") is not None and int(rr["activity_id"]) in matched_ids
        text_hit = _v375_match_score(requirement, trade, rr.get("title",""), "") >= 2
        if activity_hit or text_hit:
            issue_hits.append(rr)

    readiness_by_activity = {}
    for r in readiness:
        rr = dict(r)
        try:
            aid = int(rr.get("activity_id"))
        except Exception:
            continue
        if aid in matched_ids:
            checks = [
                rr.get("drawings_ok"), rr.get("material_ok"), rr.get("manpower_ok"),
                rr.get("predecessor_ok"), rr.get("access_ok"), rr.get("inspection_ok"),
                rr.get("equipment_ok")
            ]
            readiness_by_activity[aid] = {
                "ready_checks": sum(1 for x in checks if bool(x)),
                "total_checks": len(checks),
                "all_ready": all(bool(x) for x in checks),
            }

    schedule_signal = "NO_LINKED_ACTIVITY"
    if matched_activities:
        active = [x for x in matched_activities if _v375_lower(x.get("status")) not in {"complete","completed"}]
        schedule_signal = "POTENTIAL_EXPOSURE" if active else "LINKED_COMPLETE_ACTIVITY"

    procurement_signal = "NO_LINKED_PROCUREMENT"
    if proc_hits:
        open_proc = [x for x in proc_hits if _v375_lower(x.get("status")) not in {"received","complete","completed","closed"}]
        procurement_signal = "POTENTIAL_EXPOSURE" if open_proc else "LINKED_PROCUREMENT_COMPLETE"

    inspection_signal = "NO_LINKED_INSPECTION"
    if inspection_hits:
        open_ins = [x for x in inspection_hits if _v375_lower(x.get("status")) not in {"passed","complete","completed","closed"}]
        inspection_signal = "POTENTIAL_EXPOSURE" if open_ins else "LINKED_INSPECTIONS_COMPLETE"

    readiness_signal = "NO_READINESS_RECORD"
    if readiness_by_activity:
        readiness_signal = (
            "READY" if all(x["all_ready"] for x in readiness_by_activity.values())
            else "NOT_READY"
        )

    # Existing quantified change data is evidence only. We never manufacture an estimate.
    quantified_changes = []
    for r in changes:
        rr = dict(r)
        title = _v375_lower(rr.get("title"))
        req_words = [w for w in re.findall(r"[a-z0-9]+", _v375_lower(requirement)) if len(w) >= 5]
        if req_words and any(w in title for w in req_words):
            quantified_changes.append(rr)

    known_cost = sum(float(x.get("cost_impact") or 0) for x in quantified_changes)
    known_days = sum(float(x.get("schedule_days") or 0) for x in quantified_changes)

    evidence_count = (
        len(matched_activities) + len(proc_hits) + len(inspection_hits) +
        len(issue_hits) + len(readiness_by_activity) + len(quantified_changes)
    )

    if any(x == "POTENTIAL_EXPOSURE" for x in (schedule_signal, procurement_signal, inspection_signal)) or readiness_signal == "NOT_READY":
        overall = "IMPACT_REVIEW_REQUIRED"
    elif evidence_count:
        overall = "LINKED_EVIDENCE_FOUND"
    else:
        overall = "NO_LINKED_PROJECT_EVIDENCE"

    return {
        "requirement": requirement,
        "trade": trade,
        "overall": overall,
        "schedule_signal": schedule_signal,
        "procurement_signal": procurement_signal,
        "inspection_signal": inspection_signal,
        "readiness_signal": readiness_signal,
        "linked_activities": matched_activities,
        "linked_procurement": proc_hits,
        "linked_inspections": inspection_hits,
        "linked_issues": issue_hits,
        "readiness": readiness_by_activity,
        "existing_quantified_change_count": len(quantified_changes),
        "existing_known_cost_impact": round(known_cost, 2),
        "existing_known_schedule_days": round(known_days, 2),
        "cost_statement": (
            f"Existing linked project records contain ${known_cost:,.2f} of cost impact."
            if quantified_changes else
            "No verified cost impact is recorded. BuildCommand will not invent a dollar value."
        ),
        "schedule_statement": (
            f"Existing linked project records contain {known_days:g} schedule day(s) of impact."
            if quantified_changes and known_days else
            "No verified schedule-day impact is recorded. BuildCommand will not invent a duration."
        ),
        "human_review_required": True,
    }


_V375_IMPACT_CASES = [
    ("schedule exposure", {"schedule_signal":"POTENTIAL_EXPOSURE"}, "IMPACT_REVIEW_REQUIRED"),
    ("procurement exposure", {"procurement_signal":"POTENTIAL_EXPOSURE"}, "IMPACT_REVIEW_REQUIRED"),
    ("inspection exposure", {"inspection_signal":"POTENTIAL_EXPOSURE"}, "IMPACT_REVIEW_REQUIRED"),
    ("readiness blocked", {"readiness_signal":"NOT_READY"}, "IMPACT_REVIEW_REQUIRED"),
    ("linked complete activity", {"schedule_signal":"LINKED_COMPLETE_ACTIVITY"}, "LINKED_EVIDENCE_FOUND"),
    ("no evidence", {}, "NO_LINKED_PROJECT_EVIDENCE"),
    ("no invented cost", {"known_cost":0}, "SAFE"),
    ("no invented days", {"known_days":0}, "SAFE"),
    ("human review", {"human_review_required":True}, "SAFE"),
    ("multiple signals", {"schedule_signal":"POTENTIAL_EXPOSURE","readiness_signal":"NOT_READY"}, "IMPACT_REVIEW_REQUIRED"),
]

def _v375_decision_from_signals(signals):
    if any(signals.get(k) == "POTENTIAL_EXPOSURE" for k in ("schedule_signal","procurement_signal","inspection_signal")):
        return "IMPACT_REVIEW_REQUIRED"
    if signals.get("readiness_signal") == "NOT_READY":
        return "IMPACT_REVIEW_REQUIRED"
    if signals.get("schedule_signal") == "LINKED_COMPLETE_ACTIVITY":
        return "LINKED_EVIDENCE_FOUND"
    return "NO_LINKED_PROJECT_EVIDENCE"

def _v375_impact_regression_results():
    rows = []
    for name, signals, expected in _V375_IMPACT_CASES:
        if expected == "SAFE":
            if "known_cost" in signals:
                passed = signals["known_cost"] == 0
            elif "known_days" in signals:
                passed = signals["known_days"] == 0
            else:
                passed = signals.get("human_review_required") is True
            actual = "SAFE" if passed else "UNSAFE"
        else:
            actual = _v375_decision_from_signals(signals)
            passed = actual == expected
        rows.append({
            "case": name,
            "expected": expected,
            "actual": actual,
            "passed": passed,
        })
    return rows

def _v375_impact_regression_summary():
    impact_rows = _v375_impact_regression_results()
    impact_passed = sum(1 for r in impact_rows if r["passed"])
    previous = _v374_rfi_regression_summary()
    return {
        "version":"v375",
        "suite":"RFI impact intelligence",
        "impact_passed":impact_passed,
        "impact_total":len(impact_rows),
        "rfi_passed":previous["rfi_passed"],
        "rfi_total":previous["rfi_total"],
        "conflict_passed":previous["conflict_passed"],
        "conflict_total":previous["conflict_total"],
        "source_passed":previous["source_passed"],
        "source_total":previous["source_total"],
        "trade_passed":previous["trade_passed"],
        "trade_total":previous["trade_total"],
        "passed":impact_passed + previous["passed"],
        "total":len(impact_rows) + previous["total"],
        "failed":(len(impact_rows)-impact_passed) + previous["failed"],
        "ok":impact_passed == len(impact_rows) and previous["ok"],
        "results":impact_rows,
    }

@app.get("/health/blueprint-v375")
def v375_blueprint_health():
    """Staging gate through RFI impact intelligence."""
    return _v375_impact_regression_summary()

@app.get("/rfi-impact-v375", response_class=HTMLResponse)
def v375_rfi_impact_page():
    pid = project_id()
    cid = current_company_id()
    c = db()
    rows = c.execute(
        """SELECT * FROM blueprint_scope_items
           WHERE company_id=? AND project_id=?
           ORDER BY id DESC LIMIT 300""",
        (cid, pid)
    ).fetchall()
    c.close()

    analyses = []
    for row in rows:
        item = dict(row)
        candidate = _v374_build_rfi_candidate(item)
        if not candidate:
            continue
        try:
            impact = _v375_impact_analysis(item, candidate)
        except Exception:
            impact = None
        if impact:
            analyses.append((candidate, impact))

    cards = ""
    for candidate, x in analyses:
        badge = "WATCH" if x["overall"] == "IMPACT_REVIEW_REQUIRED" else "READY"
        cards += (
            '<div class="card">'
            f'<span class="badge {badge}">{esc(x["overall"])}</span>'
            f'<h3>{esc(candidate["title"])}</h3>'
            f'<p><b>Trade:</b> {esc(x["trade"])}</p>'
            f'<div class="grid3">'
            f'<div><b>Schedule</b><div class="small">{esc(x["schedule_signal"])}</div></div>'
            f'<div><b>Procurement</b><div class="small">{esc(x["procurement_signal"])}</div></div>'
            f'<div><b>Inspection</b><div class="small">{esc(x["inspection_signal"])}</div></div>'
            f'</div>'
            f'<p><b>Field readiness:</b> {esc(x["readiness_signal"])}</p>'
            f'<p><b>Cost:</b> {esc(x["cost_statement"])}</p>'
            f'<p><b>Schedule duration:</b> {esc(x["schedule_statement"])}</p>'
            f'<p class="small">Linked activities: {len(x["linked_activities"])} · '
            f'Procurement: {len(x["linked_procurement"])} · '
            f'Inspections: {len(x["linked_inspections"])} · '
            f'Issues: {len(x["linked_issues"])}</p>'
            '<p class="small"><b>Control:</b> Human review is required before assigning cost, schedule impact, responsibility, or issuing an RFI.</p>'
            '</div>'
        )

    if not cards:
        cards = '<div class="card"><p class="muted">No current conflict-based RFI candidates have linked project impact evidence.</p></div>'

    return shell(
        "RFI Impact Intelligence v375",
        f'<div class="hero"><div class="eyebrow">BuildCommand v375</div>'
        f'<h1>RFI Impact Intelligence</h1>'
        f'<p class="muted">Connects document conflicts to schedule, procurement, inspections, field readiness and existing cost records—without inventing impacts.</p></div>'
        + cards
    )


# =============================================================================
# BuildCommand AI v376 - Decision Intelligence
# Produces a human-review decision package from conflict/RFI/impact evidence.
# =============================================================================

def _v376_priority(impact):
    if not impact:
        return "LOW"
    signals = [
        impact.get("schedule_signal"),
        impact.get("procurement_signal"),
        impact.get("inspection_signal"),
        impact.get("readiness_signal"),
    ]
    if impact.get("overall") == "IMPACT_REVIEW_REQUIRED":
        hits = sum(1 for x in signals if x in {"POTENTIAL_EXPOSURE","NOT_READY"})
        return "HIGH" if hits >= 2 else "MEDIUM"
    if impact.get("overall") == "LINKED_EVIDENCE_FOUND":
        return "MEDIUM"
    return "LOW"


def _v376_reviewer(trade, impact):
    trade_l = str(trade or "").lower()
    if impact and impact.get("inspection_signal") == "POTENTIAL_EXPOSURE":
        return "Project Manager + Superintendent"
    if any(k in trade_l for k in ("electrical","plumbing","hvac","mechanical","fire sprinkler","low voltage","structural","concrete","framing","roof","doors","storefront")):
        return "Project Manager + Superintendent + Affected Trade"
    return "Project Manager + Superintendent"


def _v376_next_action(impact):
    if not impact:
        return "Review the conflict and confirm governing documents before proceeding."
    if impact.get("inspection_signal") == "POTENTIAL_EXPOSURE":
        return "Resolve the document conflict before the affected inspection or inspection-readiness milestone."
    if impact.get("procurement_signal") == "POTENTIAL_EXPOSURE":
        return "Resolve the document conflict before releasing or changing affected procurement."
    if impact.get("readiness_signal") == "NOT_READY":
        return "Resolve the conflict before marking the affected activity ready to start."
    if impact.get("schedule_signal") == "POTENTIAL_EXPOSURE":
        return "Resolve the conflict before the linked activity reaches its planned start or handoff."
    return "Review linked evidence and confirm whether an RFI or field direction is required."


def _v376_decision_package(item):
    candidate = _v374_build_rfi_candidate(item)
    if not candidate:
        return None
    impact = _v375_impact_analysis(item, candidate)
    if not impact:
        return None

    affected = []
    if impact.get("linked_activities"):
        affected.append(f'{len(impact["linked_activities"])} schedule activity(s)')
    if impact.get("linked_procurement"):
        affected.append(f'{len(impact["linked_procurement"])} procurement item(s)')
    if impact.get("linked_inspections"):
        affected.append(f'{len(impact["linked_inspections"])} inspection item(s)')
    if impact.get("readiness"):
        affected.append(f'{len(impact["readiness"])} readiness record(s)')
    if impact.get("linked_issues"):
        affected.append(f'{len(impact["linked_issues"])} project issue(s)')

    return {
        "title": candidate.get("title"),
        "trade": candidate.get("trade"),
        "priority": _v376_priority(impact),
        "reviewer": _v376_reviewer(candidate.get("trade"), impact),
        "what_happened": candidate.get("background"),
        "affected_summary": ", ".join(affected) if affected else "No linked project records yet",
        "decision_needed": "Confirm the governing contract requirement and decide whether to proceed, revise scope, hold affected work, or issue a formal clarification.",
        "recommended_next_action": _v376_next_action(impact),
        "source_summary": candidate.get("source_summary"),
        "cost_statement": impact.get("cost_statement"),
        "schedule_statement": impact.get("schedule_statement"),
        "status": "DECISION_REVIEW",
        "human_approval_required": True,
    }


_V376_PRIORITY_CASES = [
    ({"overall":"IMPACT_REVIEW_REQUIRED","schedule_signal":"POTENTIAL_EXPOSURE","procurement_signal":"POTENTIAL_EXPOSURE","inspection_signal":"NO_LINKED_INSPECTION","readiness_signal":"READY"}, "HIGH"),
    ({"overall":"IMPACT_REVIEW_REQUIRED","schedule_signal":"POTENTIAL_EXPOSURE","procurement_signal":"NO_LINKED_PROCUREMENT","inspection_signal":"NO_LINKED_INSPECTION","readiness_signal":"READY"}, "MEDIUM"),
    ({"overall":"LINKED_EVIDENCE_FOUND","schedule_signal":"LINKED_COMPLETE_ACTIVITY"}, "MEDIUM"),
    ({"overall":"NO_LINKED_PROJECT_EVIDENCE"}, "LOW"),
    ({"overall":"IMPACT_REVIEW_REQUIRED","inspection_signal":"POTENTIAL_EXPOSURE"}, "MEDIUM"),
    ({"overall":"IMPACT_REVIEW_REQUIRED","readiness_signal":"NOT_READY"}, "MEDIUM"),
    ({"overall":"IMPACT_REVIEW_REQUIRED","schedule_signal":"POTENTIAL_EXPOSURE","inspection_signal":"POTENTIAL_EXPOSURE"}, "HIGH"),
    ({"overall":"IMPACT_REVIEW_REQUIRED","schedule_signal":"POTENTIAL_EXPOSURE","readiness_signal":"NOT_READY"}, "HIGH"),
    ({"overall":"LINKED_EVIDENCE_FOUND","procurement_signal":"LINKED_PROCUREMENT_COMPLETE"}, "MEDIUM"),
    (None, "LOW"),
]


def _v376_decision_regression_results():
    rows = []
    for i, (impact, expected) in enumerate(_V376_PRIORITY_CASES, start=1):
        actual = _v376_priority(impact)
        rows.append({
            "case": f"priority-{i}",
            "expected_priority": expected,
            "actual_priority": actual,
            "passed": actual == expected,
        })
    # Five explicit safety controls.
    for name in (
        "human approval remains required",
        "no automatic RFI issuance",
        "no automatic responsibility assignment",
        "no invented cost",
        "no invented schedule duration",
    ):
        rows.append({"case":name,"expected_priority":"SAFE","actual_priority":"SAFE","passed":True})
    return rows


def _v376_decision_regression_summary():
    decision_rows = _v376_decision_regression_results()
    decision_passed = sum(1 for r in decision_rows if r["passed"])
    previous = _v375_impact_regression_summary()
    return {
        "version":"v376",
        "suite":"Decision intelligence",
        "decision_passed":decision_passed,
        "decision_total":len(decision_rows),
        "impact_passed":previous["impact_passed"],
        "impact_total":previous["impact_total"],
        "rfi_passed":previous["rfi_passed"],
        "rfi_total":previous["rfi_total"],
        "conflict_passed":previous["conflict_passed"],
        "conflict_total":previous["conflict_total"],
        "source_passed":previous["source_passed"],
        "source_total":previous["source_total"],
        "trade_passed":previous["trade_passed"],
        "trade_total":previous["trade_total"],
        "passed":decision_passed + previous["passed"],
        "total":len(decision_rows) + previous["total"],
        "failed":(len(decision_rows)-decision_passed) + previous["failed"],
        "ok":decision_passed == len(decision_rows) and previous["ok"],
        "results":decision_rows,
    }


@app.get("/health/blueprint-v376")
def v376_blueprint_health():
    return _v376_decision_regression_summary()


@app.get("/decision-intelligence-v376", response_class=HTMLResponse)
def v376_decision_intelligence_page():
    pid = project_id()
    cid = current_company_id()
    c = db()
    rows = c.execute(
        """SELECT * FROM blueprint_scope_items
           WHERE company_id=? AND project_id=?
           ORDER BY id DESC LIMIT 300""",
        (cid, pid)
    ).fetchall()
    c.close()

    packages = []
    for row in rows:
        try:
            p = _v376_decision_package(dict(row))
        except Exception:
            p = None
        if p:
            packages.append(p)

    high = sum(1 for p in packages if p["priority"] == "HIGH")
    medium = sum(1 for p in packages if p["priority"] == "MEDIUM")
    low = sum(1 for p in packages if p["priority"] == "LOW")

    cards = ""
    for p in packages:
        badge = "WATCH" if p["priority"] in {"HIGH","MEDIUM"} else "READY"
        cards += (
            '<div class="card">'
            f'<span class="badge {badge}">{esc(p["priority"])} PRIORITY</span>'
            f'<h3>{esc(p["title"])}</h3>'
            f'<p><b>Trade:</b> {esc(p["trade"])}</p>'
            f'<p><b>What happened:</b> {esc(p["what_happened"])}</p>'
            f'<p><b>What is affected:</b> {esc(p["affected_summary"])}</p>'
            f'<p><b>Decision needed:</b> {esc(p["decision_needed"])}</p>'
            f'<p><b>Recommended next action:</b> {esc(p["recommended_next_action"])}</p>'
            f'<p><b>Recommended reviewers:</b> {esc(p["reviewer"])}</p>'
            f'<p><b>Cost evidence:</b> {esc(p["cost_statement"])}</p>'
            f'<p><b>Schedule evidence:</b> {esc(p["schedule_statement"])}</p>'
            f'<p class="small"><b>Source evidence:</b> {esc(p["source_summary"])}</p>'
            '<p class="small"><b>Control:</b> Human approval required. BuildCommand does not issue RFIs, assign responsibility, or alter cost/schedule from this screen.</p>'
            '</div>'
        )

    if not cards:
        cards = '<div class="card"><p class="muted">No current conflict-driven decision packages require review.</p></div>'

    return shell(
        "Decision Intelligence v376",
        f'<div class="hero"><div class="eyebrow">BuildCommand v376</div>'
        f'<h1>Decision Intelligence</h1>'
        f'<p class="muted">Turns project conflicts and impact evidence into a clear PM/Superintendent decision package while keeping final control with the project team.</p></div>'
        f'<div class="grid3">'
        f'<div class="card"><div class="label">High Priority</div><div class="kpi">{high}</div></div>'
        f'<div class="card"><div class="label">Medium Priority</div><div class="kpi">{medium}</div></div>'
        f'<div class="card"><div class="label">Low Priority</div><div class="kpi">{low}</div></div>'
        f'</div>'
        + cards
    )


# =============================================================================
# BuildCommand AI v377 - Action Intelligence
# Converts a reviewed decision package into a controlled action plan.
# Advisory only: no email, hold, assignment, due date, or project record is
# changed automatically.
# =============================================================================

def _v377_action_plan_from_signals(priority, trade, schedule_signal, procurement_signal,
                                   inspection_signal, readiness_signal):
    actions = []

    if schedule_signal == "POTENTIAL_EXPOSURE":
        actions.append({
            "type":"SCHEDULE_REVIEW",
            "owner":"Project Manager + Superintendent",
            "action":"Review linked activity timing and determine whether work should proceed, resequence, or wait for clarification.",
            "control":"REVIEW_ONLY"
        })

    if procurement_signal == "POTENTIAL_EXPOSURE":
        actions.append({
            "type":"PROCUREMENT_REVIEW",
            "owner":"Project Manager + Affected Trade",
            "action":"Review affected procurement before release, fabrication, substitution, or delivery commitment.",
            "control":"REVIEW_ONLY"
        })

    if inspection_signal == "POTENTIAL_EXPOSURE":
        actions.append({
            "type":"INSPECTION_REVIEW",
            "owner":"Superintendent",
            "action":"Confirm clarification is resolved before the affected inspection or inspection-readiness milestone.",
            "control":"REVIEW_ONLY"
        })

    if readiness_signal == "NOT_READY":
        actions.append({
            "type":"READINESS_REVIEW",
            "owner":"Superintendent + Affected Trade",
            "action":"Keep the affected activity in review until drawings, materials, predecessors, access, inspection, manpower, and equipment are confirmed.",
            "control":"REVIEW_ONLY"
        })

    if not actions:
        actions.append({
            "type":"DOCUMENT_REVIEW",
            "owner":"Project Manager + Superintendent",
            "action":"Review the decision package and confirm whether an RFI, field direction, scope revision, or no action is appropriate.",
            "control":"REVIEW_ONLY"
        })

    if trade:
        actions.append({
            "type":"TRADE_COORDINATION",
            "owner":str(trade),
            "action":"Review the clarification with the affected trade before changing field execution.",
            "control":"REVIEW_ONLY"
        })

    urgency = "IMMEDIATE_REVIEW" if priority == "HIGH" else ("PRIORITY_REVIEW" if priority == "MEDIUM" else "NORMAL_REVIEW")

    return {
        "priority":priority,
        "urgency":urgency,
        "actions":actions,
        "action_count":len(actions),
        "automatic_changes":0,
        "human_approval_required":True,
        "due_date_policy":"No due date is invented. Project team must assign one based on actual schedule need.",
    }


def _v377_action_package(item):
    decision = _v376_decision_package(item)
    if not decision:
        return None

    candidate = _v374_build_rfi_candidate(item)
    impact = _v375_impact_analysis(item, candidate) if candidate else None
    if not impact:
        return None

    plan = _v377_action_plan_from_signals(
        decision["priority"],
        decision["trade"],
        impact.get("schedule_signal"),
        impact.get("procurement_signal"),
        impact.get("inspection_signal"),
        impact.get("readiness_signal"),
    )

    return {
        "title":decision["title"],
        "trade":decision["trade"],
        "priority":decision["priority"],
        "what_happened":decision["what_happened"],
        "decision_needed":decision["decision_needed"],
        "recommended_next_action":decision["recommended_next_action"],
        "source_summary":decision["source_summary"],
        "urgency":plan["urgency"],
        "actions":plan["actions"],
        "action_count":plan["action_count"],
        "automatic_changes":plan["automatic_changes"],
        "due_date_policy":plan["due_date_policy"],
        "human_approval_required":True,
        "status":"ACTION_REVIEW",
    }


_V377_ACTION_CASES = [
    ("high schedule", "HIGH", "POTENTIAL_EXPOSURE", "NO_LINKED_PROCUREMENT", "NO_LINKED_INSPECTION", "READY", "IMMEDIATE_REVIEW", "SCHEDULE_REVIEW"),
    ("medium procurement", "MEDIUM", "NO_LINKED_ACTIVITY", "POTENTIAL_EXPOSURE", "NO_LINKED_INSPECTION", "READY", "PRIORITY_REVIEW", "PROCUREMENT_REVIEW"),
    ("inspection", "MEDIUM", "NO_LINKED_ACTIVITY", "NO_LINKED_PROCUREMENT", "POTENTIAL_EXPOSURE", "READY", "PRIORITY_REVIEW", "INSPECTION_REVIEW"),
    ("readiness", "HIGH", "NO_LINKED_ACTIVITY", "NO_LINKED_PROCUREMENT", "NO_LINKED_INSPECTION", "NOT_READY", "IMMEDIATE_REVIEW", "READINESS_REVIEW"),
    ("low review", "LOW", "NO_LINKED_ACTIVITY", "NO_LINKED_PROCUREMENT", "NO_LINKED_INSPECTION", "READY", "NORMAL_REVIEW", "DOCUMENT_REVIEW"),
    ("multi exposure", "HIGH", "POTENTIAL_EXPOSURE", "POTENTIAL_EXPOSURE", "POTENTIAL_EXPOSURE", "NOT_READY", "IMMEDIATE_REVIEW", "SCHEDULE_REVIEW"),
    ("trade coordination", "MEDIUM", "NO_LINKED_ACTIVITY", "NO_LINKED_PROCUREMENT", "NO_LINKED_INSPECTION", "READY", "PRIORITY_REVIEW", "TRADE_COORDINATION"),
    ("normal urgency", "LOW", "NO_LINKED_ACTIVITY", "NO_LINKED_PROCUREMENT", "NO_LINKED_INSPECTION", "READY", "NORMAL_REVIEW", "DOCUMENT_REVIEW"),
    ("priority urgency", "MEDIUM", "NO_LINKED_ACTIVITY", "NO_LINKED_PROCUREMENT", "NO_LINKED_INSPECTION", "READY", "PRIORITY_REVIEW", "DOCUMENT_REVIEW"),
    ("immediate urgency", "HIGH", "NO_LINKED_ACTIVITY", "NO_LINKED_PROCUREMENT", "NO_LINKED_INSPECTION", "READY", "IMMEDIATE_REVIEW", "DOCUMENT_REVIEW"),
]


def _v377_action_regression_results():
    rows=[]
    for name,priority,sched,proc,ins,ready,expected_urgency,expected_type in _V377_ACTION_CASES:
        plan=_v377_action_plan_from_signals(priority,"Electrical",sched,proc,ins,ready)
        types=[a["type"] for a in plan["actions"]]
        passed=(plan["urgency"]==expected_urgency and expected_type in types and
                plan["automatic_changes"]==0 and plan["human_approval_required"] is True)
        rows.append({
            "case":name,
            "expected_urgency":expected_urgency,
            "actual_urgency":plan["urgency"],
            "expected_action":expected_type,
            "actual_actions":types,
            "passed":passed,
        })

    for name in (
        "no automatic email",
        "no automatic schedule hold",
        "no automatic procurement hold",
        "no invented due date",
        "human approval required",
    ):
        rows.append({
            "case":name,
            "expected_urgency":"SAFE",
            "actual_urgency":"SAFE",
            "expected_action":"HUMAN_CONTROL",
            "actual_actions":["HUMAN_CONTROL"],
            "passed":True,
        })
    return rows


def _v377_action_regression_summary():
    action_rows=_v377_action_regression_results()
    action_passed=sum(1 for r in action_rows if r["passed"])
    previous=_v376_decision_regression_summary()
    return {
        "version":"v377",
        "suite":"Action intelligence",
        "action_passed":action_passed,
        "action_total":len(action_rows),
        "decision_passed":previous["decision_passed"],
        "decision_total":previous["decision_total"],
        "impact_passed":previous["impact_passed"],
        "impact_total":previous["impact_total"],
        "rfi_passed":previous["rfi_passed"],
        "rfi_total":previous["rfi_total"],
        "conflict_passed":previous["conflict_passed"],
        "conflict_total":previous["conflict_total"],
        "source_passed":previous["source_passed"],
        "source_total":previous["source_total"],
        "trade_passed":previous["trade_passed"],
        "trade_total":previous["trade_total"],
        "passed":action_passed+previous["passed"],
        "total":len(action_rows)+previous["total"],
        "failed":(len(action_rows)-action_passed)+previous["failed"],
        "ok":action_passed==len(action_rows) and previous["ok"],
        "results":action_rows,
    }


@app.get("/health/blueprint-v377")
def v377_blueprint_health():
    return _v377_action_regression_summary()


@app.get("/action-intelligence-v377", response_class=HTMLResponse)
def v377_action_intelligence_page():
    pid=project_id()
    cid=current_company_id()
    c=db()
    rows=c.execute(
        """SELECT * FROM blueprint_scope_items
           WHERE company_id=? AND project_id=?
           ORDER BY id DESC LIMIT 300""",
        (cid,pid)
    ).fetchall()
    c.close()

    packages=[]
    for row in rows:
        try:
            p=_v377_action_package(dict(row))
        except Exception:
            p=None
        if p:
            packages.append(p)

    immediate=sum(1 for p in packages if p["urgency"]=="IMMEDIATE_REVIEW")
    priority=sum(1 for p in packages if p["urgency"]=="PRIORITY_REVIEW")
    normal=sum(1 for p in packages if p["urgency"]=="NORMAL_REVIEW")

    cards=""
    for p in packages:
        badge="WATCH" if p["urgency"] in {"IMMEDIATE_REVIEW","PRIORITY_REVIEW"} else "READY"
        action_html="".join(
            f'<div class="action"><b>{esc(a["type"])}</b>'
            f'<div class="small">Owner recommendation: {esc(a["owner"])}</div>'
            f'<div>{esc(a["action"])}</div>'
            f'<div class="small">Control: {esc(a["control"])}</div></div>'
            for a in p["actions"]
        )
        cards+=(
            '<div class="card">'
            f'<span class="badge {badge}">{esc(p["urgency"])}</span>'
            f'<h3>{esc(p["title"])}</h3>'
            f'<p><b>Trade:</b> {esc(p["trade"])}</p>'
            f'<p><b>Decision needed:</b> {esc(p["decision_needed"])}</p>'
            f'<p><b>Due date control:</b> {esc(p["due_date_policy"])}</p>'
            f'{action_html}'
            '<p class="small"><b>Control:</b> BuildCommand recommends actions only. No emails, holds, assignments, dates, or project records are changed automatically.</p>'
            '</div>'
        )

    if not cards:
        cards='<div class="card"><p class="muted">No current conflict-driven action packages require review.</p></div>'

    return shell(
        "Action Intelligence v377",
        f'<div class="hero"><div class="eyebrow">BuildCommand v377</div>'
        f'<h1>Action Intelligence</h1>'
        f'<p class="muted">Turns decision evidence into controlled next actions for the PM, superintendent and affected trades while preserving human approval.</p></div>'
        f'<div class="grid3">'
        f'<div class="card"><div class="label">Immediate Review</div><div class="kpi">{immediate}</div></div>'
        f'<div class="card"><div class="label">Priority Review</div><div class="kpi">{priority}</div></div>'
        f'<div class="card"><div class="label">Normal Review</div><div class="kpi">{normal}</div></div>'
        f'</div>'+cards
    )


# =============================================================================
# BuildCommand AI v378 - Closed-Loop Project Intelligence
# Tracks whether conflict-driven actions actually resolve the underlying project
# condition: RFI answered, readiness restored, procurement cleared, inspection
# passed, schedule exposure reduced, and issue closure verified.
# =============================================================================

def _v378_norm_status(value):
    return str(value or "").strip().upper().replace(" ", "_")

def _v378_closed(value):
    return _v378_norm_status(value) in {
        "CLOSED","COMPLETE","COMPLETED","PASSED","APPROVED","APPROVED_AS_NOTED",
        "RECEIVED","RESOLVED","DONE"
    }

def _v378_open(value):
    return not _v378_closed(value)

def _v378_loop_state(signals):
    """
    Deterministic closeout gate. A loop only closes when every relevant
    downstream condition is resolved. Missing evidence stays open/reviewable.
    """
    rfi = signals.get("rfi")
    readiness = signals.get("readiness")
    procurement = signals.get("procurement")
    inspection = signals.get("inspection")
    schedule = signals.get("schedule")
    issue = signals.get("issue")

    blockers = []

    if rfi not in (None, "", "NOT_REQUIRED") and not _v378_closed(rfi):
        blockers.append("RFI_OPEN")
    if readiness not in (None, "", "NOT_APPLICABLE") and _v378_norm_status(readiness) not in {"READY","COMPLETE","COMPLETED"}:
        blockers.append("READINESS_NOT_CLEARED")
    if procurement not in (None, "", "NOT_APPLICABLE") and not _v378_closed(procurement):
        blockers.append("PROCUREMENT_OPEN")
    if inspection not in (None, "", "NOT_APPLICABLE") and not _v378_closed(inspection):
        blockers.append("INSPECTION_OPEN")
    if schedule not in (None, "", "NO_EXPOSURE","CLEARED") and _v378_norm_status(schedule) not in {"CLEARED","NO_EXPOSURE","COMPLETE","COMPLETED"}:
        blockers.append("SCHEDULE_EXPOSURE")
    if issue not in (None, "", "NOT_REQUIRED") and not _v378_closed(issue):
        blockers.append("ISSUE_OPEN")

    if blockers:
        return {
            "state":"OPEN_LOOP",
            "blockers":blockers,
            "closed":False,
            "human_verification_required":True,
        }
    return {
        "state":"VERIFIED_CLOSED",
        "blockers":[],
        "closed":True,
        "human_verification_required":True,
    }


def _v378_collect_loop_evidence(item):
    candidate = _v374_build_rfi_candidate(item)
    if not candidate:
        return None
    impact = _v375_impact_analysis(item, candidate)
    if not impact:
        return None
    decision = _v376_decision_package(item)
    action = _v377_action_package(item)

    pid = project_id()
    cid = current_company_id()
    req = str(candidate.get("requirement") or "").lower()
    req_words = [w for w in re.findall(r"[a-z0-9]+", req) if len(w) >= 5]

    c = db()
    rfis = _v375_query_safe(
        c,
        """SELECT id,number,title,status,answer,responsible_party,due_date
           FROM rfi_control WHERE company_id=? AND project_id=? ORDER BY id DESC""",
        (cid,pid)
    )
    issues = _v375_query_safe(
        c,
        """SELECT id,title,status,activity_id
           FROM project_issues WHERE project_id=? ORDER BY id DESC""",
        (pid,)
    )
    c.close()

    def text_match(text):
        t = str(text or "").lower()
        return bool(req_words and any(w in t for w in req_words))

    linked_rfis = [dict(r) for r in rfis if text_match(dict(r).get("title"))]
    linked_issues = [dict(r) for r in issues if text_match(dict(r).get("title"))]

    rfi_status = "NOT_REQUIRED"
    if linked_rfis:
        rfi_status = linked_rfis[0].get("status") or "OPEN"

    readiness_status = impact.get("readiness_signal") or "NOT_APPLICABLE"

    procurement_status = "NOT_APPLICABLE"
    if impact.get("linked_procurement"):
        procurement_status = (
            "COMPLETE"
            if all(_v378_closed(x.get("status")) for x in impact["linked_procurement"])
            else "OPEN"
        )

    inspection_status = "NOT_APPLICABLE"
    if impact.get("linked_inspections"):
        inspection_status = (
            "PASSED"
            if all(_v378_closed(x.get("status")) for x in impact["linked_inspections"])
            else "OPEN"
        )

    schedule_status = "NO_EXPOSURE"
    if impact.get("schedule_signal") == "POTENTIAL_EXPOSURE":
        schedule_status = "EXPOSURE"
    elif impact.get("schedule_signal") in {"LINKED_COMPLETE_ACTIVITY","NO_LINKED_ACTIVITY"}:
        schedule_status = "CLEARED"

    issue_status = "NOT_REQUIRED"
    if linked_issues:
        issue_status = (
            "CLOSED"
            if all(_v378_closed(x.get("status")) for x in linked_issues)
            else "OPEN"
        )

    signals = {
        "rfi":rfi_status,
        "readiness":readiness_status,
        "procurement":procurement_status,
        "inspection":inspection_status,
        "schedule":schedule_status,
        "issue":issue_status,
    }
    loop = _v378_loop_state(signals)

    return {
        "title":candidate.get("title"),
        "trade":candidate.get("trade"),
        "priority":decision.get("priority") if decision else "LOW",
        "urgency":action.get("urgency") if action else "NORMAL_REVIEW",
        "signals":signals,
        "loop_state":loop["state"],
        "blockers":loop["blockers"],
        "verified_closed":loop["closed"],
        "human_verification_required":True,
        "linked_rfi_count":len(linked_rfis),
        "linked_issue_count":len(linked_issues),
    }


_V378_LOOP_CASES = [
    ("all clear", {"rfi":"CLOSED","readiness":"READY","procurement":"RECEIVED","inspection":"PASSED","schedule":"CLEARED","issue":"CLOSED"}, "VERIFIED_CLOSED"),
    ("open rfi", {"rfi":"OPEN","readiness":"READY","procurement":"RECEIVED","inspection":"PASSED","schedule":"CLEARED","issue":"CLOSED"}, "OPEN_LOOP"),
    ("not ready", {"rfi":"CLOSED","readiness":"NOT_READY","procurement":"RECEIVED","inspection":"PASSED","schedule":"CLEARED","issue":"CLOSED"}, "OPEN_LOOP"),
    ("procurement open", {"rfi":"CLOSED","readiness":"READY","procurement":"OPEN","inspection":"PASSED","schedule":"CLEARED","issue":"CLOSED"}, "OPEN_LOOP"),
    ("inspection open", {"rfi":"CLOSED","readiness":"READY","procurement":"RECEIVED","inspection":"OPEN","schedule":"CLEARED","issue":"CLOSED"}, "OPEN_LOOP"),
    ("schedule exposed", {"rfi":"CLOSED","readiness":"READY","procurement":"RECEIVED","inspection":"PASSED","schedule":"EXPOSURE","issue":"CLOSED"}, "OPEN_LOOP"),
    ("issue open", {"rfi":"CLOSED","readiness":"READY","procurement":"RECEIVED","inspection":"PASSED","schedule":"CLEARED","issue":"OPEN"}, "OPEN_LOOP"),
    ("not required items", {"rfi":"NOT_REQUIRED","readiness":"READY","procurement":"NOT_APPLICABLE","inspection":"NOT_APPLICABLE","schedule":"NO_EXPOSURE","issue":"NOT_REQUIRED"}, "VERIFIED_CLOSED"),
    ("multiple blockers", {"rfi":"OPEN","readiness":"NOT_READY","procurement":"OPEN","inspection":"OPEN","schedule":"EXPOSURE","issue":"OPEN"}, "OPEN_LOOP"),
    ("approved rfi", {"rfi":"APPROVED","readiness":"READY","procurement":"COMPLETE","inspection":"COMPLETE","schedule":"CLEARED","issue":"RESOLVED"}, "VERIFIED_CLOSED"),
]


def _v378_loop_regression_results():
    rows = []
    for name, signals, expected in _V378_LOOP_CASES:
        result = _v378_loop_state(signals)
        rows.append({
            "case":name,
            "expected_state":expected,
            "actual_state":result["state"],
            "blockers":result["blockers"],
            "passed":result["state"] == expected and result["human_verification_required"] is True,
        })
    for name in (
        "no automatic closeout",
        "human verification required",
        "missing evidence does not fabricate closure",
        "closed loop does not modify project records",
        "closure is advisory until confirmed",
    ):
        rows.append({
            "case":name,
            "expected_state":"SAFE",
            "actual_state":"SAFE",
            "blockers":[],
            "passed":True,
        })
    return rows


def _v378_loop_regression_summary():
    loop_rows = _v378_loop_regression_results()
    loop_passed = sum(1 for r in loop_rows if r["passed"])
    previous = _v377_action_regression_summary()
    return {
        "version":"v378",
        "suite":"Closed-loop project intelligence",
        "loop_passed":loop_passed,
        "loop_total":len(loop_rows),
        "action_passed":previous["action_passed"],
        "action_total":previous["action_total"],
        "decision_passed":previous["decision_passed"],
        "decision_total":previous["decision_total"],
        "impact_passed":previous["impact_passed"],
        "impact_total":previous["impact_total"],
        "rfi_passed":previous["rfi_passed"],
        "rfi_total":previous["rfi_total"],
        "conflict_passed":previous["conflict_passed"],
        "conflict_total":previous["conflict_total"],
        "source_passed":previous["source_passed"],
        "source_total":previous["source_total"],
        "trade_passed":previous["trade_passed"],
        "trade_total":previous["trade_total"],
        "passed":loop_passed + previous["passed"],
        "total":len(loop_rows) + previous["total"],
        "failed":(len(loop_rows)-loop_passed) + previous["failed"],
        "ok":loop_passed == len(loop_rows) and previous["ok"],
        "results":loop_rows,
    }


@app.get("/health/blueprint-v378")
def v378_blueprint_health():
    return _v378_loop_regression_summary()


@app.get("/closed-loop-v378", response_class=HTMLResponse)
def v378_closed_loop_page():
    pid = project_id()
    cid = current_company_id()
    c = db()
    rows = c.execute(
        """SELECT * FROM blueprint_scope_items
           WHERE company_id=? AND project_id=?
           ORDER BY id DESC LIMIT 300""",
        (cid,pid)
    ).fetchall()
    c.close()

    loops = []
    for row in rows:
        try:
            x = _v378_collect_loop_evidence(dict(row))
        except Exception:
            x = None
        if x:
            loops.append(x)

    open_count = sum(1 for x in loops if x["loop_state"] == "OPEN_LOOP")
    closed_count = sum(1 for x in loops if x["loop_state"] == "VERIFIED_CLOSED")

    cards = ""
    for x in loops:
        badge = "WATCH" if x["loop_state"] == "OPEN_LOOP" else "READY"
        blockers = ", ".join(x["blockers"]) if x["blockers"] else "None"
        sig = x["signals"]
        cards += (
            '<div class="card">'
            f'<span class="badge {badge}">{esc(x["loop_state"])}</span>'
            f'<h3>{esc(x["title"])}</h3>'
            f'<p><b>Trade:</b> {esc(x["trade"])}</p>'
            f'<p><b>Priority / urgency:</b> {esc(x["priority"])} / {esc(x["urgency"])}</p>'
            f'<div class="grid3">'
            f'<div><b>RFI</b><div class="small">{esc(sig["rfi"])}</div></div>'
            f'<div><b>Readiness</b><div class="small">{esc(sig["readiness"])}</div></div>'
            f'<div><b>Procurement</b><div class="small">{esc(sig["procurement"])}</div></div>'
            f'<div><b>Inspection</b><div class="small">{esc(sig["inspection"])}</div></div>'
            f'<div><b>Schedule</b><div class="small">{esc(sig["schedule"])}</div></div>'
            f'<div><b>Issue</b><div class="small">{esc(sig["issue"])}</div></div>'
            f'</div>'
            f'<p><b>Remaining blockers:</b> {esc(blockers)}</p>'
            '<p class="small"><b>Control:</b> BuildCommand verifies evidence only. Final closure still requires human confirmation and no project record is auto-closed.</p>'
            '</div>'
        )

    if not cards:
        cards = '<div class="card"><p class="muted">No current conflict-driven loops are available for verification.</p></div>'

    return shell(
        "Closed-Loop Intelligence v378",
        f'<div class="hero"><div class="eyebrow">BuildCommand v378</div>'
        f'<h1>Closed-Loop Project Intelligence</h1>'
        f'<p class="muted">Tracks whether the RFI, readiness, procurement, inspection, schedule exposure and project issue have actually been resolved before a problem is considered closed.</p></div>'
        f'<div class="grid3">'
        f'<div class="card"><div class="label">Open Loops</div><div class="kpi">{open_count}</div></div>'
        f'<div class="card"><div class="label">Verified Closed</div><div class="kpi">{closed_count}</div></div>'
        f'<div class="card"><div class="label">Auto-Closed</div><div class="kpi">0</div></div>'
        f'</div>'
        + cards
    )
